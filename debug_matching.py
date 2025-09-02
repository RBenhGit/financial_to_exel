#!/usr/bin/env python3
"""
Debug Column Matching Logic
"""

import sys
from pathlib import Path

# Add the project root to Python path
project_root = Path(__file__).parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from openpyxl import load_workbook

def debug_column_matching():
    """Debug why columns aren't matching"""
    print("Debugging Column Matching")
    print("=" * 40)
    
    # Step 1: Load and populate registry
    from core.data_processing.financial_variable_registry import get_registry, VariableCategory
    from core.data_processing.standard_financial_variables import register_all_variables
    
    registry = get_registry()
    register_all_variables(registry)
    
    # Get income statement variables
    all_vars = registry.list_all_variables()
    income_vars = []
    
    for var_name in all_vars:
        var_def = registry.get_variable_definition(var_name)
        if var_def and var_def.category == VariableCategory.INCOME_STATEMENT:
            income_vars.append(var_def)
    
    print(f"Income statement variables in registry: {len(income_vars)}")
    
    # Step 2: Show some variable details with aliases
    print("\nSample income statement variables:")
    for var_def in income_vars[:5]:
        print(f"  {var_def.name}: {var_def.description}")
        if hasattr(var_def, 'aliases') and var_def.aliases:
            print(f"    Aliases: {var_def.aliases}")
    
    # Step 3: Read Excel file and show variable names
    print("\nExcel file variable names:")
    test_file = "data/companies/GOOG/FY/Alphabet Inc Class C - Income Statement.xlsx"
    
    wb = load_workbook(filename=test_file)
    sheet = wb.active
    
    # Convert to list of lists
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    # Find header row
    header_row_idx = None
    for i, row in enumerate(data):
        if row and any('FY-' in str(cell) or 'FY' == str(cell) for cell in row if cell):
            header_row_idx = i
            break
    
    # Show variable names (first column of data rows)
    excel_vars = []
    for row in data[header_row_idx + 1:]:
        if row and row[0] and str(row[0]).strip():
            var_name = str(row[0]).strip()
            excel_vars.append(var_name)
    
    print(f"Found {len(excel_vars)} variable names in Excel")
    print("First 10 Excel variables:")
    for var in excel_vars[:10]:
        print(f"  '{var}'")
    
    # Step 4: Test matching logic
    print("\nTesting matching logic:")
    
    # Simulate the matching function
    def test_match_row_to_variable(row_label, relevant_variables):
        """Test version of the matching logic"""
        row_label_clean = row_label.lower().strip()
        
        print(f"\nTesting '{row_label}' (clean: '{row_label_clean}')")
        
        # Try exact name match first
        for var_def in relevant_variables:
            if var_def.name.lower() == row_label_clean:
                print(f"  EXACT MATCH: {var_def.name}")
                return var_def
        
        # Try alias matching
        for var_def in relevant_variables:
            if hasattr(var_def, 'aliases') and var_def.aliases:
                # Check excel-specific aliases
                if 'excel' in var_def.aliases:
                    excel_aliases = var_def.aliases['excel']
                    if isinstance(excel_aliases, str):
                        excel_aliases = [excel_aliases]
                    
                    for alias in excel_aliases:
                        if alias.lower().strip() == row_label_clean:
                            print(f"  EXCEL ALIAS MATCH: {var_def.name} (alias: {alias})")
                            return var_def
                
                # Check generic aliases
                for alias_source, aliases in var_def.aliases.items():
                    if isinstance(aliases, str):
                        aliases = [aliases]
                    elif aliases is None:
                        continue
                    
                    for alias in aliases:
                        if alias and alias.lower().strip() == row_label_clean:
                            print(f"  GENERIC ALIAS MATCH: {var_def.name} (source: {alias_source}, alias: {alias})")
                            return var_def
        
        # Try partial matching
        for var_def in relevant_variables:
            var_name_clean = var_def.name.lower().replace('_', ' ')
            row_label_words = set(row_label_clean.replace('_', ' ').split())
            var_name_words = set(var_name_clean.split())
            
            common_words = row_label_words & var_name_words
            if len(common_words) >= min(2, len(var_name_words) * 0.7):
                print(f"  PARTIAL MATCH: {var_def.name} (common words: {common_words})")
                return var_def
        
        print(f"  NO MATCH FOUND")
        return None
    
    # Test matching with first few Excel variables
    print(f"\nTesting matches for first 5 Excel variables:")
    for excel_var in excel_vars[:5]:
        test_match_row_to_variable(excel_var, income_vars)

if __name__ == "__main__":
    debug_column_matching()