#!/usr/bin/env python3
"""
Analyze FCF calculations in Excel file vs App calculations
"""

import pandas as pd
import openpyxl
from pathlib import Path
import sys
import os

def analyze_excel_fcf():
    """Analyze FCF calculations in the Excel file"""
    
    excel_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/MSFT/FCF_Analysis_Microsoft_Corporation.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"Excel file not found: {excel_path}")
        return
    
    print("=== EXCEL FILE FCF ANALYSIS ===")
    print(f"File: {excel_path}")
    print()
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        print("Available sheets:", workbook.sheetnames)
        print()
        
        # Analyze each sheet
        for sheet_name in workbook.sheetnames:
            print(f"--- Sheet: {sheet_name} ---")
            sheet = workbook[sheet_name]
            
            # Look for FCF-related data
            fcf_data = {}
            years = []
            
            # Scan for headers and data
            for row in range(1, min(50, sheet.max_row + 1)):  # First 50 rows
                for col in range(1, min(20, sheet.max_column + 1)):  # First 20 columns
                    cell = sheet.cell(row, col)
                    if cell.value:
                        cell_str = str(cell.value).strip().lower()
                        
                        # Look for FCF-related keywords
                        if any(keyword in cell_str for keyword in ['fcf', 'free cash flow', 'cash flow']):
                            print(f"  Found FCF-related cell at {cell.coordinate}: {cell.value}")
                            
                            # Try to get the row data
                            row_data = []
                            for c in range(col, min(col + 15, sheet.max_column + 1)):
                                val = sheet.cell(row, c).value
                                if val is not None:
                                    row_data.append(val)
                                else:
                                    row_data.append("")
                            print(f"    Row data: {row_data}")
                        
                        # Look for years
                        if isinstance(cell.value, (int, float)) and 2010 <= cell.value <= 2030:
                            years.append((row, col, cell.value))
            
            if years:
                print(f"  Found years: {[y[2] for y in years]}")
            print()
    
    except Exception as e:
        print(f"Error reading Excel file: {e}")

def analyze_app_fcf():
    """Analyze how the app calculates FCF"""
    
    print("=== APP FCF CALCULATION METHODS ===")
    
    # Import the financial calculator
    try:
        from financial_calculations import FinancialCalculator
        
        # Show the FCF calculation methods
        print("App FCF Calculation Methods:")
        print()
        
        # Read the source code to understand calculations
        with open('financial_calculations.py', 'r') as f:
            content = f.read()
            
        # Look for FCF calculation methods
        lines = content.split('\n')
        in_fcf_method = False
        current_method = ""
        
        for i, line in enumerate(lines):
            if 'def calculate_' in line and 'fcf' in line.lower():
                in_fcf_method = True
                current_method = line.strip()
                print(f"Method: {current_method}")
                print("---")
                
            elif in_fcf_method and line.strip().startswith('def ') and 'fcf' not in line.lower():
                in_fcf_method = False
                current_method = ""
                print()
                
            elif in_fcf_method and line.strip():
                # Print method content
                if not line.strip().startswith('#') and not line.strip().startswith('"""'):
                    print(f"  {line}")
                    
    except Exception as e:
        print(f"Error analyzing app calculations: {e}")

def load_msft_data_and_calculate():
    """Load MSFT data and show actual calculations"""
    
    print("=== ACTUAL MSFT DATA CALCULATIONS ===")
    
    try:
        from financial_calculations import FinancialCalculator
        
        msft_folder = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/MSFT"
        
        if os.path.exists(msft_folder):
            print(f"Loading MSFT data from: {msft_folder}")
            
            calc = FinancialCalculator(msft_folder)
            fcf_results = calc.calculate_all_fcf_types()
            
            print("FCF Results from App:")
            for fcf_type, values in fcf_results.items():
                if values:
                    print(f"  {fcf_type}: {values} (in millions)")
            print()
            
            # Show intermediate calculations
            print("Intermediate Financial Data:")
            for statement_type, data in calc.financial_data.items():
                print(f"  {statement_type}:")
                if isinstance(data, dict):
                    for year, year_data in data.items():
                        if isinstance(year_data, dict):
                            relevant_items = {}
                            for key, value in year_data.items():
                                if any(keyword in str(key).lower() for keyword in 
                                      ['ebit', 'tax', 'depreciation', 'amortization', 'capex', 'cash flow', 'net income']):
                                    relevant_items[key] = value
                            if relevant_items:
                                print(f"    {year}: {relevant_items}")
                print()
                
        else:
            print(f"MSFT folder not found: {msft_folder}")
            
    except Exception as e:
        print(f"Error loading MSFT data: {e}")

if __name__ == "__main__":
    analyze_excel_fcf()
    print("\n" + "="*80 + "\n")
    analyze_app_fcf()
    print("\n" + "="*80 + "\n")
    load_msft_data_and_calculate()