import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

def diagnose_year9_fcf_issue():
    """
    Specifically diagnose why year 9 shows N/A for FCFF and FCFE
    """
    file_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/V/FCF_Analysis_Visa_Inc_Class_A.xlsx"
    
    try:
        # Load the workbook
        wb = load_workbook(file_path, data_only=True)
        print(f"Available sheets: {wb.sheetnames}")
        
        # Check the FCF DATA sheet
        if 'FCF DATA' in wb.sheetnames:
            ws = wb['FCF DATA']
            print(f"\n=== Year 9 Diagnosis ===")
            
            # First, identify which column represents year 9
            # Based on the previous output, it looks like years start from 2016
            # So year 9 would be 2024 (2016 + 8 = 2024)
            
            print("Looking for 2024 data (which should be year 9)...")
            
            # Find all rows with 2024
            year_2024_rows = []
            for row_num in range(1, ws.max_row + 1):
                for col_num in range(1, 5):  # Check first few columns for year
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value == 2024:
                        year_2024_rows.append(row_num)
                        break
            
            print(f"Found 2024 data in rows: {year_2024_rows}")
            
            # Examine each 2024 row
            for row_num in year_2024_rows:
                print(f"\nRow {row_num} (2024 data):")
                row_data = []
                for col_num in range(1, 16):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data.append(cell_value)
                print(f"Data: {row_data}")
            
            # Now look for the actual FCFF and FCFE calculations
            print(f"\n=== FCFF/FCFE Calculation Search ===")
            
            # Look through all sheets for FCFF/FCFE
            for sheet_name in wb.sheetnames:
                if hasattr(wb[sheet_name], 'iter_rows'):  # Skip chart sheets
                    sheet = wb[sheet_name]
                    print(f"\nChecking sheet: {sheet_name}")
                    
                    for row_num in range(1, min(sheet.max_row + 1, 100)):
                        for col_num in range(1, min(sheet.max_column + 1, 20)):
                            cell_value = sheet.cell(row=row_num, column=col_num).value
                            if cell_value and isinstance(cell_value, str):
                                if 'FCFF' in cell_value.upper() or 'FCFE' in cell_value.upper():
                                    print(f"  Found '{cell_value}' at {sheet_name}!{sheet.cell(row=row_num, column=col_num).coordinate}")
                                    
                                    # Get the entire row
                                    row_data = []
                                    for c in range(1, 16):
                                        val = sheet.cell(row=row_num, column=c).value
                                        row_data.append(val)
                                    print(f"    Row data: {row_data}")
                                    
                                    # Check specifically for N/A values
                                    na_found = []
                                    for i, val in enumerate(row_data):
                                        if val and str(val).upper() in ['N/A', '#N/A', 'NA']:
                                            na_found.append(f"Col{i+1}")
                                    if na_found:
                                        print(f"    N/A values found in: {na_found}")
        
        # Also check if we can read the file with pandas to get a different view
        print(f"\n=== Pandas Analysis ===")
        try:
            # Read FCF DATA sheet with pandas
            df = pd.read_excel(file_path, sheet_name='FCF DATA', header=None)
            print(f"DataFrame shape: {df.shape}")
            
            # Look for FCFF/FCFE rows
            fcf_rows = []
            for idx, row in df.iterrows():
                for col_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str) and ('FCFF' in cell.upper() or 'FCFE' in cell.upper()):
                        fcf_rows.append(idx)
                        print(f"Found FCF calculation at row {idx}: {cell}")
                        print(f"Row data: {row.tolist()}")
                        break
            
            if not fcf_rows:
                print("No FCFF/FCFE rows found in FCF DATA sheet")
                
                # Let's look for any mentions of these terms across all data
                for idx, row in df.iterrows():
                    row_str = ' '.join([str(cell) for cell in row if pd.notna(cell)])
                    if 'FCF' in row_str.upper():
                        print(f"Row {idx} contains FCF reference: {row_str[:100]}...")
        
        except Exception as e:
            print(f"Pandas analysis failed: {e}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error in diagnosis: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    diagnose_year9_fcf_issue()