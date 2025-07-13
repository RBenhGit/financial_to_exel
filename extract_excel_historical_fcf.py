#!/usr/bin/env python3
"""
Extract historical FCF data from Excel FCF DATA sheet
"""

import pandas as pd
import openpyxl
import os

def extract_historical_fcf():
    """Extract historical FCF calculations from Excel file"""
    
    excel_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/MSFT/FCF_Analysis_Microsoft_Corporation.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        
        if 'FCF DATA' in workbook.sheetnames:
            sheet = workbook['FCF DATA']
            print("=== EXTRACTING HISTORICAL FCF DATA ===")
            
            # Extract simple FCF (Cash from Operations - CapEx)
            print("\n1. SIMPLE FCF (Cash from Operations - CapEx):")
            print("   Found at rows 2-15")
            
            # Find year columns
            year_row = 2  # Based on our analysis
            years = []
            year_cols = []
            
            for col in range(1, 15):
                cell = sheet.cell(year_row, col)
                if cell.value and isinstance(cell.value, (int, float)) and 2010 <= cell.value <= 2030:
                    years.append(cell.value)
                    year_cols.append(col)
            
            print(f"   Years found: {years}")
            
            # Extract Cash from Operations (row 3)
            cfo_values = []
            for col in year_cols:
                val = sheet.cell(3, col).value
                cfo_values.append(val)
            
            # Extract CapEx (row 4) 
            capex_values = []
            for col in year_cols:
                val = sheet.cell(4, col).value
                capex_values.append(val)
            
            # Extract FCF (row 5)
            fcf_values = []
            for col in year_cols:
                val = sheet.cell(5, col).value
                fcf_values.append(val)
            
            print(f"   Cash from Operations: {cfo_values}")
            print(f"   CapEx: {capex_values}")
            print(f"   FCF (CFO - CapEx): {fcf_values}")
            
            # Calculate our own FCF to verify
            calculated_fcf = []
            for i in range(len(cfo_values)):
                if cfo_values[i] and capex_values[i]:
                    calc_fcf = cfo_values[i] - abs(capex_values[i])
                    calculated_fcf.append(calc_fcf)
                else:
                    calculated_fcf.append(None)
            
            print(f"   Calculated FCF: {calculated_fcf}")
            
            # Extract Levered FCF (around rows 16-30)
            print("\n2. LEVERED FCF (EBIT*(1-Tax Rate) - Interest):")
            print("   Found at rows 17-30")
            
            # Find levered FCF data
            levered_years = []
            levered_year_cols = []
            
            for col in range(1, 15):
                cell = sheet.cell(17, col)  # Year row for levered FCF
                if cell.value and isinstance(cell.value, (int, float)) and 2010 <= cell.value <= 2030:
                    levered_years.append(cell.value)
                    levered_year_cols.append(col)
            
            print(f"   Years found: {levered_years}")
            
            # Extract Levered FCF values (row 23)
            levered_fcf_values = []
            for col in levered_year_cols:
                val = sheet.cell(23, col).value
                levered_fcf_values.append(val)
            
            print(f"   Levered FCF: {levered_fcf_values}")
            
            # Extract FCFE (around rows 32-45)
            print("\n3. FCFE (Free Cash Flow to Equity):")
            print("   Found at rows 34-45")
            
            # Find FCFE data
            fcfe_years = []
            fcfe_year_cols = []
            
            for col in range(1, 15):
                cell = sheet.cell(34, col)  # Year row for FCFE
                if cell.value and isinstance(cell.value, (int, float)) and 2010 <= cell.value <= 2030:
                    fcfe_years.append(cell.value)
                    fcfe_year_cols.append(col)
            
            print(f"   Years found: {fcfe_years}")
            
            # Extract FCFE values (row 41 based on header structure)
            fcfe_values = []
            for col in fcfe_year_cols:
                val = sheet.cell(41, col).value
                fcfe_values.append(val)
            
            print(f"   FCFE: {fcfe_values}")
            
            # Return structured data
            return {
                'simple_fcf': {
                    'years': years,
                    'cash_from_operations': cfo_values,
                    'capex': capex_values,
                    'fcf': fcf_values,
                    'calculated_fcf': calculated_fcf
                },
                'levered_fcf': {
                    'years': levered_years,
                    'values': levered_fcf_values
                },
                'fcfe': {
                    'years': fcfe_years,
                    'values': fcfe_values
                }
            }
            
    except Exception as e:
        print(f"Error: {e}")
        return {}

def compare_with_app():
    """Compare Excel historical data with app calculations"""
    
    from financial_calculations import FinancialCalculator
    
    # Get Excel data
    excel_data = extract_historical_fcf()
    
    # Get App data
    msft_folder = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/MSFT"
    calc = FinancialCalculator(msft_folder)
    app_data = calc.calculate_all_fcf_types()
    
    print("\n" + "="*100)
    print("COMPREHENSIVE FCF COMPARISON: EXCEL HISTORICAL vs APP")
    print("="*100)
    
    # Create detailed comparison table
    print("\n📊 DETAILED COMPARISON TABLE")
    print("-" * 140)
    print(f"{'FCF Type':<15} {'Source':<10} {'Method':<50} {'Recent Values (Millions)':<50} {'Match?':<15}")
    print("-" * 140)
    
    # Simple FCF comparison
    if excel_data.get('simple_fcf'):
        excel_simple = excel_data['simple_fcf']['fcf'][-5:]  # Last 5 values
        app_lfcf = app_data.get('LFCF', [])[-5:]  # Last 5 values
        
        print(f"{'Simple FCF':<15} {'Excel':<10} {'Cash from Operations - CapEx':<50} {str(excel_simple):<50} {'Reference':<15}")
        print(f"{'LFCF':<15} {'App':<10} {'Operating Cash Flow - CapEx':<50} {str(app_lfcf):<50} {'Similar':<15}")
    
    # Levered FCF comparison
    if excel_data.get('levered_fcf'):
        excel_levered = excel_data['levered_fcf']['values'][-5:]
        
        print(f"{'Levered FCF':<15} {'Excel':<10} {'EBIT*(1-Tax Rate) - Interest':<50} {str(excel_levered):<50} {'Different':<15}")
        print(f"{'FCFF':<15} {'App':<10} {'EBIT(1-Tax) + D&A - ΔWC - CapEx':<50} {str(app_data.get('FCFF', [])[-5:]):<50} {'Different':<15}")
    
    # FCFE comparison
    if excel_data.get('fcfe'):
        excel_fcfe = excel_data['fcfe']['values'][-5:]
        app_fcfe = app_data.get('FCFE', [])[-5:]
        
        print(f"{'FCFE':<15} {'Excel':<10} {'NI - NonCash + D&A - ΔWC - CapEx - Debt + NewDebt':<50} {str(excel_fcfe):<50} {'Complex':<15}")
        print(f"{'FCFE':<15} {'App':<10} {'NI + D&A - ΔWC - CapEx + Net Borrowing':<50} {str(app_fcfe):<50} {'Similar':<15}")
    
    print("-" * 140)
    
    return excel_data, app_data

if __name__ == "__main__":
    excel_data, app_data = compare_with_app()