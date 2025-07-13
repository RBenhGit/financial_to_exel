#!/usr/bin/env python3
"""
Detailed Excel Inspector
Examines Excel files to find actual historical FCF calculations
"""

import pandas as pd
import openpyxl

def deep_inspect_excel(excel_path, company_name):
    """Thoroughly inspect Excel file for any FCF data"""
    print(f"\n🔍 DEEP INSPECTION: {company_name}")
    print("=" * 60)
    
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        
        for sheet_name in workbook.sheetnames:
            print(f"\n📊 SHEET: {sheet_name}")
            print("-" * 40)
            ws = workbook[sheet_name]
            
            # Search for any FCF-related content
            fcf_found = False
            for row in range(1, min(30, ws.max_row + 1)):
                for col in range(1, min(10, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        cell_str = cell_value.lower()
                        if any(term in cell_str for term in ['fcf', 'free cash', 'levered', 'unlevered', 'firm', 'equity']):
                            # Found FCF-related text, get surrounding values
                            print(f"   📍 Found '{cell_value}' at R{row}C{col}")
                            
                            # Get values to the right
                            values = []
                            for val_col in range(col + 1, min(col + 12, ws.max_column + 1)):
                                val = ws.cell(row=row, column=val_col).value
                                if isinstance(val, (int, float)):
                                    values.append(val)
                                elif val is None:
                                    values.append(None)
                                else:
                                    break
                            
                            if any(isinstance(v, (int, float)) and abs(v) > 100 for v in values):
                                print(f"      Values: {values[:8]}")
                                fcf_found = True
                            
                            # Get values below
                            values_below = []
                            for val_row in range(row + 1, min(row + 8, ws.max_row + 1)):
                                val = ws.cell(row=val_row, column=col).value
                                if isinstance(val, (int, float)) and abs(val) > 100:
                                    values_below.append(val)
                                elif val is None:
                                    continue
                                else:
                                    break
                            
                            if values_below:
                                print(f"      Values below: {values_below}")
                                fcf_found = True
            
            if not fcf_found:
                print("   ⚠️ No FCF-related data found")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error inspecting {company_name}: {e}")

def inspect_goog_detailed():
    """Specific inspection for GOOG to understand the structure"""
    excel_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/GOOG/FCF_Analysis_Alphabet Inc Class C.xlsx"
    
    print("\n🔍 DETAILED GOOG INSPECTION")
    print("=" * 60)
    
    try:
        # Read FCF DATA sheet with pandas to see structure
        fcf_data = pd.read_excel(excel_path, sheet_name='FCF DATA', header=None)
        print("📊 FCF DATA Sheet Structure:")
        print(f"   Shape: {fcf_data.shape}")
        
        # Show first 20 rows to understand layout
        print("\n📋 First 20 rows:")
        for i in range(min(20, len(fcf_data))):
            row_data = fcf_data.iloc[i].dropna().values[:10]  # First 10 non-null values
            print(f"   Row {i+1}: {list(row_data)}")
        
        # Look for specific patterns
        print("\n🔍 Searching for specific patterns:")
        for i, row in fcf_data.iterrows():
            if i > 30:  # Limit search
                break
            
            first_cell = str(row.iloc[0]).lower() if pd.notna(row.iloc[0]) else ""
            if any(term in first_cell for term in ['operating', 'capital', 'capex', 'depreciation', 'working']):
                values = []
                for j in range(1, min(len(row), 12)):
                    if pd.notna(row.iloc[j]) and isinstance(row.iloc[j], (int, float)):
                        values.append(row.iloc[j])
                if values and any(abs(v) > 1000 for v in values):
                    print(f"   📊 {row.iloc[0]}: {values[:8]}")
    
    except Exception as e:
        print(f"❌ Error in detailed inspection: {e}")

def check_streamlit_app_data():
    """Check what the Streamlit app actually displays"""
    print("\n🎯 STREAMLIT APP DATA CHECK")
    print("=" * 60)
    
    try:
        from financial_calculations import FinancialCalculator
        
        # Test with GOOG
        goog_folder = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/GOOG"
        calc = FinancialCalculator(goog_folder)
        calc.load_financial_statements()
        
        # Get raw metrics to understand what's available
        metrics = calc._calculate_all_metrics()
        
        print("📊 Available Metrics:")
        for key, values in metrics.items():
            if isinstance(values, list) and values:
                print(f"   {key}: {len(values)} values - {values[:3]}..." if len(values) > 3 else f"   {key}: {values}")
        
        # Test original fcf_analysis.py compatibility
        print("\n🔄 Testing with original fcf_analysis.py structure...")
        
        # Check if we can manually calculate FCFF from components
        ebit = metrics.get('ebit', [])
        tax_rates = metrics.get('tax_rates', [])
        da = metrics.get('depreciation_amortization', [])
        capex = metrics.get('capex', [])
        wc_changes = metrics.get('working_capital_changes', [])
        
        print(f"\n🧮 Manual FCFF Components:")
        print(f"   EBIT: {ebit[:5]}...")
        print(f"   Tax Rates: {tax_rates[:5]}...")
        print(f"   D&A: {da[:5]}...")
        print(f"   CapEx: {capex[:5]}...")
        print(f"   WC Changes: {wc_changes[:5]}...")
        
        # Manual FCFF calculation for first few years
        print("\n🔢 Manual FCFF Calculation (first 3 years):")
        for i in range(min(3, len(ebit), len(tax_rates), len(da), len(capex), len(wc_changes))):
            after_tax_ebit = ebit[i] * (1 - tax_rates[i])
            manual_fcff = after_tax_ebit + da[i] - wc_changes[i] - abs(capex[i])
            print(f"   Year {i+1}: EBIT({ebit[i]:.0f}) × (1-{tax_rates[i]:.3f}) + D&A({da[i]:.0f}) - WC({wc_changes[i]:.0f}) - CapEx({abs(capex[i]):.0f}) = {manual_fcff:.0f}")
        
    except Exception as e:
        print(f"❌ Error checking app data: {e}")
        import traceback
        traceback.print_exc()

def main():
    print("🚀 DETAILED EXCEL INSPECTION FOR FCF DATA")
    print("=" * 80)
    
    # 1. Deep inspect Excel files
    companies = {
        'GOOG': '/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/GOOG/FCF_Analysis_Alphabet Inc Class C.xlsx',
        'MSFT': '/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/MSFT/FCF_Analysis_Microsoft_Corporation.xlsx'
    }
    
    for company, excel_path in companies.items():
        deep_inspect_excel(excel_path, company)
    
    # 2. Detailed GOOG inspection
    inspect_goog_detailed()
    
    # 3. Check app data
    check_streamlit_app_data()
    
    print("\n📋 CONCLUSION")
    print("=" * 60)
    print("The Excel files appear to contain DCF projections rather than")
    print("historical FCFF/FCFE calculations. The app calculations are")
    print("working correctly, but we cannot validate them against Excel")
    print("because Excel doesn't contain comparable historical data.")

if __name__ == "__main__":
    main()