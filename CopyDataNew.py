import os
from openpyxl import load_workbook
from datetime import date, datetime
from tkinter import filedialog, Tk
import logging
from data_validator import create_enhanced_copy_validation, DataQualityReport
from config import get_config, get_excel_config, get_financial_metrics_config
from excel_utils import get_company_name_from_excel, get_period_dates_from_excel, ExcelDataExtractor
from error_handler import (
    EnhancedLogger, ExcelDataError, ValidationError, 
    with_error_handling, validate_excel_file, validate_financial_data,
    handle_calculation_error, create_error_summary
)

# Set up enhanced logging
logger = EnhancedLogger(__name__, log_file="financial_analysis.log")

# Initialize data quality tracking
data_quality_report = DataQualityReport()
copy_errors = []
validation_warnings = []

def select_files(default="",title=""):
    """
    Open a file dialog to select one or more files.
    
    Args:
        title (str): The title of the file dialog window
        
    Returns:
        tuple: Selected file paths
    """
    root = Tk()
    root.withdraw()  # Hide the main window
    files = filedialog.askopenfilenames(initialdir=default,title=title)
    if not files:
        logger.warning(f"No files selected for: {title}")
    return files

def categorize_financial_files(files, file_type_label):
    """
    Categorize financial files into Balance Sheet, Cash Flow, and Income Statement.
    
    Args:
        files (tuple): List of file paths
        file_type_label (str): Label for logging purposes (e.g., "FY" or "LTM")
        
    Returns:
        dict: Dictionary containing categorized file paths
    """
    result = {}
    
    for file_path in files:
        if "Balance" in file_path:
            result["balance"] = file_path
            logger.info(f"Found {file_type_label} Balance Sheet: {os.path.basename(file_path)}")
        elif "Cash" in file_path:
            result["cash_flow"] = file_path
            logger.info(f"Found {file_type_label} Cash Flow Statement: {os.path.basename(file_path)}")
        elif "Income" in file_path:
            result["income"] = file_path
            logger.info(f"Found {file_type_label} Income Statement: {os.path.basename(file_path)}")
    
    return result

# Select DCF template file
logger.info("Starting financial data extraction process")
try:
    DCF_file = select_files(title="Please Select DCF template file")[0]
    logger.info(f"Selected DCF template: {os.path.basename(DCF_file)}")
except IndexError:
    logger.error("No DCF template file selected. Exiting.")
    raise SystemExit("Program terminated: DCF template file is required")

# Select and categorize Fiscal Year (FY) files
Selected_Files_FY = select_files("Please Select Fiscal Year financial files")
fy_files = categorize_financial_files(Selected_Files_FY, "FY")

# Extract individual FY file paths
try:
    Balance_Sheet = fy_files["balance"]
    Cash_Flow_Statement = fy_files["cash_flow"]
    Income_Statement = fy_files["income"]
except KeyError as e:
    logger.error(f"Missing required FY file: {e}")
    raise SystemExit(f"Program terminated: Missing required FY file: {e}")

# Select and categorize Latest Twelve Months (LTM) files
Selected_Files_LTM = select_files("Please Select Latest Twelve Months financial files")
ltm_files = categorize_financial_files(Selected_Files_LTM, "LTM")

# Extract individual LTM file paths
try:
    Balance_Sheet_Q = ltm_files["balance"]
    Cash_Flow_Statement_LTM = ltm_files["cash_flow"]
    Income_Statement_LTM = ltm_files["income"]
except KeyError as e:
    logger.error(f"Missing required LTM file: {e}")
    raise SystemExit(f"Program terminated: Missing required LTM file: {e}")

@with_error_handling(error_type=ExcelDataError, re_raise=True)
def load_workbooks():
    """
    Load all required workbooks and their first worksheets with enhanced error handling.
    
    Returns:
        tuple: Tuple containing target workbook and all source worksheets
    """
    # Validate all Excel files before loading
    files_to_validate = [
        ("DCF Template", DCF_file),
        ("Income Statement FY", Income_Statement),
        ("Income Statement LTM", Income_Statement_LTM),
        ("Balance Sheet FY", Balance_Sheet),
        ("Balance Sheet LTM", Balance_Sheet_Q),
        ("Cash Flow Statement FY", Cash_Flow_Statement),
        ("Cash Flow Statement LTM", Cash_Flow_Statement_LTM)
    ]
    
    for file_type, file_path in files_to_validate:
        try:
            validate_excel_file(file_path)
            logger.info(f"✓ {file_type} file validated: {os.path.basename(file_path)}")
        except ExcelDataError as e:
            logger.error(f"✗ {file_type} file validation failed", error=e, 
                        context={'file_path': file_path})
            raise
    
    try:
        # Optimized workbook loading with batch processing
        workbook_configs = [
            ("DCF Template", DCF_file, "target"),
            ("Income Statement FY", Income_Statement, "income"),
            ("Income Statement LTM", Income_Statement_LTM, "income_ltm"),
            ("Balance Sheet FY", Balance_Sheet, "balance"),
            ("Balance Sheet LTM", Balance_Sheet_Q, "balance_q"),
            ("Cash Flow FY", Cash_Flow_Statement, "cashflow"),
            ("Cash Flow LTM", Cash_Flow_Statement_LTM, "cashflow_ltm")
        ]
        
        loaded_workbooks = {}
        loaded_sheets = {}
        
        # Batch load with optimized settings
        logger.info("Batch loading Excel workbooks with optimization...")
        for desc, filepath, key in workbook_configs:
            try:
                # Load with optimization flags
                workbook = load_workbook(
                    filename=filepath,
                    read_only=False,  # We need write access for DCF template
                    keep_vba=False,   # Skip VBA for performance
                    data_only=True,   # Get calculated values, not formulas
                    keep_links=False  # Don't preserve external links
                )
                loaded_workbooks[key] = workbook
                loaded_sheets[key] = workbook.worksheets[0]
                logger.debug(f"Loaded {desc}: {os.path.basename(filepath)}")
            except Exception as e:
                logger.error(f"Failed to load {desc} from {filepath}: {e}")
                raise
        
        # Extract individual components for backward compatibility
        target_file = loaded_workbooks["target"]
        target_sheet = loaded_sheets["target"]
        income_sheet = loaded_sheets["income"]
        income_sheet_ltm = loaded_sheets["income_ltm"]
        balance_sheet = loaded_sheets["balance"]
        balance_sheet_q = loaded_sheets["balance_q"]
        cash_flow_sheet = loaded_sheets["cashflow"]
        cash_flow_sheet_ltm = loaded_sheets["cashflow_ltm"]
        
        logger.info("✓ All workbooks loaded successfully")
        
        return (
            target_file, 
            target_sheet, 
            income_sheet, 
            income_sheet_ltm, 
            balance_sheet, 
            balance_sheet_q, 
            cash_flow_sheet, 
            cash_flow_sheet_ltm
        )
    except Exception as e:
        handle_calculation_error("load_workbooks", e, context={
            'files': [f[1] for f in files_to_validate]
        })
        raise ExcelDataError(f"Failed to load workbooks: {e}") from e

# Load all workbooks
(
    TargetFile, 
    wb1, 
    Income_wb, 
    Income_wb_LTM, 
    Balance_wb, 
    Balance_wb_Q, 
    Cash_Flow_wb, 
    Cash_Flow_wb_LTM
) = load_workbooks()

# Set basic information in target file
wb1['c1'] = date.today()

# Use dynamic company name extraction instead of hardcoded cell reference
Company_Name = get_company_name_from_excel(Income_Statement)
if not Company_Name:
    from config import get_unknown_company_name
    logger.warning(f"Company name not found using dynamic extraction, using '{get_unknown_company_name()}'")
    Company_Name = get_unknown_company_name()
    
wb1['c2'] = Company_Name
logger.info(f"Processing data for company: {Company_Name}")

def extract_period_end_dates(workbook_path):
    """
    Extract Period End Date values from financial statement using dynamic extraction
    
    Args:
        workbook_path: Path to Excel workbook containing financial data
        
    Returns:
        list: List of date strings extracted from Period End Date row
    """
    try:
        # Use dynamic period date extraction instead of hardcoded positions
        dates = get_period_dates_from_excel(workbook_path)
        
        if dates:
            logger.info(f"Extracted {len(dates)} period end dates: {dates}")
        else:
            logger.warning("No period end dates found using dynamic extraction")
            
        return dates
        
    except Exception as e:
        logger.error(f"Error extracting period end dates: {str(e)}")
        return []

def parse_date_year(date_string):
    """
    Parse year from date string in YYYY-MM-DD format
    
    Args:
        date_string: Date string in format "YYYY-MM-DD"
        
    Returns:
        int: Year as integer, or None if parsing fails
    """
    try:
        if isinstance(date_string, str) and len(date_string) >= 4:
            # Extract year from YYYY-MM-DD format
            year_str = date_string[:4]
            return int(year_str)
    except (ValueError, TypeError):
        logger.warning(f"Could not parse year from date: {date_string}")
    return None

def validated_cell_copy(source_cell, target_cell, context_info):
    """
    Copy cell value with enhanced validation and error tracking
    
    Args:
        source_cell: Source cell from Excel workbook
        target_cell: Target cell in DCF workbook
        context_info: Description for error reporting
    """
    global copy_errors, validation_warnings
    
    try:
        original_value = source_cell.value
        validated_value, is_valid = create_enhanced_copy_validation(original_value, context_info)
        
        if is_valid:
            target_cell.value = validated_value
            logger.debug(f"Successfully copied {context_info}: {original_value} -> {validated_value}")
        else:
            target_cell.value = 0  # Safe fallback
            copy_errors.append({
                "context": context_info,
                "original_value": original_value,
                "fallback_value": 0,
                "error": "Validation failed"
            })
            logger.warning(f"Copy validation failed for {context_info}, using fallback value 0")
            
        # Track empty/zero values for reporting
        if original_value is None or original_value == "" or original_value == 0:
            validation_warnings.append({
                "context": context_info,
                "issue": "Empty or zero value",
                "value": original_value
            })
            
    except Exception as e:
        copy_errors.append({
            "context": context_info,
            "original_value": getattr(source_cell, 'value', 'N/A'),
            "error": str(e)
        })
        target_cell.value = 0  # Safe fallback
        logger.error(f"Error copying {context_info}: {e}")

def extract_financial_data():
    """
    Extract financial data from source files and populate the target DCF file.
    Enhanced with comprehensive validation and error tracking.
    """
    global copy_errors, validation_warnings
    logger.info("Starting enhanced financial data extraction with validation")
    
    # Reset tracking variables
    copy_errors = []
    validation_warnings = []
    
    # Extract Period End Dates from Income Statement to determine dynamic year range
    logger.info("Extracting Period End Dates for dynamic year calculation")
    fy_dates = extract_period_end_dates(Income_Statement)
    ltm_dates = extract_period_end_dates(Income_Statement_LTM)
    
    # Parse years from dates
    fy_years = [parse_date_year(date_str) for date_str in fy_dates]
    fy_years = [year for year in fy_years if year is not None]
    
    ltm_years = [parse_date_year(date_str) for date_str in ltm_dates]
    ltm_years = [year for year in ltm_years if year is not None]
    
    if fy_years:
        logger.info(f"Dynamic FY years extracted: {fy_years}")
    else:
        logger.warning("No valid FY years extracted, using fallback logic")
        
    if ltm_years:
        logger.info(f"Dynamic LTM years extracted: {ltm_years}")
    else:
        logger.warning("No valid LTM years extracted, using fallback logic")
    
    # Get financial metrics from configuration instead of hardcoded values
    config = get_config()
    financial_metrics = {
        "income": [
            {"name": name, **details} for name, details in config.financial_metrics.income_metrics.items()
        ],
        "balance": [
            {"name": name, **details} for name, details in config.financial_metrics.balance_metrics.items()
        ],
        "cashflow": [
            {"name": name, **details} for name, details in config.financial_metrics.cashflow_metrics.items()
        ]
    }
    
    # Track metrics found vs. expected
    metrics_found = {"income": 0, "balance": 0, "cashflow": 0}
    metrics_expected = {"income": len(financial_metrics["income"]), 
                       "balance": len(financial_metrics["balance"]), 
                       "cashflow": len(financial_metrics["cashflow"])}
    
    # Create a mapping of all financial data rows using configuration
    logger.info("Scanning financial statements for required metrics")
    excel_config = get_excel_config()
    row_val = []
    for row in range(0, excel_config.max_scan_rows):  # Use configurable scan range
        row_val.append({
            'income_index': row+1,
            'income_value': Income_wb.cell(row+1, 3).value,
            'balance_index': row+1,
            'balance_value': Balance_wb.cell(row+1, 3).value,
            'cashflow_index': row+1,
            'cashflow_value': Cash_Flow_wb.cell(row+1, 3).value
        })
    
    # Process Income Statement metrics
    logger.info("Processing Income Statement metrics with validation")
    for metric in financial_metrics["income"]:
        metric_found = False
        for i in row_val:
            if i['income_value'] == metric["name"]:
                row_index = i['income_index']
                logger.info(f"Found {metric['name']} at row {row_index}")
                metrics_found["income"] += 1
                metric_found = True
                
                # Copy historical data (9 years) with validation using configuration
                for j in range(9):
                    context = f"Income.{metric['name']}.FY-{j+1}"
                    source_cell = Income_wb.cell(row_index, column=excel_config.data_start_column+j)
                    target_cell = wb1.cell(row=15-j, column=metric["target_column"])
                    validated_cell_copy(source_cell, target_cell, context)
                
                # Copy LTM data with validation using configuration
                context = f"Income.{metric['name']}.LTM"
                source_cell = Income_wb_LTM.cell(row_index, column=excel_config.ltm_column)
                target_cell = wb1.cell(row=16, column=metric["target_column"])
                validated_cell_copy(source_cell, target_cell, context)
                
                # Special case for Period End Date
                if metric.get("set_c3", False):
                    wb1['C3'] = wb1.cell(row=16, column=1).value
                    
                    # Save extracted dates information for other modules
                    try:
                        # Save FY and LTM dates to a metadata file for other modules to use
                        import json
                        metadata = {
                            "fy_dates": fy_dates,
                            "ltm_dates": ltm_dates,
                            "fy_years": fy_years,
                            "ltm_years": ltm_years,
                            "last_updated": str(datetime.now())
                        }
                        with open("dates_metadata.json", "w") as f:
                            json.dump(metadata, f, indent=2)
                        logger.info("Saved dates metadata to dates_metadata.json")
                    except Exception as e:
                        logger.warning(f"Could not save dates metadata: {e}")
                break
        
        if not metric_found:
            logger.error(f"Required Income Statement metric not found: {metric['name']}")
            copy_errors.append({
                "context": f"Income.{metric['name']}",
                "error": "Metric not found in source data",
                "impact": "All values for this metric will be missing"
            })
    
    # Process Balance Sheet metrics
    logger.info("Processing Balance Sheet metrics with validation")
    for metric in financial_metrics["balance"]:
        metric_found = False
        for i in row_val:
            if i['balance_value'] == metric["name"]:
                row_index = i['balance_index']
                logger.info(f"Found {metric['name']} at row {row_index}")
                metrics_found["balance"] += 1
                metric_found = True
                
                # Copy historical data (9 years) with validation using configuration
                for j in range(9):
                    context = f"Balance.{metric['name']}.FY-{j+1}"
                    source_cell = Balance_wb.cell(row_index, column=excel_config.data_start_column+j)
                    target_cell = wb1.cell(row=15-j, column=metric["target_column"])
                    validated_cell_copy(source_cell, target_cell, context)
                
                # Copy LTM data with validation using configuration
                context = f"Balance.{metric['name']}.LTM"
                source_cell = Balance_wb_Q.cell(row_index, column=excel_config.ltm_column)
                target_cell = wb1.cell(row=16, column=metric["target_column"])
                validated_cell_copy(source_cell, target_cell, context)
                break
        
        if not metric_found:
            logger.error(f"Required Balance Sheet metric not found: {metric['name']}")
            copy_errors.append({
                "context": f"Balance.{metric['name']}",
                "error": "Metric not found in source data",
                "impact": "All values for this metric will be missing"
            })
    
    # Process Cash Flow Statement metrics
    logger.info("Processing Cash Flow Statement metrics with validation")
    for metric in financial_metrics["cashflow"]:
        metric_found = False
        for i in row_val:
            if i['cashflow_value'] == metric["name"]:
                row_index = i['cashflow_index']
                logger.info(f"Found {metric['name']} at row {row_index}")
                metrics_found["cashflow"] += 1
                metric_found = True
                
                # Copy historical data (9 years) with validation using configuration
                for j in range(9):
                    context = f"CashFlow.{metric['name']}.FY-{j+1}"
                    source_cell = Cash_Flow_wb.cell(row_index, column=excel_config.data_start_column+j)
                    target_cell = wb1.cell(row=15-j, column=metric["target_column"])
                    validated_cell_copy(source_cell, target_cell, context)
                
                # Copy LTM data with validation using configuration
                context = f"CashFlow.{metric['name']}.LTM"
                source_cell = Cash_Flow_wb_LTM.cell(row_index, column=excel_config.ltm_column)
                target_cell = wb1.cell(row=16, column=metric["target_column"])
                validated_cell_copy(source_cell, target_cell, context)
                break
        
        if not metric_found:
            logger.error(f"Required Cash Flow Statement metric not found: {metric['name']}")
            copy_errors.append({
                "context": f"CashFlow.{metric['name']}",
                "error": "Metric not found in source data",
                "impact": "All values for this metric will be missing"
            })
    
    # Generate data quality summary
    generate_data_quality_summary(metrics_found, metrics_expected)
    
    # Generate comprehensive error summary
    error_summary = create_error_summary(copy_errors, validation_warnings)
    logger.info("Data extraction completed", context={
        'metrics_found': metrics_found,
        'metrics_expected': metrics_expected,
        'error_summary': error_summary
    })
    
    # Save error log if there are significant issues
    if error_summary['total_errors'] > 0 or error_summary['total_warnings'] > 10:
        logger.save_error_log(f"data_extraction_issues_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")

def generate_data_quality_summary(metrics_found, metrics_expected):
    """
    Generate and log comprehensive data quality summary
    """
    global copy_errors, validation_warnings, data_quality_report
    
    logger.info("\n" + "="*60)
    logger.info("DATA EXTRACTION QUALITY SUMMARY")
    logger.info("="*60)
    
    # Metrics completion summary
    for statement_type in metrics_found:
        found = metrics_found[statement_type]
        expected = metrics_expected[statement_type]
        completion_rate = (found / expected * 100) if expected > 0 else 0
        logger.info(f"{statement_type.upper()} Statement: {found}/{expected} metrics found ({completion_rate:.1f}%)")
        
        if completion_rate < 100:
            data_quality_report.add_error(f"Incomplete {statement_type} data: {found}/{expected} metrics found")
    
    # Error summary
    if copy_errors:
        logger.warning(f"\nDATA COPY ERRORS: {len(copy_errors)} total")
        error_contexts = {}
        for error in copy_errors:
            context_type = error['context'].split('.')[0] if '.' in error['context'] else 'Unknown'
            error_contexts[context_type] = error_contexts.get(context_type, 0) + 1
        
        for context, count in error_contexts.items():
            logger.warning(f"  - {context}: {count} errors")
            data_quality_report.add_error(f"{count} copy errors in {context} data")
    
    # Warning summary
    if validation_warnings:
        logger.info(f"\nDATA QUALITY WARNINGS: {len(validation_warnings)} total")
        warning_types = {}
        for warning in validation_warnings:
            issue_type = warning['issue']
            warning_types[issue_type] = warning_types.get(issue_type, 0) + 1
        
        for issue_type, count in warning_types.items():
            logger.info(f"  - {issue_type}: {count} instances")
            data_quality_report.add_warning(f"{count} instances of {issue_type}")
    
    # Overall assessment
    total_errors = len(copy_errors)
    total_warnings = len(validation_warnings)
    
    if total_errors == 0 and total_warnings == 0:
        logger.info("\n✓ DATA QUALITY: EXCELLENT - No issues detected")
        data_quality_report.add_recommendation("Data quality is excellent. Proceed with confidence.", "low")
    elif total_errors == 0 and total_warnings < 5:
        logger.info("\n✓ DATA QUALITY: GOOD - Minor warnings only")
        data_quality_report.add_recommendation("Data quality is good with minor warnings.", "low")
    elif total_errors < 3 and total_warnings < 10:
        logger.warning("\n⚠ DATA QUALITY: FAIR - Some issues detected")
        data_quality_report.add_recommendation("Review data quality issues before proceeding with analysis.", "medium")
    else:
        logger.error("\n✗ DATA QUALITY: POOR - Significant issues detected")
        data_quality_report.add_recommendation("Address data quality issues before using for analysis.", "high")
    
    logger.info("\nRECOMMENDATIONS:")
    if total_errors > 0:
        logger.info("  • Review source Excel files for missing or corrupted data")
        logger.info("  • Verify file naming conventions match expected patterns")
        logger.info("  • Check that all required financial statements are included")
    
    if total_warnings > 5:
        logger.info("  • Consider data source quality review")
        logger.info("  • Validate calculations manually for critical metrics")
    
    logger.info("="*60)

# Extract and populate financial data
extract_financial_data()
def save_output_file():
    """
    Save the populated DCF file to a user-selected directory.
    """
    logger.info("Requesting output directory selection")
    path = filedialog.askdirectory(title="Select directory to save the output file")
    
    if not path:
        logger.warning("No directory selected, using current directory")
        path = os.getcwd()
    
    # Create sanitized company name for filename
    safe_company_name = ''.join(c if c.isalnum() else '_' for c in Company_Name)
    file_name = os.path.join(path, f'FCF_Analysis_{safe_company_name}.xlsx')
    
    try:
        TargetFile.save(filename=file_name)
        logger.info(f"File saved successfully: {file_name}")
        print(f'File saved successfully: {file_name}')
        return file_name
    except Exception as e:
        logger.error(f"Error saving file: {e}")
        # Try alternative filename
        alt_file_name = os.path.join(path, 'FCF_Analysis.xlsx')
        try:
            TargetFile.save(filename=alt_file_name)
            logger.info(f"Alternative file saved: {alt_file_name}")
            print(f'Alternative file saved: {alt_file_name}')
            return alt_file_name
        except Exception as e:
            logger.error(f"Error saving alternative file: {e}")
            print(f"Error: Could not save file - {e}")
            return None

# Save the final output file
if __name__ == "__main__":
    output_file = save_output_file()
    if output_file:
        logger.info("Financial data extraction completed successfully")
    else:
        logger.error("Failed to save output file")
