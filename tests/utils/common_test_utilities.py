"""
Common Test Utilities
====================

Centralized test utilities to resolve cross-test dependencies and provide
shared functionality for all test modules.

This module consolidates:
- Company discovery functions
- Date extraction utilities  
- Excel processing helpers
- Common test data operations

Usage:
------
    from tests.utils.common_test_utilities import (
        discover_companies,
        test_company_data_ordering,
        extract_period_end_dates,
        parse_date_year
    )
"""

import os
import glob
import sys
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import logging
from typing import Dict, Any, List, Optional

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def discover_companies(base_path=None):
    """
    Dynamically discover all available companies in the dataset.
    
    Args:
        base_path: Optional base directory path. If None, uses current working directory.
        
    Returns:
        List of company folder names that contain both FY and LTM subdirectories.
    """
    if base_path is None:
        base_path = os.getcwd()
    
    # Check if we're in tests directory and need to go up to root
    current_path = Path(base_path).resolve()
    if 'tests' in current_path.parts:
        # Navigate up to project root
        while current_path.name != 'financial_to_exel' and current_path.parent != current_path:
            current_path = current_path.parent
        base_path = str(current_path)
    
    company_folders = []
    
    # Look for folders that contain both FY and LTM subdirectories
    for item in os.listdir(base_path):
        item_path = os.path.join(base_path, item)
        if os.path.isdir(item_path) and not item.startswith('.') and not item.startswith('_'):
            fy_path = os.path.join(item_path, 'FY')
            ltm_path = os.path.join(item_path, 'LTM')
            if os.path.exists(fy_path) and os.path.exists(ltm_path):
                company_folders.append(item)
    
    return sorted(company_folders)


def test_company_data_ordering(company_symbol, expected_pattern=None, base_path=None):
    """
    Test data ordering for a specific company.
    
    Args:
        company_symbol: Company ticker symbol to test
        expected_pattern: Optional expected date pattern for validation
        base_path: Optional base directory path
        
    Returns:
        bool: True if test passes, False otherwise
    """
    if base_path is None:
        base_path = os.getcwd()
    
    # Check if we're in tests directory and need to go up to root  
    current_path = Path(base_path).resolve()
    if 'tests' in current_path.parts:
        while current_path.name != 'financial_to_exel' and current_path.parent != current_path:
            current_path = current_path.parent
        base_path = str(current_path)
    
    print(f"Testing data ordering for {company_symbol}...")
    
    try:
        # Check FY Income Statement
        fy_path = os.path.join(base_path, company_symbol, "FY", f"{company_symbol}_income_statement.xlsx")
        if not os.path.exists(fy_path):
            fy_path = os.path.join(base_path, company_symbol, "FY", "Income Statement.xlsx")
        
        if os.path.exists(fy_path):
            workbook = load_workbook(fy_path)
            dates = extract_period_end_dates(workbook)
            workbook.close()
            
            if dates:
                print(f"  ✅ Found {len(dates)} dates in FY Income Statement: {dates[:3]}...")
                return True
            else:
                print(f"  ❌ No dates found in FY Income Statement")
                return False
        else:
            print(f"  ❌ FY Income Statement not found for {company_symbol}")
            return False
            
    except Exception as e:
        print(f"  ❌ Error testing {company_symbol}: {str(e)}")
        return False


def extract_period_end_dates(workbook):
    """
    Extract Period End Date values from financial statement
    
    Args:
        workbook: Excel workbook containing financial data
        
    Returns:
        list: List of date strings extracted from Period End Date row
    """
    try:
        # Get the active sheet from the workbook
        sheet = workbook.active
        
        # Period End Date is always at row 10, column 3 according to analysis
        period_end_row = 10
        label_column = 3
        
        # Check if Period End Date exists at expected location
        if sheet.cell(period_end_row, label_column).value != "Period End Date":
            logger.warning("Period End Date not found at expected location (row 10, column 3)")
            return []
        
        # Extract dates starting from column 4
        dates = []
        for col in range(4, 16):  # Columns 4-15 (up to 12 periods)
            cell_value = sheet.cell(period_end_row, col).value
            if cell_value is not None:
                dates.append(str(cell_value))
            else:
                break  # Stop when we hit empty cells
        
        logger.info(f"Extracted {len(dates)} period end dates: {dates}")
        return dates
        
    except Exception as e:
        logger.error(f"Error extracting period end dates: {str(e)}")
        return []


def parse_date_year(date_string):
    """
    Parse year from date string in YYYY-MM-DD format
    
    Args:
        date_string: Date string in format "YYYY-MM-DD"
        
    Returns:
        int: Year as integer, or None if parsing fails
    """
    try:
        if isinstance(date_string, str) and len(date_string) >= 4:
            # Extract year from YYYY-MM-DD format
            year_str = date_string[:4]
            return int(year_str)
    except (ValueError, TypeError):
        logger.warning(f"Could not parse year from date: {date_string}")
    return None


def get_project_root():
    """
    Get the project root directory from any location within the project.
    
    Returns:
        Path: Path object pointing to project root
    """
    current_path = Path(__file__).resolve()
    
    # Navigate up until we find the project root
    while current_path.name != 'financial_to_exel' and current_path.parent != current_path:
        current_path = current_path.parent
        
    if current_path.name == 'financial_to_exel':
        return current_path
    else:
        # Fallback to current working directory
        return Path.cwd()


def discover_company_excel_files(company_symbol, base_path=None):
    """
    Discover all Excel files for a specific company.
    
    Args:
        company_symbol: Company ticker symbol
        base_path: Optional base directory path
        
    Returns:
        dict: Dictionary with file types and their paths
    """
    if base_path is None:
        base_path = get_project_root()
    
    files = {}
    company_path = Path(base_path) / company_symbol
    
    if not company_path.exists():
        return files
    
    # Check FY directory
    fy_path = company_path / "FY"
    if fy_path.exists():
        files['fy'] = {}
        for file_path in fy_path.glob("*.xlsx"):
            if not file_path.name.startswith('~'):  # Skip temp files
                file_type = file_path.stem.lower().replace(' ', '_')
                files['fy'][file_type] = str(file_path)
    
    # Check LTM directory  
    ltm_path = company_path / "LTM"
    if ltm_path.exists():
        files['ltm'] = {}
        for file_path in ltm_path.glob("*.xlsx"):
            if not file_path.name.startswith('~'):  # Skip temp files
                file_type = file_path.stem.lower().replace(' ', '_')
                files['ltm'][file_type] = str(file_path)
    
    return files


def validate_excel_structure(excel_path, expected_sheets=None):
    """
    Validate Excel file structure and content.
    
    Args:
        excel_path: Path to Excel file
        expected_sheets: Optional list of expected sheet names
        
    Returns:
        dict: Validation results with status and details
    """
    result = {
        'valid': False,
        'sheets': [],
        'errors': [],
        'warnings': []
    }
    
    try:
        workbook = load_workbook(excel_path)
        result['sheets'] = workbook.sheetnames
        
        # Check expected sheets if provided
        if expected_sheets:
            missing_sheets = set(expected_sheets) - set(workbook.sheetnames)
            if missing_sheets:
                result['errors'].append(f"Missing expected sheets: {missing_sheets}")
        
        # Basic structure validation
        sheet = workbook.active
        if sheet.max_row < 10:
            result['warnings'].append("Sheet has fewer than 10 rows")
        
        if sheet.max_column < 4:
            result['warnings'].append("Sheet has fewer than 4 columns")
        
        # Check for Period End Date row
        if sheet.cell(10, 3).value == "Period End Date":
            result['valid'] = True
        else:
            result['warnings'].append("Period End Date not found at expected location (row 10, col 3)")
            result['valid'] = True  # Still valid, just a warning
        
        workbook.close()
        
    except Exception as e:
        result['errors'].append(f"Error validating Excel file: {str(e)}")
    
    return result


def create_test_excel_structure(test_dir, company_symbol="TEST"):
    """
    Create test Excel structure for integration testing.

    Args:
        test_dir: Directory to create test structure in
        company_symbol: Company symbol to use for test data

    Returns:
        dict: Paths to created test files
    """
    from openpyxl import Workbook

    try:
        # Create company directory structure
        company_dir = Path(test_dir) / "data" / "companies" / company_symbol
        fy_dir = company_dir / "FY"
        ltm_dir = company_dir / "LTM"

        # Create directories
        fy_dir.mkdir(parents=True, exist_ok=True)
        ltm_dir.mkdir(parents=True, exist_ok=True)

        # Standard Excel files to create
        excel_files = {
            "Income Statement.xlsx": {
                "sheet_name": "Income Statement",
                "headers": ["Item", "2021", "2022", "2023"],
                "data": [
                    ["Revenue", 100000, 110000, 120000],
                    ["Cost of Revenue", 60000, 65000, 70000],
                    ["Gross Profit", 40000, 45000, 50000],
                    ["Operating Expenses", 25000, 27000, 29000],
                    ["Operating Income", 15000, 18000, 21000],
                    ["Net Income", 12000, 14500, 17000],
                    ["Earnings Per Share", 1.20, 1.45, 1.70]
                ]
            },
            "Balance Sheet.xlsx": {
                "sheet_name": "Balance Sheet",
                "headers": ["Item", "2021", "2022", "2023"],
                "data": [
                    ["Total Assets", 200000, 220000, 250000],
                    ["Current Assets", 80000, 90000, 100000],
                    ["Cash and Cash Equivalents", 30000, 35000, 40000],
                    ["Total Liabilities", 120000, 130000, 140000],
                    ["Total Equity", 80000, 90000, 110000],
                    ["Shareholders Equity", 80000, 90000, 110000],
                    ["Book Value", 80000, 90000, 110000]
                ]
            },
            "Cash Flow Statement.xlsx": {
                "sheet_name": "Cash Flow Statement",
                "headers": ["Item", "2021", "2022", "2023"],
                "data": [
                    ["Operating Cash Flow", 18000, 20000, 23000],
                    ["Capital Expenditures", 8000, 9000, 10000],
                    ["Free Cash Flow", 10000, 11000, 13000],
                    ["Dividends Paid", 2000, 2500, 3000],
                    ["Financing Cash Flow", -5000, -6000, -7000],
                    ["Investing Cash Flow", -8000, -9000, -10000]
                ]
            }
        }

        created_files = {"FY": {}, "LTM": {}}

        # Create files in both FY and LTM directories
        for file_name, file_config in excel_files.items():
            for period_type, period_dir in [("FY", fy_dir), ("LTM", ltm_dir)]:
                file_path = period_dir / file_name

                # Create Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = file_config["sheet_name"]

                # Add headers
                for col, header in enumerate(file_config["headers"], 1):
                    ws.cell(row=1, column=col, value=header)

                # Add data
                for row_idx, row_data in enumerate(file_config["data"], 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                # Save workbook
                wb.save(file_path)
                created_files[period_type][file_name] = str(file_path)

                logger.info(f"Created test Excel file: {file_path}")

        return {
            'success': True,
            'company_symbol': company_symbol,
            'company_dir': str(company_dir),
            'fy_dir': str(fy_dir),
            'ltm_dir': str(ltm_dir),
            'files': created_files
        }

    except Exception as e:
        logger.error(f"Error creating test Excel structure: {e}")
        return {
            'success': False,
            'error': str(e),
            'company_symbol': company_symbol,
            'files': {}
        }


def get_test_financial_data(ticker: str = "MSFT") -> Dict[str, Any]:
    """
    Get test financial data for testing purposes.

    Args:
        ticker: Stock ticker symbol

    Returns:
        Dictionary containing test financial data
    """
    return {
        'ticker': ticker,
        'revenue': 198270000000,  # $198.27B
        'net_income': 72361000000,  # $72.36B
        'total_cash_flow_from_operating_activities': 89035000000,  # $89.04B
        'capital_expenditures': -28107000000,  # $28.11B
        'free_cash_flow': 60928000000,  # $60.93B
        'shares_outstanding': 7430000000,  # 7.43B shares
        'market_cap': 3040000000000,  # $3.04T
        'current_price': 409.12,
        'book_value': 206192000000,  # $206.19B
        'total_equity': 206192000000,
        'dividend_per_share': 2.72,
        'last_updated': datetime.now().isoformat()
    }


def get_test_companies() -> List[str]:
    """
    Get list of test companies for testing purposes.

    Returns:
        List of test company ticker symbols
    """
    return ["AAPL", "MSFT", "GOOGL", "AMZN", "TSLA", "META", "NVDA", "NFLX"]


def create_mock_financial_data(ticker: str = "MSFT") -> Dict[str, Any]:
    """
    Create mock financial data for testing purposes.

    Args:
        ticker: Stock ticker symbol

    Returns:
        Dictionary containing mock financial data
    """
    multiplier_map = {
        "AAPL": 1.2,
        "MSFT": 1.0,
        "GOOGL": 0.9,
        "AMZN": 1.1,
        "TSLA": 0.8,
        "META": 0.7,
        "NVDA": 1.5,
        "NFLX": 0.6
    }

    multiplier = multiplier_map.get(ticker, 1.0)

    return {
        'ticker': ticker,
        'revenue': int(200000000000 * multiplier),
        'net_income': int(50000000000 * multiplier),
        'total_cash_flow_from_operating_activities': int(65000000000 * multiplier),
        'capital_expenditures': int(-15000000000 * multiplier),
        'free_cash_flow': int(50000000000 * multiplier),
        'shares_outstanding': int(7500000000 * multiplier),
        'market_cap': int(2500000000000 * multiplier),
        'current_price': 333.33 * multiplier,
        'book_value': int(180000000000 * multiplier),
        'total_equity': int(180000000000 * multiplier),
        'dividend_per_share': 2.5 * multiplier,
        'sector': 'Technology',
        'industry': 'Software',
        'last_updated': datetime.now().isoformat()
    }


class TestDataGenerator:
    """Test data generator for financial test scenarios"""

    @staticmethod
    def generate_test_company_data(ticker: str = "TEST") -> Dict[str, Any]:
        """Generate comprehensive test company data"""
        return {
            'ticker': ticker,
            'company_name': f"{ticker} Corporation",
            'revenue': 100000000000,
            'net_income': 25000000000,
            'total_cash_flow_from_operating_activities': 30000000000,
            'capital_expenditures': -5000000000,
            'free_cash_flow': 25000000000,
            'shares_outstanding': 5000000000,
            'market_cap': 500000000000,
            'current_price': 100.0,
            'book_value': 150000000000,
            'total_equity': 150000000000,
            'dividend_per_share': 2.0,
            'sector': 'Technology',
            'industry': 'Software',
            'last_updated': datetime.now().isoformat()
        }

    @staticmethod
    def generate_test_market_data(ticker: str = "TEST") -> Dict[str, Any]:
        """Generate test market data"""
        return {
            'ticker': ticker,
            'current_price': 100.0,
            'market_cap': 500000000000,
            'shares_outstanding': 5000000000,
            'volume': 1000000,
            'avg_volume': 1200000,
            'pe_ratio': 20.0,
            'eps': 5.0,
            'dividend_yield': 0.02,
            'beta': 1.1,
            '52_week_high': 120.0,
            '52_week_low': 80.0
        }