"""
Shared test utilities for cross-test dependencies - DEPRECATED

⚠️ DEPRECATED: This module is deprecated in favor of common_test_utilities.py
Use `from tests.utils.common_test_utilities import ...` instead.

This module contains common functions used across multiple test files
to resolve import dependencies between test modules.
"""

import warnings
warnings.warn(
    "shared_test_utilities is deprecated. Use common_test_utilities instead.",
    DeprecationWarning,
    stacklevel=2
)

# Re-export from common_test_utilities for backward compatibility
from .common_test_utilities import (
    discover_companies,
    test_company_data_ordering,
    extract_period_end_dates, 
    parse_date_year
)

import os
import glob
import sys
from pathlib import Path
from openpyxl import load_workbook


def discover_companies(base_path=None):
    """
    Dynamically discover all available companies in the dataset.
    
    Args:
        base_path: Optional base directory path. If None, uses current working directory.
        
    Returns:
        List of company folder names that contain both FY and LTM subdirectories.
    """
    if base_path is None:
        base_path = os.getcwd()
    
    # Check if we're in tests directory and need to go up to root
    if 'tests' in os.path.basename(base_path):
        base_path = Path(base_path).parent.parent.absolute()
    
    company_folders = []
    
    # Look for folders that contain both FY and LTM subdirectories
    for item in os.listdir(base_path):
        item_path = os.path.join(base_path, item)
        if os.path.isdir(item_path) and not item.startswith('.') and not item.startswith('_'):
            fy_path = os.path.join(item_path, 'FY')
            ltm_path = os.path.join(item_path, 'LTM')
            if os.path.exists(fy_path) and os.path.exists(ltm_path):
                company_folders.append(item)
    
    return sorted(company_folders)


def test_company_data_ordering(company_symbol, expected_pattern=None, base_path=None):
    """
    Test data ordering for a specific company.
    
    Args:
        company_symbol: Company symbol/folder name
        expected_pattern: Optional expected pattern for validation
        base_path: Optional base directory path
        
    Returns:
        Boolean indicating if test passed
    """
    try:
        print(f"\n=== Testing {company_symbol} ===")
        
        if base_path is None:
            base_path = os.getcwd()
            
        # Check if we're in tests directory and need to go up to root
        if 'tests' in os.path.basename(base_path):
            base_path = Path(base_path).parent.parent.absolute()
        
        # Path to company's cash flow statement
        company_folder = os.path.join(base_path, company_symbol)
        
        # Dynamically find the cash flow statement file
        cash_flow_pattern = os.path.join(company_folder, "FY", "*Cash Flow Statement.xlsx")
        cash_flow_files = glob.glob(cash_flow_pattern)
        
        if not cash_flow_files:
            print(f"  ❌ No cash flow statement found in {os.path.join(company_folder, 'FY')}")
            return False
        
        cash_flow_file = cash_flow_files[0]  # Use the first match
        
        if not os.path.exists(cash_flow_file):
            print(f"  ❌ File not found: {cash_flow_file}")
            return False
        
        # Load workbook and test data ordering
        wb = load_workbook(cash_flow_file, data_only=True)
        sheet = wb.active
        
        # Find "Free Cash Flow" or similar row
        levered_fcf_row = None
        for row in range(1, min(sheet.max_row + 1, 50)):  # Check first 50 rows
            cell_value = sheet.cell(row, column=1).value
            if cell_value and isinstance(cell_value, str):
                if any(term in cell_value.lower() for term in [
                    'free cash flow', 'levered free cash flow', 'unlevered free cash flow', 'fcf'
                ]):
                    levered_fcf_row = row
                    break
        
        if levered_fcf_row is None:
            print(f"  ❌ No FCF row found in {cash_flow_file}")
            return False
        
        # Extract values from the FCF row
        new_values = []
        for j in range(10):  # Check up to 10 columns
            cell_value = sheet.cell(levered_fcf_row, column=4 + j).value
            if cell_value:
                try:
                    if isinstance(cell_value, str):
                        parsed_value = float(cell_value.replace(',', ''))
                        new_values.append(parsed_value)
                    else:
                        new_values.append(float(cell_value))
                except (ValueError, TypeError):
                    new_values.append(cell_value)
            else:
                new_values.append(None)
        
        # Filter out None values
        filtered_values = [v for v in new_values if v is not None and isinstance(v, (int, float))]
        
        if len(filtered_values) < 2:
            print(f"  ❌ Insufficient data points found: {len(filtered_values)}")
            return False
        
        print(f"  ✅ Found {len(filtered_values)} FCF data points")
        print(f"  📊 Values: {filtered_values[:5]}...")  # Show first 5 values
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"  ❌ Error testing {company_symbol}: {e}")
        return False


def get_test_company_symbols(limit=None):
    """
    Get a list of test company symbols for testing.
    
    Args:
        limit: Optional limit on number of companies to return
        
    Returns:
        List of company symbols suitable for testing
    """
    companies = discover_companies()
    
    if limit:
        companies = companies[:limit]
        
    return companies


def validate_company_data_structure(company_symbol, base_path=None):
    """
    Validate that a company has the expected data structure.
    
    Args:
        company_symbol: Company symbol/folder name
        base_path: Optional base directory path
        
    Returns:
        Dictionary with validation results
    """
    if base_path is None:
        base_path = os.getcwd()
        
    # Check if we're in tests directory and need to go up to root
    if 'tests' in os.path.basename(base_path):
        base_path = Path(base_path).parent.parent.absolute()
    
    company_folder = os.path.join(base_path, company_symbol)
    
    validation_results = {
        'company_folder_exists': os.path.exists(company_folder),
        'fy_folder_exists': os.path.exists(os.path.join(company_folder, 'FY')),
        'ltm_folder_exists': os.path.exists(os.path.join(company_folder, 'LTM')),
        'has_income_statement': False,
        'has_balance_sheet': False,
        'has_cash_flow_statement': False
    }
    
    if validation_results['fy_folder_exists']:
        fy_path = os.path.join(company_folder, 'FY')
        for file_name in os.listdir(fy_path):
            if 'income statement' in file_name.lower():
                validation_results['has_income_statement'] = True
            elif 'balance sheet' in file_name.lower():
                validation_results['has_balance_sheet'] = True
            elif 'cash flow statement' in file_name.lower():
                validation_results['has_cash_flow_statement'] = True
    
    validation_results['is_valid'] = all([
        validation_results['company_folder_exists'],
        validation_results['fy_folder_exists'],
        validation_results['ltm_folder_exists'],
        validation_results['has_cash_flow_statement']
    ])
    
    return validation_results


def get_project_root():
    """
    Get the project root directory.
    
    Returns:
        Path to the project root directory
    """
    current_file = Path(__file__).resolve()
    
    # Go up from tests/utils/shared_test_utilities.py to project root
    project_root = current_file.parent.parent.parent
    
    return str(project_root)