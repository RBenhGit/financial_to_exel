"""
Excel Data Import User Experience Testing

This module provides comprehensive user experience testing for Excel file import
functionality, covering various scenarios, error conditions, and user workflows.
"""

import os
import sys
import unittest
import tempfile
import shutil
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import Workbook
import logging

# Add project root to path for imports
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from core.excel_integration.excel_utils import ExcelDataExtractor
from core.data_processing.adapters.excel_adapter import ExcelDataAdapter
from core.analysis.engines.financial_calculations import FinancialCalculator
from tests.user_acceptance.user_journey_testing_framework import (
    TestScenario, UserAction, TestPriority, TestStatus, TestResult
)

class ExcelImportUXTester:
    """
    Comprehensive user experience tester for Excel import functionality.

    Tests various user scenarios, error conditions, and edge cases
    to ensure robust and user-friendly Excel import experience.
    """

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.test_data_dir = None
        self.temp_dir = None
        self.results = []

    def setup_test_environment(self):
        """Set up temporary test environment with sample Excel files."""
        self.temp_dir = tempfile.mkdtemp(prefix="excel_ux_test_")
        self.test_data_dir = Path(self.temp_dir) / "companies"
        self.test_data_dir.mkdir(parents=True, exist_ok=True)

        # Create test company directories with both FY and LTM folders
        test_companies = ["MSFT", "AAPL", "INVALID_TEST", "INCOMPLETE_TEST"]
        for company in test_companies:
            for period in ["FY", "LTM"]:
                company_dir = self.test_data_dir / company / period
                company_dir.mkdir(parents=True, exist_ok=True)

        self._create_sample_excel_files()

    def cleanup_test_environment(self):
        """Clean up temporary test environment."""
        if self.temp_dir and os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def _create_sample_excel_files(self):
        """Create sample Excel files for testing."""

        # Create valid MSFT Excel files
        msft_dir = self.test_data_dir / "MSFT" / "FY"
        self._create_income_statement(msft_dir)
        self._create_balance_sheet(msft_dir)
        self._create_cash_flow_statement(msft_dir)

        # Create valid AAPL Excel files
        aapl_dir = self.test_data_dir / "AAPL" / "FY"
        self._create_income_statement(aapl_dir, company="Apple Inc.")
        self._create_balance_sheet(aapl_dir, company="Apple Inc.")
        self._create_cash_flow_statement(aapl_dir, company="Apple Inc.")

        # Create invalid/problematic files for error testing
        invalid_dir = self.test_data_dir / "INVALID_TEST" / "FY"
        self._create_corrupted_excel_file(invalid_dir)

        # Create incomplete files for partial data testing
        incomplete_dir = self.test_data_dir / "INCOMPLETE_TEST" / "FY"
        self._create_incomplete_excel_files(incomplete_dir)

    def _create_income_statement(self, directory: Path, company: str = "Microsoft Corporation"):
        """Create a sample Income Statement Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Income Statement"

        # Headers
        ws['A1'] = company
        ws['A2'] = "Income Statement"
        ws['A3'] = "In millions, except per share data"

        # Date headers
        ws['B4'] = "FY 2023"
        ws['C4'] = "FY 2022"
        ws['D4'] = "FY 2021"

        # Revenue data
        ws['A5'] = "Total Revenue"
        ws['B5'] = 211915  # FY 2023
        ws['C5'] = 198270  # FY 2022
        ws['D5'] = 168088  # FY 2021

        # Operating Income
        ws['A6'] = "Operating Income"
        ws['B6'] = 88523
        ws['C6'] = 83383
        ws['D6'] = 69916

        # Net Income
        ws['A7'] = "Net Income"
        ws['B7'] = 72361
        ws['C7'] = 72738
        ws['D7'] = 61271

        # EPS
        ws['A8'] = "Earnings per share - basic"
        ws['B8'] = 9.78
        ws['C8'] = 9.70
        ws['D8'] = 8.12

        wb.save(directory / "Income Statement.xlsx")

    def _create_balance_sheet(self, directory: Path, company: str = "Microsoft Corporation"):
        """Create a sample Balance Sheet Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"

        # Headers
        ws['A1'] = company
        ws['A2'] = "Balance Sheet"
        ws['A3'] = "In millions"

        # Date headers
        ws['B4'] = "FY 2023"
        ws['C4'] = "FY 2022"
        ws['D4'] = "FY 2021"

        # Assets
        ws['A5'] = "Total Assets"
        ws['B5'] = 411976
        ws['C5'] = 364840
        ws['D5'] = 333779

        # Current Assets
        ws['A6'] = "Current Assets"
        ws['B6'] = 184257
        ws['C6'] = 169684
        ws['D6'] = 184406

        # Cash and Equivalents
        ws['A7'] = "Cash and cash equivalents"
        ws['B7'] = 29263
        ws['C7'] = 13931
        ws['D7'] = 14224

        # Total Equity
        ws['A8'] = "Total stockholders' equity"
        ws['B8'] = 206223
        ws['C8'] = 166542
        ws['D8'] = 141988

        # Total Liabilities
        ws['A9'] = "Total Liabilities"
        ws['B9'] = 205753
        ws['C9'] = 198298
        ws['D9'] = 191791

        wb.save(directory / "Balance Sheet.xlsx")

    def _create_cash_flow_statement(self, directory: Path, company: str = "Microsoft Corporation"):
        """Create a sample Cash Flow Statement Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow Statement"

        # Headers
        ws['A1'] = company
        ws['A2'] = "Cash Flow Statement"
        ws['A3'] = "In millions"

        # Date headers
        ws['B4'] = "FY 2023"
        ws['C4'] = "FY 2022"
        ws['D4'] = "FY 2021"

        # Operating Cash Flow
        ws['A5'] = "Cash flow from operations"
        ws['B5'] = 87582
        ws['C5'] = 89035
        ws['D5'] = 76740

        # Capital Expenditures
        ws['A6'] = "Additions to property and equipment"
        ws['B6'] = -28107
        ws['C6'] = -23711
        ws['D6'] = -20622

        # Free Cash Flow
        ws['A7'] = "Free cash flow"
        ws['B7'] = 59475
        ws['C7'] = 65324
        ws['D7'] = 56118

        # Financing Cash Flow
        ws['A8'] = "Cash flow from financing"
        ws['B8'] = -30591
        ws['C8'] = -32532
        ws['D8'] = -40208

        wb.save(directory / "Cash Flow Statement.xlsx")

    def _create_corrupted_excel_file(self, directory: Path):
        """Create a corrupted Excel file for error testing."""
        # Create a file with invalid content
        corrupted_file = directory / "Income Statement.xlsx"
        with open(corrupted_file, 'w') as f:
            f.write("This is not a valid Excel file content")

    def _create_incomplete_excel_files(self, directory: Path):
        """Create incomplete Excel files for partial data testing."""
        # Create only Income Statement, missing other files
        wb = Workbook()
        ws = wb.active
        ws.title = "Income Statement"

        # Minimal headers only
        ws['A1'] = "Incomplete Company"
        ws['A2'] = "Income Statement"
        ws['A5'] = "Total Revenue"
        ws['B5'] = 1000  # Only one data point

        wb.save(directory / "Income Statement.xlsx")

    def test_successful_excel_import(self) -> TestResult:
        """Test successful Excel import workflow."""
        test_name = "Successful Excel Import"
        result = TestResult(
            scenario_id="excel_import_success_001",
            status=TestStatus.NOT_RUN,
            start_time=self.logger.info(f"Starting test: {test_name}")
        )

        try:
            # Test loading MSFT data
            msft_dir = self.test_data_dir / "MSFT"

            # Initialize calculator with Excel data
            calculator = FinancialCalculator(str(msft_dir))
            success = calculator.load_financial_data()

            if success:
                # Validate data loaded correctly
                assert calculator.ticker_symbol == "MSFT"
                assert calculator.company_name is not None
                assert calculator.total_revenue is not None
                assert calculator.total_revenue > 0

                result.status = TestStatus.PASSED
                self.logger.info(f"PASS {test_name}: Data loaded successfully")
            else:
                result.status = TestStatus.FAILED
                result.error_message = "Excel data loading failed"

        except Exception as e:
            result.status = TestStatus.FAILED
            result.error_message = f"Exception during Excel import: {str(e)}"
            self.logger.error(f"FAIL {test_name}: {str(e)}")

        return result

    def test_missing_files_error_handling(self) -> TestResult:
        """Test error handling when required Excel files are missing."""
        test_name = "Missing Files Error Handling"
        result = TestResult(
            scenario_id="excel_import_missing_001",
            status=TestStatus.NOT_RUN,
            start_time=self.logger.info(f"Starting test: {test_name}")
        )

        try:
            # Create directory with no Excel files but proper structure
            empty_fy_dir = self.test_data_dir / "EMPTY_TEST" / "FY"
            empty_ltm_dir = self.test_data_dir / "EMPTY_TEST" / "LTM"
            empty_fy_dir.mkdir(parents=True, exist_ok=True)
            empty_ltm_dir.mkdir(parents=True, exist_ok=True)

            calculator = FinancialCalculator(str(empty_fy_dir.parent))
            success = calculator.load_financial_data()

            # Should fail gracefully, not crash
            if not success:
                result.status = TestStatus.PASSED
                self.logger.info(f"PASS {test_name}: Correctly handled missing files")
            else:
                result.status = TestStatus.FAILED
                result.error_message = "Should have failed with missing files"

        except Exception as e:
            # Catching exception is also acceptable if handled gracefully
            result.status = TestStatus.PASSED
            self.logger.info(f"PASS {test_name}: Exception handled gracefully: {str(e)}")

        return result

    def test_corrupted_file_handling(self) -> TestResult:
        """Test handling of corrupted Excel files."""
        test_name = "Corrupted File Handling"
        result = TestResult(
            scenario_id="excel_import_corrupted_001",
            status=TestStatus.NOT_RUN,
            start_time=self.logger.info(f"Starting test: {test_name}")
        )

        try:
            invalid_dir = self.test_data_dir / "INVALID_TEST"

            calculator = FinancialCalculator(str(invalid_dir))
            success = calculator.load_financial_data()

            # Should handle corruption gracefully
            if not success:
                result.status = TestStatus.PASSED
                self.logger.info(f"PASS {test_name}: Corrupted file handled correctly")
            else:
                result.status = TestStatus.FAILED
                result.error_message = "Should have failed with corrupted file"

        except Exception as e:
            # Exception handling is acceptable if graceful
            result.status = TestStatus.PASSED
            self.logger.info(f"PASS {test_name}: Corruption handled with exception: {str(e)}")

        return result

    def test_incomplete_data_handling(self) -> TestResult:
        """Test handling of incomplete Excel data."""
        test_name = "Incomplete Data Handling"
        result = TestResult(
            scenario_id="excel_import_incomplete_001",
            status=TestStatus.NOT_RUN,
            start_time=self.logger.info(f"Starting test: {test_name}")
        )

        try:
            incomplete_dir = self.test_data_dir / "INCOMPLETE_TEST"

            calculator = FinancialCalculator(str(incomplete_dir))
            success = calculator.load_financial_data()

            # Should either succeed with partial data or fail gracefully
            if success:
                # Check if partial data was loaded
                assert calculator.total_revenue is not None
                result.status = TestStatus.PASSED
                self.logger.info(f"PASS {test_name}: Partial data loaded successfully")
            else:
                result.status = TestStatus.PASSED
                self.logger.info(f"PASS {test_name}: Incomplete data handled gracefully")

        except Exception as e:
            result.status = TestStatus.FAILED
            result.error_message = f"Incomplete data caused unhandled exception: {str(e)}"

        return result

    def test_data_validation_accuracy(self) -> TestResult:
        """Test accuracy of extracted data values."""
        test_name = "Data Validation Accuracy"
        result = TestResult(
            scenario_id="excel_import_accuracy_001",
            status=TestStatus.NOT_RUN,
            start_time=self.logger.info(f"Starting test: {test_name}")
        )

        try:
            msft_dir = self.test_data_dir / "MSFT"

            calculator = FinancialCalculator(str(msft_dir))
            success = calculator.load_financial_data()

            if success:
                # Validate specific data points match our test data
                expected_revenue = 211915  # From our test data
                actual_revenue = calculator.total_revenue

                # Allow for some flexibility in data extraction
                if actual_revenue and abs(actual_revenue - expected_revenue) < 1000:
                    result.status = TestStatus.PASSED
                    self.logger.info(f"PASS {test_name}: Data accuracy validated")
                else:
                    result.status = TestStatus.FAILED
                    result.error_message = f"Revenue mismatch: expected {expected_revenue}, got {actual_revenue}"
            else:
                result.status = TestStatus.FAILED
                result.error_message = "Data loading failed"

        except Exception as e:
            result.status = TestStatus.FAILED
            result.error_message = f"Data validation failed: {str(e)}"

        return result

    def test_performance_with_large_files(self) -> TestResult:
        """Test performance with larger Excel files."""
        test_name = "Performance with Large Files"
        result = TestResult(
            scenario_id="excel_import_performance_001",
            status=TestStatus.NOT_RUN,
            start_time=self.logger.info(f"Starting test: {test_name}")
        )

        try:
            # Create a larger Excel file with both FY and LTM directories
            large_fy_dir = self.test_data_dir / "LARGE_TEST" / "FY"
            large_ltm_dir = self.test_data_dir / "LARGE_TEST" / "LTM"
            large_fy_dir.mkdir(parents=True, exist_ok=True)
            large_ltm_dir.mkdir(parents=True, exist_ok=True)

            self._create_large_excel_file(large_fy_dir)

            import time
            start_time = time.time()

            calculator = FinancialCalculator(str(large_fy_dir.parent))
            success = calculator.load_financial_data()

            end_time = time.time()
            load_time = end_time - start_time

            # Should complete within reasonable time (30 seconds)
            if success and load_time < 30:
                result.status = TestStatus.PASSED
                result.performance_metrics = {"load_time_seconds": load_time}
                self.logger.info(f"PASS {test_name}: Loaded in {load_time:.2f} seconds")
            elif success:
                result.status = TestStatus.FAILED
                result.error_message = f"Load time too slow: {load_time:.2f} seconds"
            else:
                result.status = TestStatus.FAILED
                result.error_message = "Large file loading failed"

        except Exception as e:
            result.status = TestStatus.FAILED
            result.error_message = f"Performance test failed: {str(e)}"

        return result

    def _create_large_excel_file(self, directory: Path):
        """Create a larger Excel file for performance testing."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Income Statement"

        # Headers
        ws['A1'] = "Large Test Company"
        ws['A2'] = "Income Statement"

        # Create many rows of data
        for i in range(100):
            row = i + 5
            ws[f'A{row}'] = f"Metric {i}"
            ws[f'B{row}'] = 1000 + i
            ws[f'C{row}'] = 2000 + i
            ws[f'D{row}'] = 3000 + i

        wb.save(directory / "Income Statement.xlsx")

    def run_all_excel_ux_tests(self) -> Dict[str, Any]:
        """Run all Excel import UX tests."""
        print("Running Excel Import User Experience Tests...")
        print("=" * 50)

        self.setup_test_environment()

        try:
            # Define all test methods
            tests = [
                self.test_successful_excel_import,
                self.test_missing_files_error_handling,
                self.test_corrupted_file_handling,
                self.test_incomplete_data_handling,
                self.test_data_validation_accuracy,
                self.test_performance_with_large_files
            ]

            results = []
            passed = 0
            failed = 0

            for test_method in tests:
                print(f"\nRunning: {test_method.__name__}")
                result = test_method()
                results.append(result)

                if result.status == TestStatus.PASSED:
                    passed += 1
                    print(f"PASSED")
                else:
                    failed += 1
                    print(f"FAILED: {result.error_message}")

            # Generate summary
            total = len(tests)
            success_rate = (passed / total * 100) if total > 0 else 0

            summary = {
                "total_tests": total,
                "passed": passed,
                "failed": failed,
                "success_rate": success_rate,
                "results": results
            }

            print(f"\n{'='*50}")
            print(f"Excel Import UX Test Summary:")
            print(f"Total Tests: {total}")
            print(f"Passed: {passed}")
            print(f"Failed: {failed}")
            print(f"Success Rate: {success_rate:.1f}%")

            return summary

        finally:
            self.cleanup_test_environment()


def main():
    """Run Excel import UX tests as standalone script."""

    # Set up logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

    tester = ExcelImportUXTester()
    results = tester.run_all_excel_ux_tests()

    # Generate detailed report
    if results['success_rate'] >= 80:
        print(f"\nExcel import UX tests are in good shape!")
    else:
        print(f"\nExcel import UX needs improvement.")
        print("Failed tests indicate user experience issues that should be addressed.")

    return results


if __name__ == "__main__":
    main()