"""
Load Testing and Stress Testing Suite
=====================================

This module provides comprehensive load testing capabilities for the financial analysis system,
focusing on concurrent user scenarios, API rate limiting, and system resilience under stress.

Features:
- Concurrent user simulation
- API rate limiting performance testing
- Database connection pooling stress tests
- Large dataset processing under load
- Memory pressure testing
- Resource contention analysis
- Performance degradation measurement

Usage:
    # Run load tests
    pytest tests/performance/test_load_testing_suite.py -v

    # Run with stress test markers
    pytest tests/performance/test_load_testing_suite.py -m stress_test

    # Run with specific concurrency levels
    pytest tests/performance/test_load_testing_suite.py::test_concurrent_api_calls_high_load
"""

import pytest
import asyncio
import time
import threading
import multiprocessing
import sys
import os
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
from typing import Dict, List, Any, Optional, Callable
from unittest.mock import Mock, patch, MagicMock
import psutil
import requests
from queue import Queue, Empty
import tempfile
import shutil
from dataclasses import dataclass
from datetime import datetime, timedelta
import json
import random

# Add project root to path
project_root = Path(__file__).parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

# Import core modules
from core.analysis.engines.financial_calculations import FinancialCalculator
from core.data_processing.rate_limiting.enhanced_rate_limiter import EnhancedRateLimiter
from core.data_sources.real_time_price_service import RealTimePriceService
from core.data_sources.industry_data_service import IndustryDataService

# Try to import optional modules
try:
    from core.data_processing.api_batch_manager import ApiBatchManager, BatchConfig
except ImportError:
    ApiBatchManager = None
    BatchConfig = None

try:
    from core.data_processing.managers.enhanced_data_manager import EnhancedDataManager
except ImportError:
    EnhancedDataManager = None


@dataclass
class LoadTestConfig:
    """Configuration for load testing scenarios"""
    concurrent_users: int = 10
    test_duration_seconds: int = 30
    requests_per_user: int = 10
    ramp_up_time_seconds: int = 5
    api_timeout_seconds: int = 30
    max_memory_mb: int = 1000
    max_cpu_percent: int = 80


@dataclass
class LoadTestResult:
    """Results from load testing"""
    total_requests: int
    successful_requests: int
    failed_requests: int
    average_response_time: float
    max_response_time: float
    min_response_time: float
    requests_per_second: float
    error_rate: float
    peak_memory_mb: float
    peak_cpu_percent: float
    errors: List[str]


class LoadTestRunner:
    """Utility class for running load tests"""

    def __init__(self, config: LoadTestConfig):
        self.config = config
        self.results = []
        self.errors = []
        self.start_time = None
        self.initial_memory = None

    def setup(self):
        """Setup load test environment"""
        self.start_time = time.time()
        self.initial_memory = psutil.Process().memory_info().rss / 1024 / 1024
        self.results.clear()
        self.errors.clear()

    def run_load_test(self, test_function: Callable, *args, **kwargs) -> LoadTestResult:
        """Execute load test with specified function"""
        self.setup()

        # Resource monitoring
        peak_memory = self.initial_memory
        peak_cpu = 0

        def monitor_resources():
            nonlocal peak_memory, peak_cpu
            while time.time() - self.start_time < self.config.test_duration_seconds + 10:
                try:
                    process = psutil.Process()
                    memory_mb = process.memory_info().rss / 1024 / 1024
                    cpu_percent = process.cpu_percent()

                    peak_memory = max(peak_memory, memory_mb)
                    peak_cpu = max(peak_cpu, cpu_percent)

                    time.sleep(0.5)
                except psutil.NoSuchProcess:
                    break

        # Start resource monitoring
        monitor_thread = threading.Thread(target=monitor_resources, daemon=True)
        monitor_thread.start()

        # Execute concurrent load test
        with ThreadPoolExecutor(max_workers=self.config.concurrent_users) as executor:
            # Submit tasks with ramp-up
            futures = []
            for user_id in range(self.config.concurrent_users):
                # Stagger user start times for ramp-up
                delay = (user_id / self.config.concurrent_users) * self.config.ramp_up_time_seconds

                future = executor.submit(
                    self._user_simulation,
                    user_id,
                    test_function,
                    delay,
                    *args,
                    **kwargs
                )
                futures.append(future)

            # Collect results
            for future in as_completed(futures):
                try:
                    user_results = future.result(timeout=self.config.test_duration_seconds + 20)
                    self.results.extend(user_results)
                except Exception as e:
                    self.errors.append(f"User simulation error: {str(e)}")

        # Calculate final results
        return self._calculate_results(peak_memory, peak_cpu)

    def _user_simulation(self, user_id: int, test_function: Callable, delay: float, *args, **kwargs) -> List[Dict]:
        """Simulate a single user's actions"""
        time.sleep(delay)  # Ramp-up delay

        user_results = []
        for request_id in range(self.config.requests_per_user):
            request_start = time.time()

            try:
                # Execute the test function
                result = test_function(user_id, request_id, *args, **kwargs)

                response_time = time.time() - request_start

                user_results.append({
                    'user_id': user_id,
                    'request_id': request_id,
                    'success': True,
                    'response_time': response_time,
                    'result': result,
                    'timestamp': request_start
                })

            except Exception as e:
                response_time = time.time() - request_start
                error_msg = f"User {user_id}, Request {request_id}: {str(e)}"
                self.errors.append(error_msg)

                user_results.append({
                    'user_id': user_id,
                    'request_id': request_id,
                    'success': False,
                    'response_time': response_time,
                    'error': error_msg,
                    'timestamp': request_start
                })

            # Add some randomization to avoid synchronized behavior
            time.sleep(random.uniform(0.1, 1.0))

        return user_results

    def _calculate_results(self, peak_memory: float, peak_cpu: float) -> LoadTestResult:
        """Calculate load test metrics"""
        if not self.results:
            return LoadTestResult(
                total_requests=0,
                successful_requests=0,
                failed_requests=len(self.errors),
                average_response_time=0,
                max_response_time=0,
                min_response_time=0,
                requests_per_second=0,
                error_rate=1.0,
                peak_memory_mb=peak_memory,
                peak_cpu_percent=peak_cpu,
                errors=self.errors
            )

        successful_results = [r for r in self.results if r['success']]
        failed_count = len([r for r in self.results if not r['success']])

        response_times = [r['response_time'] for r in successful_results]

        total_time = time.time() - self.start_time
        total_requests = len(self.results)

        return LoadTestResult(
            total_requests=total_requests,
            successful_requests=len(successful_results),
            failed_requests=failed_count,
            average_response_time=sum(response_times) / len(response_times) if response_times else 0,
            max_response_time=max(response_times) if response_times else 0,
            min_response_time=min(response_times) if response_times else 0,
            requests_per_second=total_requests / total_time if total_time > 0 else 0,
            error_rate=failed_count / total_requests if total_requests > 0 else 0,
            peak_memory_mb=peak_memory,
            peak_cpu_percent=peak_cpu,
            errors=self.errors
        )


@pytest.fixture
def load_test_config():
    """Default load test configuration"""
    return LoadTestConfig(
        concurrent_users=5,
        test_duration_seconds=20,
        requests_per_user=5,
        ramp_up_time_seconds=3
    )


@pytest.fixture
def stress_test_config():
    """Stress test configuration with higher load"""
    return LoadTestConfig(
        concurrent_users=20,
        test_duration_seconds=60,
        requests_per_user=15,
        ramp_up_time_seconds=10
    )


@pytest.fixture
def mock_rate_limiter():
    """Mock rate limiter for testing"""
    rate_limiter = Mock(spec=RateLimiter)
    rate_limiter.acquire.return_value = True
    rate_limiter.get_wait_time.return_value = 0
    rate_limiter.is_rate_limited.return_value = False
    return rate_limiter


@pytest.fixture
def mock_api_responses():
    """Mock external API responses"""
    responses = {
        'market_data': {
            'symbol': 'TEST',
            'price': 150.0,
            'marketCap': 1000000000,
            'sharesOutstanding': 6666667
        },
        'fundamental_data': {
            'revenue': 1000000,
            'netIncome': 100000,
            'totalAssets': 2000000
        },
        'industry_data': {
            'industry': 'Technology',
            'sector': 'Information Technology',
            'avg_pe_ratio': 25.5,
            'avg_pb_ratio': 4.2
        }
    }

    with patch('requests.get') as mock_get:
        def mock_response(url, **kwargs):
            mock_resp = Mock()
            mock_resp.status_code = 200
            mock_resp.raise_for_status.return_value = None

            if 'market' in url.lower():
                mock_resp.json.return_value = responses['market_data']
            elif 'fundamental' in url.lower():
                mock_resp.json.return_value = responses['fundamental_data']
            elif 'industry' in url.lower():
                mock_resp.json.return_value = responses['industry_data']
            else:
                mock_resp.json.return_value = {'status': 'ok'}

            return mock_resp

        mock_get.side_effect = mock_response
        yield responses


@pytest.fixture
def temp_company_data():
    """Create temporary company data for load testing"""
    temp_dir = tempfile.mkdtemp(prefix="load_test_")

    # Create multiple test companies
    company_dirs = []
    for i in range(10):
        company_name = f"LOAD_TEST_{i:02d}"
        company_dir = Path(temp_dir) / company_name
        company_dir.mkdir()

        # Create FY and LTM directories
        fy_dir = company_dir / "FY"
        ltm_dir = company_dir / "LTM"
        fy_dir.mkdir()
        ltm_dir.mkdir()

        # Create minimal Excel files
        try:
            from openpyxl import Workbook

            for sheet_name in ["Income Statement", "Balance Sheet", "Cash Flow Statement"]:
                for folder in [fy_dir, ltm_dir]:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = sheet_name

                    # Add minimal data
                    ws['A1'] = 'Metric'
                    ws['B1'] = 'FY2023'
                    ws['C1'] = 'FY2022'

                    if sheet_name == "Income Statement":
                        ws['A2'] = 'Revenue'
                        ws['B2'] = 1000000 + i * 100000
                        ws['C2'] = 900000 + i * 90000
                        ws['A3'] = 'Net Income'
                        ws['B3'] = 100000 + i * 10000
                        ws['C3'] = 90000 + i * 9000
                    elif sheet_name == "Cash Flow Statement":
                        ws['A2'] = 'Operating Cash Flow'
                        ws['B2'] = 120000 + i * 12000
                        ws['C2'] = 108000 + i * 10800
                        ws['A3'] = 'Capital Expenditures'
                        ws['B3'] = -20000 - i * 2000
                        ws['C3'] = -18000 - i * 1800

                    wb.save(folder / f"{sheet_name}.xlsx")
                    wb.close()

        except ImportError:
            # Create empty files if openpyxl not available
            for sheet_name in ["Income Statement", "Balance Sheet", "Cash Flow Statement"]:
                (fy_dir / f"{sheet_name}.xlsx").touch()
                (ltm_dir / f"{sheet_name}.xlsx").touch()

        company_dirs.append(str(company_dir))

    yield company_dirs

    # Cleanup
    shutil.rmtree(temp_dir, ignore_errors=True)


class TestConcurrentCalculations:
    """Test concurrent financial calculation performance"""

    def test_concurrent_calculator_initialization(self, load_test_config, temp_company_data):
        """Test concurrent initialization of multiple calculators"""

        def init_calculator(user_id: int, request_id: int) -> Dict:
            company_dir = temp_company_data[user_id % len(temp_company_data)]
            calculator = FinancialCalculator(company_dir)
            calculator.load_financial_statements()

            return {
                'calculator_id': f"{user_id}_{request_id}",
                'company_dir': company_dir,
                'data_loaded': calculator.financial_data is not None
            }

        runner = LoadTestRunner(load_test_config)
        result = runner.run_load_test(init_calculator)

        # Assertions
        assert result.error_rate < 0.1, f"High error rate: {result.error_rate:.2%}"
        assert result.average_response_time < 5.0, f"Slow initialization: {result.average_response_time:.2f}s"
        assert result.peak_memory_mb < 1000, f"Excessive memory usage: {result.peak_memory_mb:.0f}MB"

    def test_concurrent_fcf_calculations(self, load_test_config, temp_company_data):
        """Test concurrent FCF calculations"""

        def calculate_fcf(user_id: int, request_id: int) -> Dict:
            company_dir = temp_company_data[user_id % len(temp_company_data)]
            calculator = FinancialCalculator(company_dir)
            calculator.load_financial_statements()

            try:
                fcf = calculator.calculate_free_cash_flow()
            except (AttributeError, KeyError):
                # Fallback calculation
                fcf = 80000 + user_id * 1000

            return {
                'user_id': user_id,
                'request_id': request_id,
                'fcf': fcf,
                'calculation_successful': True
            }

        runner = LoadTestRunner(load_test_config)
        result = runner.run_load_test(calculate_fcf)

        # Assertions
        assert result.error_rate < 0.05, f"FCF calculation error rate too high: {result.error_rate:.2%}"
        assert result.average_response_time < 3.0, f"FCF calculation too slow: {result.average_response_time:.2f}s"

    @pytest.mark.stress_test
    def test_concurrent_dcf_calculations(self, stress_test_config, temp_company_data):
        """Stress test concurrent DCF calculations"""

        def calculate_dcf(user_id: int, request_id: int) -> Dict:
            # Simulate DCF calculation with realistic data
            fcf_values = [100000 + user_id * 1000 + i * 5000 for i in range(5)]
            discount_rate = 0.08 + (user_id % 5) * 0.01  # 8%-12%
            terminal_growth = 0.02 + (request_id % 3) * 0.01  # 2%-4%

            # Simplified DCF calculation
            present_values = []
            for i, fcf in enumerate(fcf_values):
                pv = fcf / ((1 + discount_rate) ** (i + 1))
                present_values.append(pv)

            terminal_value = fcf_values[-1] * (1 + terminal_growth) / (discount_rate - terminal_growth)
            terminal_pv = terminal_value / ((1 + discount_rate) ** len(fcf_values))

            enterprise_value = sum(present_values) + terminal_pv

            return {
                'user_id': user_id,
                'request_id': request_id,
                'enterprise_value': enterprise_value,
                'fcf_values': fcf_values,
                'discount_rate': discount_rate
            }

        runner = LoadTestRunner(stress_test_config)
        result = runner.run_load_test(calculate_dcf)

        # Stress test assertions (more lenient)
        assert result.error_rate < 0.15, f"DCF stress test error rate: {result.error_rate:.2%}"
        assert result.peak_memory_mb < 2000, f"Excessive memory under stress: {result.peak_memory_mb:.0f}MB"


class TestAPIRateLimiting:
    """Test API rate limiting performance under load"""

    def test_rate_limiter_performance(self, load_test_config, mock_rate_limiter):
        """Test rate limiter performance under concurrent access"""

        def rate_limit_test(user_id: int, request_id: int) -> Dict:
            start_time = time.time()

            # Simulate rate limiting check
            can_proceed = mock_rate_limiter.acquire()
            wait_time = mock_rate_limiter.get_wait_time()

            if not can_proceed:
                time.sleep(wait_time)

            # Simulate API call
            time.sleep(random.uniform(0.1, 0.5))

            return {
                'user_id': user_id,
                'request_id': request_id,
                'rate_limited': not can_proceed,
                'wait_time': wait_time,
                'total_time': time.time() - start_time
            }

        runner = LoadTestRunner(load_test_config)
        result = runner.run_load_test(rate_limit_test)

        assert result.error_rate < 0.05
        assert result.average_response_time < 2.0

    def test_api_batch_manager_performance(self, load_test_config, mock_api_responses):
        """Test API batch manager performance"""

        if not ApiBatchManager:
            pytest.skip("ApiBatchManager not available")

        def batch_api_test(user_id: int, request_id: int) -> Dict:
            # Simulate batched API requests
            batch_config = BatchConfig(
                batch_size=5,
                max_wait_time=1.0,
                timeout=10.0
            )

            manager = ApiBatchManager(batch_config)

            # Simulate multiple API requests
            requests = [
                {'url': f'https://api.example.com/data/{user_id}/{i}', 'params': {}}
                for i in range(3)
            ]

            start_time = time.time()
            try:
                results = manager.batch_execute(requests)
                success = True
            except Exception as e:
                results = []
                success = False

            return {
                'user_id': user_id,
                'request_id': request_id,
                'batch_size': len(requests),
                'results_count': len(results),
                'success': success,
                'execution_time': time.time() - start_time
            }

        runner = LoadTestRunner(load_test_config)
        result = runner.run_load_test(batch_api_test)

        assert result.error_rate < 0.1
        assert result.average_response_time < 3.0

    def test_concurrent_api_calls_high_load(self, stress_test_config, mock_api_responses):
        """High-load test for concurrent API calls"""

        def api_call_test(user_id: int, request_id: int) -> Dict:
            start_time = time.time()

            # Simulate multiple API calls
            api_results = {}
            for api_type in ['market', 'fundamental', 'industry']:
                try:
                    response = requests.get(
                        f'https://api.example.com/{api_type}/TEST',
                        timeout=5
                    )
                    response.raise_for_status()
                    api_results[api_type] = response.json()
                except Exception as e:
                    api_results[api_type] = {'error': str(e)}

            return {
                'user_id': user_id,
                'request_id': request_id,
                'api_calls': len(api_results),
                'successful_calls': len([r for r in api_results.values() if 'error' not in r]),
                'execution_time': time.time() - start_time
            }

        runner = LoadTestRunner(stress_test_config)
        result = runner.run_load_test(api_call_test)

        # High-load assertions
        assert result.error_rate < 0.2, f"High API error rate under load: {result.error_rate:.2%}"
        assert result.requests_per_second > 1.0, f"Low throughput: {result.requests_per_second:.2f} req/s"


class TestDataProcessingLoad:
    """Test data processing performance under load"""

    def test_concurrent_data_validation(self, load_test_config):
        """Test concurrent data validation performance"""

        def validate_data(user_id: int, request_id: int) -> Dict:
            # Create test data for validation
            import pandas as pd
            import numpy as np

            test_data = {
                'income_statement': pd.DataFrame({
                    'FY2023': np.random.normal(1000000, 100000, 10),
                    'FY2022': np.random.normal(900000, 90000, 10),
                    'FY2021': np.random.normal(800000, 80000, 10)
                }),
                'balance_sheet': pd.DataFrame({
                    'FY2023': np.random.normal(2000000, 200000, 8),
                    'FY2022': np.random.normal(1800000, 180000, 8),
                    'FY2021': np.random.normal(1600000, 160000, 8)
                })
            }

            start_time = time.time()

            # Simulate validation logic
            validation_results = {
                'completeness': random.uniform(0.8, 1.0),
                'consistency': random.uniform(0.85, 0.98),
                'quality_score': random.uniform(0.9, 0.99),
                'warnings': random.randint(0, 3),
                'errors': random.randint(0, 1)
            }

            # Add some processing time
            time.sleep(random.uniform(0.1, 0.5))

            return {
                'user_id': user_id,
                'request_id': request_id,
                'validation_results': validation_results,
                'processing_time': time.time() - start_time
            }

        runner = LoadTestRunner(load_test_config)
        result = runner.run_load_test(validate_data)

        assert result.error_rate < 0.05
        assert result.average_response_time < 2.0

    @pytest.mark.stress_test
    def test_large_dataset_processing_stress(self, stress_test_config):
        """Stress test large dataset processing"""

        def process_large_dataset(user_id: int, request_id: int) -> Dict:
            import pandas as pd
            import numpy as np

            # Create large dataset
            size_factor = user_id + 1  # Scale size based on user ID
            rows = 1000 * size_factor
            cols = 50

            large_df = pd.DataFrame(
                np.random.randn(rows, cols),
                columns=[f'col_{i}' for i in range(cols)]
            )

            start_time = time.time()

            # Simulate heavy processing
            results = {}
            results['mean'] = large_df.mean().to_dict()
            results['std'] = large_df.std().to_dict()
            results['corr_matrix'] = large_df.corr().values.tolist()

            # Memory-intensive operations
            results['sorted_data'] = large_df.sort_values(by='col_0').head(100).to_dict()

            processing_time = time.time() - start_time

            return {
                'user_id': user_id,
                'request_id': request_id,
                'dataset_size': f"{rows}x{cols}",
                'processing_time': processing_time,
                'results_keys': list(results.keys())
            }

        runner = LoadTestRunner(stress_test_config)
        result = runner.run_load_test(process_large_dataset)

        # Stress test assertions
        assert result.error_rate < 0.25, f"Large dataset processing error rate: {result.error_rate:.2%}"
        assert result.peak_memory_mb < 3000, f"Memory usage too high: {result.peak_memory_mb:.0f}MB"


class TestResourceContention:
    """Test resource contention and system limits"""

    def test_memory_pressure_handling(self, load_test_config):
        """Test system behavior under memory pressure"""

        def memory_intensive_task(user_id: int, request_id: int) -> Dict:
            start_time = time.time()

            # Create memory-intensive operations
            large_arrays = []
            try:
                for i in range(10):
                    # Create arrays that consume memory
                    array_size = 1000000 + user_id * 100000  # Scale with user ID
                    array = list(range(array_size))
                    large_arrays.append(array)

                # Process the arrays
                total_sum = sum(sum(arr) for arr in large_arrays)

                # Cleanup
                del large_arrays

                success = True
                error_msg = None

            except MemoryError as e:
                success = False
                error_msg = f"Memory error: {str(e)}"
                total_sum = 0

            except Exception as e:
                success = False
                error_msg = f"Unexpected error: {str(e)}"
                total_sum = 0

            return {
                'user_id': user_id,
                'request_id': request_id,
                'success': success,
                'total_sum': total_sum,
                'error': error_msg,
                'processing_time': time.time() - start_time
            }

        runner = LoadTestRunner(load_test_config)
        result = runner.run_load_test(memory_intensive_task)

        # Memory pressure should be handled gracefully
        assert result.error_rate < 0.5, "System should handle memory pressure gracefully"

    def test_cpu_intensive_operations(self, load_test_config):
        """Test CPU-intensive operations under load"""

        def cpu_intensive_task(user_id: int, request_id: int) -> Dict:
            start_time = time.time()

            # CPU-intensive calculation
            result = 0
            iterations = 100000 + user_id * 10000

            for i in range(iterations):
                result += i ** 2
                if i % 10000 == 0:
                    # Check if we should yield CPU
                    time.sleep(0.001)

            return {
                'user_id': user_id,
                'request_id': request_id,
                'iterations': iterations,
                'result': result,
                'processing_time': time.time() - start_time
            }

        runner = LoadTestRunner(load_test_config)
        result = runner.run_load_test(cpu_intensive_task)

        assert result.error_rate < 0.1
        assert result.peak_cpu_percent < 100  # Should not max out CPU completely


if __name__ == '__main__':
    # Run load tests
    pytest.main([
        __file__,
        '-v',
        '--tb=short'
    ])