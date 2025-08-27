"""
Configuration and fixtures for performance tests.
"""

import pytest
from unittest.mock import Mock, patch
import time
import psutil
import os


@pytest.fixture
def mock_external_api():
    """Mock external API calls for performance testing."""
    with patch('yfinance.Ticker') as mock_ticker:
        mock_instance = Mock()
        mock_instance.info = {
            'marketCap': 1000000000,
            'shares': 1000000,
            'beta': 1.2
        }
        mock_instance.financials = Mock()
        mock_instance.balance_sheet = Mock()
        mock_instance.cashflow = Mock()
        mock_ticker.return_value = mock_instance
        yield mock_instance


@pytest.fixture
def performance_monitor():
    """Monitor system performance during tests."""
    class PerformanceMonitor:
        def __init__(self):
            self.start_time = None
            self.start_memory = None
            self.process = psutil.Process()
        
        def start(self):
            self.start_time = time.time()
            self.start_memory = self.process.memory_info().rss
        
        def stop(self):
            end_time = time.time()
            end_memory = self.process.memory_info().rss
            
            return {
                'execution_time': end_time - self.start_time,
                'memory_delta': end_memory - self.start_memory,
                'peak_memory': max(self.start_memory, end_memory)
            }
    
    return PerformanceMonitor()


@pytest.fixture
def sample_performance_data():
    """Provide optimized sample data for performance testing."""
    import pandas as pd
    import numpy as np
    
    return {
        'ticker': 'TEST',
        'fcf_data': [100, 110, 121, 133, 146],  # Growing FCF
        'years': [2019, 2020, 2021, 2022, 2023],
        'market_data': {
            'price': 150.0,
            'shares': 1000000,
            'market_cap': 150000000,
            'beta': 1.2,
            'currentPrice': 150.0,
            'regularMarketPrice': 150.0
        },
        'financials_df': pd.DataFrame({
            'Date': pd.date_range('2023-01-01', periods=5),
            'Revenue': [1000, 1100, 1210, 1331, 1464],
            'Operating Income': [200, 220, 242, 266, 293],
            'Net Income': [150, 165, 182, 200, 220]
        }),
        'balance_sheet_df': pd.DataFrame({
            'Date': pd.date_range('2023-01-01', periods=5),
            'Total Assets': [5000, 5500, 6050, 6655, 7320],
            'Total Debt': [2000, 2100, 2205, 2315, 2431],
            'Shareholders Equity': [2500, 2750, 3025, 3328, 3660]
        }),
        'cashflow_df': pd.DataFrame({
            'Date': pd.date_range('2023-01-01', periods=5),
            'Operating Cash Flow': [120, 132, 145, 160, 176],
            'Capital Expenditures': [-20, -22, -24, -27, -30],
            'Free Cash Flow': [100, 110, 121, 133, 146]
        })
    }


@pytest.fixture(scope="session")
def temp_test_data(tmp_path_factory):
    """Create temporary test data directory structure with minimal Excel files."""
    temp_dir = tmp_path_factory.mktemp("test_data")
    
    # Create company structure
    company_dir = temp_dir / "TEST_COMPANY"
    company_dir.mkdir()
    
    fy_dir = company_dir / "FY"
    ltm_dir = company_dir / "LTM"
    fy_dir.mkdir()
    ltm_dir.mkdir()
    
    # Create minimal Excel files for testing
    excel_files = [
        "Income Statement.xlsx",
        "Balance Sheet.xlsx", 
        "Cash Flow Statement.xlsx"
    ]
    
    # Create lightweight Excel files with openpyxl
    try:
        from openpyxl import Workbook
        for excel_file in excel_files:
            for folder in [fy_dir, ltm_dir]:
                wb = Workbook()
                ws = wb.active
                ws.title = "Sheet1"
                
                # Add minimal header data
                ws['A1'] = 'Date'
                ws['B1'] = 'Value' 
                ws['A2'] = '2023-12-31'
                ws['B2'] = 1000
                
                wb.save(folder / excel_file)
                wb.close()
    except ImportError:
        # Fallback to empty files if openpyxl not available
        for excel_file in excel_files:
            (fy_dir / excel_file).touch()
            (ltm_dir / excel_file).touch()
    
    return str(company_dir)


@pytest.fixture
def fast_timeout():
    """Provide a fast timeout for performance tests."""
    return 10  # 10 seconds


@pytest.fixture
def medium_timeout():
    """Provide a medium timeout for performance tests."""
    return 30  # 30 seconds


@pytest.fixture
def slow_timeout():
    """Provide a slow timeout for performance tests."""
    return 60  # 1 minute - reduced from 2 minutes


@pytest.fixture
def mock_yfinance():
    """Mock yfinance completely for performance tests."""
    with patch('yfinance.Ticker') as mock_ticker, \
         patch('yfinance.download') as mock_download:
        
        # Mock ticker instance
        mock_instance = Mock()
        mock_instance.info = {
            'symbol': 'TEST',
            'shortName': 'Test Company Inc.',
            'marketCap': 1000000000,
            'shares': 1000000,
            'beta': 1.2,
            'currentPrice': 150.0,
            'regularMarketPrice': 150.0,
            'sharesOutstanding': 1000000
        }
        
        # Mock financial data
        mock_instance.financials = Mock()
        mock_instance.balance_sheet = Mock() 
        mock_instance.cashflow = Mock()
        mock_instance.history = Mock(return_value=Mock())
        
        mock_ticker.return_value = mock_instance
        mock_download.return_value = Mock()
        
        yield mock_instance


@pytest.fixture
def mock_requests():
    """Mock all requests for external APIs."""
    with patch('requests.get') as mock_get, \
         patch('requests.post') as mock_post:
        
        # Mock successful response
        mock_response = Mock()
        mock_response.status_code = 200
        mock_response.json.return_value = {
            'symbol': 'TEST',
            'price': 150.0,
            'market_cap': 1000000000
        }
        mock_response.raise_for_status.return_value = None
        
        mock_get.return_value = mock_response
        mock_post.return_value = mock_response
        
        yield mock_response


@pytest.fixture
def mock_excel_data():
    """Mock Excel data loading for performance tests."""
    with patch('openpyxl.load_workbook') as mock_wb, \
         patch('pandas.read_excel') as mock_read:
        
        # Mock workbook
        mock_workbook = Mock()
        mock_worksheet = Mock()
        mock_workbook.active = mock_worksheet
        mock_wb.return_value = mock_workbook
        
        # Mock pandas DataFrame
        import pandas as pd
        import numpy as np
        
        # Create small test data
        test_data = pd.DataFrame({
            'Date': pd.date_range('2023-01-01', periods=5),
            'Revenue': [1000, 1100, 1210, 1331, 1464],
            'Free Cash Flow': [100, 110, 121, 133, 146],
            'Total Assets': [5000, 5500, 6050, 6655, 7320]
        })
        
        mock_read.return_value = test_data
        
        yield test_data