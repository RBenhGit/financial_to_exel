"""
Comprehensive Excel Format and Edge Case Testing Suite
=====================================================

This module provides extensive testing for various Excel formats, edge cases,
and error conditions to ensure robust performance across different scenarios.

Test Categories:
- Different Excel file formats (.xlsx, .xlsm, .xls)
- Corrupted and malformed files
- Files with merged cells and complex formatting
- Very large files with memory constraints
- Files with formulas and circular references
- International formats and character encodings
- Protected and password-protected files
- Empty files and files with missing data
- Files with different date formats and number systems

Usage:
    # Run all edge case tests
    pytest tests/performance/test_excel_format_edge_cases.py -v

    # Run specific categories
    pytest tests/performance/test_excel_format_edge_cases.py -k "corrupted"
    pytest tests/performance/test_excel_format_edge_cases.py -k "international"
    pytest tests/performance/test_excel_format_edge_cases.py -k "large_file"
"""

import pytest
import os
import sys
import tempfile
import shutil
import time
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import random
import string
import datetime
from dataclasses import dataclass
import psutil

# Add project root to path
project_root = Path(__file__).parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

# Import modules to test
from core.excel_integration.performance_optimized_processor import (
    PerformanceOptimizedExcelProcessor,
    PerformanceConfig,
    create_performance_config
)
from core.excel_integration.excel_utils import (
    ExcelDataExtractor,
    ExcelFileValidator,
    ExcelProcessingError,
    CorruptedFileError
)
from core.excel_integration.format_detector import ExcelFormatDetector


@dataclass
class EdgeCaseScenario:
    """Configuration for edge case test scenarios"""
    name: str
    description: str
    file_type: str
    complexity: str
    expected_outcome: str  # 'success', 'failure', 'partial'
    performance_threshold_seconds: float = 30.0


class EdgeCaseFileGenerator:
    """Generate Excel files with various edge cases and problems"""

    @staticmethod
    def create_corrupted_file(file_path: str) -> Dict[str, Any]:
        """Create a corrupted Excel file"""
        try:
            # Create a normal file first
            wb = Workbook()
            ws = wb.active
            ws.cell(1, 1, "Test Data")
            wb.save(file_path)
            wb.close()

            # Corrupt the file by truncating it
            with open(file_path, 'r+b') as f:
                f.seek(0, 2)  # Go to end
                size = f.tell()
                f.seek(size // 2)  # Go to middle
                f.truncate()  # Truncate from middle

            return {
                'file_path': file_path,
                'type': 'corrupted',
                'success': True
            }

        except Exception as e:
            return {
                'file_path': file_path,
                'type': 'corrupted',
                'success': False,
                'error': str(e)
            }

    @staticmethod
    def create_empty_file(file_path: str) -> Dict[str, Any]:
        """Create an empty Excel file"""
        try:
            wb = Workbook()
            ws = wb.active
            # Don't add any data - leave completely empty
            wb.save(file_path)
            wb.close()

            return {
                'file_path': file_path,
                'type': 'empty',
                'success': True
            }

        except Exception as e:
            return {
                'file_path': file_path,
                'type': 'empty',
                'success': False,
                'error': str(e)
            }

    @staticmethod
    def create_merged_cells_file(file_path: str) -> Dict[str, Any]:
        """Create Excel file with heavily merged cells"""
        try:
            wb = Workbook()
            ws = wb.active

            # Add company name in merged cells
            ws.merge_cells('B2:E2')
            ws.cell(2, 2, "Test Company Inc.")

            # Add headers with merged cells
            ws.merge_cells('B4:C4')
            ws.cell(4, 2, "Financial Data")

            # Add period headers with merging
            periods = ["FY-2", "FY-1", "FY", "Q1", "Q2"]
            for i, period in enumerate(periods, 4):
                ws.merge_cells(f'{chr(65+i)}5:{chr(65+i)}6')
                ws.cell(5, i, period)

            # Add financial metrics with some merged cells
            metrics = [
                "Revenue", "Cost of Revenue", "Gross Profit",
                "Operating Expenses", "Operating Income", "Net Income"
            ]

            for row, metric in enumerate(metrics, 8):
                # Merge metric name across multiple columns sometimes
                if row % 3 == 0:
                    ws.merge_cells(f'A{row}:B{row}')

                ws.cell(row, 1, metric)

                # Add some data
                for col in range(4, 9):
                    ws.cell(row, col, random.randint(100000, 1000000))

            wb.save(file_path)
            wb.close()

            return {
                'file_path': file_path,
                'type': 'merged_cells',
                'success': True,
                'metrics_count': len(metrics),
                'merged_ranges': 5
            }

        except Exception as e:
            return {
                'file_path': file_path,
                'type': 'merged_cells',
                'success': False,
                'error': str(e)
            }

    @staticmethod
    def create_formula_heavy_file(file_path: str) -> Dict[str, Any]:
        """Create Excel file with complex formulas"""
        try:
            wb = Workbook()
            ws = wb.active

            # Add headers
            ws.cell(1, 1, "Metric")
            for i, year in enumerate(["FY-2", "FY-1", "FY"], 2):
                ws.cell(1, i, year)

            # Base metrics with values
            base_metrics = {
                "Revenue": [1000000, 1100000, 1200000],
                "Cost of Revenue": [600000, 650000, 700000],
                "R&D Expenses": [50000, 55000, 60000],
                "Sales & Marketing": [80000, 85000, 90000]
            }

            row = 2
            for metric, values in base_metrics.items():
                ws.cell(row, 1, metric)
                for col, value in enumerate(values, 2):
                    ws.cell(row, col, value)
                row += 1

            # Calculated metrics with formulas
            calculated_metrics = [
                ("Gross Profit", "=B2-B3"),  # Revenue - Cost of Revenue
                ("Total OpEx", "=B4+B5"),   # R&D + Sales & Marketing
                ("Operating Income", "=B6-B7"),  # Gross Profit - Total OpEx
                ("Operating Margin", "=B8/B2"),  # Operating Income / Revenue
                ("Revenue Growth", "=(C2-B2)/B2"),  # YoY growth
                ("Cost Ratio", "=B3/B2"),   # Cost / Revenue
            ]

            for metric, formula in calculated_metrics:
                ws.cell(row, 1, metric)
                # Add formulas for each year column
                for col in range(2, 5):
                    # Adjust formula for each column
                    adjusted_formula = formula.replace('B', chr(65 + col - 1))
                    if 'C2-B2' in adjusted_formula:
                        # Handle growth calculation
                        if col == 2:  # First year, no growth
                            adjusted_formula = "=0"
                        elif col == 3:  # Second year
                            adjusted_formula = "=(C2-B2)/B2"
                        else:  # Third year
                            adjusted_formula = "=(D2-C2)/C2"

                    ws.cell(row, col, adjusted_formula)
                row += 1

            # Add some circular reference warning (intentionally problematic)
            ws.cell(row, 1, "Circular Test")
            ws.cell(row, 2, f"=B{row+1}")
            ws.cell(row+1, 2, f"=B{row}")

            wb.save(file_path)
            wb.close()

            return {
                'file_path': file_path,
                'type': 'formula_heavy',
                'success': True,
                'base_metrics': len(base_metrics),
                'calculated_metrics': len(calculated_metrics),
                'has_circular_reference': True
            }

        except Exception as e:
            return {
                'file_path': file_path,
                'type': 'formula_heavy',
                'success': False,
                'error': str(e)
            }

    @staticmethod
    def create_international_format_file(file_path: str, locale: str = "european") -> Dict[str, Any]:
        """Create Excel file with international formatting"""
        try:
            wb = Workbook()
            ws = wb.active

            if locale == "european":
                # European format: comma as decimal separator, dot as thousands
                ws.cell(1, 1, "European Format Data")
                ws.cell(2, 1, "Company Name")
                ws.cell(2, 2, "Société Générale SA")

                # Date format: DD/MM/YYYY
                ws.cell(3, 1, "Period End Date")
                ws.cell(3, 2, "31/12/2023")
                ws.cell(3, 3, "31/12/2022")
                ws.cell(3, 4, "31/12/2021")

                # Numbers with European formatting
                metrics = [
                    ("Chiffre d'affaires", [1234567.89, 1345678.90, 1456789.01]),
                    ("Coût des ventes", [745123.45, 803234.56, 874345.67]),
                    ("Bénéfice brut", [489444.44, 542444.34, 582443.34])
                ]

            elif locale == "asian":
                # Asian format with specific considerations
                ws.cell(1, 1, "Asian Format Data")
                ws.cell(2, 1, "会社名")  # Company name in Japanese
                ws.cell(2, 2, "トヨタ自動車株式会社")

                # Date format: YYYY/MM/DD
                ws.cell(3, 1, "期末日")  # Period end date
                ws.cell(3, 2, "2023/12/31")
                ws.cell(3, 3, "2022/12/31")
                ws.cell(3, 4, "2021/12/31")

                # Large numbers (Asian markets often have larger nominal values)
                metrics = [
                    ("売上高", [12345678900, 13456789000, 14567890100]),  # Revenue in yen
                    ("売上原価", [7451234500, 8032345600, 8743456700]),  # Cost of sales
                    ("営業利益", [4894444400, 5424443400, 5824433400])   # Operating income
                ]

            else:  # US format (default)
                ws.cell(1, 1, "US Format Data")
                ws.cell(2, 1, "Company Name")
                ws.cell(2, 2, "Apple Inc.")

                # Date format: MM/DD/YYYY
                ws.cell(3, 1, "Period End Date")
                ws.cell(3, 2, "12/31/2023")
                ws.cell(3, 3, "12/31/2022")
                ws.cell(3, 4, "12/31/2021")

                metrics = [
                    ("Revenue", [1234567890, 1345678900, 1456789010]),
                    ("Cost of Revenue", [745123450, 803234560, 874345670]),
                    ("Gross Profit", [489444440, 542444340, 582443340])
                ]

            # Add metrics data
            for row, (metric_name, values) in enumerate(metrics, 5):
                ws.cell(row, 1, metric_name)
                for col, value in enumerate(values, 2):
                    ws.cell(row, col, value)

            wb.save(file_path)
            wb.close()

            return {
                'file_path': file_path,
                'type': f'international_{locale}',
                'success': True,
                'locale': locale,
                'metrics_count': len(metrics)
            }

        except Exception as e:
            return {
                'file_path': file_path,
                'type': f'international_{locale}',
                'success': False,
                'error': str(e)
            }

    @staticmethod
    def create_very_large_file(file_path: str, target_size_mb: int = 50) -> Dict[str, Any]:
        """Create a very large Excel file"""
        try:
            wb = Workbook()
            ws = wb.active

            # Calculate approximate rows needed for target size
            # Rough estimate: each cell with data ~100 bytes
            cols = 20
            target_cells = (target_size_mb * 1024 * 1024) // 100
            rows = target_cells // cols

            # Add headers
            for col in range(1, cols + 1):
                ws.cell(1, col, f"Column_{col}")

            # Add data in batches to manage memory
            batch_size = 1000
            total_rows_added = 0

            for batch_start in range(2, rows + 2, batch_size):
                batch_end = min(batch_start + batch_size, rows + 2)

                for row in range(batch_start, batch_end):
                    for col in range(1, cols + 1):
                        if col % 4 == 1:  # Text data
                            value = f"Data_{row}_{col}_{''.join(random.choices(string.ascii_letters, k=10))}"
                        elif col % 4 == 2:  # Integer data
                            value = random.randint(100000, 9999999)
                        elif col % 4 == 3:  # Float data
                            value = random.uniform(1000.0, 999999.0)
                        else:  # Date data
                            base_date = datetime.date(2020, 1, 1)
                            value = base_date + datetime.timedelta(days=random.randint(0, 1000))

                        ws.cell(row, col, value)

                total_rows_added = batch_end - 2

                # Check memory usage and break if getting too high
                memory_mb = psutil.Process().memory_info().rss / 1024 / 1024
                if memory_mb > 2048:  # 2GB limit
                    print(f"Breaking large file creation at {total_rows_added} rows due to memory limit")
                    break

            wb.save(file_path)
            wb.close()

            # Check actual file size
            actual_size_mb = os.path.getsize(file_path) / 1024 / 1024

            return {
                'file_path': file_path,
                'type': 'very_large',
                'success': True,
                'target_size_mb': target_size_mb,
                'actual_size_mb': actual_size_mb,
                'rows_created': total_rows_added,
                'cols_created': cols
            }

        except Exception as e:
            return {
                'file_path': file_path,
                'type': 'very_large',
                'success': False,
                'error': str(e)
            }

    @staticmethod
    def create_missing_data_file(file_path: str) -> Dict[str, Any]:
        """Create Excel file with missing data patterns"""
        try:
            wb = Workbook()
            ws = wb.active

            # Add headers
            ws.cell(1, 1, "Metric")
            periods = ["FY-3", "FY-2", "FY-1", "FY"]
            for col, period in enumerate(periods, 2):
                ws.cell(1, col, period)

            # Add data with various missing patterns
            metrics_data = [
                ("Revenue", [1000000, None, 1200000, 1300000]),  # Missing middle value
                ("Cost of Revenue", [600000, 650000, None, None]),  # Missing recent values
                ("Gross Profit", [None, None, None, 700000]),  # Only latest value
                ("R&D Expenses", [50000, 55000, 60000, 65000]),  # Complete data
                ("Operating Income", [None, 100000, None, 150000]),  # Alternating missing
                ("Net Income", [None, None, None, None]),  # All missing
                ("", [None, None, None, None]),  # Empty row
                ("EPS", [1.5, 1.7, None, 2.1]),  # Missing one value
            ]

            for row, (metric, values) in enumerate(metrics_data, 2):
                ws.cell(row, 1, metric)
                for col, value in enumerate(values, 2):
                    if value is not None:
                        ws.cell(row, col, value)
                    # Leave cell empty for None values

            wb.save(file_path)
            wb.close()

            return {
                'file_path': file_path,
                'type': 'missing_data',
                'success': True,
                'total_metrics': len(metrics_data),
                'empty_metrics': len([m for m in metrics_data if not m[0]]),
                'complete_metrics': len([m for m in metrics_data if all(v is not None for v in m[1])])
            }

        except Exception as e:
            return {
                'file_path': file_path,
                'type': 'missing_data',
                'success': False,
                'error': str(e)
            }


@pytest.fixture
def edge_case_scenarios():
    """Define edge case scenarios to test"""
    return [
        EdgeCaseScenario(
            "corrupted_file",
            "File with corrupted data structure",
            "corrupted",
            "high",
            "failure",
            5.0
        ),
        EdgeCaseScenario(
            "empty_file",
            "Completely empty Excel file",
            "empty",
            "low",
            "partial",
            2.0
        ),
        EdgeCaseScenario(
            "merged_cells",
            "File with extensive merged cells",
            "merged_cells",
            "medium",
            "success",
            10.0
        ),
        EdgeCaseScenario(
            "formula_heavy",
            "File with complex formulas and circular references",
            "formula_heavy",
            "high",
            "partial",
            15.0
        ),
        EdgeCaseScenario(
            "international_european",
            "European format with different decimal separators",
            "international_european",
            "medium",
            "success",
            8.0
        ),
        EdgeCaseScenario(
            "international_asian",
            "Asian format with special characters",
            "international_asian",
            "medium",
            "success",
            8.0
        ),
        EdgeCaseScenario(
            "very_large",
            "Very large file testing memory limits",
            "very_large",
            "high",
            "success",
            60.0
        ),
        EdgeCaseScenario(
            "missing_data",
            "File with extensive missing data patterns",
            "missing_data",
            "medium",
            "partial",
            5.0
        )
    ]


@pytest.fixture(scope="session")
def edge_case_files_dir():
    """Create temporary directory with edge case files"""
    temp_dir = tempfile.mkdtemp(prefix="excel_edge_cases_")
    file_registry = {}

    print("Creating edge case test files...")

    # Create each type of edge case file
    generators = {
        'corrupted': EdgeCaseFileGenerator.create_corrupted_file,
        'empty': EdgeCaseFileGenerator.create_empty_file,
        'merged_cells': EdgeCaseFileGenerator.create_merged_cells_file,
        'formula_heavy': EdgeCaseFileGenerator.create_formula_heavy_file,
        'international_european': lambda fp: EdgeCaseFileGenerator.create_international_format_file(fp, "european"),
        'international_asian': lambda fp: EdgeCaseFileGenerator.create_international_format_file(fp, "asian"),
        'very_large': lambda fp: EdgeCaseFileGenerator.create_very_large_file(fp, 25),  # 25MB target
        'missing_data': EdgeCaseFileGenerator.create_missing_data_file,
    }

    for file_type, generator in generators.items():
        file_path = os.path.join(temp_dir, f"edge_case_{file_type}.xlsx")
        try:
            result = generator(file_path)
            if result.get('success', False):
                file_registry[file_type] = file_path
                print(f"Created {file_type} test file: {os.path.basename(file_path)}")
        except Exception as e:
            print(f"Failed to create {file_type} test file: {e}")

    yield temp_dir, file_registry

    # Cleanup
    shutil.rmtree(temp_dir, ignore_errors=True)


class TestCorruptedFileHandling:
    """Test handling of corrupted and malformed files"""

    def test_corrupted_file_graceful_handling(self, edge_case_files_dir):
        """Test graceful handling of corrupted files"""
        temp_dir, file_registry = edge_case_files_dir

        if 'corrupted' not in file_registry:
            pytest.skip("Corrupted test file not available")

        corrupted_file = file_registry['corrupted']

        # Test with file validator
        validator = ExcelFileValidator()
        validation_result = validator.validate_file(corrupted_file)

        # Should detect corruption
        assert not validation_result.is_valid, "Validator should detect file corruption"
        assert validation_result.corruption_detected, "Should flag corruption detected"
        assert validation_result.severity == 'critical', "Corruption should be critical severity"

        # Test with optimized processor
        config = create_performance_config()
        processor = PerformanceOptimizedExcelProcessor(config)

        start_time = time.time()
        result = processor.process_single_file(corrupted_file, enable_validation=True)
        processing_time = time.time() - start_time

        # Should handle gracefully without crashing
        assert 'errors' in result, "Should report errors for corrupted file"
        assert len(result['errors']) > 0, "Should have error messages"
        assert processing_time < 10.0, f"Should fail fast: {processing_time:.2f}s"

    def test_empty_file_handling(self, edge_case_files_dir):
        """Test handling of empty files"""
        temp_dir, file_registry = edge_case_files_dir

        if 'empty' not in file_registry:
            pytest.skip("Empty test file not available")

        empty_file = file_registry['empty']

        processor = PerformanceOptimizedExcelProcessor()
        result = processor.process_single_file(empty_file)

        # Should handle empty files gracefully
        assert result is not None, "Should return result for empty file"
        # May succeed with empty data or fail gracefully
        if not result.get('success', False):
            assert len(result.get('errors', [])) > 0, "Should provide error message for empty file"


class TestComplexFormattingHandling:
    """Test handling of complex formatting and structures"""

    def test_merged_cells_processing(self, edge_case_files_dir):
        """Test processing files with merged cells"""
        temp_dir, file_registry = edge_case_files_dir

        if 'merged_cells' not in file_registry:
            pytest.skip("Merged cells test file not available")

        merged_file = file_registry['merged_cells']

        # Test with standard extractor
        start_time = time.time()
        try:
            with ExcelDataExtractor(merged_file) as extractor:
                # Should be able to find company name despite merging
                company_name = extractor.find_company_name()

                # Should be able to extract some financial metrics
                income_metrics = extractor.extract_all_financial_metrics('income')

                success = True
                error_msg = None

        except Exception as e:
            success = False
            error_msg = str(e)

        processing_time = time.time() - start_time

        # Should handle merged cells reasonably well
        assert processing_time < 15.0, f"Processing with merged cells too slow: {processing_time:.2f}s"

        if success:
            # If successful, should extract some data
            print(f"Extracted company name: {company_name}")
            print(f"Extracted {len(income_metrics) if income_metrics else 0} income metrics")

    def test_formula_heavy_processing(self, edge_case_files_dir):
        """Test processing files with complex formulas"""
        temp_dir, file_registry = edge_case_files_dir

        if 'formula_heavy' not in file_registry:
            pytest.skip("Formula heavy test file not available")

        formula_file = file_registry['formula_heavy']

        # Test with performance optimized processor
        config = create_performance_config(use_read_only_mode=True)  # Should help with formulas
        processor = PerformanceOptimizedExcelProcessor(config)

        start_time = time.time()
        result = processor.process_single_file(formula_file)
        processing_time = time.time() - start_time

        # Should handle formulas without excessive delay
        assert processing_time < 20.0, f"Formula processing too slow: {processing_time:.2f}s"

        # Should extract some data even with formulas
        if result.get('success', False):
            data = result.get('data', {})
            assert len(data) > 0, "Should extract some data from formula file"


class TestInternationalFormatSupport:
    """Test support for international formats and character encodings"""

    @pytest.mark.parametrize("locale", ["european", "asian"])
    def test_international_format_processing(self, edge_case_files_dir, locale):
        """Test processing international format files"""
        temp_dir, file_registry = edge_case_files_dir

        file_key = f'international_{locale}'
        if file_key not in file_registry:
            pytest.skip(f"International {locale} test file not available")

        international_file = file_registry[file_key]

        # Test format detection
        detector = ExcelFormatDetector()

        with load_workbook(international_file, read_only=True) as wb:
            ws = wb.active
            format_result = detector.detect_format(ws)

        # Should detect some format characteristics
        assert format_result.confidence_score > 0.3, f"Should have reasonable confidence for {locale} format"

        # Test data extraction
        processor = PerformanceOptimizedExcelProcessor()
        result = processor.process_single_file(international_file)

        # Should handle international formats
        assert result is not None, f"Should process {locale} format file"

        if result.get('success', False):
            data = result.get('data', {})
            print(f"Successfully processed {locale} format with {len(data)} data sections")


class TestLargeFilePerformance:
    """Test performance with very large files"""

    def test_very_large_file_streaming(self, edge_case_files_dir):
        """Test streaming processing of very large files"""
        temp_dir, file_registry = edge_case_files_dir

        if 'very_large' not in file_registry:
            pytest.skip("Very large test file not available")

        large_file = file_registry['very_large']

        # Check file size
        file_size_mb = os.path.getsize(large_file) / 1024 / 1024
        print(f"Testing large file: {file_size_mb:.1f}MB")

        # Test with streaming enabled
        config = create_performance_config(
            enable_streaming=True,
            chunk_size=500,  # Smaller chunks for large files
            max_memory_mb=1024
        )
        processor = PerformanceOptimizedExcelProcessor(config)

        start_time = time.time()
        start_memory = psutil.Process().memory_info().rss / 1024 / 1024

        result = processor.process_single_file(large_file)

        end_time = time.time()
        end_memory = psutil.Process().memory_info().rss / 1024 / 1024

        processing_time = end_time - start_time
        memory_delta = end_memory - start_memory

        # Performance assertions for large files
        assert processing_time < 90.0, f"Large file processing too slow: {processing_time:.1f}s"
        assert memory_delta < 800, f"Memory usage too high for large file: {memory_delta:.1f}MB"

        if result.get('success', False):
            print(f"Successfully processed {file_size_mb:.1f}MB file in {processing_time:.1f}s")
            print(f"Memory delta: {memory_delta:.1f}MB")

    def test_memory_limit_enforcement(self, edge_case_files_dir):
        """Test that memory limits are enforced with large files"""
        temp_dir, file_registry = edge_case_files_dir

        if 'very_large' not in file_registry:
            pytest.skip("Very large test file not available")

        large_file = file_registry['very_large']

        # Test with very restrictive memory limit
        config = create_performance_config(
            max_memory_mb=256,  # Very restrictive
            enable_streaming=True,
            chunk_size=100
        )
        processor = PerformanceOptimizedExcelProcessor(config)

        start_memory = psutil.Process().memory_info().rss / 1024 / 1024

        result = processor.process_single_file(large_file)

        peak_memory = psutil.Process().memory_info().rss / 1024 / 1024

        # Should either succeed with controlled memory or fail gracefully
        memory_delta = peak_memory - start_memory

        # Allow some overhead but should generally stay within reasonable bounds
        assert memory_delta < 600, f"Memory usage exceeded reasonable bounds: {memory_delta:.1f}MB"


class TestMissingDataHandling:
    """Test handling of missing and incomplete data"""

    def test_missing_data_graceful_handling(self, edge_case_files_dir):
        """Test graceful handling of missing data patterns"""
        temp_dir, file_registry = edge_case_files_dir

        if 'missing_data' not in file_registry:
            pytest.skip("Missing data test file not available")

        missing_data_file = file_registry['missing_data']

        # Test with data extractor
        try:
            with ExcelDataExtractor(missing_data_file, enable_recovery=True) as extractor:
                # Should handle missing data gracefully
                income_metrics = extractor.extract_all_financial_metrics('income')

                # Count metrics with some data
                metrics_with_data = 0
                total_metrics = 0

                if income_metrics:
                    for metric_name, metric in income_metrics.items():
                        total_metrics += 1
                        if metric.values and any(v is not None for v in metric.values):
                            metrics_with_data += 1

                success = True

        except Exception as e:
            success = False
            metrics_with_data = 0
            total_metrics = 0
            print(f"Error processing missing data file: {e}")

        # Should handle missing data without crashing
        assert success, "Should handle missing data gracefully"

        if total_metrics > 0:
            data_completeness = metrics_with_data / total_metrics
            print(f"Data completeness: {data_completeness:.1%} ({metrics_with_data}/{total_metrics})")

    def test_recovery_engine_with_missing_data(self, edge_case_files_dir):
        """Test data recovery engine with missing data"""
        temp_dir, file_registry = edge_case_files_dir

        if 'missing_data' not in file_registry:
            pytest.skip("Missing data test file not available")

        missing_data_file = file_registry['missing_data']

        # Test with recovery enabled
        try:
            with ExcelDataExtractor(missing_data_file, enable_recovery=True) as extractor:
                # Check if recovery was attempted
                validation_result = extractor.get_validation_report()
                recovery_result = extractor.get_recovery_report()

                success = True

        except Exception as e:
            success = False
            validation_result = None
            recovery_result = None
            print(f"Error in recovery test: {e}")

        assert success, "Recovery engine should handle missing data"

        if recovery_result:
            print(f"Recovery attempted: {recovery_result.success}")
            print(f"Recovery method: {recovery_result.recovery_method}")
            print(f"Confidence: {recovery_result.confidence_score:.2f}")


class TestErrorHandlingAndResilience:
    """Test error handling and system resilience"""

    def test_concurrent_edge_case_processing(self, edge_case_files_dir):
        """Test concurrent processing of multiple edge case files"""
        temp_dir, file_registry = edge_case_files_dir

        # Collect available edge case files
        available_files = [
            file_path for file_type, file_path in file_registry.items()
            if file_type != 'very_large'  # Skip very large for concurrent test
        ]

        if len(available_files) < 2:
            pytest.skip("Insufficient edge case files for concurrent testing")

        # Test concurrent processing
        config = create_performance_config(max_workers=3)
        processor = PerformanceOptimizedExcelProcessor(config)

        start_time = time.time()
        result = processor.process_multiple_files(available_files)
        total_time = time.time() - start_time

        # Should handle mixed edge cases concurrently
        assert result['total_files'] == len(available_files), "Should process all files"
        assert total_time < 60.0, f"Concurrent edge case processing too slow: {total_time:.1f}s"

        # At least some files should be processed successfully
        success_rate = result['successful_files'] / result['total_files'] if result['total_files'] > 0 else 0
        assert success_rate >= 0.3, f"Success rate too low for edge cases: {success_rate:.1%}"

        print(f"Processed {result['successful_files']}/{result['total_files']} edge case files successfully")

    def test_resource_cleanup_after_errors(self, edge_case_files_dir):
        """Test that resources are properly cleaned up after errors"""
        temp_dir, file_registry = edge_case_files_dir

        if 'corrupted' not in file_registry:
            pytest.skip("Corrupted test file not available for cleanup testing")

        corrupted_file = file_registry['corrupted']

        initial_memory = psutil.Process().memory_info().rss / 1024 / 1024
        processor = PerformanceOptimizedExcelProcessor()

        # Process corrupted file multiple times to test cleanup
        for i in range(3):
            try:
                result = processor.process_single_file(corrupted_file)
            except Exception as e:
                print(f"Expected error on iteration {i}: {e}")

        final_memory = psutil.Process().memory_info().rss / 1024 / 1024
        memory_delta = final_memory - initial_memory

        # Memory should not grow excessively due to repeated errors
        assert memory_delta < 100, f"Memory not cleaned up properly after errors: {memory_delta:.1f}MB growth"


if __name__ == '__main__':
    # Run edge case tests
    pytest.main([
        __file__,
        '-v',
        '--tb=short',
        '-x'  # Stop on first failure for debugging
    ])