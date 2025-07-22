"""
Excel test helpers for centralized Excel file operations in tests.

This module eliminates duplicate Excel loading and creation logic across test files.
"""

import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any, Optional, List
import tempfile
import numpy as np
from datetime import datetime, date


class ExcelTestHelper:
    """Centralized Excel operations for testing"""
    
    @staticmethod
    def create_sample_excel_file(file_path: str, statement_type: str) -> None:
        """
        Create a sample Excel file for testing
        
        Args:
            file_path (str): Path where to create the file
            statement_type (str): Type of financial statement
        """
        wb = Workbook()
        ws = wb.active
        
        # Create sample data based on statement type
        if 'Income' in statement_type:
            data = ExcelTestHelper._create_income_statement_data()
        elif 'Balance' in statement_type:
            data = ExcelTestHelper._create_balance_sheet_data()
        elif 'Cash Flow' in statement_type:
            data = ExcelTestHelper._create_cashflow_statement_data()
        else:
            data = ExcelTestHelper._create_generic_financial_data()
        
        # Write data to worksheet
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        wb.save(file_path)
    
    @staticmethod
    def _create_income_statement_data() -> List[List[Any]]:
        """Create sample income statement data"""
        current_year = datetime.now().year
        years = [current_year - 4, current_year - 3, current_year - 2, current_year - 1, current_year]
        
        return [
            ['Test Company Inc', '', '', '', '', ''],
            ['Income Statement', '', '', '', '', ''],
            ['Period End Date', '', ''] + [f'12/31/{year}' for year in years[-3:]],
            ['', '', '', '', '', ''],
            ['Revenue', '', '', 1000, 1100, 1200],
            ['Cost of Goods Sold', '', '', -600, -650, -700],
            ['Gross Profit', '', '', 400, 450, 500],
            ['Operating Expenses', '', '', -200, -220, -240],
            ['EBIT', '', '', 200, 230, 260],
            ['Net Interest Expenses', '', '', -10, -12, -14],
            ['EBT, Incl. Unusual Items', '', '', 190, 218, 246],
            ['Income Tax Expense', '', '', -38, -44, -49],
            ['Net Income to Company', '', '', 152, 174, 197],
        ]
    
    @staticmethod
    def _create_balance_sheet_data() -> List[List[Any]]:
        """Create sample balance sheet data"""
        current_year = datetime.now().year
        years = [current_year - 4, current_year - 3, current_year - 2, current_year - 1, current_year]
        
        return [
            ['Test Company Inc', '', '', '', '', ''],
            ['Balance Sheet', '', '', '', '', ''],
            ['Period End Date', '', ''] + [f'12/31/{year}' for year in years[-3:]],
            ['', '', '', '', '', ''],
            ['ASSETS', '', '', '', '', ''],
            ['Current Assets', '', '', '', '', ''],
            ['Cash and Cash Equivalents', '', '', 100, 120, 140],
            ['Total Current Assets', '', '', 300, 350, 400],
            ['Total Assets', '', '', 1000, 1200, 1400],
            ['', '', '', '', '', ''],
            ['LIABILITIES', '', '', '', '', ''],
            ['Total Current Liabilities', '', '', 200, 220, 250],
            ['Total Liabilities', '', '', 500, 550, 600],
            ['Total Shareholders\' Equity', '', '', 500, 650, 800],
        ]
    
    @staticmethod
    def _create_cashflow_statement_data() -> List[List[Any]]:
        """Create sample cash flow statement data"""
        current_year = datetime.now().year
        years = [current_year - 4, current_year - 3, current_year - 2, current_year - 1, current_year]
        
        return [
            ['Test Company Inc', '', '', '', '', ''],
            ['Cash Flow Statement', '', '', '', '', ''],
            ['Period End Date', '', ''] + [f'12/31/{year}' for year in years[-3:]],
            ['', '', '', '', '', ''],
            ['OPERATING ACTIVITIES', '', '', '', '', ''],
            ['Net Income', '', '', 152, 174, 197],
            ['Depreciation & Amortization (CF)', '', '', 50, 55, 60],
            ['Cash from Operations', '', '', 220, 250, 280],
            ['', '', '', '', '', ''],
            ['INVESTING ACTIVITIES', '', '', '', '', ''],
            ['Capital Expenditures', '', '', -80, -90, -100],
            ['Cash from Investing', '', '', -80, -90, -100],
            ['', '', '', '', '', ''],
            ['FINANCING ACTIVITIES', '', '', '', '', ''],
            ['Cash from Financing', '', '', -50, -60, -70],
        ]
    
    @staticmethod
    def _create_generic_financial_data() -> List[List[Any]]:
        """Create generic financial data for testing"""
        current_year = datetime.now().year
        years = [current_year - 2, current_year - 1, current_year]
        
        return [
            ['Test Company', '', ''] + [f'Year {year}' for year in years],
            ['Metric 1', '', ''] + [100 * (i + 1) for i in range(len(years))],
            ['Metric 2', '', ''] + [200 * (i + 1) for i in range(len(years))],
            ['Metric 3', '', ''] + [300 * (i + 1) for i in range(len(years))],
        ]
    
    @staticmethod
    def load_test_workbook(file_path: str) -> pd.DataFrame:
        """
        Standardized Excel file loading for tests
        
        Args:
            file_path (str): Path to Excel file
            
        Returns:
            pd.DataFrame: Loaded data
        """
        try:
            # Load workbook and get active sheet
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Convert to DataFrame
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            
            return pd.DataFrame(data)
            
        except Exception as e:
            raise ValueError(f"Failed to load Excel file {file_path}: {e}")
    
    @staticmethod
    def create_sample_data() -> Dict[str, pd.DataFrame]:
        """
        Create sample DataFrames for testing without file dependencies
        
        Returns:
            Dict[str, pd.DataFrame]: Sample financial data
        """
        current_year = datetime.now().year
        years = [current_year - 2, current_year - 1, current_year]
        
        # Income statement data
        income_data = pd.DataFrame({
            'Metric': ['Revenue', 'EBIT', 'Net Income to Company', 'Income Tax Expense'],
            'Values_2022': [1000, 200, 152, -38],
            'Values_2023': [1100, 230, 174, -44],
            'Values_2024': [1200, 260, 197, -49]
        })
        
        # Balance sheet data
        balance_data = pd.DataFrame({
            'Metric': ['Total Current Assets', 'Total Current Liabilities', 'Total Assets'],
            'Values_2022': [300, 200, 1000],
            'Values_2023': [350, 220, 1200],
            'Values_2024': [400, 250, 1400]
        })
        
        # Cash flow data
        cashflow_data = pd.DataFrame({
            'Metric': ['Cash from Operations', 'Capital Expenditures', 'Depreciation & Amortization (CF)'],
            'Values_2022': [220, -80, 50],
            'Values_2023': [250, -90, 55],
            'Values_2024': [280, -100, 60]
        })
        
        return {
            'income': income_data,
            'balance': balance_data,
            'cashflow': cashflow_data,
            'years': years
        }
    
    @staticmethod
    def find_metric_in_dataframe(df: pd.DataFrame, metric_name: str, 
                                search_columns: List[int] = None) -> Optional[int]:
        """
        Find a specific metric in a DataFrame (centralized metric finding logic)
        
        Args:
            df (pd.DataFrame): DataFrame to search
            metric_name (str): Name of metric to find
            search_columns (List[int]): Columns to search in
            
        Returns:
            Optional[int]: Row index if found, None otherwise
        """
        if search_columns is None:
            search_columns = [0, 1, 2]
        
        for idx, row in df.iterrows():
            for col_idx in search_columns:
                if len(row) > col_idx and pd.notna(row.iloc[col_idx]):
                    cell_text = str(row.iloc[col_idx]).strip()
                    if metric_name.lower() == cell_text.lower():
                        return idx
        
        return None
    
    @staticmethod
    def extract_values_from_row(df: pd.DataFrame, row_idx: int, 
                               start_col: int = 3, num_values: int = 3) -> List[float]:
        """
        Extract numeric values from a specific row
        
        Args:
            df (pd.DataFrame): DataFrame to extract from
            row_idx (int): Row index
            start_col (int): Starting column for values
            num_values (int): Number of values to extract
            
        Returns:
            List[float]: Extracted numeric values
        """
        values = []
        row = df.iloc[row_idx]
        
        for col_idx in range(start_col, start_col + num_values):
            if col_idx < len(row):
                cell_value = row.iloc[col_idx]
                if pd.notna(cell_value) and isinstance(cell_value, (int, float)):
                    values.append(float(cell_value))
                else:
                    # Try to convert string to float
                    try:
                        values.append(float(str(cell_value)))
                    except (ValueError, TypeError):
                        values.append(None)
            else:
                values.append(None)
        
        return values
    
    @staticmethod
    def create_temp_excel_file(data: Dict[str, Any]) -> str:
        """
        Create a temporary Excel file for testing
        
        Args:
            data (Dict[str, Any]): Data to write to Excel
            
        Returns:
            str: Path to temporary file
        """
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        wb = Workbook()
        ws = wb.active
        
        # Write data to worksheet
        if isinstance(data, dict) and 'rows' in data:
            for row_idx, row_data in enumerate(data['rows'], start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        wb.save(temp_file.name)
        return temp_file.name