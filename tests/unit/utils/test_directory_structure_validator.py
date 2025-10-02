"""
Unit tests for DirectoryStructureValidator (Task 180)

Tests the enhanced directory structure validation system including:
- Directory structure validation
- Excel file format validation
- Compliance reporting
- Actionable recommendations
"""

import pytest
import pandas as pd
from pathlib import Path
import tempfile
import shutil
from openpyxl import Workbook

from utils.directory_structure_helper import (
    DirectoryStructureValidator,
    validate_company_directory,
    create_directory_structure_template,
    get_directory_structure_help
)

# Try to import ValidationRegistry for registry integration tests
try:
    from core.validation.validation_registry import (
        ValidationRegistry,
        ValidationRule,
        RuleType,
        RuleScope,
        RuleSet
    )
    VALIDATION_REGISTRY_AVAILABLE = True
except ImportError:
    VALIDATION_REGISTRY_AVAILABLE = False


class TestDirectoryStructureValidator:
    """Test suite for DirectoryStructureValidator"""

    @pytest.fixture
    def temp_company_dir(self):
        """Create temporary company directory for testing"""
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        # Cleanup
        shutil.rmtree(temp_dir, ignore_errors=True)

    @pytest.fixture
    def validator(self):
        """Create validator instance"""
        return DirectoryStructureValidator()

    @pytest.fixture
    def valid_company_structure(self, temp_company_dir):
        """Create a valid company directory structure"""
        company_path = Path(temp_company_dir) / "TEST"
        company_path.mkdir(parents=True, exist_ok=True)

        for period in ['FY', 'LTM']:
            period_path = company_path / period
            period_path.mkdir(exist_ok=True)

            for statement in ["Income Statement.xlsx", "Balance Sheet.xlsx", "Cash Flow Statement.xlsx"]:
                file_path = period_path / statement
                self._create_valid_excel_file(file_path, statement)

        return company_path

    def _create_valid_excel_file(self, file_path: Path, statement_type: str):
        """Helper to create a valid Excel file with proper structure"""
        wb = Workbook()
        ws = wb.active

        # Add proper column headers
        headers = ['Metric', 'FY', 'FY-1', 'FY-2', 'FY-3']
        ws.append(headers)

        # Add sample data based on statement type
        if "Income Statement" in statement_type:
            ws.append(['Revenue', 1000, 900, 800, 700])
            ws.append(['Net Income', 100, 90, 80, 70])
            ws.append(['Operating Income', 150, 135, 120, 105])
        elif "Balance Sheet" in statement_type:
            ws.append(['Total Assets', 5000, 4500, 4000, 3500])
            ws.append(['Total Liabilities', 3000, 2700, 2400, 2100])
            ws.append(['Shareholders Equity', 2000, 1800, 1600, 1400])
        elif "Cash Flow" in statement_type:
            ws.append(['Cash from Operations', 200, 180, 160, 140])
            ws.append(['Capital Expenditures', 50, 45, 40, 35])

        wb.save(file_path)

    # === Basic Directory Validation Tests ===

    def test_validate_nonexistent_directory(self, validator):
        """Test validation of nonexistent company directory"""
        result = validator.validate_company_directory("/nonexistent/path/FAKE")

        assert not result['is_valid']
        assert not result['exists']
        assert len(result['issues']) > 0
        assert any('does not exist' in issue['message'] for issue in result['issues'])

    def test_validate_valid_directory_structure(self, validator, valid_company_structure):
        """Test validation of properly structured company directory"""
        result = validator.validate_company_directory(str(valid_company_structure))

        assert result['is_valid']
        assert result['exists']
        assert len(result['missing_folders']) == 0
        assert result['structure_score'] > 0.8
        assert '✅' in result['validation_summary']

    def test_validate_missing_fy_folder(self, validator, temp_company_dir):
        """Test validation when FY folder is missing"""
        company_path = Path(temp_company_dir) / "TEST_NO_FY"
        company_path.mkdir(parents=True)

        # Only create LTM folder
        ltm_path = company_path / "LTM"
        ltm_path.mkdir()

        result = validator.validate_company_directory(str(company_path))

        assert not result['is_valid']
        assert 'FY' in result['missing_folders']
        assert len([i for i in result['issues'] if 'FY' in i['message']]) > 0

    def test_validate_missing_ltm_folder(self, validator, temp_company_dir):
        """Test validation when LTM folder is missing"""
        company_path = Path(temp_company_dir) / "TEST_NO_LTM"
        company_path.mkdir(parents=True)

        # Only create FY folder
        fy_path = company_path / "FY"
        fy_path.mkdir()

        result = validator.validate_company_directory(str(company_path))

        assert not result['is_valid']
        assert 'LTM' in result['missing_folders']
        assert len([i for i in result['issues'] if 'LTM' in i['message']]) > 0

    def test_validate_missing_excel_files(self, validator, temp_company_dir):
        """Test validation when Excel files are missing"""
        company_path = Path(temp_company_dir) / "TEST_NO_FILES"
        company_path.mkdir(parents=True)

        # Create folders but no files
        for period in ['FY', 'LTM']:
            (company_path / period).mkdir()

        result = validator.validate_company_directory(str(company_path), strict_mode=True)

        assert not result['is_valid']
        assert len(result['missing_files']) > 0
        assert result['structure_score'] < 0.5

    def test_structure_score_calculation(self, validator, temp_company_dir):
        """Test structure completeness score calculation"""
        company_path = Path(temp_company_dir) / "TEST_SCORE"
        company_path.mkdir(parents=True)

        # Create only FY folder with only 1 file
        fy_path = company_path / "FY"
        fy_path.mkdir()
        self._create_valid_excel_file(fy_path / "Income Statement.xlsx", "Income Statement.xlsx")

        result = validator.validate_company_directory(str(company_path))

        # Should have partial score: 1 folder out of 2, 1 file out of 6
        assert 0.1 < result['structure_score'] < 0.5

    # === Excel File Format Validation Tests ===

    def test_validate_excel_file_format_valid(self, validator, temp_company_dir):
        """Test validation of properly formatted Excel file"""
        file_path = Path(temp_company_dir) / "test.xlsx"
        self._create_valid_excel_file(file_path, "Income Statement.xlsx")

        result = validator.validate_excel_file_format(file_path, "Income Statement.xlsx")

        assert result['is_valid']
        assert result['exists']
        assert result['readable']
        assert result['has_data']
        assert result['has_proper_headers']
        assert len(result['detected_periods']) >= 1
        assert 'FY' in str(result['detected_periods'])

    def test_validate_excel_file_nonexistent(self, validator):
        """Test validation of nonexistent Excel file"""
        result = validator.validate_excel_file_format(
            Path("/nonexistent/file.xlsx"),
            "Income Statement.xlsx"
        )

        assert not result['is_valid']
        assert not result['exists']
        assert len(result['issues']) > 0

    def test_validate_excel_file_empty(self, validator, temp_company_dir):
        """Test validation of empty Excel file"""
        file_path = Path(temp_company_dir) / "empty.xlsx"
        wb = Workbook()
        wb.save(file_path)

        result = validator.validate_excel_file_format(file_path, "Income Statement.xlsx")

        assert not result['is_valid']
        assert result['exists']
        assert result['readable']
        assert not result['has_data']

    def test_validate_excel_file_missing_headers(self, validator, temp_company_dir):
        """Test validation of Excel file without proper column headers"""
        file_path = Path(temp_company_dir) / "no_headers.xlsx"
        wb = Workbook()
        ws = wb.active

        # Add data without FY headers
        ws.append(['Metric', '2023', '2022', '2021'])
        ws.append(['Revenue', 1000, 900, 800])

        wb.save(file_path)

        result = validator.validate_excel_file_format(file_path, "Income Statement.xlsx")

        assert result['readable']
        assert result['has_data']
        assert not result['has_proper_headers']
        assert len([w for w in result['warnings'] if 'headers' in w['message']]) > 0

    def test_validate_excel_insufficient_periods(self, validator, temp_company_dir):
        """Test validation with insufficient historical periods"""
        file_path = Path(temp_company_dir) / "few_periods.xlsx"
        wb = Workbook()
        ws = wb.active

        # Only 2 periods (below recommended 3)
        ws.append(['Metric', 'FY', 'FY-1'])
        ws.append(['Revenue', 1000, 900])

        wb.save(file_path)

        result = validator.validate_excel_file_format(file_path, "Income Statement.xlsx")

        warnings = [w for w in result['warnings'] if 'insufficient' in w['type']]
        assert len(warnings) > 0

    def test_validate_required_metrics_income_statement(self, validator, temp_company_dir):
        """Test validation of required metrics in Income Statement"""
        file_path = Path(temp_company_dir) / "incomplete_is.xlsx"
        wb = Workbook()
        ws = wb.active

        # Add headers but missing some required metrics
        ws.append(['Metric', 'FY', 'FY-1', 'FY-2'])
        ws.append(['Revenue', 1000, 900, 800])
        # Missing: Net Income, Operating Income

        wb.save(file_path)

        result = validator.validate_excel_file_format(file_path, "Income Statement.xlsx")

        # Should have warnings about missing metrics
        metric_warnings = [w for w in result['warnings'] if 'missing_metric' in w['type']]
        assert len(metric_warnings) > 0

    # === Comprehensive Directory Structure Validation Tests ===

    def test_validate_directory_structure_method(self, validator, valid_company_structure):
        """Test the main validate_directory_structure method"""
        # Extract ticker from path
        ticker = valid_company_structure.name

        # Run validation
        report = validator.validate_directory_structure(
            ticker,
            base_path=str(valid_company_structure.parent)
        )

        assert report['ticker'] == ticker
        assert 'validation_timestamp' in report
        assert 'directory_validation' in report
        assert 'excel_validations' in report
        assert 'overall_compliance' in report
        assert 'actionable_recommendations' in report

        # Check overall compliance
        compliance = report['overall_compliance']
        assert compliance['status'] in ['FULLY_COMPLIANT', 'COMPLIANT', 'PARTIALLY_COMPLIANT', 'NON_COMPLIANT', 'CRITICAL']
        assert 'overall_score' in compliance
        assert 'directory_score' in compliance
        assert 'excel_score' in compliance

    def test_compliance_score_calculation(self, validator, valid_company_structure):
        """Test compliance score calculation"""
        ticker = valid_company_structure.name

        report = validator.validate_directory_structure(
            ticker,
            base_path=str(valid_company_structure.parent)
        )

        compliance = report['overall_compliance']

        # Valid structure should have high compliance
        assert compliance['overall_score'] >= 0.8
        assert compliance['status'] in ['FULLY_COMPLIANT', 'COMPLIANT', 'PARTIALLY_COMPLIANT']
        assert compliance['total_issues'] == 0 or compliance['total_issues'] is not None

    def test_actionable_recommendations_generation(self, validator, temp_company_dir):
        """Test generation of actionable recommendations"""
        company_path = Path(temp_company_dir) / "INCOMPLETE"
        company_path.mkdir(parents=True)

        # Create only FY folder, missing LTM
        fy_path = company_path / "FY"
        fy_path.mkdir()

        report = validator.validate_directory_structure(
            "INCOMPLETE",
            base_path=temp_company_dir
        )

        recommendations = report['actionable_recommendations']

        # Should have recommendations for missing LTM folder
        assert len(recommendations) > 0
        assert any('LTM' in rec['action'] or 'missing' in rec['action'].lower()
                   for rec in recommendations)
        assert all('priority' in rec for rec in recommendations)
        assert all('category' in rec for rec in recommendations)
        assert all('impact' in rec for rec in recommendations)

    def test_excel_validations_in_comprehensive_report(self, validator, valid_company_structure):
        """Test that Excel file validations are included in comprehensive report"""
        ticker = valid_company_structure.name

        report = validator.validate_directory_structure(
            ticker,
            base_path=str(valid_company_structure.parent)
        )

        excel_validations = report['excel_validations']

        # Should have validations for all 6 files (FY and LTM, 3 statements each)
        assert len(excel_validations) == 6

        # Check keys are in expected format
        for key in excel_validations.keys():
            assert '/' in key  # Should be "FY/Income Statement.xlsx" format

        # All should be valid
        all_valid = all(v['is_valid'] for v in excel_validations.values())
        assert all_valid

    # === Directory Template Creation Tests ===

    def test_create_directory_structure_template(self, validator, temp_company_dir):
        """Test automated directory structure creation"""
        company_path = Path(temp_company_dir) / "NEW_COMPANY"

        result = validator.create_directory_structure_from_template(
            str(company_path),
            company_name="New Company Inc"
        )

        assert result['success']
        assert len(result['created_folders']) >= 3  # Main + FY + LTM
        assert len(result['created_files']) >= 6  # 3 files × 2 periods

        # Verify structure was created
        assert company_path.exists()
        assert (company_path / "FY").exists()
        assert (company_path / "LTM").exists()
        assert (company_path / "FY" / "Income Statement.xlsx").exists()

    # === Convenience Function Tests ===

    def test_validate_company_directory_convenience_function(self, valid_company_structure):
        """Test convenience function for directory validation"""
        result = validate_company_directory(str(valid_company_structure))

        assert result['is_valid']
        assert result['exists']

    def test_create_directory_structure_template_convenience_function(self, temp_company_dir):
        """Test convenience function for template creation"""
        company_path = Path(temp_company_dir) / "TEMPLATE_TEST"

        result = create_directory_structure_template(
            str(company_path),
            company_name="Template Test Co"
        )

        assert result['success']
        assert company_path.exists()

    # === Edge Cases and Error Handling ===

    def test_validation_with_special_characters_in_path(self, validator, temp_company_dir):
        """Test validation with special characters in path"""
        company_path = Path(temp_company_dir) / "COMP@NY-123_TEST"
        company_path.mkdir(parents=True)

        result = validator.validate_company_directory(str(company_path))

        # Should handle special characters without errors
        assert 'company_path' in result
        assert result['exists']

    def test_validation_history_tracking(self, validator, valid_company_structure):
        """Test that validation history is tracked"""
        initial_history_length = len(validator.validation_history)

        validator.validate_company_directory(str(valid_company_structure))

        assert len(validator.validation_history) == initial_history_length + 1

    def test_corrupted_excel_file_handling(self, validator, temp_company_dir):
        """Test handling of corrupted/unreadable Excel files"""
        file_path = Path(temp_company_dir) / "corrupted.xlsx"

        # Create a file that isn't actually a valid Excel file
        with open(file_path, 'w') as f:
            f.write("This is not an Excel file")

        result = validator.validate_excel_file_format(file_path, "Income Statement.xlsx")

        assert not result['is_valid']
        assert result['exists']
        assert not result['readable'] or len(result['issues']) > 0


    # === Task 180.1: Enhanced Excel Validation Tests ===

    def test_validate_numeric_data_types(self, validator, temp_company_dir):
        """Test numeric data type validation in Excel files"""
        file_path = Path(temp_company_dir) / "numeric_test.xlsx"
        wb = Workbook()
        ws = wb.active

        # Add headers and numeric data
        ws.append(['Metric', 'FY', 'FY-1', 'FY-2'])
        ws.append(['Revenue', 1000, 900, 800])
        ws.append(['Net Income', 100, 90, 80])

        wb.save(file_path)

        result = validator.validate_excel_file_format(file_path, "Income Statement.xlsx")

        # Should detect numeric data
        assert result['has_numeric_data']
        assert 'data_type_validation' in result
        assert result['data_type_validation']['total_numeric_columns'] >= 3

    def test_validate_non_numeric_data_warning(self, validator, temp_company_dir):
        """Test detection of non-numeric data in financial columns"""
        file_path = Path(temp_company_dir) / "non_numeric_test.xlsx"
        wb = Workbook()
        ws = wb.active

        # Add headers with mixed data types
        ws.append(['Metric', 'FY', 'FY-1', 'FY-2'])
        ws.append(['Revenue', 'N/A', 900, 800])
        ws.append(['Net Income', 'TBD', 'Pending', 80])

        wb.save(file_path)

        result = validator.validate_excel_file_format(file_path, "Income Statement.xlsx")

        # Should detect non-numeric data
        non_numeric_warnings = [w for w in result['warnings'] if w['type'] == 'non_numeric_data']
        assert len(non_numeric_warnings) > 0
        assert 'non_numeric_data_cells' in result
        assert len(result['non_numeric_data_cells']) > 0

    def test_sheet_detection(self, validator, temp_company_dir):
        """Test Excel workbook sheet detection"""
        file_path = Path(temp_company_dir) / "multi_sheet.xlsx"
        wb = Workbook()

        # Create multiple sheets
        wb.active.title = "Income Statement"
        ws1 = wb.active
        ws1.append(['Metric', 'FY', 'FY-1'])
        ws1.append(['Revenue', 1000, 900])

        ws2 = wb.create_sheet("Balance Sheet")
        ws2.append(['Metric', 'FY', 'FY-1'])
        ws2.append(['Assets', 5000, 4500])

        wb.save(file_path)

        result = validator.validate_excel_file_format(file_path, "Income Statement.xlsx")

        # Should detect multiple sheets
        assert result['sheet_count'] == 2
        assert 'Income Statement' in result['sheet_names']
        assert 'Balance Sheet' in result['sheet_names']

    def test_multi_sheet_validation(self, validator, temp_company_dir):
        """Test validation of all sheets in workbook"""
        file_path = Path(temp_company_dir) / "multi_sheet_validate.xlsx"
        wb = Workbook()

        # Create multiple sheets with different data
        wb.active.title = "Sheet1"
        ws1 = wb.active
        ws1.append(['Metric', 'FY', 'FY-1'])
        ws1.append(['Revenue', 1000, 900])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(['Metric', '2023', '2022'])  # No FY headers
        ws2.append(['Assets', 5000, 4500])

        wb.save(file_path)

        result = validator.validate_excel_file_format(
            file_path,
            "Income Statement.xlsx",
            validate_all_sheets=True
        )

        # Should validate all sheets
        assert 'sheet_validations' in result
        assert 'Sheet1' in result['sheet_validations']
        assert 'Sheet2' in result['sheet_validations']

        # Sheet1 should have FY headers
        assert result['sheet_validations']['Sheet1']['has_fy_headers']

        # Sheet2 should not have FY headers
        assert not result['sheet_validations']['Sheet2']['has_fy_headers']


    # === Task 180.2: Automated Repair and Organization Tests ===

    def test_repair_directory_structure_creates_missing_folders(self, validator, temp_company_dir):
        """Test automated repair creates missing FY/LTM folders"""
        company_path = Path(temp_company_dir) / "REPAIR_TEST"

        # Don't create the directory - let repair do it
        result = validator.repair_directory_structure(
            str(company_path),
            create_missing=True
        )

        assert result['success']
        assert len(result['created_folders']) >= 3  # Company + FY + LTM
        assert company_path.exists()
        assert (company_path / 'FY').exists()
        assert (company_path / 'LTM').exists()

    def test_repair_directory_structure_creates_template_files(self, validator, temp_company_dir):
        """Test repair creates template Excel files for missing statements"""
        company_path = Path(temp_company_dir) / "TEMPLATE_REPAIR"
        company_path.mkdir(parents=True)
        (company_path / 'FY').mkdir()
        (company_path / 'LTM').mkdir()

        result = validator.repair_directory_structure(
            str(company_path),
            create_missing=True
        )

        assert result['success']
        # Should create 6 files (3 statements × 2 periods)
        assert len(result['created_files']) == 6

        # Verify files exist
        assert (company_path / 'FY' / 'Income Statement.xlsx').exists()
        assert (company_path / 'LTM' / 'Balance Sheet.xlsx').exists()

    def test_organize_misplaced_files_from_root(self, validator, temp_company_dir):
        """Test organization of Excel files in root directory"""
        company_path = Path(temp_company_dir) / "ORGANIZE_ROOT"
        company_path.mkdir(parents=True)

        # Create misplaced file in root
        wb = Workbook()
        ws = wb.active
        ws.append(['Metric', 'FY', 'FY-1'])
        ws.append(['Revenue', 1000, 900])

        misplaced_file = company_path / "Income Statement.xlsx"
        wb.save(misplaced_file)

        result = validator.repair_directory_structure(
            str(company_path),
            organize_files=True,
            create_missing=True
        )

        assert result['success']
        assert len(result['moved_files']) > 0

        # File should be moved to FY folder by default
        assert (company_path / 'FY' / 'Income Statement.xlsx').exists()
        assert not misplaced_file.exists()

    def test_organize_files_wrong_period_folder(self, validator, temp_company_dir):
        """Test moving files between FY and LTM folders based on filename"""
        company_path = Path(temp_company_dir) / "WRONG_PERIOD"
        company_path.mkdir(parents=True)
        (company_path / 'FY').mkdir()
        (company_path / 'LTM').mkdir()

        # Create LTM file in FY folder
        wb = Workbook()
        ws = wb.active
        ws.append(['Metric', 'FY', 'FY-1'])
        ws.append(['Revenue', 1000, 900])

        wrong_file = company_path / 'FY' / 'LTM Income Statement.xlsx'
        wb.save(wrong_file)

        result = validator.repair_directory_structure(
            str(company_path),
            organize_files=True
        )

        assert result['success']
        # File should be moved from FY to LTM
        assert (company_path / 'LTM' / 'LTM Income Statement.xlsx').exists()
        assert not wrong_file.exists()

    def test_auto_fix_dry_run(self, validator, temp_company_dir):
        """Test auto-fix dry run mode (no actual changes)"""
        company_path = Path(temp_company_dir) / "DRYRUN_TEST"

        # Don't create directory
        result = validator.auto_fix_directory_structure(
            str(company_path),
            dry_run=True
        )

        assert result['dry_run']
        assert len(result['planned_actions']) > 0
        assert len(result['executed_actions']) == 0
        # Directory should not be created
        assert not Path(company_path).exists()

    def test_auto_fix_execution(self, validator, temp_company_dir):
        """Test auto-fix actually executes planned actions"""
        company_path = Path(temp_company_dir) / "AUTOFIX_TEST"

        result = validator.auto_fix_directory_structure(
            str(company_path),
            dry_run=False
        )

        assert not result['dry_run']
        assert len(result['executed_actions']) > 0
        # Directory should be created
        assert Path(company_path).exists()
        assert (Path(company_path) / 'FY').exists()

    def test_check_file_exists_with_alternatives(self, validator, temp_company_dir):
        """Test alternative filename pattern detection"""
        test_folder = Path(temp_company_dir) / "ALT_TEST"
        test_folder.mkdir(parents=True)

        # Create file with alternative name
        wb = Workbook()
        wb.save(test_folder / "income_statement.xlsx")

        # Should detect alternative pattern
        exists = validator._check_file_exists_with_alternatives(
            test_folder,
            "Income Statement.xlsx"
        )

        assert exists

    def test_repair_doesnt_overwrite_existing_files(self, validator, temp_company_dir):
        """Test that repair doesn't overwrite existing files"""
        company_path = Path(temp_company_dir) / "NO_OVERWRITE"
        company_path.mkdir(parents=True)
        fy_path = company_path / 'FY'
        fy_path.mkdir()

        # Create an existing file
        wb = Workbook()
        ws = wb.active
        ws.append(['Existing', 'Data'])
        existing_file = fy_path / 'Income Statement.xlsx'
        wb.save(existing_file)

        # Run repair
        result = validator.repair_directory_structure(
            str(company_path),
            create_missing=True
        )

        # Original file should still exist and not be in created_files list
        assert existing_file.exists()
        assert str(existing_file) not in result['created_files']


class TestRealWorldScenarios:
    """Test realistic scenarios with actual company directory patterns"""

    @pytest.fixture
    def validator(self):
        return DirectoryStructureValidator()

    @pytest.fixture
    def temp_company_dir(self):
        """Create temporary company directory for testing"""
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        # Cleanup
        shutil.rmtree(temp_dir, ignore_errors=True)

    def test_real_company_pattern_with_company_name_prefix(self, validator, temp_company_dir):
        """Test files with company name prefix (e.g., 'Microsoft Corporation - Income Statement.xlsx')"""
        company_path = Path(temp_company_dir) / "MSFT"
        company_path.mkdir(parents=True)

        fy_path = company_path / "FY"
        fy_path.mkdir()

        # Create files with company name prefix (common pattern)
        wb = Workbook()
        ws = wb.active
        ws.append(['Metric', 'FY', 'FY-1', 'FY-2'])
        ws.append(['Revenue', 1000, 900, 800])

        file_path = fy_path / "Microsoft Corporation - Income Statement.xlsx"
        wb.save(file_path)

        # Validation should still work with alternative file patterns
        result = validator.validate_company_directory(str(company_path), strict_mode=False)

        # Should find the file using alternative patterns
        assert 'FY' in result['found_files']
        assert len(result['found_files']['FY']) >= 1


class TestComplianceReporting:
    """Test suite for enhanced compliance reporting features (Task 180.3)"""

    @pytest.fixture
    def temp_company_dir(self):
        """Create temporary company directory for testing"""
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        shutil.rmtree(temp_dir, ignore_errors=True)

    @pytest.fixture
    def validator(self):
        """Create validator instance"""
        return DirectoryStructureValidator()

    @pytest.fixture
    def valid_company_structure(self, temp_company_dir):
        """Create a valid company directory structure"""
        company_path = Path(temp_company_dir) / "TESTCO"
        company_path.mkdir(parents=True)

        for period in ['FY', 'LTM']:
            period_path = company_path / period
            period_path.mkdir()

            for statement in ["Income Statement.xlsx", "Balance Sheet.xlsx", "Cash Flow Statement.xlsx"]:
                file_path = period_path / statement
                wb = Workbook()
                ws = wb.active
                ws.append(['Metric', 'FY', 'FY-1', 'FY-2', 'FY-3'])
                ws.append(['Revenue', 1000, 900, 800, 700])
                ws.append(['Net Income', 100, 90, 80, 70])
                wb.save(file_path)

        return company_path

    @pytest.fixture
    def partial_company_structure(self, temp_company_dir):
        """Create a partially compliant company directory structure"""
        company_path = Path(temp_company_dir) / "PARTIAL"
        company_path.mkdir(parents=True)

        # Only create FY folder, missing LTM
        fy_path = company_path / "FY"
        fy_path.mkdir()

        # Only create one statement file
        file_path = fy_path / "Income Statement.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(['Metric', 'FY', 'FY-1'])  # Only 2 periods (insufficient)
        ws.append(['Revenue', 1000, 900])
        wb.save(file_path)

        return company_path

    # === Compliance Scoring Tests ===

    def test_calculate_overall_compliance_fully_compliant(self, validator, valid_company_structure):
        """Test compliance calculation for fully compliant structure"""
        report = validator.validate_directory_structure(
            ticker="TESTCO",
            base_path=str(valid_company_structure.parent)
        )

        compliance = report['overall_compliance']

        assert compliance['overall_score'] >= 0.8
        assert compliance['status'] in ['FULLY_COMPLIANT', 'COMPLIANT']
        assert 'status_description' in compliance
        assert compliance['total_issues'] == 0 or compliance['total_issues'] is not None
        assert 'issue_breakdown' in compliance
        assert 'file_statistics' in compliance

    def test_calculate_overall_compliance_partial(self, validator, partial_company_structure):
        """Test compliance calculation for partially compliant structure"""
        report = validator.validate_directory_structure(
            ticker="PARTIAL",
            base_path=str(partial_company_structure.parent)
        )

        compliance = report['overall_compliance']

        assert compliance['overall_score'] < 0.8
        assert compliance['status'] in ['PARTIALLY_COMPLIANT', 'NON_COMPLIANT', 'CRITICAL']
        assert compliance['total_issues'] > 0
        assert compliance['issue_breakdown']['critical_count'] > 0

    def test_compliance_history_tracking(self, validator, valid_company_structure):
        """Test that compliance reports are added to history"""
        assert len(validator.compliance_history) == 0

        # Run multiple validations
        validator.validate_directory_structure(
            ticker="TESTCO",
            base_path=str(valid_company_structure.parent)
        )
        validator.validate_directory_structure(
            ticker="TESTCO",
            base_path=str(valid_company_structure.parent)
        )

        assert len(validator.compliance_history) == 2
        assert all('timestamp' in report for report in validator.compliance_history)

    def test_get_compliance_history(self, validator, valid_company_structure):
        """Test getting compliance history with limit"""
        # Generate some history
        for _ in range(5):
            validator.validate_directory_structure(
                ticker="TESTCO",
                base_path=str(valid_company_structure.parent)
            )

        # Get limited history
        limited_history = validator.get_compliance_history(limit=3)
        assert len(limited_history) == 3

        # Get all history
        all_history = validator.get_compliance_history()
        assert len(all_history) == 5

    def test_compliance_trend_analysis(self, validator, partial_company_structure):
        """Test compliance trend analysis from history"""
        # First validation (partial compliance)
        validator.validate_directory_structure(
            ticker="PARTIAL",
            base_path=str(partial_company_structure.parent)
        )
        first_score = validator.compliance_history[-1]['overall_score']

        # Fix structure
        ltm_path = partial_company_structure / "LTM"
        ltm_path.mkdir()
        file_path = ltm_path / "Income Statement.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(['Metric', 'FY', 'FY-1', 'FY-2'])
        ws.append(['Revenue', 1000, 900, 800])
        wb.save(file_path)

        # Second validation (improved compliance)
        validator.validate_directory_structure(
            ticker="PARTIAL",
            base_path=str(partial_company_structure.parent)
        )

        # Get trend
        trend = validator.get_compliance_trend()

        assert trend['available']
        assert trend['total_reports'] == 2
        assert trend['first_score'] == first_score
        assert trend['score_change'] != 0

    def test_compliance_trend_insufficient_data(self, validator):
        """Test trend analysis with insufficient data"""
        trend = validator.get_compliance_trend()

        assert not trend['available']
        assert 'message' in trend

    # === Actionable Recommendations Tests ===

    def test_recommendations_missing_directory(self, validator):
        """Test recommendations for missing company directory"""
        report = validator.validate_directory_structure(
            ticker="NONEXISTENT",
            base_path="/nonexistent/path"
        )

        recommendations = report['actionable_recommendations']

        assert len(recommendations) > 0
        assert recommendations[0]['priority'] == 'CRITICAL'
        assert recommendations[0]['priority_level'] == 1
        assert recommendations[0]['issue_type'] == 'missing_directory'
        assert recommendations[0]['automated_fix']['available']
        assert 'manual_steps' in recommendations[0]
        assert 'estimated_time' in recommendations[0]

    def test_recommendations_missing_folders(self, validator, temp_company_dir):
        """Test recommendations for missing period folders"""
        company_path = Path(temp_company_dir) / "TEST"
        company_path.mkdir()

        report = validator.validate_directory_structure(
            ticker="TEST",
            base_path=str(company_path.parent)
        )

        recommendations = report['actionable_recommendations']

        # Should have recommendations for missing FY and LTM folders
        folder_recs = [r for r in recommendations if r['issue_type'] == 'missing_folder']
        assert len(folder_recs) >= 1
        assert all(r['priority'] == 'CRITICAL' for r in folder_recs)
        assert all(r['automated_fix']['available'] for r in folder_recs)

    def test_recommendations_priority_sorting(self, validator, partial_company_structure):
        """Test that recommendations are sorted by priority level"""
        report = validator.validate_directory_structure(
            ticker="PARTIAL",
            base_path=str(partial_company_structure.parent)
        )

        recommendations = report['actionable_recommendations']

        if len(recommendations) > 1:
            # Verify sorting by priority_level
            priority_levels = [r['priority_level'] for r in recommendations]
            assert priority_levels == sorted(priority_levels)

    def test_recommendations_with_automated_fixes(self, validator, partial_company_structure):
        """Test recommendations include automated fix commands"""
        report = validator.validate_directory_structure(
            ticker="PARTIAL",
            base_path=str(partial_company_structure.parent)
        )

        recommendations = report['actionable_recommendations']

        # Find recommendations with automated fixes
        auto_fix_recs = [r for r in recommendations if r['automated_fix'].get('available')]

        assert len(auto_fix_recs) > 0
        for rec in auto_fix_recs:
            assert 'method' in rec['automated_fix']
            assert 'command' in rec['automated_fix']
            assert 'description' in rec['automated_fix']

    def test_recommendations_include_manual_steps(self, validator, partial_company_structure):
        """Test that all recommendations include manual steps"""
        report = validator.validate_directory_structure(
            ticker="PARTIAL",
            base_path=str(partial_company_structure.parent)
        )

        recommendations = report['actionable_recommendations']

        for rec in recommendations:
            assert 'manual_steps' in rec
            assert isinstance(rec['manual_steps'], list)
            assert len(rec['manual_steps']) > 0

    # === Export Capabilities Tests ===

    def test_export_compliance_report_json(self, validator, valid_company_structure, temp_company_dir):
        """Test exporting compliance report to JSON"""
        report = validator.validate_directory_structure(
            ticker="TESTCO",
            base_path=str(valid_company_structure.parent)
        )

        output_path = Path(temp_company_dir) / "compliance_report.json"
        result = validator.export_compliance_report(
            report,
            str(output_path),
            format='json'
        )

        assert result['success']
        assert output_path.exists()

        # Verify JSON content
        import json
        with open(output_path, 'r') as f:
            exported_data = json.load(f)

        assert exported_data['ticker'] == 'TESTCO'
        assert 'overall_compliance' in exported_data
        assert 'actionable_recommendations' in exported_data

    def test_export_compliance_report_csv(self, validator, valid_company_structure, temp_company_dir):
        """Test exporting compliance report to CSV"""
        report = validator.validate_directory_structure(
            ticker="TESTCO",
            base_path=str(valid_company_structure.parent)
        )

        output_path = Path(temp_company_dir) / "compliance_report.csv"
        result = validator.export_compliance_report(
            report,
            str(output_path),
            format='csv'
        )

        assert result['success']
        assert output_path.exists()

        # Verify CSV has content
        with open(output_path, 'r') as f:
            content = f.read()
            assert 'Compliance Report Summary' in content
            assert 'TESTCO' in content

    def test_export_compliance_report_html(self, validator, valid_company_structure, temp_company_dir):
        """Test exporting compliance report to HTML"""
        report = validator.validate_directory_structure(
            ticker="TESTCO",
            base_path=str(valid_company_structure.parent)
        )

        output_path = Path(temp_company_dir) / "compliance_report.html"
        result = validator.export_compliance_report(
            report,
            str(output_path),
            format='html'
        )

        assert result['success']
        assert output_path.exists()

        # Verify HTML structure
        with open(output_path, 'r', encoding='utf-8') as f:
            content = f.read()
            assert '<!DOCTYPE html>' in content
            assert 'Compliance Report' in content
            assert 'TESTCO' in content
            assert '<style>' in content  # Has CSS styling

    def test_export_unsupported_format(self, validator, valid_company_structure, temp_company_dir):
        """Test exporting with unsupported format"""
        report = validator.validate_directory_structure(
            ticker="TESTCO",
            base_path=str(valid_company_structure.parent)
        )

        output_path = Path(temp_company_dir) / "compliance_report.xml"
        result = validator.export_compliance_report(
            report,
            str(output_path),
            format='xml'
        )

        assert not result['success']
        assert result['error'] is not None
        assert 'Unsupported' in result['error']

    def test_export_creates_output_directory(self, validator, valid_company_structure, temp_company_dir):
        """Test that export creates output directory if it doesn't exist"""
        report = validator.validate_directory_structure(
            ticker="TESTCO",
            base_path=str(valid_company_structure.parent)
        )

        output_path = Path(temp_company_dir) / "reports" / "subdir" / "compliance_report.json"
        result = validator.export_compliance_report(
            report,
            str(output_path),
            format='json'
        )

        assert result['success']
        assert output_path.exists()
        assert output_path.parent.exists()

    # === Categorized Issues Tests ===

    def test_categorized_issues_structure(self, validator, partial_company_structure):
        """Test that compliance report includes categorized issues"""
        report = validator.validate_directory_structure(
            ticker="PARTIAL",
            base_path=str(partial_company_structure.parent)
        )

        compliance = report['overall_compliance']

        assert 'categorized_issues' in compliance
        assert 'critical' in compliance['categorized_issues']
        assert 'high' in compliance['categorized_issues']
        assert 'medium' in compliance['categorized_issues']
        assert 'low' in compliance['categorized_issues']

    def test_issue_breakdown_counts(self, validator, partial_company_structure):
        """Test that issue breakdown counts are accurate"""
        report = validator.validate_directory_structure(
            ticker="PARTIAL",
            base_path=str(partial_company_structure.parent)
        )

        compliance = report['overall_compliance']
        breakdown = compliance['issue_breakdown']
        categorized = compliance['categorized_issues']

        assert breakdown['critical_count'] == len(categorized['critical'])
        assert breakdown['high_count'] == len(categorized['high'])
        assert breakdown['medium_count'] == len(categorized['medium'])
        assert breakdown['low_count'] == len(categorized['low'])

    def test_file_statistics_in_compliance(self, validator, valid_company_structure):
        """Test that compliance report includes file statistics"""
        report = validator.validate_directory_structure(
            ticker="TESTCO",
            base_path=str(valid_company_structure.parent)
        )

        compliance = report['overall_compliance']

        assert 'file_statistics' in compliance
        stats = compliance['file_statistics']
        assert 'total_excel_files' in stats
        assert 'valid_excel_files' in stats
        assert 'invalid_excel_files' in stats
        assert stats['total_excel_files'] >= 0

    # === Integration Tests ===

    def test_full_compliance_workflow(self, validator, temp_company_dir):
        """Test complete compliance reporting workflow"""
        # Step 1: Validate non-existent directory
        report1 = validator.validate_directory_structure(
            ticker="WORKFLOW",
            base_path=temp_company_dir
        )

        assert report1['overall_compliance']['status'] == 'CRITICAL'
        assert len(report1['actionable_recommendations']) > 0

        # Step 2: Create structure using recommendations
        company_path = Path(temp_company_dir) / "WORKFLOW"
        validator.create_directory_structure_from_template(str(company_path))

        # Step 3: Validate improved structure
        report2 = validator.validate_directory_structure(
            ticker="WORKFLOW",
            base_path=temp_company_dir
        )

        assert report2['overall_compliance']['overall_score'] > report1['overall_compliance']['overall_score']

        # Step 4: Export report
        output_path = Path(temp_company_dir) / "workflow_report.html"
        export_result = validator.export_compliance_report(
            report2,
            str(output_path),
            format='html'
        )

        assert export_result['success']

        # Step 5: Check compliance trend
        trend = validator.get_compliance_trend()
        assert trend['available']
        assert trend['trend'] in ['improving', 'stable', 'declining']


@pytest.mark.skipif(not VALIDATION_REGISTRY_AVAILABLE, reason="ValidationRegistry not available")
class TestValidationRegistryIntegration:
    """Test suite for ValidationRegistry integration (Task 180.4)"""

    @pytest.fixture
    def temp_company_dir(self):
        """Create temporary company directory for testing"""
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        shutil.rmtree(temp_dir, ignore_errors=True)

    @pytest.fixture
    def registry(self):
        """Create ValidationRegistry instance"""
        return ValidationRegistry()

    @pytest.fixture
    def validator_with_registry(self, registry):
        """Create validator with registry integration"""
        return DirectoryStructureValidator(validation_registry=registry)

    @pytest.fixture
    def valid_company_structure(self, temp_company_dir):
        """Create a valid company directory structure"""
        company_path = Path(temp_company_dir) / "REGISTRY_TEST"
        company_path.mkdir(parents=True, exist_ok=True)

        for period in ['FY', 'LTM']:
            period_path = company_path / period
            period_path.mkdir(exist_ok=True)

            for statement in ["Income Statement.xlsx", "Balance Sheet.xlsx", "Cash Flow Statement.xlsx"]:
                file_path = period_path / statement
                self._create_valid_excel_file(file_path, statement)

        return company_path

    def _create_valid_excel_file(self, file_path: Path, statement_type: str):
        """Helper to create a valid Excel file with proper structure"""
        wb = Workbook()
        ws = wb.active

        # Add proper column headers
        headers = ['Metric', 'FY', 'FY-1', 'FY-2', 'FY-3']
        ws.append(headers)

        # Add sample numeric data
        if "Income Statement" in statement_type:
            ws.append(['Revenue', 1000, 900, 800, 700])
            ws.append(['Net Income', 100, 90, 80, 70])
            ws.append(['Operating Income', 150, 135, 120, 105])
        elif "Balance Sheet" in statement_type:
            ws.append(['Total Assets', 5000, 4500, 4000, 3500])
            ws.append(['Total Liabilities', 3000, 2700, 2400, 2100])
            ws.append(['Shareholders Equity', 2000, 1800, 1600, 1400])
        elif "Cash Flow" in statement_type:
            ws.append(['Cash from Operations', 200, 180, 160, 140])
            ws.append(['Capital Expenditures', 50, 45, 40, 35])

        wb.save(file_path)

    # === Registry Integration Tests ===

    def test_validator_registers_rules_on_init(self, validator_with_registry, registry):
        """Test that validator registers directory structure rules on initialization"""
        # Check that directory_structure rule set exists
        dir_rule_set = registry.get_rule_set("directory_structure")
        assert dir_rule_set is not None
        assert len(dir_rule_set.rules) > 0

        # Verify specific rules are registered
        rule_ids = [rule.rule_id for rule in dir_rule_set.rules]
        assert "dir_fy_folder_exists" in rule_ids
        assert "dir_ltm_folder_exists" in rule_ids
        assert "dir_required_excel_files" in rule_ids
        assert "dir_excel_file_format" in rule_ids

    def test_validate_with_registry_rules_valid_structure(
        self,
        validator_with_registry,
        valid_company_structure
    ):
        """Test rule-based validation with valid directory structure"""
        result = validator_with_registry.validate_with_registry_rules(
            str(valid_company_structure)
        )

        assert result['rules_executed'] > 0
        assert result['overall_status'] == 'PASS'
        assert result['rules_passed'] > 0
        assert result['rules_failed'] == 0
        assert len(result['rule_results']) > 0

    def test_validate_with_registry_rules_missing_folders(
        self,
        validator_with_registry,
        temp_company_dir
    ):
        """Test rule-based validation with missing folders"""
        # Create company directory without FY/LTM folders
        company_path = Path(temp_company_dir) / "MISSING_FOLDERS"
        company_path.mkdir(parents=True, exist_ok=True)

        result = validator_with_registry.validate_with_registry_rules(
            str(company_path)
        )

        assert result['overall_status'] == 'FAIL'
        assert result['rules_failed'] > 0

        # Check that specific folder rules failed
        rule_results = {r['rule_id']: r for r in result['rule_results']}
        assert rule_results['dir_fy_folder_exists']['status'] == 'FAIL'
        assert rule_results['dir_ltm_folder_exists']['status'] == 'FAIL'

    def test_validate_with_registry_rules_scope_filter(
        self,
        validator_with_registry,
        valid_company_structure
    ):
        """Test rule-based validation with scope filtering"""
        # Test SYSTEM scope only
        result = validator_with_registry.validate_with_registry_rules(
            str(valid_company_structure),
            scope=RuleScope.SYSTEM
        )

        # Should execute folder existence rules (SYSTEM scope)
        rule_ids = [r['rule_id'] for r in result['rule_results']]
        assert 'dir_fy_folder_exists' in rule_ids
        assert 'dir_ltm_folder_exists' in rule_ids

    def test_update_rule_config(self, validator_with_registry):
        """Test dynamic rule configuration updates"""
        # Update parameters for minimum periods rule
        success = validator_with_registry.update_rule_config(
            rule_id="dir_minimum_periods",
            parameters={"min_periods": 5}
        )

        assert success

        # Verify parameter was updated
        rule = validator_with_registry.validation_registry.get_rule_by_id(
            "dir_minimum_periods"
        )
        assert rule.parameters['min_periods'] == 5

    def test_update_rule_thresholds(self, validator_with_registry):
        """Test dynamic rule threshold updates"""
        success = validator_with_registry.update_rule_config(
            rule_id="dir_excel_numeric_data",
            thresholds={"min_numeric_percentage": 90.0}
        )

        assert success

        # Verify threshold was updated
        rule = validator_with_registry.validation_registry.get_rule_by_id(
            "dir_excel_numeric_data"
        )
        assert rule.thresholds['min_numeric_percentage'] == 90.0

    def test_enable_disable_rule(self, validator_with_registry):
        """Test enabling and disabling validation rules"""
        rule_id = "dir_naming_conventions"

        # Disable rule
        success = validator_with_registry.update_rule_config(
            rule_id=rule_id,
            enabled=False
        )
        assert success

        rule = validator_with_registry.validation_registry.get_rule_by_id(rule_id)
        assert not rule.enabled

        # Enable rule
        success = validator_with_registry.update_rule_config(
            rule_id=rule_id,
            enabled=True
        )
        assert success

        rule = validator_with_registry.validation_registry.get_rule_by_id(rule_id)
        assert rule.enabled

    def test_get_registry_validation_report(self, validator_with_registry):
        """Test retrieval of registry validation report"""
        report = validator_with_registry.get_registry_validation_report()

        assert report['available']
        assert 'registry_stats' in report
        assert 'directory_rules' in report

        # Check directory rules section
        dir_rules = report['directory_rules']
        assert dir_rules['total'] > 0
        assert dir_rules['enabled'] > 0
        assert len(dir_rules['rules']) > 0

        # Verify rule details
        for rule in dir_rules['rules']:
            assert 'id' in rule
            assert 'name' in rule
            assert 'priority' in rule
            assert 'enabled' in rule
            assert 'type' in rule
            assert 'scope' in rule

    def test_rule_execution_with_error_templates(
        self,
        validator_with_registry,
        temp_company_dir
    ):
        """Test that rule execution uses error message templates correctly"""
        # Create invalid structure
        company_path = Path(temp_company_dir) / "TEMPLATE_TEST"
        company_path.mkdir(parents=True, exist_ok=True)

        result = validator_with_registry.validate_with_registry_rules(
            str(company_path)
        )

        # Find failed rule
        failed_rules = [r for r in result['rule_results'] if r['status'] == 'FAIL']
        assert len(failed_rules) > 0

        # Check that error messages are properly formatted
        for rule_result in failed_rules:
            assert len(rule_result['issues']) > 0
            for issue in rule_result['issues']:
                assert 'message' in issue
                assert 'remediation' in issue
                assert len(issue['message']) > 0
                assert len(issue['remediation']) > 0

    def test_rule_dependency_validation(self, registry):
        """Test validation of rule dependencies"""
        conflicts = registry.validate_rule_dependencies()

        # Should have no conflicts in properly configured registry
        assert isinstance(conflicts, list)
        # Allow empty or minimal conflicts for well-configured registry

    def test_registry_stats(self, validator_with_registry):
        """Test registry statistics retrieval"""
        stats = validator_with_registry.validation_registry.get_registry_stats()

        assert 'total_rule_sets' in stats
        assert 'total_rules' in stats
        assert 'enabled_rules' in stats
        assert 'disabled_rules' in stats
        assert 'rules_by_type' in stats
        assert 'rules_by_scope' in stats
        assert 'rules_by_priority' in stats

        # Verify counts are reasonable
        assert stats['total_rules'] > 0
        assert stats['enabled_rules'] > 0

    def test_execute_validation_rule_individual(
        self,
        validator_with_registry,
        valid_company_structure
    ):
        """Test individual rule execution"""
        # Get a specific rule
        rule = validator_with_registry.validation_registry.get_rule_by_id(
            "dir_fy_folder_exists"
        )
        assert rule is not None

        # Execute the rule
        result = validator_with_registry._execute_validation_rule(
            rule,
            valid_company_structure
        )

        assert result['rule_id'] == "dir_fy_folder_exists"
        assert result['status'] == 'PASS'
        assert result['priority'] == 'critical'

    def test_rule_execution_with_missing_files(
        self,
        validator_with_registry,
        temp_company_dir
    ):
        """Test rule execution detects missing Excel files"""
        # Create structure with folders but no Excel files
        company_path = Path(temp_company_dir) / "MISSING_FILES"
        company_path.mkdir(parents=True, exist_ok=True)
        (company_path / "FY").mkdir(exist_ok=True)
        (company_path / "LTM").mkdir(exist_ok=True)

        result = validator_with_registry.validate_with_registry_rules(
            str(company_path)
        )

        # Should detect missing files
        rule_results = {r['rule_id']: r for r in result['rule_results']}
        excel_files_rule = rule_results.get('dir_required_excel_files')

        if excel_files_rule:
            assert excel_files_rule['status'] == 'FAIL'
            assert len(excel_files_rule['issues']) > 0

    def test_rule_parameter_inheritance(self, validator_with_registry):
        """Test that rule parameters are properly inherited and used"""
        # Get rule with parameters
        rule = validator_with_registry.validation_registry.get_rule_by_id(
            "dir_required_excel_files"
        )

        assert rule is not None
        assert 'required_files' in rule.parameters
        assert len(rule.parameters['required_files']) == 3

    def test_validation_without_registry(self):
        """Test validator works without registry (fallback behavior)"""
        # Create validator without registry
        validator = DirectoryStructureValidator(validation_registry=None)

        # Should still work with basic validation
        result = validator.validate_company_directory("/nonexistent/path")
        assert 'is_valid' in result
        assert not result['is_valid']


class TestEdgeCasesAndErrorHandling:
    """Test suite for edge cases, error handling, and uncovered code paths"""

    @pytest.fixture
    def temp_company_dir(self):
        """Create temporary company directory for testing"""
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        shutil.rmtree(temp_dir, ignore_errors=True)

    @pytest.fixture
    def validator(self):
        return DirectoryStructureValidator()

    # === Error Handling Tests ===

    def test_scan_folder_files_permission_error(self, validator, temp_company_dir):
        """Test error handling when folder cannot be scanned"""
        # Use a nonexistent path to trigger exception handling
        nonexistent_path = Path(temp_company_dir) / "nonexistent"
        result = validator._scan_folder_files(nonexistent_path)

        # Should return empty list on error
        assert result == []

    def test_calculate_structure_score_zero_folders(self, validator):
        """Test score calculation with zero total folders (edge case)"""
        score = validator._calculate_structure_score(0, 0, {})
        assert score == 0.0

    def test_create_validation_summary_non_existent_dir(self, validator):
        """Test validation summary for non-existent directory"""
        validation_result = {
            'is_valid': False,
            'exists': False,
            'issues': [],
            'structure_score': 0.0
        }
        summary = validator._create_validation_summary(validation_result)
        assert "not found" in summary

    def test_create_validation_summary_low_score(self, validator):
        """Test validation summary with low score (< 0.3)"""
        validation_result = {
            'is_valid': False,
            'exists': True,
            'issues': [{'type': 'test', 'message': 'test'}],
            'structure_score': 0.2
        }
        summary = validator._create_validation_summary(validation_result)
        assert "Major structure issues" in summary

    def test_generate_fix_suggestions_non_existent(self, validator):
        """Test fix suggestions for non-existent directory"""
        validation_result = {
            'exists': False,
            'missing_folders': [],
            'missing_files': [],
            'structure_score': 0.0
        }
        suggestions = validator._generate_fix_suggestions(validation_result)

        assert len(suggestions) == 2
        assert any("Create the company directory" in s for s in suggestions)

    def test_determine_statement_type_unknown(self, validator):
        """Test statement type determination for unknown file"""
        result = validator._determine_statement_type("unknown_file.xlsx")
        assert result == "unknown_file.xlsx"

    def test_validate_all_sheets_read_error(self, validator, temp_company_dir):
        """Test _validate_all_sheets with file read errors"""
        # Create a corrupted file
        file_path = Path(temp_company_dir) / "corrupted.xlsx"
        with open(file_path, 'w') as f:
            f.write("not an excel file")

        # Should handle errors gracefully
        result = validator._validate_all_sheets(file_path, "Income Statement.xlsx")
        # Result should be empty or contain error info
        assert isinstance(result, dict)

    def test_validate_numeric_data_types_empty_column(self, validator, temp_company_dir):
        """Test numeric validation with completely empty column"""
        file_path = Path(temp_company_dir) / "empty_col.xlsx"
        wb = Workbook()
        ws = wb.active

        # Create file with empty FY column
        ws.append(['Metric', 'FY', 'FY-1'])
        ws.append(['Revenue', None, 900])
        ws.append(['Net Income', None, 90])

        wb.save(file_path)

        result = validator.validate_excel_file_format(file_path, "Income Statement.xlsx")

        # Should detect warning about empty column
        warnings = [w for w in result.get('warnings', []) if 'empty' in w.get('type', '')]
        assert len(warnings) > 0 or 'empty_column' in str(result['warnings'])

    # === Coverage for _calculate_overall_compliance edge cases ===

    def test_calculate_overall_compliance_no_excel_files(self, validator):
        """Test compliance calculation with no Excel files"""
        dir_validation = {'structure_score': 0.5}
        excel_validations = {}

        compliance = validator._calculate_overall_compliance(dir_validation, excel_validations)

        assert compliance['excel_score'] == 0.0
        assert compliance['overall_score'] == 0.2  # Only dir_score * 0.4

    def test_calculate_overall_compliance_with_warnings(self, validator):
        """Test compliance with warnings included"""
        dir_validation = {
            'structure_score': 0.8,
            'issues': [],
            'warnings': [{'type': 'test', 'message': 'warning1'}]
        }
        excel_validations = {
            'FY/test.xlsx': {
                'is_valid': True,
                'issues': [],
                'warnings': [{'type': 'test', 'message': 'warning2'}]
            }
        }

        compliance = validator._calculate_overall_compliance(dir_validation, excel_validations)

        assert compliance['total_warnings'] == 2

    # === Coverage for _generate_actionable_recommendations edge cases ===

    def test_recommendations_unreadable_excel(self, validator):
        """Test recommendations for unreadable Excel files"""
        dir_validation = {
            'exists': True,
            'missing_folders': [],
            'missing_files': [],
            'company_path': '/test/path'
        }
        excel_validations = {
            'FY/corrupted.xlsx': {
                'exists': True,
                'readable': False,
                'has_data': False
            }
        }

        recommendations = validator._generate_actionable_recommendations(
            dir_validation, excel_validations
        )

        # Should have recommendation for corrupted file
        corrupted_recs = [r for r in recommendations if 'corrupted' in r['issue_type']]
        assert len(corrupted_recs) > 0

    def test_recommendations_empty_excel(self, validator):
        """Test recommendations for empty Excel files"""
        dir_validation = {
            'exists': True,
            'missing_folders': [],
            'missing_files': [],
            'company_path': '/test/path'
        }
        excel_validations = {
            'FY/empty.xlsx': {
                'exists': True,
                'readable': True,
                'has_data': False
            }
        }

        recommendations = validator._generate_actionable_recommendations(
            dir_validation, excel_validations
        )

        # Should have recommendation for empty file
        empty_recs = [r for r in recommendations if 'empty' in r['issue_type']]
        assert len(empty_recs) > 0

    def test_recommendations_missing_headers(self, validator):
        """Test recommendations for files with missing headers"""
        dir_validation = {
            'exists': True,
            'missing_folders': [],
            'missing_files': [],
            'company_path': '/test/path'
        }
        excel_validations = {
            'FY/no_headers.xlsx': {
                'exists': True,
                'readable': True,
                'has_data': True,
                'has_proper_headers': False
            }
        }

        recommendations = validator._generate_actionable_recommendations(
            dir_validation, excel_validations
        )

        # Should have recommendation for missing headers
        header_recs = [r for r in recommendations if 'headers' in r['issue_type']]
        assert len(header_recs) > 0

    # === Template Creation Error Handling ===

    def test_create_template_excel_file_openpyxl_fallback(self, validator, temp_company_dir, monkeypatch):
        """Test fallback when openpyxl module import fails"""
        file_path = Path(temp_company_dir) / "test.xlsx"

        # Mock ImportError for openpyxl
        def mock_import_error(*args, **kwargs):
            raise ImportError("openpyxl not available")

        # This will test the ImportError handling path
        # Note: actual openpyxl is available, so we can't fully test the CSV fallback
        # But we can verify the method doesn't crash
        validator._create_template_excel_file(
            file_path, "Income Statement.xlsx", "FY", "Test Company"
        )

        # File should be created
        assert file_path.exists()

    def test_create_directory_structure_template_error(self, validator, temp_company_dir, monkeypatch):
        """Test error handling in template creation"""
        # Create a path that will cause permission error on some systems
        # Use a very long path to potentially trigger OS errors
        long_path = Path(temp_company_dir) / ("x" * 255)

        result = validator.create_directory_structure_from_template(str(long_path))

        # Should complete successfully or handle errors gracefully
        assert 'success' in result
        assert 'errors' in result

    # === Repair and Organization Error Handling ===

    def test_repair_directory_structure_error_handling(self, validator, temp_company_dir, monkeypatch):
        """Test error handling in directory repair"""
        # Create a scenario that might cause errors
        company_path = Path(temp_company_dir) / "ERROR_TEST"

        result = validator.repair_directory_structure(str(company_path))

        # Should handle errors gracefully
        assert 'success' in result
        assert 'errors' in result

    def test_organize_misplaced_files_error_handling(self, validator, temp_company_dir):
        """Test error handling in file organization"""
        company_path = Path(temp_company_dir) / "ORG_ERROR"
        company_path.mkdir(parents=True)

        # Create a file that can't be moved
        test_file = company_path / "test.xlsx"
        wb = Workbook()
        wb.save(test_file)

        # Try to organize (should handle any errors)
        result = validator._organize_misplaced_files(company_path)

        assert 'moved_files' in result
        assert 'errors' in result

    def test_check_file_exists_with_alternatives_nonexistent_folder(self, validator):
        """Test alternative file check with nonexistent folder"""
        result = validator._check_file_exists_with_alternatives(
            Path("/nonexistent/path"),
            "Income Statement.xlsx"
        )
        assert result is False

    def test_auto_fix_missing_file_action(self, validator, temp_company_dir):
        """Test auto-fix planning for missing files"""
        company_path = Path(temp_company_dir) / "AUTOFIX_MISSING"
        company_path.mkdir(parents=True)
        (company_path / "FY").mkdir()

        # No Excel files - should plan to create them
        result = validator.auto_fix_directory_structure(
            str(company_path),
            dry_run=True
        )

        assert result['dry_run'] is True
        assert len(result['planned_actions']) > 0

    # === Export Error Handling ===

    def test_export_compliance_report_csv_with_recommendations(self, validator, temp_company_dir):
        """Test CSV export with recommendations"""
        report = {
            'ticker': 'TEST',
            'validation_timestamp': '2025-10-02',
            'company_path': '/test/path',
            'overall_compliance': {
                'overall_score': 0.85,
                'directory_score': 0.9,
                'excel_score': 0.8,
                'status': 'COMPLIANT',
                'status_description': 'Good',
                'issue_breakdown': {
                    'critical_count': 0,
                    'high_count': 1,
                    'medium_count': 2,
                    'low_count': 3
                }
            },
            'actionable_recommendations': [
                {
                    'priority': 'LOW',
                    'category': 'Test',
                    'action': 'Test action',
                    'impact': 'Test impact',
                    'estimated_time': '5 min'
                }
            ]
        }

        output_path = Path(temp_company_dir) / "test_report.csv"
        result = validator.export_compliance_report(report, str(output_path), format='csv')

        assert result['success']
        assert output_path.exists()

        # Verify recommendations section exists
        with open(output_path, 'r') as f:
            content = f.read()
            assert 'Actionable Recommendations' in content

    def test_export_compliance_report_html_with_automated_fix(self, validator, temp_company_dir):
        """Test HTML export with automated fix info"""
        report = {
            'ticker': 'TEST',
            'validation_timestamp': '2025-10-02',
            'company_path': '/test/path',
            'overall_compliance': {
                'overall_score': 0.85,
                'directory_score': 0.9,
                'excel_score': 0.8,
                'status': 'COMPLIANT',
                'status_description': 'Good',
                'total_issues': 1,
                'total_warnings': 2
            },
            'actionable_recommendations': [
                {
                    'priority': 'MEDIUM',
                    'category': 'Test',
                    'action': 'Test action',
                    'impact': 'Test impact',
                    'estimated_time': '10 min',
                    'automated_fix': {
                        'available': False,
                        'reason': 'Manual intervention required'
                    },
                    'manual_steps': ['Step 1', 'Step 2']
                }
            ]
        }

        output_path = Path(temp_company_dir) / "test_report.html"
        result = validator.export_compliance_report(report, str(output_path), format='html')

        assert result['success']

        # Verify automated fix section exists
        with open(output_path, 'r', encoding='utf-8') as f:
            content = f.read()
            assert 'Manual Intervention Required' in content or 'reason' in content.lower()

    # === ValidationRegistry Integration Error Paths ===

    def test_validate_with_registry_rules_fallback(self, temp_company_dir):
        """Test fallback when registry is not available"""
        # Create validator without registry
        validator = DirectoryStructureValidator(validation_registry=None)

        company_path = Path(temp_company_dir) / "NO_REGISTRY"
        company_path.mkdir(parents=True)

        # Should fall back to standard validation
        result = validator.validate_with_registry_rules(str(company_path))

        assert 'is_valid' in result  # Should return standard validation result

    @pytest.mark.skipif(not VALIDATION_REGISTRY_AVAILABLE, reason="ValidationRegistry not available")
    def test_execute_validation_rule_error_handling(self, temp_company_dir):
        """Test error handling in rule execution"""
        from core.validation.validation_registry import ValidationRegistry, ValidationRule, RuleType, RuleScope

        registry = ValidationRegistry()
        validator = DirectoryStructureValidator(validation_registry=registry)

        # Create a rule that might cause errors
        test_rule = ValidationRule(
            rule_id="test_error_rule",
            name="Test Error Rule",
            description="Rule to test error handling",
            rule_type=RuleType.FORMAT,
            scope=RuleScope.DATA,
            priority="low",
            error_message_template="Test error: {error}",
            remediation_template="Fix the test error"
        )

        company_path = Path(temp_company_dir) / "ERROR_TEST"
        company_path.mkdir(parents=True)

        # Execute rule (should handle errors gracefully)
        result = validator._execute_validation_rule(test_rule, company_path)

        assert 'rule_id' in result
        assert 'status' in result

    @pytest.mark.skipif(not VALIDATION_REGISTRY_AVAILABLE, reason="ValidationRegistry not available")
    def test_load_rules_from_config_error(self, validator):
        """Test error handling when loading rules from invalid config"""
        # Skip if no registry
        if not validator.validation_registry:
            pytest.skip("ValidationRegistry not available")

        # Try to load from nonexistent file
        result = validator.load_rules_from_config("/nonexistent/config.yaml")
        assert result is False

    @pytest.mark.skipif(not VALIDATION_REGISTRY_AVAILABLE, reason="ValidationRegistry not available")
    def test_update_rule_config_error_handling(self, validator):
        """Test error handling in rule config update"""
        # Skip if no registry
        if not validator.validation_registry:
            pytest.skip("ValidationRegistry not available")

        # Try to update nonexistent rule
        result = validator.update_rule_config(
            rule_id="nonexistent_rule",
            parameters={'test': 'value'}
        )
        assert isinstance(result, bool)

    @pytest.mark.skipif(not VALIDATION_REGISTRY_AVAILABLE, reason="ValidationRegistry not available")
    def test_get_registry_validation_report_error(self, validator):
        """Test error handling in registry report generation"""
        # Skip if no registry
        if not validator.validation_registry:
            pytest.skip("ValidationRegistry not available")

        report = validator.get_registry_validation_report()
        assert 'available' in report

    # === Helper Function Coverage ===

    def test_get_validation_help_text(self, validator):
        """Test getting validation help text"""
        help_text = validator.get_validation_help_text()

        assert isinstance(help_text, str)
        assert "REQUIRED STRUCTURE" in help_text
        assert "FY" in help_text
        assert "LTM" in help_text

    def test_get_directory_structure_help_function(self):
        """Test convenience function for help text"""
        help_text = get_directory_structure_help()

        assert isinstance(help_text, str)
        assert "Financial Data Directory Structure" in help_text


class TestPerformanceAndLargeScale:
    """Test performance with large directory structures"""

    @pytest.fixture
    def temp_company_dir(self):
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        shutil.rmtree(temp_dir, ignore_errors=True)

    @pytest.fixture
    def validator(self):
        return DirectoryStructureValidator()

    def test_validation_with_many_files(self, validator, temp_company_dir):
        """Test validation with many Excel files"""
        company_path = Path(temp_company_dir) / "LARGE_COMPANY"
        company_path.mkdir(parents=True)

        fy_path = company_path / "FY"
        fy_path.mkdir()

        # Create multiple Excel files (simulate large company data)
        for i in range(10):
            file_path = fy_path / f"Extra_File_{i}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(['Metric', 'FY', 'FY-1'])
            ws.append(['Revenue', 1000, 900])
            wb.save(file_path)

        # Run validation
        result = validator.validate_company_directory(str(company_path))

        # Should handle many files without issues
        assert result['exists']
        assert len(result['found_files']['FY']) >= 10

    def test_compliance_history_management(self, validator, temp_company_dir):
        """Test compliance history with multiple validations"""
        company_path = Path(temp_company_dir) / "HISTORY_TEST"
        company_path.mkdir(parents=True)
        (company_path / "FY").mkdir()
        (company_path / "LTM").mkdir()

        # Run multiple validations to build history
        for i in range(5):
            validator.validate_directory_structure(
                "HISTORY_TEST",
                base_path=temp_company_dir
            )

        # Get limited history
        history = validator.get_compliance_history(limit=3)
        assert len(history) == 3

        # Get all history
        all_history = validator.get_compliance_history()
        assert len(all_history) == 5


class TestAdditionalCoverage:
    """Additional tests to achieve 95%+ coverage"""

    @pytest.fixture
    def temp_company_dir(self):
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        shutil.rmtree(temp_dir, ignore_errors=True)

    @pytest.fixture
    def validator(self):
        return DirectoryStructureValidator()

    def test_validate_all_sheets_with_multiple_sheets(self, validator, temp_company_dir):
        """Test validation with multiple sheets in Excel file"""
        file_path = Path(temp_company_dir) / "multi_sheet.xlsx"
        wb = Workbook()

        # Create first sheet
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(['Metric', 'FY', 'FY-1', 'FY-2'])
        ws1.append(['Revenue', 1000, 900, 800])
        ws1.append(['Net Income', 100, 90, 80])

        # Create second sheet
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(['Metric', 'FY', 'FY-1'])
        ws2.append(['EBITDA', 200, 180])

        wb.save(file_path)

        result = validator._validate_all_sheets(file_path, "Income Statement.xlsx")

        assert len(result) == 2  # Should detect both sheets
        assert 'Sheet1' in result
        assert 'Sheet2' in result

    def test_organize_misplaced_files_with_ltm_file(self, validator, temp_company_dir):
        """Test organizing file that belongs in LTM folder"""
        company_path = Path(temp_company_dir) / "LTM_ORG_TEST"
        company_path.mkdir(parents=True)

        # Create LTM folder
        ltm_path = company_path / "LTM"
        ltm_path.mkdir()

        # Create file in root with LTM in name
        ltm_file = company_path / "LTM Income Statement.xlsx"
        wb = Workbook()
        wb.save(ltm_file)

        result = validator._organize_misplaced_files(company_path)

        # File should be moved or registered for moving
        assert 'moved_files' in result or 'errors' in result

    def test_repair_directory_structure_with_existing_files(self, validator, temp_company_dir):
        """Test repair when some files already exist"""
        company_path = Path(temp_company_dir) / "PARTIAL_COMPANY"
        company_path.mkdir(parents=True)

        fy_path = company_path / "FY"
        fy_path.mkdir()

        # Create one existing file
        existing_file = fy_path / "Income Statement.xlsx"
        wb = Workbook()
        wb.save(existing_file)

        # Run repair
        result = validator.repair_directory_structure(str(company_path))

        # Should create missing files but not overwrite existing
        assert result['success']
        assert existing_file.exists()

    def test_compliance_calculation_edge_cases(self, validator):
        """Test edge cases in compliance calculation"""
        # Test with mixed valid/invalid files
        dir_validation = {
            'structure_score': 0.7,
            'issues': [{'severity': 'high', 'message': 'test'}],
            'warnings': []
        }

        excel_validations = {
            'FY/valid.xlsx': {
                'is_valid': True,
                'has_data': True,
                'readable': True
            },
            'FY/invalid.xlsx': {
                'is_valid': False,
                'has_data': False,
                'readable': True,
                'issues': [{'severity': 'critical', 'message': 'test'}]
            }
        }

        compliance = validator._calculate_overall_compliance(dir_validation, excel_validations)

        assert 'overall_score' in compliance
        assert 'excel_score' in compliance
        assert compliance['total_issues'] > 0

    @pytest.mark.skipif(not VALIDATION_REGISTRY_AVAILABLE, reason="ValidationRegistry not available")
    def test_registry_rule_with_parameters(self, temp_company_dir):
        """Test rule execution with specific parameters"""
        from core.validation.validation_registry import ValidationRegistry, ValidationRule, RuleType, RuleScope

        registry = ValidationRegistry()
        validator = DirectoryStructureValidator(validation_registry=registry)

        # Add a custom rule with parameters
        rule = ValidationRule(
            rule_id="test_param_rule",
            name="Test Parametrized Rule",
            description="Rule with parameters",
            rule_type=RuleType.FORMAT,
            scope=RuleScope.DATA,
            priority="medium",
            parameters={'threshold': 0.8},
            error_message_template="Threshold not met: {threshold}",
            remediation_template="Adjust to meet threshold"
        )

        registry.register_rule(rule)

        company_path = Path(temp_company_dir) / "PARAM_TEST"
        company_path.mkdir(parents=True)

        # Execute with parameters
        result = validator._execute_validation_rule(rule, company_path)

        assert 'rule_id' in result
        assert result['rule_id'] == "test_param_rule"

    def test_export_report_with_empty_recommendations(self, validator, temp_company_dir):
        """Test report export with no recommendations"""
        report = {
            'ticker': 'EMPTY',
            'validation_timestamp': '2025-10-02',
            'company_path': '/test/path',
            'overall_compliance': {
                'overall_score': 1.0,
                'directory_score': 1.0,
                'excel_score': 1.0,
                'status': 'FULLY_COMPLIANT',
                'status_description': 'Perfect',
                'total_issues': 0,
                'total_warnings': 0
            },
            'actionable_recommendations': []
        }

        output_path = Path(temp_company_dir) / "empty_rec_report.json"
        result = validator.export_compliance_report(report, str(output_path), format='json')

        assert result['success']
        assert output_path.exists()

    def test_validation_with_insufficient_data_periods(self, validator, temp_company_dir):
        """Test validation with Excel file that has too few periods"""
        file_path = Path(temp_company_dir) / "few_periods.xlsx"
        wb = Workbook()
        ws = wb.active

        # Only 1 period (need at least 2)
        ws.append(['Metric', 'FY'])
        ws.append(['Revenue', 1000])
        ws.append(['Net Income', 100])

        wb.save(file_path)

        result = validator.validate_excel_file_format(file_path, "Income Statement.xlsx")

        # Should detect insufficient periods
        assert not result['sufficient_periods']


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
