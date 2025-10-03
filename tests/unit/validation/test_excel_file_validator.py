"""
Unit Tests for Excel File Validator
====================================

Tests the ExcelFileValidator to ensure it correctly validates Excel financial
statement files and detects common issues.
"""

import os
import tempfile
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from core.validation.excel_file_validator import (
    ExcelFileValidator,
    FileValidationResult,
    BatchValidationResult,
    ValidationSeverity,
    FileType,
    validate_excel_file,
    validate_company
)


class TestExcelFileValidator:
    """Test suite for ExcelFileValidator"""

    @pytest.fixture
    def validator(self):
        """Create validator instance"""
        return ExcelFileValidator()

    @pytest.fixture
    def strict_validator(self):
        """Create strict validator instance"""
        return ExcelFileValidator(strict_mode=True)

    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory for test files"""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir

    @pytest.fixture
    def valid_income_statement(self, temp_dir):
        """Create a valid Income Statement Excel file"""
        file_path = os.path.join(temp_dir, "Income Statement.xlsx")
        wb = Workbook()
        ws = wb.active

        # Add headers
        ws.append(['Item', 'FY', 'FY-1', 'FY-2', 'FY-3', 'FY-4'])

        # Add data rows
        ws.append(['Revenue', 100000, 95000, 90000, 85000, 80000])
        ws.append(['Cost of Revenue', 60000, 57000, 54000, 51000, 48000])
        ws.append(['Gross Profit', 40000, 38000, 36000, 34000, 32000])
        ws.append(['Operating Expenses', 20000, 19000, 18000, 17000, 16000])
        ws.append(['Operating Income', 20000, 19000, 18000, 17000, 16000])
        ws.append(['Interest Expense', 1000, 950, 900, 850, 800])
        ws.append(['EBT', 19000, 18050, 17100, 16150, 15200])
        ws.append(['Income Tax', 5700, 5415, 5130, 4845, 4560])
        ws.append(['Net Income', 13300, 12635, 11970, 11305, 10640])
        ws.append(['Shares Outstanding', 1000, 1000, 1000, 1000, 1000])
        ws.append(['EPS', 13.30, 12.64, 11.97, 11.31, 10.64])

        wb.save(file_path)
        return file_path

    @pytest.fixture
    def invalid_no_fy_columns(self, temp_dir):
        """Create Excel file without FY columns"""
        file_path = os.path.join(temp_dir, "Invalid Income Statement.xlsx")
        wb = Workbook()
        ws = wb.active

        # Add headers without FY columns
        ws.append(['Item', '2023', '2022', '2021'])
        ws.append(['Revenue', 100000, 95000, 90000])
        ws.append(['Cost of Revenue', 60000, 57000, 54000])

        wb.save(file_path)
        return file_path

    @pytest.fixture
    def incomplete_data_file(self, temp_dir):
        """Create Excel file with incomplete data"""
        file_path = os.path.join(temp_dir, "Incomplete Balance Sheet.xlsx")
        wb = Workbook()
        ws = wb.active

        # Add headers
        ws.append(['Item', 'FY', 'FY-1', 'FY-2'])

        # Add rows with missing data
        ws.append(['Total Assets', 500000, None, 450000])
        ws.append(['Current Assets', None, 200000, None])
        ws.append(['Total Liabilities', 300000, None, None])

        wb.save(file_path)
        return file_path

    @pytest.fixture
    def company_directory(self, temp_dir, valid_income_statement):
        """Create company directory structure with Excel files"""
        company_dir = os.path.join(temp_dir, "TEST")
        fy_dir = os.path.join(company_dir, "FY")
        ltm_dir = os.path.join(company_dir, "LTM")

        os.makedirs(fy_dir, exist_ok=True)
        os.makedirs(ltm_dir, exist_ok=True)

        # Create valid files in FY
        for file_type in ["Income Statement", "Balance Sheet", "Cash Flow Statement"]:
            file_path = os.path.join(fy_dir, f"{file_type}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.append(['Item', 'FY', 'FY-1', 'FY-2', 'FY-3'])
            for i in range(15):
                ws.append([f'Item {i}', 100 * i, 95 * i, 90 * i, 85 * i])
            wb.save(file_path)

        # Create valid files in LTM
        for file_type in ["Income Statement", "Balance Sheet", "Cash Flow Statement"]:
            file_path = os.path.join(ltm_dir, f"{file_type}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.append(['Item', 'FY', 'FY-1', 'FY-2'])
            for i in range(15):
                ws.append([f'Item {i}', 100 * i, 95 * i, 90 * i])
            wb.save(file_path)

        return company_dir

    # Test basic validation

    def test_validate_file_not_found(self, validator):
        """Test validation of non-existent file"""
        result = validator.validate_file("nonexistent_file.xlsx")
        assert not result.is_valid
        assert len(result.errors) > 0
        assert any("not found" in e.message.lower() for e in result.errors)

    def test_validate_valid_income_statement(self, validator, valid_income_statement):
        """Test validation of valid income statement"""
        result = validator.validate_file(valid_income_statement)
        assert result.is_valid
        assert result.file_type == FileType.INCOME_STATEMENT
        assert result.has_fy_columns
        assert result.fy_column_count >= 3
        assert len(result.errors) == 0

    def test_validate_file_type_detection(self, validator, valid_income_statement):
        """Test file type detection"""
        result = validator.validate_file(valid_income_statement)
        assert result.file_type == FileType.INCOME_STATEMENT

    def test_validate_no_fy_columns(self, validator, invalid_no_fy_columns):
        """Test validation of file without FY columns"""
        result = validator.validate_file(invalid_no_fy_columns)
        assert not result.is_valid
        assert not result.has_fy_columns
        assert any("header row" in e.message.lower() for e in result.errors)

    def test_validate_incomplete_data(self, validator, incomplete_data_file):
        """Test validation of file with incomplete data"""
        result = validator.validate_file(incomplete_data_file)
        # Should have warnings about data completeness
        assert len(result.warnings) > 0 or len(result.errors) > 0

    # Test strict mode

    def test_strict_mode_higher_requirements(self, strict_validator, valid_income_statement):
        """Test that strict mode enforces higher requirements"""
        # Create file with only 3 FY columns (below strict minimum of 5)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            ws.append(['Item', 'FY', 'FY-1', 'FY-2'])
            for i in range(10):
                ws.append([f'Item {i}', 100, 95, 90])
            wb.save(tmp.name)
            tmp_path = tmp.name

        try:
            result = strict_validator.validate_file(tmp_path)
            # Should have error in strict mode about insufficient FY columns
            assert len(result.errors) > 0 or len(result.warnings) > 0
        finally:
            os.unlink(tmp_path)

    # Test FY column validation

    def test_fy_column_detection(self, validator, valid_income_statement):
        """Test FY column detection and counting"""
        result = validator.validate_file(valid_income_statement)
        assert result.fy_column_count >= 3
        assert 'FY' in result.periods_found
        assert any('FY-' in p for p in result.periods_found)

    def test_fy_column_sequence(self, validator, temp_dir):
        """Test FY column sequence validation"""
        file_path = os.path.join(temp_dir, "Test.xlsx")
        wb = Workbook()
        ws = wb.active

        # Non-sequential FY columns
        ws.append(['Item', 'FY', 'FY-2', 'FY-4'])  # Missing FY-1, FY-3
        for i in range(10):
            ws.append([f'Item {i}', 100, 95, 90])
        wb.save(file_path)

        result = validator.validate_file(file_path)
        # Should have warning about non-sequential columns
        assert any("sequential" in w.message.lower() for w in result.warnings)

    # Test data quality validation

    def test_data_completeness_calculation(self, validator, valid_income_statement):
        """Test data completeness score calculation"""
        result = validator.validate_file(valid_income_statement)
        assert result.data_completeness_score > 0.9  # Should be high for valid file

    def test_low_completeness_warning(self, validator, incomplete_data_file):
        """Test warning for low data completeness"""
        result = validator.validate_file(incomplete_data_file)
        # Should flag low completeness
        assert result.data_completeness_score < 1.0
        assert len(result.warnings) > 0 or len(result.errors) > 0

    def test_numeric_data_validation(self, validator, temp_dir):
        """Test numeric data validation"""
        file_path = os.path.join(temp_dir, "Text_Data.xlsx")
        wb = Workbook()
        ws = wb.active

        ws.append(['Item', 'FY', 'FY-1', 'FY-2'])
        # Mostly text data in numeric columns (should trigger warning)
        for i in range(15):
            ws.append([f'Item {i}', 'high', 'medium', 'low'])
        wb.save(file_path)

        result = validator.validate_file(file_path)
        # Should warn about non-numeric data (less than 80% numeric)
        assert any("numeric" in w.message.lower() for w in result.warnings)

    # Test batch validation

    def test_validate_company_directory(self, validator, company_directory):
        """Test validation of entire company directory"""
        result = validator.validate_company_directory(company_directory)
        assert result.files_validated == 6  # 3 files in FY + 3 in LTM
        assert result.is_valid
        assert result.files_passed == 6

    def test_validate_missing_folder(self, validator, temp_dir):
        """Test validation with missing FY or LTM folder"""
        company_dir = os.path.join(temp_dir, "INCOMPLETE")
        os.makedirs(company_dir, exist_ok=True)

        result = validator.validate_company_directory(company_dir)
        assert "FY folder not found" in result.missing_files
        assert "LTM folder not found" in result.missing_files
        assert not result.is_valid

    def test_validate_missing_files(self, validator, temp_dir):
        """Test validation with missing required files"""
        company_dir = os.path.join(temp_dir, "PARTIAL")
        fy_dir = os.path.join(company_dir, "FY")
        os.makedirs(fy_dir, exist_ok=True)

        # Create only Income Statement, missing Balance Sheet and Cash Flow
        file_path = os.path.join(fy_dir, "Income Statement.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(['Item', 'FY', 'FY-1', 'FY-2'])
        for i in range(10):
            ws.append([f'Item {i}', 100, 95, 90])
        wb.save(file_path)

        result = validator.validate_company_directory(company_dir, validate_ltm=False)
        assert any("Balance Sheet" in mf for mf in result.missing_files)
        assert any("Cash Flow" in mf for mf in result.missing_files)

    # Test report generation

    def test_generate_validation_report(self, validator, valid_income_statement):
        """Test validation report generation"""
        result = validator.validate_file(valid_income_statement)
        report = validator.generate_validation_report(result)

        assert "VALIDATION REPORT" in report
        assert result.file_path in report
        assert "PASSED" in report or "FAILED" in report

    def test_save_validation_report(self, validator, valid_income_statement, temp_dir):
        """Test saving validation report to file"""
        result = validator.validate_file(valid_income_statement)
        output_path = os.path.join(temp_dir, "validation_report.txt")

        report = validator.generate_validation_report(result, output_path)
        assert os.path.exists(output_path)

        with open(output_path, 'r', encoding='utf-8') as f:
            content = f.read()
            assert content == report

    # Test convenience functions

    def test_validate_excel_file_convenience(self, valid_income_statement):
        """Test convenience function for file validation"""
        result = validate_excel_file(valid_income_statement)
        assert isinstance(result, FileValidationResult)
        assert result.is_valid

    def test_validate_company_convenience(self, company_directory):
        """Test convenience function for company validation"""
        result = validate_company(company_directory)
        assert isinstance(result, BatchValidationResult)
        assert result.is_valid

    # Test edge cases

    def test_empty_excel_file(self, validator, temp_dir):
        """Test validation of empty Excel file"""
        file_path = os.path.join(temp_dir, "Empty.xlsx")
        wb = Workbook()
        wb.save(file_path)

        result = validator.validate_file(file_path)
        assert not result.is_valid
        assert len(result.errors) > 0

    def test_excel_file_with_multiple_sheets(self, validator, temp_dir):
        """Test validation of file with multiple sheets"""
        file_path = os.path.join(temp_dir, "Multi_Sheet.xlsx")
        wb = Workbook()

        # Add first sheet (active)
        ws1 = wb.active
        ws1.title = "Main Data"
        ws1.append(['Item', 'FY', 'FY-1', 'FY-2'])
        for i in range(10):
            ws1.append([f'Item {i}', 100, 95, 90])

        # Add second sheet
        ws2 = wb.create_sheet("Supplemental")
        ws2.append(['Extra Data'])

        wb.save(file_path)

        result = validator.validate_file(file_path)
        assert result.is_valid
        assert len(result.sheets_found) == 2
        # Should have info message about multiple sheets
        assert any("sheet" in i.message.lower() for i in result.info)

    def test_malformed_excel_file(self, validator, temp_dir):
        """Test validation of malformed Excel file"""
        file_path = os.path.join(temp_dir, "Malformed.xlsx")

        # Create a text file with .xlsx extension
        with open(file_path, 'w') as f:
            f.write("This is not a valid Excel file")

        result = validator.validate_file(file_path)
        assert not result.is_valid
        # Should have error about invalid file or failed to open
        assert any("invalid" in e.message.lower() or "failed" in e.message.lower() for e in result.errors)

    def test_very_sparse_data(self, validator, temp_dir):
        """Test validation of file with very sparse data"""
        file_path = os.path.join(temp_dir, "Sparse.xlsx")
        wb = Workbook()
        ws = wb.active

        ws.append(['Item', 'FY', 'FY-1', 'FY-2', 'FY-3'])
        # Only 10% of cells have data
        for i in range(20):
            if i % 10 == 0:
                ws.append([f'Item {i}', 100, None, None, None])
            else:
                ws.append([f'Item {i}', None, None, None, None])

        wb.save(file_path)

        result = validator.validate_file(file_path)
        # Should flag low completeness
        assert result.data_completeness_score < 0.3
        assert len(result.warnings) > 0 or len(result.errors) > 0


class TestBatchValidationResult:
    """Test BatchValidationResult functionality"""

    def test_batch_result_summary(self):
        """Test batch result summary generation"""
        batch_result = BatchValidationResult(directory_path="/test/path")
        batch_result.files_validated = 6
        batch_result.files_passed = 5
        batch_result.files_failed = 1
        batch_result.total_errors = 3
        batch_result.total_warnings = 7

        summary = batch_result.get_summary()
        assert summary["files_validated"] == 6
        assert summary["files_passed"] == 5
        assert summary["files_failed"] == 1
        assert summary["total_errors"] == 3
        assert summary["total_warnings"] == 7
        assert "83.3%" in summary["success_rate"]

    def test_batch_result_is_valid(self):
        """Test batch validation status"""
        batch_result = BatchValidationResult(directory_path="/test/path")

        # All passed, no missing files
        batch_result.files_validated = 6
        batch_result.files_passed = 6
        batch_result.files_failed = 0
        assert batch_result.is_valid

        # Has failures
        batch_result.files_failed = 1
        assert not batch_result.is_valid

        # Has missing files
        batch_result.files_failed = 0
        batch_result.missing_files.append("Missing file")
        assert not batch_result.is_valid


@pytest.mark.integration
class TestRealFileValidation:
    """Integration tests with real company data files"""

    def test_validate_real_nvda_files(self):
        """Test validation of real NVDA company files"""
        nvda_path = "data/companies/NVDA"

        if not os.path.exists(nvda_path):
            pytest.skip("NVDA data not available")

        validator = ExcelFileValidator()
        result = validator.validate_company_directory(nvda_path)

        # Should have validated files
        assert result.files_validated > 0

        # Print summary for debugging
        print("\nNVDA Validation Summary:")
        for key, value in result.get_summary().items():
            print(f"  {key}: {value}")

    def test_validate_real_income_statement(self):
        """Test validation of real Income Statement file"""
        file_path = "data/companies/NVDA/FY/NVIDIA Corporation - Income Statement.xlsx"

        if not os.path.exists(file_path):
            pytest.skip("NVDA Income Statement not available")

        result = validate_excel_file(file_path)

        # Should detect as Income Statement
        assert result.file_type == FileType.INCOME_STATEMENT

        # Should have FY columns
        assert result.has_fy_columns
        assert result.fy_column_count > 0

        # Print details for debugging
        print(f"\nIncome Statement Validation:")
        print(f"  Valid: {result.is_valid}")
        print(f"  FY Columns: {result.fy_column_count}")
        print(f"  Periods: {result.periods_found}")
        print(f"  Completeness: {result.data_completeness_score:.1%}")
        print(f"  Errors: {len(result.errors)}")
        print(f"  Warnings: {len(result.warnings)}")

        if result.errors:
            print("\nErrors:")
            for error in result.errors:
                print(f"  - {error}")

        if result.warnings:
            print("\nWarnings:")
            for warning in result.warnings:
                print(f"  - {warning}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
