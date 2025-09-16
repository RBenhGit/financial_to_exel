"""
Test Suite for Excel Format Detection System
===========================================

Comprehensive tests for the ExcelFormatDetector and related functionality.
"""

import pytest
import logging
from unittest.mock import Mock, patch
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.excel_integration.format_detector import (
    ExcelFormatDetector,
    FormatDetectionResult,
    FormatType,
    StatementType,
    FormatSignature,
    ExcelStructureInfo,
    detect_excel_format,
    validate_excel_format
)

# Configure logging for tests
logging.basicConfig(level=logging.DEBUG)


class TestFormatDetector:
    """Test class for ExcelFormatDetector functionality"""

    @pytest.fixture
    def detector(self):
        """Create a format detector instance for testing"""
        return ExcelFormatDetector()

    @pytest.fixture
    def mock_worksheet(self):
        """Create a mock worksheet for testing"""
        ws = Mock(spec=Worksheet)
        ws.title = "test_sheet"
        ws.max_row = 48
        ws.max_column = 13
        return ws

    @pytest.fixture
    def premium_export_worksheet(self):
        """Create a worksheet that matches Premium Export format"""
        wb = Workbook()
        ws = wb.active
        ws.title = "premium_export_test"

        # Set up Premium Export format structure
        ws.cell(2, 3, "Test Company Inc")
        ws.cell(3, 3, "Premium Export")
        ws.cell(6, 3, "Income Statement")
        ws.cell(8, 4, "FY-9")
        ws.cell(8, 5, "FY-8")
        ws.cell(8, 6, "FY-7")
        ws.cell(10, 3, "Period End Date")
        ws.cell(10, 4, "2015-12-31")
        ws.cell(10, 5, "2016-12-31")
        ws.cell(10, 6, "2017-12-31")
        ws.cell(12, 3, "Revenue")
        ws.cell(12, 4, "1000.0")
        ws.cell(12, 5, "1100.0")

        return ws

    def test_detector_initialization(self, detector):
        """Test that detector initializes correctly"""
        assert isinstance(detector, ExcelFormatDetector)
        assert detector.confidence_threshold == 0.7
        assert FormatType.PREMIUM_EXPORT in detector.format_signatures
        assert FormatType.STANDARD_TEMPLATE in detector.format_signatures

    def test_format_signatures_structure(self, detector):
        """Test that format signatures are properly structured"""
        premium_sig = detector.format_signatures[FormatType.PREMIUM_EXPORT]

        assert isinstance(premium_sig, FormatSignature)
        assert premium_sig.name == "Premium Export Format"
        assert premium_sig.format_type == FormatType.PREMIUM_EXPORT
        assert len(premium_sig.identifier_patterns) >= 3
        assert len(premium_sig.company_name_positions) >= 1
        assert premium_sig.expected_max_rows > 0
        assert premium_sig.expected_max_columns > 0

    def test_is_likely_company_name(self, detector):
        """Test company name detection logic"""
        # Valid company names
        assert detector._is_likely_company_name("Apple Inc")
        assert detector._is_likely_company_name("Microsoft Corporation")
        assert detector._is_likely_company_name("Alphabet Inc Class C")
        assert detector._is_likely_company_name("Johnson & Johnson")

        # Invalid company names
        assert not detector._is_likely_company_name("Premium Export")
        assert not detector._is_likely_company_name("Period End Date")
        assert not detector._is_likely_company_name("2023")
        assert not detector._is_likely_company_name("Q1")
        assert not detector._is_likely_company_name("Revenue")
        assert not detector._is_likely_company_name("")
        assert not detector._is_likely_company_name("A")

    def test_detect_statement_type(self, detector):
        """Test financial statement type detection"""
        # Create mock worksheet with different statement types
        ws_income = Mock(spec=Worksheet)
        ws_income.max_row = 20
        ws_income.max_column = 10
        ws_income.cell.return_value.value = "Income Statement"

        # Mock cell access for income statement
        def mock_cell_income(row, col):
            mock_cell = Mock()
            if row == 6 and col == 3:
                mock_cell.value = "Income Statement"
            else:
                mock_cell.value = None
            return mock_cell

        ws_income.cell = mock_cell_income
        assert detector._detect_statement_type(ws_income) == StatementType.INCOME

        # Test balance sheet
        ws_balance = Mock(spec=Worksheet)
        ws_balance.max_row = 20
        ws_balance.max_column = 10

        def mock_cell_balance(row, col):
            mock_cell = Mock()
            if row == 6 and col == 3:
                mock_cell.value = "Balance Sheet"
            else:
                mock_cell.value = None
            return mock_cell

        ws_balance.cell = mock_cell_balance
        assert detector._detect_statement_type(ws_balance) == StatementType.BALANCE

    def test_extract_period_dates(self, detector):
        """Test period date extraction"""
        ws = Mock(spec=Worksheet)
        ws.max_row = 20
        ws.max_column = 15

        def mock_cell_dates(row, col):
            mock_cell = Mock()
            if row == 10 and col == 3:
                mock_cell.value = "Period End Date"
            elif row == 10 and col == 4:
                mock_cell.value = "2020-12-31"
            elif row == 10 and col == 5:
                mock_cell.value = "2021-12-31"
            elif row == 10 and col == 6:
                mock_cell.value = "2022-12-31"
            elif row == 10 and col == 7:
                mock_cell.value = None  # End of dates
            else:
                mock_cell.value = None
            return mock_cell

        ws.cell = mock_cell_dates
        dates = detector._extract_period_dates(ws)

        assert len(dates) == 3
        assert "2020-12-31" in dates
        assert "2021-12-31" in dates
        assert "2022-12-31" in dates

    def test_premium_export_detection(self, detector, premium_export_worksheet):
        """Test detection of Premium Export format"""
        result = detector.detect_format(premium_export_worksheet)

        assert isinstance(result, FormatDetectionResult)
        assert result.format_type == FormatType.PREMIUM_EXPORT
        assert result.confidence_score >= 0.7
        assert "identifier_3_3" in result.detected_patterns
        assert result.detected_patterns["identifier_3_3"] == "Premium Export"

    def test_extract_structure_info(self, detector, premium_export_worksheet):
        """Test extraction of structure information"""
        structure_info = detector._extract_structure_info(premium_export_worksheet)

        assert isinstance(structure_info, ExcelStructureInfo)
        assert structure_info.sheet_name == "premium_export_test"
        assert structure_info.company_name == "Test Company Inc"
        assert structure_info.statement_type == StatementType.INCOME
        assert len(structure_info.period_dates) >= 3
        assert structure_info.data_start_row > 0
        assert structure_info.data_start_column > 0

    def test_test_format_signature(self, detector, premium_export_worksheet):
        """Test format signature testing logic"""
        premium_sig = detector.format_signatures[FormatType.PREMIUM_EXPORT]
        structure_info = detector._extract_structure_info(premium_export_worksheet)

        confidence, patterns, errors = detector._test_format_signature(
            premium_export_worksheet, premium_sig, structure_info
        )

        assert confidence >= 0.5  # Should have reasonable confidence
        assert isinstance(patterns, dict)
        assert isinstance(errors, list)
        assert "identifier_3_3" in patterns

    def test_validate_format_with_expected_format(self, detector, premium_export_worksheet):
        """Test format validation with expected format specified"""
        result = detector.validate_format(
            premium_export_worksheet,
            expected_format=FormatType.PREMIUM_EXPORT
        )

        assert result.format_type == FormatType.PREMIUM_EXPORT
        assert result.confidence_score >= 0.6

    def test_validate_format_unknown_format(self, detector, premium_export_worksheet):
        """Test validation with unknown format type"""
        # Use an invalid format type
        from enum import Enum, auto

        class InvalidFormat(Enum):
            INVALID = auto()

        result = detector.validate_format(
            premium_export_worksheet,
            expected_format=InvalidFormat.INVALID
        )

        assert result.format_type == FormatType.UNKNOWN
        assert result.confidence_score == 0.0
        assert len(result.validation_errors) > 0

    def test_get_format_info(self, detector):
        """Test getting format information"""
        format_info = detector.get_format_info(FormatType.PREMIUM_EXPORT)

        assert isinstance(format_info, FormatSignature)
        assert format_info.name == "Premium Export Format"
        assert format_info.format_type == FormatType.PREMIUM_EXPORT

        # Test unknown format
        unknown_info = detector.get_format_info(FormatType.UNKNOWN)
        assert unknown_info is None

    def test_list_supported_formats(self, detector):
        """Test listing supported formats"""
        formats = detector.list_supported_formats()

        assert isinstance(formats, list)
        assert len(formats) >= 2
        assert "Premium Export Format" in formats
        assert "Standard Template Format" in formats

    def test_generate_suggestions(self, detector):
        """Test suggestion generation"""
        # Create structure info with missing elements
        structure_info = ExcelStructureInfo(
            sheet_name="test",
            max_row=10,
            max_column=5,
            company_name=None,  # Missing
            statement_type=StatementType.UNKNOWN,  # Unknown
            period_dates=[],  # Missing
            data_start_row=1,
            data_start_column=1,
            format_indicators={}
        )

        suggestions = detector._generate_suggestions(
            structure_info=structure_info,
            detected_format=FormatType.UNKNOWN,
            confidence=0.3,  # Low confidence
            errors=["Error 1", "Error 2"]
        )

        assert isinstance(suggestions, dict)
        assert "low_confidence_warning" in suggestions
        assert "missing_company_name" in suggestions
        assert "missing_period_dates" in suggestions
        assert "unknown_statement_type" in suggestions
        assert "unknown_format" in suggestions

    def test_find_data_start_position(self, detector):
        """Test finding data start position"""
        ws = Mock(spec=Worksheet)
        ws.max_row = 30
        ws.max_column = 10

        def mock_cell_data_start(row, col):
            mock_cell = Mock()
            if row == 15 and col == 3:
                mock_cell.value = "Revenue"
            else:
                mock_cell.value = None
            return mock_cell

        ws.cell = mock_cell_data_start

        start_row, start_col = detector._find_data_start_position(ws)
        assert start_row == 15
        assert start_col == 3

        # Test fallback when no financial keywords found
        def mock_cell_no_keywords(row, col):
            mock_cell = Mock()
            mock_cell.value = "Random Text"
            return mock_cell

        ws.cell = mock_cell_no_keywords
        start_row, start_col = detector._find_data_start_position(ws)
        assert start_row == 12  # Default fallback
        assert start_col == 3


class TestConvenienceFunctions:
    """Test convenience functions"""

    def test_detect_excel_format_function(self, premium_export_worksheet):
        """Test the detect_excel_format convenience function"""
        result = detect_excel_format(premium_export_worksheet)

        assert isinstance(result, FormatDetectionResult)
        assert result.format_type == FormatType.PREMIUM_EXPORT

    def test_validate_excel_format_function(self, premium_export_worksheet):
        """Test the validate_excel_format convenience function"""
        result = validate_excel_format(premium_export_worksheet)

        assert isinstance(result, FormatDetectionResult)
        assert result.format_type == FormatType.PREMIUM_EXPORT

        # Test with expected format
        result_expected = validate_excel_format(
            premium_export_worksheet,
            FormatType.PREMIUM_EXPORT
        )

        assert result_expected.format_type == FormatType.PREMIUM_EXPORT


class TestEdgeCases:
    """Test edge cases and error conditions"""

    @pytest.fixture
    def empty_worksheet(self):
        """Create an empty worksheet for testing"""
        wb = Workbook()
        ws = wb.active
        ws.title = "empty_test"
        return ws

    @pytest.fixture
    def malformed_worksheet(self):
        """Create a malformed worksheet for testing"""
        wb = Workbook()
        ws = wb.active
        ws.title = "malformed_test"

        # Add some random data that doesn't match any format
        ws.cell(1, 1, "Random")
        ws.cell(2, 2, "Data")
        ws.cell(3, 3, 12345)

        return ws

    def test_empty_worksheet_detection(self, empty_worksheet):
        """Test detection on empty worksheet"""
        result = detect_excel_format(empty_worksheet)

        assert result.format_type == FormatType.UNKNOWN
        assert result.confidence_score < 0.5
        assert len(result.validation_errors) > 0

    def test_malformed_worksheet_detection(self, malformed_worksheet):
        """Test detection on malformed worksheet"""
        result = detect_excel_format(malformed_worksheet)

        assert result.format_type == FormatType.UNKNOWN
        assert result.confidence_score < 0.3

    def test_worksheet_with_errors(self):
        """Test handling of worksheet access errors"""
        detector = ExcelFormatDetector()

        # Mock worksheet that raises exceptions
        ws = Mock(spec=Worksheet)
        ws.title = "error_test"
        ws.max_row = 10
        ws.max_column = 10
        ws.cell.side_effect = Exception("Mock error")

        result = detector.detect_format(ws)

        assert result.format_type == FormatType.UNKNOWN
        assert result.confidence_score == 0.0

    def test_large_worksheet(self):
        """Test handling of very large worksheets"""
        detector = ExcelFormatDetector()

        ws = Mock(spec=Worksheet)
        ws.title = "large_test"
        ws.max_row = 10000  # Very large
        ws.max_column = 1000  # Very large

        def mock_cell_large(row, col):
            mock_cell = Mock()
            mock_cell.value = None
            return mock_cell

        ws.cell = mock_cell_large

        result = detector.detect_format(ws)

        # Should still work but detect unknown format
        assert isinstance(result, FormatDetectionResult)
        assert result.format_type == FormatType.UNKNOWN


if __name__ == "__main__":
    # Run tests
    pytest.main([__file__, "-v"])