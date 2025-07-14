import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

def analyze_visa_fcf_file():
    """
    Analyze the Visa FCF Analysis Excel file to identify why FCFF and FCFE show N/A in year 9
    """
    file_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/V/FCF_Analysis_Visa_Inc_Class_A.xlsx"
    
    try:
        # Load the workbook to examine all sheets
        wb = load_workbook(file_path, data_only=True)
        print(f"Available sheets: {wb.sheetnames}")
        
        # Examine each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n=== Analyzing Sheet: {sheet_name} ===")
            ws = wb[sheet_name]
            
            # Skip chart sheets
            if hasattr(ws, 'iter_rows'):
                # Convert to pandas for easier analysis
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
            else:
                print(f"Skipping chart sheet: {sheet_name}")
                continue
            
            # Find non-empty rows
            non_empty_rows = []
            for i, row in enumerate(data):
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    non_empty_rows.append((i+1, row))
            
            print(f"Sheet has {len(non_empty_rows)} non-empty rows")
            
            # Look for FCF-related sections
            fcf_sections = []
            for row_num, row in non_empty_rows:
                for j, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        if any(keyword in cell.upper() for keyword in ['FCFF', 'FCFE', 'FREE CASH FLOW', 'FCF']):
                            fcf_sections.append((row_num, j, cell))
            
            if fcf_sections:
                print(f"Found FCF-related sections:")
                for row_num, col_num, content in fcf_sections:
                    print(f"  Row {row_num}, Col {col_num}: {content}")
                
                # Look around FCF sections for year 9 data
                for row_num, col_num, content in fcf_sections:
                    print(f"\n--- Examining area around '{content}' at Row {row_num} ---")
                    
                    # Check 10 rows around this section
                    start_row = max(0, row_num - 5)
                    end_row = min(len(data), row_num + 10)
                    
                    for i in range(start_row, end_row):
                        if i < len(data):
                            row_data = data[i]
                            # Look for year headers or data that might represent year 9
                            relevant_cells = []
                            for j, cell in enumerate(row_data[:15]):  # Check first 15 columns
                                if cell is not None:
                                    relevant_cells.append(f"Col{j}: {cell}")
                            
                            if relevant_cells:
                                print(f"  Row {i+1}: {' | '.join(relevant_cells)}")
        
        # Now let's specifically look for the data structure
        print("\n=== Detailed Analysis of Main Sheet ===")
        main_sheet = wb.active  # Usually the first sheet
        
        # Look for year headers (typically in row 1 or 2)
        print("\nSearching for year headers...")
        for row_num in range(1, 6):  # Check first 5 rows
            row_data = []
            for col_num in range(1, 16):  # Check first 15 columns
                cell_value = main_sheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    row_data.append(f"Col{col_num}: {cell_value}")
            
            if row_data:
                print(f"Row {row_num}: {' | '.join(row_data)}")
        
        # Look for FCF calculation rows
        print("\nSearching for FCF calculation rows...")
        for row_num in range(1, 100):  # Search first 100 rows
            cell_a = main_sheet.cell(row=row_num, column=1).value
            if cell_a and isinstance(cell_a, str) and any(keyword in cell_a.upper() for keyword in ['FCFF', 'FCFE', 'FREE CASH FLOW']):
                print(f"\nFound FCF row {row_num}: {cell_a}")
                
                # Get the entire row data
                row_values = []
                for col_num in range(1, 16):
                    cell_value = main_sheet.cell(row=row_num, column=col_num).value
                    row_values.append(cell_value)
                
                print(f"Row data: {row_values}")
                
                # Check if there are N/A values
                na_positions = []
                for i, val in enumerate(row_values):
                    if val and str(val).upper() in ['N/A', '#N/A', 'NA']:
                        na_positions.append(i+1)
                
                if na_positions:
                    print(f"N/A values found in columns: {na_positions}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error analyzing file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_visa_fcf_file()