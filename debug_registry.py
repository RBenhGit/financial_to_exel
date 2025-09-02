#!/usr/bin/env python3
"""
Debug the Financial Variable Registry and Excel structure
"""

import sys
from pathlib import Path

# Add the project root to Python path
project_root = Path(__file__).parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from core.data_processing.financial_variable_registry import get_registry
from openpyxl import load_workbook

def debug_registry():
    """Debug the financial variable registry"""
    print("=== Financial Variable Registry Debug ===")
    
    registry = get_registry()
    all_vars = registry.list_all_variables()
    
    print(f"Total variables in registry: {len(all_vars)}")
    print(f"First 10 variables: {all_vars[:10]}")
    
    # Check categories
    from core.data_processing.financial_variable_registry import VariableCategory
    for category in VariableCategory:
        vars_in_cat = [var for var in all_vars 
                       if registry.get_variable_definition(var) and 
                       registry.get_variable_definition(var).category == category]
        print(f"{category.value}: {len(vars_in_cat)} variables")
        if vars_in_cat:
            print(f"  Examples: {vars_in_cat[:3]}")
    
    # Check some specific variable definitions
    test_vars = ['revenue', 'net_income', 'total_assets', 'operating_cash_flow']
    print("\nVariable definitions:")
    for var in test_vars:
        var_def = registry.get_variable_definition(var)
        if var_def:
            print(f"  {var}: {var_def.description}")
            if hasattr(var_def, 'aliases') and var_def.aliases:
                print(f"    Aliases: {var_def.aliases}")
        else:
            print(f"  {var}: NOT FOUND")

def debug_excel_structure():
    """Debug the Excel file structure"""
    print("\n=== Excel File Structure Debug ===")
    
    test_file = "data/companies/GOOG/FY/Alphabet Inc Class C - Income Statement.xlsx"
    
    try:
        wb = load_workbook(filename=test_file)
        sheet = wb.active
        
        # Convert to list of lists
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        print(f"Total rows in Excel: {len(data)}")
        
        # Find header row
        header_row_idx = None
        for i, row in enumerate(data):
            if row and any('FY-' in str(cell) or 'FY' == str(cell) for cell in row if cell):
                header_row_idx = i
                break
        
        if header_row_idx is not None:
            print(f"Header row found at index: {header_row_idx}")
            headers = data[header_row_idx]
            print(f"Headers: {headers}")
            
            # Show first 20 data rows with their first column (variable names)
            print("\nFirst 20 variable names from Excel:")
            for i, row in enumerate(data[header_row_idx + 1:header_row_idx + 21]):
                if row and row[0]:
                    print(f"  Row {i+1}: '{row[0]}'")
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")

if __name__ == "__main__":
    debug_registry()
    debug_excel_structure()