import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

def check_visa_source_files():
    """
    Check the original Visa financial source files to understand what data is available
    """
    base_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/V"
    
    # Check FY folder files
    fy_path = os.path.join(base_path, "FY")
    print("=== Checking FY (Fiscal Year) Source Files ===")
    
    for filename in os.listdir(fy_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(fy_path, filename)
            print(f"\nAnalyzing: {filename}")
            
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                
                # Look for year headers and data range
                print("Year data found:")
                for row_num in range(1, 10):
                    row_values = []
                    for col_num in range(1, 15):
                        cell_value = ws.cell(row=row_num, column=col_num).value
                        if cell_value and isinstance(cell_value, (int, float)) and 2010 <= cell_value <= 2030:
                            row_values.append(f"Col{col_num}: {cell_value}")
                        elif cell_value and isinstance(cell_value, str) and any(year in cell_value for year in ['2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', '2023', '2024', '2025']):
                            row_values.append(f"Col{col_num}: {cell_value}")
                    
                    if row_values:
                        print(f"  Row {row_num}: {' | '.join(row_values)}")
                
                # Look for specific financial metrics
                print("Key financial metrics:")
                metrics_to_find = ['EBIT', 'NET INCOME', 'CAPEX', 'CASH FROM OPERATIONS', 'DEPRECIATION', 'TAX']
                
                for row_num in range(1, min(100, ws.max_row + 1)):
                    cell_a = ws.cell(row=row_num, column=1).value
                    if cell_a and isinstance(cell_a, str):
                        for metric in metrics_to_find:
                            if metric in cell_a.upper():
                                # Get the data for this metric
                                metric_data = []
                                for col_num in range(1, 15):
                                    cell_value = ws.cell(row=row_num, column=col_num).value
                                    metric_data.append(cell_value)
                                print(f"  {metric} (Row {row_num}): {metric_data[:10]}")  # Show first 10 values
                                break
                
                wb.close()
                
            except Exception as e:
                print(f"Error reading {filename}: {e}")

def analyze_fcf_formulas():
    """
    Analyze the FCF formulas to understand what might be causing N/A values
    """
    file_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/V/FCF_Analysis_Visa_Inc_Class_A.xlsx"
    
    try:
        # Load with formulas (not just values)
        wb = load_workbook(file_path, data_only=False)
        
        if 'FCF DATA' in wb.sheetnames:
            ws = wb['FCF DATA']
            print(f"\n=== FCF Formula Analysis ===")
            
            # Look for FCFF and FCFE calculations with formulas
            for row_num in range(1, ws.max_row + 1):
                cell_a = ws.cell(row=row_num, column=1).value
                if cell_a and isinstance(cell_a, str):
                    if any(fcf_type in cell_a.upper() for fcf_type in ['FCFF', 'FCFE', 'FREE CASH FLOW TO FIRM', 'FREE CASH FLOW TO EQUITY']):
                        print(f"\nFound FCF calculation at row {row_num}: {cell_a}")
                        
                        # Check the formulas in this row
                        for col_num in range(2, 15):  # Skip first column (label)
                            cell = ws.cell(row=row_num, column=col_num)
                            if cell.value is not None:
                                print(f"  Col {col_num}: Value={cell.value}, Formula={cell.coordinate}={getattr(cell, 'value', 'N/A')}")
                                
                                # If it's a formula, show it
                                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                                    print(f"    Formula: {cell.value}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error analyzing formulas: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_visa_source_files()
    analyze_fcf_formulas()