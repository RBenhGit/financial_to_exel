import os
from openpyxl import Workbook as workbook
from openpyxl import load_workbook
import openpyxl
import json
import datetime
from dataclasses import dataclass
from datetime import date
from tkinter import filedialog, Tk

def select_files(title=""):
    root = Tk()
    root.withdraw()  # Hide the main window
    files = filedialog.askopenfilenames(title=title)
    return files

DCF_file = select_files("Pleas Select DCF template file")[0]
Selected_Files_FY = select_files("Pleas Select Fiscal Year financial files")

# Find the correct document for the analysis
for i in Selected_Files_FY:
    if i.find("Balance")>0:
        Balance_Sheet = i
    elif i.find("Cash")>0:
        Cash_Flow_Statement = i
    elif i.find("Income")>0:
        Income_Statement = i

Selected_Files_LTM = select_files("Pleas Select Latest Twelve Months financial files")

# Find the correct document for the analysis
for i in Selected_Files_LTM:
    if i.find("Balance")>0:
        Balance_Sheet_Q = i
    elif i.find("Cash")>0:
        Cash_Flow_Statement_LTM = i
    elif i.find("Income")>0:
        Income_Statement_LTM = i

#CompanyName = 'Visa'



TargetFile = load_workbook(filename = DCF_file)
wb1 = TargetFile.worksheets[0]  # בחירה מפורשת של הגיליון הראשון

wb1['c1']=date.today()

Income_File = load_workbook(filename = Income_Statement)
Income_wb = Income_File.worksheets[0]  # בחירה מפורשת של הגיליון הראשון בקובץ השני
Income_File_LTM = load_workbook(filename = Income_Statement_LTM)
Income_wb_LTM = Income_File_LTM.worksheets[0]  # בחירה מפורשת של הגיליון הראשון בקובץ השני

Balance_File = load_workbook(filename = Balance_Sheet)
Balance_wb = Balance_File.worksheets[0]  # בחירה מפורשת של הגיליון הראשון בקובץ השני
Balance_File_Q = load_workbook(filename = Balance_Sheet_Q)
Balance_wb_Q = Balance_File_Q.worksheets[0]  # בחירה מפורשת של הגיליון הראשון בקובץ השני

Cash_Flow_File = load_workbook(filename = Cash_Flow_Statement)
Cash_Flow_wb = Cash_Flow_File.worksheets[0]  # בחירה מפורשת של הגיליון הראשון בקובץ השני
Cash_Flow_File_LTM = load_workbook(filename = Cash_Flow_Statement_LTM)
Cash_Flow_wb_LTM = Cash_Flow_File_LTM.worksheets[0]  # בחירה מפורשת של הגיליון הראשון בקובץ השני

Company_Name = Income_wb.cell(2,3).value #Assign CompanyName to target file
wb1['c2'] = Company_Name

row_val = []
for row in range(0,59):
    row_val.append(
        {
            'income_index':  row+1,
            'income_value':  Income_wb.cell(row+1,3).value,
            'balance_index':  row+1,
            'balance_value':  Balance_wb.cell(row+1,3).value,
            'cashflow_index':  row+1,
            'cashflow_value':  Cash_Flow_wb.cell(row+1,3).value
        }
    )
        
for i in row_val:
    if i['income_value']=='Period End Date':
        row_index = i['income_index']
        for j in range(9):
            wb1.cell(row=15-j, column=1).value = Income_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=1).value = Income_wb_LTM.cell(row_index, column=15).value
        wb1['C3']= wb1.cell(row=16, column=1).value   
    elif  i['income_value']=='Net Interest Expenses':
        row_index = i['income_index']
        for j in range(9):
            wb1.cell(row=15-j, column=2).value = Income_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=2).value = Income_wb_LTM.cell(row_index, column=15).value
    elif  i['income_value']=='EBT, Incl. Unusual Items':
        row_index = i['income_index']
        for j in range(9):
            wb1.cell(row=15-j, column=3).value = Income_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=3).value = Income_wb_LTM.cell(row_index, column=15).value
    elif  i['income_value']=='Income Tax Expense':
        row_index = i['income_index']
        for j in range(9):
            wb1.cell(row=15-j, column=4).value = Income_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=4).value = Income_wb_LTM.cell(row_index, column=15).value    
    elif  i['income_value']=='Net Income to Company':
        row_index = i['income_index']
        for j in range(9):
            wb1.cell(row=15-j, column=5).value = Income_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=5).value = Income_wb_LTM.cell(row_index, column=15).value    
    elif  i['income_value']=='EBIT':
        row_index = i['income_index']
        for j in range(9):
            wb1.cell(row=15-j, column=6).value = Income_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=6).value = Income_wb_LTM.cell(row_index, column=15).value
for i in row_val:
    if  i['balance_value']=='Total Current Assets':
        row_index = i['balance_index']
        for j in range(9):
            wb1.cell(row=15-j, column=7).value = Balance_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=7).value = Balance_wb_Q.cell(row_index, column=15).value
    elif  i['balance_value']=='Total Current Liabilities':
        row_index = i['balance_index']
        for j in range(9):
            wb1.cell(row=15-j, column=8).value = Balance_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=8).value = Balance_wb_Q.cell(row_index, column=15).value
for i in row_val:
    if  i['cashflow_value']=='Depreciation & Amortization (CF)':
        row_index = i['cashflow_index']
        for j in range(9):
            wb1.cell(row=15-j, column=11).value = Cash_Flow_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=11).value = Cash_Flow_wb_LTM.cell(row_index, column=15).value
    elif  i['cashflow_value']=='Amortization of Deferred Charges (CF)':
        row_index = i['cashflow_index']
        for j in range(9):
            wb1.cell(row=15-j, column=12).value = Cash_Flow_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=12).value = Cash_Flow_wb_LTM.cell(row_index, column=15).value
    elif  i['cashflow_value']=='Cash from Operations':
        row_index = i['cashflow_index']
        for j in range(9):
            wb1.cell(row=15-j, column=13).value = Cash_Flow_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=13).value = Cash_Flow_wb_LTM.cell(row_index, column=15).value
    elif  i['cashflow_value']=='Capital Expenditures':
        row_index = i['cashflow_index']
        for j in range(9):
            wb1.cell(row=15-j, column=14).value = Cash_Flow_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=14).value = Cash_Flow_wb_LTM.cell(row_index, column=15).value
    elif  i['cashflow_value']=='Cash from Financing':
        row_index = i['cashflow_index']
        for j in range(9):
            wb1.cell(row=15-j, column=15).value = Cash_Flow_wb.cell(row_index, column=13-j).value
        wb1.cell(row=16, column=15).value = Cash_Flow_wb_LTM.cell(row_index, column=15).value
 
# לא לשכוח לשמור את השינויים
path = filedialog.askdirectory()
file_name = path + '/FCF_Analysis_' + Company_Name + '.xlsx'
try:
    TargetFile.save(filename = file_name) 
    print('File Saved')
except:
    file_name = path + '/FCF_Analysis' + '.xlsx'
    TargetFile.save(filename = file_name)
    print('Aletrnative File Saved')
