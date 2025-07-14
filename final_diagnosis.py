import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

def final_year9_diagnosis():
    """
    Final diagnosis to identify the exact issue with year 9 FCFF and FCFE
    """
    file_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/V/FCF_Analysis_Visa_Inc_Class_A.xlsx"
    
    try:
        # Read the FCF DATA sheet with pandas
        df = pd.read_excel(file_path, sheet_name='FCF DATA', header=None)
        
        print("=== FCFE Data Analysis ===")
        print("FCFE header is at row 34 (index 33)")
        print("FCFE data starts at row 35 (index 34)")
        
        # Extract FCFE data section (rows 34-44 based on previous analysis)
        fcfe_header = df.iloc[33]  # Row 34
        print(f"FCFE Header: {fcfe_header.tolist()}")
        
        print(f"\nFCFE Data (rows 35-44):")
        for i in range(34, 44):  # Rows 35-44
            if i < len(df):
                row_data = df.iloc[i]
                year = row_data[0] if pd.notna(row_data[0]) else "N/A"
                fcfe_value = row_data[7] if pd.notna(row_data[7]) else "N/A"  # FCFE is in column H (index 7)
                print(f"  Year {year}: FCFE = {fcfe_value}")
                
                # If this is 2024 (year 9), show all components
                if year == 2024:
                    print(f"    2024 (Year 9) Components:")
                    print(f"      Net Income: {row_data[2] if pd.notna(row_data[2]) else 'MISSING'}")
                    print(f"      D&A: {row_data[3] if pd.notna(row_data[3]) else 'MISSING'}")
                    print(f"      Change in Working Cap: {row_data[4] if pd.notna(row_data[4]) else 'MISSING'}")
                    print(f"      CAPEX: {row_data[5] if pd.notna(row_data[5]) else 'MISSING'}")
                    print(f"      Cash from Financing: {row_data[6] if pd.notna(row_data[6]) else 'MISSING'}")
                    print(f"      Full row: {row_data.tolist()}")
        
        # Now look for FCFF data
        print(f"\n=== FCFF Data Analysis ===")
        print("Looking for FCFF calculations...")
        
        # Search for FCFF mentions
        fcff_found = False
        for idx, row in df.iterrows():
            for col_idx, cell in enumerate(row):
                if cell and isinstance(cell, str) and 'FCFF' in cell.upper():
                    print(f"Found FCFF reference at row {idx+1}, col {col_idx+1}: {cell}")
                    fcff_found = True
        
        if not fcff_found:
            print("No explicit FCFF calculations found. FCFF might be calculated differently or missing.")
        
        # Check if there are any summary rows at the bottom
        print(f"\n=== Summary Section Analysis ===")
        print("Checking rows 60-72 for summary calculations...")
        
        for i in range(59, min(72, len(df))):  # Rows 60-72
            row_data = df.iloc[i]
            if any(pd.notna(cell) for cell in row_data):
                print(f"Row {i+1}: {row_data.tolist()}")
        
        # Now check other sheets for FCFF/FCFE calculations
        print(f"\n=== Other Sheets Analysis ===")
        
        # Check Data Entry sheet
        try:
            df_entry = pd.read_excel(file_path, sheet_name='Data Entry', header=None)
            print(f"Data Entry sheet shape: {df_entry.shape}")
            
            for idx, row in df_entry.iterrows():
                for col_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str) and ('FCFF' in cell.upper() or 'FCFE' in cell.upper()):
                        print(f"Data Entry - Found at row {idx+1}, col {col_idx+1}: {cell}")
                        print(f"  Row data: {row.tolist()}")
        except:
            print("Could not read Data Entry sheet")
        
        # Check if there's a calculation issue by examining the formula structure
        print(f"\n=== Formula Analysis ===")
        wb = load_workbook(file_path, data_only=False)  # Load with formulas
        
        if 'FCF DATA' in wb.sheetnames:
            ws = wb['FCF DATA']
            
            # Check row 43 (2024 FCFE row) for formulas
            print("Checking 2024 FCFE calculation (row 43):")
            for col_num in range(1, 16):
                cell = ws.cell(row=43, column=col_num)
                if cell.value is not None:
                    coord = cell.coordinate
                    value = cell.value
                    print(f"  {coord}: {value}")
                    
                    # If it's column H (FCFE calculation), check if it's a formula
                    if col_num == 8:  # Column H
                        if hasattr(cell, 'data_type') and cell.data_type == 'f':
                            print(f"    This is a formula: {cell.value}")
                        elif isinstance(value, str) and value.upper() in ['N/A', '#N/A']:
                            print(f"    This cell contains N/A!")
        
        wb.close()
        
    except Exception as e:
        print(f"Error in final diagnosis: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    final_year9_diagnosis()