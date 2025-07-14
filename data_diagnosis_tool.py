#!/usr/bin/env python3
"""
Data Acquisition Diagnostic Tool

Comprehensive tool to identify and fix data acquisition, cleaning, and indexing issues.
"""

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import logging
from typing import Dict, List, Tuple, Any
import re
from difflib import SequenceMatcher

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DataDiagnosticTool:
    """
    Comprehensive tool for diagnosing data acquisition and processing issues
    """
    
    def __init__(self):
        self.issues_found = []
        self.recommendations = []
        self.data_coverage = {}
        
    def diagnose_company_folder(self, company_folder: str) -> Dict:
        """
        Comprehensive diagnosis of company data folder
        
        Args:
            company_folder: Path to company folder
            
        Returns:
            Dictionary with diagnosis results
        """
        logger.info(f"Starting comprehensive diagnosis for: {company_folder}")
        
        diagnosis = {
            'company_folder': company_folder,
            'folder_structure': self._check_folder_structure(company_folder),
            'excel_files': self._analyze_excel_files(company_folder),
            'metric_coverage': self._analyze_metric_coverage(company_folder),
            'data_quality': self._assess_data_quality(company_folder),
            'missing_data': self._identify_missing_data(company_folder),
            'recommendations': self._generate_recommendations()
        }
        
        return diagnosis
    
    def _check_folder_structure(self, company_folder: str) -> Dict:
        """Check folder structure requirements"""
        logger.info("Checking folder structure...")
        
        structure = {
            'valid': True,
            'missing_folders': [],
            'missing_files': [],
            'extra_files': []
        }
        
        required_folders = ['FY', 'LTM']
        required_files = {
            'FY': ['Income Statement', 'Balance Sheet', 'Cash Flow Statement'],
            'LTM': ['Income Statement', 'Balance Sheet', 'Cash Flow Statement']
        }
        
        # Check folders
        for folder in required_folders:
            folder_path = os.path.join(company_folder, folder)
            if not os.path.exists(folder_path):
                structure['missing_folders'].append(folder)
                structure['valid'] = False
        
        # Check files in each folder
        for folder, file_patterns in required_files.items():
            folder_path = os.path.join(company_folder, folder)
            if os.path.exists(folder_path):
                files_in_folder = os.listdir(folder_path)
                for pattern in file_patterns:
                    found = any(pattern in file_name for file_name in files_in_folder)
                    if not found:
                        structure['missing_files'].append(f"{folder}/{pattern}")
                        structure['valid'] = False
        
        return structure
    
    def _analyze_excel_files(self, company_folder: str) -> Dict:
        """Analyze Excel files for data structure and content"""
        logger.info("Analyzing Excel file structures...")
        
        analysis = {
            'files_analyzed': 0,
            'files_readable': 0,
            'metric_counts': {},
            'sheet_structures': {},
            'data_ranges': {}
        }
        
        for subfolder in ['FY', 'LTM']:
            subfolder_path = os.path.join(company_folder, subfolder)
            if not os.path.exists(subfolder_path):
                continue
                
            for file_name in os.listdir(subfolder_path):
                if file_name.endswith(('.xlsx', '.xls')):
                    file_path = os.path.join(subfolder_path, file_name)
                    analysis['files_analyzed'] += 1
                    
                    try:
                        file_analysis = self._analyze_single_excel_file(file_path)
                        analysis['sheet_structures'][f"{subfolder}/{file_name}"] = file_analysis
                        analysis['files_readable'] += 1
                        
                        # Count metrics found
                        metric_count = len(file_analysis.get('potential_metrics', []))
                        analysis['metric_counts'][f"{subfolder}/{file_name}"] = metric_count
                        
                    except Exception as e:
                        logger.error(f"Error analyzing {file_path}: {e}")
                        self.issues_found.append(f"Cannot read Excel file: {file_path} - {e}")
        
        return analysis
    
    def _analyze_single_excel_file(self, file_path: str) -> Dict:
        """Analyze a single Excel file for structure and metrics"""
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        
        analysis = {
            'file_path': file_path,
            'sheet_name': sheet.title,
            'dimensions': f"{sheet.max_row} rows x {sheet.max_column} columns",
            'potential_metrics': [],
            'data_structure': {},
            'header_locations': []
        }
        
        # Look for header patterns and metrics
        financial_keywords = [
            'revenue', 'income', 'ebit', 'cash', 'assets', 'liabilities',
            'depreciation', 'amortization', 'capex', 'expenditure', 'tax'
        ]
        
        # Scan first 100 rows for metrics
        for row_idx in range(1, min(101, sheet.max_row + 1)):
            for col_idx in range(1, min(6, sheet.max_column + 1)):  # First 5 columns
                cell_value = sheet.cell(row_idx, col_idx).value
                if cell_value and isinstance(cell_value, str):
                    cell_text = cell_value.lower().strip()
                    
                    # Check if it looks like a financial metric
                    if any(keyword in cell_text for keyword in financial_keywords):
                        analysis['potential_metrics'].append({
                            'text': cell_value,
                            'location': f"Row {row_idx}, Col {col_idx}",
                            'keywords_matched': [kw for kw in financial_keywords if kw in cell_text]
                        })
                    
                    # Check for year headers (e.g., 2024, FY-1, etc.)
                    if re.match(r'(20\d{2}|FY)', str(cell_value)):
                        analysis['header_locations'].append(f"Row {row_idx}, Col {col_idx}: {cell_value}")
        
        return analysis
    
    def _analyze_metric_coverage(self, company_folder: str) -> Dict:
        """Analyze coverage of required financial metrics"""
        logger.info("Analyzing metric coverage...")
        
        required_metrics = {
            'income': [
                'Net Interest Expenses', 'EBT, Incl. Unusual Items', 
                'Income Tax Expense', 'Net Income to Company', 'EBIT'
            ],
            'balance': [
                'Total Current Assets', 'Total Current Liabilities'
            ],
            'cashflow': [
                'Depreciation & Amortization (CF)', 'Amortization of Deferred Charges (CF)',
                'Cash from Operations', 'Capital Expenditures', 'Cash from Financing'
            ]
        }
        
        coverage = {
            'required_metrics': required_metrics,
            'found_metrics': {},
            'missing_metrics': {},
            'similar_metrics': {},
            'coverage_percentage': {}
        }
        
        # Use enhanced search to find metrics
        for subfolder in ['FY', 'LTM']:
            for statement_type in ['income', 'balance', 'cashflow']:
                file_pattern_map = {
                    'income': 'Income Statement',
                    'balance': 'Balance Sheet', 
                    'cashflow': 'Cash Flow Statement'
                }
                
                pattern = file_pattern_map[statement_type]
                file_path = self._find_file_by_pattern(company_folder, subfolder, pattern)
                
                if file_path:
                    found, missing, similar = self._search_metrics_in_file(
                        file_path, required_metrics[statement_type]
                    )
                    
                    key = f"{subfolder}_{statement_type}"
                    coverage['found_metrics'][key] = found
                    coverage['missing_metrics'][key] = missing
                    coverage['similar_metrics'][key] = similar
                    
                    # Calculate coverage percentage
                    total_required = len(required_metrics[statement_type])
                    found_count = len(found)
                    coverage['coverage_percentage'][key] = (found_count / total_required) * 100
        
        return coverage
    
    def _find_file_by_pattern(self, company_folder: str, subfolder: str, pattern: str) -> str:
        """Find file by pattern in subfolder"""
        subfolder_path = os.path.join(company_folder, subfolder)
        if not os.path.exists(subfolder_path):
            return None
            
        for file_name in os.listdir(subfolder_path):
            if pattern in file_name and file_name.endswith(('.xlsx', '.xls')):
                return os.path.join(subfolder_path, file_name)
        
        return None
    
    def _search_metrics_in_file(self, file_path: str, required_metrics: List[str]) -> Tuple[List, List, List]:
        """Search for metrics in Excel file with fuzzy matching"""
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        
        found_metrics = []
        missing_metrics = []
        similar_metrics = []
        
        # Extract all text from first 3 columns, first 100 rows
        all_texts = []
        for row_idx in range(1, min(101, sheet.max_row + 1)):
            for col_idx in range(1, 4):  # First 3 columns
                cell_value = sheet.cell(row_idx, col_idx).value
                if cell_value and isinstance(cell_value, str):
                    all_texts.append({
                        'text': cell_value.strip(),
                        'location': f"R{row_idx}C{col_idx}"
                    })
        
        for required_metric in required_metrics:
            best_match = None
            best_similarity = 0
            
            # Try exact match first
            for text_info in all_texts:
                if required_metric.lower() == text_info['text'].lower():
                    found_metrics.append({
                        'required': required_metric,
                        'found': text_info['text'],
                        'location': text_info['location'],
                        'match_type': 'exact'
                    })
                    best_match = text_info
                    break
            
            if not best_match:
                # Try fuzzy matching
                for text_info in all_texts:
                    similarity = SequenceMatcher(None, required_metric.lower(), text_info['text'].lower()).ratio()
                    if similarity > best_similarity:
                        best_similarity = similarity
                        best_match = text_info
                
                if best_similarity > 0.6:  # 60% similarity threshold
                    similar_metrics.append({
                        'required': required_metric,
                        'found': best_match['text'],
                        'location': best_match['location'],
                        'similarity': best_similarity,
                        'match_type': 'fuzzy'
                    })
                else:
                    missing_metrics.append(required_metric)
        
        return found_metrics, missing_metrics, similar_metrics
    
    def _assess_data_quality(self, company_folder: str) -> Dict:
        """Assess overall data quality"""
        logger.info("Assessing data quality...")
        
        quality = {
            'completeness_score': 0,
            'consistency_score': 0,
            'accessibility_score': 0,
            'overall_score': 0,
            'quality_issues': [],
            'data_gaps': []
        }
        
        try:
            # Test actual data extraction
            from financial_calculations import FinancialCalculator
            
            calc = FinancialCalculator(company_folder)
            calc.load_financial_statements()
            metrics = calc._calculate_all_metrics()
            
            # Assess completeness
            total_expected_metrics = 12  # Approximate number of key metrics
            available_metrics = sum(1 for values in metrics.values() if values and len(values) > 0)
            quality['completeness_score'] = (available_metrics / total_expected_metrics) * 100
            
            # Assess data consistency (years of data)
            year_counts = [len(values) for values in metrics.values() if values]
            if year_counts:
                min_years = min(year_counts)
                max_years = max(year_counts)
                consistency = (min_years / max_years) * 100 if max_years > 0 else 0
                quality['consistency_score'] = consistency
            
            # Check for data gaps
            for metric_name, values in metrics.items():
                if not values:
                    quality['data_gaps'].append(f"No data for {metric_name}")
                elif len(values) < 3:
                    quality['data_gaps'].append(f"Insufficient data for {metric_name} ({len(values)} years)")
                elif values.count(0) > len(values) * 0.5:
                    quality['data_gaps'].append(f"Too many zero values in {metric_name}")
            
            # Calculate accessibility (can we load the data?)
            quality['accessibility_score'] = 100 if metrics else 0
            
        except Exception as e:
            logger.error(f"Error assessing data quality: {e}")
            quality['quality_issues'].append(f"Data extraction failed: {e}")
            quality['accessibility_score'] = 0
        
        # Calculate overall score
        quality['overall_score'] = (
            quality['completeness_score'] * 0.4 +
            quality['consistency_score'] * 0.3 +
            quality['accessibility_score'] * 0.3
        )
        
        return quality
    
    def _identify_missing_data(self, company_folder: str) -> Dict:
        """Identify specific missing data points"""
        logger.info("Identifying missing data...")
        
        missing_data = {
            'critical_missing': [],
            'optional_missing': [],
            'data_source_issues': [],
            'suggested_fixes': []
        }
        
        # Critical metrics that must be present
        critical_metrics = [
            'Net Income', 'EBIT', 'Cash from Operations', 
            'Capital Expenditures', 'Total Current Assets', 'Total Current Liabilities'
        ]
        
        try:
            from financial_calculations import FinancialCalculator
            calc = FinancialCalculator(company_folder)
            calc.load_financial_statements()
            metrics = calc._calculate_all_metrics()
            
            # Check critical metrics
            for metric in critical_metrics:
                metric_key = self._map_metric_to_key(metric)
                if metric_key not in metrics or not metrics[metric_key]:
                    missing_data['critical_missing'].append(metric)
                    missing_data['suggested_fixes'].append(
                        f"Review Excel files for '{metric}' - may need fuzzy matching"
                    )
            
            # Check data source files
            for subfolder in ['FY', 'LTM']:
                for statement in ['Income Statement', 'Balance Sheet', 'Cash Flow Statement']:
                    file_path = self._find_file_by_pattern(company_folder, subfolder, statement)
                    if not file_path:
                        missing_data['data_source_issues'].append(f"Missing {subfolder}/{statement}")
            
        except Exception as e:
            missing_data['data_source_issues'].append(f"Cannot analyze data: {e}")
        
        return missing_data
    
    def _map_metric_to_key(self, metric_name: str) -> str:
        """Map display metric name to internal key"""
        mapping = {
            'Net Income': 'net_income',
            'EBIT': 'ebit',
            'Cash from Operations': 'operating_cash_flow',
            'Capital Expenditures': 'capex',
            'Total Current Assets': 'current_assets',
            'Total Current Liabilities': 'current_liabilities'
        }
        return mapping.get(metric_name, metric_name.lower().replace(' ', '_'))
    
    def _generate_recommendations(self) -> List[str]:
        """Generate actionable recommendations"""
        recommendations = []
        
        if len(self.issues_found) > 0:
            recommendations.append("Critical issues found - review data sources")
        
        recommendations.extend([
            "Implement fuzzy string matching for metric names",
            "Expand row scanning beyond 59 rows",
            "Add validation for LTM vs FY data integration", 
            "Implement cross-validation between FCF calculation methods",
            "Add automated data completeness reporting"
        ])
        
        return recommendations
    
    def print_diagnosis_report(self, diagnosis: Dict):
        """Print comprehensive diagnosis report"""
        print("\n" + "="*80)
        print("📊 COMPREHENSIVE DATA ACQUISITION DIAGNOSIS REPORT")
        print("="*80)
        
        company_name = os.path.basename(diagnosis['company_folder'])
        print(f"Company: {company_name}")
        print(f"Folder: {diagnosis['company_folder']}")
        
        # Folder Structure
        print(f"\n📁 FOLDER STRUCTURE:")
        structure = diagnosis['folder_structure']
        if structure['valid']:
            print("   ✅ Valid folder structure")
        else:
            print("   ❌ Invalid folder structure")
            if structure['missing_folders']:
                print(f"   Missing folders: {structure['missing_folders']}")
            if structure['missing_files']:
                print(f"   Missing files: {structure['missing_files']}")
        
        # Excel Files Analysis
        print(f"\n📄 EXCEL FILES ANALYSIS:")
        excel_analysis = diagnosis['excel_files']
        print(f"   Files analyzed: {excel_analysis['files_analyzed']}")
        print(f"   Files readable: {excel_analysis['files_readable']}")
        
        for file_key, metric_count in excel_analysis['metric_counts'].items():
            print(f"   {file_key}: {metric_count} potential metrics found")
        
        # Metric Coverage
        print(f"\n📈 METRIC COVERAGE:")
        coverage = diagnosis['metric_coverage']
        for key, percentage in coverage['coverage_percentage'].items():
            status = "✅" if percentage >= 90 else "⚠️" if percentage >= 70 else "❌"
            print(f"   {key}: {percentage:.1f}% {status}")
        
        # Data Quality
        print(f"\n🎯 DATA QUALITY ASSESSMENT:")
        quality = diagnosis['data_quality']
        print(f"   Overall Score: {quality['overall_score']:.1f}/100")
        print(f"   Completeness: {quality['completeness_score']:.1f}/100")
        print(f"   Consistency: {quality['consistency_score']:.1f}/100")
        print(f"   Accessibility: {quality['accessibility_score']:.1f}/100")
        
        if quality['data_gaps']:
            print(f"   Data Gaps Found: {len(quality['data_gaps'])}")
            for gap in quality['data_gaps'][:5]:  # Show first 5
                print(f"      • {gap}")
        
        # Missing Data
        print(f"\n🔍 MISSING DATA ANALYSIS:")
        missing = diagnosis['missing_data']
        if missing['critical_missing']:
            print(f"   Critical Missing Metrics: {missing['critical_missing']}")
        if missing['data_source_issues']:
            print(f"   Data Source Issues: {missing['data_source_issues']}")
        
        # Recommendations
        print(f"\n💡 RECOMMENDATIONS:")
        for i, rec in enumerate(diagnosis['recommendations'], 1):
            print(f"   {i}. {rec}")
        
        print("\n" + "="*80)

def main():
    """Main function to run diagnosis on all companies"""
    print("🚀 DATA ACQUISITION DIAGNOSIS TOOL")
    print("="*50)
    
    # Define company folders
    base_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel"
    companies = ['GOOG', 'MSFT', 'NVDA', 'TSLA', 'V']
    
    diagnostic_tool = DataDiagnosticTool()
    
    for company in companies:
        company_folder = os.path.join(base_path, company)
        if os.path.exists(company_folder):
            print(f"\n🔍 Diagnosing {company}...")
            diagnosis = diagnostic_tool.diagnose_company_folder(company_folder)
            diagnostic_tool.print_diagnosis_report(diagnosis)
        else:
            print(f"❌ Company folder not found: {company_folder}")

if __name__ == "__main__":
    main()