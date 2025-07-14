import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

def analyze_fcf_data_sheet():
    """
    Focus on the FCF DATA sheet to find the issue with year 9 N/A values
    """
    file_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/V/FCF_Analysis_Visa_Inc_Class_A.xlsx"
    
    try:
        # Load the workbook
        wb = load_workbook(file_path, data_only=True)
        print(f"Available sheets: {wb.sheetnames}")
        
        # Focus on FCF DATA sheet
        if 'FCF DATA' in wb.sheetnames:
            print(f"\n=== Analyzing FCF DATA Sheet ===")
            ws = wb['FCF DATA']
            
            # Read all data into a structured format
            print("Reading all data from FCF DATA sheet...")
            
            # Get the maximum dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"Sheet dimensions: {max_row} rows x {max_col} columns")
            
            # Read all data
            all_data = []
            for row_num in range(1, min(max_row + 1, 50)):  # Limit to first 50 rows for analysis
                row_data = []
                for col_num in range(1, min(max_col + 1, 20)):  # Limit to first 20 columns
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data.append(cell_value)
                all_data.append(row_data)
            
            # Display the data structure
            print("\nFCF DATA Sheet Structure:")
            for i, row_data in enumerate(all_data):
                # Only show rows with significant data
                non_empty_cells = [cell for cell in row_data if cell is not None and str(cell).strip() != ""]
                if non_empty_cells:
                    formatted_row = []
                    for j, cell in enumerate(row_data[:15]):  # Show first 15 columns
                        if cell is not None:
                            formatted_row.append(f"Col{j+1}: {cell}")
                    if formatted_row:
                        print(f"Row {i+1}: {' | '.join(formatted_row)}")
            
            # Look for FCFF and FCFE specifically
            print(f"\n=== Searching for FCFF and FCFE calculations ===")
            fcff_row = None
            fcfe_row = None
            
            for row_num in range(1, max_row + 1):
                cell_a = ws.cell(row=row_num, column=1).value
                if cell_a and isinstance(cell_a, str):
                    if 'FCFF' in cell_a.upper():
                        fcff_row = row_num
                        print(f"Found FCFF at row {row_num}: {cell_a}")
                    elif 'FCFE' in cell_a.upper():
                        fcfe_row = row_num
                        print(f"Found FCFE at row {row_num}: {cell_a}")
            
            # Analyze FCFF and FCFE rows in detail
            if fcff_row:
                print(f"\n=== FCFF Row {fcff_row} Analysis ===")
                fcff_data = []
                for col_num in range(1, 15):
                    cell_value = ws.cell(row=fcff_row, column=col_num).value
                    fcff_data.append(cell_value)
                print(f"FCFF data: {fcff_data}")
                
                # Check for N/A values and their positions
                na_positions = []
                for i, val in enumerate(fcff_data):
                    if val and (str(val).upper() in ['N/A', '#N/A', 'NA'] or pd.isna(val)):
                        na_positions.append(i+1)
                if na_positions:
                    print(f"N/A values in FCFF at columns: {na_positions}")
            
            if fcfe_row:
                print(f"\n=== FCFE Row {fcfe_row} Analysis ===")
                fcfe_data = []
                for col_num in range(1, 15):
                    cell_value = ws.cell(row=fcfe_row, column=col_num).value
                    fcfe_data.append(cell_value)
                print(f"FCFE data: {fcfe_data}")
                
                # Check for N/A values and their positions
                na_positions = []
                for i, val in enumerate(fcfe_data):
                    if val and (str(val).upper() in ['N/A', '#N/A', 'NA'] or pd.isna(val)):
                        na_positions.append(i+1)
                if na_positions:
                    print(f"N/A values in FCFE at columns: {na_positions}")
            
            # Look for year headers to understand which column is year 9
            print(f"\n=== Year Header Analysis ===")
            for row_num in range(1, 10):
                row_data = []
                for col_num in range(1, 15):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value and (isinstance(cell_value, (int, float)) and 2010 <= cell_value <= 2030):
                        row_data.append(f"Col{col_num}: {cell_value}")
                    elif cell_value and isinstance(cell_value, str) and any(year_str in cell_value for year_str in ['2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', '2023', '2024']):
                        row_data.append(f"Col{col_num}: {cell_value}")
                
                if row_data:
                    print(f"Year indicators in Row {row_num}: {' | '.join(row_data)}")
            
            # Look for the component data that feeds into FCF calculations
            print(f"\n=== FCF Component Analysis ===")
            components = ['EBIT', 'NET INCOME', 'CAPEX', 'DEPRECIATION', 'AMORTIZATION', 'WORKING CAPITAL', 'TAX']
            
            for component in components:
                for row_num in range(1, max_row + 1):
                    cell_a = ws.cell(row=row_num, column=1).value
                    if cell_a and isinstance(cell_a, str) and component in cell_a.upper():
                        print(f"\nFound {component} at row {row_num}: {cell_a}")
                        component_data = []
                        for col_num in range(1, 15):
                            cell_value = ws.cell(row=row_num, column=col_num).value
                            component_data.append(cell_value)
                        print(f"{component} data: {component_data}")
                        
                        # Check for missing values
                        missing_positions = []
                        for i, val in enumerate(component_data[1:], 1):  # Skip first column (label)
                            if val is None or (isinstance(val, str) and val.strip() == ""):
                                missing_positions.append(i+1)
                        if missing_positions:
                            print(f"Missing values in {component} at columns: {missing_positions}")
                        break
        
        wb.close()
        
    except Exception as e:
        print(f"Error analyzing file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_fcf_data_sheet()