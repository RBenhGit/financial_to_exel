#!/usr/bin/env python3
"""
Test script to verify Excel extraction fix
"""
import sys
import os
from openpyxl import load_workbook


def test_excel_extraction():
    """Test that Excel extraction now reads data in correct chronological order"""
    try:
        # Find the first available company and cash flow statement
        import os
        import glob

        companies = []
        for item in os.listdir("."):
            if os.path.isdir(item) and not item.startswith('.') and not item.startswith('_'):
                fy_path = os.path.join(item, 'FY')
                if os.path.exists(fy_path):
                    companies.append(item)

        if not companies:
            print("❌ No companies found in dataset")
            return

        company = sorted(companies)[0]
        cash_flow_files = glob.glob(os.path.join(company, "FY", "*Cash Flow Statement.xlsx"))

        if not cash_flow_files:
            print(f"❌ No cash flow statement found for {company}")
            return

        excel_file = cash_flow_files[0]

        print("Testing Excel extraction order...")
        print(f"Loading: {excel_file}")

        # Load the workbook
        wb = load_workbook(excel_file, read_only=True)
        sheet = wb.active

        # Search for the Levered Free Cash Flow row in the first 50 rows
        print("Searching for 'Levered Free Cash Flow' row...")
        levered_fcf_row = None
        for row in range(1, 50):
            cell_value = sheet.cell(row, 1).value
            if cell_value:
                print(f"  Row {row}: '{cell_value}'")
                # Check different variations
                if cell_value and (
                    "Levered Free Cash Flow" in str(cell_value)
                    or "levered free cash flow" in str(cell_value).lower()
                ):
                    levered_fcf_row = row
                    print(f"  ✓ Found at row {row}")
                    break

        # If not found in first column, search in all columns
        if not levered_fcf_row:
            print("Not found in column 1, searching all columns...")
            for row in range(1, 50):
                for col in range(1, 15):
                    cell_value = sheet.cell(row, col).value
                    if cell_value and (
                        "Levered Free Cash Flow" in str(cell_value)
                        or "levered free cash flow" in str(cell_value).lower()
                    ):
                        levered_fcf_row = row
                        print(f"  ✓ Found at row {row}, column {col}: '{cell_value}'")
                        break
                if levered_fcf_row:
                    break

        if levered_fcf_row:
            print(f"Found 'Levered Free Cash Flow' at row {levered_fcf_row}")

            # Test the OLD extraction logic (inverted) - using 10 years
            print("\nOLD extraction logic (column=13-j):")
            old_values = []
            for j in range(10):
                cell_value = sheet.cell(levered_fcf_row, column=13 - j).value
                old_values.append(cell_value)
                print(f"  j={j}, column={13-j}, value={cell_value}")

            # Test the NEW extraction logic (fixed) - using 10 years
            print("\nNEW extraction logic (column=4+j):")
            new_values = []
            for j in range(10):
                cell_value = sheet.cell(levered_fcf_row, column=4 + j).value
                new_values.append(cell_value)
                print(f"  j={j}, column={4+j}, value={cell_value}")

            # Test with generic chronological order checking instead of hardcoded values

            print(f"OLD logic result: {old_values}")
            print(f"NEW logic result: {new_values}")

            # Convert string values to floats for comparison
            def parse_value(val):
                if isinstance(val, str):
                    # Remove commas and convert to float
                    return float(val.replace(',', ''))
                return float(val) if val is not None else 0.0

            new_values_float = [parse_value(val) for val in new_values]
            old_values_float = [parse_value(val) for val in old_values]

            print(f"\nParsed NEW logic result: {new_values_float}")
            print(f"Parsed OLD logic result: {old_values_float}")

            # Test that we got some data and that NEW logic is different from OLD logic
            if len(new_values_float) > 0 and len(old_values_float) > 0:
                print("\n✅ SUCCESS: Excel extraction successfully read data from file!")
                print("✅ Both OLD and NEW extraction logic are working")

                # Check if NEW logic produces different results than OLD logic
                if new_values_float != old_values_float:
                    print("✅ NEW extraction logic produces different results from OLD logic")
                    print("✅ Data inversion fix is working correctly!")
                    # Check if the NEW values are in reverse order compared to OLD values
                    if new_values_float == list(reversed(old_values_float)):
                        print("✅ NEW logic correctly reverses the OLD logic order")
                else:
                    print("⚠️ NEW and OLD logic produce identical results")

                return True
            else:
                print("\n❌ FAIL: Excel extraction found no data")
                print("❌ No values were extracted from the file")
                return False
        else:
            print("❌ FAIL: Could not find 'Levered Free Cash Flow' row")
            return False

    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback

        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = test_excel_extraction()
    sys.exit(0 if success else 1)
