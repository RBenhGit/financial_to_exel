#!/usr/bin/env python3

"""
Test script for date extraction functionality
"""

import os
from openpyxl import load_workbook
from datetime import datetime
import logging

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_period_end_dates(workbook):
    """
    Extract Period End Date values from financial statement
    
    Args:
        workbook: Excel workbook containing financial data
        
    Returns:
        list: List of date strings extracted from Period End Date row
    """
    try:
        # Get the active sheet from the workbook
        sheet = workbook.active
        
        # Period End Date is always at row 10, column 3 according to analysis
        period_end_row = 10
        label_column = 3
        
        # Check if Period End Date exists at expected location
        if sheet.cell(period_end_row, label_column).value != "Period End Date":
            logger.warning("Period End Date not found at expected location (row 10, column 3)")
            return []
        
        # Extract dates starting from column 4
        dates = []
        for col in range(4, 16):  # Columns 4-15 (up to 12 periods)
            cell_value = sheet.cell(period_end_row, col).value
            if cell_value is not None:
                dates.append(str(cell_value))
            else:
                break  # Stop when we hit empty cells
        
        logger.info(f"Extracted {len(dates)} period end dates: {dates}")
        return dates
        
    except Exception as e:
        logger.error(f"Error extracting period end dates: {str(e)}")
        return []

def parse_date_year(date_string):
    """
    Parse year from date string in YYYY-MM-DD format
    
    Args:
        date_string: Date string in format "YYYY-MM-DD"
        
    Returns:
        int: Year as integer, or None if parsing fails
    """
    try:
        if isinstance(date_string, str) and len(date_string) >= 4:
            # Extract year from YYYY-MM-DD format
            year_str = date_string[:4]
            return int(year_str)
    except (ValueError, TypeError):
        logger.warning(f"Could not parse year from date: {date_string}")
    return None

def test_date_extraction():
    """Test the date extraction with sample data"""
    
    # Dynamically discover companies in the dataset
    companies = []
    for item in os.listdir("."):
        if os.path.isdir(item) and not item.startswith('.') and not item.startswith('_'):
            fy_path = os.path.join(item, 'FY')
            if os.path.exists(fy_path):
                companies.append(item)
    
    if not companies:
        print("❌ No companies found in dataset")
        return
    
    print(f"📊 Found {len(companies)} companies: {', '.join(companies)}")
    companies = sorted(companies)
    
    for company in companies:
        print(f"\n=== Testing {company} ===")
        
        try:
            # Test FY Income Statement
            fy_path = f"{company}/FY/{company}_income_statement.xlsx"
            
            # Find actual file (names might vary)
            for file in os.listdir(f"{company}/FY/"):
                if "Income" in file:
                    fy_path = f"{company}/FY/{file}"
                    break
            
            if os.path.exists(fy_path):
                wb = load_workbook(fy_path)
                dates = extract_period_end_dates(wb)
                print(f"FY dates: {dates}")
                
                years = [parse_date_year(date_str) for date_str in dates]
                years = [year for year in years if year is not None]
                print(f"FY years: {years}")
            else:
                print(f"File not found: {fy_path}")
                
        except Exception as e:
            print(f"Error testing {company}: {e}")

if __name__ == "__main__":
    test_date_extraction()