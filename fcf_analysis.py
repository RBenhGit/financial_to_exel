import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.widgets import Button, TextBox
from openpyxl import load_workbook
import logging
from datetime import datetime
from tkinter import filedialog, Tk
from scipy import stats
import yfinance as yf

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class FCFAnalyzer:
    """
    Analyzes Free Cash Flow using three different methods:
    1. Free Cash Flow to Firm (FCFF)
    2. Free Cash Flow to Equity (FCFE) 
    3. Levered Free Cash Flow (LFCF)
    """
    
    def __init__(self, company_folder):
        """
        Initialize FCF analyzer with company folder path
        
        Args:
            company_folder (str): Path to company folder containing FY and LTM subfolders
        """
        self.company_folder = company_folder
        self.company_name = os.path.basename(company_folder) if company_folder else "Unknown"
        self.financial_data = {}
        self.fcf_results = {}
        # Store current DCF assumptions for interactive updates
        self.current_dcf_assumptions = {}
        # Store ticker data
        self.ticker_symbol = None
        self.current_stock_price = None
        self.market_cap = None
        
    def load_financial_statements(self):
        """
        Load financial statements from Excel files
        """
        try:
            # Define file paths
            fy_folder = os.path.join(self.company_folder, 'FY')
            ltm_folder = os.path.join(self.company_folder, 'LTM')
            
            # Load FY statements
            for file_name in os.listdir(fy_folder):
                if 'Income Statement' in file_name:
                    self.financial_data['income_fy'] = self._load_excel_data(os.path.join(fy_folder, file_name))
                elif 'Balance Sheet' in file_name:
                    self.financial_data['balance_fy'] = self._load_excel_data(os.path.join(fy_folder, file_name))
                elif 'Cash Flow Statement' in file_name:
                    self.financial_data['cashflow_fy'] = self._load_excel_data(os.path.join(fy_folder, file_name))
            
            # Load LTM statements
            for file_name in os.listdir(ltm_folder):
                if 'Income Statement' in file_name:
                    self.financial_data['income_ltm'] = self._load_excel_data(os.path.join(ltm_folder, file_name))
                elif 'Balance Sheet' in file_name:
                    self.financial_data['balance_ltm'] = self._load_excel_data(os.path.join(ltm_folder, file_name))
                elif 'Cash Flow Statement' in file_name:
                    self.financial_data['cashflow_ltm'] = self._load_excel_data(os.path.join(ltm_folder, file_name))
                    
            logger.info("Financial statements loaded successfully")
            
        except Exception as e:
            logger.error(f"Error loading financial statements: {e}")
            raise
    
    def _load_excel_data(self, file_path):
        """
        Load Excel data and convert to DataFrame
        
        Args:
            file_path (str): Path to Excel file
            
        Returns:
            pd.DataFrame: Financial data
        """
        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active
            
            # Convert to list of lists, then to DataFrame
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            
            # Create DataFrame with first row as headers if possible
            if len(data) > 1:
                df = pd.DataFrame(data[1:], columns=data[0])
            else:
                df = pd.DataFrame(data)
                
            return df
            
        except Exception as e:
            logger.error(f"Error loading Excel file {file_path}: {e}")
            raise
    
    def fetch_ticker_data(self, ticker_symbol=None):
        """
        Fetch current stock price and market cap from yfinance
        
        Args:
            ticker_symbol (str): Stock ticker symbol (e.g., 'GOOGL', 'AAPL')
        """
        try:
            # If no ticker provided, try to infer from company name
            if ticker_symbol is None:
                ticker_symbol = self._infer_ticker_from_company_name()
            
            if ticker_symbol is None:
                logger.warning("No ticker symbol provided and couldn't infer from company name")
                return False
            
            self.ticker_symbol = ticker_symbol.upper()
            
            # Fetch ticker data
            ticker = yf.Ticker(self.ticker_symbol)
            info = ticker.info
            
            # Get current price
            self.current_stock_price = info.get('currentPrice') or info.get('regularMarketPrice')
            if self.current_stock_price is None:
                # Try to get from history
                hist = ticker.history(period='1d')
                if not hist.empty:
                    self.current_stock_price = hist['Close'].iloc[-1]
            
            # Get market cap
            self.market_cap = info.get('marketCap')
            
            # Calculate shares outstanding from market cap and price
            if self.current_stock_price and self.market_cap:
                self.shares_outstanding = self.market_cap / self.current_stock_price
                logger.info(f"Fetched data for {self.ticker_symbol}: Price=${self.current_stock_price:.2f}, Market Cap=${self.market_cap/1000000:.0f}M, Shares={self.shares_outstanding/1000000:.1f}M")
                return True
            else:
                logger.warning(f"Could not fetch complete data for {self.ticker_symbol}")
                return False
                
        except Exception as e:
            logger.error(f"Error fetching ticker data for {ticker_symbol}: {e}")
            return False
    
    def _infer_ticker_from_company_name(self):
        """
        Try to infer ticker symbol from company folder name
        """
        # Common mappings
        ticker_mappings = {
            'GOOGL': 'GOOGL',
            'GOOGLE': 'GOOGL',
            'ALPHABET': 'GOOGL',
            'AAPL': 'AAPL',
            'APPLE': 'AAPL',
            'MSFT': 'MSFT',
            'MICROSOFT': 'MSFT',
            'TSLA': 'TSLA',
            'TESLA': 'TSLA',
            'AMZN': 'AMZN',
            'AMAZON': 'AMZN',
            'META': 'META',
            'FACEBOOK': 'META',
            'NFLX': 'NFLX',
            'NETFLIX': 'NFLX'
        }
        
        company_name_upper = self.company_name.upper()
        
        # Check direct mappings
        for key, ticker in ticker_mappings.items():
            if key in company_name_upper:
                return ticker
        
        # If company name looks like a ticker (3-5 uppercase letters)
        if len(self.company_name) <= 5 and self.company_name.isalpha():
            return self.company_name.upper()
        
        return None
    
    def extract_financial_metrics(self):
        """
        Extract key financial metrics from statements
        """
        try:
            # Extract years from row 10 (Period End Date) columns 4-12
            years = []
            income_sheet = self.financial_data['income_fy']
            
            # Find the row with "Period End Date"
            period_row = None
            for row_idx in range(len(income_sheet)):
                if income_sheet.iloc[row_idx, 2] == 'Period End Date':
                    period_row = row_idx
                    break
            
            if period_row is None:
                raise ValueError("Could not find Period End Date row")
            
            # Extract years from columns 4-12 (corresponds to FY-8 to FY)
            for col_idx in range(4, 13):  # Columns 4-12
                if col_idx < len(income_sheet.columns):
                    year_val = income_sheet.iloc[period_row, col_idx]
                    if year_val and str(year_val) != 'nan':
                        # Extract year from date string like '2024-12-31'
                        year = str(year_val)[:4] if '-' in str(year_val) else str(year_val)
                        years.append(year)
            
            # Extract metrics for each year
            metrics = {}
            for year_idx, year in enumerate(years):
                col_idx = year_idx + 4  # Start from column 4
                year_metrics = {}
                
                # Extract from Income Statement
                year_metrics.update(self._extract_from_excel_sheet(
                    self.financial_data['income_fy'], col_idx, 'income'
                ))
                
                # Extract from Balance Sheet
                year_metrics.update(self._extract_from_excel_sheet(
                    self.financial_data['balance_fy'], col_idx, 'balance'
                ))
                
                # Extract from Cash Flow Statement
                year_metrics.update(self._extract_from_excel_sheet(
                    self.financial_data['cashflow_fy'], col_idx, 'cashflow'
                ))
                
                metrics[year] = year_metrics
            
            # Extract shares outstanding for per-share calculations
            self._extract_shares_outstanding()
            
            self.years = years
            self.metrics = metrics
            logger.info(f"Extracted metrics for years: {years}")
            logger.info(f"Sample metrics for {years[0] if years else 'N/A'}: {metrics.get(years[0], {}) if years else {}}")
            logger.info(f"Shares outstanding: {getattr(self, 'shares_outstanding', 'Not found')}")
            
        except Exception as e:
            logger.error(f"Error extracting financial metrics: {e}")
            raise
    
    def _extract_from_excel_sheet(self, df, col_idx, statement_type):
        """
        Extract specific metrics from a financial statement Excel sheet
        
        Args:
            df (pd.DataFrame): Financial statement data
            col_idx (int): Column index for the year
            statement_type (str): Type of statement ('income', 'balance', 'cashflow')
            
        Returns:
            dict: Extracted metrics
        """
        metrics = {}
        
        try:
            # Define metric mappings with partial string matching
            metric_mappings = {
                'income': {
                    'Net Income to Company': 'net_income',
                    'Operating Income': 'ebit',  # Use Operating Income as EBIT proxy
                    'Income Tax Expense': 'tax_expense',
                    'Net Interest Expenses': 'interest_expense',
                    'Revenue': 'revenue'
                },
                'balance': {
                    'Total Current Assets': 'current_assets',
                    'Total Current Liabilities': 'current_liabilities',
                    'Total Assets': 'total_assets',
                    'Total Debt': 'total_debt'
                },
                'cashflow': {
                    'Cash from Operations': 'operating_cash_flow',
                    'Capital Expenditures': 'capex',
                    'Depreciation': 'depreciation_amortization',
                    'Cash from Financing': 'financing_cash_flow'
                }
            }
            
            if statement_type in metric_mappings:
                for search_term, metric_name in metric_mappings[statement_type].items():
                    # Search for the metric in column 3 (index 2)
                    for idx, row in df.iterrows():
                        cell_value = row.iloc[2] if len(row) > 2 else None
                        if pd.notna(cell_value) and search_term in str(cell_value):
                            if col_idx < len(df.columns):
                                raw_value = row.iloc[col_idx]
                                if pd.notna(raw_value) and str(raw_value) not in ['None', 'nan', '']:
                                    try:
                                        # Handle string values like '74,989.0' or '-3,303.0'
                                        if isinstance(raw_value, str):
                                            # Remove commas and convert to float
                                            clean_value = raw_value.replace(',', '').replace('%', '')
                                            # Handle negative values in parentheses
                                            if '(' in clean_value and ')' in clean_value:
                                                clean_value = '-' + clean_value.replace('(', '').replace(')', '')
                                            metrics[metric_name] = float(clean_value)
                                        else:
                                            metrics[metric_name] = float(raw_value)
                                        
                                        logger.debug(f"Found {metric_name}: {metrics[metric_name]} for {search_term}")
                                    except (ValueError, TypeError) as e:
                                        logger.warning(f"Could not convert {raw_value} to float for {metric_name}: {e}")
                                        metrics[metric_name] = 0
                                break
                            
        except Exception as e:
            logger.warning(f"Error extracting from {statement_type} statement: {e}")
            
        return metrics
    
    def _extract_shares_outstanding(self):
        """
        Extract shares outstanding from yfinance first, then fall back to financial statements
        """
        try:
            # First try to fetch from yfinance
            if self.fetch_ticker_data():
                logger.info(f"Using yfinance data for shares outstanding: {self.shares_outstanding / 1000000:.1f}M")
                return
            
            # Fall back to financial statements
            shares_found = False
            
            # Try multiple sources for shares outstanding
            search_sources = [
                ('income_fy', 'Income Statement'),
                ('balance_fy', 'Balance Sheet'),
                ('cashflow_fy', 'Cash Flow Statement')
            ]
            
            # Enhanced search terms for shares outstanding
            search_terms = [
                'Shares Outstanding',
                'Outstanding Shares',
                'Common Shares Outstanding', 
                'Weighted Average Shares',
                'Shares Used in Computing',
                'Basic Shares Outstanding',
                'Diluted Shares Outstanding',
                'Common Stock Outstanding',
                'Number of Shares',
                'Share Count',
                'Ordinary Shares'
            ]
            
            for source_key, source_name in search_sources:
                if source_key in self.financial_data and not shares_found:
                    sheet = self.financial_data[source_key]
                    
                    for search_term in search_terms:
                        for idx, row in sheet.iterrows():
                            cell_value = row.iloc[2] if len(row) > 2 else None
                            if pd.notna(cell_value) and search_term.lower() in str(cell_value).lower():
                                # Get the most recent year's shares (last column with data)
                                for col_idx in range(len(sheet.columns) - 1, 3, -1):
                                    raw_value = row.iloc[col_idx]
                                    if pd.notna(raw_value) and str(raw_value) not in ['None', 'nan', '']:
                                        try:
                                            if isinstance(raw_value, str):
                                                clean_value = raw_value.replace(',', '').replace('%', '').replace('(', '').replace(')', '')
                                                # Handle different units (thousands, millions)
                                                value = float(clean_value)
                                                # If value is less than 10,000, it's likely in millions
                                                if value < 10000:
                                                    self.shares_outstanding = value * 1000000
                                                else:
                                                    # If value is large, it's likely already in actual shares
                                                    self.shares_outstanding = value
                                            else:
                                                value = float(raw_value)
                                                # Apply same logic for numeric values
                                                if value < 10000:
                                                    self.shares_outstanding = value * 1000000
                                                else:
                                                    self.shares_outstanding = value
                                            
                                            shares_found = True
                                            logger.info(f"Found shares outstanding in {source_name}: {self.shares_outstanding / 1000000:.1f}M from '{search_term}'")
                                            break
                                        except (ValueError, TypeError):
                                            continue
                                if shares_found:
                                    break
                        if shares_found:
                            break
            
            # If not found, estimate from financial metrics
            if not shares_found:
                self.shares_outstanding = self._estimate_shares_from_metrics()
                logger.warning(f"Shares outstanding not found directly, using estimate: {self.shares_outstanding / 1000000:.1f}M")
                
        except Exception as e:
            logger.warning(f"Error extracting shares outstanding: {e}")
            self.shares_outstanding = 100000000  # 100M shares default
    
    def _estimate_shares_from_metrics(self):
        """
        Estimate shares outstanding from available financial metrics
        """
        try:
            # Try to estimate from Net Income and company size
            latest_year = self.years[-1] if self.years else None
            if latest_year:
                latest_metrics = self.metrics.get(latest_year, {})
                net_income = latest_metrics.get('net_income', 0) * 1000000  # Convert to actual value
                revenue = latest_metrics.get('revenue', 0) * 1000000
                
                # Estimate shares based on company size
                if revenue > 10000000000:  # > $10B revenue
                    return 1000000000  # 1B shares
                elif revenue > 1000000000:  # > $1B revenue
                    return 500000000   # 500M shares
                else:
                    return 100000000   # 100M shares
            
            return 100000000  # Default 100M shares
            
        except Exception as e:
            logger.warning(f"Error estimating shares from metrics: {e}")
            return 100000000
    
    def calculate_fcf_to_firm(self):
        """
        Calculate Free Cash Flow to Firm (FCFF)
        Formula: FCFF = EBIT(1 - Tax Rate) + Depreciation & Amortization - Changes in Working Capital - CapEx
        """
        fcff_values = []
        
        for year in self.years:
            try:
                metrics = self.metrics[year]
                
                # Calculate components
                ebit = metrics.get('ebit', 0)
                revenue = metrics.get('revenue', 0)
                tax_expense = metrics.get('tax_expense', 0)
                
                # Calculate tax rate from revenue and tax expense
                tax_rate = abs(tax_expense) / revenue if revenue != 0 else 0.25  # Default 25%
                
                depreciation = metrics.get('depreciation_amortization', 0)
                capex = abs(metrics.get('capex', 0))  # CapEx is usually negative
                
                # Calculate working capital change (simplified - assume 2% of revenue growth)
                wc_change = revenue * 0.02
                
                # Calculate FCFF (multiply by 1M since values are in millions)
                fcff = (ebit * (1 - tax_rate) + depreciation - wc_change - capex) * 1000000
                fcff_values.append(fcff)
                
                logger.debug(f"FCFF {year}: EBIT={ebit}, Tax Rate={tax_rate:.2%}, Dep={depreciation}, WC Change={wc_change}, CapEx={capex}, FCFF={fcff/1000000:.2f}M")
                
            except Exception as e:
                logger.warning(f"Error calculating FCFF for {year}: {e}")
                fcff_values.append(0)
        
        return fcff_values
    
    def calculate_fcf_to_equity(self):
        """
        Calculate Free Cash Flow to Equity (FCFE)
        Formula: FCFE = Net Income + Depreciation & Amortization - Changes in Working Capital - CapEx + Net Borrowing
        """
        fcfe_values = []
        
        for year in self.years:
            try:
                metrics = self.metrics[year]
                
                # Calculate components
                net_income = metrics.get('net_income', 0)
                depreciation = metrics.get('depreciation_amortization', 0)
                capex = abs(metrics.get('capex', 0))
                revenue = metrics.get('revenue', 0)
                
                # Calculate working capital change (simplified - assume 2% of revenue)
                wc_change = revenue * 0.02
                
                # Estimate net borrowing from financing cash flow
                financing_cf = metrics.get('financing_cash_flow', 0)
                net_borrowing = max(0, financing_cf)  # Only positive financing (new debt)
                
                # Calculate FCFE (multiply by 1M since values are in millions)
                fcfe = (net_income + depreciation - wc_change - capex + net_borrowing) * 1000000
                fcfe_values.append(fcfe)
                
                logger.debug(f"FCFE {year}: Net Income={net_income}, Dep={depreciation}, WC Change={wc_change}, CapEx={capex}, Net Borrowing={net_borrowing}, FCFE={fcfe/1000000:.2f}M")
                
            except Exception as e:
                logger.warning(f"Error calculating FCFE for {year}: {e}")
                fcfe_values.append(0)
        
        return fcfe_values
    
    def calculate_levered_fcf(self):
        """
        Calculate Levered Free Cash Flow (LFCF)
        Formula: LFCF = Cash Flow from Operations - CapEx
        """
        lfcf_values = []
        
        for year in self.years:
            try:
                metrics = self.metrics[year]
                
                # Calculate components
                operating_cf = metrics.get('operating_cash_flow', 0)
                capex = abs(metrics.get('capex', 0))
                
                # Calculate LFCF (multiply by 1M since values are in millions)
                lfcf = (operating_cf - capex) * 1000000
                lfcf_values.append(lfcf)
                
                logger.debug(f"LFCF {year}: Operating CF={operating_cf}, CapEx={capex}, LFCF={lfcf/1000000:.2f}M")
                
            except Exception as e:
                logger.warning(f"Error calculating LFCF for {year}: {e}")
                lfcf_values.append(0)
        
        return lfcf_values
    
    def calculate_all_fcf_types(self):
        """
        Calculate all three types of FCF
        """
        self.fcf_results['FCFF'] = self.calculate_fcf_to_firm()
        self.fcf_results['FCFE'] = self.calculate_fcf_to_equity()
        self.fcf_results['LFCF'] = self.calculate_levered_fcf()
        
        logger.info("All FCF calculations completed")
    
    def plot_fcf_comparison(self, interactive=True):
        """
        Create an interactive comparison plot with FCF and DCF analysis tabs
        
        Args:
            interactive (bool): If True, display interactive plot; if False, save as image
        """
        try:
            # Initialize the figure with tab functionality
            self.current_tab = 'FCF'  # Track current tab
            fig = plt.figure(figsize=(20, 12))
            
            # Create tab buttons
            ax_fcf_button = plt.axes([0.1, 0.95, 0.1, 0.04])
            ax_dcf_button = plt.axes([0.21, 0.95, 0.1, 0.04])
            
            self.fcf_button = Button(ax_fcf_button, 'FCF Analysis')
            self.dcf_button = Button(ax_dcf_button, 'DCF Analysis')
            
            # Store figure reference for callbacks
            self.fig = fig
            
            # Create initial FCF tab content
            self._create_fcf_tab()
            
            # Set up button callbacks
            self.fcf_button.on_clicked(self._show_fcf_tab)
            self.dcf_button.on_clicked(self._show_dcf_tab)
            
            if interactive:
                plt.show()
                logger.info("Interactive plot displayed")
            else:
                plot_path = f"fcf_comparison_{self.company_name}.png"
                plt.savefig(plot_path, dpi=300, bbox_inches='tight')
                logger.info(f"Plot saved to {plot_path}")
                
        except Exception as e:
            logger.error(f"Error creating plot: {e}")
            raise
    
    def _calculate_relative_slopes(self, x_values, y_values):
        """
        Calculate relative slopes for 1-10 year periods
        
        Args:
            x_values (array): X coordinates
            y_values (array): Y coordinates
            
        Returns:
            dict: Relative slopes in percentage
        """
        slopes = {}
        n = len(x_values)
        
        # Define periods from 1 to 10 years (or max available data)
        for period_years in range(1, min(11, n + 1)):
            period_name = f'{period_years}y'
            
            if period_years == 1:
                # For 1-year, calculate year-over-year change if we have at least 2 points
                if n >= 2:
                    # Use last two points for 1-year change
                    current_value = y_values[-1]
                    previous_value = y_values[-2]
                    if previous_value != 0:
                        yoy_change = ((current_value - previous_value) / abs(previous_value)) * 100
                        slopes[period_name] = yoy_change
                    else:
                        slopes[period_name] = 0
                else:
                    slopes[period_name] = 0
            else:
                # For multi-year periods, use linear regression
                if period_years <= n:
                    # Take the last period_years points
                    period_x = x_values[-period_years:]
                    period_y = y_values[-period_years:]
                    
                    if len(period_x) >= 2:
                        slope, _, _, _, _ = stats.linregress(period_x, period_y)
                        
                        # Calculate annualized relative slope as percentage
                        initial_value = period_y[0] if period_y[0] != 0 else 1
                        # Annualize the slope (slope per year)
                        annualized_slope = slope / abs(initial_value) * 100
                        slopes[period_name] = annualized_slope
                    else:
                        slopes[period_name] = 0
                else:
                    slopes[period_name] = 0
                
        return slopes
    
    def _create_slope_table(self, ax, slope_data):
        """
        Create a table showing relative slopes for each FCF type from 1-10 years
        
        Args:
            ax: Matplotlib axis for the table
            slope_data (dict): Slope data for each FCF type
        """
        ax.axis('off')
        ax.set_title('Annualized Slope Analysis (%)', fontsize=14, fontweight='bold', pad=20)
        
        # Get all available periods dynamically (1y to 10y)
        all_periods = set()
        for fcf_type in slope_data:
            all_periods.update(slope_data[fcf_type].keys())
        
        # Sort periods numerically (1y, 2y, 3y, etc.)
        periods = sorted(all_periods, key=lambda x: int(x.replace('y', '')))
        
        # Prepare table data
        table_data = []
        for fcf_type in ['FCFF', 'FCFE', 'LFCF']:
            row = [fcf_type]
            for period in periods:
                slope_val = slope_data[fcf_type].get(period, 0)
                if slope_val == 0:
                    row.append("N/A")
                else:
                    row.append(f"{slope_val:.1f}%")
            table_data.append(row)
        
        # Add average row
        avg_row = ["Average"]
        for period in periods:
            period_slopes = []
            for fcf_type in ['FCFF', 'FCFE', 'LFCF']:
                slope_val = slope_data[fcf_type].get(period, 0)
                if slope_val != 0:  # Only include valid slopes
                    period_slopes.append(slope_val)
            
            if period_slopes:
                avg_slope = sum(period_slopes) / len(period_slopes)
                avg_row.append(f"{avg_slope:.1f}%")
            else:
                avg_row.append("N/A")
        
        table_data.append(avg_row)
        
        # Create table with dynamic column count
        headers = ['FCF Type'] + periods
        table = ax.table(cellText=table_data,
                        colLabels=headers,
                        cellLoc='center',
                        loc='center',
                        bbox=[0, 0.1, 1, 0.8])
        
        # Style the table
        table.auto_set_font_size(False)
        table.set_fontsize(9)  # Smaller font for more columns
        table.scale(1.0, 1.8)  # Adjust scaling for more columns
        
        # Color code the cells based on slope values
        for i in range(len(table_data)):
            for j in range(1, len(headers)):
                cell_val = table_data[i][j-1]  # Get the actual cell value
                if j > 0 and cell_val != "N/A":  # Skip FCF Type column and N/A values
                    try:
                        slope_val = float(cell_val.replace('%', ''))
                        if i == len(table_data) - 1:  # Average row
                            if slope_val > 0:
                                table[(i+1, j)].set_facecolor('#FFD700')  # Gold for positive average
                            elif slope_val < 0:
                                table[(i+1, j)].set_facecolor('#FFA500')  # Orange for negative average
                            else:
                                table[(i+1, j)].set_facecolor('#F0F0F0')  # Light gray for zero
                        else:  # Regular FCF rows
                            if slope_val > 0:
                                table[(i+1, j)].set_facecolor('#90EE90')  # Light green for positive
                            elif slope_val < 0:
                                table[(i+1, j)].set_facecolor('#FFB6C1')  # Light red for negative
                            else:
                                table[(i+1, j)].set_facecolor('#F0F0F0')  # Light gray for zero
                    except ValueError:
                        table[(i+1, j)].set_facecolor('#F0F0F0')  # Light gray for N/A
                elif cell_val == "N/A":
                    if i == len(table_data) - 1:  # Average row
                        table[(i+1, j)].set_facecolor('#E0E0E0')  # Darker gray for N/A in average
                    else:
                        table[(i+1, j)].set_facecolor('#E0E0E0')  # Darker gray for N/A
        
        # Style header row
        for j in range(len(headers)):
            table[(0, j)].set_facecolor('#4472C4')
            table[(0, j)].set_text_props(weight='bold', color='white')
        
        # Style average row label
        table[(len(table_data), 0)].set_text_props(weight='bold')
        table[(len(table_data), 0)].set_facecolor('#D3D3D3')  # Light gray background for average row label
            
        # Add explanation text
        ax.text(0.5, 0.05, '1y = Year-over-Year change, 2y+ = Annualized linear regression slope', 
                ha='center', va='bottom', fontsize=10, style='italic', transform=ax.transAxes)
    
    def _create_slope_graph(self, ax, slope_data):
        """
        Create a graph showing annualized slopes and their average
        
        Args:
            ax: Matplotlib axis for the graph
            slope_data (dict): Slope data for each FCF type
        """
        # Get all available periods dynamically (excluding 1y since it's YoY change, not annualized slope)
        all_periods = set()
        for fcf_type in slope_data:
            all_periods.update(slope_data[fcf_type].keys())
        
        # Sort periods numerically and exclude 1y for slope analysis
        periods = sorted([p for p in all_periods if p != '1y'], key=lambda x: int(x.replace('y', '')))
        
        if not periods:
            ax.text(0.5, 0.5, 'Insufficient data for slope analysis', 
                   ha='center', va='center', fontsize=14, transform=ax.transAxes)
            ax.axis('off')
            return
        
        # Extract period numbers for x-axis
        period_numbers = [int(p.replace('y', '')) for p in periods]
        
        # Plot slopes for each FCF type
        colors = {'FCFF': '#1f77b4', 'FCFE': '#ff7f0e', 'LFCF': '#2ca02c'}
        markers = {'FCFF': 'o', 'FCFE': 's', 'LFCF': '^'}
        
        avg_slopes = []
        
        for period in periods:
            period_slopes = []
            for fcf_type in ['FCFF', 'FCFE', 'LFCF']:
                slope_val = slope_data[fcf_type].get(period, 0)
                if slope_val != 0:  # Only include valid slopes
                    period_slopes.append(slope_val)
            
            # Calculate average for this period
            if period_slopes:
                avg_slopes.append(sum(period_slopes) / len(period_slopes))
            else:
                avg_slopes.append(0)
        
        # Plot individual FCF slopes
        for fcf_type in ['FCFF', 'FCFE', 'LFCF']:
            y_values = []
            for period in periods:
                slope_val = slope_data[fcf_type].get(period, 0)
                y_values.append(slope_val if slope_val != 0 else None)
            
            # Plot line, handling None values
            valid_indices = [i for i, val in enumerate(y_values) if val is not None]
            if valid_indices:
                valid_x = [period_numbers[i] for i in valid_indices]
                valid_y = [y_values[i] for i in valid_indices]
                
                ax.plot(valid_x, valid_y, marker=markers[fcf_type], linewidth=2, 
                       markersize=8, label=fcf_type, color=colors[fcf_type], alpha=0.8)
        
        # Plot average line
        valid_avg_indices = [i for i, val in enumerate(avg_slopes) if val != 0]
        if valid_avg_indices:
            valid_avg_x = [period_numbers[i] for i in valid_avg_indices]
            valid_avg_y = [avg_slopes[i] for i in valid_avg_indices]
            
            ax.plot(valid_avg_x, valid_avg_y, marker='D', linewidth=3, 
                   markersize=10, label='Average', color='red', alpha=0.9, linestyle='--')
        
        # Customize the graph
        ax.set_xlabel('Time Period (Years)', fontsize=12)
        ax.set_ylabel('Annualized Slope (%)', fontsize=12)
        ax.set_title('Annualized Linear Regression Slopes by Time Period', fontsize=14, fontweight='bold')
        ax.legend(fontsize=11)
        ax.grid(True, alpha=0.3)
        
        # Set x-axis ticks to show all available periods
        ax.set_xticks(period_numbers)
        ax.set_xticklabels([f'{p}y' for p in period_numbers])
        
        # Add horizontal line at y=0 for reference
        ax.axhline(y=0, color='black', linestyle='-', alpha=0.3, linewidth=1)
    
    def _create_fcf_tab(self):
        """Create FCF analysis tab content"""
        # Clear existing axes (except buttons)
        for ax in self.fig.axes[2:]:  # Keep first two axes (buttons)
            ax.remove()
            
        # Create FCF tab layout
        ax1 = plt.subplot2grid((2, 2), (0, 0), rowspan=2)  # FCF plot - spans full left vertical
        ax2 = plt.subplot2grid((2, 2), (0, 1), colspan=1)  # Slope table - top right
        ax3 = plt.subplot2grid((2, 2), (1, 1), colspan=1)  # Slope graph - bottom right
        
        # Convert values to millions for better readability
        fcff_millions = [x / 1000000 for x in self.fcf_results['FCFF']]
        fcfe_millions = [x / 1000000 for x in self.fcf_results['FCFE']]
        lfcf_millions = [x / 1000000 for x in self.fcf_results['LFCF']]
        
        # Create numeric x values for calculations
        x_numeric = np.arange(len(self.years))
        
        # Plot lines instead of bars
        ax1.plot(x_numeric, fcff_millions, marker='o', linewidth=2, markersize=8, 
                label='FCFF (Free Cash Flow to Firm)', color='#1f77b4')
        ax1.plot(x_numeric, fcfe_millions, marker='s', linewidth=2, markersize=8, 
                label='FCFE (Free Cash Flow to Equity)', color='#ff7f0e')
        ax1.plot(x_numeric, lfcf_millions, marker='^', linewidth=2, markersize=8, 
                label='LFCF (Levered Free Cash Flow)', color='#2ca02c')
        
        # Add linear fits
        fcf_data = {'FCFF': fcff_millions, 'FCFE': fcfe_millions, 'LFCF': lfcf_millions}
        colors = {'FCFF': '#1f77b4', 'FCFE': '#ff7f0e', 'LFCF': '#2ca02c'}
        slope_data = {}
        
        for fcf_type, values in fcf_data.items():
            slope, intercept, r_value, p_value, std_err = stats.linregress(x_numeric, values)
            line = slope * x_numeric + intercept
            ax1.plot(x_numeric, line, '--', color=colors[fcf_type], alpha=0.7, linewidth=1.5)
            
            # Calculate relative slopes for different periods
            slope_data[fcf_type] = self._calculate_relative_slopes(x_numeric, values)
            
        # Customize plot
        ax1.set_xlabel('Year', fontsize=12)
        ax1.set_ylabel('Free Cash Flow (Millions USD)', fontsize=12)
        ax1.set_title(f'Free Cash Flow Analysis with Linear Fits - {self.company_name}', fontsize=16, fontweight='bold')
        ax1.set_xticks(x_numeric)
        ax1.set_xticklabels(self.years, rotation=45)
        ax1.legend(fontsize=11)
        ax1.grid(True, alpha=0.3)
        
        # Create slope analysis table
        self._create_slope_table(ax2, slope_data)
        
        # Create slope analysis graph
        self._create_slope_graph(ax3, slope_data)
    
    def _create_dcf_tab(self):
        """Create DCF analysis tab content - Excel-style DCF model"""
        # Clear existing axes (except buttons)
        for ax in self.fig.axes[2:]:  # Keep first two axes (buttons)
            ax.remove()
            
        # Create main layout - single full canvas for Excel-style layout
        ax_main = plt.subplot2grid((1, 1), (0, 0))
        ax_main.set_xlim(0, 16)  # Wider canvas for better spacing
        ax_main.set_ylim(0, 14)  # Taller canvas for better spacing
        ax_main.axis('off')
        
        # Initialize DCF assumptions if not already set
        if not hasattr(self, 'dcf_assumptions'):
            # Calculate slope data to get the 5-year average slope
            fcff_millions = [x / 1000000 for x in self.fcf_results['FCFF']]
            fcfe_millions = [x / 1000000 for x in self.fcf_results['FCFE']]
            lfcf_millions = [x / 1000000 for x in self.fcf_results['LFCF']]
            x_numeric = np.arange(len(self.years))
            
            # Calculate slopes for each FCF type
            slope_data = {}
            slope_data['FCFF'] = self._calculate_relative_slopes(x_numeric, fcff_millions)
            slope_data['FCFE'] = self._calculate_relative_slopes(x_numeric, fcfe_millions)
            slope_data['LFCF'] = self._calculate_relative_slopes(x_numeric, lfcf_millions)
            
            # Get 5-year average slope
            five_year_slopes = []
            for fcf_type in ['FCFF', 'FCFE', 'LFCF']:
                if '5y' in slope_data[fcf_type] and slope_data[fcf_type]['5y'] != 0:
                    five_year_slopes.append(slope_data[fcf_type]['5y'])
            
            # Calculate average 5-year slope
            if five_year_slopes:
                avg_5yr_slope = sum(five_year_slopes) / len(five_year_slopes)
                # Convert from percentage to decimal
                growth_rate_yr1_5 = avg_5yr_slope / 100.0
                growth_rate_yr5_10 = growth_rate_yr1_5 / 2.0  # Half of 5yr average slope
            else:
                # Fallback to default values
                growth_rate_yr1_5 = 0.07  # 7%
                growth_rate_yr5_10 = 0.05  # 5%
            
            self.dcf_assumptions = {
                'growth_rate_yr1_5': growth_rate_yr1_5,
                'growth_rate_yr5_10': growth_rate_yr5_10,
                'discount_rate': 0.12,  # 12%
                'terminal_growth_rate': 0.03,  # 3%
                'terminal_method': 'gordon_growth',  # 'gordon_growth' or 'terminal_multiple'
                'terminal_multiple_input': 11.44  # Used if terminal_method is 'terminal_multiple'
            }
        
        # Calculate DCF projections and metrics
        dcf_data = self._calculate_dcf_projections_excel()
        
        # Store reference to main axis for updates
        self.dcf_main_ax = ax_main
        
        # Create all components of the Excel-style layout with proper spacing
        self._create_growth_estimates_box(ax_main, dcf_data)  # Top left - slope table
        self._create_dcf_growth_table_interactive(ax_main, dcf_data)  # Top right - DCF growth
        self._create_assumption_boxes_interactive(ax_main, dcf_data)  # Middle right - assumptions
        self._create_main_dcf_table(ax_main, dcf_data)  # Center - main DCF table
        self._create_company_info_box(ax_main, dcf_data)  # Bottom left - company info
    
    def _calculate_dcf_projections_excel(self):
        """Calculate DCF projections and metrics for Excel-style layout"""
        try:
            # Get base FCF (use FCFF as primary)
            fcff_millions = [x / 1000000 for x in self.fcf_results['FCFF']]
            base_fcf = fcff_millions[-1] if fcff_millions else 100  # Latest FCF
            
            # Calculate historical growth rates
            growth_rates = self._calculate_historical_growth_rates_excel(fcff_millions)
            
            # DCF assumptions - use dynamic values if available
            if hasattr(self, 'dcf_assumptions'):
                assumptions = {
                    'base_fcf': base_fcf,
                    'growth_rate_yr1_5': self.dcf_assumptions['growth_rate_yr1_5'],
                    'growth_rate_yr5_10': self.dcf_assumptions['growth_rate_yr5_10'],
                    'terminal_growth_rate': self.dcf_assumptions['terminal_growth_rate'],
                    'discount_rate': self.dcf_assumptions['discount_rate'],
                    'shares_outstanding': self.shares_outstanding,
                    'current_stock_price': self.current_stock_price if self.current_stock_price else 178.70,  # Use fetched price or default
                    'projection_years': 10
                }
            else:
                assumptions = {
                    'base_fcf': base_fcf,
                    'growth_rate_yr1_5': growth_rates.get('3yr', 0.07),  # Use 3yr growth for yr 1-5
                    'growth_rate_yr5_10': growth_rates.get('5yr', 0.05),  # Use 5yr growth for yr 5-10
                    'terminal_growth_rate': 0.03,  # 3% terminal growth
                    'discount_rate': 0.12,  # 12% required rate of return
                    'shares_outstanding': self.shares_outstanding,
                    'current_stock_price': self.current_stock_price if self.current_stock_price else 178.70,  # Use fetched price or default
                    'projection_years': 10
                }
            
            # Calculate 10-year projections
            projected_fcf = []
            current_fcf = base_fcf
            
            for year in range(1, 11):  # Years 1-10
                if year <= 5:
                    growth_rate = assumptions['growth_rate_yr1_5']
                else:
                    growth_rate = assumptions['growth_rate_yr5_10']
                    
                current_fcf *= (1 + growth_rate)
                projected_fcf.append(current_fcf)
            
            # Calculate terminal value using selected method
            terminal_fcf = projected_fcf[-1] * (1 + assumptions['terminal_growth_rate'])
            
            # Check if user has set a terminal method preference
            terminal_method = assumptions.get('terminal_method', 'gordon_growth')
            
            if terminal_method == 'terminal_multiple' and 'terminal_multiple_input' in assumptions:
                # Use user-specified terminal multiple
                terminal_multiple = assumptions['terminal_multiple_input']
                terminal_value = terminal_fcf * terminal_multiple
            else:
                # Use Gordon Growth Model (default)
                terminal_value = terminal_fcf / (assumptions['discount_rate'] - assumptions['terminal_growth_rate'])
                # Calculate terminal multiple for display consistency
                # Terminal Multiple = Terminal Value / Terminal FCF
                # This represents the multiple of terminal FCF that the terminal value represents
                terminal_multiple = terminal_value / terminal_fcf if terminal_fcf != 0 else 0
            
            # Add terminal multiple to assumptions for display
            assumptions['terminal_multiple'] = terminal_multiple
            
            # Calculate present values
            pv_fcf = []
            discount_factors = []
            
            for i in range(1, 11):
                discount_factor = 1 / ((1 + assumptions['discount_rate']) ** i)
                discount_factors.append(discount_factor)
                pv = projected_fcf[i-1] * discount_factor
                pv_fcf.append(pv)
            
            # Terminal value PV
            terminal_discount_factor = 1 / ((1 + assumptions['discount_rate']) ** 10)
            pv_terminal = terminal_value * terminal_discount_factor
            
            # Calculate enterprise and equity value
            enterprise_value = sum(pv_fcf) + pv_terminal
            net_debt = 0  # Simplified - would need actual data
            equity_value = enterprise_value - net_debt
            
            # Per share calculations
            equity_value_per_share = equity_value * 1000000 / assumptions['shares_outstanding']
            
            # Create projection years
            current_year = int(self.years[-1]) if self.years else 2025
            projection_years = list(range(current_year + 1, current_year + 11))
            
            return {
                'assumptions': assumptions,
                'growth_rates': growth_rates,
                'projection_years': projection_years,
                'projected_fcf': projected_fcf,
                'pv_fcf': pv_fcf,
                'discount_factors': discount_factors,
                'terminal_value': terminal_value,
                'terminal_multiple': terminal_multiple,
                'terminal_fcf': terminal_fcf,
                'pv_terminal': pv_terminal,
                'enterprise_value': enterprise_value,
                'equity_value': equity_value,
                'equity_value_per_share': equity_value_per_share,
                'base_fcf': base_fcf
            }
            
        except Exception as e:
            logger.error(f"Error calculating DCF projections: {e}")
            return None
    
    def _calculate_historical_growth_rates_excel(self, fcf_values):
        """Calculate historical growth rates for different periods"""
        growth_rates = {}
        
        if len(fcf_values) >= 3:
            # 3-year growth rate
            start_val = fcf_values[-4] if len(fcf_values) >= 4 else fcf_values[0]
            end_val = fcf_values[-1]
            if start_val != 0:
                years = min(3, len(fcf_values) - 1)
                growth_rates['3yr'] = (end_val / abs(start_val)) ** (1/years) - 1
            
        if len(fcf_values) >= 5:
            # 5-year growth rate
            start_val = fcf_values[-6] if len(fcf_values) >= 6 else fcf_values[0]
            end_val = fcf_values[-1]
            if start_val != 0:
                years = min(5, len(fcf_values) - 1)
                growth_rates['5yr'] = (end_val / abs(start_val)) ** (1/years) - 1
                
        if len(fcf_values) >= 10:
            # 10-year growth rate
            start_val = fcf_values[0]
            end_val = fcf_values[-1]
            if start_val != 0:
                years = len(fcf_values) - 1
                growth_rates['10yr'] = (end_val / abs(start_val)) ** (1/years) - 1
        
        return growth_rates
    
    def _create_main_dcf_table(self, ax, dcf_data):
        """Create the main DCF projection table"""
        if not dcf_data:
            return
            
        # Extract assumptions from dcf_data
        assumptions = dcf_data['assumptions']
        
        # Table position and size - center of canvas with more space
        table_x = 0.5
        table_y = 8.0
        table_width = 12.0
        table_height = 3.0
        
        # Create table data
        headers = ['period', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'Terminal\nvalue']
        
        # Year row
        year_row = ['Year', str(dcf_data['projection_years'][0] - 1)] + [str(y) for y in dcf_data['projection_years']] + ['']
        
        # FCF row
        fcf_row = ['FCF', f"{dcf_data['base_fcf']:.2f}"] + [f"{fcf:.2f}" for fcf in dcf_data['projected_fcf']] + [f"{dcf_data['terminal_value']:.2f}"]
        
        # Growth rate row
        growth_row = ['Growth rate', ''] + [f"{assumptions['growth_rate_yr1_5']*100:.1f}%"] * 5 + [f"{assumptions['growth_rate_yr5_10']*100:.1f}%"] * 5 + [f"{assumptions['terminal_growth_rate']*100:.1f}%"]
        
        # Discount rate row
        discount_factor = 1 + assumptions['discount_rate']
        discount_row = ['Discount rate'] + [f"{discount_factor:.2f}"] + [f"{discount_factor**i:.2f}" for i in range(1, 11)] + [f"{discount_factor**10:.2f}"]
        
        # Discounted CF row
        discounted_row = ['Discounted CF', f"{dcf_data['base_fcf']:.2f}"] + [f"{pv:.2f}" for pv in dcf_data['pv_fcf']] + [f"{dcf_data['pv_terminal']:.2f}"]
        
        table_data = [year_row, fcf_row, growth_row, discount_row, discounted_row]
        
        # Create table
        table = ax.table(cellText=table_data,
                        colLabels=headers,
                        cellLoc='center',
                        loc='center',
                        bbox=[table_x/16, table_y/14, table_width/16, table_height/14])
        
        # Style the table
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1.0, 1.8)
        
        # Color coding
        for i in range(len(headers)):
            table[(0, i)].set_facecolor('#4472C4')
            table[(0, i)].set_text_props(weight='bold', color='white')
        
        # Alternate row colors
        for i in range(1, len(table_data) + 1):
            for j in range(len(headers)):
                if i % 2 == 0:
                    table[(i, j)].set_facecolor('#F2F2F2')
                else:
                    table[(i, j)].set_facecolor('#FFFFFF')
    
    def _create_growth_estimates_box(self, ax, dcf_data):
        """Create slope estimates box displaying 1yr, 3yr, 5yr, 8yr for all FCF types"""
        if not dcf_data:
            return
            
        # Box position - top left with better spacing
        box_x = 0.5
        box_y = 11.5
        
        # Calculate slope data for all FCF types
        fcff_millions = [x / 1000000 for x in self.fcf_results['FCFF']]
        fcfe_millions = [x / 1000000 for x in self.fcf_results['FCFE']]
        lfcf_millions = [x / 1000000 for x in self.fcf_results['LFCF']]
        x_numeric = np.arange(len(self.years))
        
        # Calculate slopes for each FCF type
        slope_data = {}
        slope_data['FCFF'] = self._calculate_relative_slopes(x_numeric, fcff_millions)
        slope_data['FCFE'] = self._calculate_relative_slopes(x_numeric, fcfe_millions)
        slope_data['LFCF'] = self._calculate_relative_slopes(x_numeric, lfcf_millions)
        
        # Create slope estimates table: Period, FCFF, FCFE, LFCF, Average
        slope_display_data = []
        desired_periods = ['1y', '3y', '5y', '8y']
        
        for period in desired_periods:
            row = [period]
            period_slopes = []
            
            # Add FCFF, FCFE, LFCF slopes
            for fcf_type in ['FCFF', 'FCFE', 'LFCF']:
                if period in slope_data[fcf_type]:
                    slope_val = slope_data[fcf_type][period]
                    row.append(f"{slope_val:.1f}%")
                    if slope_val != 0:  # Only include valid slopes in average
                        period_slopes.append(slope_val)
                else:
                    row.append("N/A")
            
            # Calculate and add average
            if period_slopes:
                avg_slope = sum(period_slopes) / len(period_slopes)
                row.append(f"{avg_slope:.1f}%")
            else:
                row.append("N/A")
            
            slope_display_data.append(row)
        
        if slope_display_data:
            slope_table = ax.table(cellText=slope_display_data,
                                  colLabels=['Period', 'FCFF', 'FCFE', 'LFCF', 'Average'],
                                  cellLoc='center',
                                  loc='center',
                                  bbox=[box_x/16, box_y/14, 5.0/16, 1.8/14])
            
            slope_table.auto_set_font_size(False)
            slope_table.set_fontsize(10)
            slope_table.scale(1.0, 1.6)
            
            # Style header
            for j in range(5):
                slope_table[(0, j)].set_facecolor('#4472C4')
                slope_table[(0, j)].set_text_props(weight='bold', color='white')
            
            # Style data rows with color coding based on slope values
            for i in range(1, len(slope_display_data) + 1):
                for j in range(5):
                    if j == 0:  # Period column
                        slope_table[(i, j)].set_facecolor('#F0F0F0')
                    else:
                        cell_val = slope_display_data[i-1][j]
                        if cell_val != "N/A":
                            try:
                                slope_val = float(cell_val.replace('%', ''))
                                if j == 4:  # Average column
                                    if slope_val > 0:
                                        slope_table[(i, j)].set_facecolor('#FFD700')  # Gold for positive average
                                    elif slope_val < 0:
                                        slope_table[(i, j)].set_facecolor('#FFA500')  # Orange for negative average
                                    else:
                                        slope_table[(i, j)].set_facecolor('#F0F0F0')  # Light gray for zero
                                else:  # FCF type columns
                                    if slope_val > 0:
                                        slope_table[(i, j)].set_facecolor('#90EE90')  # Light green for positive
                                    elif slope_val < 0:
                                        slope_table[(i, j)].set_facecolor('#FFB6C1')  # Light red for negative
                                    else:
                                        slope_table[(i, j)].set_facecolor('#F0F0F0')  # Light gray for zero
                            except ValueError:
                                slope_table[(i, j)].set_facecolor('#E0E0E0')  # Gray for N/A
                        else:
                            slope_table[(i, j)].set_facecolor('#E0E0E0')  # Gray for N/A
    
    def _create_dcf_growth_table_interactive(self, ax, dcf_data):
        """Create interactive DCF growth estimate table with text boxes"""
        if not dcf_data:
            return
            
        # Table position - top right corner with better spacing
        table_x = 12.5
        table_y = 12.0
        
        # Create DCF growth table with current values
        dcf_growth_data = [
            ['yr 1-5', f"{self.dcf_assumptions['growth_rate_yr1_5']*100:.1f}%"],
            ['yr 5-10', f"{self.dcf_assumptions['growth_rate_yr5_10']*100:.1f}%"]
        ]
        
        dcf_growth_table = ax.table(cellText=dcf_growth_data,
                                   colLabels=['', 'DCF Growth estimate'],
                                   cellLoc='center',
                                   loc='center',
                                   bbox=[table_x/16, table_y/14, 3.0/16, 1.5/14])
        
        dcf_growth_table.auto_set_font_size(False)
        dcf_growth_table.set_fontsize(10)
        dcf_growth_table.scale(1.0, 1.5)
        
        # Style header
        dcf_growth_table[(0, 0)].set_facecolor('#D4E6F1')
        dcf_growth_table[(0, 1)].set_facecolor('#D4E6F1')
        dcf_growth_table[(0, 0)].set_text_props(weight='bold')
        dcf_growth_table[(0, 1)].set_text_props(weight='bold')
        
        # Style data rows
        for i in range(1, len(dcf_growth_data) + 1):
            dcf_growth_table[(i, 0)].set_facecolor('#EBF3FD')
            dcf_growth_table[(i, 1)].set_facecolor('#EBF3FD')
        
        # Add text boxes for interactive input - positioned directly next to the growth values
        # Year 1-5 growth rate text box - positioned next to the yr 1-5 row
        ax_textbox_yr1_5 = plt.axes([0.87, 0.79, 0.06, 0.03])
        self.textbox_yr1_5 = TextBox(ax_textbox_yr1_5, '', 
                                     initial=f"{self.dcf_assumptions['growth_rate_yr1_5']*100:.1f}")
        self.textbox_yr1_5.on_submit(self._update_growth_yr1_5)
        
        # Year 5-10 growth rate text box - positioned next to the yr 5-10 row
        ax_textbox_yr5_10 = plt.axes([0.87, 0.74, 0.06, 0.03])
        self.textbox_yr5_10 = TextBox(ax_textbox_yr5_10, '', 
                                      initial=f"{self.dcf_assumptions['growth_rate_yr5_10']*100:.1f}")
        self.textbox_yr5_10.on_submit(self._update_growth_yr5_10)
    
    def _create_dcf_growth_table(self, ax, dcf_data):
        """Create DCF growth estimate table"""
        if not dcf_data:
            return
            
        # Table position
        table_x = 4.0
        table_y = 8.0
        
        # Create DCF growth table
        dcf_growth_data = [
            ['yr 1-5', '7.0%'],
            ['yr 5-10', '5.0%']
        ]
        
        dcf_growth_table = ax.table(cellText=dcf_growth_data,
                                   colLabels=['', 'DCF Growth estimate'],
                                   cellLoc='center',
                                   loc='center',
                                   bbox=[table_x/16, table_y/14, 2.0/16, 1.0/14])
        
        dcf_growth_table.auto_set_font_size(False)
        dcf_growth_table.set_fontsize(10)
        dcf_growth_table.scale(1.0, 1.5)
        
        # Style header
        dcf_growth_table[(0, 0)].set_facecolor('#D4E6F1')
        dcf_growth_table[(0, 1)].set_facecolor('#D4E6F1')
        dcf_growth_table[(0, 0)].set_text_props(weight='bold')
        dcf_growth_table[(0, 1)].set_text_props(weight='bold')
        
        # Style data rows
        for i in range(1, len(dcf_growth_data) + 1):
            dcf_growth_table[(i, 0)].set_facecolor('#EBF3FD')
            dcf_growth_table[(i, 1)].set_facecolor('#EBF3FD')
    
    def _create_assumption_boxes(self, ax, dcf_data):
        """Create assumption boxes for terminal multiple, required rate, etc."""
        if not dcf_data:
            return
            
        assumptions = dcf_data['assumptions']
        
        # Terminal multiple box
        terminal_box_x = 7.0
        terminal_box_y = 8.0
        ax.text(terminal_box_x/16, terminal_box_y/14, 'Terminal multiple', 
                fontsize=10, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#D4E6F1'))
        ax.text(terminal_box_x/16, (terminal_box_y-0.5)/14, f"{assumptions['terminal_multiple']:.2f}", 
                fontsize=12, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#EBF3FD'))
        
        # Required rate of return box
        required_rate_x = 4.0
        required_rate_y = 3.0
        ax.text(required_rate_x/16, required_rate_y/14, 'Required rate of return', 
                fontsize=10, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#D4E6F1'))
        ax.text(required_rate_x/16, (required_rate_y-0.5)/14, f"{assumptions['discount_rate']*100:.0f}%", 
                fontsize=12, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#D5E8D4'))
        
        # Perpetual growth rate box
        perpetual_growth_x = 7.0
        perpetual_growth_y = 3.0
        ax.text(perpetual_growth_x/16, perpetual_growth_y/14, 'Perpetual growth rate', 
                fontsize=10, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#D4E6F1'))
        ax.text(perpetual_growth_x/16, (perpetual_growth_y-0.5)/14, f"{assumptions['terminal_growth_rate']*100:.2f}%", 
                fontsize=12, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#CCE5FF'))
    
    def _create_assumption_boxes_interactive(self, ax, dcf_data):
        """Create interactive assumption boxes for terminal multiple, required rate, etc."""
        if not dcf_data:
            return
            
        assumptions = dcf_data['assumptions']
        
        # Terminal multiple box - positioned right side with better spacing
        terminal_box_x = 13.0
        terminal_box_y = 9.0
        
        # Determine method used for terminal value calculation
        terminal_method = dcf_data.get('assumptions', {}).get('terminal_method', 'gordon_growth')
        method_text = 'Terminal multiple\n(from Gordon Growth)' if terminal_method == 'gordon_growth' else 'Terminal multiple\n(user input)'
        
        ax.text(terminal_box_x/16, terminal_box_y/14, method_text, 
                fontsize=10, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#D4E6F1'), ha='center')
        ax.text(terminal_box_x/16, (terminal_box_y-0.8)/14, f"{assumptions['terminal_multiple']:.2f}", 
                fontsize=14, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#EBF3FD'), ha='center')
        
        # Required rate of return box with text input - positioned with better spacing
        required_rate_x = 3.0
        required_rate_y = 5.0
        ax.text(required_rate_x/16, required_rate_y/14, 'Required rate of return', 
                fontsize=10, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#D4E6F1'), ha='center')
        ax.text(required_rate_x/16, (required_rate_y-0.8)/14, f"{assumptions['discount_rate']*100:.0f}%", 
                fontsize=12, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#D5E8D4'), ha='center')
        
        # Add text box for required rate of return - positioned next to the value
        ax_textbox_discount = plt.axes([0.25, 0.32, 0.06, 0.03])
        self.textbox_discount = TextBox(ax_textbox_discount, '', 
                                       initial=f"{self.dcf_assumptions['discount_rate']*100:.0f}")
        self.textbox_discount.on_submit(self._update_discount_rate)
        
        # Perpetual growth rate box with text input - positioned with better spacing
        perpetual_growth_x = 8.0
        perpetual_growth_y = 5.0
        ax.text(perpetual_growth_x/16, perpetual_growth_y/14, 'Perpetual growth rate', 
                fontsize=10, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#D4E6F1'), ha='center')
        ax.text(perpetual_growth_x/16, (perpetual_growth_y-0.8)/14, f"{assumptions['terminal_growth_rate']*100:.2f}%", 
                fontsize=12, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#CCE5FF'), ha='center')
        
        # Add text box for perpetual growth rate - positioned next to the value
        ax_textbox_terminal = plt.axes([0.55, 0.32, 0.06, 0.03])
        self.textbox_terminal = TextBox(ax_textbox_terminal, '', 
                                       initial=f"{self.dcf_assumptions['terminal_growth_rate']*100:.2f}")
        self.textbox_terminal.on_submit(self._update_terminal_growth)
    
    def _create_company_info_box(self, ax, dcf_data):
        """Create company information box"""
        if not dcf_data:
            return
            
        assumptions = dcf_data['assumptions']
        
        # Company info box position - bottom left with better spacing
        info_x = 0.5
        info_y = 3.5
        
        # Company name and symbol - use dynamic ticker symbol
        ticker_display = self.ticker_symbol if self.ticker_symbol else 'N/A'
        ax.text(info_x/16, info_y/14, ticker_display, 
                fontsize=16, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#90EE90'))
        ax.text(info_x/16, (info_y-0.8)/14, self.company_name.upper(), 
                fontsize=12, weight='bold', transform=ax.transAxes)
        
        # Price and shares
        ax.text(info_x/16, (info_y-1.3)/14, f'Price: ${assumptions["current_stock_price"]:.2f}', 
                fontsize=10, transform=ax.transAxes)
        ax.text(info_x/16, (info_y-1.6)/14, f'Shares: {assumptions["shares_outstanding"]/1000000:.0f},000,000', 
                fontsize=10, transform=ax.transAxes)
        
        # Market cap and fair value
        market_cap = assumptions['current_stock_price'] * assumptions['shares_outstanding'] / 1000000
        ax.text(info_x/16, (info_y-2.2)/14, f'Market CAP: {market_cap:.0f}', 
                fontsize=10, transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#FFE4B5'))
        ax.text(info_x/16, (info_y-2.8)/14, f'Fair Value: ${dcf_data["equity_value_per_share"]:.2f}', 
                fontsize=10, transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='#FFE4B5'))
        
        # Upside/Downside
        upside_pct = (dcf_data['equity_value_per_share'] - assumptions['current_stock_price']) / assumptions['current_stock_price'] * 100
        upside_color = '#90EE90' if upside_pct > 0 else '#FFB6C1'
        ax.text(info_x/16, (info_y-3.4)/14, f'Upside/Downside: {upside_pct:.2f}%', 
                fontsize=10, weight='bold', transform=ax.transAxes,
                bbox=dict(boxstyle="round,pad=0.3", facecolor=upside_color))
    
    def _update_growth_yr1_5(self, text):
        """Update growth rate for years 1-5"""
        try:
            value = float(text)
            self.dcf_assumptions['growth_rate_yr1_5'] = value / 100.0  # Convert percentage to decimal
            self._refresh_dcf_tab()
        except ValueError:
            logger.warning(f"Invalid input for growth rate yr 1-5: {text}")
    
    def _update_growth_yr5_10(self, text):
        """Update growth rate for years 5-10"""
        try:
            value = float(text)
            self.dcf_assumptions['growth_rate_yr5_10'] = value / 100.0  # Convert percentage to decimal
            self._refresh_dcf_tab()
        except ValueError:
            logger.warning(f"Invalid input for growth rate yr 5-10: {text}")
    
    def _update_discount_rate(self, text):
        """Update discount rate"""
        try:
            value = float(text)
            self.dcf_assumptions['discount_rate'] = value / 100.0  # Convert percentage to decimal
            self._refresh_dcf_tab()
        except ValueError:
            logger.warning(f"Invalid input for discount rate: {text}")
    
    def _update_terminal_growth(self, text):
        """Update terminal growth rate"""
        try:
            value = float(text)
            self.dcf_assumptions['terminal_growth_rate'] = value / 100.0  # Convert percentage to decimal
            self._refresh_dcf_tab()
        except ValueError:
            logger.warning(f"Invalid input for terminal growth rate: {text}")
    
    def _refresh_dcf_tab(self):
        """Refresh the DCF tab with updated values"""
        try:
            # Clear the main axis
            self.dcf_main_ax.clear()
            self.dcf_main_ax.set_xlim(0, 16)
            self.dcf_main_ax.set_ylim(0, 14)
            self.dcf_main_ax.axis('off')
            
            # Recalculate DCF projections
            dcf_data = self._calculate_dcf_projections_excel()
            
            # Recreate all components
            self._create_main_dcf_table(self.dcf_main_ax, dcf_data)
            self._create_growth_estimates_box(self.dcf_main_ax, dcf_data)
            self._create_dcf_growth_table_interactive(self.dcf_main_ax, dcf_data)
            self._create_assumption_boxes_interactive(self.dcf_main_ax, dcf_data)
            self._create_company_info_box(self.dcf_main_ax, dcf_data)
            
            # Refresh the plot
            self.fig.canvas.draw()
            
        except Exception as e:
            logger.error(f"Error refreshing DCF tab: {e}")
    
    def _show_fcf_tab(self, event):
        """Callback to show FCF tab"""
        self.current_tab = 'FCF'
        self._create_fcf_tab()
        self.fig.canvas.draw()
    
    def _show_dcf_tab(self, event):
        """Callback to show DCF tab"""
        self.current_tab = 'DCF'
        self._create_dcf_tab()
        self.fig.canvas.draw()
    
    def _calculate_dcf_valuation(self):
        """Calculate DCF valuation using FCF projections"""
        try:
            # Use FCFF for DCF calculation (most common approach)
            fcff_millions = [x / 1000000 for x in self.fcf_results['FCFF']]
            
            # Calculate growth rates from historical data
            if len(fcff_millions) >= 3:
                # Use 3-year average growth rate
                recent_fcf = fcff_millions[-3:]
                growth_rates = []
                for i in range(1, len(recent_fcf)):
                    if recent_fcf[i-1] != 0:
                        growth_rate = (recent_fcf[i] - recent_fcf[i-1]) / abs(recent_fcf[i-1])
                        growth_rates.append(growth_rate)
                
                avg_growth_rate = sum(growth_rates) / len(growth_rates) if growth_rates else 0.05
            else:
                avg_growth_rate = 0.05  # Default 5% growth
            
            # DCF assumptions - use stored values if available, otherwise defaults
            assumptions = {
                'base_fcf': fcff_millions[-1] if fcff_millions else 100,  # Latest FCF in millions
                'base_fcf_per_share': (fcff_millions[-1] * 1000000 / self.shares_outstanding) if fcff_millions else (100 * 1000000 / self.shares_outstanding),  # Latest FCF per share
                'growth_rate_5yr': self.current_dcf_assumptions.get('growth_rate_5yr', max(min(avg_growth_rate, 0.15), -0.05)),
                'terminal_growth_rate': self.current_dcf_assumptions.get('terminal_growth_rate', 0.025),
                'discount_rate': self.current_dcf_assumptions.get('discount_rate', 0.10),
                'projection_years': 5,
                'shares_outstanding': self.current_dcf_assumptions.get('shares_outstanding', self.shares_outstanding)
            }
            
            # Store current assumptions
            self.current_dcf_assumptions = assumptions.copy()
            
            # Project future FCF (both total and per share)
            projected_fcf = []
            projected_fcf_per_share = []
            current_fcf = assumptions['base_fcf']
            current_fcf_per_share = assumptions['base_fcf_per_share']
            
            for year in range(1, assumptions['projection_years'] + 1):
                current_fcf *= (1 + assumptions['growth_rate_5yr'])
                current_fcf_per_share *= (1 + assumptions['growth_rate_5yr'])
                projected_fcf.append(current_fcf)
                projected_fcf_per_share.append(current_fcf_per_share)
            
            # Calculate terminal value (both total and per share)
            terminal_fcf = projected_fcf[-1] * (1 + assumptions['terminal_growth_rate'])
            terminal_fcf_per_share = projected_fcf_per_share[-1] * (1 + assumptions['terminal_growth_rate'])
            terminal_value = terminal_fcf / (assumptions['discount_rate'] - assumptions['terminal_growth_rate'])
            terminal_value_per_share = terminal_fcf_per_share / (assumptions['discount_rate'] - assumptions['terminal_growth_rate'])
            
            # Calculate present values (both total and per share)
            pv_fcf = []
            pv_fcf_per_share = []
            for i, (fcf, fcf_per_share) in enumerate(zip(projected_fcf, projected_fcf_per_share), 1):
                pv = fcf / ((1 + assumptions['discount_rate']) ** i)
                pv_per_share = fcf_per_share / ((1 + assumptions['discount_rate']) ** i)
                pv_fcf.append(pv)
                pv_fcf_per_share.append(pv_per_share)
            
            pv_terminal = terminal_value / ((1 + assumptions['discount_rate']) ** assumptions['projection_years'])
            pv_terminal_per_share = terminal_value_per_share / ((1 + assumptions['discount_rate']) ** assumptions['projection_years'])
            
            # Calculate enterprise value and equity value (both total and per share)
            enterprise_value = sum(pv_fcf) + pv_terminal
            enterprise_value_per_share = sum(pv_fcf_per_share) + pv_terminal_per_share
            
            # Estimate net debt (simplified - use financing cash flow as proxy)
            financing_cf = [x / 1000000 for x in [self.metrics.get(year, {}).get('financing_cash_flow', 0) for year in self.years]]
            avg_financing_cf = sum(financing_cf) / len(financing_cf) if financing_cf else 0
            net_debt = max(0, -avg_financing_cf * 3)  # Rough estimate
            net_debt_per_share = net_debt * 1000000 / assumptions['shares_outstanding']
            
            equity_value = enterprise_value - net_debt
            equity_value_per_share = enterprise_value_per_share - net_debt_per_share
            
            return {
                'assumptions': assumptions,
                'projected_fcf': projected_fcf,
                'projected_fcf_per_share': projected_fcf_per_share,
                'pv_fcf': pv_fcf,
                'pv_fcf_per_share': pv_fcf_per_share,
                'terminal_value': terminal_value,
                'terminal_value_per_share': terminal_value_per_share,
                'pv_terminal': pv_terminal,
                'pv_terminal_per_share': pv_terminal_per_share,
                'enterprise_value': enterprise_value,
                'enterprise_value_per_share': enterprise_value_per_share,
                'net_debt': net_debt,
                'net_debt_per_share': net_debt_per_share,
                'equity_value': equity_value,
                'equity_value_per_share': equity_value_per_share,
                'years': list(range(int(self.years[-1]) + 1, int(self.years[-1]) + assumptions['projection_years'] + 1))
            }
            
        except Exception as e:
            logger.error(f"Error calculating DCF valuation: {e}")
            return None
    
    def _create_dcf_visualization(self, ax, dcf_results):
        """Create DCF waterfall chart and summary"""
        if not dcf_results:
            ax.text(0.5, 0.5, 'DCF calculation failed', ha='center', va='center', fontsize=14)
            ax.axis('off')
            return
            
        # Create waterfall chart with per-share values
        categories = ['PV of FCF\n(Years 1-5)', 'Terminal Value', 'Enterprise Value', 'Less: Net Debt', 'Equity Value']
        values = [
            sum(dcf_results['pv_fcf_per_share']),
            dcf_results['pv_terminal_per_share'],
            dcf_results['enterprise_value_per_share'],
            -dcf_results['net_debt_per_share'],
            dcf_results['equity_value_per_share']
        ]
        
        colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd']
        bars = ax.bar(categories, values, color=colors, alpha=0.8)
        
        # Add value labels on bars
        for bar, value in zip(bars, values):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'${value:.2f}', ha='center', va='bottom', fontsize=10, fontweight='bold')
        
        ax.set_title(f'DCF Valuation Summary (Per Share) - {self.company_name}', fontsize=16, fontweight='bold')
        ax.set_ylabel('Value Per Share (USD)', fontsize=12)
        ax.grid(True, alpha=0.3)
        plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
    
    def _create_dcf_assumptions_table(self, ax, dcf_results):
        """Create DCF assumptions table with interactive text boxes"""
        if not dcf_results:
            ax.axis('off')
            return
            
        ax.clear()
        ax.axis('off')
        ax.set_title('DCF Assumptions (Click to Edit)', fontsize=14, fontweight='bold', pad=20)
        
        assumptions = dcf_results['assumptions']
        
        # Create simple table without interactive text boxes for now
        # Display assumptions in a clean table format
        table_data = []
        table_labels = []
        
        # Editable assumptions
        editable_items = [
            ('Shares Outstanding (M)', f"{assumptions['shares_outstanding']/1000000:.1f}"),
            ('5-Year Growth Rate (%)', f"{assumptions['growth_rate_5yr']*100:.1f}"),
            ('Terminal Growth Rate (%)', f"{assumptions['terminal_growth_rate']*100:.1f}"),
            ('Discount Rate/WACC (%)', f"{assumptions['discount_rate']*100:.1f}")
        ]
        
        # Create table
        for i, (label, value) in enumerate(editable_items):
            table_labels.append(label)
            table_data.append([value])
        
        # Create table using matplotlib table
        table = ax.table(cellText=table_data,
                        rowLabels=table_labels,
                        cellLoc='center',
                        loc='center',
                        colWidths=[0.3])
        
        # Style the table
        table.auto_set_font_size(False)
        table.set_fontsize(12)
        table.scale(1.2, 2)
        
        # Color coding
        for i in range(len(table_data)):
            table[(i, 0)].set_facecolor('#E8F4FD')
            table[(i, -1)].set_facecolor('#F0F0F0')
        
        # Display read-only values below the table
        readonly_y = 0.1
        ax.text(0.5, readonly_y, 'Read-Only Values:', fontsize=12, fontweight='bold', ha='center', va='center', transform=ax.transAxes)
        
        readonly_items = [
            f"Base FCF Per Share: ${assumptions['base_fcf_per_share']:.2f}",
            f"Enterprise Value Per Share: ${dcf_results['enterprise_value_per_share']:.2f}",
            f"Less: Net Debt Per Share: ${dcf_results['net_debt_per_share']:.2f}",
            f"Equity Value Per Share: ${dcf_results['equity_value_per_share']:.2f}"
        ]
        
        for i, item in enumerate(readonly_items):
            y_pos = readonly_y - (i + 1) * 0.06
            ax.text(0.5, y_pos, item, fontsize=10, ha='center', va='center', transform=ax.transAxes,
                   bbox=dict(boxstyle="round,pad=0.3", facecolor='lightblue', alpha=0.7))
    
    def _update_dcf_assumption(self, key, value_str):
        """Update DCF assumption and recalculate"""
        try:
            value = float(value_str)
            
            # Convert percentage inputs back to decimals
            if key in ['growth_rate_5yr', 'terminal_growth_rate', 'discount_rate']:
                value = value / 100.0
            elif key == 'shares_outstanding':
                value = value * 1000000  # Convert millions to actual shares
            
            # Update the assumption
            self.current_dcf_assumptions[key] = value
            
            # Recalculate and update displays
            self._refresh_dcf_displays()
            
        except ValueError:
            logger.warning(f"Invalid input for {key}: {value_str}")
    
    def _refresh_dcf_displays(self):
        """Refresh all DCF displays with updated assumptions"""
        try:
            # Recalculate DCF with new assumptions
            dcf_results = self._calculate_dcf_valuation()
            
            # Clear and recreate all DCF visualizations
            self._create_dcf_visualization(self.dcf_ax1, dcf_results)
            self._create_dcf_assumptions_table(self.dcf_ax2, dcf_results)
            self._create_dcf_sensitivity_analysis(self.dcf_ax3, dcf_results)
            
            # Refresh the plot
            self.fig.canvas.draw()
            
        except Exception as e:
            logger.error(f"Error refreshing DCF displays: {e}")
    
    def _create_dcf_sensitivity_analysis(self, ax, dcf_results):
        """Create DCF sensitivity analysis"""
        if not dcf_results:
            ax.axis('off')
            return
            
        # Create sensitivity analysis for discount rate and growth rate
        base_discount_rate = dcf_results['assumptions']['discount_rate']
        base_growth_rate = dcf_results['assumptions']['growth_rate_5yr']
        base_equity_value = dcf_results['equity_value']
        
        # Range of discount rates and growth rates
        discount_rates = np.linspace(base_discount_rate - 0.02, base_discount_rate + 0.02, 5)
        growth_rates = np.linspace(base_growth_rate - 0.02, base_growth_rate + 0.02, 5)
        
        # Calculate sensitivity matrix (per share values)
        sensitivity_matrix = np.zeros((len(growth_rates), len(discount_rates)))
        
        for i, growth in enumerate(growth_rates):
            for j, discount in enumerate(discount_rates):
                # Recalculate equity value per share with new assumptions
                projected_fcf_per_share = []
                current_fcf_per_share = dcf_results['assumptions']['base_fcf_per_share']
                
                for year in range(1, dcf_results['assumptions']['projection_years'] + 1):
                    current_fcf_per_share *= (1 + growth)
                    projected_fcf_per_share.append(current_fcf_per_share)
                
                terminal_fcf_per_share = projected_fcf_per_share[-1] * (1 + dcf_results['assumptions']['terminal_growth_rate'])
                terminal_value_per_share = terminal_fcf_per_share / (discount - dcf_results['assumptions']['terminal_growth_rate'])
                
                pv_fcf_per_share = [fcf / ((1 + discount) ** (idx + 1)) for idx, fcf in enumerate(projected_fcf_per_share)]
                pv_terminal_per_share = terminal_value_per_share / ((1 + discount) ** dcf_results['assumptions']['projection_years'])
                
                enterprise_value_per_share = sum(pv_fcf_per_share) + pv_terminal_per_share
                equity_value_per_share = enterprise_value_per_share - dcf_results['net_debt_per_share']
                
                sensitivity_matrix[i, j] = equity_value_per_share
        
        # Create heatmap
        im = ax.imshow(sensitivity_matrix, cmap='RdYlGn', aspect='auto')
        
        # Set ticks and labels
        ax.set_xticks(range(len(discount_rates)))
        ax.set_yticks(range(len(growth_rates)))
        ax.set_xticklabels([f'{dr:.1%}' for dr in discount_rates])
        ax.set_yticklabels([f'{gr:.1%}' for gr in growth_rates])
        
        ax.set_xlabel('Discount Rate', fontsize=12)
        ax.set_ylabel('Growth Rate', fontsize=12)
        ax.set_title('Sensitivity Analysis\n(Equity Value Per Share)', fontsize=14, fontweight='bold')
        
        # Add text annotations
        for i in range(len(growth_rates)):
            for j in range(len(discount_rates)):
                text = ax.text(j, i, f'${sensitivity_matrix[i, j]:.2f}',
                             ha="center", va="center", color="black", fontsize=8)
        
        # Add colorbar
        plt.colorbar(im, ax=ax, shrink=0.8)
    
    def print_fcf_summary(self):
        """
        Print a summary of FCF calculations
        """
        print("\n" + "="*60)
        print("FREE CASH FLOW ANALYSIS SUMMARY")
        print("="*60)
        
        print(f"\nYears analyzed: {', '.join(self.years)}")
        
        for fcf_type, values in self.fcf_results.items():
            print(f"\n{fcf_type} (in millions):")
            for year, value in zip(self.years, values):
                print(f"  {year}: ${value/1000000:.2f}M")
        
        # Calculate averages
        print(f"\nAverage FCF (in millions):")
        for fcf_type, values in self.fcf_results.items():
            avg_value = sum(values) / len(values) if values else 0
            print(f"  {fcf_type}: ${avg_value/1000000:.2f}M")

def select_company_folder():
    """
    Allow user to select company folder containing financial statements
    
    Returns:
        str: Selected folder path
    """
    root = Tk()
    root.withdraw()  # Hide the main window
    
    folder_path = filedialog.askdirectory(
        title="Select Company Folder (containing FY and LTM subfolders)",
        initialdir=os.getcwd()
    )
    
    if not folder_path:
        logger.warning("No folder selected")
        return None
        
    # Validate folder structure
    fy_folder = os.path.join(folder_path, 'FY')
    ltm_folder = os.path.join(folder_path, 'LTM')
    
    if not os.path.exists(fy_folder):
        logger.error(f"FY subfolder not found in {folder_path}")
        return None
        
    if not os.path.exists(ltm_folder):
        logger.error(f"LTM subfolder not found in {folder_path}")
        return None
        
    logger.info(f"Selected company folder: {folder_path}")
    return folder_path

def main():
    """
    Main function to run FCF analysis
    """
    try:
        print("Free Cash Flow Analysis Tool")
        print("="*40)
        
        # Select company folder
        company_folder = select_company_folder()
        if not company_folder:
            print("No valid company folder selected. Exiting.")
            return
            
        # Initialize analyzer
        analyzer = FCFAnalyzer(company_folder)
        
        # Try to fetch ticker data with user input if needed
        if not analyzer.fetch_ticker_data():
            # Ask user for ticker symbol
            ticker_input = input(f"\nCould not automatically determine ticker symbol for '{analyzer.company_name}'. Please enter ticker symbol (or press Enter to skip): ").strip()
            if ticker_input:
                analyzer.fetch_ticker_data(ticker_input)
        
        # Load and analyze financial data
        logger.info("Starting FCF analysis...")
        print(f"\nAnalyzing {analyzer.company_name}...")
        
        analyzer.load_financial_statements()
        analyzer.extract_financial_metrics()
        analyzer.calculate_all_fcf_types()
        
        # Display results
        analyzer.print_fcf_summary()
        
        # Create interactive plot
        print("\nGenerating interactive plot...")
        analyzer.plot_fcf_comparison(interactive=True)
        
        logger.info("FCF analysis completed successfully!")
        
    except Exception as e:
        logger.error(f"Error in main execution: {e}")
        print(f"Error: {e}")
        raise

if __name__ == "__main__":
    main()