#!/usr/bin/env python3
"""
Direct FCF Comparison: Excel vs App (Post-Fix)
Extracts actual FCFF and FCFE values from Excel and compares with app calculations
"""

import pandas as pd
import numpy as np
import openpyxl
from financial_calculations import FinancialCalculator

def extract_fcf_from_excel(excel_path, company_name):
    """Extract all FCF types from Excel file"""
    print(f"\n🔍 EXTRACTING FCF DATA FROM {company_name} EXCEL")
    print("=" * 60)
    
    fcf_data = {}
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        print(f"📋 Available sheets: {workbook.sheetnames}")
        
        # Check FCF DATA sheet first
        if 'FCF DATA' in workbook.sheetnames:
            ws = workbook['FCF DATA']
            print(f"📊 FCF DATA sheet: {ws.max_row} rows x {ws.max_column} cols")
            
            # Find years row (usually in first few rows)
            years_row = None
            for row in range(1, 6):
                row_values = []
                for col in range(1, min(15, ws.max_column + 1)):
                    val = ws.cell(row=row, column=col).value
                    if isinstance(val, (int, float)) and 2010 <= val <= 2030:
                        row_values.append(int(val))
                if len(row_values) >= 5:  # At least 5 years
                    years_row = row
                    fcf_data['years'] = row_values
                    print(f"📅 Found years in row {row}: {row_values}")
                    break
            
            if years_row:
                # Look for FCF data rows
                fcf_patterns = {
                    'LFCF': ['lfcf', 'levered free cash', 'simple fcf', 'free cash flow'],
                    'FCFF': ['fcff', 'free cash flow to firm', 'unlevered free cash'],
                    'FCFE': ['fcfe', 'free cash flow to equity']
                }
                
                for row in range(1, min(50, ws.max_row + 1)):
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value and isinstance(cell_value, str):
                        cell_str = cell_value.lower().strip()
                        
                        # Check each FCF type
                        for fcf_type, patterns in fcf_patterns.items():
                            if any(pattern in cell_str for pattern in patterns):
                                # Extract values
                                values = []
                                for col in range(2, min(2 + len(fcf_data['years']), ws.max_column + 1)):
                                    val = ws.cell(row=row, column=col).value
                                    if isinstance(val, (int, float)) and abs(val) > 10:  # Filter small values
                                        values.append(float(val))
                                
                                if values and len(values) >= 3:
                                    fcf_data[fcf_type] = {
                                        'values': values,
                                        'label': cell_value,
                                        'row': row
                                    }
                                    print(f"💰 {fcf_type} ({cell_value}): {values[:5]}..." if len(values) > 5 else f"💰 {fcf_type} ({cell_value}): {values}")
        
        # If no FCF DATA, try other sheets
        if not fcf_data and 'DCF' in workbook.sheetnames:
            print("🔄 Trying DCF sheet...")
            ws = workbook['DCF']
            
            # Look for historical FCF values in DCF sheet
            for row in range(1, min(30, ws.max_row + 1)):
                for col in range(1, min(5, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        cell_str = cell_value.lower()
                        if any(term in cell_str for term in ['fcf', 'free cash', 'historical']):
                            # Try to extract values from this row
                            values = []
                            for val_col in range(col + 1, min(col + 12, ws.max_column + 1)):
                                val = ws.cell(row=row, column=val_col).value
                                if isinstance(val, (int, float)) and abs(val) > 1000:
                                    values.append(float(val))
                            
                            if len(values) >= 3:
                                fcf_data['Historical_FCF'] = {
                                    'values': values,
                                    'label': cell_value,
                                    'row': row
                                }
                                print(f"📈 Historical FCF: {values}")
        
        workbook.close()
        return fcf_data
        
    except Exception as e:
        print(f"❌ Error extracting from Excel: {e}")
        return {}

def calculate_app_fcf(company_folder, company_name):
    """Calculate FCF using app with detailed output"""
    print(f"\n🤖 CALCULATING {company_name} APP FCF")
    print("=" * 60)
    
    try:
        calc = FinancialCalculator(company_folder)
        calc.load_financial_statements()
        
        results = {}
        
        # LFCF
        lfcf_values = calc.calculate_levered_fcf()
        if lfcf_values:
            results['LFCF'] = {
                'values': lfcf_values,
                'count': len(lfcf_values),
                'method': 'Operating CF - CapEx'
            }
            print(f"💰 LFCF: {lfcf_values}")
        
        # FCFF
        fcff_values = calc.calculate_fcf_to_firm()
        if fcff_values:
            results['FCFF'] = {
                'values': fcff_values,
                'count': len(fcff_values),
                'method': 'EBIT(1-Tax) + D&A - ΔWC - CapEx'
            }
            print(f"💼 FCFF: {fcff_values}")
        
        # FCFE
        fcfe_values = calc.calculate_fcf_to_equity()
        if fcfe_values:
            results['FCFE'] = {
                'values': fcfe_values,
                'count': len(fcfe_values),
                'method': 'NI + D&A - ΔWC - CapEx + Net Borrowing'
            }
            print(f"🏦 FCFE: {fcfe_values}")
        
        return results
        
    except Exception as e:
        print(f"❌ Error calculating app FCF: {e}")
        import traceback
        traceback.print_exc()
        return {}

def compare_fcf_values(excel_data, app_data, company_name):
    """Compare Excel vs App FCF values"""
    print(f"\n🔄 COMPARING {company_name}: EXCEL vs APP")
    print("=" * 60)
    
    comparison_results = {}
    
    for fcf_type in ['LFCF', 'FCFF', 'FCFE']:
        print(f"\n📊 {fcf_type} COMPARISON:")
        
        excel_values = excel_data.get(fcf_type, {}).get('values', [])
        app_values = app_data.get(fcf_type, {}).get('values', [])
        
        if not excel_values and not app_values:
            print(f"   ⚠️ No data available for {fcf_type}")
            comparison_results[fcf_type] = {'status': 'no_data'}
            continue
        
        if not excel_values:
            print(f"   ⚠️ No Excel data for {fcf_type}")
            print(f"   🤖 App values: {app_values}")
            comparison_results[fcf_type] = {'status': 'excel_missing', 'app_values': app_values}
            continue
        
        if not app_values:
            print(f"   ⚠️ No App data for {fcf_type}")
            print(f"   📊 Excel values: {excel_values}")
            comparison_results[fcf_type] = {'status': 'app_missing', 'excel_values': excel_values}
            continue
        
        # Both have data - compare
        print(f"   📊 Excel: {excel_values}")
        print(f"   🤖 App:   {app_values}")
        
        # Try different alignments (Excel might be newest-first, app oldest-first)
        app_reversed = list(reversed(app_values))
        
        # Compare direct alignment
        min_len_direct = min(len(excel_values), len(app_values))
        differences_direct = []
        for i in range(min_len_direct):
            diff = abs(excel_values[i] - app_values[i])
            differences_direct.append(diff)
        
        # Compare reversed alignment
        min_len_reversed = min(len(excel_values), len(app_reversed))
        differences_reversed = []
        for i in range(min_len_reversed):
            diff = abs(excel_values[i] - app_reversed[i])
            differences_reversed.append(diff)
        
        # Choose best alignment
        avg_diff_direct = sum(differences_direct) / len(differences_direct) if differences_direct else float('inf')
        avg_diff_reversed = sum(differences_reversed) / len(differences_reversed) if differences_reversed else float('inf')
        
        if avg_diff_reversed < avg_diff_direct:
            print(f"   🔄 Using reversed app values for better alignment")
            differences = differences_reversed
            app_aligned = app_reversed[:min_len_reversed]
            excel_aligned = excel_values[:min_len_reversed]
        else:
            differences = differences_direct
            app_aligned = app_values[:min_len_direct]
            excel_aligned = excel_values[:min_len_direct]
        
        # Calculate match statistics
        max_diff = max(differences) if differences else 0
        avg_diff = sum(differences) / len(differences) if differences else 0
        tolerance = 1000  # Allow 1M difference for rounding
        
        matches = sum(1 for diff in differences if diff < tolerance)
        match_percentage = (matches / len(differences)) * 100 if differences else 0
        
        print(f"   📈 Comparison (aligned {len(differences)} values):")
        print(f"      Max difference: ${max_diff:,.0f}")
        print(f"      Avg difference: ${avg_diff:,.0f}")
        print(f"      Matches within {tolerance}: {matches}/{len(differences)} ({match_percentage:.1f}%)")
        
        if match_percentage > 90:
            status = "✅ EXCELLENT MATCH"
        elif match_percentage > 70:
            status = "⚠️ GOOD MATCH"
        elif match_percentage > 50:
            status = "❌ POOR MATCH"
        else:
            status = "❌ NO MATCH"
        
        print(f"   🎯 Status: {status}")
        
        comparison_results[fcf_type] = {
            'status': 'compared',
            'excel_values': excel_aligned,
            'app_values': app_aligned,
            'differences': differences,
            'max_diff': max_diff,
            'avg_diff': avg_diff,
            'match_percentage': match_percentage,
            'assessment': status
        }
    
    return comparison_results

def main():
    print("🚀 FCF EXCEL vs APP COMPARISON (POST-FIX VALIDATION)")
    print("=" * 80)
    
    companies = {
        'GOOG': '/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/GOOG/FCF_Analysis_Alphabet Inc Class C.xlsx',
        'MSFT': '/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/MSFT/FCF_Analysis_Microsoft_Corporation.xlsx'
    }
    
    all_results = {}
    
    for company, excel_path in companies.items():
        print(f"\n🏢 ANALYZING {company}")
        print("=" * 80)
        
        # Extract Excel data
        excel_data = extract_fcf_from_excel(excel_path, company)
        
        # Calculate app data
        company_folder = f"/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/{company}"
        app_data = calculate_app_fcf(company_folder, company)
        
        # Compare
        comparison = compare_fcf_values(excel_data, app_data, company)
        
        all_results[company] = {
            'excel_data': excel_data,
            'app_data': app_data,
            'comparison': comparison
        }
    
    # Final summary
    print(f"\n📋 FINAL SUMMARY")
    print("=" * 80)
    
    for company, results in all_results.items():
        print(f"\n🏢 {company}:")
        comparison = results['comparison']
        for fcf_type, result in comparison.items():
            status = result.get('status', 'unknown')
            if status == 'compared':
                assessment = result.get('assessment', 'unknown')
                match_pct = result.get('match_percentage', 0)
                print(f"   {fcf_type}: {assessment} ({match_pct:.1f}% match)")
            else:
                print(f"   {fcf_type}: {status}")
    
    return all_results

if __name__ == "__main__":
    results = main()