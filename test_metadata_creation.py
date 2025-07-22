#!/usr/bin/env python3

"""
Test metadata creation functionality
"""

import json
from datetime import datetime
from openpyxl import load_workbook
from test_date_extraction import extract_period_end_dates, parse_date_year

def test_metadata_creation():
    """Test creating the dates metadata file"""
    
    # Test with first available company data
    try:
        # Find the first available company
        companies = []
        for item in os.listdir("."):
            if os.path.isdir(item) and not item.startswith('.') and not item.startswith('_'):
                fy_path = os.path.join(item, 'FY')
                ltm_path = os.path.join(item, 'LTM')
                if os.path.exists(fy_path) and os.path.exists(ltm_path):
                    companies.append(item)
        
        if not companies:
            print("❌ No companies found in dataset")
            return
        
        company = sorted(companies)[0]
        print(f"Testing metadata creation with {company} data...")
        
        # Find income statement files dynamically
        import glob
        fy_files = glob.glob(os.path.join(company, "FY", "*Income Statement.xlsx"))
        ltm_files = glob.glob(os.path.join(company, "LTM", "*Income Statement.xlsx"))
        
        if not fy_files or not ltm_files:
            print(f"❌ Income statement files not found for {company}")
            return
        
        # Load FY and LTM data
        fy_wb = load_workbook(fy_files[0])
        ltm_wb = load_workbook(ltm_files[0])
        
        # Extract dates
        fy_dates = extract_period_end_dates(fy_wb)
        ltm_dates = extract_period_end_dates(ltm_wb)
        
        # Parse years
        fy_years = [parse_date_year(date_str) for date_str in fy_dates]
        fy_years = [year for year in fy_years if year is not None]
        
        ltm_years = [parse_date_year(date_str) for date_str in ltm_dates]
        ltm_years = [year for year in ltm_years if year is not None]
        
        print(f"FY dates: {fy_dates}")
        print(f"LTM dates: {ltm_dates}")
        print(f"FY years: {fy_years}")
        print(f"LTM years: {ltm_years}")
        
        # Create metadata
        metadata = {
            "fy_dates": fy_dates,
            "ltm_dates": ltm_dates,
            "fy_years": fy_years,
            "ltm_years": ltm_years,
            "last_updated": str(datetime.now())
        }
        
        # Save metadata
        with open("dates_metadata.json", "w") as f:
            json.dump(metadata, f, indent=2)
        
        print("Metadata saved successfully!")
        
        # Read back and verify
        with open("dates_metadata.json", "r") as f:
            loaded_metadata = json.load(f)
        
        print(f"Loaded metadata: {loaded_metadata}")
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    test_metadata_creation()