"""
Excel file inspector to understand the actual structure of your files
"""

import pandas as pd
from openpyxl import load_workbook
import sys

def inspect_excel_structure(file_path):
    """Inspect the actual structure of an Excel file"""
    print(f"🔍 Inspecting: {file_path}")
    print("=" * 60)
    
    try:
        # Method 1: Load with openpyxl (raw data)
        print("📊 Raw Excel structure (first 10 rows, first 5 columns):")
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx >= 10:  # Only show first 10 rows
                break
            # Show first 5 columns
            row_preview = []
            for col_idx, cell in enumerate(row):
                if col_idx >= 5:
                    break
                cell_str = str(cell) if cell is not None else "None"
                # Truncate long strings
                if len(cell_str) > 20:
                    cell_str = cell_str[:17] + "..."
                row_preview.append(cell_str)
            
            print(f"  Row {row_idx:2d}: {row_preview}")
        
        # Method 2: Try pandas with different approaches
        print(f"\n📋 Trying pandas read_excel with different parameters:")
        
        # Try different header rows
        for header_row in [0, 1, 2, 3, None]:
            try:
                df = pd.read_excel(file_path, header=header_row)
                print(f"  ✅ header={header_row}: Shape {df.shape}")
                
                # Show first column content
                if not df.empty and len(df.columns) > 0:
                    first_col = df.iloc[:, 0]
                    non_null_values = first_col.dropna().head(10)
                    print(f"     First column sample: {list(non_null_values)}")
                    
            except Exception as e:
                print(f"  ❌ header={header_row}: {str(e)[:50]}...")
        
        # Method 3: Look for potential metric names anywhere in the sheet
        print(f"\n🎯 Searching for financial metrics anywhere in the file:")
        financial_terms = [
            "revenue", "income", "ebit", "assets", "liabilities", 
            "cash", "depreciation", "capex", "expenditure", "operations"
        ]
        
        found_terms = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            for col_idx, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    cell_lower = cell.lower()
                    for term in financial_terms:
                        if term in cell_lower:
                            found_terms.append(f"Row {row_idx}, Col {col_idx}: '{cell}'")
                            break
                if len(found_terms) >= 10:  # Limit output
                    break
            if len(found_terms) >= 10:
                break
        
        if found_terms:
            print("  Found these financial terms:")
            for term in found_terms:
                print(f"    {term}")
        else:
            print("  ❌ No obvious financial terms found")
            
    except Exception as e:
        print(f"❌ Error inspecting file: {e}")

def main():
    if len(sys.argv) != 2:
        print("Usage: python inspect_excel.py <path_to_excel_file>")
        print("\nExample:")
        print("  python inspect_excel.py 'GOOG/FY/Alphabet Inc Class C - Income Statement.xlsx'")
        sys.exit(1)
    
    file_path = sys.argv[1]
    inspect_excel_structure(file_path)

if __name__ == "__main__":
    main()