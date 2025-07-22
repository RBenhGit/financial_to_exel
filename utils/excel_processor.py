"""
Unified Excel Processor

This module consolidates Excel processing logic that was duplicated across
financial_calculations.py, centralized_data_processor.py, and excel_utils.py.
"""

import os
import pandas as pd
from openpyxl import load_workbook
import logging
from typing import Dict, List, Optional, Tuple, Any, Union
import numpy as np

logger = logging.getLogger(__name__)


class UnifiedExcelProcessor:
    """
    Centralized Excel processing to eliminate duplicate Excel operations
    found across multiple modules.
    """
    
    def __init__(self):
        """Initialize the Excel processor"""
        self.default_search_columns = [0, 1, 2]  # Columns to search for metrics
        self.data_start_column = 3  # Default data starting column
        self.max_scan_rows = 100  # Maximum rows to scan
        
        # Cache for loaded workbooks to avoid reloading
        self._workbook_cache = {}
    
    def load_excel_to_dataframe(self, file_path: str, use_cache: bool = True) -> pd.DataFrame:
        """
        Load Excel file to DataFrame with caching and error handling
        
        Args:
            file_path: Path to Excel file
            use_cache: Whether to use workbook cache
            
        Returns:
            DataFrame with Excel data
        """
        try:
            # Check cache first if enabled
            if use_cache and file_path in self._workbook_cache:
                return self._workbook_cache[file_path].copy()
            
            # Load workbook
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Convert to DataFrame
            data = []
            for row in sheet.iter_rows(values_only=True, max_row=self.max_scan_rows):
                data.append(list(row))
            
            df = pd.DataFrame(data)
            
            # Cache if enabled
            if use_cache:
                self._workbook_cache[file_path] = df.copy()
            
            logger.info(f"Loaded Excel file: {file_path} ({len(df)} rows, {len(df.columns)} columns)")
            return df
            
        except Exception as e:
            logger.error(f"Error loading Excel file {file_path}: {e}")
            raise ValueError(f"Failed to load Excel file: {e}")
    
    def find_metric_row(self, df: pd.DataFrame, metric_name: str, 
                       search_columns: Optional[List[int]] = None,
                       case_sensitive: bool = False) -> Optional[int]:
        """
        Find the row index for a specific metric in the DataFrame
        
        This consolidates the metric finding logic from:
        - financial_calculations.py (lines 636-651)
        - centralized_data_processor.py (lines 292-304)
        
        Args:
            df: DataFrame to search
            metric_name: Name of metric to find
            search_columns: Columns to search in
            case_sensitive: Whether to use case-sensitive matching
            
        Returns:
            Row index if found, None otherwise
        """
        if search_columns is None:
            search_columns = self.default_search_columns
        
        target_name = metric_name if case_sensitive else metric_name.lower()
        
        for idx, row in df.iterrows():
            for col_idx in search_columns:
                if col_idx < len(row) and pd.notna(row.iloc[col_idx]):
                    cell_text = str(row.iloc[col_idx]).strip()
                    compare_text = cell_text if case_sensitive else cell_text.lower()
                    
                    if target_name == compare_text:
                        return idx
        
        return None
    
    def extract_numeric_values(self, df: pd.DataFrame, row_idx: int, 
                             start_col: Optional[int] = None,
                             num_values: Optional[int] = None) -> List[Optional[float]]:
        """
        Extract numeric values from a specific row
        
        Args:
            df: DataFrame to extract from
            row_idx: Row index to extract from
            start_col: Starting column index
            num_values: Number of values to extract
            
        Returns:
            List of numeric values (None for non-numeric)
        """
        if start_col is None:
            start_col = self.data_start_column
        
        if row_idx >= len(df):
            return []
        
        row = df.iloc[row_idx]
        values = []
        
        # Determine end column
        if num_values is None:
            end_col = len(row)
        else:
            end_col = min(start_col + num_values, len(row))
        
        for col_idx in range(start_col, end_col):
            if col_idx < len(row):
                cell_value = row.iloc[col_idx]
                numeric_value = self._convert_to_numeric(cell_value)
                values.append(numeric_value)
            else:
                values.append(None)
        
        return values
    
    def extract_financial_metric(self, df: pd.DataFrame, metric_name: str,
                                start_col: Optional[int] = None,
                                num_values: Optional[int] = None) -> List[Optional[float]]:
        """
        Find and extract values for a specific financial metric
        
        This combines metric finding and value extraction in one operation
        
        Args:
            df: DataFrame containing financial data
            metric_name: Name of the metric to find
            start_col: Starting column for data extraction
            num_values: Number of values to extract
            
        Returns:
            List of numeric values for the metric
        """
        row_idx = self.find_metric_row(df, metric_name)
        
        if row_idx is None:
            logger.warning(f"Metric '{metric_name}' not found in DataFrame")
            return []
        
        return self.extract_numeric_values(df, row_idx, start_col, num_values)
    
    def extract_multiple_metrics(self, df: pd.DataFrame, 
                                metric_names: List[str],
                                start_col: Optional[int] = None,
                                num_values: Optional[int] = None) -> Dict[str, List[Optional[float]]]:
        """
        Extract values for multiple financial metrics
        
        Args:
            df: DataFrame containing financial data
            metric_names: List of metric names to extract
            start_col: Starting column for data extraction  
            num_values: Number of values to extract per metric
            
        Returns:
            Dictionary mapping metric names to value lists
        """
        results = {}
        
        for metric_name in metric_names:
            values = self.extract_financial_metric(df, metric_name, start_col, num_values)
            results[metric_name] = values
        
        return results
    
    def find_company_name(self, df: pd.DataFrame, 
                         search_positions: Optional[List[Tuple[int, int]]] = None) -> Optional[str]:
        """
        Find company name in Excel file
        
        Args:
            df: DataFrame to search
            search_positions: List of (row, col) positions to check
            
        Returns:
            Company name if found
        """
        if search_positions is None:
            search_positions = [(0, 0), (1, 0), (2, 0), (0, 1), (1, 1)]  # Common positions
        
        for row_idx, col_idx in search_positions:
            if row_idx < len(df) and col_idx < len(df.columns):
                cell_value = df.iloc[row_idx, col_idx]
                if pd.notna(cell_value) and isinstance(cell_value, str):
                    cleaned_name = str(cell_value).strip()
                    if cleaned_name and len(cleaned_name) > 3:  # Reasonable company name
                        return cleaned_name
        
        return None
    
    def extract_period_dates(self, df: pd.DataFrame) -> List[str]:
        """
        Extract period end dates from Excel file
        
        Args:
            df: DataFrame to search for dates
            
        Returns:
            List of date strings found
        """
        dates = []
        
        # Look for "Period End Date" row or similar
        date_indicators = ["Period End Date", "Date", "Period", "End Date"]
        
        for indicator in date_indicators:
            row_idx = self.find_metric_row(df, indicator)
            if row_idx is not None:
                date_values = self.extract_numeric_values(df, row_idx, self.data_start_column)
                # Convert to strings and filter valid dates
                for value in date_values:
                    if value is not None:
                        date_str = str(value)
                        if self._looks_like_date(date_str):
                            dates.append(date_str)
                break
        
        return dates
    
    def load_company_financial_statements(self, company_folder: str) -> Dict[str, Dict[str, pd.DataFrame]]:
        """
        Load all financial statements for a company
        
        Args:
            company_folder: Path to company folder containing FY and LTM subfolders
            
        Returns:
            Nested dictionary: period -> statement_type -> DataFrame
        """
        statements = {}
        
        for period in ['FY', 'LTM']:
            period_folder = os.path.join(company_folder, period)
            if not os.path.exists(period_folder):
                continue
                
            statements[period] = {}
            
            for filename in os.listdir(period_folder):
                if filename.endswith('.xlsx'):
                    file_path = os.path.join(period_folder, filename)
                    
                    try:
                        df = self.load_excel_to_dataframe(file_path)
                        
                        # Determine statement type from filename
                        statement_type = self._determine_statement_type(filename)
                        statements[period][statement_type] = df
                        
                    except Exception as e:
                        logger.warning(f"Failed to load {file_path}: {e}")
                        continue
        
        return statements
    
    def _convert_to_numeric(self, value: Any) -> Optional[float]:
        """
        Convert a cell value to numeric, handling various formats
        
        Args:
            value: Cell value to convert
            
        Returns:
            Numeric value or None if conversion fails
        """
        if pd.isna(value) or value is None:
            return None
        
        if isinstance(value, (int, float)):
            if np.isfinite(value):
                return float(value)
            else:
                return None
        
        if isinstance(value, str):
            # Clean up string
            cleaned = str(value).strip()
            if not cleaned:
                return None
            
            # Handle percentage strings
            if cleaned.endswith('%'):
                try:
                    return float(cleaned[:-1]) / 100
                except ValueError:
                    pass
            
            # Handle parentheses (negative numbers)
            if cleaned.startswith('(') and cleaned.endswith(')'):
                try:
                    return -float(cleaned[1:-1])
                except ValueError:
                    pass
            
            # Try direct conversion
            try:
                return float(cleaned)
            except ValueError:
                pass
        
        return None
    
    def _looks_like_date(self, date_str: str) -> bool:
        """
        Check if a string looks like a date
        
        Args:
            date_str: String to check
            
        Returns:
            True if string looks like a date
        """
        if not isinstance(date_str, str):
            return False
        
        date_str = date_str.strip()
        
        # Check for common date patterns
        date_patterns = [
            lambda s: len(s.split('/')) == 3,  # MM/DD/YYYY or similar
            lambda s: len(s.split('-')) == 3,  # YYYY-MM-DD or similar
            lambda s: len(s) == 4 and s.isdigit(),  # YYYY
        ]
        
        return any(pattern(date_str) for pattern in date_patterns)
    
    def _determine_statement_type(self, filename: str) -> str:
        """
        Determine financial statement type from filename
        
        Args:
            filename: Name of the Excel file
            
        Returns:
            Statement type string
        """
        filename_lower = filename.lower()
        
        if 'income' in filename_lower:
            return 'income'
        elif 'balance' in filename_lower:
            return 'balance'
        elif 'cash' in filename_lower or 'cashflow' in filename_lower:
            return 'cashflow'
        else:
            return 'unknown'
    
    def clear_cache(self):
        """Clear the workbook cache"""
        self._workbook_cache.clear()
        logger.info("Excel processor cache cleared")
    
    def get_cache_stats(self) -> Dict[str, Any]:
        """Get statistics about the cache"""
        return {
            'cached_files': len(self._workbook_cache),
            'cache_keys': list(self._workbook_cache.keys())
        }