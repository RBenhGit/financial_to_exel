"""
Directory Structure Validation Helper

Provides comprehensive validation and guidance for FY/LTM directory structure requirements.
This module centralizes directory structure validation logic and provides helpful error messages.

Enhanced in Task 180 to include:
- Integration with ValidationRegistry
- Excel file format and content validation
- Automated compliance reporting
- Directory repair and organization tools
"""

import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import logging
import pandas as pd
from datetime import datetime

logger = logging.getLogger(__name__)

# Import validation registry if available
try:
    from core.validation.validation_registry import (
        ValidationRegistry, ValidationRule, RuleType, RuleScope
    )
    VALIDATION_REGISTRY_AVAILABLE = True
except ImportError:
    VALIDATION_REGISTRY_AVAILABLE = False
    logger.warning("ValidationRegistry not available - using standalone validation")


class DirectoryStructureValidator:
    """
    Comprehensive directory structure validator for financial data folders.

    Validates the required FY/LTM directory structure and provides
    detailed guidance for fixing any issues found.
    """

    # Required financial statement files
    REQUIRED_STATEMENT_FILES = [
        "Income Statement.xlsx",
        "Balance Sheet.xlsx",
        "Cash Flow Statement.xlsx"
    ]

    # Alternative acceptable file patterns
    ALTERNATIVE_FILE_PATTERNS = {
        "Income Statement.xlsx": [
            "income_statement.xlsx",
            "Income Statement.xlsx",
            "income.xlsx",
            "IS.xlsx"
        ],
        "Balance Sheet.xlsx": [
            "balance_sheet.xlsx",
            "Balance Sheet.xlsx",
            "balance.xlsx",
            "BS.xlsx"
        ],
        "Cash Flow Statement.xlsx": [
            "cash_flow_statement.xlsx",
            "Cash Flow Statement.xlsx",
            "cash_flow.xlsx",
            "CF.xlsx",
            "cashflow.xlsx"
        ]
    }

    def __init__(self, validation_registry: Optional['ValidationRegistry'] = None):
        """
        Initialize the validator

        Args:
            validation_registry: Optional ValidationRegistry for rule-based validation
        """
        self.validation_history = []
        self.compliance_history = []
        self.validation_registry = validation_registry

        # Register directory structure validation rules if registry available
        if VALIDATION_REGISTRY_AVAILABLE and self.validation_registry:
            self._register_directory_validation_rules()

    def validate_company_directory(
        self,
        company_path: str,
        strict_mode: bool = False,
        suggest_fixes: bool = True
    ) -> Dict[str, Any]:
        """
        Comprehensive validation of company directory structure.

        Args:
            company_path: Path to the company directory
            strict_mode: If True, requires exact file names
            suggest_fixes: If True, includes fix suggestions in response

        Returns:
            Dictionary with detailed validation results and guidance
        """
        company_path = Path(company_path)
        validation_result = {
            'is_valid': False,
            'company_path': str(company_path),
            'exists': False,
            'issues': [],
            'warnings': [],
            'missing_folders': [],
            'missing_files': [],
            'found_files': {},
            'suggestions': [],
            'structure_score': 0.0,  # 0-1 score indicating structure completeness
            'validation_summary': '',
        }

        # Check if company directory exists
        if not company_path.exists():
            validation_result['issues'].append({
                'type': 'missing_directory',
                'message': f"Company directory does not exist: {company_path}",
                'severity': 'error',
                'fix_suggestion': self._get_directory_creation_guide(company_path) if suggest_fixes else None
            })
            validation_result['validation_summary'] = "Company directory not found"
            return validation_result

        validation_result['exists'] = True

        # Check for required period folders
        period_folders = ['FY', 'LTM']
        existing_folders = []

        for period in period_folders:
            period_path = company_path / period
            if period_path.exists():
                existing_folders.append(period)
                validation_result['found_files'][period] = self._scan_folder_files(period_path)
            else:
                validation_result['missing_folders'].append(period)
                validation_result['issues'].append({
                    'type': 'missing_folder',
                    'message': f"Missing required folder: {period}/",
                    'severity': 'error',
                    'fix_suggestion': self._get_folder_creation_guide(period) if suggest_fixes else None
                })

        # Validate files in existing folders
        for period in existing_folders:
            period_path = company_path / period
            missing_statements = self._validate_period_folder(
                period_path, period, strict_mode
            )

            if missing_statements:
                for missing_statement in missing_statements:
                    validation_result['missing_files'].append(f"{period}/{missing_statement}")
                    validation_result['issues'].append({
                        'type': 'missing_file',
                        'message': f"Missing statement file: {period}/{missing_statement}",
                        'severity': 'error' if strict_mode else 'warning',
                        'fix_suggestion': self._get_file_guidance(missing_statement) if suggest_fixes else None
                    })

        # Calculate structure completeness score
        validation_result['structure_score'] = self._calculate_structure_score(
            len(existing_folders), len(period_folders),
            validation_result['found_files']
        )

        # Determine overall validity
        validation_result['is_valid'] = (
            len(validation_result['missing_folders']) == 0 and
            (len(validation_result['missing_files']) == 0 or not strict_mode)
        )

        # Generate helpful suggestions
        if suggest_fixes and not validation_result['is_valid']:
            validation_result['suggestions'] = self._generate_fix_suggestions(validation_result)

        # Create validation summary
        validation_result['validation_summary'] = self._create_validation_summary(validation_result)

        # Store validation for history tracking
        self.validation_history.append(validation_result)

        return validation_result

    def _scan_folder_files(self, folder_path: Path) -> List[str]:
        """Scan folder and return list of Excel files found"""
        try:
            return [f.name for f in folder_path.iterdir()
                   if f.is_file() and f.suffix.lower() in ['.xlsx', '.xls']]
        except Exception as e:
            logger.warning(f"Error scanning folder {folder_path}: {e}")
            return []

    def _validate_period_folder(self, period_path: Path, period: str, strict_mode: bool) -> List[str]:
        """
        Validate files in a specific period folder (FY or LTM).

        Returns list of missing required statement files.
        """
        missing_statements = []
        found_files = self._scan_folder_files(period_path)

        for required_file in self.REQUIRED_STATEMENT_FILES:
            file_found = False

            if strict_mode:
                # Exact match required
                file_found = required_file in found_files
            else:
                # Check for alternative patterns
                patterns = self.ALTERNATIVE_FILE_PATTERNS.get(required_file, [required_file])
                for pattern in patterns:
                    if any(pattern.lower() in found_file.lower() for found_file in found_files):
                        file_found = True
                        break

            if not file_found:
                missing_statements.append(required_file)

        return missing_statements

    def _calculate_structure_score(self, existing_folders: int, total_folders: int, found_files: Dict) -> float:
        """Calculate a 0-1 score representing directory structure completeness"""
        if total_folders == 0:
            return 0.0

        # Folder score (40% of total)
        folder_score = existing_folders / total_folders * 0.4

        # File score (60% of total)
        file_score = 0.0
        total_possible_files = len(self.REQUIRED_STATEMENT_FILES) * total_folders
        total_found_files = sum(len(files) for files in found_files.values())

        if total_possible_files > 0:
            file_score = min(total_found_files / total_possible_files, 1.0) * 0.6

        return min(folder_score + file_score, 1.0)

    def _get_directory_creation_guide(self, company_path: Path) -> str:
        """Generate guide for creating the company directory structure"""
        return f"""
To create the required directory structure:

1. Create the main company folder:
   {company_path}

2. Inside the company folder, create two subfolders:
   • FY/   (for Full Year financial statements)
   • LTM/  (for Last Twelve Months statements)

3. Add the required Excel files to each subfolder:
   • Income Statement.xlsx
   • Balance Sheet.xlsx
   • Cash Flow Statement.xlsx

Complete structure:
{company_path}/
├── FY/
│   ├── Income Statement.xlsx
│   ├── Balance Sheet.xlsx
│   └── Cash Flow Statement.xlsx
└── LTM/
    ├── Income Statement.xlsx
    ├── Balance Sheet.xlsx
    └── Cash Flow Statement.xlsx
        """

    def _get_folder_creation_guide(self, folder_name: str) -> str:
        """Generate guide for creating a specific period folder"""
        purpose = {
            'FY': 'Full Year financial statements (annual data)',
            'LTM': 'Last Twelve Months financial statements (rolling 12-month data)'
        }.get(folder_name, 'Financial statements')

        return f"""
Create the {folder_name}/ folder for {purpose}:

1. Create folder: {folder_name}/
2. Add required Excel files:
   • Income Statement.xlsx
   • Balance Sheet.xlsx
   • Cash Flow Statement.xlsx

Each Excel file should contain:
   • Column headers like 'FY', 'FY-1', 'FY-2', etc.
   • Financial data organized by year/period
   • Proper metric labels in the first column
        """

    def _get_file_guidance(self, filename: str) -> str:
        """Generate guidance for creating a specific Excel file"""
        file_guidance = {
            "Income Statement.xlsx": """
Income Statement should contain:
• Revenue/Sales figures
• Cost of revenues/COGS
• Operating expenses (R&D, SG&A)
• Operating income/EBIT
• Interest expenses
• Net income
• Column headers: FY, FY-1, FY-2, etc.
            """,
            "Balance Sheet.xlsx": """
Balance Sheet should contain:
• Current assets (cash, receivables, inventory)
• Total assets
• Current liabilities
• Long-term debt
• Shareholders' equity
• Column headers: FY, FY-1, FY-2, etc.
            """,
            "Cash Flow Statement.xlsx": """
Cash Flow Statement should contain:
• Cash from operations
• Capital expenditures
• Free cash flow
• Debt issuance/repayment
• Dividend payments
• Column headers: FY, FY-1, FY-2, etc.
            """
        }

        return file_guidance.get(filename, f"Create {filename} with proper financial data and FY column headers.")

    def _generate_fix_suggestions(self, validation_result: Dict) -> List[str]:
        """Generate actionable fix suggestions based on validation results"""
        suggestions = []

        if not validation_result['exists']:
            suggestions.append("Create the company directory at the specified path")
            suggestions.append("Set up the complete FY/LTM folder structure")
            return suggestions

        if validation_result['missing_folders']:
            for folder in validation_result['missing_folders']:
                suggestions.append(f"Create the {folder}/ subfolder")
                suggestions.append(f"Add required Excel files to the {folder}/ folder")

        if validation_result['missing_files']:
            file_types = set()
            for missing_file in validation_result['missing_files']:
                file_type = missing_file.split('/')[-1]
                file_types.add(file_type)

            for file_type in file_types:
                suggestions.append(f"Create or fix the {file_type} file")

        # Add general suggestions based on structure score
        score = validation_result['structure_score']
        if score < 0.3:
            suggestions.append("Consider using the provided directory structure template")
        elif score < 0.7:
            suggestions.append("Review Excel file formats and column headers")

        return suggestions

    def _create_validation_summary(self, validation_result: Dict) -> str:
        """Create a concise validation summary message"""
        if validation_result['is_valid']:
            return "✅ Directory structure is valid and complete"

        issues = len(validation_result['issues'])
        score = validation_result['structure_score']

        if not validation_result['exists']:
            return "❌ Company directory not found"
        elif score < 0.3:
            return f"❌ Major structure issues found ({issues} problems)"
        elif score < 0.7:
            return f"⚠️ Structure partially complete ({issues} issues remaining)"
        else:
            return f"⚠️ Minor structure issues ({issues} issues to resolve)"

    def _register_directory_validation_rules(self):
        """Register comprehensive directory structure validation rules with the registry"""
        if not VALIDATION_REGISTRY_AVAILABLE or not self.validation_registry:
            return

        from core.validation.validation_registry import RuleSet

        # Create directory structure rule set
        dir_rules = RuleSet(
            name="directory_structure",
            description="Directory structure validation rules for financial data",
            version="1.0.0"
        )

        # Rule: FY folder exists
        fy_folder_rule = ValidationRule(
            rule_id="dir_fy_folder_exists",
            name="FY Folder Exists",
            description="Validates that FY/ folder exists in company directory",
            rule_type=RuleType.COMPLETENESS,
            scope=RuleScope.SYSTEM,
            priority="critical",
            parameters={"folder_name": "FY"},
            error_message_template="Missing required FY/ folder in {company_path}",
            remediation_template="Create FY/ folder for Full Year financial statements"
        )
        dir_rules.add_rule(fy_folder_rule)

        # Rule: LTM folder exists
        ltm_folder_rule = ValidationRule(
            rule_id="dir_ltm_folder_exists",
            name="LTM Folder Exists",
            description="Validates that LTM/ folder exists in company directory",
            rule_type=RuleType.COMPLETENESS,
            scope=RuleScope.SYSTEM,
            priority="critical",
            parameters={"folder_name": "LTM"},
            error_message_template="Missing required LTM/ folder in {company_path}",
            remediation_template="Create LTM/ folder for Last Twelve Months statements"
        )
        dir_rules.add_rule(ltm_folder_rule)

        # Rule: Required Excel files exist
        excel_files_rule = ValidationRule(
            rule_id="dir_required_excel_files",
            name="Required Excel Files Exist",
            description="Validates all required Excel statement files are present",
            rule_type=RuleType.COMPLETENESS,
            scope=RuleScope.DATA,
            priority="critical",
            parameters={
                "required_files": self.REQUIRED_STATEMENT_FILES
            },
            error_message_template="Missing required Excel file: {file_name}",
            remediation_template="Add {file_name} to the {period}/ folder"
        )
        dir_rules.add_rule(excel_files_rule)

        # Rule: Excel file format validation
        excel_format_rule = ValidationRule(
            rule_id="dir_excel_file_format",
            name="Excel File Format Validation",
            description="Validates Excel files are readable and contain proper format",
            rule_type=RuleType.FORMAT,
            scope=RuleScope.DATA,
            priority="high",
            parameters={
                "check_readability": True,
                "check_headers": True,
                "check_numeric_data": True
            },
            error_message_template="Excel file format issue in {file_path}: {issue}",
            remediation_template="Ensure Excel file has FY headers and numeric financial data"
        )
        dir_rules.add_rule(excel_format_rule)

        # Rule: Excel numeric data validation
        excel_numeric_rule = ValidationRule(
            rule_id="dir_excel_numeric_data",
            name="Excel Numeric Data Validation",
            description="Validates Excel financial columns contain numeric data",
            rule_type=RuleType.QUALITY,
            scope=RuleScope.DATA,
            priority="medium",
            parameters={
                "min_numeric_ratio": 0.8
            },
            thresholds={
                "min_numeric_percentage": 80.0
            },
            error_message_template="Non-numeric data in {file_path}: {details}",
            remediation_template="Convert text to numbers and remove invalid characters"
        )
        dir_rules.add_rule(excel_numeric_rule)

        # Rule: Minimum historical periods
        historical_periods_rule = ValidationRule(
            rule_id="dir_minimum_periods",
            name="Minimum Historical Periods",
            description="Validates minimum number of historical periods in Excel files",
            rule_type=RuleType.COMPLETENESS,
            scope=RuleScope.DATA,
            priority="medium",
            parameters={
                "min_periods": 3
            },
            thresholds={
                "recommended_periods": 5
            },
            error_message_template="Insufficient periods in {file_path}: {detected} < {required}",
            remediation_template="Add more historical data (recommend 3-5 years minimum)"
        )
        dir_rules.add_rule(historical_periods_rule)

        # Rule: Directory naming conventions
        naming_convention_rule = ValidationRule(
            rule_id="dir_naming_conventions",
            name="Directory Naming Conventions",
            description="Validates directory and file naming follow conventions",
            rule_type=RuleType.FORMAT,
            scope=RuleScope.SYSTEM,
            priority="low",
            parameters={
                "enforce_case_sensitivity": False,
                "allow_alternative_names": True
            },
            error_message_template="Naming convention violation: {path}",
            remediation_template="Use standard names: 'Income Statement.xlsx', 'Balance Sheet.xlsx', 'Cash Flow Statement.xlsx'"
        )
        dir_rules.add_rule(naming_convention_rule)

        # Register the rule set
        self.validation_registry.register_rule_set(dir_rules)
        logger.info(f"Registered directory structure rule set with {len(dir_rules.rules)} rules")

    def validate_excel_file_format(
        self,
        file_path: Path,
        statement_type: str,
        validate_all_sheets: bool = False
    ) -> Dict[str, Any]:
        """
        Enhanced Excel file format and content validation with comprehensive checks.

        Args:
            file_path: Path to Excel file
            statement_type: Type of financial statement
            validate_all_sheets: If True, validates all sheets in workbook

        Returns:
            Validation results with format compliance details
        """
        validation_result = {
            'is_valid': False,
            'file_path': str(file_path),
            'exists': False,
            'readable': False,
            'has_data': False,
            'has_proper_headers': False,
            'has_numeric_data': False,
            'issues': [],
            'warnings': [],
            'column_count': 0,
            'row_count': 0,
            'detected_periods': [],
            'missing_periods': [],
            'sheet_count': 0,
            'sheet_names': [],
            'data_type_validation': {},
            'numeric_columns': [],
            'non_numeric_data_cells': []
        }

        # Check file exists
        if not file_path.exists():
            validation_result['issues'].append({
                'type': 'missing_file',
                'message': f"Excel file does not exist: {file_path}",
                'severity': 'error'
            })
            return validation_result

        validation_result['exists'] = True

        # Try to read the Excel file with enhanced validation
        try:
            # First, get workbook metadata
            import openpyxl
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                validation_result['sheet_names'] = wb.sheetnames
                validation_result['sheet_count'] = len(wb.sheetnames)
                wb.close()
            except Exception as e:
                logger.warning(f"Could not read workbook metadata with openpyxl: {e}")

            # Read the primary sheet
            df = pd.read_excel(file_path, sheet_name=0)
            validation_result['readable'] = True
            validation_result['row_count'] = len(df)
            validation_result['column_count'] = len(df.columns)

            # Check if file has data
            if df.empty:
                validation_result['issues'].append({
                    'type': 'empty_file',
                    'message': f"Excel file is empty: {file_path.name}",
                    'severity': 'error'
                })
                return validation_result

            validation_result['has_data'] = True

            # Validate column headers (looking for FY pattern)
            fy_pattern_columns = [col for col in df.columns if 'FY' in str(col).upper()]
            validation_result['detected_periods'] = fy_pattern_columns
            validation_result['numeric_columns'] = fy_pattern_columns

            if len(fy_pattern_columns) >= 1:
                validation_result['has_proper_headers'] = True
            else:
                validation_result['warnings'].append({
                    'type': 'missing_headers',
                    'message': f"No FY column headers detected in {file_path.name}",
                    'severity': 'warning'
                })

            # Check for minimum number of periods (at least 3 years recommended)
            if len(fy_pattern_columns) < 3:
                validation_result['warnings'].append({
                    'type': 'insufficient_periods',
                    'message': f"Only {len(fy_pattern_columns)} periods found, recommend at least 3",
                    'severity': 'warning'
                })

            # Enhanced: Validate numeric data types in financial columns
            if fy_pattern_columns:
                numeric_validation = self._validate_numeric_data_types(
                    df, fy_pattern_columns, validation_result
                )
                validation_result['has_numeric_data'] = numeric_validation

            # Statement-specific validations
            if "Income Statement" in statement_type:
                required_metrics = ["Revenue", "Net Income", "Operating Income"]
                self._validate_required_metrics(df, required_metrics, validation_result)
            elif "Balance Sheet" in statement_type:
                required_metrics = ["Total Assets", "Total Liabilities", "Shareholders Equity"]
                self._validate_required_metrics(df, required_metrics, validation_result)
            elif "Cash Flow" in statement_type:
                required_metrics = ["Cash from Operations", "Capital Expenditures"]
                self._validate_required_metrics(df, required_metrics, validation_result)

            # Enhanced: Multi-sheet validation if requested
            if validate_all_sheets and validation_result['sheet_count'] > 1:
                sheet_validations = self._validate_all_sheets(file_path, statement_type)
                validation_result['sheet_validations'] = sheet_validations

            # Overall validity determination
            validation_result['is_valid'] = (
                validation_result['exists'] and
                validation_result['readable'] and
                validation_result['has_data'] and
                validation_result['has_proper_headers'] and
                validation_result['has_numeric_data']
            )

        except Exception as e:
            validation_result['issues'].append({
                'type': 'read_error',
                'message': f"Failed to read Excel file {file_path.name}: {str(e)}",
                'severity': 'error'
            })
            logger.error(f"Error reading Excel file {file_path}: {e}")

        return validation_result

    def _validate_numeric_data_types(
        self,
        df: pd.DataFrame,
        numeric_columns: List[str],
        validation_result: Dict
    ) -> bool:
        """
        Validate that financial data columns contain numeric data.

        Args:
            df: DataFrame to validate
            numeric_columns: List of columns that should contain numeric data
            validation_result: Validation result dict to append warnings/issues

        Returns:
            True if numeric data is valid, False otherwise
        """
        has_numeric_data = False
        non_numeric_cells = []

        for col in numeric_columns:
            if col not in df.columns:
                continue

            # Check data types in this column
            col_data = df[col]

            # Count numeric vs non-numeric values (excluding NaN)
            numeric_count = pd.to_numeric(col_data, errors='coerce').notna().sum()
            total_non_null = col_data.notna().sum()

            if total_non_null > 0:
                numeric_ratio = numeric_count / total_non_null

                # At least 80% of non-null values should be numeric
                if numeric_ratio < 0.8:
                    # Find non-numeric cells
                    for idx, val in col_data.items():
                        if pd.notna(val):
                            try:
                                float(val)
                            except (ValueError, TypeError):
                                non_numeric_cells.append({
                                    'column': col,
                                    'row': idx + 2,  # +2 for header and 0-indexing
                                    'value': str(val)
                                })

                    validation_result['warnings'].append({
                        'type': 'non_numeric_data',
                        'message': f"Column '{col}' contains non-numeric data ({(1-numeric_ratio)*100:.1f}% non-numeric)",
                        'severity': 'warning'
                    })
                else:
                    has_numeric_data = True
            else:
                validation_result['warnings'].append({
                    'type': 'empty_column',
                    'message': f"Column '{col}' has no data",
                    'severity': 'warning'
                })

        validation_result['non_numeric_data_cells'] = non_numeric_cells[:10]  # Limit to first 10

        # Store detailed data type validation results
        validation_result['data_type_validation'] = {
            'total_numeric_columns': len(numeric_columns),
            'columns_with_issues': len([w for w in validation_result['warnings']
                                       if w['type'] == 'non_numeric_data']),
            'non_numeric_cell_count': len(non_numeric_cells)
        }

        return has_numeric_data

    def _validate_all_sheets(
        self,
        file_path: Path,
        statement_type: str
    ) -> Dict[str, Any]:
        """
        Validate all sheets in an Excel workbook.

        Args:
            file_path: Path to Excel file
            statement_type: Type of financial statement

        Returns:
            Dictionary with validation results for each sheet
        """
        sheet_validations = {}

        try:
            # Get all sheet names
            xl_file = pd.ExcelFile(file_path)
            sheet_names = xl_file.sheet_names

            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)

                    sheet_validation = {
                        'sheet_name': sheet_name,
                        'row_count': len(df),
                        'column_count': len(df.columns),
                        'has_data': not df.empty,
                        'has_fy_headers': False,
                        'issues': []
                    }

                    if not df.empty:
                        # Check for FY headers
                        fy_cols = [col for col in df.columns if 'FY' in str(col).upper()]
                        sheet_validation['has_fy_headers'] = len(fy_cols) > 0
                        sheet_validation['detected_periods'] = fy_cols
                    else:
                        sheet_validation['issues'].append({
                            'type': 'empty_sheet',
                            'message': f"Sheet '{sheet_name}' is empty"
                        })

                    sheet_validations[sheet_name] = sheet_validation

                except Exception as e:
                    sheet_validations[sheet_name] = {
                        'sheet_name': sheet_name,
                        'error': str(e),
                        'readable': False
                    }
                    logger.warning(f"Could not read sheet '{sheet_name}': {e}")

        except Exception as e:
            logger.error(f"Error reading Excel sheets from {file_path}: {e}")

        return sheet_validations

    def _validate_required_metrics(
        self,
        df: pd.DataFrame,
        required_metrics: List[str],
        validation_result: Dict
    ):
        """Validate that required metrics are present in the dataframe"""
        first_column = df.iloc[:, 0].astype(str).str.strip().str.lower()

        for metric in required_metrics:
            metric_lower = metric.lower()
            if not any(metric_lower in cell.lower() for cell in first_column):
                validation_result['warnings'].append({
                    'type': 'missing_metric',
                    'message': f"Recommended metric '{metric}' not found",
                    'severity': 'warning'
                })

    def _determine_statement_type(self, file_name: str) -> str:
        """Determine statement type from filename"""
        file_lower = file_name.lower()

        if 'income' in file_lower or 'income statement' in file_lower:
            return "Income Statement.xlsx"
        elif 'balance' in file_lower or 'balance sheet' in file_lower:
            return "Balance Sheet.xlsx"
        elif 'cash' in file_lower and 'flow' in file_lower:
            return "Cash Flow Statement.xlsx"
        else:
            return file_name  # Return original if can't determine

    def validate_directory_structure(
        self,
        ticker: str,
        base_path: str = "data/companies"
    ) -> Dict[str, Any]:
        """
        Comprehensive validation method for data/companies/{TICKER}/ structure

        This is the main validation method requested in Task 180.

        Args:
            ticker: Company ticker symbol
            base_path: Base path to companies data directory

        Returns:
            Comprehensive validation report with compliance status
        """
        company_path = Path(base_path) / ticker

        # Run basic directory validation
        dir_validation = self.validate_company_directory(
            str(company_path),
            strict_mode=False,
            suggest_fixes=True
        )

        # Enhanced validation with Excel file format checks
        excel_validations = {}
        if dir_validation['exists']:
            for period in ['FY', 'LTM']:
                period_path = company_path / period
                if period_path.exists():
                    # Get actual files found in the directory
                    found_files = dir_validation.get('found_files', {}).get(period, [])

                    # Validate all found Excel files
                    for file_name in found_files:
                        file_path = period_path / file_name

                        # Determine statement type from filename
                        statement_type = self._determine_statement_type(file_name)

                        if file_path.exists():
                            excel_key = f"{period}/{file_name}"
                            excel_validations[excel_key] = self.validate_excel_file_format(
                                file_path,
                                statement_type
                            )

        # Compile comprehensive report
        compliance_report = {
            'ticker': ticker,
            'company_path': str(company_path),
            'validation_timestamp': datetime.now().isoformat(),
            'directory_validation': dir_validation,
            'excel_validations': excel_validations,
            'overall_compliance': self._calculate_overall_compliance(
                dir_validation,
                excel_validations
            ),
            'actionable_recommendations': self._generate_actionable_recommendations(
                dir_validation,
                excel_validations
            )
        }

        return compliance_report

    def _calculate_overall_compliance(
        self,
        dir_validation: Dict,
        excel_validations: Dict
    ) -> Dict[str, Any]:
        """
        Calculate overall compliance score and status with detailed breakdown and history tracking.

        Args:
            dir_validation: Directory validation results
            excel_validations: Excel file validation results

        Returns:
            Comprehensive compliance report with scores, status, and categorized issues
        """
        dir_score = dir_validation.get('structure_score', 0.0)

        # Calculate Excel file compliance
        total_excel_files = len(excel_validations)
        valid_excel_files = sum(
            1 for v in excel_validations.values() if v.get('is_valid', False)
        )

        excel_score = valid_excel_files / total_excel_files if total_excel_files > 0 else 0.0

        # Overall score (weighted average)
        overall_score = (dir_score * 0.4) + (excel_score * 0.6)

        # Determine compliance status with more granular levels
        if overall_score >= 0.95:
            compliance_status = "FULLY_COMPLIANT"
            status_description = "Excellent - all requirements met"
        elif overall_score >= 0.8:
            compliance_status = "COMPLIANT"
            status_description = "Good - meets core requirements"
        elif overall_score >= 0.6:
            compliance_status = "PARTIALLY_COMPLIANT"
            status_description = "Fair - significant issues need attention"
        elif overall_score >= 0.3:
            compliance_status = "NON_COMPLIANT"
            status_description = "Poor - major structural issues"
        else:
            compliance_status = "CRITICAL"
            status_description = "Critical - requires immediate attention"

        # Categorize issues by severity
        critical_issues = []
        high_issues = []
        medium_issues = []
        low_issues = []

        # Process directory issues
        for issue in dir_validation.get('issues', []):
            severity = issue.get('severity', 'medium')
            if severity == 'error' or issue.get('type') in ['missing_directory', 'missing_folder']:
                critical_issues.append(issue)
            elif severity == 'critical':
                critical_issues.append(issue)
            elif severity == 'high':
                high_issues.append(issue)
            elif severity == 'medium':
                medium_issues.append(issue)
            else:
                low_issues.append(issue)

        # Process Excel file issues
        for file_key, excel_val in excel_validations.items():
            for issue in excel_val.get('issues', []):
                issue_with_file = {**issue, 'file': file_key}
                severity = issue.get('severity', 'medium')
                if severity == 'error' or issue.get('type') in ['missing_file', 'read_error']:
                    critical_issues.append(issue_with_file)
                elif severity == 'critical':
                    critical_issues.append(issue_with_file)
                elif severity == 'high':
                    high_issues.append(issue_with_file)
                elif severity == 'medium':
                    medium_issues.append(issue_with_file)
                else:
                    low_issues.append(issue_with_file)

        # Build comprehensive compliance report
        compliance_report = {
            'overall_score': round(overall_score, 2),
            'directory_score': round(dir_score, 2),
            'excel_score': round(excel_score, 2),
            'status': compliance_status,
            'status_description': status_description,
            'total_issues': len(dir_validation.get('issues', [])) + \
                          sum(len(v.get('issues', [])) for v in excel_validations.values()),
            'total_warnings': len(dir_validation.get('warnings', [])) + \
                            sum(len(v.get('warnings', [])) for v in excel_validations.values()),
            'categorized_issues': {
                'critical': critical_issues,
                'high': high_issues,
                'medium': medium_issues,
                'low': low_issues
            },
            'issue_breakdown': {
                'critical_count': len(critical_issues),
                'high_count': len(high_issues),
                'medium_count': len(medium_issues),
                'low_count': len(low_issues)
            },
            'file_statistics': {
                'total_excel_files': total_excel_files,
                'valid_excel_files': valid_excel_files,
                'invalid_excel_files': total_excel_files - valid_excel_files
            },
            'timestamp': datetime.now().isoformat()
        }

        # Add to compliance history
        self.compliance_history.append(compliance_report)

        return compliance_report

    def _generate_actionable_recommendations(
        self,
        dir_validation: Dict,
        excel_validations: Dict
    ) -> List[Dict[str, Any]]:
        """
        Generate prioritized actionable recommendations with automated fix suggestions.

        Args:
            dir_validation: Directory validation results
            excel_validations: Excel file validation results

        Returns:
            List of recommendations with priority, actions, and automated fix commands
        """
        recommendations = []
        company_path = dir_validation.get('company_path', '')

        # Critical: Missing company directory
        if not dir_validation.get('exists', False):
            recommendations.append({
                'priority': 'CRITICAL',
                'priority_level': 1,
                'category': 'Directory Structure',
                'issue_type': 'missing_directory',
                'action': f"Create company directory: {company_path}",
                'impact': 'Completely blocks all financial data processing',
                'automated_fix': {
                    'available': True,
                    'method': 'create_directory_structure_from_template',
                    'command': f'validator.create_directory_structure_from_template("{company_path}")',
                    'description': 'Automatically creates complete directory structure with templates'
                },
                'manual_steps': [
                    f'Create directory: {company_path}',
                    'Add FY/ and LTM/ subfolders',
                    'Add required Excel statement files'
                ],
                'estimated_time': '5-10 minutes'
            })
            # If directory doesn't exist, other recommendations are less relevant
            return recommendations

        # Critical: Missing period folders
        if dir_validation.get('missing_folders'):
            for folder in dir_validation['missing_folders']:
                recommendations.append({
                    'priority': 'CRITICAL',
                    'priority_level': 1,
                    'category': 'Directory Structure',
                    'issue_type': 'missing_folder',
                    'action': f"Create missing {folder}/ folder",
                    'impact': f'Blocks {folder} period data processing and analysis',
                    'automated_fix': {
                        'available': True,
                        'method': 'repair_directory_structure',
                        'command': f'validator.repair_directory_structure("{company_path}", create_missing=True)',
                        'description': f'Creates {folder}/ folder with template Excel files'
                    },
                    'manual_steps': [
                        f'Create folder: {company_path}/{folder}/',
                        f'Add Income Statement.xlsx to {folder}/',
                        f'Add Balance Sheet.xlsx to {folder}/',
                        f'Add Cash Flow Statement.xlsx to {folder}/'
                    ],
                    'estimated_time': '10-15 minutes'
                })

        # Critical: Missing Excel statement files
        if dir_validation.get('missing_files'):
            for missing_file in dir_validation['missing_files']:
                period = missing_file.split('/')[0] if '/' in missing_file else 'Unknown'
                file_name = missing_file.split('/')[-1]
                recommendations.append({
                    'priority': 'CRITICAL',
                    'priority_level': 1,
                    'category': 'Required Files',
                    'issue_type': 'missing_file',
                    'action': f"Add missing statement file: {missing_file}",
                    'impact': 'Required for complete financial analysis and calculations',
                    'automated_fix': {
                        'available': True,
                        'method': 'repair_directory_structure',
                        'command': f'validator.repair_directory_structure("{company_path}", create_missing=True)',
                        'description': f'Creates template {file_name} with proper structure'
                    },
                    'manual_steps': [
                        f'Obtain {file_name} from financial data source',
                        f'Ensure file has FY column headers (FY, FY-1, FY-2, etc.)',
                        f'Place file in {company_path}/{period}/'
                    ],
                    'estimated_time': '30-60 minutes per file'
                })

        # High: Unreadable or corrupted Excel files
        for file_key, excel_val in excel_validations.items():
            if not excel_val.get('readable', True) and excel_val.get('exists', False):
                recommendations.append({
                    'priority': 'HIGH',
                    'priority_level': 2,
                    'category': 'File Format',
                    'issue_type': 'corrupted_file',
                    'action': f"Fix corrupted Excel file: {file_key}",
                    'impact': 'File cannot be processed - data is inaccessible',
                    'automated_fix': {
                        'available': False,
                        'reason': 'File corruption requires manual intervention'
                    },
                    'manual_steps': [
                        f'Attempt to repair file using Excel\'s built-in repair tool',
                        'Re-export data from original source',
                        'Verify file opens correctly in Excel before placing in directory'
                    ],
                    'estimated_time': '15-30 minutes'
                })

        # High: Excel files with no data
        for file_key, excel_val in excel_validations.items():
            if excel_val.get('readable', False) and not excel_val.get('has_data', True):
                recommendations.append({
                    'priority': 'HIGH',
                    'priority_level': 2,
                    'category': 'Data Completeness',
                    'issue_type': 'empty_file',
                    'action': f"Add data to empty Excel file: {file_key}",
                    'impact': 'No financial data available for analysis',
                    'automated_fix': {
                        'available': False,
                        'reason': 'Data must be obtained from actual financial sources'
                    },
                    'manual_steps': [
                        'Obtain financial statement data from company filings or database',
                        'Populate Excel file with historical data (minimum 3 years recommended)',
                        'Ensure proper FY column headers and metric labels'
                    ],
                    'estimated_time': '1-2 hours per statement'
                })

        # Medium: Missing FY column headers
        for file_key, excel_val in excel_validations.items():
            if excel_val.get('has_data', False) and not excel_val.get('has_proper_headers', True):
                recommendations.append({
                    'priority': 'MEDIUM',
                    'priority_level': 3,
                    'category': 'Data Format',
                    'issue_type': 'missing_headers',
                    'action': f"Add proper FY column headers to {file_key}",
                    'impact': 'May cause data parsing errors and incorrect calculations',
                    'automated_fix': {
                        'available': False,
                        'reason': 'Header correction requires understanding of data periods'
                    },
                    'manual_steps': [
                        'Open Excel file',
                        'Add column headers: FY, FY-1, FY-2, etc. (or FY-9 through FY)',
                        'Ensure headers match the actual fiscal periods of the data',
                        'Save file'
                    ],
                    'estimated_time': '5 minutes'
                })

        # Medium: Non-numeric data in financial columns
        for file_key, excel_val in excel_validations.items():
            if excel_val.get('has_proper_headers', False) and not excel_val.get('has_numeric_data', True):
                non_numeric_cells = excel_val.get('non_numeric_data_cells', [])
                recommendations.append({
                    'priority': 'MEDIUM',
                    'priority_level': 3,
                    'category': 'Data Quality',
                    'issue_type': 'non_numeric_data',
                    'action': f"Fix non-numeric data in {file_key}",
                    'impact': 'Calculations will fail or produce incorrect results',
                    'details': f'Found {len(non_numeric_cells)} cells with non-numeric data',
                    'automated_fix': {
                        'available': False,
                        'reason': 'Data cleaning requires manual review and correction'
                    },
                    'manual_steps': [
                        'Review flagged cells with non-numeric data',
                        'Convert text values to numbers where appropriate',
                        'Remove or replace invalid characters (e.g., $, commas, text)',
                        'Use consistent units (millions, thousands, etc.)'
                    ],
                    'problem_cells': non_numeric_cells[:5],  # Show first 5 problematic cells
                    'estimated_time': '10-20 minutes'
                })

        # Low: Insufficient historical periods
        for file_key, excel_val in excel_validations.items():
            detected_periods = excel_val.get('detected_periods', [])
            if 1 <= len(detected_periods) < 3:
                recommendations.append({
                    'priority': 'LOW',
                    'priority_level': 4,
                    'category': 'Data Completeness',
                    'issue_type': 'insufficient_periods',
                    'action': f"Add more historical periods to {file_key} (currently {len(detected_periods)} periods)",
                    'impact': 'Limited historical data reduces analysis accuracy and trend analysis capability',
                    'automated_fix': {
                        'available': False,
                        'reason': 'Historical data must be obtained from financial sources'
                    },
                    'manual_steps': [
                        'Obtain additional historical financial data (recommend 3-5 years)',
                        'Add columns for previous periods (FY-2, FY-3, etc.)',
                        'Populate historical metric values'
                    ],
                    'estimated_time': '30-60 minutes'
                })

        # Sort recommendations by priority level
        recommendations.sort(key=lambda x: x.get('priority_level', 999))

        return recommendations

    def create_directory_structure_from_template(
        self,
        company_path: str,
        company_name: str = None
    ) -> Dict[str, Any]:
        """
        Create the complete directory structure from template.

        Args:
            company_path: Path where to create the structure
            company_name: Optional company name for folder naming

        Returns:
            Dictionary with creation results
        """
        company_path = Path(company_path)
        creation_result = {
            'success': False,
            'created_folders': [],
            'created_files': [],
            'errors': []
        }

        try:
            # Create main company folder
            company_path.mkdir(parents=True, exist_ok=True)
            creation_result['created_folders'].append(str(company_path))

            # Create period folders
            for period in ['FY', 'LTM']:
                period_path = company_path / period
                period_path.mkdir(exist_ok=True)
                creation_result['created_folders'].append(str(period_path))

                # Create placeholder Excel files with basic structure
                for statement_file in self.REQUIRED_STATEMENT_FILES:
                    file_path = period_path / statement_file
                    if not file_path.exists():
                        self._create_template_excel_file(file_path, statement_file, period, company_name)
                        creation_result['created_files'].append(str(file_path))

            creation_result['success'] = True
            logger.info(f"Successfully created directory structure at {company_path}")

        except Exception as e:
            error_msg = f"Failed to create directory structure: {str(e)}"
            creation_result['errors'].append(error_msg)
            logger.error(error_msg)

        return creation_result

    def _create_template_excel_file(self, file_path: Path, statement_type: str, period: str, company_name: str = None):
        """Create a template Excel file with basic structure"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = statement_type.replace('.xlsx', '')

            # Add headers
            headers = ['Metric'] + [f'FY-{i}' for i in range(9, -1, -1)] + ['FY']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            # Add sample rows based on statement type
            if "Income Statement" in statement_type:
                sample_rows = ["Revenue", "Cost of Revenues", "Gross Profit", "Operating Expenses",
                              "Operating Income", "Interest Expense", "Net Income"]
            elif "Balance Sheet" in statement_type:
                sample_rows = ["Total Current Assets", "Total Assets", "Total Current Liabilities",
                              "Long-term Debt", "Total Shareholders Equity"]
            elif "Cash Flow" in statement_type:
                sample_rows = ["Cash from Operations", "Capital Expenditures", "Free Cash Flow",
                              "Net Borrowings", "Dividends Paid"]
            else:
                sample_rows = ["Sample Metric 1", "Sample Metric 2", "Sample Metric 3"]

            # Add sample rows
            for row, metric in enumerate(sample_rows, 2):
                ws.cell(row=row, column=1, value=metric)

            # Add instruction note
            instruction = f"""
INSTRUCTIONS:
1. Replace sample metrics with actual {statement_type.replace('.xlsx', '')} data
2. Fill in historical data under FY-9 through FY columns
3. Ensure all financial figures are in consistent units (e.g., millions)
4. {period} = {'Full Year data' if period == 'FY' else 'Last Twelve Months data'}
{f'5. Company: {company_name}' if company_name else ''}

This is a template file - please replace with your actual financial data.
            """

            ws.cell(row=len(sample_rows) + 4, column=1, value=instruction)

            # Save the file
            wb.save(file_path)
            logger.debug(f"Created template Excel file: {file_path}")

        except ImportError:
            # If openpyxl not available, create a simple CSV
            try:
                import csv
                csv_path = file_path.with_suffix('.csv')
                with open(csv_path, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    headers = ['Metric'] + [f'FY-{i}' for i in range(9, -1, -1)] + ['FY']
                    writer.writerow(headers)
                    writer.writerow(['Sample Metric', ''] * len(headers[1:]))
                logger.debug(f"Created template CSV file: {csv_path} (Excel not available)")
            except Exception as e:
                logger.warning(f"Could not create template file {file_path}: {e}")

        except Exception as e:
            logger.warning(f"Could not create template Excel file {file_path}: {e}")

    def repair_directory_structure(
        self,
        company_path: str,
        auto_fix: bool = False,
        create_missing: bool = True,
        organize_files: bool = True
    ) -> Dict[str, Any]:
        """
        Automated repair of directory structure issues.

        Args:
            company_path: Path to company directory
            auto_fix: If True, automatically fix detected issues
            create_missing: Create missing folders and template files
            organize_files: Reorganize misplaced files to correct folders

        Returns:
            Dictionary with repair results and actions taken
        """
        company_path = Path(company_path)
        repair_result = {
            'success': False,
            'issues_found': [],
            'actions_taken': [],
            'warnings': [],
            'created_folders': [],
            'created_files': [],
            'moved_files': [],
            'errors': []
        }

        try:
            # First, validate to identify issues
            validation = self.validate_company_directory(str(company_path))

            # Create company directory if it doesn't exist
            if not validation['exists'] and create_missing:
                company_path.mkdir(parents=True, exist_ok=True)
                repair_result['created_folders'].append(str(company_path))
                repair_result['actions_taken'].append(f"Created company directory: {company_path}")

            # Create missing period folders
            if create_missing:
                for period in ['FY', 'LTM']:
                    period_path = company_path / period
                    if not period_path.exists():
                        period_path.mkdir(parents=True, exist_ok=True)
                        repair_result['created_folders'].append(str(period_path))
                        repair_result['actions_taken'].append(f"Created {period}/ folder")

            # Organize misplaced files if requested
            if organize_files and company_path.exists():
                organized = self._organize_misplaced_files(company_path)
                repair_result['moved_files'].extend(organized['moved_files'])
                repair_result['actions_taken'].extend(organized['actions'])

            # Create missing statement files with templates
            if create_missing:
                for period in ['FY', 'LTM']:
                    period_path = company_path / period
                    if period_path.exists():
                        for statement_file in self.REQUIRED_STATEMENT_FILES:
                            file_path = period_path / statement_file

                            # Check if file exists or if alternative exists
                            file_exists = self._check_file_exists_with_alternatives(
                                period_path, statement_file
                            )

                            if not file_exists:
                                self._create_template_excel_file(
                                    file_path, statement_file, period
                                )
                                repair_result['created_files'].append(str(file_path))
                                repair_result['actions_taken'].append(
                                    f"Created template: {period}/{statement_file}"
                                )

            repair_result['success'] = True
            repair_result['issues_found'] = validation.get('issues', [])

        except Exception as e:
            error_msg = f"Error during directory repair: {str(e)}"
            repair_result['errors'].append(error_msg)
            logger.error(error_msg)

        return repair_result

    def _organize_misplaced_files(self, company_path: Path) -> Dict[str, Any]:
        """
        Organize misplaced Excel files into correct period folders.

        Args:
            company_path: Path to company directory

        Returns:
            Dictionary with organization results
        """
        result = {
            'moved_files': [],
            'actions': [],
            'errors': []
        }

        try:
            # Scan root directory for Excel files that should be in FY/LTM folders
            root_excel_files = [f for f in company_path.iterdir()
                               if f.is_file() and f.suffix.lower() in ['.xlsx', '.xls']]

            for file_path in root_excel_files:
                # Determine which statement type this is
                statement_type = self._determine_statement_type(file_path.name)

                # Try to determine if it's FY or LTM based on filename
                target_period = 'FY'  # Default to FY
                if 'LTM' in file_path.name.upper() or 'TTM' in file_path.name.upper():
                    target_period = 'LTM'

                # Move file to appropriate folder
                target_path = company_path / target_period / statement_type

                # Ensure target folder exists
                target_path.parent.mkdir(parents=True, exist_ok=True)

                # Move the file
                if not target_path.exists():
                    file_path.rename(target_path)
                    result['moved_files'].append({
                        'from': str(file_path),
                        'to': str(target_path)
                    })
                    result['actions'].append(
                        f"Moved {file_path.name} to {target_period}/"
                    )

            # Check for files in wrong period folders
            for period in ['FY', 'LTM']:
                period_path = company_path / period
                if not period_path.exists():
                    continue

                for file_path in period_path.iterdir():
                    if not file_path.is_file() or file_path.suffix.lower() not in ['.xlsx', '.xls']:
                        continue

                    # Check if file belongs to other period based on name
                    opposite_period = 'LTM' if period == 'FY' else 'FY'

                    if opposite_period in file_path.name.upper():
                        # This file is in the wrong period folder
                        target_path = company_path / opposite_period / file_path.name
                        target_path.parent.mkdir(parents=True, exist_ok=True)

                        if not target_path.exists():
                            file_path.rename(target_path)
                            result['moved_files'].append({
                                'from': str(file_path),
                                'to': str(target_path)
                            })
                            result['actions'].append(
                                f"Moved {file_path.name} from {period}/ to {opposite_period}/"
                            )

        except Exception as e:
            error_msg = f"Error organizing files: {str(e)}"
            result['errors'].append(error_msg)
            logger.error(error_msg)

        return result

    def _check_file_exists_with_alternatives(
        self,
        folder_path: Path,
        required_file: str
    ) -> bool:
        """
        Check if a file exists using alternative naming patterns.

        Args:
            folder_path: Folder to check
            required_file: Required filename

        Returns:
            True if file or alternative exists
        """
        if not folder_path.exists():
            return False

        # Check exact match first
        if (folder_path / required_file).exists():
            return True

        # Check alternative patterns
        patterns = self.ALTERNATIVE_FILE_PATTERNS.get(required_file, [])
        folder_files = [f.name for f in folder_path.iterdir() if f.is_file()]

        for pattern in patterns:
            for file_name in folder_files:
                if pattern.lower() in file_name.lower():
                    return True

        return False

    def auto_fix_directory_structure(
        self,
        company_path: str,
        dry_run: bool = False
    ) -> Dict[str, Any]:
        """
        Automatically fix all detected directory structure issues.

        Args:
            company_path: Path to company directory
            dry_run: If True, only report what would be done without making changes

        Returns:
            Dictionary with planned or executed fixes
        """
        result = {
            'dry_run': dry_run,
            'planned_actions': [],
            'executed_actions': [],
            'errors': []
        }

        # Validate current structure
        validation = self.validate_company_directory(str(company_path))

        # Plan fixes based on issues
        for issue in validation.get('issues', []):
            issue_type = issue.get('type')

            if issue_type == 'missing_directory':
                action = f"Create company directory at {company_path}"
                result['planned_actions'].append(action)

            elif issue_type == 'missing_folder':
                folder_name = issue.get('message', '').split(':')[-1].strip()
                action = f"Create missing folder: {folder_name}"
                result['planned_actions'].append(action)

            elif issue_type == 'missing_file':
                file_name = issue.get('message', '').split(':')[-1].strip()
                action = f"Create template file: {file_name}"
                result['planned_actions'].append(action)

        # Execute fixes if not dry run
        if not dry_run and result['planned_actions']:
            repair_result = self.repair_directory_structure(
                company_path,
                auto_fix=True,
                create_missing=True,
                organize_files=True
            )
            result['executed_actions'] = repair_result.get('actions_taken', [])
            result['errors'] = repair_result.get('errors', [])

        return result

    def export_compliance_report(
        self,
        compliance_report: Dict[str, Any],
        output_path: str,
        format: str = 'json'
    ) -> Dict[str, Any]:
        """
        Export compliance report to various formats.

        Args:
            compliance_report: Compliance report dictionary
            output_path: Path where to save the export
            format: Export format - 'json', 'csv', or 'html'

        Returns:
            Dictionary with export results
        """
        export_result = {
            'success': False,
            'format': format,
            'output_path': output_path,
            'error': None
        }

        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            if format.lower() == 'json':
                self._export_report_json(compliance_report, output_path)
            elif format.lower() == 'csv':
                self._export_report_csv(compliance_report, output_path)
            elif format.lower() == 'html':
                self._export_report_html(compliance_report, output_path)
            else:
                raise ValueError(f"Unsupported export format: {format}")

            export_result['success'] = True
            logger.info(f"Successfully exported compliance report to {output_path}")

        except Exception as e:
            error_msg = f"Failed to export report: {str(e)}"
            export_result['error'] = error_msg
            logger.error(error_msg)

        return export_result

    def _export_report_json(self, report: Dict[str, Any], output_path: Path):
        """Export report as JSON"""
        import json

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False)

    def _export_report_csv(self, report: Dict[str, Any], output_path: Path):
        """Export report as CSV"""
        import csv

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Write summary section
            writer.writerow(['Compliance Report Summary'])
            writer.writerow(['Ticker', report.get('ticker', 'N/A')])
            writer.writerow(['Timestamp', report.get('validation_timestamp', 'N/A')])
            writer.writerow(['Company Path', report.get('company_path', 'N/A')])
            writer.writerow([])

            # Write compliance scores
            overall_compliance = report.get('overall_compliance', {})
            writer.writerow(['Compliance Scores'])
            writer.writerow(['Overall Score', overall_compliance.get('overall_score', 'N/A')])
            writer.writerow(['Directory Score', overall_compliance.get('directory_score', 'N/A')])
            writer.writerow(['Excel Score', overall_compliance.get('excel_score', 'N/A')])
            writer.writerow(['Status', overall_compliance.get('status', 'N/A')])
            writer.writerow(['Status Description', overall_compliance.get('status_description', 'N/A')])
            writer.writerow([])

            # Write issue breakdown
            issue_breakdown = overall_compliance.get('issue_breakdown', {})
            writer.writerow(['Issue Breakdown'])
            writer.writerow(['Critical Issues', issue_breakdown.get('critical_count', 0)])
            writer.writerow(['High Priority Issues', issue_breakdown.get('high_count', 0)])
            writer.writerow(['Medium Priority Issues', issue_breakdown.get('medium_count', 0)])
            writer.writerow(['Low Priority Issues', issue_breakdown.get('low_count', 0)])
            writer.writerow([])

            # Write recommendations
            recommendations = report.get('actionable_recommendations', [])
            if recommendations:
                writer.writerow(['Actionable Recommendations'])
                writer.writerow(['Priority', 'Category', 'Action', 'Impact', 'Estimated Time'])
                for rec in recommendations:
                    writer.writerow([
                        rec.get('priority', ''),
                        rec.get('category', ''),
                        rec.get('action', ''),
                        rec.get('impact', ''),
                        rec.get('estimated_time', '')
                    ])

    def _export_report_html(self, report: Dict[str, Any], output_path: Path):
        """Export report as HTML"""
        overall_compliance = report.get('overall_compliance', {})
        recommendations = report.get('actionable_recommendations', [])

        # Generate HTML
        html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Compliance Report - {report.get('ticker', 'Unknown')}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            border-radius: 10px;
            margin-bottom: 20px;
        }}
        .header h1 {{
            margin: 0 0 10px 0;
        }}
        .score-section {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 20px;
        }}
        .score-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .score-value {{
            font-size: 2em;
            font-weight: bold;
            color: #667eea;
        }}
        .score-label {{
            color: #666;
            margin-top: 5px;
        }}
        .status {{
            display: inline-block;
            padding: 5px 15px;
            border-radius: 20px;
            font-weight: bold;
        }}
        .status-FULLY_COMPLIANT {{ background-color: #4caf50; color: white; }}
        .status-COMPLIANT {{ background-color: #8bc34a; color: white; }}
        .status-PARTIALLY_COMPLIANT {{ background-color: #ff9800; color: white; }}
        .status-NON_COMPLIANT {{ background-color: #f44336; color: white; }}
        .status-CRITICAL {{ background-color: #d32f2f; color: white; }}
        .recommendations {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin-top: 20px;
        }}
        .recommendation {{
            border-left: 4px solid #ccc;
            padding: 15px;
            margin-bottom: 15px;
            background-color: #fafafa;
        }}
        .recommendation.CRITICAL {{ border-left-color: #d32f2f; }}
        .recommendation.HIGH {{ border-left-color: #ff9800; }}
        .recommendation.MEDIUM {{ border-left-color: #ffc107; }}
        .recommendation.LOW {{ border-left-color: #4caf50; }}
        .recommendation-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
        }}
        .priority-badge {{
            padding: 3px 10px;
            border-radius: 3px;
            font-size: 0.85em;
            font-weight: bold;
        }}
        .priority-CRITICAL {{ background-color: #d32f2f; color: white; }}
        .priority-HIGH {{ background-color: #ff9800; color: white; }}
        .priority-MEDIUM {{ background-color: #ffc107; color: black; }}
        .priority-LOW {{ background-color: #4caf50; color: white; }}
        .automated-fix {{
            background-color: #e3f2fd;
            padding: 10px;
            border-radius: 5px;
            margin-top: 10px;
        }}
        .manual-steps {{
            margin-top: 10px;
        }}
        .manual-steps li {{
            margin: 5px 0;
        }}
        .timestamp {{
            color: #666;
            font-size: 0.9em;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📊 Directory Structure Compliance Report</h1>
        <p><strong>Ticker:</strong> {report.get('ticker', 'N/A')}</p>
        <p><strong>Path:</strong> {report.get('company_path', 'N/A')}</p>
        <p class="timestamp">Generated: {report.get('validation_timestamp', 'N/A')}</p>
    </div>

    <div class="score-section">
        <div class="score-card">
            <div class="score-value">{overall_compliance.get('overall_score', 'N/A')}</div>
            <div class="score-label">Overall Score</div>
        </div>
        <div class="score-card">
            <div class="score-value">{overall_compliance.get('directory_score', 'N/A')}</div>
            <div class="score-label">Directory Score</div>
        </div>
        <div class="score-card">
            <div class="score-value">{overall_compliance.get('excel_score', 'N/A')}</div>
            <div class="score-label">Excel Score</div>
        </div>
    </div>

    <div class="score-card">
        <h3>Compliance Status</h3>
        <span class="status status-{overall_compliance.get('status', '')}">{overall_compliance.get('status', 'UNKNOWN')}</span>
        <p>{overall_compliance.get('status_description', '')}</p>
        <p><strong>Issues:</strong> {overall_compliance.get('total_issues', 0)} Critical/High,
           {overall_compliance.get('total_warnings', 0)} Medium/Low</p>
    </div>

    <div class="recommendations">
        <h2>📋 Actionable Recommendations</h2>
        <p>Total recommendations: {len(recommendations)}</p>
"""

        # Add each recommendation
        for i, rec in enumerate(recommendations, 1):
            priority = rec.get('priority', 'MEDIUM')
            category = rec.get('category', 'General')
            action = rec.get('action', 'No action specified')
            impact = rec.get('impact', '')
            estimated_time = rec.get('estimated_time', 'Unknown')
            automated_fix = rec.get('automated_fix', {})
            manual_steps = rec.get('manual_steps', [])

            html_content += f"""
        <div class="recommendation {priority}">
            <div class="recommendation-header">
                <span><strong>#{i}:</strong> {category}</span>
                <span class="priority-badge priority-{priority}">{priority}</span>
            </div>
            <p><strong>Action:</strong> {action}</p>
            <p><strong>Impact:</strong> {impact}</p>
            <p><strong>Estimated Time:</strong> {estimated_time}</p>
"""

            if automated_fix.get('available'):
                html_content += f"""
            <div class="automated-fix">
                <strong>✅ Automated Fix Available</strong>
                <p>{automated_fix.get('description', '')}</p>
                <code>{automated_fix.get('command', '')}</code>
            </div>
"""
            elif automated_fix.get('reason'):
                html_content += f"""
            <div class="automated-fix">
                <strong>ℹ️ Manual Intervention Required</strong>
                <p>{automated_fix.get('reason', '')}</p>
            </div>
"""

            if manual_steps:
                html_content += """
            <div class="manual-steps">
                <strong>Manual Steps:</strong>
                <ol>
"""
                for step in manual_steps:
                    html_content += f"                    <li>{step}</li>\n"
                html_content += """
                </ol>
            </div>
"""

            html_content += "        </div>\n"

        html_content += """
    </div>
</body>
</html>
"""

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

    def get_compliance_history(
        self,
        limit: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """
        Get compliance history with optional limit.

        Args:
            limit: Maximum number of history entries to return (most recent first)

        Returns:
            List of compliance reports from history
        """
        if limit:
            return self.compliance_history[-limit:]
        return self.compliance_history

    def get_compliance_trend(self) -> Dict[str, Any]:
        """
        Analyze compliance trend from history.

        Returns:
            Dictionary with trend analysis
        """
        if len(self.compliance_history) < 2:
            return {
                'available': False,
                'message': 'Insufficient history for trend analysis (minimum 2 reports required)'
            }

        # Get first and last reports
        first_report = self.compliance_history[0]
        last_report = self.compliance_history[-1]

        first_score = first_report.get('overall_score', 0)
        last_score = last_report.get('overall_score', 0)

        score_change = last_score - first_score
        trend = 'improving' if score_change > 0 else 'declining' if score_change < 0 else 'stable'

        return {
            'available': True,
            'total_reports': len(self.compliance_history),
            'first_score': first_score,
            'last_score': last_score,
            'score_change': round(score_change, 2),
            'trend': trend,
            'first_timestamp': first_report.get('timestamp', 'Unknown'),
            'last_timestamp': last_report.get('timestamp', 'Unknown')
        }

    def validate_with_registry_rules(
        self,
        company_path: str,
        scope: Optional['RuleScope'] = None
    ) -> Dict[str, Any]:
        """
        Execute rule-based validation using ValidationRegistry rules.

        Args:
            company_path: Path to company directory
            scope: Optional scope filter for rules (SYSTEM, DATA, etc.)

        Returns:
            Comprehensive validation results with rule execution details
        """
        if not VALIDATION_REGISTRY_AVAILABLE or not self.validation_registry:
            logger.warning("ValidationRegistry not available, falling back to standard validation")
            return self.validate_company_directory(company_path)

        from core.validation.validation_registry import RuleScope

        # Get applicable rules
        if scope:
            rules = self.validation_registry.get_rules_by_scope(scope, enabled_only=True)
        else:
            # Get all directory structure rules
            dir_rule_set = self.validation_registry.get_rule_set("directory_structure")
            rules = dir_rule_set.get_enabled_rules() if dir_rule_set else []

        validation_result = {
            'company_path': company_path,
            'rules_executed': 0,
            'rules_passed': 0,
            'rules_failed': 0,
            'rule_results': [],
            'overall_status': 'PASS',
            'timestamp': datetime.now().isoformat()
        }

        company_path_obj = Path(company_path)

        # Execute each rule
        for rule in rules:
            rule_result = self._execute_validation_rule(rule, company_path_obj)
            validation_result['rule_results'].append(rule_result)
            validation_result['rules_executed'] += 1

            if rule_result['status'] == 'FAIL':
                validation_result['rules_failed'] += 1
                if rule.priority in ['critical', 'high']:
                    validation_result['overall_status'] = 'FAIL'
            else:
                validation_result['rules_passed'] += 1

        return validation_result

    def _execute_validation_rule(
        self,
        rule: 'ValidationRule',
        company_path: Path
    ) -> Dict[str, Any]:
        """
        Execute a single validation rule.

        Args:
            rule: ValidationRule to execute
            company_path: Path to company directory

        Returns:
            Rule execution result dictionary
        """
        rule_result = {
            'rule_id': rule.rule_id,
            'rule_name': rule.name,
            'priority': rule.priority,
            'status': 'PASS',
            'issues': [],
            'details': {}
        }

        try:
            # Route to appropriate validation method based on rule_id
            if rule.rule_id == "dir_fy_folder_exists":
                exists = (company_path / "FY").exists()
                if not exists:
                    rule_result['status'] = 'FAIL'
                    rule_result['issues'].append({
                        'message': rule.error_message_template.format(company_path=company_path),
                        'remediation': rule.remediation_template
                    })

            elif rule.rule_id == "dir_ltm_folder_exists":
                exists = (company_path / "LTM").exists()
                if not exists:
                    rule_result['status'] = 'FAIL'
                    rule_result['issues'].append({
                        'message': rule.error_message_template.format(company_path=company_path),
                        'remediation': rule.remediation_template
                    })

            elif rule.rule_id == "dir_required_excel_files":
                required_files = rule.parameters.get('required_files', self.REQUIRED_STATEMENT_FILES)
                for period in ['FY', 'LTM']:
                    period_path = company_path / period
                    if period_path.exists():
                        for file_name in required_files:
                            if not (period_path / file_name).exists():
                                # Check for alternative names
                                has_alternative = self._check_file_exists_with_alternatives(
                                    period_path, file_name
                                )
                                if not has_alternative:
                                    rule_result['status'] = 'FAIL'
                                    rule_result['issues'].append({
                                        'message': rule.error_message_template.format(file_name=file_name),
                                        'remediation': rule.remediation_template.format(
                                            file_name=file_name,
                                            period=period
                                        )
                                    })

            elif rule.rule_id == "dir_excel_file_format":
                # Validate Excel file formats
                for period in ['FY', 'LTM']:
                    period_path = company_path / period
                    if period_path.exists():
                        excel_files = [f for f in period_path.iterdir()
                                     if f.suffix.lower() in ['.xlsx', '.xls']]
                        for excel_file in excel_files:
                            format_result = self.validate_excel_file_format(
                                excel_file,
                                self._determine_statement_type(excel_file.name)
                            )
                            if not format_result.get('is_valid', False):
                                rule_result['status'] = 'FAIL'
                                for issue in format_result.get('issues', []):
                                    rule_result['issues'].append({
                                        'message': rule.error_message_template.format(
                                            file_path=excel_file,
                                            issue=issue.get('message', '')
                                        ),
                                        'remediation': rule.remediation_template
                                    })

            elif rule.rule_id == "dir_excel_numeric_data":
                # Check numeric data quality
                min_ratio = rule.parameters.get('min_numeric_ratio', 0.8)
                for period in ['FY', 'LTM']:
                    period_path = company_path / period
                    if period_path.exists():
                        excel_files = [f for f in period_path.iterdir()
                                     if f.suffix.lower() in ['.xlsx', '.xls']]
                        for excel_file in excel_files:
                            format_result = self.validate_excel_file_format(
                                excel_file,
                                self._determine_statement_type(excel_file.name)
                            )
                            if not format_result.get('has_numeric_data', True):
                                non_numeric_cells = format_result.get('non_numeric_data_cells', [])
                                rule_result['status'] = 'FAIL'
                                rule_result['issues'].append({
                                    'message': rule.error_message_template.format(
                                        file_path=excel_file,
                                        details=f"{len(non_numeric_cells)} non-numeric cells"
                                    ),
                                    'remediation': rule.remediation_template
                                })

            elif rule.rule_id == "dir_minimum_periods":
                # Check minimum historical periods
                min_periods = rule.parameters.get('min_periods', 3)
                for period in ['FY', 'LTM']:
                    period_path = company_path / period
                    if period_path.exists():
                        excel_files = [f for f in period_path.iterdir()
                                     if f.suffix.lower() in ['.xlsx', '.xls']]
                        for excel_file in excel_files:
                            format_result = self.validate_excel_file_format(
                                excel_file,
                                self._determine_statement_type(excel_file.name)
                            )
                            detected_periods = len(format_result.get('detected_periods', []))
                            if 0 < detected_periods < min_periods:
                                rule_result['status'] = 'FAIL'
                                rule_result['issues'].append({
                                    'message': rule.error_message_template.format(
                                        file_path=excel_file,
                                        detected=detected_periods,
                                        required=min_periods
                                    ),
                                    'remediation': rule.remediation_template
                                })

        except Exception as e:
            rule_result['status'] = 'ERROR'
            rule_result['issues'].append({
                'message': f"Rule execution error: {str(e)}",
                'remediation': 'Review rule configuration and validation logic'
            })
            logger.error(f"Error executing rule {rule.rule_id}: {e}")

        return rule_result

    def load_rules_from_config(self, config_file: str) -> bool:
        """
        Load validation rules from configuration file.

        Args:
            config_file: Path to YAML or JSON configuration file

        Returns:
            True if rules loaded successfully
        """
        if not VALIDATION_REGISTRY_AVAILABLE or not self.validation_registry:
            logger.error("ValidationRegistry not available")
            return False

        try:
            self.validation_registry.load_config(config_file)
            logger.info(f"Loaded validation rules from {config_file}")
            return True
        except Exception as e:
            logger.error(f"Failed to load rules from {config_file}: {e}")
            return False

    def update_rule_config(
        self,
        rule_id: str,
        parameters: Optional[Dict[str, Any]] = None,
        thresholds: Optional[Dict[str, float]] = None,
        enabled: Optional[bool] = None
    ) -> bool:
        """
        Update configuration for a specific validation rule.

        Args:
            rule_id: ID of the rule to update
            parameters: New parameters to set
            thresholds: New thresholds to set
            enabled: Enable/disable the rule

        Returns:
            True if update successful
        """
        if not VALIDATION_REGISTRY_AVAILABLE or not self.validation_registry:
            logger.error("ValidationRegistry not available")
            return False

        try:
            success = True

            if parameters:
                success = success and self.validation_registry.update_rule_parameters(
                    rule_id, parameters
                )

            if thresholds:
                success = success and self.validation_registry.update_rule_thresholds(
                    rule_id, thresholds
                )

            if enabled is not None:
                if enabled:
                    success = success and self.validation_registry.enable_rule(rule_id)
                else:
                    success = success and self.validation_registry.disable_rule(rule_id)

            return success

        except Exception as e:
            logger.error(f"Failed to update rule {rule_id}: {e}")
            return False

    def get_registry_validation_report(self) -> Dict[str, Any]:
        """
        Get comprehensive validation report from registry.

        Returns:
            Dictionary with registry statistics and rule information
        """
        if not VALIDATION_REGISTRY_AVAILABLE or not self.validation_registry:
            return {
                'available': False,
                'message': 'ValidationRegistry not available'
            }

        try:
            stats = self.validation_registry.get_registry_stats()
            dir_rules = self.validation_registry.get_rule_set("directory_structure")

            report = {
                'available': True,
                'registry_stats': stats,
                'directory_rules': {
                    'total': len(dir_rules.rules) if dir_rules else 0,
                    'enabled': len(dir_rules.get_enabled_rules()) if dir_rules else 0,
                    'rules': [
                        {
                            'id': rule.rule_id,
                            'name': rule.name,
                            'priority': rule.priority,
                            'enabled': rule.enabled,
                            'type': rule.rule_type.value,
                            'scope': rule.scope.value
                        }
                        for rule in (dir_rules.rules if dir_rules else [])
                    ]
                }
            }

            return report

        except Exception as e:
            logger.error(f"Failed to generate registry report: {e}")
            return {
                'available': False,
                'error': str(e)
            }

    def get_validation_help_text(self) -> str:
        """Get comprehensive help text for directory structure requirements"""
        return """
📁 Financial Data Directory Structure Requirements

REQUIRED STRUCTURE:
CompanyName/
├── FY/                     (Full Year Financial Statements)
│   ├── Income Statement.xlsx
│   ├── Balance Sheet.xlsx
│   └── Cash Flow Statement.xlsx
└── LTM/                    (Last Twelve Months Statements)
    ├── Income Statement.xlsx
    ├── Balance Sheet.xlsx
    └── Cash Flow Statement.xlsx

📋 EXCEL FILE REQUIREMENTS:
• Column headers: 'FY', 'FY-1', 'FY-2', etc. (historical periods)
• First column: Metric names (e.g., "Revenue", "Net Income")
• Data rows: Financial values for each period
• Format: .xlsx (Excel 2007+)

📊 STATEMENT CONTENT GUIDELINES:

Income Statement:
• Revenue, Cost of Revenues, Gross Profit
• Operating Expenses (R&D, SG&A)
• Operating Income, Interest Expense, Net Income

Balance Sheet:
• Current Assets, Total Assets
• Current Liabilities, Long-term Debt
• Shareholders' Equity

Cash Flow Statement:
• Cash from Operations, Capital Expenditures
• Free Cash Flow, Debt Changes, Dividends

💡 TIPS:
• Ensure consistent units (e.g., all values in millions)
• Include at least 3-5 years of historical data
• Use standard accounting terminology
• Keep metric names consistent across periods
        """


# Convenience functions for common operations

def validate_company_directory(company_path: str, **kwargs) -> Dict[str, Any]:
    """
    Convenience function to validate a company directory structure.

    Args:
        company_path: Path to company directory
        **kwargs: Additional arguments for DirectoryStructureValidator

    Returns:
        Validation results dictionary
    """
    validator = DirectoryStructureValidator()
    return validator.validate_company_directory(company_path, **kwargs)


def create_directory_structure_template(company_path: str, company_name: str = None) -> Dict[str, Any]:
    """
    Convenience function to create directory structure from template.

    Args:
        company_path: Path where to create structure
        company_name: Optional company name

    Returns:
        Creation results dictionary
    """
    validator = DirectoryStructureValidator()
    return validator.create_directory_structure_from_template(company_path, company_name)


def get_directory_structure_help() -> str:
    """
    Get comprehensive help text for directory structure requirements.

    Returns:
        Help text string
    """
    validator = DirectoryStructureValidator()
    return validator.get_validation_help_text()