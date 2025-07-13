#!/usr/bin/env python3
"""
GOOG FCF Validation Script
Specifically validates GOOG Excel data vs App calculations
"""

import pandas as pd
import numpy as np
import os
from financial_calculations import FinancialCalculator

def analyze_goog_excel():
    """Analyze GOOG Excel file in detail"""
    excel_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/GOOG/FCF_Analysis_Alphabet Inc Class C.xlsx"
    
    print("🔍 ANALYZING GOOG EXCEL FILE")
    print("=" * 50)
    
    # Read all sheets to understand structure
    excel_file = pd.ExcelFile(excel_path)
    print(f"📊 Available sheets: {excel_file.sheet_names}")
    
    # Try to read FCF DATA sheet
    try:
        fcf_data = pd.read_excel(excel_path, sheet_name='FCF DATA')
        print(f"\n📈 FCF DATA Sheet Shape: {fcf_data.shape}")
        print("📋 First 10 rows of FCF DATA:")
        print(fcf_data.head(10))
        
        # Look for actual FCF values
        print("\n🔍 Searching for FCF patterns in FCF DATA...")
        for idx, row in fcf_data.iterrows():
            if idx > 20:  # Limit search
                break
            row_str = str(row.iloc[0]).lower() if pd.notna(row.iloc[0]) else ""
            if any(term in row_str for term in ['fcf', 'free cash', 'lfcf', 'cash flow']):
                print(f"Row {idx}: {row.values[:8]}")  # Show first 8 columns
        
    except Exception as e:
        print(f"❌ Error reading FCF DATA sheet: {e}")
    
    # Try DCF sheet
    try:
        dcf_data = pd.read_excel(excel_path, sheet_name='DCF')
        print(f"\n📊 DCF Sheet Shape: {dcf_data.shape}")
        print("📋 First 10 rows of DCF:")
        print(dcf_data.head(10))
        
    except Exception as e:
        print(f"❌ Error reading DCF sheet: {e}")

def calculate_goog_app_fcf():
    """Calculate GOOG FCF using the app"""
    print("\n🤖 CALCULATING GOOG APP FCF")
    print("=" * 50)
    
    try:
        goog_folder = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/GOOG"
        calc = FinancialCalculator(goog_folder)
        calc.load_financial_statements()
        
        # Calculate all FCF types
        print("💰 LFCF (Levered Free Cash Flow):")
        lfcf_values = calc.calculate_levered_fcf()
        print(f"   Values: {lfcf_values}")
        print(f"   Count: {len(lfcf_values)} years")
        
        print("\n💼 FCFF (Free Cash Flow to Firm):")
        fcff_values = calc.calculate_fcf_to_firm()
        print(f"   Values: {fcff_values}")
        print(f"   Count: {len(fcff_values)} years")
        
        print("\n🏦 FCFE (Free Cash Flow to Equity):")
        fcfe_values = calc.calculate_fcf_to_equity()
        print(f"   Values: {fcfe_values}")
        print(f"   Count: {len(fcfe_values)} years")
        
        # Get underlying metrics for validation
        print("\n🔍 UNDERLYING METRICS:")
        metrics = calc._calculate_all_metrics()
        print(f"   Operating Cash Flow: {metrics.get('operating_cash_flow', [])}")
        print(f"   CapEx: {metrics.get('capex', [])}")
        print(f"   EBIT: {metrics.get('ebit', [])[:5]}...")  # First 5 values
        print(f"   Tax Rates: {metrics.get('tax_rates', [])[:5]}...")
        print(f"   Working Capital Changes: {metrics.get('working_capital_changes', [])[:5]}...")
        
        return {
            'LFCF': lfcf_values,
            'FCFF': fcff_values,
            'FCFE': fcfe_values,
            'metrics': metrics
        }
        
    except Exception as e:
        print(f"❌ Error calculating GOOG app FCF: {e}")
        import traceback
        traceback.print_exc()
        return None

def manual_excel_extraction():
    """Manually extract GOOG FCF data from specific cells"""
    print("\n🔧 MANUAL EXCEL EXTRACTION")
    print("=" * 50)
    
    excel_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/GOOG/FCF_Analysis_Alphabet Inc Class C.xlsx"
    
    try:
        # Read with openpyxl for better control
        import openpyxl
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        
        print(f"📋 Available worksheets: {workbook.sheetnames}")
        
        # Check FCF DATA sheet
        if 'FCF DATA' in workbook.sheetnames:
            ws = workbook['FCF DATA']
            print(f"\n📊 FCF DATA sheet dimensions: {ws.max_row} x {ws.max_column}")
            
            # Look for year headers
            years = []
            fcf_data = {}
            
            # Search for years in first few rows
            for row in range(1, 6):
                for col in range(1, 15):
                    cell_value = ws.cell(row=row, column=col).value
                    if isinstance(cell_value, (int, float)) and 2010 <= cell_value <= 2030:
                        years.append((int(cell_value), row, col))
            
            print(f"📅 Found years: {[(year, f'R{row}C{col}') for year, row, col in years]}")
            
            # Search for FCF labels and values
            for row in range(1, min(50, ws.max_row + 1)):
                cell_a = ws.cell(row=row, column=1).value
                if cell_a and isinstance(cell_a, str):
                    cell_str = cell_a.lower()
                    if any(term in cell_str for term in ['fcf', 'free cash', 'operating cash', 'capex']):
                        row_values = []
                        for col in range(2, min(15, ws.max_column + 1)):
                            val = ws.cell(row=row, column=col).value
                            if isinstance(val, (int, float)):
                                row_values.append(val)
                        if row_values:
                            fcf_data[cell_a] = row_values
                            print(f"💰 {cell_a}: {row_values[:8]}")  # First 8 values
        
        workbook.close()
        return fcf_data
        
    except Exception as e:
        print(f"❌ Error in manual extraction: {e}")
        return {}

def validate_calculations():
    """Validate app calculations against expected methodology"""
    print("\n✅ VALIDATION TESTS")
    print("=" * 50)
    
    try:
        goog_folder = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/GOOG"
        calc = FinancialCalculator(goog_folder)
        calc.load_financial_statements()
        
        metrics = calc._calculate_all_metrics()
        operating_cf = metrics.get('operating_cash_flow', [])
        capex = metrics.get('capex', [])
        
        print(f"🔍 Manual LFCF calculation check:")
        print(f"   Operating CF: {operating_cf}")
        print(f"   CapEx: {capex}")
        
        # Manual LFCF calculation
        manual_lfcf = []
        min_len = min(len(operating_cf), len(capex))
        for i in range(min_len):
            lfcf = operating_cf[i] - abs(capex[i])
            manual_lfcf.append(lfcf)
        
        print(f"   Manual LFCF: {manual_lfcf}")
        
        # App LFCF
        app_lfcf = calc.calculate_levered_fcf()
        print(f"   App LFCF: {app_lfcf}")
        
        # Compare
        if len(manual_lfcf) == len(app_lfcf):
            differences = [abs(m - a) for m, a in zip(manual_lfcf, app_lfcf)]
            max_diff = max(differences) if differences else 0
            print(f"   Max difference: {max_diff}")
            if max_diff < 1000:  # Allow for rounding
                print("   ✅ LFCF calculation VALIDATED")
            else:
                print("   ❌ LFCF calculation has differences")
        else:
            print(f"   ❌ Length mismatch: manual {len(manual_lfcf)} vs app {len(app_lfcf)}")
        
    except Exception as e:
        print(f"❌ Validation error: {e}")

def main():
    print("🚀 GOOG FCF COMPREHENSIVE VALIDATION")
    print("=" * 70)
    
    # 1. Analyze Excel structure
    analyze_goog_excel()
    
    # 2. Calculate app FCF
    app_results = calculate_goog_app_fcf()
    
    # 3. Manual Excel extraction
    excel_data = manual_excel_extraction()
    
    # 4. Validate calculations
    validate_calculations()
    
    # 5. Summary
    print("\n📋 SUMMARY")
    print("=" * 50)
    if app_results:
        print(f"✅ App calculated {len(app_results.get('LFCF', []))} years of LFCF")
        print(f"✅ App calculated {len(app_results.get('FCFF', []))} years of FCFF")
        print(f"✅ App calculated {len(app_results.get('FCFE', []))} years of FCFE")
    
    if excel_data:
        print(f"✅ Excel contains {len(excel_data)} FCF-related data series")
        for key in excel_data.keys():
            print(f"   - {key}")
    
    print("\n🎯 NEXT STEPS:")
    print("1. Compare specific Excel FCF values with app LFCF")
    print("2. Investigate FCFF calculation differences")
    print("3. Validate FCFE net borrowing methodology")

if __name__ == "__main__":
    main()