"""
Financial Calculations Module
============================

This module provides comprehensive financial calculation capabilities including Free Cash Flow
(FCF) analysis, financial metrics computation, and integration with multiple data sources for
robust financial analysis and valuation modeling.

Key Features
------------
- **Multi-Type FCF Calculations**: FCFE, FCFF, and Levered FCF with proper tax adjustments
- **Automated Data Loading**: Excel-based financial statement processing with error handling
- **Multi-Source Integration**: Enhanced data manager support for API-based data sources
- **Comprehensive Validation**: Data quality assessment and financial metric validation
- **Market Data Integration**: Real-time stock price and market data fetching
- **Currency Support**: Multi-currency financial analysis with exchange rate handling
- **Performance Optimization**: Caching and efficient calculation algorithms
- **Error Recovery**: Robust error handling with detailed logging and fallback mechanisms

Classes
-------
FinancialCalculator
    Main class for financial calculations and data management

Free Cash Flow Types
-------------------

**FCFE (Free Cash Flow to Equity)**
    Cash flow available to equity holders after all expenses, reinvestment, and debt payments
    Formula: FCFE = Net Income + Depreciation - CapEx - Change in Working Capital - Net Debt Payments

**FCFF (Free Cash Flow to Firm)**
    Cash flow available to all capital providers (debt and equity holders)
    Formula: FCFF = EBIT(1-Tax Rate) + Depreciation - CapEx - Change in Working Capital

**Levered FCF (LFCF)**
    Traditional FCF calculation including debt effects
    Formula: LFCF = Operating Cash Flow - Capital Expenditures

Usage Example
-------------
>>> from core.analysis.engines.financial_calculations import FinancialCalculator
>>> from enhanced_data_manager import EnhancedDataManager
>>>
>>> # Initialize with Excel data
>>> calc = FinancialCalculator("./data/AAPL")
>>>
>>> # Load financial statements
>>> calc.load_financial_statements()
>>>
>>> # Calculate all FCF types
>>> fcf_results = calc.calculate_all_fcf_types()
>>> print(f"FCFE: {fcf_results['FCFE']}")
>>> print(f"FCFF: {fcf_results['FCFF']}")
>>> print(f"LFCF: {fcf_results['LFCF']}")
>>>
>>> # Get comprehensive financial metrics
>>> metrics = calc.get_financial_metrics()
>>> print(f"ROIC: {metrics['roic']:.2%}")
>>> print(f"ROE: {metrics['roe']:.2%}")
>>>
>>> # Enhanced data integration
>>> enhanced_manager = EnhancedDataManager()
>>> calc_enhanced = FinancialCalculator("./data/MSFT", enhanced_manager)

Data Sources and Integration
---------------------------

**Primary Data Source: Excel Files**
- Annual financial statements (FY folders)
- Last Twelve Months data (LTM folders)
- Standardized format with automatic column detection
- Multiple sheet support (Income Statement, Balance Sheet, Cash Flow)

**Enhanced Data Sources**
- Yahoo Finance integration for market data
- Multi-API support (FMP, Alpha Vantage, Polygon)
- Real-time price feeds and fundamental data
- Automatic data validation and reconciliation

**Data Validation Framework**
- Pre-calculation data quality checks
- Cross-validation between data sources
- Outlier detection and correction suggestions
- Comprehensive data quality reporting

Financial Metrics Calculation
-----------------------------

**Profitability Metrics**
- Return on Invested Capital (ROIC)
- Return on Equity (ROE)
- Return on Assets (ROA)
- Gross, Operating, and Net Profit Margins

**Efficiency Metrics**
- Asset Turnover Ratios
- Working Capital Metrics
- Cash Conversion Cycle
- Inventory and Receivables Turnover

**Leverage Metrics**
- Debt-to-Equity Ratios
- Interest Coverage Ratios
- Debt Service Coverage
- Financial Leverage Analysis

**Growth Metrics**
- Revenue Growth Rates (1Y, 3Y, 5Y CAGR)
- Earnings Growth Analysis
- Free Cash Flow Growth Trends
- Sustainable Growth Rate Calculation

Error Handling and Validation
-----------------------------

**Robust Error Recovery**
- Graceful handling of missing data
- Fallback calculation methods
- Clear error reporting with actionable suggestions
- Validation checkpoints throughout calculation process

**Data Quality Assessment**
- Completeness scoring for financial statements
- Consistency checks across time periods
- Reasonableness tests for financial ratios
- Data source reliability scoring

**Calculation Validation**
- Cross-checks between different FCF methods
- Sanity tests for calculated metrics
- Historical trend validation
- Peer comparison validation (when data available)

Currency and International Support
---------------------------------

**Multi-Currency Analysis**
- Automatic currency detection from financial statements
- Exchange rate integration for USD conversion
- TASE (Tel Aviv Stock Exchange) specific handling
- Currency consistency validation across periods

**Regional Adaptations**
- Different accounting standards (GAAP, IFRS)
- Market-specific adjustments
- Local tax rate considerations
- Currency volatility impact analysis

Performance and Optimization
----------------------------

**Calculation Efficiency**
- Vectorized operations using NumPy/Pandas
- Intelligent caching of intermediate results
- Lazy evaluation for expensive operations
- Memory-efficient data processing

**Data Loading Optimization**
- Efficient Excel file processing
- Selective data loading based on requirements
- Background data fetching for real-time updates
- Compressed data storage for caching

Integration Points
-----------------
The module integrates seamlessly with:

**Valuation Modules**
- DCF Valuation: Provides FCF projections and historical data
- DDM Valuation: Supplies earnings and payout ratio data
- P/B Valuation: Delivers book value and financial metrics

**Analysis Tools**
- Centralized Data Manager: Unified data source management
- Analysis Capture: Persistent storage of calculation results
- Streamlit Interface: Real-time financial analysis dashboard

**External Systems**
- Excel-based financial models
- API-based data providers
- Database systems for historical analysis
- Reporting and visualization tools

Configuration Options
--------------------
Extensive configuration support for different use cases:

```python
config = {
    'calculation_precision': 6,          # Decimal places for calculations
    'fcf_scale_factor': 1,              # Scale factor for FCF values
    'validation_strictness': 'moderate', # Data validation level
    'currency_conversion': True,         # Enable currency conversion
    'cache_calculations': True,          # Cache expensive calculations
    'market_data_refresh': 300           # Market data refresh interval (seconds)
}
```

Notes
-----
- All financial calculations follow standard financial theory and best practices
- Results are optimized for integration with valuation models
- Supports both development and production environments
- Comprehensive logging for audit trails and debugging
- Thread-safe operations for concurrent calculations
- Cross-platform compatibility with consistent results
"""

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import logging
from datetime import datetime
from functools import lru_cache
from scipy import stats
import yfinance as yf
import re
import warnings
from typing import Dict, Any, Optional, List, Union, Tuple, Callable, Type
from core.data_processing.data_validator import FinancialDataValidator, validate_financial_calculation_input
from core.analysis.fcf_date_correlation import (
    CorrelatedFCFResults, 
    ComprehensiveFCFResults,
    FCFDataPoint,
    create_correlated_fcf_from_legacy
)
from config import get_config, get_dcf_config, get_unknown_company_name
from utils.error_handler import (
    FinancialAnalysisError,
    CalculationError,
    ValidationError,
    ExcelDataError,
    EnhancedLogger,
    with_error_handling,
)
import functools
import time
import requests

# Configure pandas warnings management for financial data processing
warnings.filterwarnings('ignore', category=pd.errors.SettingWithCopyWarning)
warnings.filterwarnings('ignore', category=FutureWarning, module='pandas')
warnings.filterwarnings('ignore', category=pd.errors.PerformanceWarning)

# Configure NumPy warnings and error handling
# Note: Some warning categories removed in NumPy 2.x
try:
    warnings.filterwarnings('ignore', category=np.VisibleDeprecationWarning)
except AttributeError:
    pass  # VisibleDeprecationWarning not available in NumPy 2.x
try:
    warnings.filterwarnings('ignore', category=np.ComplexWarning)
except AttributeError:
    pass  # ComplexWarning not available in NumPy 2.x

# Set NumPy error handling for financial calculations
np.seterr(divide='warn', over='warn', under='ignore', invalid='warn')

# Configure pandas options for better financial data handling
pd.set_option(
    'mode.chained_assignment', None
)  # Disable chained assignment warnings for financial calculations
pd.set_option('display.float_format', '{:.2f}'.format)  # Better float display for financial data

# Set up enhanced logging
logger = EnhancedLogger(__name__)


def retry_with_exponential_backoff(
    max_retries: int = 3,
    base_delay: float = 1.0,
    max_delay: float = 60.0,
    backoff_factor: float = 2.0,
    retryable_exceptions: Optional[Tuple[Type[Exception], ...]] = None,
) -> Callable:
    """
    Decorator for retrying functions with exponential backoff

    Args:
        max_retries (int): Maximum number of retry attempts
        base_delay (float): Initial delay between retries (seconds)
        max_delay (float): Maximum delay between retries (seconds)
        backoff_factor (float): Multiplier for exponential backoff
        retryable_exceptions (tuple): Tuple of exception types to retry on
    """
    if retryable_exceptions is None:
        retryable_exceptions = (ConnectionError, TimeoutError, requests.RequestException)

    def decorator(func: Callable) -> Callable:
        @functools.wraps(func)
        def wrapper(*args: Any, **kwargs: Any) -> Any:
            last_exception = None
            delay = base_delay

            for attempt in range(max_retries + 1):
                try:
                    return func(*args, **kwargs)
                except retryable_exceptions as e:
                    last_exception = e
                    if attempt < max_retries:
                        logger.warning(f"Attempt {attempt + 1} failed for {func.__name__}: {e}")
                        logger.info(f"Retrying in {delay:.1f} seconds...")
                        time.sleep(delay)
                        delay = min(delay * backoff_factor, max_delay)
                    else:
                        logger.error(f"All {max_retries + 1} attempts failed for {func.__name__}")
                        break
                except Exception as e:
                    # Non-retryable exception, raise immediately
                    logger.error(f"Non-retryable error in {func.__name__}: {e}")
                    raise

            # If we get here, all retries failed
            raise CalculationError(
                f"Function {func.__name__} failed after {max_retries + 1} attempts"
            ) from last_exception

        return wrapper

    return decorator


# Enhanced NaN handling utilities for financial data
def safe_numeric_conversion(value: Any, default: float = 0.0, context: str = "") -> float:
    """
    Safely convert value to numeric with enhanced error handling and logging for financial data.
    Handles infinity, NaN, and various Excel/financial formatting edge cases.

    Args:
        value: Value to convert
        default: Default value if conversion fails
        context: Context for logging

    Returns:
        float: Converted numeric value or default
    """
    try:
        if pd.isna(value) or value == '' or value is None:
            return default

        # Handle string values with common financial formatting
        if isinstance(value, str):
            # Handle Excel error values explicitly
            if value in ('#VALUE!', '#N/A', '#DIV/0!', '#REF!', '#NAME?', '#NUM!', '#NULL!'):
                if context:
                    logger.debug(
                        f"Excel error value '{value}' in {context}, using default {default}"
                    )
                return default
            
            # Remove common financial formatting characters
            clean_value = value.replace(',', '').replace('$', '').replace('%', '')
            clean_value = clean_value.replace('(', '-').replace(')', '').strip()

            if clean_value == '' or clean_value == '-':
                return default

            numeric_val = pd.to_numeric(clean_value, errors='coerce')
        else:
            numeric_val = pd.to_numeric(value, errors='coerce')

        # Check for NaN after conversion
        if pd.isna(numeric_val):
            if context:
                logger.debug(
                    f"Failed to convert value '{value}' in {context}, using default {default}"
                )
            return default

        # Convert to float for final checks
        float_val = float(numeric_val)
        
        # Handle mathematical edge cases: infinity and NaN
        if not np.isfinite(float_val):
            if context:
                logger.debug(
                    f"Non-finite value (inf/NaN) '{value}' in {context}, using default {default}"
                )
            return default

        return float_val

    except (ValueError, TypeError) as e:
        if context:
            logger.warning(
                f"Error converting value '{value}' in {context}: {e}, using default {default}"
            )
        return default
    except Exception as e:
        raise CalculationError(
            f"Unexpected error in numeric conversion: {str(e)}",
            context={'value': value, 'context': context},
        ) from e


def handle_financial_nan_series(
    series: pd.Series, fill_value: float = 0.0, method: str = 'forward'
) -> pd.Series:
    """
    Enhanced NaN handling for financial data series with multiple strategies

    Args:
        series: pandas Series with potential NaN values
        fill_value: Value to use for filling NaN
        method: Method for handling NaN ('forward', 'backward', 'interpolate', 'zero')

    Returns:
        pd.Series: Series with NaN values handled appropriately
    """
    if series.empty:
        return series

    # Convert to numeric first with coercion
    numeric_series = pd.to_numeric(series, errors='coerce')

    if method == 'forward':
        # Forward fill for time series data, then fill remaining with fill_value
        filled_series = numeric_series.fillna(method='ffill').fillna(fill_value)
    elif method == 'backward':
        # Backward fill, then forward fill, then fill_value
        filled_series = (
            numeric_series.fillna(method='bfill').fillna(method='ffill').fillna(fill_value)
        )
    elif method == 'interpolate':
        # Linear interpolation for smooth financial trends
        filled_series = numeric_series.interpolate(method='linear').fillna(fill_value)
    else:  # 'zero' or default
        filled_series = numeric_series.fillna(fill_value)

    return filled_series


class FinancialCalculator:
    """
    Core financial calculations for FCF analysis and DCF valuation
    """

    def __init__(
        self, company_folder: Optional[str], enhanced_data_manager: Optional[Any] = None
    ) -> None:
        """
        Initialize financial calculator with company folder path

        Args:
            company_folder (str): Path to company folder containing FY and LTM subfolders
            enhanced_data_manager (EnhancedDataManager, optional): Enhanced data manager for multi-source data access
        """
        self.company_folder = company_folder
        self.company_name = (
            os.path.basename(company_folder) if company_folder else get_unknown_company_name()
        )
        self.financial_data = {}
        self.fcf_results = {}
        
        # Enhanced FCF results with date correlation
        self.comprehensive_fcf_results = ComprehensiveFCFResults()
        self.date_correlation_info = None  # Will store FY+LTM date correlation
        
        self.ticker_symbol = None
        self.current_stock_price = None
        self.market_cap = None
        self.shares_outstanding = None

        # Currency and exchange information
        self.currency = 'USD'
        self.financial_currency = 'USD'
        self.is_tase_stock = False

        # Enhanced data manager for multi-source data access
        self.enhanced_data_manager = enhanced_data_manager
        if enhanced_data_manager:
            logger.info(
                "FinancialCalculator initialized with enhanced multi-source data management"
            )

        # Data validation components
        self.data_validator = FinancialDataValidator()
        self.data_quality_report = None
        self.validation_enabled = True
        self._last_calculation_error = None  # Diagnostic: last error from calculate_all_fcf_types

        # Automatically extract ticker symbol from folder structure
        self._auto_extract_ticker()

        # Cached metrics to avoid redundant calculations
        self.metrics = {}
        self.metrics_calculated = False

        # Financial data scale factor - Excel data is typically in millions
        # Keep FCF results in millions to match DCF module expectations
        self.financial_scale_factor = 1

        # Date correlation tracking
        self.data_point_dates = {}  # Track dates for each metric/FCF type
        self.actual_ltm_date_used = None  # Track the actual LTM date used
        self._has_api_financial_data = False  # Flag to indicate if data came from API
        self._data_source_used = "Unknown"  # Track the actual data source

        # Initialize VarInputData integration
        self._initialize_var_input_data_integration()
        
        # Auto-load financial statements if company folder exists
        if self.company_folder and os.path.exists(self.company_folder):
            try:
                self.load_financial_statements()
                logger.info(f"Auto-loaded financial statements for {self.company_name}")
            except (FileNotFoundError, PermissionError) as e:
                logger.warning(
                    f"Could not auto-load financial statements for {self.company_name}: {e}"
                )
            except Exception as e:
                logger.error(
                    f"Unexpected error auto-loading financial statements for {self.company_name}: {e}",
                    context={'company_name': self.company_name},
                    error=e,
                )
                raise ExcelDataError(f"Failed to auto-load financial statements: {str(e)}") from e

    def load_financial_statements(self) -> None:
        """
        Load financial statements from Excel files
        """
        try:
            # Define file paths
            fy_folder = os.path.join(self.company_folder, 'FY')
            ltm_folder = os.path.join(self.company_folder, 'LTM')

            # Load FY statements
            for file_name in os.listdir(fy_folder):
                if 'Income Statement' in file_name:
                    self.financial_data['income_fy'] = self._load_excel_data(
                        os.path.join(fy_folder, file_name)
                    )
                elif 'Balance Sheet' in file_name:
                    self.financial_data['balance_fy'] = self._load_excel_data(
                        os.path.join(fy_folder, file_name)
                    )
                elif 'Cash Flow Statement' in file_name:
                    self.financial_data['cashflow_fy'] = self._load_excel_data(
                        os.path.join(fy_folder, file_name)
                    )

            # Load LTM statements
            for file_name in os.listdir(ltm_folder):
                if 'Income Statement' in file_name:
                    self.financial_data['income_ltm'] = self._load_excel_data(
                        os.path.join(ltm_folder, file_name)
                    )
                elif 'Balance Sheet' in file_name:
                    self.financial_data['balance_ltm'] = self._load_excel_data(
                        os.path.join(ltm_folder, file_name)
                    )
                elif 'Cash Flow Statement' in file_name:
                    self.financial_data['cashflow_ltm'] = self._load_excel_data(
                        os.path.join(ltm_folder, file_name)
                    )

            logger.info("Financial statements loaded successfully")

            # Clear cached metrics when new data is loaded
            self.metrics = {}
            self.metrics_calculated = False

        except (FileNotFoundError, PermissionError) as e:
            logger.error(
                f"File access error loading financial statements: {e}",
                context={'company_folder': self.company_folder},
                error=e,
            )
            raise ExcelDataError(f"Cannot access financial statement files: {str(e)}") from e
        except Exception as e:
            logger.error(
                f"Unexpected error loading financial statements: {e}",
                context={'company_folder': self.company_folder},
                error=e,
            )
            raise ExcelDataError(f"Failed to load financial statements: {str(e)}") from e

    def _load_excel_data(self, file_path: str) -> pd.DataFrame:
        """
        Load Excel data and convert to DataFrame with dynamic FY column scanning
        
        This method dynamically discovers all available FY columns in the Excel file,
        sorts them chronologically, and logs the discovered date range for transparency.

        Args:
            file_path (str): Path to Excel file

        Returns:
            pd.DataFrame: Financial data with dynamically discovered FY columns
        """
        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active

            # Convert to list of lists, then to DataFrame
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            # Find the header row (contains 'FY-N', 'FY', etc.)
            header_row_idx = None
            for i, row in enumerate(data):
                if row and any('FY-' in str(cell) or 'FY' == str(cell) for cell in row if cell):
                    header_row_idx = i
                    break

            if header_row_idx is not None:
                headers = data[header_row_idx]
                
                # Dynamic FY column discovery and analysis
                fy_columns = []
                fy_info = {}
                
                for col_idx, header in enumerate(headers):
                    if header is not None:
                        header_str = str(header).strip()
                        
                        # Match FY-N pattern (e.g., FY-9, FY-8, FY-1)
                        if header_str.startswith('FY-'):
                            try:
                                years_back = int(header_str[3:])
                                fy_info[col_idx] = {
                                    'column': header_str,
                                    'years_back': years_back,
                                    'sort_key': years_back  # Higher numbers = further back
                                }
                                fy_columns.append(col_idx)
                            except ValueError:
                                logger.warning(f"Could not parse FY column: {header_str}")
                        
                        # Match current FY pattern
                        elif header_str == 'FY':
                            fy_info[col_idx] = {
                                'column': 'FY',
                                'years_back': 0,
                                'sort_key': 0  # Current year
                            }
                            fy_columns.append(col_idx)
                
                # Sort FY columns chronologically (most recent first: FY, FY-1, FY-2, ...)
                fy_columns.sort(key=lambda idx: fy_info[idx]['sort_key'])
                
                # Log discovered FY column range for transparency
                if fy_columns:
                    fy_names = [fy_info[idx]['column'] for idx in fy_columns]
                    oldest_fy = fy_info[fy_columns[-1]]['column'] if fy_columns else 'None'
                    newest_fy = fy_info[fy_columns[0]]['column'] if fy_columns else 'None'
                    
                    logger.info(
                        f"Discovered {len(fy_columns)} FY columns in {os.path.basename(file_path)}: "
                        f"Range from {oldest_fy} to {newest_fy}",
                        context={
                            'file_path': file_path,
                            'discovered_columns': fy_names,
                            'column_count': len(fy_columns),
                            'date_range': f"{oldest_fy} to {newest_fy}"
                        }
                    )
                else:
                    logger.warning(
                        f"No FY columns found in {os.path.basename(file_path)}",
                        context={'file_path': file_path}
                    )

                data_rows = data[header_row_idx + 1:]
                df = pd.DataFrame(data_rows, columns=headers)
                
                # Store FY column metadata on the DataFrame for future reference
                if hasattr(df, 'attrs'):
                    df.attrs['fy_columns'] = fy_info
                    df.attrs['fy_column_indices'] = fy_columns
                    df.attrs['fy_date_range'] = f"{oldest_fy} to {newest_fy}" if fy_columns else "None"
            else:
                logger.warning(f"No FY header row found in {os.path.basename(file_path)}, using fallback method")
                # Fallback to old method
                if len(data) > 1:
                    df = pd.DataFrame(data[1:], columns=data[0])
                else:
                    df = pd.DataFrame(data)

            return df

        except (FileNotFoundError, PermissionError) as e:
            logger.error(
                f"File access error loading Excel file {file_path}: {e}",
                context={'file_path': file_path},
                error=e,
            )
            raise ExcelDataError(f"Cannot access Excel file: {str(e)}") from e
        except Exception as e:
            logger.error(
                f"Unexpected error loading Excel file {file_path}: {e}",
                context={'file_path': file_path},
                error=e,
            )
            raise ExcelDataError(f"Failed to load Excel file: {str(e)}") from e

    def _extract_metric_with_ltm(
        self, fy_data: pd.DataFrame, ltm_data: pd.DataFrame, metric_name: str
    ) -> List[float]:
        """
        Extract metric values combining FY historical data with LTM latest data.
        Replaces the most recent FY data point with the most recent LTM value.

        Args:
            fy_data (pd.DataFrame): Fiscal year financial data
            ltm_data (pd.DataFrame): Latest twelve months financial data
            metric_name (str): Name of the metric to extract

        Returns:
            list: Combined metric values (FY historical + LTM latest)
        """
        try:
            # Extract FY historical data
            fy_values = self._extract_metric_values(fy_data, metric_name)

            # Extract LTM data if available
            ltm_values = []
            if not ltm_data.empty:
                ltm_values = self._extract_metric_values(ltm_data, metric_name)

            # Combine FY and LTM data
            if fy_values and ltm_values:
                # Use FY historical data (all but last) + most recent LTM value
                combined_values = fy_values[:-1] + [ltm_values[-1]]
                logger.debug(
                    f"{metric_name}: Combined FY historical ({len(fy_values)-1} years) + LTM latest ({ltm_values[-1]:.1f})"
                )
                return combined_values
            elif fy_values:
                # Fallback to FY data only if no LTM available
                logger.warning(f"{metric_name}: No LTM data available, using FY data only")
                return fy_values
            else:
                # No data available
                logger.warning(f"{metric_name}: No data available in FY or LTM")
                return []

        except (KeyError, IndexError, ValueError) as e:
            logger.warning(
                f"Data processing error extracting {metric_name} with LTM integration: {e}"
            )
            # Fallback to FY data only
            return self._extract_metric_values(fy_data, metric_name)
        except Exception as e:
            logger.error(
                f"Unexpected error extracting {metric_name} with LTM integration: {e}",
                context={'metric_name': metric_name},
                error=e,
            )
            raise CalculationError(f"Failed to extract metric {metric_name}: {str(e)}") from e

    def _calculate_all_metrics(self) -> Dict[str, List[float]]:
        """
        Calculate all financial metrics needed for FCF calculations in one pass.
        This eliminates redundant calculations across different FCF methods.
        Uses LTM (Latest Twelve Months) data for the most recent data point.
        Enhanced with comprehensive validation.
        """
        if self.metrics_calculated:
            return self.metrics

        try:
            # Validate financial data before processing
            if self.validation_enabled:
                logger.info("Validating financial data before metrics calculation...")
                self.data_quality_report = self.data_validator.validate_financial_statements(
                    self.financial_data
                )

                # Log validation summary
                if self.data_quality_report.errors:
                    logger.warning(
                        f"Found {len(self.data_quality_report.errors)} validation errors"
                    )
                if self.data_quality_report.warnings:
                    logger.info(
                        f"Found {len(self.data_quality_report.warnings)} validation warnings"
                    )
            # Get financial data once
            income_data = self.financial_data.get('income_fy', pd.DataFrame())
            balance_data = self.financial_data.get('balance_fy', pd.DataFrame())
            cashflow_data = self.financial_data.get('cashflow_fy', pd.DataFrame())

            # Get LTM data for latest periods
            income_ltm = self.financial_data.get('income_ltm', pd.DataFrame())
            balance_ltm = self.financial_data.get('balance_ltm', pd.DataFrame())
            cashflow_ltm = self.financial_data.get('cashflow_ltm', pd.DataFrame())

            # Validate data availability
            missing_data = []
            if income_data.empty:
                missing_data.append("income statement")
            if balance_data.empty:
                missing_data.append("balance sheet")
            if cashflow_data.empty:
                missing_data.append("cash flow statement")

            if missing_data:
                error_msg = f"Missing required financial data: {', '.join(missing_data)}"
                logger.error(error_msg)
                self._last_calculation_error = error_msg
                if self.validation_enabled:
                    self.data_validator.report.add_error(error_msg, "Data availability")
                return {}

            logger.info("Calculating all financial metrics with LTM data integration...")

            # Extract all required metrics once
            metrics = {}

            # Income statement metrics (FY historical + LTM latest)
            metrics['ebit'] = self._extract_metric_with_ltm(income_data, income_ltm, "EBIT")
            metrics['net_income'] = self._extract_metric_with_ltm(
                income_data, income_ltm, "Net Income"
            )
            metrics['tax_expense'] = self._extract_metric_with_ltm(
                income_data, income_ltm, "Income Tax Expense"
            )
            metrics['ebt'] = self._extract_metric_with_ltm(income_data, income_ltm, "EBT")

            # Balance sheet metrics (FY historical + LTM latest)
            metrics['current_assets'] = self._extract_metric_with_ltm(
                balance_data, balance_ltm, "Total Current Assets"
            )
            metrics['current_liabilities'] = self._extract_metric_with_ltm(
                balance_data, balance_ltm, "Total Current Liabilities"
            )

            # Cash flow statement metrics (FY historical + LTM latest)
            metrics['depreciation_amortization'] = self._extract_metric_with_ltm(
                cashflow_data, cashflow_ltm, "Depreciation & Amortization"
            )
            metrics['operating_cash_flow'] = self._extract_metric_with_ltm(
                cashflow_data, cashflow_ltm, "Cash from Operations"
            )
            metrics['capex'] = self._extract_metric_with_ltm(
                cashflow_data, cashflow_ltm, "Capital Expenditure"
            )

            # Extract specific debt financing components for accurate FCFE calculation
            metrics['debt_issued'] = self._extract_metric_with_ltm(
                cashflow_data, cashflow_ltm, "Long-Term Debt Issued"
            )
            metrics['debt_repaid'] = self._extract_metric_with_ltm(
                cashflow_data, cashflow_ltm, "Long-Term Debt Repaid"
            )

            # Calculate derived metrics

            # Net Borrowing calculation (debt issued - debt repaid)
            # For financial statements, missing debt data means zero activity, not missing data
            debt_issued = metrics.get('debt_issued', [])
            debt_repaid = metrics.get('debt_repaid', [])

            # Use the length of the most data-rich metric as reference
            # This handles API sources that may omit some fields
            _primary_metrics = ['net_income', 'ebit', 'operating_cash_flow', 'ebt', 'tax_expense',
                                 'current_assets', 'capex', 'depreciation_amortization']
            _metric_lengths = [len(metrics.get(k, [])) for k in _primary_metrics]
            reference_length = max(_metric_lengths) if _metric_lengths else 0
            logger.info(
                f"Metric lengths — net_income:{len(metrics.get('net_income',[]))}, "
                f"ebit:{len(metrics.get('ebit',[]))}, ebt:{len(metrics.get('ebt',[]))}, "
                f"tax:{len(metrics.get('tax_expense',[]))}, "
                f"curr_assets:{len(metrics.get('current_assets',[]))}, "
                f"curr_liab:{len(metrics.get('current_liabilities',[]))}, "
                f"ocf:{len(metrics.get('operating_cash_flow',[]))}, "
                f"capex:{len(metrics.get('capex',[]))} → reference_length={reference_length}"
            )

            metrics['net_borrowing'] = []

            # Optimized vectorized debt calculation using pandas Series with enhanced NaN handling
            # Convert to pandas Series for efficient vectorized operations
            debt_issued_series = pd.Series(
                debt_issued + [0] * (reference_length - len(debt_issued))
            )
            debt_repaid_series = pd.Series(
                debt_repaid + [0] * (reference_length - len(debt_repaid))
            )

            # Enhanced NaN handling for financial debt data
            debt_issued_clean = handle_financial_nan_series(
                debt_issued_series, fill_value=0.0, method='zero'
            )
            debt_repaid_clean = handle_financial_nan_series(
                debt_repaid_series, fill_value=0.0, method='zero'
            )

            # Vectorized net borrowing calculation
            net_borrowing_series = (
                debt_issued_clean + debt_repaid_clean
            )  # repaid is already negative
            metrics['net_borrowing'] = net_borrowing_series.tolist()

            logger.debug(
                f"Vectorized debt calculation: {len(debt_issued_clean)} debt records processed"
            )

            # Optimized vectorized tax rates calculation with enhanced NaN handling
            ebt_raw = handle_financial_nan_series(
                pd.Series(metrics['ebt']), fill_value=1.0, method='interpolate'
            )
            tax_raw = handle_financial_nan_series(
                pd.Series(metrics['tax_expense']), fill_value=0.0, method='interpolate'
            )

            # Guard against empty or mismatched-length series (e.g. API sources missing fields)
            # np.where requires all broadcast-compatible shapes — mismatches cause ValueError
            if len(ebt_raw) == 0 or len(tax_raw) == 0:
                # Fall back to a flat default tax rate for every year in scope
                n_years = max(len(ebt_raw), len(tax_raw), reference_length)
                metrics['tax_rates'] = [0.25] * n_years
                logger.warning(
                    f"Tax rate defaulted to 0.25 for all {n_years} years "
                    f"(ebt_len={len(ebt_raw)}, tax_len={len(tax_raw)})"
                )
            else:
                # Align lengths if one is shorter (pad front with neutral defaults)
                if len(ebt_raw) != len(tax_raw):
                    target = max(len(ebt_raw), len(tax_raw))
                    if len(ebt_raw) < target:
                        ebt_raw = pd.concat(
                            [pd.Series([1.0] * (target - len(ebt_raw))), ebt_raw]
                        ).reset_index(drop=True)
                    if len(tax_raw) < target:
                        tax_raw = pd.concat(
                            [pd.Series([0.0] * (target - len(tax_raw))), tax_raw]
                        ).reset_index(drop=True)

                ebt_series = ebt_raw
                tax_series = tax_raw

                # Vectorized tax rate calculation with division by zero protection
                with np.errstate(divide='ignore', invalid='ignore'):
                    tax_rates_series = np.where(
                        np.abs(ebt_series) > 1e-6,
                        np.minimum(np.abs(tax_series) / np.abs(ebt_series), 0.35),
                        0.25,
                    )

                tax_rates_clean = (
                    pd.Series(tax_rates_series).fillna(0.25).replace([np.inf, -np.inf], 0.25)
                )
                metrics['tax_rates'] = tax_rates_clean.tolist()

            logger.debug(
                f"Vectorized tax rate calculation: {len(metrics['tax_rates'])} rates calculated"
            )

            # Optimized vectorized working capital changes calculation with enhanced NaN handling
            current_assets_series = handle_financial_nan_series(
                pd.Series(metrics['current_assets']), fill_value=0.0, method='interpolate'
            )
            current_liabilities_series = handle_financial_nan_series(
                pd.Series(metrics['current_liabilities']), fill_value=0.0, method='interpolate'
            )

            # Calculate working capital (assets - liabilities)
            working_capital = current_assets_series - current_liabilities_series

            # Vectorized year-over-year change calculation using diff()
            wc_changes = working_capital.diff()  # This gives NaN for first value
            if not wc_changes.empty:
                wc_changes.iloc[0] = 0  # Set first year to 0 (no prior year to compare)

            # Handle any remaining NaN values in working capital changes
            wc_changes_clean = wc_changes.fillna(0.0)
            metrics['working_capital_changes'] = wc_changes_clean.tolist()

            logger.debug(
                f"Vectorized working capital calculation: {len(wc_changes_clean)} changes calculated"
            )

            # Final validation of calculated metrics
            if self.validation_enabled:
                final_validation = validate_financial_calculation_input(metrics)
                if final_validation.errors:
                    logger.error(
                        f"Metrics validation failed with {len(final_validation.errors)} errors"
                    )
                    for error in final_validation.errors[:3]:  # Show first 3 errors
                        logger.error(f"  - {error['message']}")
                if final_validation.warnings:
                    logger.warning(
                        f"Metrics validation found {len(final_validation.warnings)} warnings"
                    )

            # Store metrics and mark as calculated
            self.metrics = metrics
            self.metrics_calculated = True

            logger.info(
                f"Calculated metrics for {len(metrics.get('working_capital_changes', []))} years"
            )

            # Log data completeness summary
            self._log_metrics_completeness(metrics)

            return metrics

        except (KeyError, ValueError, TypeError) as e:
            import traceback as _tb
            error_msg = f"Data processing error calculating financial metrics: {e}"
            detail = _tb.format_exc()
            logger.error(error_msg + "\n" + detail, context={'company_name': self.company_name}, error=e)
            self._last_calculation_error = f"{error_msg} | {detail}"
            if self.validation_enabled:
                self.data_validator.report.add_error(error_msg, "Metrics calculation")
            return {}
        except Exception as e:
            import traceback as _tb
            error_msg = f"Critical error calculating financial metrics: {e}"
            detail = _tb.format_exc()
            logger.error(error_msg + "\n" + detail, context={'company_name': self.company_name}, error=e)
            self._last_calculation_error = f"{error_msg} | {detail}"
            if self.validation_enabled:
                self.data_validator.report.add_error(error_msg, "Metrics calculation")
            raise CalculationError(f"Financial metrics calculation failed: {str(e)}") from e

    def _log_metrics_completeness(self, metrics: Dict[str, List[float]]) -> None:
        """
        Log a summary of metrics completeness and data quality
        """
        logger.info("\n" + "=" * 50)
        logger.info("METRICS CALCULATION SUMMARY")
        logger.info("=" * 50)

        for metric_name, values in metrics.items():
            if isinstance(values, list) and values:
                non_zero_count = sum(1 for v in values if v != 0)
                zero_count = len(values) - non_zero_count
                logger.info(
                    f"{metric_name}: {len(values)} years, {non_zero_count} non-zero, {zero_count} zero"
                )
            elif not values:
                logger.warning(f"{metric_name}: NO DATA")

        # Overall assessment
        total_metrics = len(metrics)
        complete_metrics = sum(1 for values in metrics.values() if values and len(values) > 0)
        completeness_rate = (complete_metrics / total_metrics * 100) if total_metrics > 0 else 0

        logger.info(
            f"\nOverall Completeness: {complete_metrics}/{total_metrics} metrics ({completeness_rate:.1f}%)"
        )

        if completeness_rate >= 90:
            logger.info("✓ EXCELLENT: High data completeness")
        elif completeness_rate >= 70:
            logger.warning("⚠ GOOD: Adequate data completeness")
        else:
            logger.error("✗ POOR: Low data completeness - review source data")

        logger.info("=" * 50)

    def get_data_quality_report(self) -> Any:
        """
        Get the current data quality report

        Returns:
            DataQualityReport: Current validation report
        """
        if self.data_quality_report:
            return self.data_quality_report
        else:
            return self.data_validator.report

    def get_latest_report_date(self) -> str:
        """
        Get the latest report date from the financial data with comprehensive source logging

        Returns:
            str: Latest report date in format 'YYYY-MM-DD' or 'Unknown' if not available
        """
        latest_date = "Unknown"
        date_source = "Unknown"
        
        logger.info("=== DATE SOURCE TRACKING ===")
        logger.info("Starting latest report date extraction with comprehensive source tracking")
        
        try:
            # Check if we have financial data with report dates
            if hasattr(self, 'financial_data') and self.financial_data:
                logger.info(f"Financial data available with keys: {list(self.financial_data.keys()) if isinstance(self.financial_data, dict) else 'Non-dict structure'}")
                
                # Look for report_date in the financial data
                if isinstance(self.financial_data, dict):
                    # Check for report_date directly
                    if 'report_date' in self.financial_data:
                        latest_date = str(self.financial_data['report_date'])
                        date_source = "financial_data.report_date (direct)"
                        logger.info(f"✓ Found date from direct report_date: {latest_date}")
                    # Check for quarterly data with dates
                    elif 'quarterly' in self.financial_data:
                        quarterly = self.financial_data['quarterly']
                        if isinstance(quarterly, list) and len(quarterly) > 0:
                            logger.info(f"Checking quarterly data with {len(quarterly)} entries")
                            # Get the most recent quarter (first item is usually latest)
                            latest_quarter = quarterly[0]
                            if isinstance(latest_quarter, dict):
                                if 'date' in latest_quarter:
                                    latest_date = str(latest_quarter['date'])
                                    date_source = f"financial_data.quarterly[0].date"
                                    logger.info(f"✓ Found date from quarterly data: {latest_date}")
                                elif 'report_date' in latest_quarter:
                                    latest_date = str(latest_quarter['report_date'])
                                    date_source = f"financial_data.quarterly[0].report_date"
                                    logger.info(f"✓ Found date from quarterly report_date: {latest_date}")
                    # Check for annual data with dates
                    elif 'annual' in self.financial_data:
                        annual = self.financial_data['annual']
                        if isinstance(annual, list) and len(annual) > 0:
                            logger.info(f"Checking annual data with {len(annual)} entries")
                            # Get the most recent year (first item is usually latest)
                            latest_annual = annual[0]
                            if isinstance(latest_annual, dict):
                                if 'date' in latest_annual:
                                    latest_date = str(latest_annual['date'])
                                    date_source = f"financial_data.annual[0].date"
                                    logger.info(f"✓ Found date from annual data: {latest_date}")
                                elif 'report_date' in latest_annual:
                                    latest_date = str(latest_annual['report_date'])
                                    date_source = f"financial_data.annual[0].report_date"
                                    logger.info(f"✓ Found date from annual report_date: {latest_date}")
                    
                    # Check for standardized data with report_date from API converters
                    elif 'cash_flow' in self.financial_data:
                        logger.info("Checking cash_flow data for report dates")
                        cash_flow_data = self.financial_data['cash_flow']
                        if isinstance(cash_flow_data, dict) and 'report_date' in cash_flow_data:
                            latest_date = str(cash_flow_data['report_date'])
                            date_source = "financial_data.cash_flow.report_date"
                            logger.info(f"✓ Found date from cash_flow dict: {latest_date}")
                        elif isinstance(cash_flow_data, list) and len(cash_flow_data) > 0:
                            if isinstance(cash_flow_data[0], dict) and 'report_date' in cash_flow_data[0]:
                                latest_date = str(cash_flow_data[0]['report_date'])
                                date_source = "financial_data.cash_flow[0].report_date"
                                logger.info(f"✓ Found date from cash_flow list: {latest_date}")
                    
                    # Check for income statement data with report dates
                    elif 'income_statement' in self.financial_data:
                        logger.info("Checking income_statement data for report dates")
                        income_data = self.financial_data['income_statement']
                        if isinstance(income_data, dict) and 'report_date' in income_data:
                            latest_date = str(income_data['report_date'])
                            date_source = "financial_data.income_statement.report_date"
                            logger.info(f"✓ Found date from income_statement dict: {latest_date}")
                        elif isinstance(income_data, list) and len(income_data) > 0:
                            if isinstance(income_data[0], dict) and 'report_date' in income_data[0]:
                                latest_date = str(income_data[0]['report_date'])
                                date_source = "financial_data.income_statement[0].report_date"
                                logger.info(f"✓ Found date from income_statement list: {latest_date}")
                                
            # Check in pre-calculated FCF results for report dates
            if latest_date == "Unknown" and hasattr(self, 'fcf_results') and self.fcf_results:
                logger.info("Checking pre-calculated FCF results for report dates")
                for fcf_type, fcf_data in self.fcf_results.items():
                    if isinstance(fcf_data, dict) and 'report_date' in fcf_data:
                        latest_date = str(fcf_data['report_date'])
                        date_source = f"fcf_results.{fcf_type}.report_date"
                        logger.info(f"✓ Found date from FCF results ({fcf_type}): {latest_date}")
                        break
                                
            # Try to extract from Excel files directly using Period End Date extraction
            if latest_date == "Unknown":
                logger.info("Attempting Excel file date extraction as fallback")
                try:
                    from excel_utils import get_period_dates_from_excel, get_fy_ltm_correlated_dates
                    from pathlib import Path
                    import os
                    
                    # Look for company folders and income statement files
                    current_dir = Path('.')
                    company_dirs = [d for d in current_dir.iterdir() if d.is_dir() and not d.name.startswith('.')]
                    logger.info(f"Found {len(company_dirs)} potential company directories")
                    
                    for company_dir in company_dirs:
                        logger.info(f"Searching company directory: {company_dir.name}")
                        # Check FY folder first (Full Year data)
                        fy_dir = company_dir / 'FY'
                        if fy_dir.exists():
                            logger.info(f"Found FY directory: {fy_dir}")
                            for file in fy_dir.iterdir():
                                if file.suffix.lower() == '.xlsx' and 'income' in file.name.lower():
                                    logger.info(f"Processing Excel file: {file.name}")
                                    period_dates = get_period_dates_from_excel(str(file))
                                    if period_dates:
                                        # Get the most recent date (usually first in list)
                                        latest_date = period_dates[0]
                                        date_source = f"Excel:{file.name} (FY)"
                                        logger.info(f"✓ Extracted latest date from Excel FY: {latest_date}")
                                        break
                        
                        # Also check LTM folder if needed
                        if latest_date == "Unknown":
                            ltm_dir = company_dir / 'LTM'
                            if ltm_dir.exists():
                                logger.info(f"Found LTM directory: {ltm_dir}")
                                for file in ltm_dir.iterdir():
                                    if file.suffix.lower() == '.xlsx' and 'income' in file.name.lower():
                                        logger.info(f"Processing LTM Excel file: {file.name}")
                                        period_dates = get_period_dates_from_excel(str(file))
                                        if period_dates:
                                            # Get the most recent date (usually first in list)
                                            latest_date = period_dates[0]
                                            date_source = f"Excel:{file.name} (LTM)"
                                            logger.info(f"✓ Extracted latest date from LTM Excel: {latest_date}")
                                            break
                        
                        # Break if we found a date from any company folder
                        if latest_date != "Unknown":
                            break
                            
                except Exception as e:
                    logger.warning(f"Excel date extraction failed: {e}")
            
            # If Excel extraction didn't work, keep "Unknown" status
            if latest_date == "Unknown":
                date_source = "No report date available"
                logger.warning("No report date found in any source, returning Unknown")
                            
        except Exception as e:
            logger.error(f"Error retrieving latest report date: {e}")
            date_source = f"Error: {str(e)}"
            
        # Clean up the date format if it contains extra information
        if latest_date != "Unknown":
            # Extract just the date part if it contains time information
            if ' ' in latest_date:
                latest_date = latest_date.split(' ')[0]
            # Ensure it's in YYYY-MM-DD format if possible
            if len(latest_date) == 10 and latest_date.count('-') == 2:
                pass  # Already in correct format
            elif len(latest_date) == 8 and latest_date.isdigit():
                # Convert YYYYMMDD to YYYY-MM-DD
                latest_date = f"{latest_date[:4]}-{latest_date[4:6]}-{latest_date[6:8]}"
        
        # Final logging summary
        logger.info("=== DATE SOURCE TRACKING SUMMARY ===")
        logger.info(f"Final report date: {latest_date}")
        logger.info(f"Date source: {date_source}")
        logger.info("=====================================")
                
        return latest_date

    def _extract_excel_dates(self) -> List[str]:
        """
        Extract dates from Excel financial data structure.
        
        Returns:
            List[str]: List of dates in YYYY-MM-DD format from financial statements
        """
        try:
            # Handle the case where financial_data is a dict of DataFrames (typical structure)
            if hasattr(self, 'financial_data') and isinstance(self.financial_data, dict):
                logger.info("Financial data is a dict, looking for dates in individual DataFrames")
                
                # Look through each statement DataFrame
                for statement_name, df in self.financial_data.items():
                    if isinstance(df, pd.DataFrame) and not df.empty:
                        logger.info(f"Checking {statement_name} for dates (shape: {df.shape})")
                        
                        # Look for dates in row 1 (index 1) which commonly contains period end dates
                        if len(df) > 1:
                            try:
                                date_row = df.iloc[1]  # Row index 1
                                potential_dates = []
                                
                                for value in date_row:
                                    if pd.notna(value):
                                        date_str = str(value).strip()
                                        # Check if this looks like a date
                                        if self._looks_like_date(date_str):
                                            standardized = self._standardize_excel_date(date_str)
                                            if standardized:
                                                potential_dates.append(standardized)
                                
                                if potential_dates:
                                    logger.info(f"Found dates in {statement_name} row 1: {potential_dates}")
                                    return potential_dates
                            except Exception as e:
                                logger.warning(f"Error processing {statement_name}: {e}")
                                continue
                        
                        # Also check for specific date rows by name
                        date_patterns = ['period end date', 'period_end_date', 'report date', 'reporting date']
                        for pattern in date_patterns:
                            matching_index = [idx for idx in df.index if pattern in str(idx).lower()]
                            if matching_index:
                                period_dates = df.loc[matching_index[0]]
                                potential_dates = []
                                for date in period_dates:
                                    if pd.notna(date):
                                        standardized = self._standardize_excel_date(str(date))
                                        if standardized:
                                            potential_dates.append(standardized)
                                if potential_dates:
                                    logger.info(f"Found period dates via pattern '{pattern}' in {statement_name}: {potential_dates}")
                                    return potential_dates
                
                logger.warning("No dates found in any DataFrame")
                return None
                                
            # Handle the case where financial_data is a single DataFrame
            elif hasattr(self, 'financial_data') and isinstance(self.financial_data, pd.DataFrame):
                logger.info("Financial data is a single DataFrame")
                
                # Look for dates in row 1
                if len(self.financial_data) > 1:
                    date_row = self.financial_data.iloc[1]
                    potential_dates = []
                    
                    for value in date_row:
                        if pd.notna(value):
                            date_str = str(value).strip()
                            if self._looks_like_date(date_str):
                                standardized = self._standardize_excel_date(date_str)
                                if standardized:
                                    potential_dates.append(standardized)
                    
                    if potential_dates:
                        logger.info(f"Found dates in row 1: {potential_dates}")
                        return potential_dates
                
                # Fallback: Extract years from DataFrame index
                if hasattr(self.financial_data.index, 'tolist'):
                    dates = self.financial_data.index.tolist()
                    try:
                        years = [int(d) for d in dates if str(d).isdigit() or isinstance(d, (int, float))]
                        return [f"{year}-12-31" for year in years]
                    except (ValueError, TypeError):
                        logger.warning("Could not convert index to years")
                        return None
                
                return None
            else:
                logger.warning(f"Unexpected financial_data type: {type(self.financial_data)}")
                return None
                
        except Exception as e:
            logger.warning(f"Error extracting Excel dates: {e}")
            return None

    def _looks_like_date(self, value_str: str) -> bool:
        """
        Check if a string looks like a date.
        
        Args:
            value_str (str): String to check
            
        Returns:
            bool: True if it looks like a date
        """
        import re
        
        # Common date patterns
        date_patterns = [
            r'\d{4}-\d{1,2}-\d{1,2}',      # YYYY-MM-DD
            r'\d{1,2}/\d{1,2}/\d{4}',      # MM/DD/YYYY
            r'\d{4}/\d{1,2}/\d{1,2}',      # YYYY/MM/DD
            r'\d{1,2}-\d{1,2}-\d{4}',      # MM-DD-YYYY
            r'\d{8}',                       # YYYYMMDD
            r'^\d{4}$'                      # Just year
        ]
        
        for pattern in date_patterns:
            if re.match(pattern, value_str):
                return True
        
        return False

    def _standardize_excel_date(self, date_str: str) -> str:
        """
        Convert Excel date string to standard YYYY-MM-DD format.
        
        Args:
            date_str (str): Date string from Excel (various formats possible)
            
        Returns:
            str: Standardized date in YYYY-MM-DD format or None if parsing fails
        """
        try:
            import re
            from datetime import datetime
            
            # Clean the date string
            date_str = str(date_str).strip()
            
            # Handle various date formats that might come from Excel
            date_patterns = [
                r'(\d{4})-(\d{1,2})-(\d{1,2})',  # YYYY-MM-DD or YYYY-M-D
                r'(\d{1,2})/(\d{1,2})/(\d{4})',  # MM/DD/YYYY or M/D/YYYY
                r'(\d{4})/(\d{1,2})/(\d{1,2})',  # YYYY/MM/DD or YYYY/M/D
                r'(\d{1,2})-(\d{1,2})-(\d{4})',  # MM-DD-YYYY or M-D-YYYY
                r'(\d{4})(\d{2})(\d{2})',         # YYYYMMDD
                r'^(\d{4})$'                      # Just year
            ]
            
            for pattern in date_patterns:
                match = re.match(pattern, date_str)
                if match:
                    groups = match.groups()
                    if len(groups) == 1:  # Just year
                        year = int(groups[0])
                        return f"{year}-12-31"
                    elif len(groups) == 3:
                        # Try different interpretations based on pattern
                        if pattern.startswith(r'(\d{4})-'):  # Year first (YYYY-MM-DD)
                            year, month, day = int(groups[0]), int(groups[1]), int(groups[2])
                        elif pattern.endswith(r'(\d{4})'):  # Year last (MM/DD/YYYY or MM-DD-YYYY)
                            month, day, year = int(groups[0]), int(groups[1]), int(groups[2])
                        elif pattern.startswith(r'(\d{4})/') or pattern.startswith(r'(\d{4})(\d{2})'):  # YYYY/MM/DD or YYYYMMDD
                            year, month, day = int(groups[0]), int(groups[1]), int(groups[2])
                        else:
                            # Default interpretation
                            year, month, day = int(groups[0]), int(groups[1]), int(groups[2])
                        
                        # Validate date
                        if 1 <= month <= 12 and 1 <= day <= 31:
                            return f"{year:04d}-{month:02d}-{day:02d}"
            
            # Try pandas datetime parsing as fallback
            try:
                parsed_date = pd.to_datetime(date_str)
                return parsed_date.strftime('%Y-%m-%d')
            except:
                pass
                
            logger.warning(f"Could not parse date: {date_str}")
            return None
            
        except Exception as e:
            logger.warning(f"Error standardizing Excel date '{date_str}': {e}")
            return None

    def _extract_api_dates(self) -> List[str]:
        """
        Extract dates from API financial data structure.
        
        Returns:
            List[str]: List of dates in YYYY-MM-DD format from API data
        """
        try:
            dates = []
            if hasattr(self, 'financial_data') and isinstance(self.financial_data, dict):
                # Check annual data
                if 'annual' in self.financial_data:
                    annual_data = self.financial_data['annual']
                    if isinstance(annual_data, list):
                        for entry in annual_data:
                            if isinstance(entry, dict) and 'date' in entry:
                                dates.append(entry['date'])
                
                # Check quarterly data
                if 'quarterly' in self.financial_data:
                    quarterly_data = self.financial_data['quarterly']
                    if isinstance(quarterly_data, list):
                        for entry in quarterly_data:
                            if isinstance(entry, dict) and 'date' in entry:
                                if entry['date'] not in dates:  # Avoid duplicates
                                    dates.append(entry['date'])
                
                # Sort dates in descending order (most recent first)
                dates.sort(reverse=True)
                return dates
            return []
        except Exception as e:
            logger.warning(f"Error extracting API dates: {e}")
            return []

    def _track_calculation_dates(self, metric_name: str, values: List[float]) -> None:
        """
        Track dates for calculated financial metrics.
        
        Args:
            metric_name (str): Name of the metric (e.g., 'FCFF', 'FCFE', 'LFCF')
            values (List[float]): Calculated values to track dates for
        """
        try:
            if not hasattr(self, 'data_point_dates'):
                self.data_point_dates = {}
            
            dates = []
            
            if self._has_api_financial_data:
                # Use API dates
                api_dates = self._extract_api_dates()
                # Take only as many dates as we have values
                dates = api_dates[:len(values)] if api_dates else []
            else:
                # Use Excel dates
                excel_dates = self._extract_excel_dates()
                if excel_dates:
                    # Convert years to standard date format (assume December 31st)
                    dates = [f"{year}-12-31" for year in excel_dates[:len(values)]]
            
            # Store the dates for this metric
            if dates:
                self.data_point_dates[metric_name] = dates
                logger.info(f"Tracked {len(dates)} dates for {metric_name}")
            else:
                logger.warning(f"No dates available to track for {metric_name}")
                
        except Exception as e:
            logger.error(f"Error tracking calculation dates for {metric_name}: {e}")

    def _create_enhanced_fcf_correlation(self, fcf_type: str, values: List[float]) -> None:
        """
        Create enhanced FCF correlation with proper date tracking
        
        Args:
            fcf_type (str): Type of FCF ('FCFF', 'FCFE', 'LFCF')
            values (List[float]): Calculated FCF values
        """
        try:
            # Initialize date correlation if not already done
            if self.date_correlation_info is None:
                correlation_initialized = self.initialize_enhanced_date_correlation()
                if not correlation_initialized:
                    logger.warning(f"Could not initialize date correlation for {fcf_type}, using fallback")
                    # Create basic correlation without detailed date info
                    self._create_basic_fcf_correlation(fcf_type, values)
                    return
            
            # Get correlated dates and sources
            dates = self.date_correlation_info.get('correlated_dates', [])
            sources = self.date_correlation_info.get('date_sources', [])
            source_files = self.date_correlation_info.get('source_files', [])
            
            # Ensure we have the right number of dates for our values
            if len(dates) != len(values):
                logger.warning(f"Date count ({len(dates)}) doesn't match value count ({len(values)}) for {fcf_type}")
                # Truncate or pad as needed
                if len(dates) > len(values):
                    dates = dates[-len(values):]  # Take the most recent dates
                    sources = sources[-len(values):] if sources else ['Unknown'] * len(values)
                    source_files = source_files[-len(values):] if source_files else ['Unknown'] * len(values)
                else:
                    # Pad with unknown values
                    dates.extend(['Unknown'] * (len(values) - len(dates)))
                    sources.extend(['Unknown'] * (len(values) - len(sources))) if sources else ['Unknown'] * len(values)
                    source_files.extend(['Unknown'] * (len(values) - len(source_files))) if source_files else ['Unknown'] * len(values)
            
            # Create enhanced correlated FCF results
            correlated_fcf = CorrelatedFCFResults(fcf_type=fcf_type)
            
            for i, value in enumerate(values):
                correlated_fcf.add_data_point(
                    value=value,
                    report_date=dates[i] if i < len(dates) else 'Unknown',
                    source_type=sources[i] if i < len(sources) else 'Unknown',
                    source_file=source_files[i] if i < len(source_files) else 'Unknown',
                    calculation_method='standard'
                )
            
            # Validate the correlation
            correlation_valid = correlated_fcf.validate_correlation()
            
            # Add to comprehensive results
            self.comprehensive_fcf_results.add_fcf_results(fcf_type, correlated_fcf)
            
            logger.info(f"Enhanced FCF correlation created for {fcf_type}: {len(values)} points")
            logger.info(f"Date correlation valid: {correlation_valid}")
            if correlation_valid and dates:
                logger.info(f"Date range: {min(dates)} to {max(dates)}")
                
        except Exception as e:
            logger.error(f"Error creating enhanced FCF correlation for {fcf_type}: {e}")
            # Fallback to basic correlation
            self._create_basic_fcf_correlation(fcf_type, values)

    def _create_basic_fcf_correlation(self, fcf_type: str, values: List[float]) -> None:
        """
        Create basic FCF correlation as fallback when enhanced correlation fails
        
        Args:
            fcf_type (str): Type of FCF 
            values (List[float]): Calculated FCF values
        """
        try:
            # Create basic correlation with generic dates
            basic_dates = [f"Unknown_{i+1}" for i in range(len(values))]
            basic_sources = ['Unknown'] * len(values)
            basic_files = ['Unknown'] * len(values)
            
            correlated_fcf = create_correlated_fcf_from_legacy(
                fcf_type=fcf_type,
                values=values,
                dates=basic_dates,
                sources=basic_sources,
                source_files=basic_files
            )
            
            self.comprehensive_fcf_results.add_fcf_results(fcf_type, correlated_fcf)
            logger.info(f"Basic FCF correlation created for {fcf_type}: {len(values)} points")
            
        except Exception as e:
            logger.error(f"Error creating basic FCF correlation for {fcf_type}: {e}")

    def _get_latest_report_date(self) -> str:
        """
        Wrapper for get_latest_report_date that also tracks the LTM date used.
        
        Returns:
            str: Latest report date and updates actual_ltm_date_used
        """
        latest_date = self.get_latest_report_date()
        self.actual_ltm_date_used = latest_date
        return latest_date

    def set_validation_enabled(self, enabled: bool):
        """
        Enable or disable data validation (for performance in testing)

        Args:
            enabled (bool): Whether to enable validation
        """
        self.validation_enabled = enabled
        logger.info(f"Data validation {'enabled' if enabled else 'disabled'}")

    def initialize_enhanced_date_correlation(self) -> bool:
        """
        Initialize enhanced date correlation system using FY+LTM pattern
        
        Returns:
            bool: True if initialization successful, False otherwise
        """
        try:
            if not self.company_folder:
                logger.warning("No company folder available for enhanced date correlation")
                return False
            
            from excel_utils import get_fy_ltm_correlated_dates
            
            # Get the correlated date information
            logger.info("Initializing enhanced FCF date correlation system")
            self.date_correlation_info = get_fy_ltm_correlated_dates(self.company_folder)
            
            if self.date_correlation_info['extraction_success']:
                logger.info(f"Enhanced date correlation initialized successfully")
                logger.info(f"Pattern: {self.date_correlation_info['correlation_pattern']}")
                logger.info(f"Total dates: {len(self.date_correlation_info['correlated_dates'])}")
                logger.info(f"Date sequence: {self.date_correlation_info['correlated_dates']}")
                
                # Update comprehensive FCF results metadata
                self.comprehensive_fcf_results.company_info.update({
                    'company_folder': self.company_folder,
                    'company_name': self.company_name,
                    'date_correlation_pattern': self.date_correlation_info['correlation_pattern']
                })
                
                return True
            else:
                logger.warning(f"Enhanced date correlation initialization failed: {self.date_correlation_info.get('error', 'Unknown error')}")
                return False
                
        except Exception as e:
            logger.error(f"Error initializing enhanced date correlation: {e}")
            return False

    def calculate_fcf_to_firm(self, use_var_input_data: bool = True) -> List[float]:
        """
        Calculate Free Cash Flow to Firm (FCFF)
        FCFF = EBIT(1-Tax Rate) + Depreciation & Amortization - Change in Working Capital - Capital Expenditures
        
        Args:
            use_var_input_data: Whether to use VarInputData system for inputs/outputs (default: True)
        """
        try:
            if use_var_input_data:
                # Try to initialize VarInputData if not already done
                if not hasattr(self, '_var_input_data') or self._var_input_data is None:
                    self._initialize_var_input_data_integration()
                
                if self._var_input_data is not None:
                    result = self._calculate_fcf_to_firm_with_var_data()
                    if result:
                        return result
                    # VarInputData returned empty — fall through to legacy

            # Fallback to legacy method for backward compatibility
            return self._calculate_fcf_to_firm_legacy()

        except Exception as e:
            logger.error(f"Error calculating FCFF: {e}")
            return []
    
    def _calculate_fcf_to_firm_with_var_data(self) -> List[float]:
        """Calculate FCFF using VarInputData system"""
        try:
            from core.data_processing.var_input_data import get_var_input_data
            var_data = get_var_input_data()
            
            # Get company symbol for data retrieval
            symbol = self.ticker_symbol or "UNKNOWN"
            
            # Get required financial variables from var_input_data
            # Try to get historical data (last 10 years)
            ebit_data = var_data.get_historical_data(symbol, "ebit", years=10)
            tax_expense_data = var_data.get_historical_data(symbol, "tax_expense", years=10) 
            revenue_data = var_data.get_historical_data(symbol, "revenue", years=10)  # For tax rate calc
            da_data = var_data.get_historical_data(symbol, "depreciation_amortization", years=10)
            capex_data = var_data.get_historical_data(symbol, "capital_expenditures", years=10)
            
            # Get working capital components for change calculation  
            current_assets_data = var_data.get_historical_data(symbol, "current_assets", years=10)
            current_liabilities_data = var_data.get_historical_data(symbol, "current_liabilities", years=10)
            
            if not ebit_data:
                logger.warning(f"No EBIT data available for {symbol} in VarInputData")
                return []
            
            # Convert historical data to period-indexed dictionaries
            ebit_by_period = {period: value for period, value in ebit_data}
            tax_by_period = {period: value for period, value in tax_expense_data}
            revenue_by_period = {period: value for period, value in revenue_data}
            da_by_period = {period: value for period, value in da_data}
            capex_by_period = {period: value for period, value in capex_data}
            current_assets_by_period = {period: value for period, value in current_assets_data}
            current_liab_by_period = {period: value for period, value in current_liabilities_data}
            
            # Get common periods (sorted most recent first)
            common_periods = sorted(set(ebit_by_period.keys()), reverse=True)
            
            fcff_values = []
            periods_calculated = []
            
            # Calculate working capital changes and FCFF for each period
            previous_working_capital = None
            
            for period in reversed(common_periods):  # Calculate chronologically for WC changes
                try:
                    # Get values for this period (with defaults)
                    ebit = ebit_by_period.get(period, 0)
                    tax_expense = tax_by_period.get(period, 0)
                    revenue = revenue_by_period.get(period, 1)  # Avoid division by zero
                    da = da_by_period.get(period, 0)
                    capex = capex_by_period.get(period, 0)
                    
                    current_assets = current_assets_by_period.get(period, 0)
                    current_liabilities = current_liab_by_period.get(period, 0)
                    
                    # Calculate tax rate (tax expense / pre-tax income)
                    # Use revenue as proxy if EBIT is too low, or default to 25%
                    if abs(ebit) > 0.01:  # Avoid division by very small numbers
                        tax_rate = min(abs(tax_expense / ebit), 0.5)  # Cap at 50%
                    else:
                        tax_rate = 0.25  # Default corporate tax rate
                    
                    # Calculate working capital change
                    working_capital = current_assets - current_liabilities
                    working_capital_change = 0
                    
                    if previous_working_capital is not None:
                        working_capital_change = working_capital - previous_working_capital
                    
                    previous_working_capital = working_capital
                    
                    # Calculate FCFF = EBIT(1-Tax Rate) + DA - WC Change - CapEx
                    after_tax_ebit = ebit * (1 - tax_rate)
                    fcff = after_tax_ebit + da - working_capital_change - abs(capex)
                    
                    fcff_values.append(fcff)
                    periods_calculated.append(period)
                    
                    logger.debug(f"FCFF[{period}]: EBIT({ebit:.1f}) * (1-{tax_rate:.2%}) + DA({da:.1f}) - WC_Change({working_capital_change:.1f}) - CapEx({abs(capex):.1f}) = {fcff:.1f}")
                    
                except Exception as e:
                    logger.warning(f"Error calculating FCFF for period {period}: {e}")
                    continue
            
            # Reverse to get most recent first (matches legacy behavior)
            fcff_values.reverse()
            periods_calculated.reverse()
            
            # Store results back to VarInputData with metadata
            self._store_fcf_results_to_var_data("fcff", fcff_values, periods_calculated, symbol, 
                                               calculation_method="EBIT(1-Tax Rate) + DA - WC Change - CapEx")
            
            # Store in legacy results for backward compatibility
            self.fcf_results['FCFF'] = fcff_values
            
            logger.info(f"FCFF calculated using VarInputData: {len(fcff_values)} periods for {symbol}")
            return fcff_values
            
        except Exception as e:
            logger.error(f"Error in VarInputData FCFF calculation: {e}")
            # Fallback to legacy method
            return self._calculate_fcf_to_firm_legacy()
    
    def _calculate_fcf_to_firm_legacy(self) -> List[float]:
        """Legacy FCFF calculation method (for backward compatibility)"""
        try:
            # Get pre-calculated metrics
            metrics = self._calculate_all_metrics()
            if not metrics:
                logger.warning("No metrics available for FCFF calculation")
                return []

            # Extract required metrics
            ebit_values = metrics.get('ebit', [])
            tax_rates = metrics.get('tax_rates', [])
            da_values = metrics.get('depreciation_amortization', [])
            capex_values = metrics.get('capex', [])
            working_capital_changes = metrics.get('working_capital_changes', [])

            # Handle missing data - use zeros if critical data is missing but other data exists
            max_length = max(
                len(ebit_values),
                len(tax_rates),
                len(da_values),
                len(capex_values),
                len(working_capital_changes),
            )

            # Pad missing arrays with zeros (special case handling)
            if len(capex_values) == 0 and max_length > 0:
                logger.info(
                    "CapEx data missing - using zeros (typical for REITs/Financial companies)"
                )
                capex_values = [0.0] * max_length

            if len(da_values) == 0 and max_length > 0:
                logger.info("Depreciation & Amortization data missing - using zeros")
                da_values = [0.0] * max_length

            # Calculate FCFF - Use minimum of available data
            fcff_values = []
            min_length = min(
                len(ebit_values),
                len(tax_rates),
                len(da_values),
                len(capex_values),
                len(working_capital_changes),
            )

            for i in range(min_length):
                # Use consistent indexing for all components
                after_tax_ebit = ebit_values[i] * (1 - tax_rates[i])
                fcff = (
                    after_tax_ebit
                    + da_values[i]
                    - working_capital_changes[i]
                    - abs(capex_values[i])
                )
                fcff_values.append(fcff)

            # Apply scale factor to maintain values in millions
            scaled_fcff_values = [value * self.financial_scale_factor for value in fcff_values]
            self.fcf_results['FCFF'] = scaled_fcff_values
            
            # Track dates for this calculation (legacy)
            self._track_calculation_dates('FCFF', scaled_fcff_values)
            
            # Enhanced date correlation tracking
            self._create_enhanced_fcf_correlation('FCFF', scaled_fcff_values)
            
            logger.info(
                f"FCFF calculated: {len(fcff_values)} years (kept in millions for DCF compatibility)"
            )
            return scaled_fcff_values

        except Exception as e:
            logger.error(f"Error calculating FCFF: {e}")
            return []

    def _store_fcf_results_to_var_data(
        self, 
        fcf_type: str, 
        values: List[float], 
        periods: List[str], 
        symbol: str,
        calculation_method: str
    ) -> None:
        """
        Store calculated FCF results back to VarInputData with metadata.
        
        Args:
            fcf_type: Type of FCF ('fcff', 'fcfe', 'levered_fcf')
            values: List of calculated FCF values
            periods: List of corresponding periods
            symbol: Stock symbol
            calculation_method: Description of how values were calculated
        """
        try:
            from core.data_processing.var_input_data import get_var_input_data, VariableMetadata
            from datetime import datetime
            
            var_data = get_var_input_data()
            
            # Store each period's value with metadata
            for period, value in zip(periods, values):
                metadata = VariableMetadata(
                    source="calculated",
                    timestamp=datetime.now(),
                    quality_score=1.0,
                    validation_passed=True,
                    calculation_method=calculation_method,
                    period=period,
                    dependencies=self._get_fcf_dependencies(fcf_type),
                    lineage_id=f"fcf_{fcf_type}_{symbol}_{period}"
                )
                
                success = var_data.set_variable(
                    symbol=symbol,
                    variable_name=fcf_type,
                    value=value,
                    period=period,
                    source="calculated", 
                    metadata=metadata,
                    validate=True,
                    emit_event=True
                )
                
                if success:
                    logger.debug(f"Stored {fcf_type.upper()}[{period}] = {value:.1f} to VarInputData")
                else:
                    logger.warning(f"Failed to store {fcf_type.upper()}[{period}] to VarInputData")
                    
        except Exception as e:
            logger.error(f"Error storing FCF results to VarInputData: {e}")
    
    def _get_fcf_dependencies(self, fcf_type: str) -> List[str]:
        """Get list of variables that a specific FCF type depends on"""
        dependencies = {
            'fcff': ['ebit', 'tax_expense', 'depreciation_amortization', 'capital_expenditures', 'current_assets', 'current_liabilities'],
            'fcfe': ['net_income', 'depreciation_amortization', 'capital_expenditures', 'current_assets', 'current_liabilities'],
            'levered_fcf': ['net_income', 'depreciation_amortization', 'capital_expenditures', 'current_assets', 'current_liabilities']
        }
        return dependencies.get(fcf_type, [])
    
    def _initialize_var_input_data_integration(self) -> None:
        """Initialize VarInputData integration for the calculator"""
        try:
            from core.data_processing.var_input_data import get_var_input_data
            self._var_input_data = get_var_input_data()
            logger.debug("VarInputData integration initialized for FinancialCalculator")
        except Exception as e:
            logger.warning(f"Could not initialize VarInputData integration: {e}")
            self._var_input_data = None

    def calculate_fcf_to_equity(self, use_var_input_data: bool = True) -> List[float]:
        """
        Calculate Free Cash Flow to Equity (FCFE)
        FCFE = Net Income + Depreciation & Amortization - Change in Working Capital - Capital Expenditures + Net Borrowing

        Note: Uses Net Borrowing (debt issued - debt repaid) instead of total financing cash flow
        to exclude equity transactions and dividend payments from the calculation.
        
        Args:
            use_var_input_data: Whether to use VarInputData system for inputs/outputs (default: True)
        """
        try:
            if use_var_input_data:
                # Try to initialize VarInputData if not already done
                if not hasattr(self, '_var_input_data') or self._var_input_data is None:
                    self._initialize_var_input_data_integration()
                
                if self._var_input_data is not None:
                    result = self._calculate_fcf_to_equity_with_var_data()
                    if result:
                        return result
                    # VarInputData returned empty — fall through to legacy

            # Fallback to legacy method for backward compatibility
            return self._calculate_fcf_to_equity_legacy()
            
        except Exception as e:
            logger.error(f"Error calculating FCFE: {e}")
            return []
    
    def _calculate_fcf_to_equity_with_var_data(self) -> List[float]:
        """Calculate FCFE using VarInputData system"""
        try:
            from core.data_processing.var_input_data import get_var_input_data
            var_data = get_var_input_data()
            
            # Get company symbol for data retrieval
            symbol = self.ticker_symbol or "UNKNOWN"
            
            # Get required financial variables from var_input_data
            net_income_data = var_data.get_historical_data(symbol, "net_income", years=10)
            da_data = var_data.get_historical_data(symbol, "depreciation_amortization", years=10)
            capex_data = var_data.get_historical_data(symbol, "capital_expenditures", years=10)
            
            # Get working capital components for change calculation  
            current_assets_data = var_data.get_historical_data(symbol, "current_assets", years=10)
            current_liabilities_data = var_data.get_historical_data(symbol, "current_liabilities", years=10)
            
            # Get debt data for net borrowing calculation
            total_debt_data = var_data.get_historical_data(symbol, "total_debt", years=10)
            
            if not net_income_data:
                logger.warning(f"No Net Income data available for {symbol} in VarInputData")
                return []
            
            # Convert historical data to period-indexed dictionaries
            net_income_by_period = {period: value for period, value in net_income_data}
            da_by_period = {period: value for period, value in da_data}
            capex_by_period = {period: value for period, value in capex_data}
            current_assets_by_period = {period: value for period, value in current_assets_data}
            current_liab_by_period = {period: value for period, value in current_liabilities_data}
            total_debt_by_period = {period: value for period, value in total_debt_data}
            
            # Get common periods (sorted most recent first)
            common_periods = sorted(set(net_income_by_period.keys()), reverse=True)
            
            fcfe_values = []
            periods_calculated = []
            
            # Calculate working capital and net borrowing changes, then FCFE for each period
            previous_working_capital = None
            previous_total_debt = None
            
            for period in reversed(common_periods):  # Calculate chronologically for changes
                try:
                    # Get values for this period (with defaults)
                    net_income = net_income_by_period.get(period, 0)
                    da = da_by_period.get(period, 0)
                    capex = capex_by_period.get(period, 0)
                    
                    current_assets = current_assets_by_period.get(period, 0)
                    current_liabilities = current_liab_by_period.get(period, 0)
                    total_debt = total_debt_by_period.get(period, 0)
                    
                    # Calculate working capital change
                    working_capital = current_assets - current_liabilities
                    working_capital_change = 0
                    
                    if previous_working_capital is not None:
                        working_capital_change = working_capital - previous_working_capital
                    
                    previous_working_capital = working_capital
                    
                    # Calculate net borrowing (change in total debt)
                    net_borrowing = 0
                    
                    if previous_total_debt is not None:
                        net_borrowing = total_debt - previous_total_debt
                    
                    previous_total_debt = total_debt
                    
                    # Calculate FCFE = Net Income + DA - WC Change - CapEx + Net Borrowing
                    fcfe = net_income + da - working_capital_change - abs(capex) + net_borrowing
                    
                    fcfe_values.append(fcfe)
                    periods_calculated.append(period)
                    
                    logger.debug(f"FCFE[{period}]: NetIncome({net_income:.1f}) + DA({da:.1f}) - WC_Change({working_capital_change:.1f}) - CapEx({abs(capex):.1f}) + NetBorrowing({net_borrowing:.1f}) = {fcfe:.1f}")
                    
                except Exception as e:
                    logger.warning(f"Error calculating FCFE for period {period}: {e}")
                    continue
            
            # Reverse to get most recent first (matches legacy behavior)
            fcfe_values.reverse()
            periods_calculated.reverse()
            
            # Store results back to VarInputData with metadata
            self._store_fcf_results_to_var_data("fcfe", fcfe_values, periods_calculated, symbol,
                                               calculation_method="Net Income + DA - WC Change - CapEx + Net Borrowing")
            
            # Store in legacy results for backward compatibility
            self.fcf_results['FCFE'] = fcfe_values
            
            logger.info(f"FCFE calculated using VarInputData: {len(fcfe_values)} periods for {symbol}")
            return fcfe_values
            
        except Exception as e:
            logger.error(f"Error in VarInputData FCFE calculation: {e}")
            # Fallback to legacy method
            return self._calculate_fcf_to_equity_legacy()
    
    def _calculate_fcf_to_equity_legacy(self) -> List[float]:
        """Legacy FCFE calculation method (for backward compatibility)"""
        try:
            # Get pre-calculated metrics
            metrics = self._calculate_all_metrics()
            if not metrics:
                logger.warning("No metrics available for FCFE calculation")
                return []

            # Extract required metrics
            net_income_values = metrics.get('net_income', [])
            da_values = metrics.get('depreciation_amortization', [])
            capex_values = metrics.get('capex', [])
            net_borrowing_values = metrics.get('net_borrowing', [])
            working_capital_changes = metrics.get('working_capital_changes', [])

            # Handle missing data - use zeros if critical data is missing but other data exists
            max_length = max(
                len(net_income_values),
                len(da_values),
                len(capex_values),
                len(net_borrowing_values),
                len(working_capital_changes),
            )

            # Pad missing arrays with zeros (special case handling)
            if len(capex_values) == 0 and max_length > 0:
                logger.info(
                    "CapEx data missing for FCFE - using zeros (typical for REITs/Financial companies)"
                )
                capex_values = [0.0] * max_length

            if len(da_values) == 0 and max_length > 0:
                logger.info("Depreciation & Amortization data missing for FCFE - using zeros")
                da_values = [0.0] * max_length

            # Calculate FCFE - Use minimum of available data
            fcfe_values = []
            min_length = min(
                len(net_income_values),
                len(da_values),
                len(capex_values),
                len(net_borrowing_values),
                len(working_capital_changes),
            )

            for i in range(min_length):
                # Use consistent indexing for all components
                fcfe = (
                    net_income_values[i]
                    + da_values[i]
                    - working_capital_changes[i]
                    - abs(capex_values[i])
                    + net_borrowing_values[i]
                )
                fcfe_values.append(fcfe)

            # Apply scale factor to maintain values in millions
            scaled_fcfe_values = [value * self.financial_scale_factor for value in fcfe_values]
            self.fcf_results['FCFE'] = scaled_fcfe_values
            
            # Track dates for this calculation (legacy)
            self._track_calculation_dates('FCFE', scaled_fcfe_values)
            
            # Enhanced date correlation tracking
            self._create_enhanced_fcf_correlation('FCFE', scaled_fcfe_values)
            
            logger.info(
                f"FCFE calculated: {len(fcfe_values)} years using Net Borrowing (kept in millions for DCF compatibility)"
            )
            return scaled_fcfe_values

        except Exception as e:
            logger.error(f"Error calculating FCFE: {e}")
            return []

    def calculate_levered_fcf(self, use_var_input_data: bool = True) -> List[float]:
        """
        Calculate Levered Free Cash Flow (LFCF)
        LFCF = Cash from Operations - Capital Expenditures
        
        Args:
            use_var_input_data: Whether to use VarInputData system for inputs/outputs (default: True)
        """
        try:
            # Get pre-calculated metrics
            metrics = self._calculate_all_metrics()
            if not metrics:
                logger.warning("No metrics available for LFCF calculation")
                return []

            # Extract required metrics
            operating_cash_flow_values = metrics.get('operating_cash_flow', [])
            capex_values = metrics.get('capex', [])

            # Handle missing CapEx data (special case for REITs/Financial companies)
            if len(capex_values) == 0 and len(operating_cash_flow_values) > 0:
                logger.info(
                    "CapEx data missing for LFCF - using zeros (typical for REITs/Financial companies)"
                )
                capex_values = [0.0] * len(operating_cash_flow_values)

            # Calculate LFCF
            lfcf_values = []
            min_length = min(len(operating_cash_flow_values), len(capex_values))

            for i in range(min_length):
                lfcf = operating_cash_flow_values[i] - abs(capex_values[i])
                lfcf_values.append(lfcf)

            # Apply scale factor to maintain values in millions
            scaled_lfcf_values = [value * self.financial_scale_factor for value in lfcf_values]
            self.fcf_results['LFCF'] = scaled_lfcf_values
            
            # Track dates for this calculation (legacy)
            self._track_calculation_dates('LFCF', scaled_lfcf_values)
            
            # Enhanced date correlation tracking
            self._create_enhanced_fcf_correlation('LFCF', scaled_lfcf_values)
            
            logger.info(
                f"LFCF calculated: {len(lfcf_values)} years (kept in millions for DCF compatibility)"
            )
            return scaled_lfcf_values

        except Exception as e:
            logger.error(f"Error calculating LFCF: {e}")
            return []

    def calculate_all_fcf_types(self) -> Dict[str, List[float]]:
        """
        Calculate all three FCF types and store results.
        This method now efficiently calculates all metrics once and then computes FCF values.
        Also automatically fetches market data if ticker is available.
        Enhanced with comprehensive validation and reporting.
        """
        try:
            logger.info("Starting comprehensive FCF calculation with validation...")

            # Load financial statements first
            self.load_financial_statements()

            # Calculate all metrics once (this will be cached)
            metrics = self._calculate_all_metrics()
            if not metrics:
                err = self._last_calculation_error or "Unknown — check logs"
                logger.error(
                    f"Failed to calculate financial metrics. Last error: {err}"
                )
                self._last_calculation_error = (
                    self._last_calculation_error
                    or "Financial metrics returned empty — income/balance/cashflow data may be missing or malformed"
                )
                return {}

            # Calculate all FCF types using pre-calculated metrics
            fcff_result = self.calculate_fcf_to_firm()
            fcfe_result = self.calculate_fcf_to_equity()
            lfcf_result = self.calculate_levered_fcf()

            # Validate FCF calculation results
            if self.validation_enabled:
                self._validate_fcf_results(
                    {'FCFF': fcff_result, 'FCFE': fcfe_result, 'LFCF': lfcf_result}
                )

            # Automatically fetch market data if ticker was extracted
            if self.ticker_symbol and not self.current_stock_price:
                logger.info(f"Automatically fetching market data for {self.ticker_symbol}...")
                market_data = self.fetch_market_data()
                if market_data:
                    logger.info(
                        f"Successfully fetched market data: Current Price=${self.current_stock_price:.2f}"
                    )
                else:
                    logger.warning(f"Could not fetch market data for {self.ticker_symbol}")

            # Generate final data quality summary
            if self.validation_enabled:
                logger.info("\nFINAL DATA QUALITY SUMMARY:")
                quality_report = self.get_data_quality_report()
                print(quality_report.get_summary())

            logger.info("All FCF calculations completed efficiently")
            return self.fcf_results

        except Exception as e:
            import traceback as _tb
            error_msg = f"Critical error in FCF calculations: {e}"
            detail = _tb.format_exc()
            logger.error(error_msg + "\n" + detail)
            self._last_calculation_error = f"{error_msg} | {detail}"
            if self.validation_enabled:
                self.data_validator.report.add_error(error_msg, "FCF calculation process")
            return {}

    def get_comprehensive_fcf_results(self) -> ComprehensiveFCFResults:
        """
        Get enhanced FCF results with complete date correlation
        
        Returns:
            ComprehensiveFCFResults: Enhanced FCF results with date correlation
        """
        # Validate all correlations before returning
        self.comprehensive_fcf_results.validate_all_correlations()
        return self.comprehensive_fcf_results

    def get_fcf_results_with_dates(self) -> Dict[str, Any]:
        """
        Get FCF results in legacy format but with enhanced date information
        
        Returns:
            Dict[str, Any]: FCF results with date correlation in legacy format
        """
        return self.comprehensive_fcf_results.to_legacy_format()

    def get_date_correlation_summary(self) -> Dict[str, Any]:
        """
        Get summary of date correlations for all FCF calculations
        
        Returns:
            Dict[str, Any]: Date correlation summary
        """
        return self.comprehensive_fcf_results.get_correlated_dates_summary()

    def get_fcf_with_dates(self, fcf_type: str) -> Dict[str, Any]:
        """
        Get specific FCF type results with date correlation
        
        Args:
            fcf_type (str): FCF type ('FCFF', 'FCFE', 'LFCF')
            
        Returns:
            Dict[str, Any]: FCF results with dates, values, and sources
        """
        fcf_results = self.comprehensive_fcf_results.get_fcf_results(fcf_type)
        if fcf_results:
            return fcf_results.to_dict()
        else:
            # Fallback to legacy format
            return {
                'values': self.fcf_results.get(fcf_type, []),
                'dates': ['Unknown'] * len(self.fcf_results.get(fcf_type, [])),
                'sources': ['Unknown'] * len(self.fcf_results.get(fcf_type, [])),
                'source_files': ['Unknown'] * len(self.fcf_results.get(fcf_type, [])),
                'metadata': {'correlation_validated': False}
            }

    def _validate_fcf_results(self, fcf_results: Dict[str, List[float]]) -> None:
        """
        Validate the calculated FCF results for reasonableness

        Args:
            fcf_results (dict): Dictionary of FCF calculation results
        """
        logger.info("Validating FCF calculation results...")

        for fcf_type, values in fcf_results.items():
            if not values:
                # Provide specific guidance on missing data
                if not hasattr(self, 'metrics') or not self.metrics:
                    error_msg = f"No {fcf_type} values calculated - missing financial metrics"
                else:
                    # Check for critical missing data that prevents FCF calculation
                    missing_critical = []
                    critical_fields = ['capex', 'depreciation_amortization', 'operating_cash_flow']
                    for field in critical_fields:
                        if field not in self.metrics or not self.metrics.get(field):
                            missing_critical.append(field.replace('_', ' ').title())

                    if missing_critical:
                        error_msg = (
                            f"No {fcf_type} values calculated due to missing critical data: {', '.join(missing_critical)}. "
                            f"Please verify these fields exist in your Excel files with valid numeric data."
                        )
                        logger.error(f"Required for FCF calculation:")
                        if 'capex' in [f.lower().replace(' ', '_') for f in missing_critical]:
                            logger.error("  • Capital Expenditure in Cash Flow Statement")
                        if 'depreciation_amortization' in [
                            f.lower().replace(' ', '_') for f in missing_critical
                        ]:
                            logger.error("  • Depreciation & Amortization in Cash Flow Statement")
                        if 'operating_cash_flow' in [
                            f.lower().replace(' ', '_') for f in missing_critical
                        ]:
                            logger.error("  • Cash from Operations in Cash Flow Statement")
                    else:
                        error_msg = f"No {fcf_type} values calculated - check calculation logic"

                self.data_validator.report.add_error(error_msg)
                continue

            # Check for extreme values
            max_val = max(abs(v) for v in values)
            if max_val > 1e12:  # > 1 trillion
                self.data_validator.report.add_warning(
                    f"{fcf_type} contains extreme values (max: {max_val:,.0f})"
                )

            # Check for all negative values
            if all(v < 0 for v in values):
                self.data_validator.report.add_warning(
                    f"{fcf_type} has all negative values - review business fundamentals"
                )

            # Check for unusual volatility
            if len(values) > 1:
                std_dev = np.std(values)
                mean_val = np.mean(values)
                if mean_val != 0 and std_dev / abs(mean_val) > 2:  # High coefficient of variation
                    self.data_validator.report.add_warning(
                        f"{fcf_type} shows high volatility (CV: {std_dev/abs(mean_val):.1f})"
                    )

        logger.info("FCF validation completed")

    def _extract_metric_values(
        self, df: pd.DataFrame, metric_name: str, reverse: bool = False
    ) -> List[float]:
        """
        Extract values for a specific metric from financial statement DataFrame
        Enhanced with comprehensive validation and error handling
        Supports both Excel format (metrics in rows) and yfinance format (metrics in columns)

        Args:
            df (pd.DataFrame): Financial statement data
            metric_name (str): Name of the metric to extract
            reverse (bool): Whether to reverse the order of values

        Returns:
            list: List of metric values
        """
        try:
            # Validate input DataFrame
            if df.empty:
                logger.error(f"Cannot extract '{metric_name}': DataFrame is empty")
                if self.validation_enabled:
                    self.data_validator.report.add_error(
                        f"Empty DataFrame for metric extraction: {metric_name}"
                    )
                return []

            # Detect data format: check if metric name is in columns (yfinance format)
            if metric_name in df.columns:
                logger.debug(f"Found '{metric_name}' as column header (yfinance format)")
                values = []
                for val in df[metric_name]:
                    if pd.isna(val):
                        values.append(0.0)
                    else:
                        try:
                            # Use pandas to_numeric with error handling for robust conversion
                            if isinstance(val, str):
                                val = (
                                    val.replace(',', '')
                                    .replace('$', '')
                                    .replace('(', '-')
                                    .replace(')', '')
                                )
                            numeric_val = pd.to_numeric(val, errors='coerce')
                            values.append(numeric_val if pd.notna(numeric_val) else 0.0)
                        except (ValueError, TypeError):
                            values.append(0.0)

                logger.debug(
                    f"Extracted {len(values)} values for '{metric_name}' from column format"
                )
                return list(reversed(values)) if reverse else values

            # Fall back to Excel format search (metrics in rows)
            logger.debug(f"Searching for '{metric_name}' in row format (Excel format)")

            # Find row containing the metric with enhanced search
            metric_row = None
            available_metrics = []
            best_match_score = 0
            best_match_row = None

            for idx, row in df.iterrows():
                # Check multiple columns for metric names (more flexible matching)
                metric_text = None
                for col_idx in [0, 1, 2]:  # Check first 3 columns
                    if len(row) > col_idx and pd.notna(row.iloc[col_idx]):
                        metric_text = str(row.iloc[col_idx]).strip()
                        break

                if metric_text:
                    available_metrics.append(metric_text)

                    # Exact match (best case)
                    if metric_name.lower() == metric_text.lower():
                        metric_row = row
                        logger.debug(f"Exact match found for '{metric_name}' at row {idx}")
                        break

                    # Partial match with scoring
                    if metric_name.lower() in metric_text.lower():
                        # Calculate match score based on how much of the target is in the text
                        match_score = len(metric_name) / len(metric_text)
                        if match_score > best_match_score:
                            best_match_score = match_score
                            best_match_row = row

            # Use best match if no exact match found
            if metric_row is None and best_match_row is not None:
                metric_row = best_match_row
                logger.info(
                    f"Using best match for '{metric_name}' with score {best_match_score:.2f}"
                )

            if metric_row is None:
                error_msg = f"Metric '{metric_name}' not found in financial data"
                logger.warning(error_msg)

                # Enhanced debugging information
                logger.debug(f"Searched in {len(available_metrics)} available metrics")
                if available_metrics:
                    # Show closest matches for debugging
                    closest_matches = [
                        m
                        for m in available_metrics
                        if any(word in m.lower() for word in metric_name.lower().split())
                    ]
                    if closest_matches:
                        logger.info(f"Possible similar metrics: {closest_matches[:5]}")
                    else:
                        logger.info(f"Available metrics (first 10): {available_metrics[:10]}")

                if self.validation_enabled:
                    self.data_validator.report.add_error(
                        error_msg, f"DataFrame with {len(available_metrics)} metrics"
                    )
                return []

            # Extract numeric values with enhanced validation
            values = []
            empty_count = 0
            invalid_count = 0

            # Skip the first 3 columns which contain metadata
            for col_idx, val in enumerate(metric_row.iloc[3:], start=3):
                context = f"{metric_name}.Column{col_idx}"

                if pd.isna(val) or val == '':
                    empty_count += 1
                    if self.validation_enabled:
                        validated_val, is_valid = self.data_validator.validate_cell_value(
                            val, float, True, context
                        )
                        values.append(validated_val)
                    else:
                        values.append(0)
                else:
                    try:
                        # Enhanced numeric conversion with modern pandas error handling
                        if self.validation_enabled:
                            validated_val, is_valid = self.data_validator.validate_cell_value(
                                val, float, True, context
                            )
                            values.append(validated_val)
                            if not is_valid:
                                invalid_count += 1
                        else:
                            # Use pandas to_numeric for robust conversion with error handling
                            if isinstance(val, str):
                                clean_val = val.replace(',', '').replace('(', '-').replace(')', '')
                                numeric_val = pd.to_numeric(clean_val, errors='coerce')
                            else:
                                numeric_val = pd.to_numeric(val, errors='coerce')

                            values.append(numeric_val if pd.notna(numeric_val) else 0.0)
                    except (ValueError, TypeError) as e:
                        invalid_count += 1
                        logger.warning(f"Invalid value in {context}: {val} -> {e}")
                        values.append(0)

            # Log data quality information
            if empty_count > 0:
                logger.info(f"'{metric_name}': {empty_count} empty values converted to 0")
            if invalid_count > 0:
                logger.warning(f"'{metric_name}': {invalid_count} invalid values converted to 0")

            # Validate the extracted series
            if self.validation_enabled and values:
                self.data_validator.validate_data_series(values, metric_name)

            # Remove trailing zeros that might represent missing future periods
            while values and values[-1] == 0:
                values.pop()

            if reverse:
                values = values[::-1]

            # Final validation
            if not values:
                logger.warning(f"No valid data extracted for '{metric_name}'")
                if self.validation_enabled:
                    self.data_validator.report.add_warning(
                        f"No valid data extracted for {metric_name}"
                    )
            else:
                logger.debug(f"Successfully extracted {len(values)} values for '{metric_name}'")

            return values

        except Exception as e:
            error_msg = f"Critical error extracting metric '{metric_name}': {e}"
            logger.error(error_msg)
            if self.validation_enabled:
                self.data_validator.report.add_error(error_msg, "Metric extraction function")
            return []

    @lru_cache(maxsize=256)
    def calculate_growth_rates(
        self, values: tuple, periods: tuple = (1, 3, 5, 10)
    ) -> Dict[str, float]:
        """
        Calculate annualized growth rates for different periods

        Args:
            values (list): List of values to calculate growth rates for
            periods (list): List of periods to calculate growth rates for

        Returns:
            dict: Dictionary of growth rates by period
        """
        growth_rates = {}

        try:
            for period in periods:
                if len(values) >= period + 1:
                    start_value = values[-(period + 1)]
                    end_value = values[-1]

                    if start_value != 0:
                        # Calculate annualized growth rate: (end/start)^(1/period) - 1
                        growth_rate = (abs(end_value) / abs(start_value)) ** (1 / period) - 1

                        # Handle negative cash flows
                        if end_value < 0 and start_value > 0:
                            growth_rate = -growth_rate
                        elif end_value > 0 and start_value < 0:
                            growth_rate = abs(growth_rate)

                        growth_rates[f"{period}Y"] = growth_rate
                    else:
                        growth_rates[f"{period}Y"] = 0

        except Exception as e:
            logger.error(f"Error calculating growth rates: {e}")

        return growth_rates

    def _auto_extract_ticker(self) -> None:
        """
        Automatically extract ticker symbol from company folder and financial data
        """
        if not self.company_folder:
            return

        # Pre-detect if this is likely a TASE stock
        is_likely_tase = self._detect_tase_stock_from_context()

        # Method 1: Extract from folder name (most reliable)
        folder_name = os.path.basename(self.company_folder.rstrip(os.sep))

        # Check if folder name looks like a ticker (1-5 uppercase letters)
        if re.match(r'^[A-Z]{1,5}$', folder_name):
            self.ticker_symbol = self._apply_tase_suffix_if_needed(folder_name, is_likely_tase)
            logger.info(f"Auto-extracted ticker from folder name: {self.ticker_symbol}")
            return

        # Method 2: Try to infer from company name using mappings
        inferred_ticker = self._infer_ticker_from_company_name()
        if inferred_ticker:
            self.ticker_symbol = self._apply_tase_suffix_if_needed(inferred_ticker, is_likely_tase)
            logger.info(f"Auto-extracted ticker from company name mapping: {self.ticker_symbol}")
            return

        # Method 3: Pattern-based extraction from folder name
        pattern_ticker = self._extract_ticker_from_pattern(folder_name)
        if pattern_ticker:
            self.ticker_symbol = self._apply_tase_suffix_if_needed(pattern_ticker, is_likely_tase)
            logger.info(f"Auto-extracted ticker from pattern analysis: {self.ticker_symbol}")
            return

        logger.warning(f"Could not auto-extract ticker for company: {self.company_name}")

    def _apply_tase_suffix_if_needed(self, ticker: str, is_likely_tase: bool) -> str:
        """
        Apply .TA suffix to ticker if it's likely a TASE stock and doesn't already have it

        Args:
            ticker (str): Base ticker symbol
            is_likely_tase (bool): Whether this is likely a TASE stock

        Returns:
            str: Ticker with .TA suffix if needed
        """
        if not ticker:
            return ticker

        # If already has .TA suffix, return as-is
        if ticker.endswith('.TA'):
            if is_likely_tase:
                logger.info(f"Ticker {ticker} already has .TA suffix for TASE stock")
            return ticker

        # If likely TASE stock, append .TA suffix
        if is_likely_tase:
            tase_ticker = f"{ticker}.TA"
            logger.info(f"Applied TASE suffix: {ticker} → {tase_ticker}")
            return tase_ticker

        return ticker

    def _detect_tase_stock_from_context(self) -> bool:
        """
        Pre-detect if this is likely a TASE stock based on company context

        Returns:
            bool: True if likely TASE stock, False otherwise
        """
        # Method 1: Check financial data for Israeli currency indicators
        if self._check_financial_data_for_israeli_indicators():
            logger.info("TASE stock detected via financial data currency indicators")
            return True

        # Method 2: Check company name for Israeli patterns
        if self._check_company_name_for_israeli_patterns():
            logger.info("TASE stock detected via company name patterns")
            return True

        # Method 3: Check folder structure for Israeli context
        if self._check_folder_context_for_israeli_indicators():
            logger.info("TASE stock detected via folder context")
            return True

        return False

    def _check_financial_data_for_israeli_indicators(self) -> bool:
        """
        Check loaded financial data for Israeli currency indicators

        Returns:
            bool: True if Israeli indicators found
        """
        try:
            # Check if we have any financial data loaded
            if not hasattr(self, 'financial_data') or not self.financial_data:
                return False

            # Look through financial data for currency indicators
            for period_type in ['FY', 'LTM']:
                if period_type in self.financial_data:
                    for statement_type in ['income', 'balance_sheet', 'cash_flow']:
                        if statement_type in self.financial_data[period_type]:
                            data = self.financial_data[period_type][statement_type]

                            # Convert to string and check for Israeli currency indicators
                            data_str = str(data).lower()
                            if any(
                                indicator in data_str
                                for indicator in ['shekel', 'ils', 'nis', '₪', 'israeli']
                            ):
                                return True

            return False

        except Exception as e:
            logger.debug(f"Error checking financial data for Israeli indicators: {e}")
            return False

    def _check_company_name_for_israeli_patterns(self) -> bool:
        """
        Check company name for Israeli business patterns

        Returns:
            bool: True if Israeli patterns found
        """
        if not self.company_name:
            return False

        company_lower = self.company_name.lower()

        # Israeli company name patterns and suffixes
        israeli_patterns = [
            'israel',
            'israeli',
            'tel aviv',
            'jerusalem',
            'haifa',
            'ltd.',
            '(israel)',
            'technologies israel',
            'israel corp',
            'telit',
            'teva',
            'check point',
            'nice',
            'elbit',
            'rafael',
        ]

        # Hebrew characters (if present)
        hebrew_chars = any('\u0590' <= char <= '\u05ff' for char in self.company_name)

        return hebrew_chars or any(pattern in company_lower for pattern in israeli_patterns)

    def _check_folder_context_for_israeli_indicators(self) -> bool:
        """
        Check folder path and structure for Israeli context

        Returns:
            bool: True if Israeli context found
        """
        if not self.company_folder:
            return False

        folder_path_lower = self.company_folder.lower()

        # Israeli context indicators in folder path
        israeli_context = [
            'israel',
            'israeli',
            'tase',
            'tel_aviv',
            'telaviv',
            'middle_east',
            'hebrew',
            'shekel',
            'ils',
        ]

        return any(context in folder_path_lower for context in israeli_context)

    def _infer_ticker_from_company_name(self) -> Optional[str]:
        """
        Infer ticker symbol from company name using dynamic pattern extraction
        """
        # Use pattern-based extraction instead of hardcoded mappings
        # This makes the system work with any company without maintenance
        return self._extract_ticker_from_pattern(self.company_name)

    def _extract_ticker_from_pattern(self, folder_name: str) -> Optional[str]:
        """
        Extract ticker using pattern analysis
        """
        # If folder name has numbers or special chars, might be a company code
        # Try to find ticker-like patterns within it

        # Look for uppercase letter sequences
        matches = re.findall(r'[A-Z]{1,5}', folder_name.upper())

        # Filter out common non-ticker patterns
        excluded_patterns = {'INC', 'CORP', 'LLC', 'LTD', 'CO', 'CLASS'}

        for match in matches:
            if match not in excluded_patterns and len(match) <= 5:
                return match

        # If folder name is mixed case/has numbers, try extracting letters only
        letters_only = re.sub(r'[^A-Za-z]', '', folder_name).upper()
        if 1 <= len(letters_only) <= 5:
            return letters_only

        return None

    def fetch_market_data(
        self, ticker_symbol: Optional[str] = None, force_reload: bool = False
    ) -> Optional[Dict[str, Any]]:
        """
        Fetch current market data using enhanced data manager with multiple sources fallback

        Args:
            ticker_symbol (str): Stock ticker symbol
            force_reload (bool): Force reload even if cached data exists

        Returns:
            dict: Market data or None if failed
        """
        if ticker_symbol:
            self.ticker_symbol = ticker_symbol.upper()

        if not self.ticker_symbol:
            logger.warning("No ticker symbol available for market data fetch")
            return None

        # Try enhanced data manager first if available
        if self.enhanced_data_manager:
            market_data = self._fetch_from_enhanced_manager(force_reload)
            if market_data:
                return market_data

        # Fallback to original yfinance implementation
        logger.info(f"Fetching market data for {self.ticker_symbol} using yfinance fallback")

        try:
            import yfinance as yf
            import time
            import requests

            # Note: yfinance v0.2.65+ handles session management internally
            # No longer need custom session configuration

            # Add retry logic for rate limiting issues with longer delays
            max_retries = 5
            retry_delay = 3  # Start with 3 seconds

            for attempt in range(max_retries):
                try:
                    # Configure yfinance - let YF handle session internally (v0.2.65+ requirement)
                    ticker = yf.Ticker(self.ticker_symbol)
                    info = ticker.info
                    break
                except Exception as e:
                    error_str = str(e).lower()
                    if (
                        "429" in error_str or "rate limit" in error_str
                    ) and attempt < max_retries - 1:
                        logger.warning(
                            f"Rate limited, retrying in {retry_delay} seconds... (attempt {attempt + 1})"
                        )
                        time.sleep(retry_delay)
                        retry_delay *= 1.5  # More gradual exponential backoff
                        continue
                    elif (
                        "timeout" in error_str or "timed out" in error_str
                    ) and attempt < max_retries - 1:
                        logger.warning(
                            f"Request timeout, retrying in {retry_delay} seconds... (attempt {attempt + 1})"
                        )
                        time.sleep(retry_delay)
                        retry_delay *= 1.5
                        continue
                    else:
                        raise e

            # Get company name
            company_name = (
                info.get('longName')
                or info.get('shortName')
                or info.get('longBusinessSummary', '').split('.')[0]
            )
            if company_name:
                self.company_name = company_name
            else:
                # Fallback to ticker symbol if no name found
                self.company_name = self.ticker_symbol

            # Get currency information for proper handling
            currency = info.get('currency', 'USD')
            financial_currency = info.get('financialCurrency', currency)

            # Detect TASE (Tel Aviv Stock Exchange) stocks
            is_tase_stock = (
                self.ticker_symbol.endswith('.TA')
                or currency in ['ILS', 'ILA']
                or financial_currency in ['ILS', 'ILA']
            )

            # Get current price
            current_price = info.get('currentPrice') or info.get('regularMarketPrice')
            if not current_price:
                # Try to get from recent history
                hist = ticker.history(period='1d')
                if not hist.empty:
                    current_price = hist['Close'].iloc[-1]

            # For TASE stocks, log currency handling information
            if is_tase_stock:
                logger.info(f"TASE stock detected: {self.ticker_symbol}")
                logger.info(f"Price currency: {currency}, Financial currency: {financial_currency}")
                logger.info(f"Raw price: {current_price} (likely in Agorot for TASE stocks)")

            # Get shares outstanding and market cap with multiple fallback options
            shares_outstanding = (
                info.get('sharesOutstanding')
                or info.get('impliedSharesOutstanding')
                or info.get('basicAvgSharesOutstanding')
                or info.get('weightedAvgSharesOutstanding')
                or 0
            )
            market_cap = (
                info.get('marketCap') or info.get('enterpriseValue') or info.get('impliedMarketCap')
            )

            # Store currency information even if price is not available
            self.currency = currency
            self.financial_currency = financial_currency
            self.is_tase_stock = is_tase_stock

            if current_price:
                self.current_stock_price = float(current_price)

                # Handle shares outstanding - either direct or calculated from market cap
                if shares_outstanding and shares_outstanding > 0:
                    self.shares_outstanding = shares_outstanding
                    self.market_cap = market_cap or (current_price * shares_outstanding)
                    logger.info(
                        f"Fetched market data for {self.ticker_symbol}: "
                        f"Price=${self.current_stock_price:.2f}, "
                        f"Shares={self.shares_outstanding/1000000:.1f}M (direct)"
                    )
                elif market_cap and market_cap > 0:
                    # Calculate shares outstanding from market cap
                    self.shares_outstanding = market_cap / current_price
                    self.market_cap = market_cap
                    logger.info(
                        f"Fetched market data for {self.ticker_symbol}: "
                        f"Price=${self.current_stock_price:.2f}, "
                        f"Shares={self.shares_outstanding/1000000:.1f}M (calculated from market cap)"
                    )
                else:
                    # Cannot determine shares outstanding - provide manual entry guidance
                    self.shares_outstanding = 0
                    self.market_cap = 0
                    logger.warning(
                        f"Cannot determine shares outstanding for {self.ticker_symbol}: "
                        f"sharesOutstanding={shares_outstanding}, marketCap={market_cap}"
                    )
                    logger.info(
                        f"Manual entry required: Visit https://finance.yahoo.com/quote/{self.ticker_symbol}/key-statistics/ "
                        f"to find shares outstanding data"
                    )

                return {
                    'current_price': self.current_stock_price,
                    'shares_outstanding': self.shares_outstanding,
                    'market_cap': self.market_cap,
                    'ticker_symbol': self.ticker_symbol,
                    'company_name': self.company_name,
                    'currency': currency,
                    'financial_currency': financial_currency,
                    'is_tase_stock': is_tase_stock,
                }
            else:
                logger.warning(f"Could not fetch current price for {self.ticker_symbol}")
                return None

        except (ConnectionError, TimeoutError, requests.RequestException) as e:
            logger.warning(f"Network error fetching market data for {self.ticker_symbol}: {e}")
            # Try TASE fallback for network errors too
            if self._should_retry_with_tase_suffix(e):
                logger.info(f"Attempting TASE fallback for network error: {self.ticker_symbol}.TA")
                return self._retry_fetch_with_tase_suffix()
            return self._create_fallback_market_data("Network connectivity issues")
        except Exception as e:
            error_str = str(e).lower()
            logger.error(
                f"Error fetching market data for {self.ticker_symbol}: {e}",
                context={'ticker': self.ticker_symbol},
                error=e,
            )

            # Check if this might be a TASE stock without .TA suffix
            if self._should_retry_with_tase_suffix(e):
                logger.warning(
                    f"Attempting fallback: trying {self.ticker_symbol}.TA for potential TASE stock"
                )
                return self._retry_fetch_with_tase_suffix()

            # Create informative error message with graceful degradation
            fallback_reason = "Unknown error"
            if "429" in error_str:
                fallback_reason = "Rate limiting - Yahoo Finance temporarily unavailable"
            elif "404" in error_str or "delisted" in error_str or "no data found" in error_str:
                fallback_reason = "Ticker not found or delisted"
            elif "timeout" in error_str:
                fallback_reason = "Request timeout"

            if self.validation_enabled:
                error_msg = f"Could not fetch market data: {fallback_reason}"
                self.data_validator.report.add_warning(error_msg)

            return self._create_fallback_market_data(fallback_reason)

    def _should_retry_with_tase_suffix(self, error: Exception) -> bool:
        """
        Check if we should retry with .TA suffix based on the error

        Args:
            error: The exception that occurred

        Returns:
            bool: True if should retry with .TA suffix
        """
        if not self.ticker_symbol or self.ticker_symbol.endswith('.TA'):
            return False

        error_str = str(error).lower()

        # Check for indicators that this might be a delisted/missing ticker
        error_indicators = [
            '404',
            'delisted',
            'no data found',
            'symbol may be delisted',
            'not found',
            'invalid symbol',
        ]

        has_error_indicator = any(indicator in error_str for indicator in error_indicators)

        # Only retry if we have a delisting/missing ticker error
        # AND we can detect some Israeli context
        if has_error_indicator:
            # Check if we can detect Israeli context even without full data
            israeli_context = self._detect_minimal_israeli_context()
            if israeli_context:
                logger.info(
                    f"Israeli context detected for failed ticker {self.ticker_symbol}, will retry with .TA suffix"
                )
                return True

        return False

    def _detect_minimal_israeli_context(self) -> bool:
        """
        Detect Israeli context with minimal information available

        Returns:
            bool: True if Israeli context detected
        """
        # Check company name for Israeli indicators
        if self.company_name:
            company_lower = self.company_name.lower()
            israeli_patterns = ['israel', 'israeli', 'teva', 'check point', 'nice', 'elbit']
            if any(pattern in company_lower for pattern in israeli_patterns):
                return True

        # Check folder name for Israeli context
        if self.company_folder:
            folder_lower = self.company_folder.lower()
            if any(context in folder_lower for context in ['israel', 'tase', 'hebrew']):
                return True

        return False

    def _retry_fetch_with_tase_suffix(self) -> Optional[Dict[str, Any]]:
        """
        Retry fetching market data with .TA suffix appended

        Returns:
            dict: Market data or None if failed
        """
        original_ticker = self.ticker_symbol
        tase_ticker = f"{original_ticker}.TA"

        try:
            # Temporarily update ticker symbol
            self.ticker_symbol = tase_ticker
            logger.info(f"Retrying market data fetch with TASE ticker: {tase_ticker}")

            # Retry the fetch with .TA suffix
            import yfinance as yf

            ticker = yf.Ticker(tase_ticker)
            info = ticker.info

            # If we get here, the .TA version worked
            logger.info(
                f"✅ Successfully fetched data with TASE suffix: {original_ticker} → {tase_ticker}"
            )

            # Continue with the same processing logic as the main method
            company_name = (
                info.get('longName')
                or info.get('shortName')
                or info.get('longBusinessSummary', '').split('.')[0]
            )
            if company_name:
                self.company_name = company_name
            else:
                self.company_name = tase_ticker

            currency = info.get('currency', 'USD')
            financial_currency = info.get('financialCurrency', currency)

            # Should be TASE stock now
            is_tase_stock = True
            self.currency = currency
            self.financial_currency = financial_currency
            self.is_tase_stock = is_tase_stock

            logger.info(f"TASE stock confirmed via fallback: {tase_ticker}")
            logger.info(f"Price currency: {currency}, Financial currency: {financial_currency}")

            # Get current price
            current_price = info.get('currentPrice') or info.get('regularMarketPrice')
            if not current_price:
                hist = ticker.history(period='1d')
                if not hist.empty:
                    current_price = hist['Close'].iloc[-1]

            if current_price:
                self.current_stock_price = float(current_price)
                logger.info(f"Raw price: {current_price} (likely in Agorot for TASE stocks)")

                # Get shares outstanding
                shares_outstanding = (
                    info.get('sharesOutstanding')
                    or info.get('impliedSharesOutstanding')
                    or info.get('basicAvgSharesOutstanding')
                    or info.get('weightedAvgSharesOutstanding')
                    or 0
                )
                market_cap = (
                    info.get('marketCap')
                    or info.get('enterpriseValue')
                    or info.get('impliedMarketCap')
                )

                if shares_outstanding and shares_outstanding > 0:
                    self.shares_outstanding = shares_outstanding
                    self.market_cap = market_cap or (current_price * shares_outstanding)
                    logger.info(
                        f"Fetched market data for {tase_ticker}: "
                        f"Price=${self.current_stock_price:.2f}, "
                        f"Shares={self.shares_outstanding/1000000:.1f}M (direct)"
                    )
                elif market_cap and market_cap > 0:
                    self.shares_outstanding = market_cap / current_price
                    self.market_cap = market_cap
                    logger.info(
                        f"Fetched market data for {tase_ticker}: "
                        f"Price=${self.current_stock_price:.2f}, "
                        f"Shares={self.shares_outstanding/1000000:.1f}M (calculated from market cap)"
                    )
                else:
                    self.shares_outstanding = 0
                    self.market_cap = 0
                    logger.warning(f"Cannot determine shares outstanding for {tase_ticker}")

                return {
                    'current_price': self.current_stock_price,
                    'shares_outstanding': self.shares_outstanding,
                    'market_cap': self.market_cap,
                    'ticker_symbol': tase_ticker,
                    'company_name': self.company_name,
                    'currency': currency,
                    'financial_currency': financial_currency,
                    'is_tase_stock': is_tase_stock,
                }
            else:
                logger.warning(f"Could not fetch current price for {tase_ticker}")
                # Restore original ticker since fallback also failed
                self.ticker_symbol = original_ticker
                return None

        except Exception as e:
            logger.error(f"Fallback with .TA suffix also failed for {tase_ticker}: {e}")
            # Restore original ticker
            self.ticker_symbol = original_ticker
            return None

    def convert_agorot_to_shekel(self, agorot_value: Optional[float]) -> Optional[float]:
        """
        Convert Agorot (ILA) to Shekel (ILS) for TASE stocks

        Args:
            agorot_value (float): Value in Agorot

        Returns:
            float: Value in Shekels (1 ILS = 100 ILA)
        """
        if agorot_value is None:
            return None
        return agorot_value / 100.0

    def convert_shekel_to_agorot(self, shekel_value: Optional[float]) -> Optional[float]:
        """
        Convert Shekel (ILS) to Agorot (ILA) for TASE stocks

        Args:
            shekel_value (float): Value in Shekels

        Returns:
            float: Value in Agorot (1 ILS = 100 ILA)
        """
        if shekel_value is None:
            return None
        return shekel_value * 100.0

    def get_price_in_shekels(self) -> Optional[float]:
        """
        Get stock price in Shekels for TASE stocks
        For TASE stocks, yfinance typically returns prices in Agorot

        Returns:
            float: Price in Shekels, or original price for non-TASE stocks
        """
        if not self.current_stock_price:
            return None

        if self.is_tase_stock:
            # TASE stock prices from yfinance are typically in Agorot
            return self.convert_agorot_to_shekel(self.current_stock_price)
        else:
            # Non-TASE stocks return price as-is
            return self.current_stock_price

    def get_price_in_agorot(self) -> Optional[float]:
        """
        Get stock price in Agorot for TASE stocks

        Returns:
            float: Price in Agorot for TASE stocks, or None for non-TASE stocks
        """
        if not self.current_stock_price:
            return None

        if self.is_tase_stock:
            # TASE stock prices from yfinance are typically already in Agorot
            return self.current_stock_price
        else:
            return None

    def normalize_financial_values_for_tase(
        self, financial_values_millions_ils: List[float]
    ) -> List[float]:
        """
        Normalize financial statement values for TASE stock calculations
        Financial statements are in millions of ILS, need to convert appropriately

        Args:
            financial_values_millions_ils (list): Financial values in millions of ILS

        Returns:
            list: Normalized values for calculations
        """
        if not self.is_tase_stock or not financial_values_millions_ils:
            return financial_values_millions_ils

        # Financial statements are already in millions of ILS - this is the correct unit
        # for DCF calculations (enterprise/equity value calculations)
        logger.info(
            f"TASE financial values normalized: {len(financial_values_millions_ils)} values in millions ILS"
        )
        return financial_values_millions_ils

    def get_currency_info(self) -> Dict[str, Any]:
        """
        Get currency information for the stock

        Returns:
            dict: Currency information including conversions for TASE stocks
        """
        info = {
            'currency': self.currency,
            'financial_currency': self.financial_currency,
            'is_tase_stock': self.is_tase_stock,
        }

        if self.is_tase_stock and self.current_stock_price:
            info.update(
                {
                    'price_in_agorot': self.get_price_in_agorot(),
                    'price_in_shekels': self.get_price_in_shekels(),
                    'currency_note': 'TASE stocks: Prices in Agorot (ILA), Financials in millions Shekels (ILS)',
                }
            )

        return info

    def _update_from_market_data(self, market_data: Optional[Dict[str, Any]]) -> None:
        """
        Update instance variables from market data dictionary

        Args:
            market_data (dict): Market data from enhanced data manager
        """
        if not market_data:
            return

        # Update core market data
        self.current_stock_price = market_data.get('current_price')
        self.shares_outstanding = market_data.get('shares_outstanding')
        self.market_cap = market_data.get('market_cap')

        # Update company information if available
        if market_data.get('company_name'):
            self.company_name = market_data['company_name']

        # Update currency information
        self.currency = market_data.get('currency', 'USD')
        self.financial_currency = market_data.get('financial_currency', self.currency)
        self.is_tase_stock = market_data.get('is_tase_stock', False)

        logger.debug(
            f"Updated FinancialCalculator with market data: "
            f"Price=${self.current_stock_price}, Shares={self.shares_outstanding}, "
            f"Currency={self.currency}, TASE={self.is_tase_stock}"
        )

    def _create_fallback_market_data(self, reason: str) -> Dict[str, Any]:
        """
        Create fallback market data when API calls fail

        Args:
            reason (str): Reason for fallback

        Returns:
            dict: Minimal market data structure to prevent calculation failures
        """
        logger.warning(f"Creating fallback market data due to: {reason}")

        fallback_data = {
            'current_price': None,
            'shares_outstanding': None,
            'market_cap': None,
            'currency': 'USD',
            'financial_currency': 'USD',
            'is_tase_stock': False,
            'fallback_reason': reason,
            'data_quality': 'degraded',
        }

        # Try to preserve any existing data
        if hasattr(self, 'current_stock_price') and self.current_stock_price:
            fallback_data['current_price'] = self.current_stock_price
            logger.info(f"Preserved existing stock price: ${self.current_stock_price:.2f}")

        if hasattr(self, 'shares_outstanding') and self.shares_outstanding:
            fallback_data['shares_outstanding'] = self.shares_outstanding

        return fallback_data

    @retry_with_exponential_backoff(max_retries=2, base_delay=2.0, max_delay=30.0)
    def _fetch_from_enhanced_manager(self, force_reload: bool = False) -> Optional[Dict[str, Any]]:
        """
        Fetch market data from enhanced data manager with retry logic

        Args:
            force_reload (bool): Force reload even if cached data exists

        Returns:
            dict: Market data or None if failed
        """
        try:
            logger.info(
                f"Fetching market data for {self.ticker_symbol} using enhanced data manager with multiple sources"
            )
            market_data = self.enhanced_data_manager.fetch_market_data(
                self.ticker_symbol, force_reload=force_reload
            )

            if market_data:
                # Update instance variables with enhanced data
                self._update_from_market_data(market_data)
                logger.info(f"Successfully fetched market data using enhanced data manager")
                return market_data
            else:
                logger.warning("Enhanced data manager returned no data, falling back to yfinance")
                return None

        except (ConnectionError, TimeoutError, requests.RequestException) as e:
            # These exceptions will be caught by the retry decorator
            logger.debug(f"Network error with enhanced data manager: {e}")
            raise
        except Exception as e:
            logger.warning(
                f"Enhanced data manager failed with non-retryable error: {e}",
                context={'ticker': self.ticker_symbol},
                error=e,
            )
            return None

    def get_standardized_financial_data(self) -> Dict[str, Any]:
        """
        Get financial data in standardized format for valuation modules

        Returns:
            dict: Standardized financial data structure
        """
        try:
            # Ensure metrics are calculated
            metrics = self._calculate_all_metrics()

            # Create standardized structure
            standardized_data = {
                # Company identification
                'ticker_symbol': self.ticker_symbol,
                'company_name': self.company_name,
                'currency': self.currency,
                'financial_currency': self.financial_currency,
                'is_tase_stock': self.is_tase_stock,
                # Market data
                'current_stock_price': self.current_stock_price,
                'shares_outstanding': self.shares_outstanding,
                'market_cap': self.market_cap,
                # Financial statements (raw data)
                'Income Statement': self.financial_data.get('income_fy', {}),
                'Balance Sheet': self.financial_data.get('balance_fy', {}),
                'Cash Flow Statement': self.financial_data.get('cashflow_fy', {}),
                # LTM data for latest periods
                'Income Statement LTM': self.financial_data.get('income_ltm', {}),
                'Balance Sheet LTM': self.financial_data.get('balance_ltm', {}),
                'Cash Flow Statement LTM': self.financial_data.get('cashflow_ltm', {}),
                # Calculated metrics
                'calculated_metrics': metrics,
                # FCF results if available
                'fcf_results': self.fcf_results,
                # Scaling information
                'financial_scale_factor': self.financial_scale_factor,
                # Data quality
                'data_quality_report': (
                    self.get_data_quality_report() if self.validation_enabled else None
                ),
            }

            logger.debug("Generated standardized financial data structure for valuation modules")
            return standardized_data

        except Exception as e:
            logger.error(f"Error generating standardized financial data: {e}")
            return {}

    def get_enhanced_data_manager(self) -> Optional[Any]:
        """
        Get the enhanced data manager instance

        Returns:
            EnhancedDataManager: Enhanced data manager instance or None
        """
        return self.enhanced_data_manager

    def set_enhanced_data_manager(self, enhanced_data_manager: Any) -> None:
        """
        Set or update the enhanced data manager

        Args:
            enhanced_data_manager (EnhancedDataManager): Enhanced data manager instance
        """
        self.enhanced_data_manager = enhanced_data_manager
        if enhanced_data_manager:
            logger.info("Enhanced data manager set for FinancialCalculator")

    def load_financial_data(self) -> Dict[str, Any]:
        """
        Load Excel financial statements from standardized paths
        
        This method serves as a bridge method that calls the existing load_financial_statements()
        method and returns a structured dictionary of the loaded financial data.
        
        Returns:
            Dict[str, Any]: Dictionary containing loaded financial data organized by statement type
            
        Raises:
            Exception: If financial data cannot be loaded
            
        Example:
            >>> calc = FinancialCalculator('./data/companies/AAPL')
            >>> data = calc.load_financial_data()
            >>> print(data['Income Statement'])
        """
        try:
            # Use existing load_financial_statements method
            self.load_financial_statements()
            
            # Return structured financial data
            financial_data = {
                'Income Statement': self.financial_data.get('Income Statement', {}),
                'Balance Sheet': self.financial_data.get('Balance Sheet', {}), 
                'Cash Flow Statement': self.financial_data.get('Cash Flow Statement', {}),
                'company_name': self.company_name,
                'ticker_symbol': self.ticker_symbol,
                'currency': self.currency,
                'is_tase_stock': self.is_tase_stock,
                'data_loaded': len(self.financial_data) > 0
            }
            
            logger.info(f"Financial data loaded successfully for {self.company_name}")
            return financial_data
            
        except Exception as e:
            logger.error(f"Error loading financial data: {e}")
            raise Exception(f"Failed to load financial data: {e}")

    def calculate_dcf_inputs(self) -> Dict[str, Any]:
        """
        Extract and prepare DCF calculation inputs from financial data
        
        This method processes the loaded financial data and calculates all necessary
        inputs for DCF valuation including FCF values, growth rates, and key metrics.
        
        Returns:
            Dict[str, Any]: Dictionary containing DCF inputs including FCF values,
                          growth rates, financial metrics, and market data
            
        Raises:
            Exception: If DCF inputs cannot be calculated
            
        Example:
            >>> calc = FinancialCalculator('./data/companies/MSFT')
            >>> calc.load_financial_data()
            >>> inputs = calc.calculate_dcf_inputs()
            >>> print(f"FCFE values: {inputs['fcf_values']['FCFE']}")
        """
        try:
            # Ensure financial data is loaded
            if not self.financial_data:
                self.load_financial_statements()
            
            # Calculate all FCF types if not already done
            if not self.fcf_results:
                fcf_results = self.calculate_all_fcf_types()
                self.fcf_results = fcf_results
            
            # Calculate growth rates for all FCF types
            growth_rates = {}
            for fcf_type, values in self.fcf_results.items():
                if values:
                    growth_rates[fcf_type] = self.calculate_growth_rates(values)
            
            # Get market data
            market_data = {}
            if self.ticker_symbol:
                try:
                    market_data = self.fetch_market_data(self.ticker_symbol)
                except Exception as e:
                    logger.warning(f"Could not fetch market data: {e}")
                    market_data = {
                        'current_price': self.current_stock_price,
                        'market_cap': self.market_cap, 
                        'shares_outstanding': self.shares_outstanding
                    }
            
            # Prepare comprehensive DCF inputs
            dcf_inputs = {
                'fcf_values': self.fcf_results.copy(),
                'growth_rates': growth_rates,
                'market_data': market_data,
                'financial_metrics': self._calculate_all_metrics(),
                'company_info': {
                    'name': self.company_name,
                    'ticker': self.ticker_symbol,
                    'currency': self.currency,
                    'is_tase_stock': self.is_tase_stock
                },
                'data_quality': self.get_data_quality_report() if self.validation_enabled else None,
                'calculation_date': self.get_latest_report_date()
            }
            
            logger.info(f"DCF inputs calculated successfully for {self.company_name}")
            logger.info(f"Available FCF types: {list(self.fcf_results.keys())}")
            
            return dcf_inputs
            
        except Exception as e:
            logger.error(f"Error calculating DCF inputs: {e}")
            raise Exception(f"Failed to calculate DCF inputs: {e}")

    def get_financial_metrics(self) -> Dict[str, Any]:
        """
        Return formatted financial metrics dictionary
        
        This method calculates and returns a comprehensive dictionary of financial metrics
        including profitability ratios, efficiency metrics, and key financial indicators.
        
        Returns:
            Dict[str, Any]: Dictionary containing formatted financial metrics with
                          calculated ratios and financial indicators
            
        Example:
            >>> calc = FinancialCalculator('./data/companies/AAPL')
            >>> calc.load_financial_data()
            >>> metrics = calc.get_financial_metrics()
            >>> print(f"ROE: {metrics['roe']:.2%}")
        """
        try:
            # Ensure financial data is loaded
            if not self.financial_data:
                self.load_financial_statements()
            
            # Calculate all basic metrics
            raw_metrics = self._calculate_all_metrics()
            
            # Format metrics into comprehensive dictionary with proper structure
            formatted_metrics = {
                # Raw calculated metrics
                'raw_metrics': raw_metrics,
                
                # Profitability Metrics (formatted as percentages where appropriate)
                'profitability': self._format_profitability_metrics(raw_metrics),
                
                # Efficiency Metrics
                'efficiency': self._format_efficiency_metrics(raw_metrics),
                
                # Liquidity Metrics  
                'liquidity': self._format_liquidity_metrics(raw_metrics),
                
                # Leverage Metrics
                'leverage': self._format_leverage_metrics(raw_metrics),
                
                # Growth Metrics
                'growth': self._calculate_growth_metrics(raw_metrics),
                
                # FCF Metrics
                'fcf_metrics': self._format_fcf_metrics(),
                
                # Company Information
                'company_info': {
                    'name': self.company_name,
                    'ticker': self.ticker_symbol,
                    'currency': self.currency,
                    'is_tase_stock': self.is_tase_stock,
                    'report_date': self.get_latest_report_date()
                },
                
                # Data Quality Information
                'data_quality': {
                    'validation_enabled': self.validation_enabled,
                    'data_quality_report': self.get_data_quality_report() if self.validation_enabled else None,
                    'metrics_completeness': self._assess_metrics_completeness(raw_metrics)
                }
            }
            
            logger.info(f"Financial metrics formatted successfully for {self.company_name}")
            return formatted_metrics
            
        except Exception as e:
            logger.error(f"Error calculating financial metrics: {e}")
            return {
                'error': str(e),
                'company_info': {
                    'name': self.company_name,
                    'ticker': self.ticker_symbol
                }
            }

    def _format_profitability_metrics(self, raw_metrics: Dict[str, List[float]]) -> Dict[str, List[float]]:
        """Format profitability metrics with proper scaling"""
        try:
            profitability = {}
            
            # Standard profitability metrics (already calculated as ratios)
            profitability_keys = ['roe', 'roa', 'roic', 'gross_margin', 'operating_margin', 'net_margin']
            
            for key in profitability_keys:
                if key in raw_metrics and raw_metrics[key]:
                    # These are already in ratio form (0.15 = 15%), keep as-is
                    profitability[key] = raw_metrics[key]
            
            return profitability
        except Exception as e:
            logger.warning(f"Error formatting profitability metrics: {e}")
            return {}

    def _format_efficiency_metrics(self, raw_metrics: Dict[str, List[float]]) -> Dict[str, List[float]]:
        """Format efficiency metrics"""
        try:
            efficiency = {}
            
            efficiency_keys = ['asset_turnover', 'inventory_turnover', 'receivables_turnover', 'working_capital_turnover']
            
            for key in efficiency_keys:
                if key in raw_metrics and raw_metrics[key]:
                    efficiency[key] = raw_metrics[key]
                    
            return efficiency
        except Exception as e:
            logger.warning(f"Error formatting efficiency metrics: {e}")
            return {}

    def _format_liquidity_metrics(self, raw_metrics: Dict[str, List[float]]) -> Dict[str, List[float]]:
        """Format liquidity metrics"""
        try:
            liquidity = {}
            
            liquidity_keys = ['current_ratio', 'quick_ratio', 'cash_ratio']
            
            for key in liquidity_keys:
                if key in raw_metrics and raw_metrics[key]:
                    liquidity[key] = raw_metrics[key]
                    
            return liquidity
        except Exception as e:
            logger.warning(f"Error formatting liquidity metrics: {e}")
            return {}

    def _format_leverage_metrics(self, raw_metrics: Dict[str, List[float]]) -> Dict[str, List[float]]:
        """Format leverage metrics"""
        try:
            leverage = {}
            
            leverage_keys = ['debt_to_equity', 'debt_to_assets', 'interest_coverage', 'debt_service_coverage']
            
            for key in leverage_keys:
                if key in raw_metrics and raw_metrics[key]:
                    leverage[key] = raw_metrics[key]
                    
            return leverage
        except Exception as e:
            logger.warning(f"Error formatting leverage metrics: {e}")
            return {}

    def _calculate_growth_metrics(self, raw_metrics: Dict[str, List[float]]) -> Dict[str, float]:
        """Calculate growth metrics from raw data"""
        try:
            growth = {}
            
            # Calculate growth rates for key metrics
            growth_metrics = ['revenue', 'net_income', 'total_assets', 'shareholders_equity']
            
            for metric in growth_metrics:
                if metric in raw_metrics and len(raw_metrics[metric]) >= 2:
                    values = raw_metrics[metric]
                    if len(values) >= 2 and values[-2] != 0:
                        growth_rate = (values[-1] / values[-2]) - 1
                        growth[f'{metric}_growth'] = growth_rate
                        
            return growth
        except Exception as e:
            logger.warning(f"Error calculating growth metrics: {e}")
            return {}

    def _format_fcf_metrics(self) -> Dict[str, Any]:
        """Format FCF-related metrics"""
        try:
            if not self.fcf_results:
                return {}
                
            fcf_metrics = {
                'fcf_results': self.fcf_results.copy(),
                'fcf_growth_rates': {}
            }
            
            # Calculate FCF growth rates
            for fcf_type, values in self.fcf_results.items():
                if values and len(values) >= 2:
                    fcf_metrics['fcf_growth_rates'][fcf_type] = self.calculate_growth_rates(values)
                    
            return fcf_metrics
        except Exception as e:
            logger.warning(f"Error formatting FCF metrics: {e}")
            return {}

    def _assess_metrics_completeness(self, raw_metrics: Dict[str, List[float]]) -> Dict[str, Any]:
        """Assess completeness of calculated metrics"""
        try:
            total_metrics = len(raw_metrics)
            complete_metrics = sum(1 for values in raw_metrics.values() if values and any(v != 0 for v in values))
            
            return {
                'total_metrics': total_metrics,
                'complete_metrics': complete_metrics,
                'completeness_ratio': complete_metrics / total_metrics if total_metrics > 0 else 0,
                'missing_metrics': [key for key, values in raw_metrics.items() if not values or all(v == 0 for v in values)]
            }
        except Exception as e:
            logger.warning(f"Error assessing metrics completeness: {e}")
            return {'error': str(e)}


# Unified FCF Calculation Functions for All APIs
def calculate_unified_fcf(standardized_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Unified Free Cash Flow calculation that works with standardized data from any API.

    Args:
        standardized_data: Dictionary with standardized field names containing:
            - operating_cash_flow: Operating cash flow value
            - capital_expenditures: Capital expenditures value
            - source: API source identifier

    Returns:
        Dict containing FCF calculation results and metadata
    """
    try:
        operating_cash_flow = standardized_data.get("operating_cash_flow")
        capital_expenditures = standardized_data.get("capital_expenditures")
        source = standardized_data.get("source", "unknown")

        logger.info(
            f"Calculating FCF from {source}: OCF={operating_cash_flow}, CapEx={capital_expenditures}"
        )

        # Validate inputs
        if operating_cash_flow is None:
            logger.error(f"Operating cash flow is None from {source}")
            return {
                "free_cash_flow": None,
                "error": "Operating cash flow not available",
                "source": source,
                "calculation_date": datetime.now().isoformat(),
            }

        if capital_expenditures is None:
            logger.error(f"Capital expenditures is None from {source}")
            return {
                "free_cash_flow": None,
                "error": "Capital expenditures not available",
                "source": source,
                "calculation_date": datetime.now().isoformat(),
            }

        # Convert to float and validate with enhanced error handling
        try:
            with np.errstate(invalid='raise', over='raise'):
                ocf = float(operating_cash_flow)
                capex = float(capital_expenditures)

                # Validate for NaN, inf, and extreme values
                if not np.isfinite(ocf):
                    raise ValueError(f"Operating cash flow is not finite: {ocf}")
                if not np.isfinite(capex):
                    raise ValueError(f"Capital expenditures is not finite: {capex}")

        except (ValueError, TypeError, FloatingPointError) as e:
            logger.error(
                f"Invalid numeric values from {source}: OCF={operating_cash_flow}, CapEx={capital_expenditures}, Error: {e}"
            )
            return {
                "free_cash_flow": None,
                "error": f"Invalid numeric values: {e}",
                "source": source,
                "calculation_date": datetime.now().isoformat(),
            }

        # Calculate FCF: Operating Cash Flow - Capital Expenditures
        # Note: Most APIs report CapEx as negative, so we subtract the absolute value
        # to ensure consistent calculation regardless of sign convention
        with np.errstate(divide='ignore', invalid='ignore'):
            free_cash_flow = ocf - abs(capex)

            # Validate FCF result
            if not np.isfinite(free_cash_flow):
                logger.warning(f"FCF calculation resulted in non-finite value: {free_cash_flow}")
                free_cash_flow = 0.0

        # Calculate FCF margin if revenue is available
        fcf_margin = None
        total_revenue = standardized_data.get("total_revenue")
        if total_revenue and total_revenue != 0:
            with np.errstate(divide='ignore', invalid='ignore'):
                revenue_float = float(total_revenue)
                if np.isfinite(revenue_float) and revenue_float != 0:
                    fcf_margin = (free_cash_flow / revenue_float) * 100
                    if not np.isfinite(fcf_margin):
                        fcf_margin = None

        result = {
            "free_cash_flow": free_cash_flow,
            "operating_cash_flow": ocf,
            "capital_expenditures": capex,
            "fcf_margin_percent": fcf_margin,
            "source": source,
            "calculation_method": "unified_standardized",
            "calculation_date": datetime.now().isoformat(),
            "success": True,
        }

        # Add additional context from standardized data
        if "report_date" in standardized_data:
            result["report_date"] = standardized_data["report_date"]
        if "report_year" in standardized_data:
            result["report_year"] = standardized_data["report_year"]
        if "report_period" in standardized_data:
            result["report_period"] = standardized_data["report_period"]

        logger.info(f"FCF calculation successful from {source}: {free_cash_flow:,.0f}")
        if fcf_margin:
            logger.info(f"FCF margin: {fcf_margin:.2f}%")

        return result

    except Exception as e:
        logger.error(f"Unified FCF calculation failed: {e}")
        return {
            "free_cash_flow": None,
            "error": f"Calculation failed: {e}",
            "source": standardized_data.get("source", "unknown"),
            "calculation_date": datetime.now().isoformat(),
            "success": False,
        }


def validate_fcf_calculation(
    fcf_result: Dict[str, Any], min_reasonable_fcf: float = -1e12, max_reasonable_fcf: float = 1e12
) -> bool:
    """
    Validate FCF calculation results for reasonableness.

    Args:
        fcf_result: Result from calculate_unified_fcf
        min_reasonable_fcf: Minimum reasonable FCF value
        max_reasonable_fcf: Maximum reasonable FCF value

    Returns:
        bool: True if FCF result appears reasonable
    """
    if not fcf_result.get("success", False):
        return False

    fcf = fcf_result.get("free_cash_flow")
    if fcf is None:
        return False

    # Check reasonable range
    if not (min_reasonable_fcf <= fcf <= max_reasonable_fcf):
        logger.warning(f"FCF value outside reasonable range: {fcf}")
        return False

    # Check that OCF and CapEx are reasonable
    ocf = fcf_result.get("operating_cash_flow", 0)
    capex = fcf_result.get("capital_expenditures", 0)

    if abs(ocf) > max_reasonable_fcf or abs(capex) > max_reasonable_fcf:
        logger.warning(f"OCF or CapEx values seem unreasonable: OCF={ocf}, CapEx={capex}")
        return False

    return True


def calculate_fcf_from_api_data(api_data: Dict[str, Any], api_type: str) -> Dict[str, Any]:
    """
    Calculate FCF from API-specific data by first converting to standardized format.

    Args:
        api_data: Raw data from specific API
        api_type: Type of API (alpha_vantage, fmp, yfinance, polygon)

    Returns:
        Dict containing FCF calculation results
    """
    try:
        # Import converters
        if api_type == "alpha_vantage":
            from alpha_vantage_converter import AlphaVantageConverter

            standardized_data = AlphaVantageConverter.convert_financial_data(api_data)
        elif api_type == "fmp":
            from fmp_converter import FMPConverter

            standardized_data = FMPConverter.convert_financial_data(api_data)
        elif api_type == "yfinance":
            from yfinance_converter import YfinanceConverter

            standardized_data = YfinanceConverter.convert_financial_data(api_data)
        elif api_type == "polygon":
            from polygon_converter import PolygonConverter

            standardized_data = PolygonConverter.convert_financial_data(api_data)
        else:
            logger.error(f"Unsupported API type: {api_type}")
            return {
                "free_cash_flow": None,
                "error": f"Unsupported API type: {api_type}",
                "source": api_type,
                "success": False,
            }

        # Calculate FCF using unified function
        fcf_result = calculate_unified_fcf(standardized_data)

        # Validate result
        if not validate_fcf_calculation(fcf_result):
            fcf_result["validation_warning"] = "FCF calculation may be unreliable"

        return fcf_result

    except Exception as e:
        logger.error(f"FCF calculation from {api_type} failed: {e}")
        return {
            "free_cash_flow": None,
            "error": f"API-specific calculation failed: {e}",
            "source": api_type,
            "success": False,
        }

    def _extract_excel_dividends(self):
        """
        Extract dividend information from Excel cash flow statements using specific field names
        
        Returns:
            dict: Dividend data with years, total_dividends, regular_dividends, special_dividends
        """
        try:
            # Get cash flow statement data
            cashflow_data = self.financial_data.get('cashflow_fy', {})
            if not cashflow_data:
                # Try alternative cash flow data keys
                for alt_key in ['Cash Flow Statement', 'cashflow_ltm']:
                    cashflow_data = self.financial_data.get(alt_key, {})
                    if cashflow_data:
                        break
            
            if not cashflow_data:
                logger.info("No cash flow data available for dividend extraction")
                return None
            
            # Enhanced dividend field matching patterns
            dividend_patterns = [
                # Primary dividend patterns
                'dividends paid',
                'dividend payments', 
                'cash dividends',
                'dividends to shareholders',
                'dividends to stockholders',
                'common dividends',
                'dividend paid to shareholders',
                'payment of dividends',
                # Alternative patterns
                'cash_dividends_paid',
                'dividend_payments_cash',
                'shareholder_dividends',
                'common_stock_dividends',
                'regular_dividends',
                'cash dividend payments',
                'dividends paid to common shareholders'
            ]
            
            # Special dividend patterns (separate tracking)
            special_dividend_patterns = [
                'special dividends',
                'special dividend payments',
                'extraordinary dividends',
                'one-time dividends',
                'special_dividends_paid'
            ]
            
            years = []
            total_dividends = []
            regular_dividends = []
            special_dividends = []
            
            # Process DataFrame format
            if isinstance(cashflow_data, pd.DataFrame):
                # Get available year columns (skip first few metadata columns)
                year_columns = [col for col in cashflow_data.columns if str(col).replace('FY-', '').replace('FY', '').isdigit()]
                
                for year_col in year_columns:
                    year_num = int(str(year_col).replace('FY-', '').replace('FY', ''))
                    dividend_total = 0
                    regular_total = 0
                    special_total = 0
                    
                    # Search for dividend entries in this year
                    for idx, row in cashflow_data.iterrows():
                        metric_name = str(row.iloc[0] if len(row) > 0 else '').lower().strip()
                        
                        if pd.notna(row.get(year_col, None)):
                            value = abs(float(row[year_col]))  # Make positive
                            
                            # Check for regular dividend patterns
                            for pattern in dividend_patterns:
                                if pattern in metric_name:
                                    dividend_total += value
                                    regular_total += value
                                    logger.debug(f"Found dividend: '{metric_name}' = {value:.2f}M for {year_num}")
                                    break
                            
                            # Check for special dividend patterns
                            for pattern in special_dividend_patterns:
                                if pattern in metric_name:
                                    special_total += value
                                    logger.debug(f"Found special dividend: '{metric_name}' = {value:.2f}M for {year_num}")
                                    break
                    
                    if dividend_total > 0 or special_total > 0:
                        years.append(year_num)
                        total_dividends.append(dividend_total + special_total)
                        regular_dividends.append(regular_total) 
                        special_dividends.append(special_total)
                        
            else:
                # Process dictionary format
                for year_key, year_data in cashflow_data.items():
                    if isinstance(year_key, str) and year_key.replace('FY-', '').replace('FY', '').isdigit():
                        year_num = int(year_key.replace('FY-', '').replace('FY', ''))
                        dividend_total = 0
                        regular_total = 0
                        special_total = 0
                        
                        if isinstance(year_data, dict):
                            for metric_name, value in year_data.items():
                                if isinstance(value, (int, float)) and value != 0:
                                    metric_lower = str(metric_name).lower().strip()
                                    abs_value = abs(value)
                                    
                                    # Check dividend patterns
                                    for pattern in dividend_patterns:
                                        if pattern in metric_lower:
                                            dividend_total += abs_value
                                            regular_total += abs_value
                                            break
                                    
                                    # Check special dividend patterns  
                                    for pattern in special_dividend_patterns:
                                        if pattern in metric_lower:
                                            special_total += abs_value
                                            break
                        
                        if dividend_total > 0 or special_total > 0:
                            years.append(year_num)
                            total_dividends.append(dividend_total + special_total)
                            regular_dividends.append(regular_total)
                            special_dividends.append(special_total)
            
            # Sort by year
            if years:
                sorted_data = sorted(zip(years, total_dividends, regular_dividends, special_dividends))
                years, total_dividends, regular_dividends, special_dividends = zip(*sorted_data)
                
                logger.info(f"Successfully extracted Excel dividends for {len(years)} years: {years}")
                return {
                    'years': list(years),
                    'total_dividends': list(total_dividends),
                    'regular_dividends': list(regular_dividends), 
                    'special_dividends': list(special_dividends)
                }
            else:
                logger.info("No dividend data found in Excel cash flow statements")
                return None
                
        except Exception as e:
            logger.error(f"Error extracting Excel dividends: {e}")
            return None
    
    def _extract_excel_shares_outstanding(self):
        """
        Extract shares outstanding from Excel balance sheet data
        
        Returns:
            float: Shares outstanding or None if not found
        """
        try:
            # Check balance sheet data
            balance_sheet = self.financial_data.get('balance_fy', {})
            if not balance_sheet:
                balance_sheet = self.financial_data.get('Balance Sheet', {})
            
            if not balance_sheet:
                logger.info("No balance sheet data available for shares outstanding extraction")
                return None
            
            shares_patterns = [
                'shares outstanding',
                'common shares outstanding', 
                'outstanding shares',
                'number of shares',
                'shares_outstanding',
                'common_shares_outstanding',
                'ordinary shares outstanding',
                'total shares outstanding'
            ]
            
            # Get most recent year's shares outstanding
            latest_shares = None
            
            if isinstance(balance_sheet, pd.DataFrame):
                # Get latest year column
                year_columns = [col for col in balance_sheet.columns if str(col).replace('FY-', '').replace('FY', '').isdigit()]
                if year_columns:
                    latest_year_col = sorted(year_columns, key=lambda x: int(str(x).replace('FY-', '').replace('FY', '')))[-1]
                    
                    for idx, row in balance_sheet.iterrows():
                        metric_name = str(row.iloc[0] if len(row) > 0 else '').lower().strip()
                        
                        for pattern in shares_patterns:
                            if pattern in metric_name:
                                shares_value = row.get(latest_year_col, None)
                                if pd.notna(shares_value) and shares_value != 0:
                                    # Convert to actual number of shares (Excel data typically in millions)
                                    latest_shares = abs(float(shares_value)) * 1_000_000
                                    logger.info(f"Found shares outstanding: {shares_value:.1f}M = {latest_shares:,.0f} shares")
                                    break
                        if latest_shares:
                            break
            
            else:
                # Process dictionary format - get most recent year
                years = [k for k in balance_sheet.keys() if isinstance(k, str) and k.replace('FY-', '').replace('FY', '').isdigit()]
                if years:
                    latest_year = sorted(years, key=lambda x: int(x.replace('FY-', '').replace('FY', '')))[-1]
                    year_data = balance_sheet.get(latest_year, {})
                    
                    if isinstance(year_data, dict):
                        for metric_name, value in year_data.items():
                            if isinstance(value, (int, float)) and value != 0:
                                metric_lower = str(metric_name).lower().strip()
                                
                                for pattern in shares_patterns:
                                    if pattern in metric_lower:
                                        latest_shares = abs(float(value)) * 1_000_000
                                        logger.info(f"Found shares outstanding: {value:.1f}M = {latest_shares:,.0f} shares") 
                                        break
                                if latest_shares:
                                    break
            
            return latest_shares
            
        except Exception as e:
            logger.error(f"Error extracting Excel shares outstanding: {e}")
            return None
