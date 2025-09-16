"""
Excel Format Detection and Validation System
============================================

This module provides comprehensive Excel format detection capabilities to automatically
identify different Excel layouts and structures, enabling robust data extraction across
various financial statement formats.

Key Features:
- Automatic format detection with confidence scoring
- Support for multiple layout variations
- Validation rules for financial statement structures
- Extensible architecture for new format types
- Integration with existing ExcelDataExtractor
"""

import logging
import re
from typing import Dict, List, Optional, Tuple, Any, NamedTuple
from dataclasses import dataclass, field
from enum import Enum, auto
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class FormatType(Enum):
    """Supported Excel format types"""
    PREMIUM_EXPORT = auto()
    STANDARD_TEMPLATE = auto()
    QUARTERLY_LAYOUT = auto()
    CUSTOM_FORMAT = auto()
    UNKNOWN = auto()


class StatementType(Enum):
    """Financial statement types"""
    INCOME = auto()
    BALANCE = auto()
    CASHFLOW = auto()
    UNKNOWN = auto()


@dataclass
class FormatSignature:
    """Defines the signature patterns for a specific Excel format"""
    name: str
    format_type: FormatType

    # Key cell patterns that identify this format
    identifier_patterns: List[Tuple[int, int, str]] = field(default_factory=list)
    company_name_positions: List[Tuple[int, int]] = field(default_factory=list)
    period_header_positions: List[Tuple[int, int]] = field(default_factory=list)
    data_start_indicators: List[Tuple[int, int, str]] = field(default_factory=list)

    # Layout characteristics
    expected_max_rows: int = 100
    expected_max_columns: int = 20
    data_columns_range: Tuple[int, int] = (4, 13)

    # Required sections
    required_sections: List[str] = field(default_factory=list)


class FormatDetectionResult(NamedTuple):
    """Result of format detection analysis"""
    format_type: FormatType
    confidence_score: float
    detected_patterns: Dict[str, Any]
    validation_errors: List[str]
    suggested_adjustments: Dict[str, Any]


@dataclass
class ExcelStructureInfo:
    """Information about Excel file structure"""
    sheet_name: str
    max_row: int
    max_column: int
    company_name: Optional[str]
    statement_type: Optional[StatementType]
    period_dates: List[str]
    data_start_row: int
    data_start_column: int
    format_indicators: Dict[str, Any]


class ExcelFormatDetector:
    """
    Advanced Excel format detection system that can identify and validate
    different Excel layouts for financial statements.
    """

    def __init__(self):
        """Initialize the format detector with predefined format signatures"""
        self.format_signatures = self._initialize_format_signatures()
        self.confidence_threshold = 0.7

    def _initialize_format_signatures(self) -> Dict[FormatType, FormatSignature]:
        """Initialize predefined format signatures"""
        signatures = {}

        # Premium Export Format (current standard format)
        premium_export = FormatSignature(
            name="Premium Export Format",
            format_type=FormatType.PREMIUM_EXPORT,
            identifier_patterns=[
                (3, 3, "Premium Export"),  # Key identifier
                (6, 3, r".*Statement"),    # Statement type indicator
                (10, 3, "Period End Date") # Period date header
            ],
            company_name_positions=[(2, 3)],
            period_header_positions=[(8, 4), (8, 5), (8, 6)],  # FY-9, FY-8, etc.
            data_start_indicators=[
                (10, 3, "Period End Date"),
                (12, 3, r"Revenue|Sales|Income")  # First financial metric
            ],
            expected_max_rows=50,
            expected_max_columns=15,
            data_columns_range=(4, 13),
            required_sections=["company_name", "statement_type", "period_dates"]
        )
        signatures[FormatType.PREMIUM_EXPORT] = premium_export

        # Standard Template Format (alternative layout)
        standard_template = FormatSignature(
            name="Standard Template Format",
            format_type=FormatType.STANDARD_TEMPLATE,
            identifier_patterns=[
                (1, 1, r".*Financial Statement"),
                (2, 1, r"\d{4}-\d{2}-\d{2}")  # Date pattern
            ],
            company_name_positions=[(1, 2), (1, 3)],
            period_header_positions=[(3, 2), (3, 3), (3, 4)],
            data_start_indicators=[
                (4, 1, r"Revenue|Sales|Income")
            ],
            expected_max_rows=100,
            expected_max_columns=20,
            data_columns_range=(2, 15),
            required_sections=["company_name", "period_dates"]
        )
        signatures[FormatType.STANDARD_TEMPLATE] = standard_template

        return signatures

    def detect_format(self, worksheet: Worksheet) -> FormatDetectionResult:
        """
        Detect the format of an Excel worksheet

        Args:
            worksheet: The Excel worksheet to analyze

        Returns:
            FormatDetectionResult with detection details and confidence score
        """
        logger.info(f"Starting format detection for worksheet: {worksheet.title}")

        # Extract basic structure information
        structure_info = self._extract_structure_info(worksheet)

        # Test each format signature
        detection_results = {}
        for format_type, signature in self.format_signatures.items():
            confidence, patterns, errors = self._test_format_signature(
                worksheet, signature, structure_info
            )
            detection_results[format_type] = (confidence, patterns, errors)

        # Find the best match
        best_format = FormatType.UNKNOWN
        best_confidence = 0.0
        best_patterns = {}
        best_errors = []

        for format_type, (confidence, patterns, errors) in detection_results.items():
            if confidence > best_confidence:
                best_confidence = confidence
                best_format = format_type
                best_patterns = patterns
                best_errors = errors

        # Generate suggestions for improvement
        suggestions = self._generate_suggestions(
            structure_info, best_format, best_confidence, best_errors
        )

        result = FormatDetectionResult(
            format_type=best_format,
            confidence_score=best_confidence,
            detected_patterns=best_patterns,
            validation_errors=best_errors,
            suggested_adjustments=suggestions
        )

        logger.info(
            f"Format detection complete: {best_format.name} "
            f"(confidence: {best_confidence:.2f})"
        )

        return result

    def _extract_structure_info(self, worksheet: Worksheet) -> ExcelStructureInfo:
        """Extract basic structural information from the worksheet"""
        max_row = worksheet.max_row if worksheet.max_row else 50
        max_column = worksheet.max_column if worksheet.max_column else 15

        structure_info = ExcelStructureInfo(
            sheet_name=worksheet.title,
            max_row=max_row,
            max_column=max_column,
            company_name=None,
            statement_type=None,
            period_dates=[],
            data_start_row=1,
            data_start_column=1,
            format_indicators={}
        )

        # Look for company name in common positions
        for row in range(1, min(6, max_row + 1)):
            for col in range(1, min(6, max_column + 1)):
                cell_value = worksheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    if self._is_likely_company_name(cell_value):
                        structure_info.company_name = cell_value.strip()
                        break
            if structure_info.company_name:
                break

        # Determine statement type from content
        structure_info.statement_type = self._detect_statement_type(worksheet)

        # Find period dates
        structure_info.period_dates = self._extract_period_dates(worksheet)

        # Identify data start position
        structure_info.data_start_row, structure_info.data_start_column = \
            self._find_data_start_position(worksheet)

        return structure_info

    def _test_format_signature(
        self,
        worksheet: Worksheet,
        signature: FormatSignature,
        structure_info: ExcelStructureInfo
    ) -> Tuple[float, Dict[str, Any], List[str]]:
        """
        Test a worksheet against a specific format signature

        Returns:
            Tuple of (confidence_score, matched_patterns, validation_errors)
        """
        confidence_score = 0.0
        matched_patterns = {}
        validation_errors = []
        total_tests = 0
        passed_tests = 0

        # Test identifier patterns
        for row, col, pattern in signature.identifier_patterns:
            total_tests += 1
            try:
                cell_value = worksheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    if re.search(pattern, cell_value, re.IGNORECASE):
                        passed_tests += 1
                        matched_patterns[f"identifier_{row}_{col}"] = cell_value
                    else:
                        validation_errors.append(
                            f"Pattern mismatch at ({row}, {col}): "
                            f"expected '{pattern}', found '{cell_value}'"
                        )
                else:
                    validation_errors.append(
                        f"No content at identifier position ({row}, {col})"
                    )
            except Exception as e:
                validation_errors.append(f"Error checking ({row}, {col}): {e}")

        # Test company name positions
        company_found = False
        for row, col in signature.company_name_positions:
            total_tests += 1
            try:
                cell_value = worksheet.cell(row, col).value
                if cell_value and self._is_likely_company_name(str(cell_value)):
                    passed_tests += 1
                    company_found = True
                    matched_patterns["company_name"] = str(cell_value)
                    break
            except Exception as e:
                validation_errors.append(f"Error checking company name at ({row}, {col}): {e}")

        if not company_found and signature.company_name_positions:
            validation_errors.append("No valid company name found in expected positions")

        # Test period header positions
        period_headers_found = 0
        for row, col in signature.period_header_positions:
            total_tests += 1
            try:
                cell_value = worksheet.cell(row, col).value
                if cell_value and str(cell_value).startswith(('FY-', 'Q', '20', 'Period')):
                    passed_tests += 1
                    period_headers_found += 1
                    matched_patterns[f"period_header_{row}_{col}"] = str(cell_value)
            except Exception as e:
                validation_errors.append(f"Error checking period header at ({row}, {col}): {e}")

        # Test data structure constraints
        total_tests += 3  # Structure tests

        # Check worksheet size is reasonable
        max_row = worksheet.max_row if worksheet.max_row else 50
        max_column = worksheet.max_column if worksheet.max_column else 15

        if (signature.expected_max_rows >= max_row and
            signature.expected_max_columns >= max_column):
            passed_tests += 1
        else:
            validation_errors.append(
                f"Worksheet size ({max_row}x{max_column}) "
                f"exceeds expected ({signature.expected_max_rows}x{signature.expected_max_columns})"
            )

        # Check data columns are in expected range
        data_start, data_end = signature.data_columns_range
        if data_start <= max_column <= data_end + 5:  # Allow some flexibility
            passed_tests += 1
        else:
            validation_errors.append(
                f"Data columns ({max_column}) outside expected range "
                f"({data_start}-{data_end})"
            )

        # Check required sections are present
        sections_found = 0
        for section in signature.required_sections:
            if section == "company_name" and structure_info.company_name:
                sections_found += 1
            elif section == "period_dates" and structure_info.period_dates:
                sections_found += 1
            elif section == "statement_type" and structure_info.statement_type != StatementType.UNKNOWN:
                sections_found += 1

        if sections_found == len(signature.required_sections):
            passed_tests += 1
        else:
            validation_errors.append(
                f"Missing required sections: {len(signature.required_sections) - sections_found} "
                f"of {len(signature.required_sections)}"
            )

        # Calculate confidence score
        if total_tests > 0:
            confidence_score = passed_tests / total_tests

        # Apply bonus for strong indicators
        if "Premium Export" in matched_patterns.get("identifier_3_3", ""):
            confidence_score *= 1.2  # Strong indicator bonus

        # Cap at 1.0
        confidence_score = min(confidence_score, 1.0)

        return confidence_score, matched_patterns, validation_errors

    def _is_likely_company_name(self, value: str) -> bool:
        """Determine if a string is likely to be a company name"""
        value = value.strip()

        if len(value) < 2 or len(value) > 100:
            return False

        # Skip common non-company patterns
        skip_patterns = [
            r'^period end date',
            r'^fiscal year',
            r'^quarter',
            r'^annual',
            r'^monthly',
            r'^\d{4}',  # Years
            r'^q[1-4]',  # Quarters
            r'^jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec',  # Months
            r'^total|sum|average|net|gross',  # Financial terms
            r'^usd|eur|gbp',  # Currencies
            r'^premium export$',  # Format identifier
        ]

        for pattern in skip_patterns:
            if re.match(pattern, value.lower()):
                return False

        # Look for company indicators
        company_indicators = [
            r'\b(inc|corp|corporation|company|ltd|llc|plc)\b',
            r'\b(technologies|tech|systems|group|holdings)\b',
            r'\b(international|global|worldwide)\b',
        ]

        for pattern in company_indicators:
            if re.search(pattern, value.lower()):
                return True

        # If it contains multiple words and capital letters, likely a company name
        if len(value.split()) > 1 and any(c.isupper() for c in value):
            return True

        return False

    def _detect_statement_type(self, worksheet: Worksheet) -> StatementType:
        """Detect the type of financial statement"""
        max_row = worksheet.max_row if worksheet.max_row else 50
        max_column = worksheet.max_column if worksheet.max_column else 15

        # Look for statement type indicators in first 20 rows
        for row in range(1, min(21, max_row + 1)):
            for col in range(1, min(10, max_column + 1)):
                try:
                    cell_value = worksheet.cell(row, col).value
                    if cell_value and isinstance(cell_value, str):
                        value_lower = cell_value.lower()
                        if 'income statement' in value_lower:
                            return StatementType.INCOME
                        elif 'balance sheet' in value_lower:
                            return StatementType.BALANCE
                        elif 'cash flow' in value_lower:
                            return StatementType.CASHFLOW
                except Exception:
                    continue

        return StatementType.UNKNOWN

    def _extract_period_dates(self, worksheet: Worksheet) -> List[str]:
        """Extract period end dates from the worksheet"""
        dates = []
        max_row = worksheet.max_row if worksheet.max_row else 50
        max_column = worksheet.max_column if worksheet.max_column else 15

        # Look for "Period End Date" row
        for row in range(1, min(21, max_row + 1)):
            for col in range(1, min(10, max_column + 1)):
                try:
                    cell_value = worksheet.cell(row, col).value
                    if cell_value and isinstance(cell_value, str) and "Period End Date" in cell_value:
                        # Extract dates from subsequent columns
                        for data_col in range(col + 1, min(col + 15, max_column + 1)):
                            date_value = worksheet.cell(row, data_col).value
                            if date_value is not None:
                                dates.append(str(date_value))
                            else:
                                break
                        return dates
                except Exception:
                    continue

        return dates

    def _find_data_start_position(self, worksheet: Worksheet) -> Tuple[int, int]:
        """Find the starting position of financial data"""
        # Look for first financial metric (usually Revenue, Sales, etc.)
        financial_keywords = [
            'revenue', 'sales', 'income', 'net sales', 'total revenue',
            'gross profit', 'operating income', 'ebitda'
        ]

        max_row = worksheet.max_row if worksheet.max_row else 50
        max_column = worksheet.max_column if worksheet.max_column else 15

        for row in range(1, min(30, max_row + 1)):
            for col in range(1, min(10, max_column + 1)):
                try:
                    cell_value = worksheet.cell(row, col).value
                    if cell_value and isinstance(cell_value, str):
                        value_lower = cell_value.lower().strip()
                        for keyword in financial_keywords:
                            if keyword in value_lower:
                                return row, col
                except Exception:
                    continue

        # Default fallback
        return 12, 3

    def _generate_suggestions(
        self,
        structure_info: ExcelStructureInfo,
        detected_format: FormatType,
        confidence: float,
        errors: List[str]
    ) -> Dict[str, Any]:
        """Generate suggestions for improving format detection"""
        suggestions = {}

        if confidence < self.confidence_threshold:
            suggestions['low_confidence_warning'] = (
                f"Low confidence score ({confidence:.2f}). "
                "Consider manual verification or format adjustment."
            )

        if not structure_info.company_name:
            suggestions['missing_company_name'] = (
                "Company name not detected. Verify company name is in expected position."
            )

        if not structure_info.period_dates:
            suggestions['missing_period_dates'] = (
                "Period end dates not found. Check for 'Period End Date' row."
            )

        if structure_info.statement_type == StatementType.UNKNOWN:
            suggestions['unknown_statement_type'] = (
                "Statement type not detected. Verify statement type is clearly labeled."
            )

        if len(errors) > 0:
            suggestions['validation_errors'] = errors[:5]  # Limit to top 5 errors

        # Format-specific suggestions
        if detected_format == FormatType.UNKNOWN:
            suggestions['unknown_format'] = (
                "Format not recognized. Consider adding custom format signature or "
                "reformatting Excel file to match supported templates."
            )

        return suggestions

    def validate_format(
        self,
        worksheet: Worksheet,
        expected_format: FormatType = None
    ) -> FormatDetectionResult:
        """
        Validate a worksheet against a specific format or detect the best match

        Args:
            worksheet: Excel worksheet to validate
            expected_format: Specific format to validate against (optional)

        Returns:
            FormatDetectionResult with validation details
        """
        if expected_format:
            # Validate against specific format
            signature = self.format_signatures.get(expected_format)
            if not signature:
                return FormatDetectionResult(
                    format_type=FormatType.UNKNOWN,
                    confidence_score=0.0,
                    detected_patterns={},
                    validation_errors=[f"Unknown format type: {expected_format}"],
                    suggested_adjustments={}
                )

            structure_info = self._extract_structure_info(worksheet)
            confidence, patterns, errors = self._test_format_signature(
                worksheet, signature, structure_info
            )
            suggestions = self._generate_suggestions(
                structure_info, expected_format, confidence, errors
            )

            return FormatDetectionResult(
                format_type=expected_format,
                confidence_score=confidence,
                detected_patterns=patterns,
                validation_errors=errors,
                suggested_adjustments=suggestions
            )
        else:
            # Detect best format
            return self.detect_format(worksheet)

    def get_format_info(self, format_type: FormatType) -> Optional[FormatSignature]:
        """Get detailed information about a specific format"""
        return self.format_signatures.get(format_type)

    def list_supported_formats(self) -> List[str]:
        """Get list of supported format names"""
        return [sig.name for sig in self.format_signatures.values()]


def detect_excel_format(worksheet: Worksheet) -> FormatDetectionResult:
    """
    Convenience function to detect Excel format

    Args:
        worksheet: Excel worksheet to analyze

    Returns:
        FormatDetectionResult with detection details
    """
    detector = ExcelFormatDetector()
    return detector.detect_format(worksheet)


def validate_excel_format(
    worksheet: Worksheet,
    expected_format: FormatType = None
) -> FormatDetectionResult:
    """
    Convenience function to validate Excel format

    Args:
        worksheet: Excel worksheet to validate
        expected_format: Expected format type (optional)

    Returns:
        FormatDetectionResult with validation details
    """
    detector = ExcelFormatDetector()
    return detector.validate_format(worksheet, expected_format)


if __name__ == "__main__":
    # Example usage and testing
    from openpyxl import load_workbook
    import sys

    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        print(f"Analyzing Excel format: {file_path}")

        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            result = detect_excel_format(ws)

            print(f"\nDetection Results:")
            print(f"Format Type: {result.format_type.name}")
            print(f"Confidence Score: {result.confidence_score:.2f}")
            print(f"Detected Patterns: {len(result.detected_patterns)}")

            if result.validation_errors:
                print(f"\nValidation Errors ({len(result.validation_errors)}):")
                for error in result.validation_errors[:5]:
                    print(f"  - {error}")

            if result.suggested_adjustments:
                print(f"\nSuggestions:")
                for key, value in result.suggested_adjustments.items():
                    print(f"  - {key}: {value}")

            wb.close()

        except Exception as e:
            print(f"Error analyzing file: {e}")
    else:
        print("Usage: python format_detector.py <path_to_excel_file>")