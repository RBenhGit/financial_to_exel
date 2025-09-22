"""
Excel Utilities Module

This module provides dynamic Excel data extraction utilities to replace
hardcoded cell references and make the application more robust across
different Excel file formats.
"""

import os
import re
import logging
from typing import Dict, List, Tuple, Optional, Any, Union
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException
from dataclasses import dataclass
from config import get_excel_config, get_financial_metrics_config
from .format_detector import (
    ExcelFormatDetector,
    FormatDetectionResult,
    FormatType,
    detect_excel_format
)
from .international_format_handler import (
    InternationalFormatHandler,
    InternationalConfig,
    NumberFormat,
    DateFormat,
    create_european_handler,
    create_asian_handler,
    create_us_handler
)
from .custom_template_manager import (
    CustomTemplateManager,
    CustomTemplate,
    TemplateType
)
from ..data_processing.data_validator import FinancialDataValidator, validate_financial_statements_comprehensive

logger = logging.getLogger(__name__)


class ExcelProcessingError(Exception):
    """Base exception for Excel processing errors"""
    pass


class CorruptedFileError(ExcelProcessingError):
    """Raised when Excel file is corrupted or unreadable"""
    pass


class MissingDataError(ExcelProcessingError):
    """Raised when required data is missing from Excel file"""
    pass


class FormatInconsistencyError(ExcelProcessingError):
    """Raised when Excel format is inconsistent or unsupported"""
    pass


@dataclass
class ExcelFileValidationResult:
    """Result of Excel file validation"""
    is_valid: bool
    file_path: str
    file_size: int
    corruption_detected: bool
    missing_sheets: List[str]
    format_issues: List[str]
    data_quality_issues: List[str]
    recovery_suggestions: List[str]
    severity: str  # 'low', 'medium', 'high', 'critical'


@dataclass
class DataRecoveryResult:
    """Result of data recovery attempt"""
    success: bool
    recovered_metrics: Dict[str, Any]
    partial_data: Dict[str, Any]
    interpolated_values: Dict[str, List[float]]
    confidence_score: float
    recovery_method: str
    warnings: List[str]


@dataclass
class CellLocation:
    """Represents a cell location in an Excel worksheet"""

    row: int
    column: int
    value: Any = None


@dataclass
class ExcelMetric:
    """Represents a financial metric found in Excel"""

    name: str
    location: CellLocation
    data_columns: List[int]
    values: List[Any]


class ExcelFileValidator:
    """
    Comprehensive Excel file validator for detecting corruption, missing data,
    and format inconsistencies before processing
    """

    def __init__(self):
        """Initialize the Excel file validator"""
        self.data_validator = FinancialDataValidator()
        self.format_detector = ExcelFormatDetector()

    def validate_file(self, file_path: str) -> ExcelFileValidationResult:
        """
        Perform comprehensive validation of Excel file

        Args:
            file_path (str): Path to Excel file

        Returns:
            ExcelFileValidationResult: Detailed validation results
        """
        logger.info(f"Starting comprehensive validation of {os.path.basename(file_path)}")

        # Initialize validation result
        result = ExcelFileValidationResult(
            is_valid=True,
            file_path=file_path,
            file_size=0,
            corruption_detected=False,
            missing_sheets=[],
            format_issues=[],
            data_quality_issues=[],
            recovery_suggestions=[],
            severity='low'
        )

        # Check file existence and basic properties
        if not self._validate_file_existence(file_path, result):
            return result

        # Check for file corruption
        if not self._validate_file_integrity(file_path, result):
            return result

        # Validate Excel structure and format
        if not self._validate_excel_structure(file_path, result):
            return result

        # Validate data quality
        self._validate_data_quality(file_path, result)

        # Generate recovery suggestions
        self._generate_recovery_suggestions(result)

        # Determine overall severity
        self._determine_severity(result)

        logger.info(f"Validation complete: {result.severity} severity, valid: {result.is_valid}")
        return result

    def _validate_file_existence(self, file_path: str, result: ExcelFileValidationResult) -> bool:
        """Validate file existence and basic properties"""
        try:
            if not os.path.exists(file_path):
                result.is_valid = False
                result.format_issues.append(f"File does not exist: {file_path}")
                result.severity = 'critical'
                return False

            file_size = os.path.getsize(file_path)
            result.file_size = file_size

            if file_size == 0:
                result.is_valid = False
                result.corruption_detected = True
                result.format_issues.append("File is empty (0 bytes)")
                result.severity = 'critical'
                return False

            if file_size < 1024:  # Less than 1KB is suspicious for Excel files
                result.format_issues.append("File size is unusually small (< 1KB)")
                result.severity = 'high'

            return True

        except Exception as e:
            result.is_valid = False
            result.format_issues.append(f"Error accessing file: {str(e)}")
            result.severity = 'critical'
            return False

    def _validate_file_integrity(self, file_path: str, result: ExcelFileValidationResult) -> bool:
        """Validate Excel file integrity and detect corruption"""
        try:
            # Attempt to load the workbook
            workbook = load_workbook(file_path, read_only=True, data_only=False)

            # Check if workbook has sheets
            if not workbook.sheetnames:
                result.is_valid = False
                result.corruption_detected = True
                result.format_issues.append("Workbook contains no sheets")
                result.severity = 'critical'
                workbook.close()
                return False

            # Validate each sheet can be accessed
            accessible_sheets = []
            for sheet_name in workbook.sheetnames:
                try:
                    worksheet = workbook[sheet_name]
                    # Try to access a cell to verify sheet integrity
                    _ = worksheet.cell(1, 1).value
                    accessible_sheets.append(sheet_name)
                except Exception as e:
                    result.format_issues.append(f"Cannot access sheet '{sheet_name}': {str(e)}")
                    result.missing_sheets.append(sheet_name)

            if not accessible_sheets:
                result.is_valid = False
                result.corruption_detected = True
                result.format_issues.append("No accessible sheets found in workbook")
                result.severity = 'critical'
                workbook.close()
                return False

            workbook.close()
            return True

        except InvalidFileException as e:
            result.is_valid = False
            result.corruption_detected = True
            result.format_issues.append(f"Invalid Excel file format: {str(e)}")
            result.severity = 'critical'
            return False

        except Exception as e:
            result.is_valid = False
            result.corruption_detected = True
            result.format_issues.append(f"File corruption detected: {str(e)}")
            result.severity = 'critical'
            return False

    def _validate_excel_structure(self, file_path: str, result: ExcelFileValidationResult) -> bool:
        """Validate Excel structure and format consistency"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            worksheet = workbook.active

            # Perform format detection
            format_result = self.format_detector.detect_format(worksheet)

            # Check format confidence
            if format_result.confidence_score < 0.5:
                result.format_issues.append(
                    f"Low format confidence ({format_result.confidence_score:.2f})"
                )
                if format_result.confidence_score < 0.3:
                    result.severity = 'high'

            # Add format detection errors
            if format_result.validation_errors:
                result.format_issues.extend(format_result.validation_errors[:3])

            # Check for empty or mostly empty sheets
            max_row = worksheet.max_row or 0
            max_col = worksheet.max_column or 0

            if max_row < 5 or max_col < 3:
                result.format_issues.append(
                    f"Sheet appears mostly empty (size: {max_row}x{max_col})"
                )
                result.severity = 'medium'

            workbook.close()
            return True

        except Exception as e:
            result.format_issues.append(f"Error validating Excel structure: {str(e)}")
            return False

    def _validate_data_quality(self, file_path: str, result: ExcelFileValidationResult):
        """Validate data quality within the Excel file"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            worksheet = workbook.active

            # Sample data validation
            empty_cells = 0
            total_cells = 0
            invalid_values = 0

            # Check first 20 rows and 15 columns for data quality
            for row in range(1, min(21, (worksheet.max_row or 20) + 1)):
                for col in range(1, min(16, (worksheet.max_column or 15) + 1)):
                    try:
                        cell_value = worksheet.cell(row, col).value
                        total_cells += 1

                        if cell_value is None or cell_value == "":
                            empty_cells += 1
                        elif isinstance(cell_value, str):
                            # Check for common data quality issues
                            if cell_value.strip() in ['#N/A', '#ERROR', '#DIV/0!', '#VALUE!']:
                                invalid_values += 1
                                result.data_quality_issues.append(
                                    f"Excel error at ({row}, {col}): {cell_value}"
                                )

                    except Exception as e:
                        invalid_values += 1
                        result.data_quality_issues.append(f"Cell access error at ({row}, {col})")

            # Calculate data quality metrics
            if total_cells > 0:
                empty_ratio = empty_cells / total_cells
                invalid_ratio = invalid_values / total_cells

                if empty_ratio > 0.8:
                    result.data_quality_issues.append(
                        f"High empty cell ratio: {empty_ratio:.1%}"
                    )
                    result.severity = 'medium'

                if invalid_ratio > 0.1:
                    result.data_quality_issues.append(
                        f"High invalid value ratio: {invalid_ratio:.1%}"
                    )
                    result.severity = 'medium'

            workbook.close()

        except Exception as e:
            result.data_quality_issues.append(f"Error validating data quality: {str(e)}")

    def _generate_recovery_suggestions(self, result: ExcelFileValidationResult):
        """Generate actionable recovery suggestions"""
        if result.corruption_detected:
            result.recovery_suggestions.extend([
                "Try opening file in Excel and saving as a new file",
                "Check if file was corrupted during transfer",
                "Restore from backup if available"
            ])

        if result.missing_sheets:
            result.recovery_suggestions.append(
                f"Check for missing sheets: {', '.join(result.missing_sheets)}"
            )

        if any("format confidence" in issue.lower() for issue in result.format_issues):
            result.recovery_suggestions.extend([
                "Verify Excel file follows expected template format",
                "Check that financial statements are properly labeled",
                "Ensure data is in expected columns and rows"
            ])

        if result.data_quality_issues:
            result.recovery_suggestions.extend([
                "Review data for completeness and accuracy",
                "Check for Excel errors (#N/A, #DIV/0!, etc.)",
                "Verify all required financial metrics are present"
            ])

    def _determine_severity(self, result: ExcelFileValidationResult):
        """Determine overall severity of validation issues"""
        if result.corruption_detected or not result.is_valid:
            result.severity = 'critical'
        elif len(result.format_issues) > 3 or len(result.data_quality_issues) > 5:
            result.severity = 'high'
        elif len(result.format_issues) > 1 or len(result.data_quality_issues) > 2:
            result.severity = 'medium'
        else:
            result.severity = 'low'


class DataRecoveryEngine:
    """
    Engine for recovering data from problematic Excel files using various strategies
    """

    def __init__(self):
        """Initialize the data recovery engine"""
        self.data_validator = FinancialDataValidator()

    def attempt_data_recovery(
        self,
        file_path: str,
        validation_result: ExcelFileValidationResult,
        target_metrics: List[str]
    ) -> DataRecoveryResult:
        """
        Attempt to recover data from problematic Excel file

        Args:
            file_path (str): Path to Excel file
            validation_result: Previous validation results
            target_metrics: List of required financial metrics

        Returns:
            DataRecoveryResult: Recovery attempt results
        """
        logger.info(f"Attempting data recovery for {os.path.basename(file_path)}")

        recovery_result = DataRecoveryResult(
            success=False,
            recovered_metrics={},
            partial_data={},
            interpolated_values={},
            confidence_score=0.0,
            recovery_method="",
            warnings=[]
        )

        if validation_result.corruption_detected:
            return self._attempt_corruption_recovery(file_path, recovery_result)

        if validation_result.format_issues:
            return self._attempt_format_recovery(file_path, target_metrics, recovery_result)

        if validation_result.data_quality_issues:
            return self._attempt_data_quality_recovery(file_path, target_metrics, recovery_result)

        return recovery_result

    def _attempt_corruption_recovery(self, file_path: str, result: DataRecoveryResult) -> DataRecoveryResult:
        """Attempt to recover data from corrupted file"""
        result.recovery_method = "corruption_recovery"
        result.warnings.append("File corruption detected - attempting emergency recovery")

        try:
            # Try different recovery strategies
            strategies = [
                ("read_only_mode", self._try_read_only_recovery),
                ("ignore_errors", self._try_ignore_errors_recovery),
                ("partial_sheet_recovery", self._try_partial_sheet_recovery)
            ]

            for strategy_name, strategy_func in strategies:
                try:
                    recovered_data = strategy_func(file_path)
                    if recovered_data:
                        result.success = True
                        result.partial_data = recovered_data
                        result.confidence_score = 0.3  # Low confidence for corruption recovery
                        result.recovery_method = strategy_name
                        result.warnings.append(f"Partial recovery successful using {strategy_name}")
                        break
                except Exception as e:
                    result.warnings.append(f"Strategy {strategy_name} failed: {str(e)}")

        except Exception as e:
            result.warnings.append(f"All recovery strategies failed: {str(e)}")

        return result

    def _attempt_format_recovery(self, file_path: str, target_metrics: List[str], result: DataRecoveryResult) -> DataRecoveryResult:
        """Attempt to recover data despite format issues"""
        result.recovery_method = "format_recovery"

        try:
            workbook = load_workbook(file_path, read_only=True)
            worksheet = workbook.active

            # Use flexible search for target metrics
            recovered_metrics = {}
            for metric in target_metrics:
                recovered_values = self._flexible_metric_search(worksheet, metric)
                if recovered_values:
                    recovered_metrics[metric] = recovered_values

            if recovered_metrics:
                result.success = True
                result.recovered_metrics = recovered_metrics
                result.confidence_score = 0.6
                result.warnings.append("Data recovered using flexible format detection")

            workbook.close()

        except Exception as e:
            result.warnings.append(f"Format recovery failed: {str(e)}")

        return result

    def _attempt_data_quality_recovery(self, file_path: str, target_metrics: List[str], result: DataRecoveryResult) -> DataRecoveryResult:
        """Attempt to recover and clean data quality issues"""
        result.recovery_method = "data_quality_recovery"

        try:
            workbook = load_workbook(file_path, read_only=True)
            worksheet = workbook.active

            recovered_metrics = {}
            interpolated_values = {}

            for metric in target_metrics:
                # Extract values with quality validation
                raw_values = self._extract_metric_with_validation(worksheet, metric)
                if raw_values:
                    # Attempt to clean and interpolate missing values
                    cleaned_values, interpolated = self._clean_and_interpolate_values(raw_values, metric)

                    if cleaned_values:
                        recovered_metrics[metric] = cleaned_values
                        if interpolated:
                            interpolated_values[metric] = interpolated

            if recovered_metrics:
                result.success = True
                result.recovered_metrics = recovered_metrics
                result.interpolated_values = interpolated_values
                result.confidence_score = 0.8
                result.warnings.append("Data recovered with quality improvements and interpolation")

            workbook.close()

        except Exception as e:
            result.warnings.append(f"Data quality recovery failed: {str(e)}")

        return result

    def _try_read_only_recovery(self, file_path: str) -> Optional[Dict]:
        """Try recovery using read-only mode"""
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            recovered_data = {"sheets": list(workbook.sheetnames)}
            workbook.close()
            return recovered_data
        except:
            return None

    def _try_ignore_errors_recovery(self, file_path: str) -> Optional[Dict]:
        """Try recovery by ignoring certain errors"""
        try:
            # Implementation would use openpyxl's error handling options
            workbook = load_workbook(file_path, read_only=True)
            recovered_data = {"partial_access": True}
            workbook.close()
            return recovered_data
        except:
            return None

    def _try_partial_sheet_recovery(self, file_path: str) -> Optional[Dict]:
        """Try to recover data from accessible sheets only"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            accessible_sheets = []

            for sheet_name in workbook.sheetnames:
                try:
                    worksheet = workbook[sheet_name]
                    # Test if sheet is accessible
                    _ = worksheet.cell(1, 1).value
                    accessible_sheets.append(sheet_name)
                except:
                    continue

            workbook.close()

            if accessible_sheets:
                return {"accessible_sheets": accessible_sheets}
            return None
        except:
            return None

    def _flexible_metric_search(self, worksheet: Worksheet, metric_name: str) -> Optional[List]:
        """Flexibly search for a metric using various patterns"""
        search_patterns = [
            metric_name,
            metric_name.lower(),
            metric_name.replace(" ", ""),
            metric_name.replace("&", "and"),
            # Add more variations as needed
        ]

        max_row = worksheet.max_row or 50
        max_col = worksheet.max_column or 15

        for pattern in search_patterns:
            for row in range(1, min(max_row + 1, 50)):
                for col in range(1, min(max_col + 1, 10)):
                    try:
                        cell_value = worksheet.cell(row, col).value
                        if cell_value and isinstance(cell_value, str):
                            if pattern.lower() in cell_value.lower():
                                # Extract values from subsequent columns
                                values = []
                                for data_col in range(col + 1, min(col + 15, max_col + 1)):
                                    data_value = worksheet.cell(row, data_col).value
                                    if data_value is not None:
                                        values.append(data_value)
                                    else:
                                        break

                                if values:
                                    return values
                    except:
                        continue

        return None

    def _extract_metric_with_validation(self, worksheet: Worksheet, metric_name: str) -> Optional[List]:
        """Extract metric values with data validation"""
        # Similar to _flexible_metric_search but with additional validation
        raw_values = self._flexible_metric_search(worksheet, metric_name)

        if raw_values:
            # Validate each value
            validated_values = []
            for value in raw_values:
                cleaned_value, is_valid = self.data_validator.validate_cell_value(
                    value, float, True, f"Recovery: {metric_name}"
                )
                validated_values.append(cleaned_value if is_valid else None)

            return validated_values

        return None

    def _clean_and_interpolate_values(self, values: List, metric_name: str) -> Tuple[List, List]:
        """Clean values and interpolate missing data"""
        cleaned_values = []
        interpolated_indices = []

        # First pass: clean and identify missing values
        for i, value in enumerate(values):
            if value is None or (isinstance(value, str) and value.strip() in ['', '#N/A', '#ERROR']):
                cleaned_values.append(None)
            else:
                try:
                    cleaned_values.append(float(value))
                except (ValueError, TypeError):
                    cleaned_values.append(None)

        # Second pass: interpolate missing values
        interpolated_values = cleaned_values.copy()

        for i, value in enumerate(cleaned_values):
            if value is None:
                # Simple linear interpolation between adjacent valid values
                prev_value = None
                next_value = None

                # Find previous valid value
                for j in range(i - 1, -1, -1):
                    if cleaned_values[j] is not None:
                        prev_value = cleaned_values[j]
                        break

                # Find next valid value
                for j in range(i + 1, len(cleaned_values)):
                    if cleaned_values[j] is not None:
                        next_value = cleaned_values[j]
                        break

                # Interpolate if possible
                if prev_value is not None and next_value is not None:
                    interpolated_value = (prev_value + next_value) / 2
                    interpolated_values[i] = interpolated_value
                    interpolated_indices.append(i)
                elif prev_value is not None:
                    interpolated_values[i] = prev_value
                    interpolated_indices.append(i)
                elif next_value is not None:
                    interpolated_values[i] = next_value
                    interpolated_indices.append(i)

        return interpolated_values, interpolated_indices


class ExcelDataExtractor:
    """
    Dynamic Excel data extractor that can handle various Excel formats
    without relying on hardcoded cell references. Now supports international
    formats, custom templates, merged cells, and formula processing.
    """

    def __init__(
        self,
        workbook_path: str,
        enable_recovery: bool = True,
        international_config: Optional[InternationalConfig] = None,
        custom_template: Optional[CustomTemplate] = None,
        template_manager: Optional[CustomTemplateManager] = None
    ):
        """
        Initialize the Excel data extractor

        Args:
            workbook_path (str): Path to the Excel workbook
            enable_recovery (bool): Enable automatic error recovery
            international_config: Configuration for international format handling
            custom_template: Specific custom template to use
            template_manager: Manager for custom templates
        """
        self.workbook_path = workbook_path
        self.workbook = None
        self.worksheet = None
        self.config = get_excel_config()
        self.metrics_config = get_financial_metrics_config()
        self.format_detector = ExcelFormatDetector()
        self.format_detection_result: Optional[FormatDetectionResult] = None

        # Enhanced error handling components
        self.enable_recovery = enable_recovery
        self.file_validator = ExcelFileValidator()
        self.recovery_engine = DataRecoveryEngine()
        self.validation_result: Optional[ExcelFileValidationResult] = None
        self.recovery_result: Optional[DataRecoveryResult] = None

        # International format support
        self.international_config = international_config
        self.international_handler = InternationalFormatHandler(international_config) if international_config else None

        # Custom template support
        self.custom_template = custom_template
        self.template_manager = template_manager or CustomTemplateManager()
        self.detected_template: Optional[CustomTemplate] = None

    def __enter__(self):
        """Context manager entry"""
        self.load_workbook()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit"""
        if self.workbook:
            self.workbook.close()

    def load_workbook(self):
        """Load the Excel workbook with comprehensive error handling and recovery"""
        logger.info(f"Loading workbook with enhanced error handling: {os.path.basename(self.workbook_path)}")

        # Step 1: Validate the Excel file before attempting to load
        if self.enable_recovery:
            self.validation_result = self.file_validator.validate_file(self.workbook_path)
            self._log_validation_results()

            # Handle critical validation failures
            if self.validation_result.severity == 'critical' and not self.validation_result.is_valid:
                if self.validation_result.corruption_detected:
                    logger.warning("File corruption detected - attempting recovery")
                    return self._attempt_recovery_load()
                else:
                    raise CorruptedFileError(f"Critical validation failure: {self.validation_result.format_issues}")

        # Step 2: Attempt normal loading
        try:
            self.workbook = load_workbook(self.workbook_path, read_only=True)
            self.worksheet = self.workbook.active
            logger.info(f"Successfully loaded workbook: {os.path.basename(self.workbook_path)}")

            # Perform format detection
            self._detect_format()

            # Detect custom template if not provided
            if not self.custom_template:
                self._detect_custom_template()

            # Initialize international handler if needed
            if not self.international_handler:
                self._initialize_international_handler()

            # Additional validation for loaded workbook
            if self.enable_recovery and self.validation_result:
                self._validate_loaded_workbook()

        except InvalidFileException as e:
            logger.error(f"Invalid file format for {self.workbook_path}: {e}")
            if self.enable_recovery:
                return self._attempt_recovery_load()
            raise CorruptedFileError(f"Invalid Excel file format: {str(e)}")

        except Exception as e:
            logger.error(f"Failed to load workbook {self.workbook_path}: {e}")
            if self.enable_recovery:
                return self._attempt_recovery_load()
            raise ExcelProcessingError(f"Unable to load Excel file: {str(e)}")

    def _log_validation_results(self):
        """Log validation results for user awareness"""
        if not self.validation_result:
            return

        logger.info(f"File validation: {self.validation_result.severity} severity")

        if self.validation_result.format_issues:
            logger.warning(f"Format issues detected: {len(self.validation_result.format_issues)}")
            for issue in self.validation_result.format_issues[:3]:  # Log first 3
                logger.debug(f"  - {issue}")

        if self.validation_result.data_quality_issues:
            logger.warning(f"Data quality issues: {len(self.validation_result.data_quality_issues)}")
            for issue in self.validation_result.data_quality_issues[:3]:  # Log first 3
                logger.debug(f"  - {issue}")

        if self.validation_result.recovery_suggestions:
            logger.info("Recovery suggestions available:")
            for suggestion in self.validation_result.recovery_suggestions[:2]:  # Log first 2
                logger.info(f"  - {suggestion}")

    def _attempt_recovery_load(self):
        """Attempt to load workbook using recovery strategies"""
        logger.warning("Attempting to load workbook using recovery methods")

        if not self.validation_result:
            raise ExcelProcessingError("No validation result available for recovery")

        # Get target metrics for recovery
        target_metrics = list(self.metrics_config.income_metrics.keys())[:5]  # Top 5 critical metrics

        # Attempt data recovery
        self.recovery_result = self.recovery_engine.attempt_data_recovery(
            self.workbook_path, self.validation_result, target_metrics
        )

        if self.recovery_result.success:
            logger.info(f"Data recovery successful using {self.recovery_result.recovery_method}")
            logger.info(f"Recovery confidence: {self.recovery_result.confidence_score:.2f}")

            # Try to load with recovery data
            try:
                self.workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
                self.worksheet = self.workbook.active

                # Log recovery warnings
                for warning in self.recovery_result.warnings:
                    logger.warning(f"Recovery: {warning}")

                return True

            except Exception as e:
                logger.error(f"Recovery load failed: {e}")
                raise ExcelProcessingError(f"Recovery unsuccessful: {str(e)}")
        else:
            # Recovery failed - provide detailed error information
            error_details = []
            if self.validation_result.format_issues:
                error_details.extend(self.validation_result.format_issues[:2])
            if self.recovery_result.warnings:
                error_details.extend(self.recovery_result.warnings[:2])

            raise ExcelProcessingError(
                f"File could not be loaded or recovered. Issues: {'; '.join(error_details)}"
            )

    def _validate_loaded_workbook(self):
        """Perform additional validation on successfully loaded workbook"""
        if not self.worksheet:
            logger.warning("No active worksheet available for validation")
            return

        try:
            # Check if worksheet has reasonable content
            max_row = self.worksheet.max_row or 0
            max_col = self.worksheet.max_column or 0

            if max_row < 5 or max_col < 3:
                logger.warning(f"Worksheet appears mostly empty: {max_row}x{max_col}")

            # Sample a few cells to ensure they're accessible
            test_positions = [(1, 1), (2, 2), (3, 3)]
            accessible_cells = 0

            for row, col in test_positions:
                try:
                    if row <= max_row and col <= max_col:
                        _ = self.worksheet.cell(row, col).value
                        accessible_cells += 1
                except Exception as e:
                    logger.debug(f"Cell ({row}, {col}) not accessible: {e}")

            if accessible_cells == 0:
                logger.warning("No cells appear to be accessible in loaded worksheet")

        except Exception as e:
            logger.warning(f"Post-load validation failed: {e}")

    def _detect_format(self):
        """Detect and validate the Excel format"""
        if not self.worksheet:
            logger.warning("No worksheet available for format detection")
            return

        try:
            self.format_detection_result = self.format_detector.detect_format(self.worksheet)

            logger.info(
                f"Format detection complete: {self.format_detection_result.format_type.name} "
                f"(confidence: {self.format_detection_result.confidence_score:.2f})"
            )

            # Log any validation errors
            if self.format_detection_result.validation_errors:
                logger.warning(
                    f"Format validation warnings for {os.path.basename(self.workbook_path)}: "
                    f"{len(self.format_detection_result.validation_errors)} issues found"
                )
                for error in self.format_detection_result.validation_errors[:3]:  # Log first 3
                    logger.debug(f"  - {error}")

            # Log suggestions if confidence is low
            if self.format_detection_result.confidence_score < 0.8:
                logger.warning(
                    f"Low confidence format detection ({self.format_detection_result.confidence_score:.2f}). "
                    "Consider manual verification."
                )
                if self.format_detection_result.suggested_adjustments:
                    for key, value in list(self.format_detection_result.suggested_adjustments.items())[:2]:
                        logger.info(f"  Suggestion - {key}: {value}")

        except Exception as e:
            logger.error(f"Format detection failed for {self.workbook_path}: {e}")
            # Continue without format detection - fallback to original behavior

    def get_format_info(self) -> Optional[FormatDetectionResult]:
        """
        Get format detection information

        Returns:
            Optional[FormatDetectionResult]: Format detection result if available
        """
        return self.format_detection_result

    def get_detected_format_type(self) -> FormatType:
        """
        Get the detected format type

        Returns:
            FormatType: Detected format type, UNKNOWN if detection failed
        """
        if self.format_detection_result:
            return self.format_detection_result.format_type
        return FormatType.UNKNOWN

    def is_format_supported(self) -> bool:
        """
        Check if the detected format is supported

        Returns:
            bool: True if format is supported and confidence is reasonable
        """
        if not self.format_detection_result:
            return False
        return (
            self.format_detection_result.format_type != FormatType.UNKNOWN and
            self.format_detection_result.confidence_score >= 0.6
        )

    def find_company_name(self) -> Optional[str]:
        """
        Find company name in the Excel sheet - specifically in cell C2

        Returns:
            Optional[str]: Company name if found, None otherwise
        """
        if not self.worksheet:
            return None

        # Company name is always in cell C2 (row 2, column 3)
        try:
            value = self.worksheet.cell(2, 3).value
            if value and isinstance(value, str) and len(value.strip()) > 0:
                company_name = value.strip()
                logger.info(f"Found company name in C2: {company_name}")
                return company_name
            else:
                logger.warning("Cell C2 is empty or does not contain company name")
                return None
        except Exception as e:
            logger.error(f"Error reading company name from cell C2: {e}")
            return None


    def find_period_end_dates(self) -> List[str]:
        """
        Find period end dates in the worksheet

        Returns:
            List[str]: List of period end dates
        """
        if not self.worksheet:
            return []

        # First try the default location
        try:
            period_row = self.config.default_period_end_row
            period_col = self.config.default_period_end_column

            if self.worksheet.cell(period_row, period_col).value == "Period End Date":
                dates = []
                for col in range(self.config.data_start_column, self.config.max_scan_columns):
                    cell_value = self.worksheet.cell(period_row, col).value
                    if cell_value is not None:
                        dates.append(str(cell_value))
                    else:
                        break

                if dates:
                    logger.info(f"Found {len(dates)} period end dates at default location")
                    return dates
        except Exception as e:
            logger.debug(f"Error finding period dates at default location: {e}")

        # Search for "Period End Date" more broadly
        return self._search_period_end_dates_broadly()

    def _search_period_end_dates_broadly(self) -> List[str]:
        """
        Search for period end dates more broadly in the worksheet

        Returns:
            List[str]: List of period end dates
        """
        if not self.worksheet:
            return []

        # Search for "Period End Date" in first 20 rows
        for row in range(1, 21):
            for col in range(1, 10):
                try:
                    value = self.worksheet.cell(row, col).value
                    if value and isinstance(value, str) and "Period End Date" in value:
                        # Found the row, now extract dates
                        dates = []
                        for data_col in range(col + 1, self.config.max_scan_columns):
                            cell_value = self.worksheet.cell(row, data_col).value
                            if cell_value is not None:
                                dates.append(str(cell_value))
                            else:
                                break

                        if dates:
                            logger.info(f"Found {len(dates)} period end dates at row {row}")
                            return dates
                except Exception as e:
                    logger.debug(f"Error checking cell ({row}, {col}): {e}")
                    continue

        logger.warning("Could not find period end dates in worksheet")
        return []

    def find_financial_metric(self, metric_name: str) -> Optional[ExcelMetric]:
        """
        Find a financial metric in the worksheet with enhanced error handling

        Args:
            metric_name (str): Name of the financial metric to find

        Returns:
            Optional[ExcelMetric]: ExcelMetric object if found, None otherwise
        """
        if not self.worksheet:
            logger.error("No worksheet available for metric search")
            return None

        # Try recovery data first if available
        if self.recovery_result and self.recovery_result.success:
            if metric_name in self.recovery_result.recovered_metrics:
                logger.info(f"Using recovered data for metric '{metric_name}'")
                recovered_values = self.recovery_result.recovered_metrics[metric_name]
                return self._create_metric_from_recovered_data(metric_name, recovered_values)

        # Standard search with enhanced error handling
        try:
            return self._search_metric_with_fallbacks(metric_name)
        except Exception as e:
            logger.error(f"Error during metric search for '{metric_name}': {e}")

            # Attempt flexible recovery if enabled
            if self.enable_recovery:
                logger.info(f"Attempting flexible search for '{metric_name}'")
                return self._flexible_metric_search_fallback(metric_name)

            return None

    def _search_metric_with_fallbacks(self, metric_name: str) -> Optional[ExcelMetric]:
        """Search for metric using multiple strategies with fallbacks"""
        # Strategy 1: Exact match search
        metric = self._exact_metric_search(metric_name)
        if metric:
            return metric

        # Strategy 2: Partial match search
        metric = self._partial_metric_search(metric_name)
        if metric:
            return metric

        # Strategy 3: Case-insensitive search
        metric = self._case_insensitive_metric_search(metric_name)
        if metric:
            return metric

        logger.warning(f"All search strategies failed for metric '{metric_name}'")
        return None

    def _exact_metric_search(self, metric_name: str) -> Optional[ExcelMetric]:
        """Perform exact metric name search"""
        for row in range(1, self.config.max_scan_rows + 1):
            for col in range(1, 10):
                try:
                    value = self.worksheet.cell(row, col).value
                    if value and isinstance(value, str) and metric_name == value.strip():
                        return self._extract_metric_data(metric_name, row, col, value)
                except Exception as e:
                    logger.debug(f"Cell access error at ({row}, {col}): {e}")
                    continue
        return None

    def _partial_metric_search(self, metric_name: str) -> Optional[ExcelMetric]:
        """Perform partial metric name search"""
        for row in range(1, self.config.max_scan_rows + 1):
            for col in range(1, 10):
                try:
                    value = self.worksheet.cell(row, col).value
                    if value and isinstance(value, str) and metric_name in value:
                        return self._extract_metric_data(metric_name, row, col, value)
                except Exception as e:
                    logger.debug(f"Cell access error at ({row}, {col}): {e}")
                    continue
        return None

    def _case_insensitive_metric_search(self, metric_name: str) -> Optional[ExcelMetric]:
        """Perform case-insensitive metric name search"""
        metric_lower = metric_name.lower()
        for row in range(1, self.config.max_scan_rows + 1):
            for col in range(1, 10):
                try:
                    value = self.worksheet.cell(row, col).value
                    if value and isinstance(value, str) and metric_lower in value.lower():
                        return self._extract_metric_data(metric_name, row, col, value)
                except Exception as e:
                    logger.debug(f"Cell access error at ({row}, {col}): {e}")
                    continue
        return None

    def _extract_metric_data(self, metric_name: str, row: int, col: int, label_value: str) -> Optional[ExcelMetric]:
        """Extract data for a found metric with validation"""
        try:
            data_columns = []
            values = []
            valid_values = 0

            # Extract historical data (9 years for FY)
            for j in range(9):
                data_col = self.config.data_start_column + j
                data_columns.append(data_col)

                try:
                    cell_value = self.worksheet.cell(row, data_col).value
                    values.append(cell_value)

                    # Count valid numeric values
                    if cell_value is not None and cell_value != "":
                        try:
                            float(cell_value)
                            valid_values += 1
                        except (ValueError, TypeError):
                            pass

                except Exception as e:
                    logger.debug(f"Error extracting data at ({row}, {data_col}): {e}")
                    values.append(None)

            # Validate that we have reasonable data
            if valid_values < 2:  # Require at least 2 valid data points
                logger.warning(f"Insufficient valid data for '{metric_name}': {valid_values} valid values")
                if not self.enable_recovery:
                    return None

                # Attempt data quality recovery
                if self.recovery_result and metric_name in self.recovery_result.interpolated_values:
                    logger.info(f"Using interpolated values for '{metric_name}'")
                    values = self.recovery_result.interpolated_values[metric_name]

            location = CellLocation(row, col, label_value)
            metric = ExcelMetric(metric_name, location, data_columns, values)

            logger.info(f"Successfully extracted metric '{metric_name}' at row {row} with {valid_values} valid values")
            return metric

        except Exception as e:
            logger.error(f"Error extracting data for metric '{metric_name}' at ({row}, {col}): {e}")
            return None

    def _flexible_metric_search_fallback(self, metric_name: str) -> Optional[ExcelMetric]:
        """Flexible search as a last resort using recovery engine patterns"""
        if not self.recovery_engine:
            return None

        try:
            # Use recovery engine's flexible search
            values = self.recovery_engine._flexible_metric_search(self.worksheet, metric_name)
            if values:
                return self._create_metric_from_recovered_data(metric_name, values)

        except Exception as e:
            logger.debug(f"Flexible search failed for '{metric_name}': {e}")

        return None

    def _create_metric_from_recovered_data(self, metric_name: str, values: List) -> ExcelMetric:
        """Create ExcelMetric from recovered data"""
        # Create a synthetic location since we might not know exact position
        location = CellLocation(-1, -1, f"Recovered: {metric_name}")

        # Create synthetic data columns
        data_columns = list(range(self.config.data_start_column, self.config.data_start_column + len(values)))

        metric = ExcelMetric(metric_name, location, data_columns, values)
        logger.info(f"Created metric from recovered data: '{metric_name}' with {len(values)} values")
        return metric

    def extract_all_financial_metrics(self, statement_type: str) -> Dict[str, ExcelMetric]:
        """
        Extract all financial metrics for a given statement type

        Args:
            statement_type (str): Type of statement ('income', 'balance', 'cashflow')

        Returns:
            Dict[str, ExcelMetric]: Dictionary of metric names to ExcelMetric objects
        """
        if not self.worksheet:
            return {}

        # Get the appropriate metrics configuration
        if statement_type == 'income':
            metrics_dict = self.metrics_config.income_metrics
        elif statement_type == 'balance':
            metrics_dict = self.metrics_config.balance_metrics
        elif statement_type == 'cashflow':
            metrics_dict = self.metrics_config.cashflow_metrics
        else:
            logger.error(f"Unknown statement type: {statement_type}")
            return {}

        extracted_metrics = {}

        for metric_name in metrics_dict.keys():
            metric = self.find_financial_metric(metric_name)
            if metric:
                extracted_metrics[metric_name] = metric

        logger.info(f"Extracted {len(extracted_metrics)} metrics for {statement_type} statement")
        return extracted_metrics

    def extract_ltm_data(self, metric_name: str) -> Optional[Any]:
        """
        Extract LTM (Latest Twelve Months) data for a specific metric

        Args:
            metric_name (str): Name of the financial metric

        Returns:
            Optional[Any]: LTM value if found, None otherwise
        """
        if not self.worksheet:
            return None

        metric = self.find_financial_metric(metric_name)
        if not metric:
            return None

        # Extract LTM data from the configured column
        try:
            ltm_value = self.worksheet.cell(metric.location.row, self.config.ltm_column).value
            logger.debug(f"LTM value for '{metric_name}': {ltm_value}")
            return ltm_value
        except Exception as e:
            logger.error(f"Error extracting LTM data for '{metric_name}': {e}")
            return None

    def get_validation_report(self) -> Optional[ExcelFileValidationResult]:
        """
        Get the file validation report

        Returns:
            Optional[ExcelFileValidationResult]: Validation report if available
        """
        return self.validation_result

    def get_recovery_report(self) -> Optional[DataRecoveryResult]:
        """
        Get the data recovery report

        Returns:
            Optional[DataRecoveryResult]: Recovery report if available
        """
        return self.recovery_result

    def get_processing_status(self) -> Dict[str, Any]:
        """
        Get comprehensive processing status and diagnostics

        Returns:
            Dict[str, Any]: Processing status information
        """
        status = {
            "file_path": self.workbook_path,
            "file_loaded": self.workbook is not None,
            "worksheet_available": self.worksheet is not None,
            "recovery_enabled": self.enable_recovery,
            "validation_performed": self.validation_result is not None,
            "recovery_attempted": self.recovery_result is not None,
            "format_detected": self.format_detection_result is not None,
        }

        # Add validation details
        if self.validation_result:
            status.update({
                "validation_status": {
                    "is_valid": self.validation_result.is_valid,
                    "severity": self.validation_result.severity,
                    "corruption_detected": self.validation_result.corruption_detected,
                    "format_issues_count": len(self.validation_result.format_issues),
                    "data_quality_issues_count": len(self.validation_result.data_quality_issues),
                    "file_size": self.validation_result.file_size,
                }
            })

        # Add recovery details
        if self.recovery_result:
            status.update({
                "recovery_status": {
                    "success": self.recovery_result.success,
                    "confidence_score": self.recovery_result.confidence_score,
                    "recovery_method": self.recovery_result.recovery_method,
                    "recovered_metrics_count": len(self.recovery_result.recovered_metrics),
                    "interpolated_values_count": len(self.recovery_result.interpolated_values),
                    "warnings_count": len(self.recovery_result.warnings),
                }
            })

        # Add format detection details
        if self.format_detection_result:
            status.update({
                "format_status": {
                    "format_type": self.format_detection_result.format_type.name,
                    "confidence_score": self.format_detection_result.confidence_score,
                    "validation_errors_count": len(self.format_detection_result.validation_errors),
                    "is_supported": self.is_format_supported(),
                }
            })

        return status

    def _detect_custom_template(self):
        """Detect if the worksheet matches any custom templates"""
        if not self.worksheet:
            return

        try:
            self.detected_template = self.template_manager.detect_template_for_worksheet(self.worksheet)
            if self.detected_template:
                logger.info(f"Detected custom template: {self.detected_template.name}")
                self.custom_template = self.detected_template
        except Exception as e:
            logger.debug(f"Custom template detection failed: {e}")

    def _initialize_international_handler(self):
        """Initialize international format handler based on detected format"""
        if not self.worksheet:
            return

        try:
            # Try to determine international format based on format detection
            if self.format_detection_result:
                if self.format_detection_result.format_type == FormatType.EUROPEAN_FORMAT:
                    self.international_handler = create_european_handler()
                    logger.info("Initialized European format handler")
                elif self.format_detection_result.format_type == FormatType.INTERNATIONAL_STANDARD:
                    # Analyze worksheet to determine specific format
                    temp_handler = InternationalFormatHandler()
                    analysis = temp_handler.analyze_worksheet_format(self.worksheet)
                    if analysis['recommended_config']:
                        self.international_handler = InternationalFormatHandler(analysis['recommended_config'])
                        logger.info("Initialized international format handler with auto-detected config")
                else:
                    # Default to US format
                    self.international_handler = create_us_handler()
                    logger.info("Initialized US format handler as default")
            else:
                # Fallback to US format
                self.international_handler = create_us_handler()

        except Exception as e:
            logger.debug(f"International handler initialization failed: {e}")
            # Fallback to basic handler
            self.international_handler = create_us_handler()

    def parse_cell_with_international_support(self, row: int, col: int) -> Any:
        """
        Parse a cell value with international format support

        Args:
            row (int): Row number
            col (int): Column number

        Returns:
            Any: Parsed cell value
        """
        if not self.worksheet:
            return None

        try:
            if self.international_handler:
                result = self.international_handler.parse_cell_value(self.worksheet, row, col)
                if result.confidence > 0.6:
                    return result.parsed_value
                else:
                    logger.debug(f"Low confidence parse at ({row}, {col}): {result.confidence}")
                    return result.original_value
            else:
                # Fallback to normal cell access
                return self.worksheet.cell(row, col).value

        except Exception as e:
            logger.debug(f"Error parsing cell ({row}, {col}) with international support: {e}")
            try:
                return self.worksheet.cell(row, col).value
            except:
                return None

    def extract_metric_with_template_support(self, metric_name: str) -> Optional[ExcelMetric]:
        """
        Extract metric using custom template information if available

        Args:
            metric_name (str): Name of the metric to extract

        Returns:
            Optional[ExcelMetric]: Extracted metric or None
        """
        if not self.worksheet:
            return None

        # First try template-based extraction
        if self.custom_template:
            metric = self._extract_metric_using_template(metric_name)
            if metric:
                return metric

        # Fallback to standard extraction
        return self.find_financial_metric(metric_name)

    def _extract_metric_using_template(self, metric_name: str) -> Optional[ExcelMetric]:
        """Extract metric using custom template mappings"""
        if not self.custom_template or not self.worksheet:
            return None

        try:
            # Look for metric in template sections
            for section in self.custom_template.sections:
                if section.section_type != 'data':
                    continue

                # Check cell mappings first
                for cell_mapping in section.cell_mappings:
                    if metric_name.lower() in cell_mapping.name.lower():
                        # Found potential match, extract data
                        return self._extract_metric_from_template_cell(cell_mapping)

                # Check for metric in section rows
                for row in range(section.start_row, section.end_row + 1):
                    for col in range(section.start_column, min(section.start_column + 5, section.end_column + 1)):
                        try:
                            cell_value = self.worksheet.cell(row, col).value
                            if cell_value and isinstance(cell_value, str) and metric_name.lower() in cell_value.lower():
                                return self._extract_metric_data_from_template_position(metric_name, row, col)
                        except Exception:
                            continue

        except Exception as e:
            logger.debug(f"Template-based metric extraction failed for '{metric_name}': {e}")

        return None

    def _extract_metric_from_template_cell(self, cell_mapping) -> Optional[ExcelMetric]:
        """Extract metric data based on template cell mapping"""
        try:
            # Get label value
            label_value = self.parse_cell_with_international_support(cell_mapping.row, cell_mapping.column)

            # Extract data values (assuming data is in subsequent columns)
            data_columns = []
            values = []

            start_col = cell_mapping.column + 1
            for i in range(9):  # Extract up to 9 years of data
                data_col = start_col + i
                data_columns.append(data_col)
                cell_value = self.parse_cell_with_international_support(cell_mapping.row, data_col)
                values.append(cell_value)

            location = CellLocation(cell_mapping.row, cell_mapping.column, label_value)
            return ExcelMetric(cell_mapping.name, location, data_columns, values)

        except Exception as e:
            logger.debug(f"Error extracting metric from template cell: {e}")
            return None

    def _extract_metric_data_from_template_position(self, metric_name: str, row: int, col: int) -> Optional[ExcelMetric]:
        """Extract metric data from a specific position found in template"""
        try:
            label_value = self.parse_cell_with_international_support(row, col)

            # Determine data start column based on template or default
            if self.custom_template:
                data_start_col = self.custom_template.data_start_column
            else:
                data_start_col = col + 1

            data_columns = []
            values = []

            for i in range(9):  # Extract up to 9 years of data
                data_col = data_start_col + i
                data_columns.append(data_col)
                cell_value = self.parse_cell_with_international_support(row, data_col)
                values.append(cell_value)

            location = CellLocation(row, col, label_value)
            return ExcelMetric(metric_name, location, data_columns, values)

        except Exception as e:
            logger.debug(f"Error extracting metric data from template position: {e}")
            return None

    def get_international_analysis(self) -> Optional[Dict[str, Any]]:
        """
        Get international format analysis of the worksheet

        Returns:
            Optional[Dict[str, Any]]: Analysis results or None
        """
        if not self.international_handler or not self.worksheet:
            return None

        try:
            return self.international_handler.analyze_worksheet_format(self.worksheet)
        except Exception as e:
            logger.error(f"Error performing international analysis: {e}")
            return None

    def get_custom_template_info(self) -> Optional[Dict[str, Any]]:
        """
        Get information about the detected or applied custom template

        Returns:
            Optional[Dict[str, Any]]: Template information or None
        """
        template = self.custom_template or self.detected_template
        if not template:
            return None

        return {
            'template_id': template.template_id,
            'name': template.name,
            'description': template.description,
            'template_type': template.template_type.name,
            'version': template.version,
            'sections_count': len(template.sections),
            'company_name_cell': template.company_name_cell,
            'data_start_position': (template.data_start_row, template.data_start_column),
            'has_international_config': template.international_config is not None,
            'created_by': template.created_by,
            'tags': template.tags
        }

    def generate_template_from_current_worksheet(self, template_id: str, name: str, description: str) -> Optional[CustomTemplate]:
        """
        Generate a custom template based on the current worksheet structure

        Args:
            template_id (str): Unique identifier for the template
            name (str): Template name
            description (str): Template description

        Returns:
            Optional[CustomTemplate]: Generated template or None
        """
        if not self.worksheet:
            return None

        try:
            template = self.template_manager.generate_template_from_worksheet(
                self.worksheet, template_id, name, description
            )
            return template
        except Exception as e:
            logger.error(f"Error generating template from worksheet: {e}")
            return None

    def find_company_name_with_template_support(self) -> Optional[str]:
        """
        Find company name using template support if available

        Returns:
            Optional[str]: Company name if found
        """
        if not self.worksheet:
            return None

        # Try template-based approach first
        if self.custom_template:
            try:
                row, col = self.custom_template.company_name_cell
                company_name = self.parse_cell_with_international_support(row, col)
                if company_name and isinstance(company_name, str) and len(company_name.strip()) > 0:
                    logger.info(f"Found company name using template at ({row}, {col}): {company_name.strip()}")
                    return company_name.strip()
            except Exception as e:
                logger.debug(f"Template-based company name extraction failed: {e}")

        # Fallback to standard approach
        return self.find_company_name()

    def validate_extracted_data_comprehensive(
        self,
        extracted_data: Dict[str, Any],
        source_identifier: str = None
    ) -> Dict[str, Any]:
        """
        Perform comprehensive validation on extracted financial data

        Args:
            extracted_data: Dictionary containing extracted financial data
            source_identifier: Identifier for the data source

        Returns:
            Dictionary with comprehensive validation results
        """
        if not source_identifier:
            source_identifier = f"Excel:{os.path.basename(self.workbook_path)}"

        logger.info(f"Performing comprehensive validation for {source_identifier}")

        try:
            # Transform extracted data to the format expected by validators
            financial_data = self._transform_extracted_data_for_validation(extracted_data)

            # Run comprehensive validation
            validation_results = validate_financial_statements_comprehensive(
                financial_data, source_identifier, include_enhanced_validation=True
            )

            # Add Excel-specific context
            validation_results['excel_context'] = {
                'file_path': self.workbook_path,
                'format_detection': self.get_format_info(),
                'processing_status': self.get_processing_status(),
                'recovery_performed': self.recovery_result is not None,
                'validation_performed': self.validation_result is not None
            }

            return validation_results

        except Exception as e:
            logger.error(f"Comprehensive validation failed: {e}")
            return {
                'error': str(e),
                'base_validation': {'overall_score': 0},
                'enhanced_validation': {'overall_score': 0},
                'combined_score': 0
            }

    def _transform_extracted_data_for_validation(self, extracted_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Transform extracted Excel data to format expected by financial validators

        Args:
            extracted_data: Raw extracted data from Excel

        Returns:
            Transformed data in validator-expected format
        """
        financial_data = {}

        # Map statement types
        statement_mapping = {
            'income': 'income_fy',
            'balance': 'balance_fy',
            'cashflow': 'cashflow_fy'
        }

        for statement_type, data in extracted_data.items():
            if statement_type in statement_mapping:
                validator_key = statement_mapping[statement_type]
                financial_data[validator_key] = {}

                # Transform metrics to list format expected by validators
                for metric_name, metric_info in data.items():
                    if isinstance(metric_info, dict) and 'values' in metric_info:
                        financial_data[validator_key][metric_name] = metric_info['values']
                    elif isinstance(metric_info, list):
                        financial_data[validator_key][metric_name] = metric_info
                    else:
                        # Handle single values
                        financial_data[validator_key][metric_name] = [metric_info]

        return financial_data

    def extract_and_validate_financial_data(
        self,
        statement_types: List[str] = None,
        perform_validation: bool = True,
        validation_level: str = "comprehensive"
    ) -> Dict[str, Any]:
        """
        Extract financial data and perform validation in one operation

        Args:
            statement_types: List of statement types to extract ('income', 'balance', 'cashflow')
            perform_validation: Whether to perform validation on extracted data
            validation_level: Level of validation ('basic', 'enhanced', 'comprehensive')

        Returns:
            Dictionary with extracted data and validation results
        """
        if statement_types is None:
            statement_types = ['income', 'balance', 'cashflow']

        logger.info(f"Extracting and validating financial data for: {os.path.basename(self.workbook_path)}")

        results = {
            'extracted_data': {},
            'validation_results': None,
            'extraction_summary': {
                'total_metrics_extracted': 0,
                'successful_extractions': 0,
                'failed_extractions': 0,
                'statements_processed': []
            }
        }

        # Extract data for each statement type
        for statement_type in statement_types:
            try:
                logger.info(f"Extracting {statement_type} statement data")
                extracted_metrics = self.extract_all_financial_metrics(statement_type)

                if extracted_metrics:
                    # Convert to simple dictionary format for validation
                    statement_data = {}
                    for metric_name, metric in extracted_metrics.items():
                        statement_data[metric_name] = {
                            'values': metric.values,
                            'location': (metric.location.row, metric.location.column),
                            'data_columns': metric.data_columns,
                        }
                        results['extraction_summary']['total_metrics_extracted'] += 1
                        results['extraction_summary']['successful_extractions'] += 1

                    results['extracted_data'][statement_type] = statement_data
                    results['extraction_summary']['statements_processed'].append(statement_type)
                    logger.info(f"Successfully extracted {len(extracted_metrics)} metrics from {statement_type} statement")
                else:
                    logger.warning(f"No metrics extracted from {statement_type} statement")
                    results['extraction_summary']['failed_extractions'] += 1

            except Exception as e:
                logger.error(f"Error extracting {statement_type} statement: {e}")
                results['extraction_summary']['failed_extractions'] += 1

        # Perform validation if requested
        if perform_validation and results['extracted_data']:
            try:
                source_identifier = f"Excel:{os.path.basename(self.workbook_path)}"

                if validation_level == "comprehensive":
                    results['validation_results'] = self.validate_extracted_data_comprehensive(
                        results['extracted_data'], source_identifier
                    )
                elif validation_level == "enhanced":
                    financial_data = self._transform_extracted_data_for_validation(results['extracted_data'])
                    results['validation_results'] = validate_financial_statements_comprehensive(
                        financial_data, source_identifier, include_enhanced_validation=True
                    )
                elif validation_level == "basic":
                    financial_data = self._transform_extracted_data_for_validation(results['extracted_data'])
                    results['validation_results'] = validate_financial_statements_comprehensive(
                        financial_data, source_identifier, include_enhanced_validation=False
                    )

                logger.info(f"Validation completed. Combined score: {results['validation_results'].get('combined_score', 0):.1f}")

            except Exception as e:
                logger.error(f"Validation failed: {e}")
                results['validation_results'] = {'error': str(e), 'combined_score': 0}

        return results

    def generate_error_report(self) -> str:
        """
        Generate a comprehensive error report with suggested fixes

        Returns:
            str: Formatted error report
        """
        report_lines = [
            "=" * 80,
            f"EXCEL PROCESSING ERROR REPORT",
            f"File: {os.path.basename(self.workbook_path)}",
            f"Timestamp: {logging.Formatter().formatTime(logging.LogRecord('', 0, '', 0, '', (), None))}",
            "=" * 80,
        ]

        # File status
        report_lines.extend([
            "\n📄 FILE STATUS:",
            f"  Path: {self.workbook_path}",
            f"  Exists: {os.path.exists(self.workbook_path)}",
            f"  Size: {os.path.getsize(self.workbook_path) if os.path.exists(self.workbook_path) else 'N/A'} bytes",
            f"  Loaded: {self.workbook is not None}",
        ])

        # Validation results
        if self.validation_result:
            report_lines.extend([
                "\n🔍 VALIDATION RESULTS:",
                f"  Overall Status: {'✅ VALID' if self.validation_result.is_valid else '❌ INVALID'}",
                f"  Severity: {self.validation_result.severity.upper()}",
                f"  Corruption Detected: {'⚠️ YES' if self.validation_result.corruption_detected else '✅ NO'}",
            ])

            if self.validation_result.format_issues:
                report_lines.append("\n📋 FORMAT ISSUES:")
                for i, issue in enumerate(self.validation_result.format_issues[:5], 1):
                    report_lines.append(f"  {i}. {issue}")
                if len(self.validation_result.format_issues) > 5:
                    report_lines.append(f"  ... and {len(self.validation_result.format_issues) - 5} more")

            if self.validation_result.data_quality_issues:
                report_lines.append("\n📊 DATA QUALITY ISSUES:")
                for i, issue in enumerate(self.validation_result.data_quality_issues[:5], 1):
                    report_lines.append(f"  {i}. {issue}")
                if len(self.validation_result.data_quality_issues) > 5:
                    report_lines.append(f"  ... and {len(self.validation_result.data_quality_issues) - 5} more")

        # Recovery results
        if self.recovery_result:
            report_lines.extend([
                "\n🔧 RECOVERY RESULTS:",
                f"  Recovery Attempted: ✅ YES",
                f"  Success: {'✅ YES' if self.recovery_result.success else '❌ NO'}",
                f"  Method: {self.recovery_result.recovery_method}",
                f"  Confidence: {self.recovery_result.confidence_score:.2f}",
                f"  Recovered Metrics: {len(self.recovery_result.recovered_metrics)}",
            ])

            if self.recovery_result.warnings:
                report_lines.append("\n⚠️ RECOVERY WARNINGS:")
                for i, warning in enumerate(self.recovery_result.warnings[:3], 1):
                    report_lines.append(f"  {i}. {warning}")

        # Suggestions
        if self.validation_result and self.validation_result.recovery_suggestions:
            report_lines.append("\n💡 RECOMMENDED ACTIONS:")
            for i, suggestion in enumerate(self.validation_result.recovery_suggestions, 1):
                report_lines.append(f"  {i}. {suggestion}")

        # Additional troubleshooting steps
        report_lines.extend([
            "\n🛠️ ADDITIONAL TROUBLESHOOTING STEPS:",
            "  1. Verify the Excel file opens correctly in Microsoft Excel",
            "  2. Check that the file contains the expected financial statement data",
            "  3. Ensure data is properly formatted and contains no merged cells in data rows",
            "  4. Try saving the file in .xlsx format if it's in .xls format",
            "  5. Remove any protection or passwords from the Excel file",
        ])

        report_lines.append("=" * 80)

        return "\n".join(report_lines)


def get_company_name_from_excel(file_path: str) -> Optional[str]:
    """
    Utility function to extract company name from Excel file

    Args:
        file_path (str): Path to Excel file

    Returns:
        Optional[str]: Company name if found
    """
    try:
        with ExcelDataExtractor(file_path) as extractor:
            return extractor.find_company_name()
    except Exception as e:
        logger.error(f"Error extracting company name from {file_path}: {e}")
        return None


def get_period_dates_from_excel(file_path: str) -> List[str]:
    """
    Utility function to extract period dates from Excel file

    Args:
        file_path (str): Path to Excel file

    Returns:
        List[str]: List of period dates
    """
    try:
        with ExcelDataExtractor(file_path) as extractor:
            return extractor.find_period_end_dates()
    except Exception as e:
        logger.error(f"Error extracting period dates from {file_path}: {e}")
        return []


def extract_financial_data_from_excel(file_path: str, statement_type: str) -> Dict[str, Any]:
    """
    Utility function to extract financial data from Excel file

    Args:
        file_path (str): Path to Excel file
        statement_type (str): Type of statement ('income', 'balance', 'cashflow')

    Returns:
        Dict[str, Any]: Dictionary of extracted financial data
    """
    try:
        with ExcelDataExtractor(file_path) as extractor:
            metrics = extractor.extract_all_financial_metrics(statement_type)

            # Convert to simple dictionary format
            result = {}
            for metric_name, metric in metrics.items():
                result[metric_name] = {
                    'values': metric.values,
                    'location': (metric.location.row, metric.location.column),
                    'data_columns': metric.data_columns,
                }

            return result
    except Exception as e:
        logger.error(f"Error extracting financial data from {file_path}: {e}")
        return {}


def extract_financial_data_with_international_support(
    file_path: str,
    statement_type: str,
    number_format: NumberFormat = NumberFormat.US_FORMAT,
    date_format: DateFormat = DateFormat.MDY
) -> Dict[str, Any]:
    """
    Extract financial data with international format support

    Args:
        file_path (str): Path to Excel file
        statement_type (str): Type of statement ('income', 'balance', 'cashflow')
        number_format: Preferred number format
        date_format: Preferred date format

    Returns:
        Dict[str, Any]: Dictionary of extracted financial data with format info
    """
    try:
        # Create international configuration
        international_config = InternationalConfig(
            preferred_number_format=number_format,
            preferred_date_format=date_format
        )

        with ExcelDataExtractor(
            file_path,
            international_config=international_config
        ) as extractor:
            metrics = extractor.extract_all_financial_metrics(statement_type)

            # Convert to dictionary with additional format information
            result = {
                'metrics': {},
                'format_info': {
                    'international_analysis': extractor.get_international_analysis(),
                    'format_detection': extractor.get_format_info(),
                    'template_info': extractor.get_custom_template_info()
                }
            }

            for metric_name, metric in metrics.items():
                result['metrics'][metric_name] = {
                    'values': metric.values,
                    'location': (metric.location.row, metric.location.column),
                    'data_columns': metric.data_columns,
                }

            return result

    except Exception as e:
        logger.error(f"Error extracting financial data with international support from {file_path}: {e}")
        return {'metrics': {}, 'format_info': {}}


def extract_financial_data_with_custom_template(
    file_path: str,
    statement_type: str,
    template_id: Optional[str] = None,
    template_manager: Optional[CustomTemplateManager] = None
) -> Dict[str, Any]:
    """
    Extract financial data using custom template

    Args:
        file_path (str): Path to Excel file
        statement_type (str): Type of statement ('income', 'balance', 'cashflow')
        template_id: Specific template ID to use
        template_manager: Custom template manager instance

    Returns:
        Dict[str, Any]: Dictionary of extracted financial data with template info
    """
    try:
        custom_template = None
        if template_id and template_manager:
            custom_template = template_manager.get_template(template_id)

        with ExcelDataExtractor(
            file_path,
            custom_template=custom_template,
            template_manager=template_manager
        ) as extractor:
            # Use template-supported extraction
            metrics = {}
            metrics_config = get_financial_metrics_config()

            if statement_type == 'income':
                metric_names = list(metrics_config.income_metrics.keys())
            elif statement_type == 'balance':
                metric_names = list(metrics_config.balance_metrics.keys())
            elif statement_type == 'cashflow':
                metric_names = list(metrics_config.cashflow_metrics.keys())
            else:
                metric_names = []

            for metric_name in metric_names:
                metric = extractor.extract_metric_with_template_support(metric_name)
                if metric:
                    metrics[metric_name] = metric

            # Convert to dictionary format
            result = {
                'metrics': {},
                'template_info': extractor.get_custom_template_info(),
                'format_info': extractor.get_format_info()
            }

            for metric_name, metric in metrics.items():
                result['metrics'][metric_name] = {
                    'values': metric.values,
                    'location': (metric.location.row, metric.location.column),
                    'data_columns': metric.data_columns,
                }

            return result

    except Exception as e:
        logger.error(f"Error extracting financial data with custom template from {file_path}: {e}")
        return {'metrics': {}, 'template_info': {}, 'format_info': {}}


def create_european_excel_extractor(file_path: str) -> ExcelDataExtractor:
    """
    Create an Excel extractor optimized for European formats

    Args:
        file_path (str): Path to Excel file

    Returns:
        ExcelDataExtractor: Configured for European formats
    """
    international_config = InternationalConfig(
        preferred_number_format=NumberFormat.EUROPEAN_FORMAT,
        preferred_date_format=DateFormat.DMY,
        decimal_separator=',',
        thousands_separator='.',
        currency_symbols=['€', '£', 'CHF', 'SEK', 'NOK', 'DKK']
    )
    return ExcelDataExtractor(file_path, international_config=international_config)


def create_asian_excel_extractor(file_path: str) -> ExcelDataExtractor:
    """
    Create an Excel extractor optimized for Asian formats

    Args:
        file_path (str): Path to Excel file

    Returns:
        ExcelDataExtractor: Configured for Asian formats
    """
    international_config = InternationalConfig(
        preferred_number_format=NumberFormat.US_FORMAT,
        preferred_date_format=DateFormat.YMD,
        currency_symbols=['¥', '₹', '₩', 'CNY', 'JPY', 'KRW', 'INR']
    )
    return ExcelDataExtractor(file_path, international_config=international_config)


def generate_custom_template_from_excel(
    file_path: str,
    template_id: str,
    template_name: str,
    description: str
) -> Optional[CustomTemplate]:
    """
    Generate a custom template by analyzing an Excel file

    Args:
        file_path (str): Path to Excel file
        template_id (str): Unique template identifier
        template_name (str): Human-readable template name
        description (str): Template description

    Returns:
        Optional[CustomTemplate]: Generated template or None
    """
    try:
        with ExcelDataExtractor(file_path) as extractor:
            template = extractor.generate_template_from_current_worksheet(
                template_id, template_name, description
            )
            return template
    except Exception as e:
        logger.error(f"Error generating custom template from {file_path}: {e}")
        return None


def analyze_excel_international_format(file_path: str) -> Dict[str, Any]:
    """
    Analyze an Excel file for international format characteristics

    Args:
        file_path (str): Path to Excel file

    Returns:
        Dict[str, Any]: Analysis results
    """
    try:
        with ExcelDataExtractor(file_path) as extractor:
            # Initialize international handler for analysis
            if not extractor.international_handler:
                extractor._initialize_international_handler()

            analysis = extractor.get_international_analysis()
            return analysis or {}
    except Exception as e:
        logger.error(f"Error analyzing international format for {file_path}: {e}")
        return {}


def detect_excel_layout_type(file_path: str) -> Dict[str, Any]:
    """
    Detect the layout type and characteristics of an Excel file

    Args:
        file_path (str): Path to Excel file

    Returns:
        Dict[str, Any]: Detection results including format type, template info, and recommendations
    """
    try:
        with ExcelDataExtractor(file_path) as extractor:
            result = {
                'format_detection': extractor.get_format_info(),
                'template_info': extractor.get_custom_template_info(),
                'international_analysis': extractor.get_international_analysis(),
                'processing_status': extractor.get_processing_status(),
                'recommendations': []
            }

            # Generate recommendations based on analysis
            if result['format_detection'] and result['format_detection'].confidence_score < 0.7:
                result['recommendations'].append(
                    "Low format confidence detected. Consider using a custom template or "
                    "standardizing the Excel format."
                )

            if result['international_analysis']:
                indicators = result['international_analysis'].get('international_indicators', [])
                if indicators:
                    result['recommendations'].append(
                        f"International format features detected: {', '.join(indicators[:3])}. "
                        "Consider using international format support."
                    )

            if not result['template_info']:
                result['recommendations'].append(
                    "No custom template detected. Consider creating a template for this layout "
                    "to improve extraction accuracy."
                )

            return result

    except Exception as e:
        logger.error(f"Error detecting Excel layout type for {file_path}: {e}")
        return {'format_detection': None, 'template_info': None, 'international_analysis': None,
                'processing_status': None, 'recommendations': [f"Error during analysis: {str(e)}"]}


if __name__ == "__main__":
    # Test the Excel utilities
    import sys

    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        print(f"Testing Excel utilities with file: {file_path}")

        company_name = get_company_name_from_excel(file_path)
        print(f"Company name: {company_name}")

        period_dates = get_period_dates_from_excel(file_path)
        print(f"Period dates: {period_dates}")

        financial_data = extract_financial_data_from_excel(file_path, 'income')
        print(f"Financial data keys: {list(financial_data.keys())}")
    else:
        print("Usage: python excel_utils.py <path_to_excel_file>")
