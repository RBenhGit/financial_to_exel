"""
International Format Handler for Excel Data Processing
=====================================================

This module provides comprehensive support for international Excel layouts,
including different number formats, date formats, and localization patterns
commonly used in financial statements across different regions and countries.

Key Features:
- European number formats (comma as decimal separator)
- International date format parsing
- Multi-language financial term recognition
- Currency format detection and parsing
- Merged cell handling and data extraction
- Formula cell processing and evaluation
"""

import re
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any, Union
from dataclasses import dataclass
from enum import Enum, auto
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

logger = logging.getLogger(__name__)


class NumberFormat(Enum):
    """Supported number format conventions"""
    US_FORMAT = auto()      # 1,234.56
    EUROPEAN_FORMAT = auto() # 1.234,56
    INDIAN_FORMAT = auto()   # 1,23,456.78
    SWISS_FORMAT = auto()    # 1'234.56
    SPACE_FORMAT = auto()    # 1 234,56


class DateFormat(Enum):
    """Supported date format conventions"""
    MDY = auto()    # MM/DD/YYYY (US)
    DMY = auto()    # DD/MM/YYYY (European)
    YMD = auto()    # YYYY/MM/DD (ISO, Asian)
    DOTTED = auto() # DD.MM.YYYY (German)
    DASHED = auto() # DD-MM-YYYY (Various)


@dataclass
class InternationalConfig:
    """Configuration for international format detection"""
    preferred_number_format: NumberFormat = NumberFormat.US_FORMAT
    preferred_date_format: DateFormat = DateFormat.MDY
    decimal_separator: str = '.'
    thousands_separator: str = ','
    currency_symbols: List[str] = None
    date_separators: List[str] = None

    def __post_init__(self):
        if self.currency_symbols is None:
            self.currency_symbols = ['$', '€', '£', '¥', '₹', 'CHF', 'SEK', 'NOK', 'DKK']
        if self.date_separators is None:
            self.date_separators = ['/', '-', '.', ' ']


@dataclass
class CellParsingResult:
    """Result of parsing a cell value with international formatting"""
    original_value: Any
    parsed_value: Any
    value_type: str  # 'number', 'date', 'text', 'formula', 'merged'
    format_detected: str
    confidence: float
    errors: List[str]


class InternationalNumberParser:
    """Parser for international number formats"""

    def __init__(self, config: InternationalConfig):
        self.config = config

    def parse_number(self, value: str) -> Tuple[Optional[float], str, float]:
        """
        Parse a number string according to international formats

        Returns:
            Tuple of (parsed_number, detected_format, confidence)
        """
        if not isinstance(value, str):
            return None, "not_string", 0.0

        value = value.strip()
        if not value:
            return None, "empty", 0.0

        # Remove currency symbols
        cleaned_value = value
        for symbol in self.config.currency_symbols:
            cleaned_value = cleaned_value.replace(symbol, '')
        cleaned_value = cleaned_value.strip()

        # Try different number formats
        formats_to_try = [
            (NumberFormat.US_FORMAT, self._parse_us_format),
            (NumberFormat.EUROPEAN_FORMAT, self._parse_european_format),
            (NumberFormat.INDIAN_FORMAT, self._parse_indian_format),
            (NumberFormat.SWISS_FORMAT, self._parse_swiss_format),
            (NumberFormat.SPACE_FORMAT, self._parse_space_format),
        ]

        best_result = None
        best_confidence = 0.0
        best_format = "unknown"

        for format_type, parser_func in formats_to_try:
            try:
                result = parser_func(cleaned_value)
                if result is not None:
                    confidence = self._calculate_number_confidence(cleaned_value, format_type)
                    if confidence > best_confidence:
                        best_result = result
                        best_confidence = confidence
                        best_format = format_type.name.lower()
            except Exception as e:
                logger.debug(f"Error parsing {cleaned_value} with {format_type.name}: {e}")
                continue

        return best_result, best_format, best_confidence

    def _parse_us_format(self, value: str) -> Optional[float]:
        """Parse US format: 1,234.56"""
        # Remove spaces and check for valid US format
        value = value.replace(' ', '')
        if re.match(r'^-?\d{1,3}(,\d{3})*(\.\d+)?$', value):
            # Valid US format
            return float(value.replace(',', ''))
        elif re.match(r'^-?\d+(\.\d+)?$', value):
            # Simple decimal number
            return float(value)
        return None

    def _parse_european_format(self, value: str) -> Optional[float]:
        """Parse European format: 1.234,56"""
        # Remove spaces and check for valid European format
        value = value.replace(' ', '')
        if re.match(r'^-?\d{1,3}(\.\d{3})*(,\d+)?$', value):
            # Valid European format: thousands separator is dot, decimal is comma
            # Convert to US format
            if ',' in value:
                integer_part, decimal_part = value.rsplit(',', 1)
                integer_part = integer_part.replace('.', '')
                return float(f"{integer_part}.{decimal_part}")
            else:
                # No decimal part, check if dots are thousands separators
                # If the number has dots and matches European thousands pattern, remove dots
                if '.' in value and len(value.replace('.', '')) > 3:
                    return float(value.replace('.', ''))
                elif '.' not in value:
                    return float(value)
                else:
                    # Single dot case - might be decimal or thousands separator
                    # If it's a 3-digit group after dot, it's thousands separator
                    parts = value.split('.')
                    if len(parts) == 2 and len(parts[1]) == 3:
                        return float(value.replace('.', ''))
                    else:
                        # Treat as decimal (fallback)
                        return float(value)
        return None

    def _parse_indian_format(self, value: str) -> Optional[float]:
        """Parse Indian format: 1,23,456.78"""
        # Remove spaces
        value = value.replace(' ', '')
        if re.match(r'^-?\d{1,3}(,\d{2})*(,\d{3})?(\.\d+)?$', value):
            # Valid Indian format
            return float(value.replace(',', ''))
        return None

    def _parse_swiss_format(self, value: str) -> Optional[float]:
        """Parse Swiss format: 1'234.56"""
        # Remove spaces
        value = value.replace(' ', '')
        if re.match(r"^-?\d{1,3}('\d{3})*(\.\d+)?$", value):
            # Valid Swiss format
            return float(value.replace("'", ''))
        return None

    def _parse_space_format(self, value: str) -> Optional[float]:
        """Parse space format: 1 234,56"""
        if re.match(r'^-?\d{1,3}( \d{3})*(,\d+)?$', value):
            # Valid space format with comma decimal
            if ',' in value:
                integer_part, decimal_part = value.rsplit(',', 1)
                integer_part = integer_part.replace(' ', '')
                return float(f"{integer_part}.{decimal_part}")
            else:
                return float(value.replace(' ', ''))
        return None

    def _calculate_number_confidence(self, value: str, format_type: NumberFormat) -> float:
        """Calculate confidence score for number format detection"""
        confidence = 0.5  # Base confidence

        # Boost confidence based on format patterns
        if format_type == NumberFormat.US_FORMAT:
            if ',' in value and '.' in value:
                # Check if comma appears before dot (US format)
                comma_pos = value.rfind(',')
                dot_pos = value.rfind('.')
                if comma_pos < dot_pos:
                    confidence += 0.3
            elif ',' in value and '.' not in value:
                confidence += 0.2

        elif format_type == NumberFormat.EUROPEAN_FORMAT:
            if '.' in value and ',' in value:
                # Check if dot appears before comma (European format)
                comma_pos = value.rfind(',')
                dot_pos = value.rfind('.')
                if dot_pos < comma_pos:
                    confidence += 0.3
            elif '.' in value and ',' not in value:
                # For European format, dots without commas are likely thousands separators
                # if the number follows European thousands pattern
                if re.match(r'^\d{1,3}(\.\d{3})+$', value):
                    confidence += 0.4  # High confidence for clear thousands pattern
                elif len(value.replace('.', '')) > 3:
                    confidence += 0.2  # Some confidence for larger numbers
            elif ',' in value and '.' not in value:
                confidence += 0.2

        # Boost confidence for preferred format
        if format_type == self.config.preferred_number_format:
            confidence += 0.1

        # Boost confidence for larger numbers (more likely to have separators)
        digit_count = sum(c.isdigit() for c in value)
        if digit_count > 6:
            confidence += 0.1

        return min(confidence, 1.0)


class InternationalDateParser:
    """Parser for international date formats"""

    def __init__(self, config: InternationalConfig):
        self.config = config

    def parse_date(self, value: str) -> Tuple[Optional[datetime], str, float]:
        """
        Parse a date string according to international formats

        Returns:
            Tuple of (parsed_date, detected_format, confidence)
        """
        if not isinstance(value, str):
            return None, "not_string", 0.0

        value = value.strip()
        if not value:
            return None, "empty", 0.0

        # Try different date formats
        formats_to_try = [
            (DateFormat.MDY, ['%m/%d/%Y', '%m-%d-%Y', '%m.%d.%Y']),
            (DateFormat.DMY, ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y']),
            (DateFormat.YMD, ['%Y/%m/%d', '%Y-%m-%d', '%Y.%m.%d']),
        ]

        best_result = None
        best_confidence = 0.0
        best_format = "unknown"

        for format_type, format_strings in formats_to_try:
            for format_string in format_strings:
                try:
                    result = datetime.strptime(value, format_string)
                    confidence = self._calculate_date_confidence(value, format_type)
                    if confidence > best_confidence:
                        best_result = result
                        best_confidence = confidence
                        best_format = format_type.name.lower()
                except ValueError:
                    continue

        return best_result, best_format, best_confidence

    def _calculate_date_confidence(self, value: str, format_type: DateFormat) -> float:
        """Calculate confidence score for date format detection"""
        confidence = 0.6  # Base confidence for successful parsing

        # Boost confidence based on regional preferences
        if format_type == self.config.preferred_date_format:
            confidence += 0.2

        # Boost confidence for ISO format (most unambiguous)
        if format_type == DateFormat.YMD:
            confidence += 0.1

        return min(confidence, 1.0)


class MergedCellHandler:
    """Handler for merged cells in Excel worksheets"""

    def __init__(self):
        self.merged_ranges_cache = {}

    def get_merged_cell_value(self, worksheet: Worksheet, row: int, col: int) -> Tuple[Any, bool]:
        """
        Get value from a potentially merged cell

        Returns:
            Tuple of (cell_value, is_merged)
        """
        try:
            cell = worksheet.cell(row, col)

            # Check if this cell is part of a merged range
            for merged_range in worksheet.merged_cells.ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):

                    # Get value from the top-left cell of the merged range
                    top_left_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
                    return top_left_cell.value, True

            # Not a merged cell, return normal value
            return cell.value, False

        except Exception as e:
            logger.debug(f"Error accessing cell ({row}, {col}): {e}")
            return None, False

    def get_merged_range_info(self, worksheet: Worksheet) -> Dict[str, Any]:
        """Get information about all merged ranges in the worksheet"""
        merged_info = {
            'total_ranges': len(worksheet.merged_cells.ranges),
            'ranges': [],
            'affected_cells': 0
        }

        for merged_range in worksheet.merged_cells.ranges:
            range_info = {
                'range': str(merged_range),
                'top_left': (merged_range.min_row, merged_range.min_col),
                'bottom_right': (merged_range.max_row, merged_range.max_col),
                'size': (merged_range.max_row - merged_range.min_row + 1,
                        merged_range.max_col - merged_range.min_col + 1),
                'cell_count': ((merged_range.max_row - merged_range.min_row + 1) *
                              (merged_range.max_col - merged_range.min_col + 1))
            }
            merged_info['ranges'].append(range_info)
            merged_info['affected_cells'] += range_info['cell_count']

        return merged_info


class FormulaHandler:
    """Handler for formula cells and calculated values"""

    def __init__(self):
        self.formula_cache = {}

    def get_formula_info(self, cell: Cell) -> Dict[str, Any]:
        """Get information about a formula cell"""
        if not hasattr(cell, 'formula') or not cell.formula:
            return {'is_formula': False}

        formula_info = {
            'is_formula': True,
            'formula': cell.formula,
            'calculated_value': cell.value,
            'formula_type': self._classify_formula(cell.formula),
            'references': self._extract_cell_references(cell.formula),
            'functions_used': self._extract_functions(cell.formula)
        }

        return formula_info

    def _classify_formula(self, formula: str) -> str:
        """Classify the type of formula"""
        formula_lower = formula.lower()

        if any(func in formula_lower for func in ['sum', 'total']):
            return 'sum'
        elif any(func in formula_lower for func in ['average', 'mean']):
            return 'average'
        elif any(func in formula_lower for func in ['if', 'choose']):
            return 'conditional'
        elif any(func in formula_lower for func in ['vlookup', 'hlookup', 'lookup']):
            return 'lookup'
        elif any(op in formula for op in ['+', '-', '*', '/']):
            return 'arithmetic'
        else:
            return 'other'

    def _extract_cell_references(self, formula: str) -> List[str]:
        """Extract cell references from formula"""
        # Pattern to match cell references like A1, B2, $A$1, etc.
        pattern = r'\$?[A-Z]+\$?\d+'
        return re.findall(pattern, formula)

    def _extract_functions(self, formula: str) -> List[str]:
        """Extract function names from formula"""
        # Pattern to match function names (word followed by opening parenthesis)
        pattern = r'([A-Z][A-Z0-9]*)\s*\('
        return re.findall(pattern, formula.upper())


class InternationalFormatHandler:
    """Main handler for international Excel formats"""

    def __init__(self, config: InternationalConfig = None):
        self.config = config or InternationalConfig()
        self.number_parser = InternationalNumberParser(self.config)
        self.date_parser = InternationalDateParser(self.config)
        self.merged_cell_handler = MergedCellHandler()
        self.formula_handler = FormulaHandler()

    def parse_cell_value(self, worksheet: Worksheet, row: int, col: int) -> CellParsingResult:
        """
        Parse a cell value considering international formats and special cases

        Returns:
            CellParsingResult with parsed value and metadata
        """
        errors = []

        try:
            # Get cell value, handling merged cells
            cell_value, is_merged = self.merged_cell_handler.get_merged_cell_value(worksheet, row, col)
            cell = worksheet.cell(row, col)

            if is_merged:
                return CellParsingResult(
                    original_value=cell_value,
                    parsed_value=cell_value,
                    value_type='merged',
                    format_detected='merged_cell',
                    confidence=1.0,
                    errors=errors
                )

            # Check if it's a formula cell
            formula_info = self.formula_handler.get_formula_info(cell)
            if formula_info['is_formula']:
                return CellParsingResult(
                    original_value=cell_value,
                    parsed_value=formula_info['calculated_value'],
                    value_type='formula',
                    format_detected=f"formula_{formula_info['formula_type']}",
                    confidence=0.9,
                    errors=errors
                )

            # Handle None or empty values
            if cell_value is None or cell_value == "":
                return CellParsingResult(
                    original_value=cell_value,
                    parsed_value=None,
                    value_type='empty',
                    format_detected='empty',
                    confidence=1.0,
                    errors=errors
                )

            # Try parsing as number first
            if isinstance(cell_value, str):
                parsed_number, number_format, number_confidence = self.number_parser.parse_number(cell_value)
                if parsed_number is not None and number_confidence > 0.6:
                    return CellParsingResult(
                        original_value=cell_value,
                        parsed_value=parsed_number,
                        value_type='number',
                        format_detected=number_format,
                        confidence=number_confidence,
                        errors=errors
                    )

                # Try parsing as date
                parsed_date, date_format, date_confidence = self.date_parser.parse_date(cell_value)
                if parsed_date is not None and date_confidence > 0.6:
                    return CellParsingResult(
                        original_value=cell_value,
                        parsed_value=parsed_date,
                        value_type='date',
                        format_detected=date_format,
                        confidence=date_confidence,
                        errors=errors
                    )

            # Default to text
            return CellParsingResult(
                original_value=cell_value,
                parsed_value=cell_value,
                value_type='text',
                format_detected='text',
                confidence=0.8,
                errors=errors
            )

        except Exception as e:
            errors.append(f"Error parsing cell ({row}, {col}): {str(e)}")
            return CellParsingResult(
                original_value=None,
                parsed_value=None,
                value_type='error',
                format_detected='error',
                confidence=0.0,
                errors=errors
            )

    def analyze_worksheet_format(self, worksheet: Worksheet) -> Dict[str, Any]:
        """Analyze the international format characteristics of a worksheet"""
        analysis = {
            'number_formats': {},
            'date_formats': {},
            'merged_cells': self.merged_cell_handler.get_merged_range_info(worksheet),
            'formula_cells': {'count': 0, 'types': {}},
            'international_indicators': [],
            'recommended_config': None
        }

        # Sample cells to analyze formats
        max_row = min(worksheet.max_row or 50, 50)
        max_col = min(worksheet.max_column or 20, 20)

        sample_cells = []
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                if len(sample_cells) < 100:  # Limit sampling
                    sample_cells.append((row, col))

        # Analyze sample cells
        for row, col in sample_cells:
            result = self.parse_cell_value(worksheet, row, col)

            if result.value_type == 'number':
                format_key = result.format_detected
                analysis['number_formats'][format_key] = analysis['number_formats'].get(format_key, 0) + 1
            elif result.value_type == 'date':
                format_key = result.format_detected
                analysis['date_formats'][format_key] = analysis['date_formats'].get(format_key, 0) + 1
            elif result.value_type == 'formula':
                analysis['formula_cells']['count'] += 1
                formula_type = result.format_detected.replace('formula_', '')
                analysis['formula_cells']['types'][formula_type] = \
                    analysis['formula_cells']['types'].get(formula_type, 0) + 1

        # Detect international indicators
        if analysis['number_formats'].get('european_format', 0) > 0:
            analysis['international_indicators'].append('European number format detected')
        if analysis['date_formats'].get('dmy', 0) > 0:
            analysis['international_indicators'].append('European date format detected')
        if analysis['merged_cells']['total_ranges'] > 0:
            analysis['international_indicators'].append('Merged cells detected')
        if analysis['formula_cells']['count'] > 0:
            analysis['international_indicators'].append('Formula cells detected')

        # Generate recommended configuration
        analysis['recommended_config'] = self._generate_recommended_config(analysis)

        return analysis

    def _generate_recommended_config(self, analysis: Dict[str, Any]) -> InternationalConfig:
        """Generate recommended configuration based on analysis"""
        config = InternationalConfig()

        # Determine preferred number format
        number_formats = analysis['number_formats']
        if number_formats.get('european_format', 0) > number_formats.get('us_format', 0):
            config.preferred_number_format = NumberFormat.EUROPEAN_FORMAT
            config.decimal_separator = ','
            config.thousands_separator = '.'

        # Determine preferred date format
        date_formats = analysis['date_formats']
        if date_formats.get('dmy', 0) > date_formats.get('mdy', 0):
            config.preferred_date_format = DateFormat.DMY
        elif date_formats.get('ymd', 0) > date_formats.get('mdy', 0):
            config.preferred_date_format = DateFormat.YMD

        return config


def create_international_handler(
    number_format: NumberFormat = NumberFormat.US_FORMAT,
    date_format: DateFormat = DateFormat.MDY,
    **kwargs
) -> InternationalFormatHandler:
    """
    Convenience function to create an international format handler

    Args:
        number_format: Preferred number format
        date_format: Preferred date format
        **kwargs: Additional configuration options

    Returns:
        Configured InternationalFormatHandler
    """
    config = InternationalConfig(
        preferred_number_format=number_format,
        preferred_date_format=date_format,
        **kwargs
    )
    return InternationalFormatHandler(config)


# Convenience functions for common regional configurations
def create_european_handler() -> InternationalFormatHandler:
    """Create handler optimized for European Excel formats"""
    return create_international_handler(
        number_format=NumberFormat.EUROPEAN_FORMAT,
        date_format=DateFormat.DMY,
        decimal_separator=',',
        thousands_separator='.',
        currency_symbols=['€', '£', 'CHF', 'SEK', 'NOK', 'DKK']
    )


def create_asian_handler() -> InternationalFormatHandler:
    """Create handler optimized for Asian Excel formats"""
    return create_international_handler(
        number_format=NumberFormat.US_FORMAT,
        date_format=DateFormat.YMD,
        currency_symbols=['¥', '₹', '₩', 'CNY', 'JPY', 'KRW', 'INR']
    )


def create_us_handler() -> InternationalFormatHandler:
    """Create handler optimized for US Excel formats"""
    return create_international_handler(
        number_format=NumberFormat.US_FORMAT,
        date_format=DateFormat.MDY,
        currency_symbols=['$', 'USD']
    )