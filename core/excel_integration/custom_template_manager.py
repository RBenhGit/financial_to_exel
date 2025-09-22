"""
Custom Template Manager for Excel Data Processing
================================================

This module provides a comprehensive system for managing custom Excel templates,
allowing users to define their own layouts and structures for financial statements
that don't conform to standard formats.

Key Features:
- Template definition and validation
- Template library management
- Dynamic template detection
- Template configuration persistence
- Template sharing and import/export
"""

import json
import logging
import os
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Union
from enum import Enum, auto
from openpyxl.worksheet.worksheet import Worksheet

from .format_detector import FormatSignature, FormatType
from .international_format_handler import InternationalFormatHandler, InternationalConfig

logger = logging.getLogger(__name__)


class TemplateType(Enum):
    """Types of custom templates"""
    USER_DEFINED = auto()
    IMPORTED = auto()
    SYSTEM_GENERATED = auto()
    SHARED = auto()


@dataclass
class CellMapping:
    """Defines a mapping for a specific cell or range"""
    name: str
    row: int
    column: int
    data_type: str = 'text'  # 'text', 'number', 'date', 'formula'
    required: bool = True
    validation_pattern: Optional[str] = None
    description: Optional[str] = None


@dataclass
class ColumnMapping:
    """Defines a mapping for a data column"""
    name: str
    column_index: int
    data_type: str = 'number'
    format_pattern: Optional[str] = None
    required: bool = True
    description: Optional[str] = None


@dataclass
class SectionMapping:
    """Defines a section of the Excel template"""
    name: str
    start_row: int
    end_row: int
    start_column: int
    end_column: int
    section_type: str = 'data'  # 'header', 'data', 'footer', 'metadata'
    cell_mappings: List[CellMapping] = field(default_factory=list)
    column_mappings: List[ColumnMapping] = field(default_factory=list)


@dataclass
class CustomTemplate:
    """Defines a complete custom Excel template"""
    template_id: str
    name: str
    description: str
    template_type: TemplateType
    version: str = "1.0"

    # Template structure
    company_name_cell: Tuple[int, int] = (2, 3)
    period_header_row: int = 8
    data_start_row: int = 12
    data_start_column: int = 4

    # Sections
    sections: List[SectionMapping] = field(default_factory=list)

    # International format configuration
    international_config: Optional[InternationalConfig] = None

    # Validation rules
    required_sheets: List[str] = field(default_factory=list)
    expected_patterns: List[Tuple[int, int, str]] = field(default_factory=list)

    # Metadata
    created_by: str = ""
    created_date: str = ""
    last_modified: str = ""
    tags: List[str] = field(default_factory=list)


class CustomTemplateManager:
    """Manager for custom Excel templates"""

    def __init__(self, templates_directory: Optional[str] = None):
        """
        Initialize the custom template manager

        Args:
            templates_directory: Directory to store template definitions
        """
        if templates_directory is None:
            # Default to application data directory
            templates_directory = os.path.join(
                os.path.expanduser("~"),
                ".financial_analysis",
                "excel_templates"
            )

        self.templates_directory = Path(templates_directory)
        self.templates_directory.mkdir(parents=True, exist_ok=True)

        self.templates: Dict[str, CustomTemplate] = {}
        self.load_templates()

    def create_template(
        self,
        template_id: str,
        name: str,
        description: str,
        template_type: TemplateType = TemplateType.USER_DEFINED,
        **kwargs
    ) -> CustomTemplate:
        """
        Create a new custom template

        Args:
            template_id: Unique identifier for the template
            name: Human-readable name
            description: Template description
            template_type: Type of template
            **kwargs: Additional template properties

        Returns:
            Created CustomTemplate object
        """
        if template_id in self.templates:
            raise ValueError(f"Template with ID '{template_id}' already exists")

        template = CustomTemplate(
            template_id=template_id,
            name=name,
            description=description,
            template_type=template_type,
            **kwargs
        )

        # Set metadata
        from datetime import datetime
        template.created_date = datetime.now().isoformat()
        template.last_modified = template.created_date

        self.templates[template_id] = template
        self.save_template(template)

        logger.info(f"Created custom template: {name} ({template_id})")
        return template

    def load_templates(self):
        """Load all templates from the templates directory"""
        template_files = list(self.templates_directory.glob("*.json"))

        for template_file in template_files:
            try:
                template = self.load_template_from_file(template_file)
                if template:
                    self.templates[template.template_id] = template
                    logger.debug(f"Loaded template: {template.name}")
            except Exception as e:
                logger.error(f"Failed to load template from {template_file}: {e}")

    def load_template_from_file(self, file_path: Union[str, Path]) -> Optional[CustomTemplate]:
        """Load a template from a JSON file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # Convert data to CustomTemplate
            template = self._dict_to_template(data)
            return template

        except Exception as e:
            logger.error(f"Error loading template from {file_path}: {e}")
            return None

    def save_template(self, template: CustomTemplate):
        """Save a template to file"""
        file_path = self.templates_directory / f"{template.template_id}.json"

        try:
            # Update last modified timestamp
            from datetime import datetime
            template.last_modified = datetime.now().isoformat()

            # Convert template to dictionary
            template_dict = self._template_to_dict(template)

            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(template_dict, f, indent=2, ensure_ascii=False)

            logger.debug(f"Saved template: {template.name} to {file_path}")

        except Exception as e:
            logger.error(f"Error saving template {template.template_id}: {e}")
            raise

    def get_template(self, template_id: str) -> Optional[CustomTemplate]:
        """Get a template by ID"""
        return self.templates.get(template_id)

    def list_templates(self) -> List[CustomTemplate]:
        """Get a list of all available templates"""
        return list(self.templates.values())

    def delete_template(self, template_id: str) -> bool:
        """Delete a template"""
        if template_id not in self.templates:
            return False

        # Remove from memory
        del self.templates[template_id]

        # Remove file
        file_path = self.templates_directory / f"{template_id}.json"
        try:
            if file_path.exists():
                file_path.unlink()
            logger.info(f"Deleted template: {template_id}")
            return True
        except Exception as e:
            logger.error(f"Error deleting template file {file_path}: {e}")
            return False

    def detect_template_for_worksheet(self, worksheet: Worksheet) -> Optional[CustomTemplate]:
        """
        Detect which custom template best matches a worksheet

        Args:
            worksheet: Excel worksheet to analyze

        Returns:
            Best matching CustomTemplate or None
        """
        best_template = None
        best_score = 0.0

        for template in self.templates.values():
            score = self._calculate_template_match_score(worksheet, template)
            if score > best_score and score > 0.6:  # Minimum confidence threshold
                best_score = score
                best_template = template

        if best_template:
            logger.info(f"Detected template: {best_template.name} (score: {best_score:.2f})")

        return best_template

    def _calculate_template_match_score(self, worksheet: Worksheet, template: CustomTemplate) -> float:
        """Calculate how well a worksheet matches a template"""
        score = 0.0
        total_checks = 0

        # Check company name position
        total_checks += 1
        try:
            company_row, company_col = template.company_name_cell
            company_value = worksheet.cell(company_row, company_col).value
            if company_value and isinstance(company_value, str) and len(company_value.strip()) > 0:
                score += 1.0
        except Exception:
            pass

        # Check expected patterns
        for row, col, pattern in template.expected_patterns:
            total_checks += 1
            try:
                cell_value = worksheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    import re
                    if re.search(pattern, cell_value, re.IGNORECASE):
                        score += 1.0
            except Exception:
                pass

        # Check section structure
        for section in template.sections:
            total_checks += 1
            section_score = self._check_section_match(worksheet, section)
            score += section_score

        # Calculate final score
        if total_checks > 0:
            return score / total_checks
        return 0.0

    def _check_section_match(self, worksheet: Worksheet, section: SectionMapping) -> float:
        """Check how well a section matches the worksheet"""
        matches = 0
        total_checks = 0

        # Check cell mappings
        for cell_mapping in section.cell_mappings:
            total_checks += 1
            try:
                cell_value = worksheet.cell(cell_mapping.row, cell_mapping.column).value
                if cell_value is not None:
                    matches += 1
            except Exception:
                pass

        if total_checks > 0:
            return matches / total_checks
        return 0.0

    def generate_template_from_worksheet(
        self,
        worksheet: Worksheet,
        template_id: str,
        name: str,
        description: str
    ) -> CustomTemplate:
        """
        Generate a custom template by analyzing a worksheet structure

        Args:
            worksheet: Excel worksheet to analyze
            template_id: ID for the new template
            name: Name for the template
            description: Description of the template

        Returns:
            Generated CustomTemplate
        """
        logger.info(f"Generating template from worksheet: {worksheet.title}")

        # Use international format handler to analyze the worksheet
        handler = InternationalFormatHandler()
        analysis = handler.analyze_worksheet_format(worksheet)

        # Detect company name position
        company_name_cell = self._detect_company_name_position(worksheet)

        # Detect data structure
        data_start_row, data_start_column = self._detect_data_start_position(worksheet)
        period_header_row = self._detect_period_header_row(worksheet)

        # Create sections based on analysis
        sections = self._generate_sections_from_analysis(worksheet, analysis)

        # Generate expected patterns
        expected_patterns = self._generate_expected_patterns(worksheet)

        # Create the template
        template = CustomTemplate(
            template_id=template_id,
            name=name,
            description=description,
            template_type=TemplateType.SYSTEM_GENERATED,
            company_name_cell=company_name_cell,
            period_header_row=period_header_row,
            data_start_row=data_start_row,
            data_start_column=data_start_column,
            sections=sections,
            international_config=analysis['recommended_config'],
            expected_patterns=expected_patterns,
            tags=['auto-generated', 'from-worksheet']
        )

        # Save the template
        self.templates[template_id] = template
        self.save_template(template)

        logger.info(f"Generated template: {name} with {len(sections)} sections")
        return template

    def _detect_company_name_position(self, worksheet: Worksheet) -> Tuple[int, int]:
        """Detect the position of the company name"""
        # Common positions to check
        positions_to_check = [(2, 3), (1, 3), (3, 3), (1, 1), (2, 2)]

        for row, col in positions_to_check:
            try:
                cell_value = worksheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 2:
                    # Simple heuristic: if it's not a number and not a common header term
                    value_lower = cell_value.lower().strip()
                    if not any(term in value_lower for term in
                              ['period', 'date', 'year', 'quarter', 'month', 'statement']):
                        return (row, col)
            except Exception:
                continue

        return (2, 3)  # Default fallback

    def _detect_data_start_position(self, worksheet: Worksheet) -> Tuple[int, int]:
        """Detect where financial data starts"""
        financial_keywords = [
            'revenue', 'sales', 'income', 'net sales', 'total revenue',
            'gross profit', 'operating income', 'ebitda', 'assets', 'liabilities'
        ]

        max_row = min(worksheet.max_row or 50, 50)
        max_col = min(worksheet.max_column or 20, 20)

        for row in range(1, max_row + 1):
            for col in range(1, min(10, max_col + 1)):
                try:
                    cell_value = worksheet.cell(row, col).value
                    if cell_value and isinstance(cell_value, str):
                        value_lower = cell_value.lower().strip()
                        for keyword in financial_keywords:
                            if keyword in value_lower:
                                return (row, col + 1)  # Data typically starts in next column
                except Exception:
                    continue

        return (12, 4)  # Default fallback

    def _detect_period_header_row(self, worksheet: Worksheet) -> int:
        """Detect the row containing period headers"""
        max_row = min(worksheet.max_row or 30, 30)
        max_col = min(worksheet.max_column or 20, 20)

        for row in range(1, max_row + 1):
            period_indicators = 0
            for col in range(1, max_col + 1):
                try:
                    cell_value = worksheet.cell(row, col).value
                    if cell_value and isinstance(cell_value, str):
                        value = cell_value.strip()
                        # Look for year patterns or period indicators
                        if (value.startswith(('20', 'FY-', 'Q', 'Year')) or
                            'period' in value.lower() or
                            len(value) == 4 and value.isdigit()):
                            period_indicators += 1
                except Exception:
                    continue

            if period_indicators >= 2:  # At least 2 period indicators in the row
                return row

        return 8  # Default fallback

    def _generate_sections_from_analysis(self, worksheet: Worksheet, analysis: Dict[str, Any]) -> List[SectionMapping]:
        """Generate section mappings based on worksheet analysis"""
        sections = []

        # Header section (typically first 10 rows)
        header_section = SectionMapping(
            name="Header",
            start_row=1,
            end_row=10,
            start_column=1,
            end_column=10,
            section_type="header"
        )
        sections.append(header_section)

        # Data section (main financial data)
        data_start_row, data_start_col = self._detect_data_start_position(worksheet)
        max_row = worksheet.max_row or 50
        max_col = worksheet.max_column or 15

        data_section = SectionMapping(
            name="Financial Data",
            start_row=data_start_row,
            end_row=min(max_row, data_start_row + 40),
            start_column=data_start_col,
            end_column=min(max_col, data_start_col + 10),
            section_type="data"
        )
        sections.append(data_section)

        return sections

    def _generate_expected_patterns(self, worksheet: Worksheet) -> List[Tuple[int, int, str]]:
        """Generate expected patterns for template validation"""
        patterns = []

        # Look for common financial statement indicators
        max_row = min(worksheet.max_row or 30, 30)
        max_col = min(worksheet.max_column or 15, 15)

        indicator_terms = [
            'statement', 'period', 'revenue', 'income', 'assets',
            'liabilities', 'cash flow', 'balance sheet'
        ]

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                try:
                    cell_value = worksheet.cell(row, col).value
                    if cell_value and isinstance(cell_value, str):
                        value_lower = cell_value.lower().strip()
                        for term in indicator_terms:
                            if term in value_lower:
                                patterns.append((row, col, f".*{term}.*"))
                                break
                except Exception:
                    continue

        return patterns[:10]  # Limit to top 10 patterns

    def export_template(self, template_id: str, export_path: str) -> bool:
        """Export a template to a file"""
        template = self.get_template(template_id)
        if not template:
            return False

        try:
            template_dict = self._template_to_dict(template)
            with open(export_path, 'w', encoding='utf-8') as f:
                json.dump(template_dict, f, indent=2, ensure_ascii=False)

            logger.info(f"Exported template {template_id} to {export_path}")
            return True
        except Exception as e:
            logger.error(f"Error exporting template {template_id}: {e}")
            return False

    def import_template(self, import_path: str) -> Optional[CustomTemplate]:
        """Import a template from a file"""
        try:
            template = self.load_template_from_file(import_path)
            if template:
                # Mark as imported
                template.template_type = TemplateType.IMPORTED

                # Ensure unique ID
                original_id = template.template_id
                counter = 1
                while template.template_id in self.templates:
                    template.template_id = f"{original_id}_{counter}"
                    counter += 1

                self.templates[template.template_id] = template
                self.save_template(template)

                logger.info(f"Imported template: {template.name} ({template.template_id})")
                return template

        except Exception as e:
            logger.error(f"Error importing template from {import_path}: {e}")

        return None

    def _template_to_dict(self, template: CustomTemplate) -> Dict[str, Any]:
        """Convert CustomTemplate to dictionary for serialization"""
        result = asdict(template)

        # Handle enum values
        result['template_type'] = template.template_type.name

        return result

    def _dict_to_template(self, data: Dict[str, Any]) -> CustomTemplate:
        """Convert dictionary to CustomTemplate"""
        # Handle enum values
        if 'template_type' in data and isinstance(data['template_type'], str):
            data['template_type'] = TemplateType[data['template_type']]

        # Convert sections
        if 'sections' in data:
            sections = []
            for section_data in data['sections']:
                # Convert cell mappings
                cell_mappings = []
                for cm_data in section_data.get('cell_mappings', []):
                    cell_mappings.append(CellMapping(**cm_data))

                # Convert column mappings
                column_mappings = []
                for col_data in section_data.get('column_mappings', []):
                    column_mappings.append(ColumnMapping(**col_data))

                section_data['cell_mappings'] = cell_mappings
                section_data['column_mappings'] = column_mappings
                sections.append(SectionMapping(**section_data))

            data['sections'] = sections

        # Convert international config
        if 'international_config' in data and data['international_config']:
            data['international_config'] = InternationalConfig(**data['international_config'])

        return CustomTemplate(**data)

    def create_format_signature_from_template(self, template: CustomTemplate) -> FormatSignature:
        """Convert a CustomTemplate to a FormatSignature for format detection"""
        return FormatSignature(
            name=template.name,
            format_type=FormatType.CUSTOM_TEMPLATE,
            identifier_patterns=template.expected_patterns,
            company_name_positions=[template.company_name_cell],
            period_header_positions=[(template.period_header_row, col) for col in range(2, 8)],
            data_start_indicators=[(template.data_start_row, template.data_start_column, ".*")],
            expected_max_rows=200,
            expected_max_columns=50,
            data_columns_range=(template.data_start_column, template.data_start_column + 15),
            required_sections=['company_name'] if template.company_name_cell else []
        )

    def update_template(self, template_id: str, **updates) -> bool:
        """Update an existing template"""
        template = self.get_template(template_id)
        if not template:
            return False

        try:
            # Update fields
            for field, value in updates.items():
                if hasattr(template, field):
                    setattr(template, field, value)

            # Update last modified timestamp
            from datetime import datetime
            template.last_modified = datetime.now().isoformat()

            # Save updated template
            self.save_template(template)

            logger.info(f"Updated template: {template_id}")
            return True

        except Exception as e:
            logger.error(f"Error updating template {template_id}: {e}")
            return False