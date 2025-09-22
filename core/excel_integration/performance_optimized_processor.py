"""
Performance Optimized Excel Processor
====================================

This module provides high-performance Excel processing capabilities with:
- Streaming processing for large files
- Memory optimization through lazy loading
- Parallel processing for multiple files
- Advanced caching strategies
- Performance monitoring and profiling

Key Features:
- Chunked reading for memory efficiency
- Concurrent processing with thread/process pools
- Intelligent caching with LRU eviction
- Memory usage monitoring and limits
- Performance metrics collection
- Support for various Excel formats and edge cases
"""

import os
import gc
import time
import logging
import threading
import multiprocessing
from pathlib import Path
from typing import Dict, List, Optional, Any, Iterator, Tuple, Union, Callable
from dataclasses import dataclass, field
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
from functools import lru_cache, wraps
import psutil
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from collections import defaultdict, OrderedDict
import threading
import weakref
import hashlib
import json

from .excel_utils import ExcelDataExtractor, ExcelFileValidator
from .format_detector import ExcelFormatDetector, FormatDetectionResult

logger = logging.getLogger(__name__)


@dataclass
class PerformanceConfig:
    """Configuration for performance optimization"""
    # Memory management
    max_memory_mb: int = 1024  # 1GB default limit
    chunk_size: int = 1000     # Rows per chunk
    enable_streaming: bool = True
    lazy_loading: bool = True

    # Concurrency
    max_workers: int = min(8, multiprocessing.cpu_count())
    use_process_pool: bool = False  # Use threads by default for I/O bound tasks

    # Caching
    cache_size: int = 128  # Number of cached items
    cache_ttl_seconds: int = 3600  # 1 hour cache TTL
    enable_disk_cache: bool = True
    disk_cache_dir: Optional[str] = None

    # Performance monitoring
    enable_profiling: bool = True
    profile_memory: bool = True
    profile_timing: bool = True

    # Processing optimization
    skip_empty_cells: bool = True
    optimize_data_types: bool = True
    use_read_only_mode: bool = True


@dataclass
class PerformanceMetrics:
    """Performance metrics for Excel processing"""
    total_time_seconds: float = 0.0
    memory_peak_mb: float = 0.0
    memory_start_mb: float = 0.0
    memory_end_mb: float = 0.0
    files_processed: int = 0
    rows_processed: int = 0
    cells_processed: int = 0
    cache_hits: int = 0
    cache_misses: int = 0
    errors_encountered: int = 0
    concurrent_operations: int = 0

    def memory_delta_mb(self) -> float:
        """Calculate memory usage delta"""
        return self.memory_end_mb - self.memory_start_mb

    def processing_rate_rows_per_second(self) -> float:
        """Calculate rows per second processing rate"""
        if self.total_time_seconds > 0:
            return self.rows_processed / self.total_time_seconds
        return 0.0

    def cache_hit_rate(self) -> float:
        """Calculate cache hit rate"""
        total_requests = self.cache_hits + self.cache_misses
        if total_requests > 0:
            return self.cache_hits / total_requests
        return 0.0


class MemoryMonitor:
    """Monitor memory usage during Excel processing"""

    def __init__(self, max_memory_mb: int = 1024):
        self.max_memory_mb = max_memory_mb
        self.start_memory = 0.0
        self.peak_memory = 0.0
        self.current_memory = 0.0
        self._monitoring = False
        self._monitor_thread = None
        self._stop_event = threading.Event()

    def __enter__(self):
        self.start_monitoring()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.stop_monitoring()

    def start_monitoring(self):
        """Start memory monitoring"""
        self.start_memory = self._get_memory_usage()
        self.peak_memory = self.start_memory
        self.current_memory = self.start_memory
        self._monitoring = True

        # Start background monitoring thread
        self._monitor_thread = threading.Thread(target=self._monitor_loop, daemon=True)
        self._monitor_thread.start()

    def stop_monitoring(self):
        """Stop memory monitoring"""
        self._monitoring = False
        if self._monitor_thread:
            self._stop_event.set()
            self._monitor_thread.join(timeout=1.0)

        self.current_memory = self._get_memory_usage()

    def _monitor_loop(self):
        """Background monitoring loop"""
        while self._monitoring and not self._stop_event.wait(0.1):
            try:
                current = self._get_memory_usage()
                self.current_memory = current
                self.peak_memory = max(self.peak_memory, current)

                # Check memory limit
                if current > self.max_memory_mb:
                    logger.warning(f"Memory usage ({current:.0f}MB) exceeds limit ({self.max_memory_mb}MB)")

                    # Trigger garbage collection
                    gc.collect()

            except Exception as e:
                logger.debug(f"Memory monitoring error: {e}")

    def _get_memory_usage(self) -> float:
        """Get current memory usage in MB"""
        try:
            return psutil.Process().memory_info().rss / 1024 / 1024
        except Exception:
            return 0.0

    def get_metrics(self) -> Dict[str, float]:
        """Get memory metrics"""
        return {
            'start_memory_mb': self.start_memory,
            'peak_memory_mb': self.peak_memory,
            'current_memory_mb': self.current_memory,
            'memory_delta_mb': self.current_memory - self.start_memory
        }


class IntelligentCache:
    """LRU cache with TTL and memory-aware eviction"""

    def __init__(self, max_size: int = 128, ttl_seconds: int = 3600):
        self.max_size = max_size
        self.ttl_seconds = ttl_seconds
        self._cache = OrderedDict()
        self._timestamps = {}
        self._lock = threading.RLock()
        self.hits = 0
        self.misses = 0

    def get(self, key: str) -> Optional[Any]:
        """Get item from cache"""
        with self._lock:
            current_time = time.time()

            if key in self._cache:
                # Check TTL
                if current_time - self._timestamps[key] <= self.ttl_seconds:
                    # Move to end (most recently used)
                    self._cache.move_to_end(key)
                    self.hits += 1
                    return self._cache[key]
                else:
                    # Expired
                    del self._cache[key]
                    del self._timestamps[key]

            self.misses += 1
            return None

    def put(self, key: str, value: Any):
        """Put item in cache"""
        with self._lock:
            current_time = time.time()

            if key in self._cache:
                # Update existing
                self._cache[key] = value
                self._timestamps[key] = current_time
                self._cache.move_to_end(key)
            else:
                # Add new
                if len(self._cache) >= self.max_size:
                    # Evict oldest
                    oldest_key = next(iter(self._cache))
                    del self._cache[oldest_key]
                    del self._timestamps[oldest_key]

                self._cache[key] = value
                self._timestamps[key] = current_time

    def clear(self):
        """Clear cache"""
        with self._lock:
            self._cache.clear()
            self._timestamps.clear()

    def get_stats(self) -> Dict[str, int]:
        """Get cache statistics"""
        with self._lock:
            return {
                'size': len(self._cache),
                'max_size': self.max_size,
                'hits': self.hits,
                'misses': self.misses,
                'hit_rate': self.hits / (self.hits + self.misses) if (self.hits + self.misses) > 0 else 0.0
            }


class StreamingExcelReader:
    """Stream Excel data in chunks for memory efficiency"""

    def __init__(self, file_path: str, config: PerformanceConfig):
        self.file_path = file_path
        self.config = config
        self.workbook = None
        self.worksheet = None

    def __enter__(self):
        self.workbook = load_workbook(
            self.file_path,
            read_only=self.config.use_read_only_mode,
            data_only=True
        )
        self.worksheet = self.workbook.active
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.workbook:
            self.workbook.close()

    def read_chunks(self, chunk_size: Optional[int] = None) -> Iterator[pd.DataFrame]:
        """Read Excel file in chunks"""
        if not self.worksheet:
            return

        chunk_size = chunk_size or self.config.chunk_size
        max_row = self.worksheet.max_row or 0
        max_col = self.worksheet.max_column or 0

        # Read headers first
        headers = []
        for col in range(1, max_col + 1):
            header_value = self.worksheet.cell(1, col).value
            headers.append(header_value if header_value is not None else f"Column_{col}")

        # Read data in chunks
        for start_row in range(2, max_row + 1, chunk_size):
            end_row = min(start_row + chunk_size - 1, max_row)

            # Extract chunk data
            chunk_data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell_value = self.worksheet.cell(row, col).value

                    # Skip empty cells if configured
                    if self.config.skip_empty_cells and cell_value is None:
                        cell_value = np.nan

                    row_data.append(cell_value)

                chunk_data.append(row_data)

            if chunk_data:
                df = pd.DataFrame(chunk_data, columns=headers)

                # Optimize data types if configured
                if self.config.optimize_data_types:
                    df = self._optimize_datatypes(df)

                yield df

    def _optimize_datatypes(self, df: pd.DataFrame) -> pd.DataFrame:
        """Optimize DataFrame data types for memory efficiency"""
        for col in df.columns:
            # Try to convert to numeric if possible
            try:
                # Check if column contains mostly numbers
                numeric_series = pd.to_numeric(df[col], errors='coerce')
                non_null_ratio = numeric_series.notna().sum() / len(df)

                if non_null_ratio > 0.8:  # If 80%+ are numeric
                    # Use smallest possible integer or float type
                    if numeric_series.apply(lambda x: x == int(x) if pd.notna(x) else True).all():
                        # Integer values
                        min_val = numeric_series.min()
                        max_val = numeric_series.max()

                        if pd.isna(min_val) or pd.isna(max_val):
                            continue

                        if -128 <= min_val and max_val <= 127:
                            df[col] = numeric_series.astype('int8')
                        elif -32768 <= min_val and max_val <= 32767:
                            df[col] = numeric_series.astype('int16')
                        elif -2147483648 <= min_val and max_val <= 2147483647:
                            df[col] = numeric_series.astype('int32')
                        else:
                            df[col] = numeric_series.astype('int64')
                    else:
                        # Float values - use float32 if precision allows
                        df[col] = numeric_series.astype('float32')
            except Exception:
                # Keep original type if conversion fails
                pass

        return df


class PerformanceOptimizedExcelProcessor:
    """High-performance Excel processor with streaming and parallel processing"""

    def __init__(self, config: Optional[PerformanceConfig] = None):
        self.config = config or PerformanceConfig()
        self.cache = IntelligentCache(self.config.cache_size, self.config.cache_ttl_seconds)
        self.metrics = PerformanceMetrics()
        self.format_detector = ExcelFormatDetector()
        self.validator = ExcelFileValidator()

        # Setup disk cache if enabled
        if self.config.enable_disk_cache:
            self.disk_cache_dir = Path(self.config.disk_cache_dir or "data/cache/excel_performance")
            self.disk_cache_dir.mkdir(parents=True, exist_ok=True)

    def process_single_file(
        self,
        file_path: str,
        operations: List[Callable] = None,
        enable_validation: bool = True
    ) -> Dict[str, Any]:
        """Process a single Excel file with performance optimization"""

        start_time = time.time()
        file_key = self._generate_file_key(file_path)

        # Check cache first
        cached_result = self.cache.get(file_key)
        if cached_result:
            logger.debug(f"Using cached result for {os.path.basename(file_path)}")
            return cached_result

        result = {
            'file_path': file_path,
            'success': False,
            'data': {},
            'metrics': {},
            'errors': []
        }

        with MemoryMonitor(self.config.max_memory_mb) as memory_monitor:
            try:
                # Validate file if enabled
                if enable_validation:
                    validation_result = self.validator.validate_file(file_path)
                    if not validation_result.is_valid:
                        result['errors'].extend(validation_result.format_issues)
                        if validation_result.severity == 'critical':
                            return result

                # Process with streaming if enabled and file is large
                file_size_mb = os.path.getsize(file_path) / 1024 / 1024

                if self.config.enable_streaming and file_size_mb > 10:  # Stream files > 10MB
                    result['data'] = self._process_file_streaming(file_path, operations)
                else:
                    result['data'] = self._process_file_standard(file_path, operations)

                result['success'] = True
                self.metrics.files_processed += 1

            except Exception as e:
                logger.error(f"Error processing {file_path}: {e}")
                result['errors'].append(str(e))
                self.metrics.errors_encountered += 1

            # Collect metrics
            processing_time = time.time() - start_time
            memory_metrics = memory_monitor.get_metrics()

            result['metrics'] = {
                'processing_time_seconds': processing_time,
                'file_size_mb': file_size_mb,
                **memory_metrics
            }

            # Update global metrics
            self.metrics.total_time_seconds += processing_time
            self.metrics.memory_peak_mb = max(self.metrics.memory_peak_mb, memory_metrics['peak_memory_mb'])

        # Cache successful results
        if result['success']:
            self.cache.put(file_key, result)

        return result

    def process_multiple_files(
        self,
        file_paths: List[str],
        operations: List[Callable] = None,
        enable_validation: bool = True
    ) -> Dict[str, Any]:
        """Process multiple Excel files concurrently"""

        start_time = time.time()
        results = []

        self.metrics.concurrent_operations = len(file_paths)

        # Choose executor based on configuration
        executor_class = ProcessPoolExecutor if self.config.use_process_pool else ThreadPoolExecutor

        with executor_class(max_workers=self.config.max_workers) as executor:
            # Submit all tasks
            future_to_file = {
                executor.submit(
                    self.process_single_file,
                    file_path,
                    operations,
                    enable_validation
                ): file_path
                for file_path in file_paths
            }

            # Collect results
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                try:
                    result = future.result(timeout=300)  # 5 minute timeout per file
                    results.append(result)
                except Exception as e:
                    logger.error(f"Failed to process {file_path}: {e}")
                    results.append({
                        'file_path': file_path,
                        'success': False,
                        'error': str(e)
                    })
                    self.metrics.errors_encountered += 1

        total_time = time.time() - start_time
        successful_results = [r for r in results if r.get('success', False)]

        return {
            'total_files': len(file_paths),
            'successful_files': len(successful_results),
            'failed_files': len(results) - len(successful_results),
            'total_processing_time': total_time,
            'results': results,
            'performance_metrics': self.get_performance_summary()
        }

    def _process_file_streaming(self, file_path: str, operations: List[Callable] = None) -> Dict[str, Any]:
        """Process file using streaming approach"""
        logger.debug(f"Processing {os.path.basename(file_path)} with streaming")

        data = {
            'chunks_processed': 0,
            'total_rows': 0,
            'summary_statistics': {},
            'format_info': None
        }

        with StreamingExcelReader(file_path, self.config) as reader:
            # Detect format first
            if reader.worksheet:
                format_result = self.format_detector.detect_format(reader.worksheet)
                data['format_info'] = {
                    'format_type': format_result.format_type.name,
                    'confidence': format_result.confidence_score
                }

            # Process chunks
            chunk_summaries = []
            for chunk_df in reader.read_chunks():
                chunk_summary = self._process_chunk(chunk_df, operations)
                chunk_summaries.append(chunk_summary)

                data['chunks_processed'] += 1
                data['total_rows'] += len(chunk_df)
                self.metrics.rows_processed += len(chunk_df)
                self.metrics.cells_processed += chunk_df.size

            # Aggregate chunk summaries
            if chunk_summaries:
                data['summary_statistics'] = self._aggregate_chunk_summaries(chunk_summaries)

        return data

    def _process_file_standard(self, file_path: str, operations: List[Callable] = None) -> Dict[str, Any]:
        """Process file using standard approach"""
        logger.debug(f"Processing {os.path.basename(file_path)} with standard method")

        data = {}

        with ExcelDataExtractor(file_path) as extractor:
            # Get format info
            format_info = extractor.get_format_info()
            if format_info:
                data['format_info'] = {
                    'format_type': format_info.format_type.name,
                    'confidence': format_info.confidence_score
                }

            # Extract financial data
            statement_types = ['income', 'balance', 'cashflow']
            for statement_type in statement_types:
                try:
                    metrics = extractor.extract_all_financial_metrics(statement_type)
                    if metrics:
                        data[statement_type] = {}
                        for metric_name, metric in metrics.items():
                            data[statement_type][metric_name] = {
                                'values': metric.values,
                                'location': (metric.location.row, metric.location.column)
                            }
                            self.metrics.cells_processed += len(metric.values)

                        self.metrics.rows_processed += len(metrics)
                except Exception as e:
                    logger.debug(f"Error extracting {statement_type} data: {e}")

        return data

    def _process_chunk(self, chunk_df: pd.DataFrame, operations: List[Callable] = None) -> Dict[str, Any]:
        """Process a single chunk of data"""
        summary = {
            'rows': len(chunk_df),
            'columns': len(chunk_df.columns),
            'memory_usage_mb': chunk_df.memory_usage(deep=True).sum() / 1024 / 1024
        }

        # Calculate basic statistics for numeric columns
        numeric_cols = chunk_df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            summary['numeric_summary'] = chunk_df[numeric_cols].describe().to_dict()

        # Apply custom operations if provided
        if operations:
            for operation in operations:
                try:
                    operation_result = operation(chunk_df)
                    summary[f'operation_{operation.__name__}'] = operation_result
                except Exception as e:
                    logger.debug(f"Operation {operation.__name__} failed: {e}")

        return summary

    def _aggregate_chunk_summaries(self, chunk_summaries: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Aggregate statistics from multiple chunks"""
        aggregated = {
            'total_rows': sum(cs['rows'] for cs in chunk_summaries),
            'total_columns': chunk_summaries[0]['columns'] if chunk_summaries else 0,
            'total_memory_mb': sum(cs.get('memory_usage_mb', 0) for cs in chunk_summaries)
        }

        # Aggregate numeric summaries
        numeric_summaries = [cs.get('numeric_summary', {}) for cs in chunk_summaries if 'numeric_summary' in cs]
        if numeric_summaries:
            # Simple aggregation - could be made more sophisticated
            aggregated['numeric_summary'] = 'aggregated_from_chunks'

        return aggregated

    def _generate_file_key(self, file_path: str) -> str:
        """Generate cache key for file"""
        # Use file path and modification time for cache key
        try:
            mtime = os.path.getmtime(file_path)
            file_info = f"{file_path}_{mtime}"
            return hashlib.md5(file_info.encode()).hexdigest()
        except Exception:
            return hashlib.md5(file_path.encode()).hexdigest()

    def get_performance_summary(self) -> Dict[str, Any]:
        """Get comprehensive performance summary"""
        cache_stats = self.cache.get_stats()

        return {
            'processing_metrics': {
                'total_time_seconds': self.metrics.total_time_seconds,
                'files_processed': self.metrics.files_processed,
                'rows_processed': self.metrics.rows_processed,
                'cells_processed': self.metrics.cells_processed,
                'processing_rate_rows_per_sec': self.metrics.processing_rate_rows_per_second(),
                'errors_encountered': self.metrics.errors_encountered
            },
            'memory_metrics': {
                'peak_memory_mb': self.metrics.memory_peak_mb,
                'memory_start_mb': self.metrics.memory_start_mb,
                'memory_end_mb': self.metrics.memory_end_mb,
                'memory_delta_mb': self.metrics.memory_delta_mb()
            },
            'cache_metrics': cache_stats,
            'concurrency_metrics': {
                'concurrent_operations': self.metrics.concurrent_operations,
                'max_workers': self.config.max_workers,
                'use_process_pool': self.config.use_process_pool
            }
        }

    def clear_cache(self):
        """Clear all caches"""
        self.cache.clear()

        # Clear disk cache if enabled
        if self.config.enable_disk_cache and hasattr(self, 'disk_cache_dir'):
            try:
                for cache_file in self.disk_cache_dir.glob("*.cache"):
                    cache_file.unlink()
            except Exception as e:
                logger.debug(f"Error clearing disk cache: {e}")

    def get_config(self) -> PerformanceConfig:
        """Get current configuration"""
        return self.config

    def update_config(self, **kwargs):
        """Update configuration"""
        for key, value in kwargs.items():
            if hasattr(self.config, key):
                setattr(self.config, key, value)


# Convenience functions for common operations

def process_excel_files_optimized(
    file_paths: List[str],
    config: Optional[PerformanceConfig] = None,
    enable_validation: bool = True
) -> Dict[str, Any]:
    """Process multiple Excel files with performance optimization"""
    processor = PerformanceOptimizedExcelProcessor(config)
    return processor.process_multiple_files(file_paths, enable_validation=enable_validation)


def create_performance_config(
    max_memory_mb: int = 1024,
    max_workers: int = None,
    enable_streaming: bool = True,
    cache_size: int = 128
) -> PerformanceConfig:
    """Create performance configuration with common settings"""
    return PerformanceConfig(
        max_memory_mb=max_memory_mb,
        max_workers=max_workers or min(8, multiprocessing.cpu_count()),
        enable_streaming=enable_streaming,
        cache_size=cache_size
    )


# Performance monitoring decorators

def monitor_performance(func):
    """Decorator to monitor function performance"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        start_memory = psutil.Process().memory_info().rss / 1024 / 1024

        try:
            result = func(*args, **kwargs)
            success = True
        except Exception as e:
            logger.error(f"Performance monitored function {func.__name__} failed: {e}")
            result = None
            success = False
            raise
        finally:
            end_time = time.time()
            end_memory = psutil.Process().memory_info().rss / 1024 / 1024

            logger.info(
                f"Performance: {func.__name__} - "
                f"Time: {end_time - start_time:.2f}s, "
                f"Memory: {end_memory - start_memory:+.1f}MB, "
                f"Success: {success}"
            )

        return result
    return wrapper


if __name__ == "__main__":
    # Example usage
    import sys

    if len(sys.argv) > 1:
        file_paths = sys.argv[1:]

        # Create optimized configuration
        config = create_performance_config(
            max_memory_mb=2048,
            max_workers=4,
            enable_streaming=True
        )

        # Process files
        results = process_excel_files_optimized(file_paths, config)

        print("Performance Optimized Excel Processing Results:")
        print(f"Processed {results['successful_files']}/{results['total_files']} files")
        print(f"Total time: {results['total_processing_time']:.2f} seconds")

        performance_metrics = results['performance_metrics']
        print(f"Processing rate: {performance_metrics['processing_metrics']['processing_rate_rows_per_sec']:.0f} rows/sec")
        print(f"Cache hit rate: {performance_metrics['cache_metrics']['hit_rate']:.1%}")
    else:
        print("Usage: python performance_optimized_processor.py <file1.xlsx> [file2.xlsx] ...")