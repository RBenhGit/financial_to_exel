"""
Excel Data Source Adapter
=========================

Extracts financial variables from Excel files into standardized VarInputData format.

This adapter extends the existing Excel loading logic from FinancialCalculator
to work with the VarInputData system and FinancialVariableRegistry. It provides:

- Column mapping from Excel headers to standard variable names using aliases
- Support for FY/ and LTM/ folder structures  
- Historical data extraction (10+ years)
- Data validation and quality scoring
- Unit normalization (millions, billions)
- Missing data identification and flagging
- Integration with VarInputData storage system

Key Features:
-------------
- **Intelligent Column Mapping**: Uses FinancialVariableRegistry aliases to map Excel column names
- **Multi-Sheet Support**: Processes Income Statement, Balance Sheet, and Cash Flow Statement
- **Historical Range Detection**: Automatically discovers FY-N columns and date ranges
- **Data Quality Assessment**: Scores data quality based on completeness and validation
- **Unit Conversion**: Handles different unit scales (thousands, millions, billions)
- **Error Recovery**: Robust error handling with detailed logging
- **Memory Efficient**: Processes data in chunks for large historical datasets

Usage Example:
--------------
>>> from excel_adapter import ExcelDataAdapter
>>> from var_input_data import get_var_input_data
>>> 
>>> # Initialize adapter
>>> adapter = ExcelDataAdapter()
>>> 
>>> # Load data for a specific company
>>> result = adapter.load_company_data("AAPL", "data/companies/AAPL")
>>> print(f"Loaded {result['variables_loaded']} variables")
>>> 
>>> # Access the data through VarInputData
>>> var_data = get_var_input_data()
>>> revenue = var_data.get_variable("AAPL", "revenue", period="2023")
>>> print(f"AAPL 2023 Revenue: ${revenue}M")
"""

import logging
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Union

import pandas as pd
from openpyxl import load_workbook

# Import base adapter and types
from .base_adapter import (
    BaseApiAdapter,
    DataSourceType,
    DataCategory,
    ApiCapabilities,
    ExtractionResult,
    DataQualityMetrics
)
from .types import (
    GeneralizedVariableDict,
    AdapterOutputMetadata,
    ValidationResult,
    AdapterException,
    REQUIRED_FIELDS
)
from .adapter_validator import AdapterValidator, ValidationLevel

# Import project dependencies
from ..var_input_data import (
    get_var_input_data,
    VarInputData,
    VariableMetadata,
    DataChangeEvent
)
from ..financial_variable_registry import (
    get_registry,
    FinancialVariableRegistry,
    VariableCategory,
    DataType,
    Units
)

# Configure logging
logger = logging.getLogger(__name__)


@dataclass
class ExcelFileInfo:
    """Information about an Excel file and its contents"""
    file_path: str
    sheet_type: str                    # "income", "balance", "cashflow"
    period_type: str                   # "FY", "LTM"
    company_symbol: str
    available_periods: List[str]       # ["FY", "FY-1", "FY-2", ...]
    header_row: int
    data_quality_score: float = 0.0
    
    def __post_init__(self):
        """Validate the file info after initialization"""
        if self.sheet_type not in ["income", "balance", "cashflow"]:
            raise ValueError(f"Invalid sheet_type: {self.sheet_type}")
        if self.period_type not in ["FY", "LTM"]:
            raise ValueError(f"Invalid period_type: {self.period_type}")


@dataclass 
class VariableExtractionResult:
    """Result of extracting a single variable from Excel"""
    variable_name: str
    excel_column_name: str
    values: Dict[str, Any]             # period -> value mapping
    metadata: VariableMetadata
    quality_score: float
    missing_periods: List[str]
    conversion_applied: Optional[str]  # Unit conversion that was applied
    
    def __post_init__(self):
        """Calculate overall quality score based on data completeness"""
        if not self.values:
            self.quality_score = 0.0
        else:
            # Base score from metadata
            base_score = self.metadata.quality_score
            
            # Adjust based on data completeness
            total_periods = len(self.values) + len(self.missing_periods)
            completeness = len(self.values) / max(1, total_periods)
            
            # Final score combines base score with completeness
            self.quality_score = base_score * 0.7 + completeness * 0.3


class ExcelDataAdapter(BaseApiAdapter):
    """
    Excel Data Source Adapter for extracting financial variables.

    This class provides the main interface for loading Excel financial data
    into the VarInputData system with proper validation and quality scoring.

    Extends BaseApiAdapter to provide consistent interface across all data sources.
    """

    def __init__(
        self,
        timeout: int = 30,
        max_retries: int = 3,
        retry_delay: float = 1.0,
        validation_level: ValidationLevel = ValidationLevel.MODERATE
    ):
        """
        Initialize the Excel adapter with required registries.

        Args:
            timeout: Not used for Excel, kept for interface compatibility
            max_retries: Not used for Excel, kept for interface compatibility
            retry_delay: Not used for Excel, kept for interface compatibility
            validation_level: Validation strictness level
        """
        # Initialize base adapter (no API key needed for Excel)
        super().__init__(
            api_key=None,
            timeout=timeout,
            max_retries=max_retries,
            retry_delay=retry_delay,
            rate_limit_delay=0.0  # No rate limiting for local files
        )

        # Initialize validator
        self.validator = AdapterValidator(level=validation_level)

        # Excel-specific statistics (extends base _stats)
        self._stats.update({
            'files_processed': 0,
            'variables_extracted': 0,
            'data_points_loaded': 0,
            'validation_failures': 0,
            'conversion_applied': 0
        })

        # Standard file name patterns for different statement types
        self._file_patterns = {
            'income': ['Income Statement', 'income_statement', 'income'],
            'balance': ['Balance Sheet', 'balance_sheet', 'balance'],
            'cashflow': ['Cash Flow Statement', 'cash_flow_statement', 'cashflow', 'cash_flow']
        }

        # Current extraction state
        self._current_symbol: Optional[str] = None
        self._current_metadata: Optional[AdapterOutputMetadata] = None

        logger.info("ExcelDataAdapter initialized successfully")

    # ========================================================================
    # BaseApiAdapter Abstract Method Implementations
    # ========================================================================

    def get_source_type(self) -> DataSourceType:
        """Return the data source type for this adapter"""
        return DataSourceType.EXCEL

    def get_capabilities(self) -> ApiCapabilities:
        """Return the capabilities of this Excel adapter"""
        return ApiCapabilities(
            source_type=DataSourceType.EXCEL,
            supported_categories=[
                DataCategory.INCOME_STATEMENT,
                DataCategory.BALANCE_SHEET,
                DataCategory.CASH_FLOW
            ],
            rate_limit_per_minute=0,  # No rate limit for local files
            rate_limit_per_day=None,
            max_historical_years=20,  # Can store 20+ years in Excel
            requires_api_key=False,
            supports_batch_requests=True,  # Can process multiple files
            real_time_data=False,  # Static file data
            cost_per_request=0.0,
            reliability_rating=0.95  # High reliability for local files
        )

    def validate_credentials(self) -> bool:
        """
        Validate Excel adapter readiness.

        Excel doesn't require credentials, but validates that required
        registries are available.

        Returns:
            True if registries are initialized
        """
        return (
            self.variable_registry is not None and
            self.var_data is not None
        )

    def get_available_data(self, symbol: str) -> Dict[str, Any]:
        """
        Check what Excel data is available for a symbol.

        Args:
            symbol: Stock symbol to check

        Returns:
            Dictionary describing available data files and periods
        """
        # This would need a company_data_path parameter in practice
        # For now, return basic structure
        return {
            'symbol': symbol.upper(),
            'source': 'excel',
            'available_categories': ['income', 'balance', 'cashflow'],
            'requires_file_path': True,
            'note': 'Use load_symbol_data() with file_path parameter'
        }

    def load_symbol_data(
        self,
        symbol: str,
        categories: Optional[List[DataCategory]] = None,
        historical_years: int = 5,
        validate_data: bool = True
    ) -> ExtractionResult:
        """
        Load financial data for a symbol from Excel files.

        This method serves as a bridge to the existing load_company_data() method
        for backward compatibility while conforming to the BaseApiAdapter interface.

        Args:
            symbol: Stock symbol (e.g., "AAPL")
            categories: List of data categories to retrieve (all if None)
            historical_years: Years of historical data to retrieve
            validate_data: Whether to validate data using registry definitions

        Returns:
            ExtractionResult with detailed results and metrics

        Note:
            This method requires that company_data_path was set via
            extract_variables() or you can use load_company_data() directly.
        """
        # This is a simplified wrapper - in practice, callers should use
        # extract_variables() or load_company_data() which have file_path params
        logger.warning(
            f"load_symbol_data() called for {symbol} - this method requires "
            "file paths. Use extract_variables() with file_path instead."
        )

        return ExtractionResult(
            source=DataSourceType.EXCEL,
            symbol=symbol.upper(),
            success=False,
            variables_extracted=0,
            data_points_stored=0,
            categories_covered=[],
            periods_covered=[],
            quality_metrics=DataQualityMetrics(
                completeness_score=0.0,
                timeliness_score=0.0,
                consistency_score=0.0,
                reliability_score=0.0,
                overall_score=0.0,
                issues=["File path required for Excel adapter"],
                metadata={}
            ),
            extraction_time=0.0,
            errors=["Excel adapter requires file_path parameter"],
            warnings=[],
            metadata={'note': 'Use extract_variables() with file_path'}
        )

    def extract_variables(
        self,
        symbol: str,
        period: str = "latest",
        historical_years: int = 10,
        file_path: Optional[str] = None,
        company_data_path: Optional[str] = None
    ) -> GeneralizedVariableDict:
        """
        Extract financial variables from Excel files and return standardized dict.

        This is the core method that extracts data from Excel files and transforms
        it into the standardized GeneralizedVariableDict format.

        Args:
            symbol: Stock symbol (e.g., "AAPL")
            period: Period identifier ("latest", "FY", "LTM", "2023", etc.)
            historical_years: Number of years of historical data to include
            file_path: Path to specific Excel file (optional)
            company_data_path: Path to company data folder containing FY/LTM subfolders

        Returns:
            GeneralizedVariableDict with standardized variable names and values

        Raises:
            AdapterException: On extraction or transformation failures
        """
        symbol = symbol.upper().strip()
        self._current_symbol = symbol

        try:
            start_time = datetime.now()

            # Initialize result dictionary with required fields
            result: GeneralizedVariableDict = {
                'ticker': symbol,
                'company_name': f"{symbol} Inc.",  # Will be updated if found
                'currency': 'USD',
                'fiscal_year_end': 'December',  # Will be updated if found
                'data_source': 'excel',
                'data_timestamp': start_time,
                'last_updated': start_time
            }

            if company_data_path:
                # Use existing comprehensive load logic
                load_results = self.load_company_data(
                    symbol=symbol,
                    company_data_path=company_data_path,
                    load_fy=(period in ["latest", "FY"]),
                    load_ltm=(period == "LTM"),
                    max_historical_years=historical_years,
                    validate_data=True
                )

                # Convert loaded data to GeneralizedVariableDict format
                # Data was already stored in VarInputData, now retrieve it
                result = self._build_generalized_dict_from_var_data(
                    symbol, period, load_results
                )

            elif file_path:
                # Load single file
                sheet_type = self._determine_sheet_type(os.path.basename(file_path))
                period_type = "FY" if "FY" in file_path else "LTM"

                if not sheet_type:
                    raise AdapterException(
                        f"Could not determine sheet type from file: {file_path}",
                        source="excel"
                    )

                file_results = self.load_single_file(
                    symbol=symbol,
                    file_path=file_path,
                    sheet_type=sheet_type,
                    period_type=period_type,
                    validate_data=True
                )

                # Build result from file data
                result = self._build_generalized_dict_from_file(
                    symbol, period, file_results
                )

            else:
                raise AdapterException(
                    "Either file_path or company_data_path must be provided",
                    source="excel"
                )

            # Calculate extraction time
            extraction_time = (datetime.now() - start_time).total_seconds()

            # Store metadata
            completeness = self.calculate_completeness_score(result)
            quality_score = result.get('data_quality_score', 0.8)

            self._current_metadata = AdapterOutputMetadata(
                source="excel",
                timestamp=start_time,
                quality_score=quality_score,
                completeness=completeness,
                validation_errors=[],
                extraction_time=extraction_time,
                cache_hit=False,
                api_calls_made=0
            )

            logger.info(
                f"Extracted variables for {symbol}: "
                f"{len([v for v in result.values() if v is not None])} fields, "
                f"completeness={completeness:.2%}"
            )

            return result

        except Exception as e:
            logger.error(f"Failed to extract variables for {symbol}: {e}")
            raise AdapterException(
                f"Excel extraction failed for {symbol}",
                source="excel",
                original_exception=e
            )

    def get_extraction_metadata(self) -> AdapterOutputMetadata:
        """
        Return metadata about the most recent extraction operation.

        Returns:
            AdapterOutputMetadata for the last extraction
        """
        if self._current_metadata is None:
            # Return default metadata if no extraction has occurred
            return AdapterOutputMetadata(
                source="excel",
                timestamp=datetime.now(),
                quality_score=0.0,
                completeness=0.0,
                validation_errors=["No extraction performed yet"],
                extraction_time=0.0,
                cache_hit=False,
                api_calls_made=0
            )
        return self._current_metadata

    def validate_output(self, variables: GeneralizedVariableDict) -> ValidationResult:
        """
        Validate that output conforms to GeneralizedVariableDict schema.

        Uses AdapterValidator to perform comprehensive validation including:
        - Required fields presence
        - Data type correctness
        - Value range validation
        - Business rule compliance

        Args:
            variables: Dictionary to validate

        Returns:
            ValidationResult with validation status and any errors
        """
        try:
            # Use the comprehensive validator
            report = self.validator.validate(variables, include_quality_score=True)

            # Convert ValidationReport to ValidationResult
            result = ValidationResult(
                valid=report.valid,
                validation_type="comprehensive"
            )

            # Add errors and warnings
            for error in report.errors:
                result.add_error(f"{error.field}: {error.message}")

            for warning in report.warnings:
                result.add_warning(f"{warning.field}: {warning.message}")

            # Add details
            result.details.update({
                'quality_score': report.quality_score,
                'completeness_score': report.completeness_score,
                'consistency_score': report.consistency_score,
                'overall_score': report.overall_score,
                'fields_validated': report.fields_validated,
                'fields_passed': report.fields_passed,
                'fields_failed': report.fields_failed,
                'validation_level': report.validation_level.value
            })

            return result

        except Exception as e:
            logger.error(f"Validation failed: {e}")
            result = ValidationResult(valid=False, validation_type="comprehensive")
            result.add_error(f"Validation exception: {str(e)}")
            return result

    def get_supported_variable_categories(self) -> List[str]:
        """
        Return list of variable categories this adapter supports.

        Returns:
            List of supported category names
        """
        return [
            'income_statement',
            'balance_sheet',
            'cash_flow'
        ]

    # ========================================================================
    # Helper Methods for BaseApiAdapter Integration
    # ========================================================================

    def _build_generalized_dict_from_var_data(
        self,
        symbol: str,
        period: str,
        load_results: Dict[str, Any]
    ) -> GeneralizedVariableDict:
        """
        Build GeneralizedVariableDict from data stored in VarInputData.

        Args:
            symbol: Stock symbol
            period: Period identifier
            load_results: Results from load_company_data()

        Returns:
            GeneralizedVariableDict with retrieved data
        """
        result: GeneralizedVariableDict = {
            'ticker': symbol,
            'company_name': f"{symbol} Inc.",
            'currency': 'USD',
            'fiscal_year_end': 'December',
            'data_source': 'excel',
            'data_timestamp': datetime.now(),
            'last_updated': datetime.now(),
            'data_quality_score': load_results.get('quality_scores', {}).get('overall', 0.8),
            'completeness_score': load_results.get('completeness_score', 0.7)
        }

        # Retrieve all loaded variables from VarInputData
        all_variables = self.variable_registry.list_all_variables()

        for var_name in all_variables:
            try:
                value = self.var_data.get_variable(
                    symbol=symbol,
                    variable_name=var_name,
                    period=period if period != "latest" else None
                )
                if value is not None:
                    result[var_name] = value  # type: ignore
            except Exception:
                # Variable not available, skip
                pass

        return result

    def _build_generalized_dict_from_file(
        self,
        symbol: str,
        period: str,
        file_results: Dict[str, Any]
    ) -> GeneralizedVariableDict:
        """
        Build GeneralizedVariableDict from single file results.

        Args:
            symbol: Stock symbol
            period: Period identifier
            file_results: Results from load_single_file()

        Returns:
            GeneralizedVariableDict with file data
        """
        # Similar to _build_generalized_dict_from_var_data but for single file
        return self._build_generalized_dict_from_var_data(symbol, period, file_results)

    # ========================================================================
    # Existing Excel-Specific Methods (Preserved for Backward Compatibility)
    # ========================================================================

    def load_company_data(
        self,
        symbol: str,
        company_data_path: str,
        load_fy: bool = True,
        load_ltm: bool = True,
        max_historical_years: int = 10,
        validate_data: bool = True
    ) -> Dict[str, Any]:
        """
        Load all financial data for a company from its Excel files.
        
        Args:
            symbol: Stock symbol (e.g., "AAPL")
            company_data_path: Path to company data folder
            load_fy: Whether to load FY (full year) data
            load_ltm: Whether to load LTM (last twelve months) data
            max_historical_years: Maximum years of historical data to load
            validate_data: Whether to validate data using registry definitions
            
        Returns:
            Dictionary with loading results and statistics
        """
        symbol = symbol.upper().strip()
        results = {
            'symbol': symbol,
            'files_processed': 0,
            'variables_loaded': 0,
            'data_points_loaded': 0,
            'errors': [],
            'quality_scores': {},
            'period_coverage': {},
            'file_details': []
        }
        
        logger.info(f"Starting data load for {symbol} from {company_data_path}")
        
        try:
            # Process FY data
            if load_fy:
                fy_path = os.path.join(company_data_path, "FY")
                if os.path.exists(fy_path):
                    fy_results = self._process_folder(
                        symbol, fy_path, "FY", max_historical_years, validate_data
                    )
                    self._merge_results(results, fy_results)
                else:
                    logger.warning(f"FY folder not found: {fy_path}")
            
            # Process LTM data
            if load_ltm:
                ltm_path = os.path.join(company_data_path, "LTM")
                if os.path.exists(ltm_path):
                    ltm_results = self._process_folder(
                        symbol, ltm_path, "LTM", max_historical_years, validate_data
                    )
                    self._merge_results(results, ltm_results)
                else:
                    logger.warning(f"LTM folder not found: {ltm_path}")
            
            # Update global statistics
            self._stats['files_processed'] += results['files_processed']
            self._stats['variables_extracted'] += results['variables_loaded']
            self._stats['data_points_loaded'] += results['data_points_loaded']
            
            logger.info(f"Completed data load for {symbol}: {results['variables_loaded']} variables, "
                       f"{results['data_points_loaded']} data points")
            
        except Exception as e:
            error_msg = f"Failed to load data for {symbol}: {str(e)}"
            logger.error(error_msg)
            results['errors'].append(error_msg)
        
        return results
    
    def load_single_file(
        self,
        symbol: str,
        file_path: str,
        sheet_type: str,
        period_type: str,
        validate_data: bool = True
    ) -> Dict[str, Any]:
        """
        Load data from a single Excel file.
        
        Args:
            symbol: Stock symbol
            file_path: Path to Excel file
            sheet_type: Type of sheet ("income", "balance", "cashflow")
            period_type: Type of period ("FY", "LTM")
            validate_data: Whether to validate data
            
        Returns:
            Dictionary with extraction results
        """
        symbol = symbol.upper().strip()
        
        logger.info(f"Loading single file for {symbol}: {file_path}")
        
        try:
            # Analyze the file structure first
            file_info = self._analyze_excel_file(file_path, symbol, sheet_type, period_type)
            
            # Extract variables from the file
            extraction_results = self._extract_variables_from_file(file_info, validate_data)
            
            # Store data in VarInputData
            storage_results = self._store_extracted_data(extraction_results)
            
            # Compile results
            results = {
                'symbol': symbol,
                'file_path': file_path,
                'sheet_type': sheet_type,
                'period_type': period_type,
                'variables_extracted': len(extraction_results),
                'data_points_stored': storage_results['stored_count'],
                'periods_covered': file_info.available_periods,
                'quality_score': file_info.data_quality_score,
                'errors': storage_results.get('errors', [])
            }
            
            logger.info(f"Successfully loaded {results['variables_extracted']} variables from {file_path}")
            return results
            
        except Exception as e:
            error_msg = f"Failed to load file {file_path}: {str(e)}"
            logger.error(error_msg)
            return {
                'symbol': symbol,
                'file_path': file_path,
                'errors': [error_msg],
                'variables_extracted': 0,
                'data_points_stored': 0
            }
    
    def get_adapter_statistics(self) -> Dict[str, Any]:
        """Get comprehensive statistics about the adapter's operations"""
        var_data_stats = self.var_data.get_statistics()
        
        return {
            'adapter_stats': dict(self._stats),
            'var_data_stats': var_data_stats,
            'registry_info': {
                'total_variables': len(self.variable_registry.list_all_variables()),
                'categories': [cat.value for cat in VariableCategory]
            }
        }
    
    # Private helper methods
    
    def _process_folder(
        self,
        symbol: str,
        folder_path: str,
        period_type: str,
        max_historical_years: int,
        validate_data: bool
    ) -> Dict[str, Any]:
        """Process all Excel files in a folder (FY or LTM)"""
        results = {
            'files_processed': 0,
            'variables_loaded': 0,
            'data_points_loaded': 0,
            'errors': [],
            'file_details': []
        }
        
        try:
            for file_name in os.listdir(folder_path):
                if not file_name.endswith(('.xlsx', '.xls')):
                    continue
                    
                file_path = os.path.join(folder_path, file_name)
                sheet_type = self._determine_sheet_type(file_name)
                
                if sheet_type:
                    file_results = self.load_single_file(
                        symbol, file_path, sheet_type, period_type, validate_data
                    )
                    
                    # Merge file results
                    results['files_processed'] += 1
                    results['variables_loaded'] += file_results.get('variables_extracted', 0)
                    results['data_points_loaded'] += file_results.get('data_points_stored', 0)
                    results['file_details'].append(file_results)
                    
                    if file_results.get('errors'):
                        results['errors'].extend(file_results['errors'])
                else:
                    logger.warning(f"Could not determine sheet type for file: {file_name}")
                    
        except Exception as e:
            error_msg = f"Error processing folder {folder_path}: {str(e)}"
            logger.error(error_msg)
            results['errors'].append(error_msg)
        
        return results
    
    def _determine_sheet_type(self, file_name: str) -> Optional[str]:
        """Determine the type of financial statement from file name"""
        file_name_lower = file_name.lower()
        
        for sheet_type, patterns in self._file_patterns.items():
            for pattern in patterns:
                if pattern.lower() in file_name_lower:
                    return sheet_type
        
        return None
    
    def _analyze_excel_file(
        self,
        file_path: str,
        symbol: str,
        sheet_type: str,
        period_type: str
    ) -> ExcelFileInfo:
        """
        Analyze Excel file structure and discover available periods.
        
        This method replicates and extends the logic from FinancialCalculator._load_excel_data()
        """
        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active
            
            # Convert to list of lists for processing
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            
            # Find the header row (contains 'FY-N', 'FY', etc.)
            header_row_idx = None
            for i, row in enumerate(data):
                if row and any('FY-' in str(cell) or 'FY' == str(cell) for cell in row if cell):
                    header_row_idx = i
                    break
            
            if header_row_idx is None:
                raise ValueError(f"Could not find header row with FY columns in {file_path}")
            
            headers = data[header_row_idx]
            
            # Discover available periods using the same logic as FinancialCalculator
            available_periods = []
            fy_info = {}
            
            for col_idx, header in enumerate(headers):
                if header is not None:
                    header_str = str(header).strip()
                    
                    # Match FY-N pattern (e.g., FY-9, FY-8, FY-1)
                    if header_str.startswith('FY-'):
                        try:
                            years_back = int(header_str[3:])
                            fy_info[col_idx] = {
                                'column': header_str,
                                'years_back': years_back,
                                'sort_key': years_back
                            }
                            available_periods.append(header_str)
                        except ValueError:
                            continue
                    
                    # Match exact 'FY' (current year)
                    elif header_str == 'FY':
                        fy_info[col_idx] = {
                            'column': 'FY',
                            'years_back': 0,
                            'sort_key': 0
                        }
                        available_periods.append('FY')
            
            # Sort periods by recency (FY, FY-1, FY-2, etc.)
            available_periods.sort(key=lambda x: int(x.split('-')[1]) if '-' in x else 0)
            
            # Calculate data quality score based on completeness and structure
            quality_score = self._calculate_file_quality_score(data, header_row_idx, fy_info)
            
            file_info = ExcelFileInfo(
                file_path=file_path,
                sheet_type=sheet_type,
                period_type=period_type,
                company_symbol=symbol,
                available_periods=available_periods,
                header_row=header_row_idx,
                data_quality_score=quality_score
            )
            
            logger.debug(f"Analyzed {file_path}: {len(available_periods)} periods found, quality={quality_score:.2f}")
            return file_info
            
        except Exception as e:
            logger.error(f"Failed to analyze Excel file {file_path}: {str(e)}")
            raise
    
    def _calculate_file_quality_score(
        self,
        data: List[Tuple],
        header_row_idx: int,
        fy_info: Dict[int, Dict]
    ) -> float:
        """Calculate data quality score for the file"""
        try:
            # Base score factors
            structure_score = 1.0    # File has proper structure
            
            # Period coverage score (more historical data = better)
            period_count = len(fy_info)
            coverage_score = min(1.0, period_count / 10.0)  # 10 years = perfect score
            
            # Data completeness score (check for missing values)
            if header_row_idx + 1 < len(data):
                data_rows = data[header_row_idx + 1:]
                total_cells = 0
                filled_cells = 0
                
                # Check data in FY columns
                fy_columns = list(fy_info.keys())
                
                for row in data_rows[:50]:  # Check first 50 data rows
                    for col_idx in fy_columns:
                        if col_idx < len(row):
                            total_cells += 1
                            if row[col_idx] is not None and str(row[col_idx]).strip():
                                filled_cells += 1
                
                completeness_score = filled_cells / max(1, total_cells)
            else:
                completeness_score = 0.0
            
            # Weighted average of quality factors
            final_score = (
                structure_score * 0.3 +
                coverage_score * 0.4 +
                completeness_score * 0.3
            )
            
            return round(final_score, 2)
            
        except Exception as e:
            logger.warning(f"Failed to calculate quality score: {str(e)}")
            return 0.5  # Default neutral score
    
    def _extract_variables_from_file(
        self,
        file_info: ExcelFileInfo,
        validate_data: bool
    ) -> List[VariableExtractionResult]:
        """Extract financial variables from the Excel file using registry aliases"""
        extraction_results = []
        
        try:
            # Load the Excel data using openpyxl (same as FinancialCalculator)
            wb = load_workbook(filename=file_info.file_path)
            sheet = wb.active
            
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            
            headers = data[file_info.header_row]
            
            # Get all variables from registry that might be in this sheet type
            relevant_variables = self._get_relevant_variables_for_sheet(file_info.sheet_type)
            
            # Find FY column indices for period mapping
            period_column_map = {}  # period_name -> column_index
            for col_idx, header in enumerate(headers):
                if header and str(header).strip() in file_info.available_periods:
                    period_column_map[str(header).strip()] = col_idx
            
            # Process each row to find variable matches
            for row_idx, row in enumerate(data[file_info.header_row + 1:], start=file_info.header_row + 1):
                if not row or len(row) == 0:
                    continue
                
                # Get the row label - try multiple columns as different Excel files may have different structures
                row_label = ""
                for col_idx in range(min(5, len(row))):  # Check first 5 columns
                    if row[col_idx] is not None and str(row[col_idx]).strip():
                        candidate = str(row[col_idx]).strip()
                        # Skip dates and pure numbers - we want variable names
                        if (not candidate.replace('-', '').replace('.', '').replace(',', '').replace('%', '').isdigit() and 
                            not any(date_part in candidate for date_part in ['2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', '2023', '2024'])):
                            row_label = candidate
                            break
                
                if not row_label:
                    continue
                
                # Try to match this row to a variable in the registry
                matched_variable = self._match_row_to_variable(
                    row_label, relevant_variables, file_info.sheet_type
                )
                
                if matched_variable:
                    # Extract values for each available period
                    values = {}
                    missing_periods = []
                    
                    for period, col_idx in period_column_map.items():
                        if col_idx < len(row):
                            cell_value = row[col_idx]
                            if cell_value is not None and str(cell_value).strip():
                                # Apply unit conversion if needed
                                converted_value, conversion_info = self._convert_units(
                                    cell_value, matched_variable
                                )
                                values[period] = converted_value
                            else:
                                missing_periods.append(period)
                    
                    if values:  # Only include if we have at least some data
                        # Create metadata
                        metadata = VariableMetadata(
                            source="excel",
                            timestamp=datetime.now(),
                            quality_score=file_info.data_quality_score,
                            validation_passed=True,  # Will be updated during validation
                            period=file_info.period_type,
                            lineage_id=f"{file_info.company_symbol}_{file_info.sheet_type}_{row_idx}"
                        )
                        
                        # Validate data if requested
                        if validate_data:
                            validation_result = self._validate_variable_data(
                                matched_variable, values, metadata
                            )
                            metadata.validation_passed = validation_result['passed']
                            if not validation_result['passed']:
                                self._stats['validation_failures'] += 1
                                metadata.quality_score *= 0.8  # Reduce quality score
                        
                        # Create extraction result
                        result = VariableExtractionResult(
                            variable_name=matched_variable.name,
                            excel_column_name=row_label,
                            values=values,
                            metadata=metadata,
                            quality_score=metadata.quality_score,
                            missing_periods=missing_periods,
                            conversion_applied=conversion_info
                        )
                        
                        extraction_results.append(result)
                        logger.debug(f"Extracted {matched_variable.name} with {len(values)} periods")
            
            logger.info(f"Extracted {len(extraction_results)} variables from {file_info.file_path}")
            return extraction_results
            
        except Exception as e:
            logger.error(f"Failed to extract variables from {file_info.file_path}: {str(e)}")
            raise
    
    def _get_relevant_variables_for_sheet(self, sheet_type: str) -> List:
        """Get variables from registry that are relevant for this sheet type"""
        category_mapping = {
            'income': [VariableCategory.INCOME_STATEMENT],
            'balance': [VariableCategory.BALANCE_SHEET],
            'cashflow': [VariableCategory.CASH_FLOW]
        }
        
        relevant_categories = category_mapping.get(sheet_type, [])
        all_variables = self.variable_registry.list_all_variables()
        
        relevant_variables = []
        for var_name in all_variables:
            var_def = self.variable_registry.get_variable_definition(var_name)
            if var_def and var_def.category in relevant_categories:
                relevant_variables.append(var_def)
        
        return relevant_variables
    
    def _match_row_to_variable(
        self,
        row_label: str,
        relevant_variables: List,
        sheet_type: str
    ):
        """Match an Excel row label to a variable definition using aliases"""
        row_label_clean = row_label.lower().strip()
        
        # Try exact name match first
        for var_def in relevant_variables:
            if var_def.name.lower() == row_label_clean:
                return var_def
        
        # Try alias matching
        for var_def in relevant_variables:
            if hasattr(var_def, 'aliases') and var_def.aliases:
                # Check excel-specific aliases
                if 'excel' in var_def.aliases:
                    excel_aliases = var_def.aliases['excel']
                    if isinstance(excel_aliases, str):
                        excel_aliases = [excel_aliases]
                    
                    for alias in excel_aliases:
                        if alias and alias.lower().strip() == row_label_clean:
                            return var_def
                
                # Check generic aliases
                for alias_source, aliases in var_def.aliases.items():
                    if isinstance(aliases, str):
                        aliases = [aliases]
                    elif aliases is None:
                        continue
                    
                    for alias in aliases:
                        if alias and alias.lower().strip() == row_label_clean:
                            return var_def
        
        # Try common Excel variations and mapping
        excel_mappings = {
            'revenue': ['revenue', 'total revenue', 'net sales', 'sales', 'net revenue'],
            'cost_of_revenue': ['cost of revenues', 'cost of sales', 'cogs', 'cost of goods sold'],
            'gross_profit': ['gross profit', 'gross income'],
            'operating_income': ['operating income', 'operating profit', 'ebit'],
            'net_income': ['net income', 'net earnings', 'profit', 'bottom line'],
            'rd_expenses': ['r&d expenses', 'research and development', 'rd expense'],
            'sga_expenses': ['selling and marketing expense', 'general & admin expenses', 'sga', 'sg&a'],
            'operating_expenses': ['operating expenses', 'total operating expenses'],
            'interest_expense': ['net interest expenses', 'interest expense'],
            'ebt': ['ebt, incl. unusual items', 'earnings before tax', 'pre-tax income'],
        }
        
        for var_name, variations in excel_mappings.items():
            for variation in variations:
                if variation == row_label_clean:
                    # Find the variable in registry
                    for var_def in relevant_variables:
                        if var_def.name == var_name:
                            logger.debug(f"Excel mapping match: '{row_label}' -> '{var_name}'")
                            return var_def
        
        # Try partial matching for common variations
        for var_def in relevant_variables:
            var_name_clean = var_def.name.lower().replace('_', ' ')
            row_label_words = set(row_label_clean.replace('_', ' ').replace('&', ' ').split())
            var_name_words = set(var_name_clean.split())
            
            # If most words match, consider it a match
            common_words = row_label_words & var_name_words
            if len(common_words) >= min(2, max(1, len(var_name_words) * 0.6)):
                logger.debug(f"Partial match: '{row_label}' -> '{var_def.name}'")
                return var_def
        
        # Log unmatched variables for debugging
        logger.debug(f"No match found for Excel variable: '{row_label}'")
        return None
    
    def _convert_units(self, value: Any, variable_def) -> Tuple[Any, Optional[str]]:
        """Convert units based on variable definition and value magnitude"""
        try:
            # Try to convert to float
            if isinstance(value, str):
                # Clean the string (remove commas, parentheses, etc.)
                cleaned = re.sub(r'[,\(\)\$]', '', value.strip())
                if cleaned.lower() in ['', '-', 'n/a', 'na', 'none']:
                    return None, None
                
                numeric_value = float(cleaned)
            elif isinstance(value, (int, float)):
                numeric_value = float(value)
            else:
                return value, None  # Can't convert
            
            # Determine unit conversion based on magnitude and variable definition
            conversion_applied = None
            
            # Check if value seems to be in thousands/millions/billions based on magnitude
            abs_value = abs(numeric_value)
            
            if hasattr(variable_def, 'units') and variable_def.units:
                target_unit = variable_def.units
                
                # Apply conversion based on value magnitude and target unit
                if target_unit == Units.MILLIONS_USD:
                    if abs_value > 1_000_000:  # Value seems to be in actual dollars
                        numeric_value = numeric_value / 1_000_000
                        conversion_applied = "dollars_to_millions"
                        self._stats['conversion_applied'] += 1
                    elif abs_value < 1000 and abs_value > 0.1:  # Value seems to be in billions
                        numeric_value = numeric_value * 1000
                        conversion_applied = "billions_to_millions" 
                        self._stats['conversion_applied'] += 1
                elif target_unit == Units.PERCENTAGE:
                    if abs_value > 1:  # Value seems to be in percentage points
                        numeric_value = numeric_value / 100
                        conversion_applied = "percentage_points_to_decimal"
                        self._stats['conversion_applied'] += 1
            
            return numeric_value, conversion_applied
            
        except (ValueError, TypeError) as e:
            logger.debug(f"Could not convert value '{value}': {str(e)}")
            return value, None
    
    def _validate_variable_data(
        self,
        variable_def,
        values: Dict[str, Any],
        metadata: VariableMetadata
    ) -> Dict[str, Any]:
        """Validate extracted variable data using registry definition"""
        validation_result = {
            'passed': True,
            'errors': [],
            'warnings': []
        }
        
        try:
            # Validate each value
            for period, value in values.items():
                if hasattr(variable_def, 'validate_value'):
                    is_valid, errors = variable_def.validate_value(value)
                    if not is_valid:
                        validation_result['passed'] = False
                        validation_result['errors'].extend([
                            f"Period {period}: {error}" for error in errors
                        ])
                
                # Additional custom validations can be added here
                # E.g., range checks, logical consistency, etc.
        
        except Exception as e:
            validation_result['passed'] = False
            validation_result['errors'].append(f"Validation error: {str(e)}")
            logger.warning(f"Validation failed for {variable_def.name}: {str(e)}")
        
        return validation_result
    
    def _store_extracted_data(
        self,
        extraction_results: List[VariableExtractionResult]
    ) -> Dict[str, Any]:
        """Store extracted data in VarInputData system"""
        storage_results = {
            'stored_count': 0,
            'failed_count': 0,
            'errors': []
        }
        
        try:
            for result in extraction_results:
                # Store each period's value
                for period, value in result.values.items():
                    success = self.var_data.set_variable(
                        symbol=result.metadata.lineage_id.split('_')[0],  # Extract symbol
                        variable_name=result.variable_name,
                        value=value,
                        period=period,
                        source=result.metadata.source,
                        metadata=result.metadata,
                        validate=False,  # Already validated in extraction
                        emit_event=True
                    )
                    
                    if success:
                        storage_results['stored_count'] += 1
                        self._stats['data_points_loaded'] += 1
                    else:
                        storage_results['failed_count'] += 1
                        storage_results['errors'].append(
                            f"Failed to store {result.variable_name}[{period}]"
                        )
            
            logger.info(f"Stored {storage_results['stored_count']} data points, "
                       f"failed: {storage_results['failed_count']}")
            
        except Exception as e:
            error_msg = f"Failed to store extracted data: {str(e)}"
            logger.error(error_msg)
            storage_results['errors'].append(error_msg)
        
        return storage_results
    
    def _merge_results(self, main_results: Dict, new_results: Dict) -> None:
        """Merge new results into main results dictionary"""
        main_results['files_processed'] += new_results.get('files_processed', 0)
        main_results['variables_loaded'] += new_results.get('variables_loaded', 0)
        main_results['data_points_loaded'] += new_results.get('data_points_loaded', 0)
        
        if new_results.get('errors'):
            main_results['errors'].extend(new_results['errors'])
        
        if new_results.get('file_details'):
            main_results['file_details'].extend(new_results['file_details'])


# Convenience functions for common operations

def load_company_excel_data(
    symbol: str,
    company_data_path: str,
    **kwargs
) -> Dict[str, Any]:
    """
    Convenience function to load Excel data for a company.
    
    Args:
        symbol: Stock symbol
        company_data_path: Path to company data folder
        **kwargs: Additional arguments for ExcelDataAdapter.load_company_data()
        
    Returns:
        Loading results dictionary
    """
    adapter = ExcelDataAdapter()
    return adapter.load_company_data(symbol, company_data_path, **kwargs)


def get_excel_adapter_stats() -> Dict[str, Any]:
    """
    Convenience function to get Excel adapter statistics.
    
    Returns:
        Statistics dictionary
    """
    adapter = ExcelDataAdapter()
    return adapter.get_adapter_statistics()