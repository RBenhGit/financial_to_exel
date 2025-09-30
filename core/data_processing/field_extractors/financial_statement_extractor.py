"""
Financial Statement Field Extractor
===================================

Comprehensive system for extracting all available fields from financial statement Excel files.
Provides base classes, data models, and specialized extractors for Income Statement,
Balance Sheet, and Cash Flow Statement.

This module implements a unified field extraction framework that:
- Automatically detects and extracts all financial line items
- Supports both FY and LTM data structures
- Provides comprehensive field mapping dictionaries
- Includes data validation and missing field detection
- Follows existing data processing patterns

Key Features:
- BaseFieldExtractor abstract class with common functionality
- Specialized extractors for each statement type
- Comprehensive field mapping and validation
- Data quality assessment and scoring
- Integration with existing openpyxl-based processing
"""

import logging
import re
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import Any, Dict, List, Optional, Set, Tuple, Union

import openpyxl
from openpyxl import Workbook, load_workbook

from ..exceptions import (
    DataSourceException,
    DataParsingException,
    DataValidationException,
    ErrorSeverity,
    ErrorCategory,
    RecoveryStrategy
)

logger = logging.getLogger(__name__)


class StatementType(Enum):
    """Types of financial statements for field extraction"""
    INCOME = "income_statement"
    BALANCE = "balance_sheet"
    CASH_FLOW = "cash_flow_statement"


class FieldValidationError(DataSourceException):
    """Exception raised when field validation fails"""

    def __init__(self, field_name: str, value: Any, validation_rule: str, **kwargs):
        message = f"Field validation failed for '{field_name}': {validation_rule} (value: {value})"
        super().__init__(
            message,
            error_category=ErrorCategory.VALIDATION,
            severity=ErrorSeverity.LOW,
            recovery_strategy=RecoveryStrategy.DEGRADE_GRACEFULLY,
            **kwargs
        )
        self.field_name = field_name
        self.value = value
        self.validation_rule = validation_rule


class MissingFieldError(DataSourceException):
    """Exception raised when required fields are missing"""

    def __init__(self, missing_fields: List[str], statement_type: StatementType, **kwargs):
        message = f"Missing required fields in {statement_type.value}: {', '.join(missing_fields)}"
        super().__init__(
            message,
            error_category=ErrorCategory.DATA_QUALITY,
            severity=ErrorSeverity.MEDIUM,
            recovery_strategy=RecoveryStrategy.DEGRADE_GRACEFULLY,
            **kwargs
        )
        self.missing_fields = missing_fields
        self.statement_type = statement_type


@dataclass
class FieldMappingDict:
    """Comprehensive field mapping for standardized field names"""

    # Core field mappings
    primary_names: Dict[str, str] = field(default_factory=dict)

    # Alternative names and aliases
    aliases: Dict[str, List[str]] = field(default_factory=dict)

    # Excel-specific variations
    excel_variations: Dict[str, List[str]] = field(default_factory=dict)

    # Common abbreviations and short forms
    abbreviations: Dict[str, List[str]] = field(default_factory=dict)

    # Regional/international variations
    international_names: Dict[str, List[str]] = field(default_factory=dict)

    def get_all_mappings_for_field(self, standard_field_name: str) -> List[str]:
        """Get all possible mappings for a standard field name"""
        mappings = [standard_field_name]

        # Add primary name if different
        if standard_field_name in self.primary_names:
            mappings.append(self.primary_names[standard_field_name])

        # Add all variations
        for mapping_dict in [self.aliases, self.excel_variations,
                           self.abbreviations, self.international_names]:
            if standard_field_name in mapping_dict:
                mappings.extend(mapping_dict[standard_field_name])

        # Remove duplicates while preserving order
        unique_mappings = []
        for mapping in mappings:
            if mapping not in unique_mappings:
                unique_mappings.append(mapping)

        return unique_mappings

    def find_standard_field_name(self, input_name: str) -> Optional[str]:
        """Find the standard field name for an input name"""
        input_clean = input_name.lower().strip()

        # Check direct match with standard names
        for standard_name in self.primary_names.keys():
            if standard_name.lower() == input_clean:
                return standard_name

        # Check all mapping dictionaries
        for mapping_dict in [self.primary_names, self.aliases, self.excel_variations,
                           self.abbreviations, self.international_names]:
            for standard_name, variations in mapping_dict.items():
                if isinstance(variations, str):
                    variations = [variations]

                for variation in variations:
                    if variation and variation.lower().strip() == input_clean:
                        return standard_name

        return None


@dataclass
class FieldExtractionResult:
    """Result of extracting fields from a financial statement"""

    # Statement information
    statement_type: StatementType
    company_symbol: str
    period_type: str  # "FY" or "LTM"
    file_path: str

    # Extraction results
    extracted_fields: Dict[str, Dict[str, Any]]  # field_name -> {period -> value}
    missing_fields: List[str]
    invalid_fields: Dict[str, str]  # field_name -> error_message

    # Quality metrics
    total_fields_attempted: int
    successful_extractions: int
    data_quality_score: float

    # Processing metadata
    extraction_timestamp: datetime = field(default_factory=datetime.now)
    processing_time_seconds: float = 0.0

    # Additional context
    available_periods: List[str] = field(default_factory=list)
    header_row_index: int = -1
    notes: List[str] = field(default_factory=list)

    @property
    def success_rate(self) -> float:
        """Calculate field extraction success rate"""
        if self.total_fields_attempted == 0:
            return 0.0
        return self.successful_extractions / self.total_fields_attempted

    @property
    def completeness_score(self) -> float:
        """Calculate data completeness score"""
        total_expected = len(self.extracted_fields) + len(self.missing_fields)
        if total_expected == 0:
            return 1.0
        return len(self.extracted_fields) / total_expected

    def get_field_value(self, field_name: str, period: str) -> Optional[Any]:
        """Get value for a specific field and period"""
        if field_name in self.extracted_fields:
            return self.extracted_fields[field_name].get(period)
        return None

    def add_note(self, note: str) -> None:
        """Add a processing note"""
        self.notes.append(f"[{datetime.now().strftime('%H:%M:%S')}] {note}")


class BaseFieldExtractor(ABC):
    """
    Abstract base class for financial statement field extractors.

    Provides common functionality for field extraction, validation, and data processing
    while allowing specialized implementations for different statement types.
    """

    def __init__(self,
                 statement_type: StatementType,
                 validate_data: bool = True,
                 min_quality_threshold: float = 0.5):
        """
        Initialize base field extractor.

        Args:
            statement_type: Type of statement this extractor handles
            validate_data: Whether to perform data validation
            min_quality_threshold: Minimum quality score to consider extraction successful
        """
        self.statement_type = statement_type
        self.validate_data = validate_data
        self.min_quality_threshold = min_quality_threshold

        # Initialize field mappings
        self.field_mappings = self._create_field_mappings()

        # Processing statistics
        self._stats = {
            'files_processed': 0,
            'fields_extracted': 0,
            'validation_failures': 0,
            'processing_time': 0.0
        }

        logger.info(f"Initialized {self.__class__.__name__} for {statement_type.value}")

    @abstractmethod
    def _create_field_mappings(self) -> FieldMappingDict:
        """Create field mappings specific to this statement type"""
        pass

    @abstractmethod
    def _get_required_fields(self) -> List[str]:
        """Get list of required fields for this statement type"""
        pass

    @abstractmethod
    def _validate_field_value(self, field_name: str, value: Any, period: str) -> Tuple[bool, str]:
        """Validate a specific field value"""
        pass

    def extract_fields(self,
                      file_path: str,
                      company_symbol: str,
                      period_type: str = "FY") -> FieldExtractionResult:
        """
        Extract fields from an Excel financial statement file.

        Args:
            file_path: Path to Excel file
            company_symbol: Company ticker symbol
            period_type: Period type ("FY" or "LTM")

        Returns:
            FieldExtractionResult with extraction details
        """
        start_time = datetime.now()

        try:
            # Load Excel file
            workbook = load_workbook(filename=file_path, data_only=True)
            worksheet = workbook.active

            # Convert to list of lists for processing
            data_rows = []
            for row in worksheet.iter_rows(values_only=True):
                data_rows.append(row)

            # Find header row and available periods
            header_info = self._find_header_row(data_rows)
            if not header_info:
                raise DataParsingException(
                    f"Could not find header row in {file_path}",
                    parsing_stage="header_detection",
                    source="excel",
                    context={'file_path': file_path}
                )

            header_row_idx, available_periods = header_info

            # Extract fields from data rows
            extracted_fields, processing_notes = self._extract_fields_from_rows(
                data_rows, header_row_idx, available_periods
            )

            # Validate extracted data
            validation_results = self._validate_extracted_data(extracted_fields)

            # Calculate quality metrics
            quality_score = self._calculate_quality_score(
                extracted_fields, validation_results, available_periods
            )

            # Create result object
            processing_time = (datetime.now() - start_time).total_seconds()

            result = FieldExtractionResult(
                statement_type=self.statement_type,
                company_symbol=company_symbol,
                period_type=period_type,
                file_path=file_path,
                extracted_fields=extracted_fields,
                missing_fields=validation_results['missing_fields'],
                invalid_fields=validation_results['invalid_fields'],
                total_fields_attempted=len(self._get_required_fields()),
                successful_extractions=len(extracted_fields),
                data_quality_score=quality_score,
                processing_time_seconds=processing_time,
                available_periods=available_periods,
                header_row_index=header_row_idx,
                notes=processing_notes
            )

            # Update statistics
            self._stats['files_processed'] += 1
            self._stats['fields_extracted'] += len(extracted_fields)
            self._stats['processing_time'] += processing_time

            logger.info(f"Successfully extracted {len(extracted_fields)} fields from {file_path}")
            return result

        except Exception as e:
            logger.error(f"Failed to extract fields from {file_path}: {str(e)}")
            raise DataParsingException(
                f"Field extraction failed: {str(e)}",
                parsing_stage="field_extraction",
                source="excel",
                context={'file_path': file_path},
                original_exception=e
            )

    def _find_header_row(self, data_rows: List[Tuple]) -> Optional[Tuple[int, List[str]]]:
        """
        Find the header row containing period columns (FY, FY-1, etc.).

        Returns:
            Tuple of (header_row_index, available_periods) or None if not found
        """
        for row_idx, row in enumerate(data_rows):
            if not row:
                continue

            # Look for FY patterns in this row
            periods_found = []
            for cell in row:
                if cell is not None:
                    cell_str = str(cell).strip()

                    # Match FY patterns
                    if cell_str == "FY":
                        periods_found.append("FY")
                    elif cell_str.startswith("FY-") and len(cell_str) > 3:
                        try:
                            years_back = int(cell_str[3:])
                            periods_found.append(cell_str)
                        except ValueError:
                            continue

            # If we found multiple periods, this is likely the header row
            if len(periods_found) >= 2:
                # Sort periods for consistent ordering (FY, FY-1, FY-2, etc.)
                periods_found.sort(key=lambda x: 0 if x == "FY" else int(x.split('-')[1]))
                return row_idx, periods_found

        return None

    def _extract_fields_from_rows(self,
                                 data_rows: List[Tuple],
                                 header_row_idx: int,
                                 available_periods: List[str]) -> Tuple[Dict[str, Dict[str, Any]], List[str]]:
        """
        Extract field values from data rows.

        Returns:
            Tuple of (extracted_fields, processing_notes)
        """
        extracted_fields = {}
        processing_notes = []

        # Create period to column index mapping
        header_row = data_rows[header_row_idx]
        period_columns = {}
        for col_idx, cell in enumerate(header_row):
            if cell and str(cell).strip() in available_periods:
                period_columns[str(cell).strip()] = col_idx

        # Process each data row
        for row_idx in range(header_row_idx + 1, len(data_rows)):
            row = data_rows[row_idx]
            if not row or len(row) == 0:
                continue

            # Extract row label (financial line item name)
            row_label = self._extract_row_label(row)
            if not row_label:
                continue

            # Try to match this row to a standard field
            standard_field_name = self.field_mappings.find_standard_field_name(row_label)
            if not standard_field_name:
                # Try fuzzy matching for partial matches
                standard_field_name = self._fuzzy_match_field(row_label)

            if standard_field_name:
                # Extract values for each period
                field_values = {}

                for period, col_idx in period_columns.items():
                    if col_idx < len(row):
                        raw_value = row[col_idx]
                        if raw_value is not None and str(raw_value).strip():
                            # Clean and convert the value
                            cleaned_value = self._clean_numeric_value(raw_value)
                            if cleaned_value is not None:
                                field_values[period] = cleaned_value

                if field_values:
                    extracted_fields[standard_field_name] = field_values
                    processing_notes.append(f"Mapped '{row_label}' to '{standard_field_name}'")
            else:
                processing_notes.append(f"No mapping found for row label: '{row_label}'")

        return extracted_fields, processing_notes

    def _extract_row_label(self, row: Tuple) -> Optional[str]:
        """Extract the financial line item name from a row"""
        # Check first few columns for the label
        for col_idx in range(min(5, len(row))):
            cell_value = row[col_idx]
            if cell_value is not None:
                cell_str = str(cell_value).strip()

                # Skip numeric values and dates
                if (cell_str and
                    not cell_str.replace('-', '').replace('.', '').replace(',', '').replace('%', '').isdigit() and
                    not any(year in cell_str for year in ['2015', '2016', '2017', '2018', '2019',
                                                        '2020', '2021', '2022', '2023', '2024', '2025'])):
                    return cell_str

        return None

    def _fuzzy_match_field(self, row_label: str) -> Optional[str]:
        """Attempt fuzzy matching for field names"""
        row_label_clean = row_label.lower().strip()
        row_words = set(re.findall(r'\w+', row_label_clean))

        best_match = None
        best_score = 0

        # Check all standard field names and their mappings
        for standard_field in self.field_mappings.primary_names.keys():
            all_mappings = self.field_mappings.get_all_mappings_for_field(standard_field)

            for mapping in all_mappings:
                mapping_clean = mapping.lower().strip()
                mapping_words = set(re.findall(r'\w+', mapping_clean))

                # Calculate word overlap score
                common_words = row_words & mapping_words
                if common_words:
                    # Score based on ratio of common words
                    score = len(common_words) / max(len(row_words), len(mapping_words))

                    # Boost score if it's a high-confidence match
                    if len(common_words) >= 2 and score >= 0.5:
                        if score > best_score:
                            best_score = score
                            best_match = standard_field

        # Only return match if confidence is high enough
        if best_score >= 0.6:
            return best_match

        return None

    def _clean_numeric_value(self, raw_value: Any) -> Optional[float]:
        """Clean and convert raw value to numeric format"""
        if raw_value is None:
            return None

        try:
            # Handle different value types
            if isinstance(raw_value, (int, float)):
                return float(raw_value)

            if isinstance(raw_value, str):
                # Clean string value
                cleaned = raw_value.strip()

                # Handle empty or null indicators
                if cleaned.lower() in ['', '-', 'n/a', 'na', 'none', 'null']:
                    return None

                # Remove formatting characters
                cleaned = re.sub(r'[\$,\(\)]', '', cleaned)

                # Handle negative values in parentheses
                if raw_value.strip().startswith('(') and raw_value.strip().endswith(')'):
                    cleaned = '-' + cleaned

                # Try to convert to float
                return float(cleaned)

            # For other types, try direct conversion
            return float(raw_value)

        except (ValueError, TypeError):
            return None

    def _validate_extracted_data(self, extracted_fields: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
        """Validate extracted field data"""
        validation_results = {
            'missing_fields': [],
            'invalid_fields': {},
            'warnings': []
        }

        # Check for required fields
        required_fields = self._get_required_fields()
        for field_name in required_fields:
            if field_name not in extracted_fields:
                validation_results['missing_fields'].append(field_name)

        # Validate individual field values
        if self.validate_data:
            for field_name, period_values in extracted_fields.items():
                for period, value in period_values.items():
                    is_valid, error_msg = self._validate_field_value(field_name, value, period)
                    if not is_valid:
                        validation_results['invalid_fields'][f"{field_name}[{period}]"] = error_msg
                        self._stats['validation_failures'] += 1

        return validation_results

    def _calculate_quality_score(self,
                                extracted_fields: Dict[str, Dict[str, Any]],
                                validation_results: Dict[str, Any],
                                available_periods: List[str]) -> float:
        """Calculate overall data quality score"""
        if not extracted_fields:
            return 0.0

        # Completeness score (required fields present)
        required_fields = self._get_required_fields()
        missing_count = len(validation_results['missing_fields'])
        completeness_score = 1.0 - (missing_count / max(1, len(required_fields)))

        # Validation score (extracted data is valid)
        total_values = sum(len(periods) for periods in extracted_fields.values())
        invalid_count = len(validation_results['invalid_fields'])
        validation_score = 1.0 - (invalid_count / max(1, total_values))

        # Coverage score (data available across periods)
        total_expected_values = len(extracted_fields) * len(available_periods)
        actual_values = sum(len(periods) for periods in extracted_fields.values())
        coverage_score = actual_values / max(1, total_expected_values)

        # Weighted average
        final_score = (
            completeness_score * 0.4 +
            validation_score * 0.4 +
            coverage_score * 0.2
        )

        return round(final_score, 3)

    def get_statistics(self) -> Dict[str, Any]:
        """Get processing statistics for this extractor"""
        return {
            'extractor_type': self.__class__.__name__,
            'statement_type': self.statement_type.value,
            'stats': dict(self._stats),
            'avg_processing_time': (
                self._stats['processing_time'] / max(1, self._stats['files_processed'])
            )
        }


# This is the base implementation - specialized extractors will be implemented next
# in subsequent commits as separate classes that inherit from BaseFieldExtractor
class IncomeStatementExtractor(BaseFieldExtractor):
    """
    Income Statement field extractor with comprehensive field mapping.

    Extracts all revenue, expense, and profit fields from Income Statement Excel files
    including revenue streams, cost components, operating expenses, and profit measures.
    """

    def __init__(self, **kwargs):
        super().__init__(StatementType.INCOME, **kwargs)

    def _create_field_mappings(self) -> FieldMappingDict:
        """Create comprehensive field mappings for Income Statement items"""

        # Primary field names (standardized)
        primary_names = {
            "revenue": "Total Revenue",
            "total_revenue": "Total Revenue",
            "gross_revenue": "Gross Revenue",
            "net_sales": "Net Sales",
            "cost_of_revenue": "Cost of Revenue",
            "cost_of_goods_sold": "Cost of Goods Sold",
            "gross_profit": "Gross Profit",
            "operating_expenses": "Operating Expenses",
            "rd_expenses": "Research and Development",
            "sga_expenses": "Selling, General & Administrative",
            "marketing_expenses": "Marketing Expenses",
            "depreciation_amortization": "Depreciation and Amortization",
            "operating_income": "Operating Income",
            "ebit": "EBIT",
            "ebitda": "EBITDA",
            "interest_expense": "Interest Expense",
            "interest_income": "Interest Income",
            "other_income": "Other Income",
            "pretax_income": "Income Before Tax",
            "tax_expense": "Tax Expense",
            "net_income": "Net Income",
            "net_income_continuing": "Net Income from Continuing Operations",
            "net_income_discontinued": "Net Income from Discontinued Operations",
            "earnings_per_share_basic": "Basic Earnings Per Share",
            "earnings_per_share_diluted": "Diluted Earnings Per Share",
            "shares_outstanding_basic": "Basic Shares Outstanding",
            "shares_outstanding_diluted": "Diluted Shares Outstanding"
        }

        # Excel-specific variations (common patterns found in real Excel files)
        excel_variations = {
            "revenue": [
                "Revenue", "Total Revenue", "Net Revenue", "Sales", "Total Sales",
                "Net Sales", "Total Net Sales", "Revenues", "Total Revenues",
                "Product Revenue", "Service Revenue", "License Revenue"
            ],
            "total_revenue": [
                "Total Revenue", "Total Revenues", "Total Net Sales", "Total Sales",
                "Consolidated Revenue", "Total Operating Revenue"
            ],
            "cost_of_revenue": [
                "Cost of Revenue", "Cost of Revenues", "Cost of Sales", "Cost of Goods Sold",
                "COGS", "Direct Costs", "Cost of Products Sold", "Product Costs",
                "Cost of Services", "Total Cost of Revenue"
            ],
            "gross_profit": [
                "Gross Profit", "Gross Income", "Total Gross Profit", "Gross Margin",
                "Gross Profit (Loss)", "Gross Earnings"
            ],
            "operating_expenses": [
                "Operating Expenses", "Total Operating Expenses", "Operating Costs",
                "Operating Expenditures", "Operating Expense", "Total Operating Expense"
            ],
            "rd_expenses": [
                "Research and Development", "R&D", "R&D Expenses", "Research & Development",
                "Product Development", "Technology Development", "R & D",
                "Research and Development Expenses", "Development Expenses"
            ],
            "sga_expenses": [
                "Selling, General & Administrative", "SG&A", "SGA", "Sales & Marketing",
                "General & Administrative", "Selling and Administrative",
                "Selling, General and Administrative", "Sales, General & Administrative",
                "Marketing and Sales", "Administrative Expenses"
            ],
            "operating_income": [
                "Operating Income", "Operating Profit", "Income from Operations",
                "Operating Earnings", "Operating Income (Loss)", "EBIT",
                "Earnings Before Interest and Tax", "Operating Profit (Loss)"
            ],
            "ebit": [
                "EBIT", "Earnings Before Interest and Tax", "Operating Income",
                "Income from Operations", "Operating Profit", "Operating Earnings"
            ],
            "ebitda": [
                "EBITDA", "Earnings Before Interest, Tax, Depreciation and Amortization",
                "EBITDA (Loss)", "Adjusted EBITDA"
            ],
            "interest_expense": [
                "Interest Expense", "Interest Expenses", "Net Interest Expense",
                "Interest and Other Financial Charges", "Financial Costs",
                "Interest on Debt", "Interest Paid", "Interest and Debt Expense"
            ],
            "interest_income": [
                "Interest Income", "Interest and Investment Income", "Investment Income",
                "Interest Earned", "Interest Revenue", "Other Interest Income"
            ],
            "other_income": [
                "Other Income", "Other Revenue", "Non-Operating Income",
                "Other Income (Expense)", "Miscellaneous Income", "Other, Net",
                "Other Income, Net", "Other Operating Income"
            ],
            "pretax_income": [
                "Income Before Tax", "Pretax Income", "Earnings Before Tax",
                "Income Before Income Tax", "Pre-Tax Income", "EBT",
                "Income (Loss) Before Tax", "Pretax Earnings"
            ],
            "tax_expense": [
                "Income Tax Expense", "Tax Expense", "Provision for Income Tax",
                "Income Taxes", "Tax Provision", "Income Tax", "Taxes"
            ],
            "net_income": [
                "Net Income", "Net Earnings", "Net Profit", "Bottom Line",
                "Net Income (Loss)", "Profit", "Net Income Attributable to Common Shareholders",
                "Net Income Available to Common Shareholders", "Earnings"
            ],
            "earnings_per_share_basic": [
                "Basic Earnings Per Share", "Basic EPS", "Earnings Per Share - Basic",
                "EPS Basic", "Basic Earnings per Share", "Earnings per Common Share - Basic"
            ],
            "earnings_per_share_diluted": [
                "Diluted Earnings Per Share", "Diluted EPS", "Earnings Per Share - Diluted",
                "EPS Diluted", "Diluted Earnings per Share", "Earnings per Common Share - Diluted"
            ]
        }

        # Common abbreviations and alternative forms
        abbreviations = {
            "revenue": ["Rev", "Sales", "Turnover"],
            "cost_of_revenue": ["COGS", "COR", "COS"],
            "gross_profit": ["GP", "Gross"],
            "operating_expenses": ["OpEx", "OPEX"],
            "rd_expenses": ["R&D", "RnD"],
            "sga_expenses": ["SG&A", "SGA"],
            "operating_income": ["OI", "EBIT"],
            "ebit": ["OI"],
            "ebitda": ["EBITDA"],
            "interest_expense": ["IntExp", "IE"],
            "pretax_income": ["EBT", "PBT"],
            "tax_expense": ["Tax", "Taxes"],
            "net_income": ["NI", "Earnings"],
            "earnings_per_share_basic": ["Basic EPS", "EPS"],
            "earnings_per_share_diluted": ["Diluted EPS", "DEPS"]
        }

        # International/regional variations
        international_names = {
            "revenue": ["Turnover", "Net Turnover", "Sales Turnover"],
            "cost_of_revenue": ["Cost of Sales", "Direct Costs"],
            "operating_income": ["Trading Profit", "Operating Result"],
            "pretax_income": ["Profit Before Tax", "PBT"],
            "net_income": ["Profit After Tax", "PAT", "Net Profit"]
        }

        return FieldMappingDict(
            primary_names=primary_names,
            excel_variations=excel_variations,
            abbreviations=abbreviations,
            international_names=international_names
        )

    def _get_required_fields(self) -> List[str]:
        """Get list of required fields for Income Statement"""
        return [
            "revenue",           # Essential for all calculations
            "cost_of_revenue",   # Needed for gross profit
            "gross_profit",      # Key profitability metric
            "operating_income",  # Core operating performance
            "net_income"         # Bottom line metric
        ]

    def _validate_field_value(self, field_name: str, value: Any, period: str) -> Tuple[bool, str]:
        """Validate Income Statement field values with business logic"""
        try:
            if value is None:
                return False, "Value cannot be None"

            # Convert to float for validation
            numeric_value = float(value)

            # General validations
            if not (-1_000_000 <= numeric_value <= 1_000_000):  # Reasonable range in millions
                return False, f"Value {numeric_value} is outside reasonable range"

            # Field-specific validations
            if field_name in ["revenue", "total_revenue", "gross_revenue"]:
                if numeric_value < 0:
                    return False, "Revenue cannot be negative"
                if numeric_value > 500_000:  # Very large revenue (500B)
                    return False, f"Revenue {numeric_value}M seems unrealistically high"

            elif field_name in ["cost_of_revenue", "cost_of_goods_sold"]:
                if numeric_value < 0:
                    return False, "Cost of revenue cannot be negative"

            elif field_name == "gross_profit":
                # Gross profit can be negative (company losing money on operations)
                pass  # No specific validation beyond range

            elif field_name in ["operating_expenses", "rd_expenses", "sga_expenses"]:
                if numeric_value < 0:
                    return False, f"{field_name} cannot be negative"

            elif field_name in ["operating_income", "ebit"]:
                # Operating income can be negative (operating losses)
                pass  # No specific validation beyond range

            elif field_name == "interest_expense":
                if numeric_value < 0:
                    return False, "Interest expense cannot be negative"

            elif field_name == "tax_expense":
                # Tax expense can be negative (tax benefits)
                pass  # No specific validation

            elif field_name == "net_income":
                # Net income can be negative (losses)
                pass  # No specific validation beyond range

            elif field_name in ["earnings_per_share_basic", "earnings_per_share_diluted"]:
                if not (-1_000 <= numeric_value <= 1_000):  # EPS range
                    return False, f"EPS {numeric_value} is outside reasonable range"

            elif field_name in ["shares_outstanding_basic", "shares_outstanding_diluted"]:
                if numeric_value <= 0:
                    return False, "Shares outstanding must be positive"
                if numeric_value > 100_000:  # 100B shares seems excessive
                    return False, f"Shares outstanding {numeric_value}M seems unrealistically high"

            return True, ""

        except (ValueError, TypeError) as e:
            return False, f"Value validation failed: {str(e)}"

    def validate_income_statement_relationships(self, extracted_fields: Dict[str, Dict[str, Any]]) -> List[str]:
        """
        Validate logical relationships between Income Statement fields.

        Returns:
            List of validation warnings/errors
        """
        warnings = []

        # Check each period for relationship validation
        all_periods = set()
        for field_data in extracted_fields.values():
            all_periods.update(field_data.keys())

        for period in all_periods:
            period_data = {}
            for field_name, field_values in extracted_fields.items():
                if period in field_values:
                    period_data[field_name] = field_values[period]

            # Validate: Gross Profit = Revenue - Cost of Revenue
            if all(key in period_data for key in ["revenue", "cost_of_revenue", "gross_profit"]):
                calculated_gp = period_data["revenue"] - period_data["cost_of_revenue"]
                actual_gp = period_data["gross_profit"]

                if abs(calculated_gp - actual_gp) > max(abs(calculated_gp) * 0.05, 1.0):  # 5% tolerance
                    warnings.append(
                        f"[{period}] Gross Profit mismatch: "
                        f"Expected {calculated_gp:.1f}M (Rev-COGS), got {actual_gp:.1f}M"
                    )

            # Validate: Cost of Revenue should typically be less than Revenue
            if "revenue" in period_data and "cost_of_revenue" in period_data:
                if period_data["cost_of_revenue"] > period_data["revenue"]:
                    warnings.append(
                        f"[{period}] Cost of Revenue ({period_data['cost_of_revenue']:.1f}M) "
                        f"exceeds Revenue ({period_data['revenue']:.1f}M)"
                    )

            # Validate: Operating Income relationship to Gross Profit
            if "gross_profit" in period_data and "operating_income" in period_data:
                if period_data["operating_income"] > period_data["gross_profit"]:
                    warnings.append(
                        f"[{period}] Operating Income ({period_data['operating_income']:.1f}M) "
                        f"exceeds Gross Profit ({period_data['gross_profit']:.1f}M) - unusual"
                    )

            # Validate EPS consistency
            if all(key in period_data for key in ["net_income", "shares_outstanding_basic", "earnings_per_share_basic"]):
                calculated_eps = period_data["net_income"] / period_data["shares_outstanding_basic"] * 1_000_000  # Convert from millions
                actual_eps = period_data["earnings_per_share_basic"]

                if abs(calculated_eps - actual_eps) > max(abs(calculated_eps) * 0.05, 0.01):  # 5% tolerance
                    warnings.append(
                        f"[{period}] Basic EPS mismatch: "
                        f"Expected {calculated_eps:.2f}, got {actual_eps:.2f}"
                    )

        return warnings


class BalanceSheetExtractor(BaseFieldExtractor):
    """
    Balance Sheet field extractor with comprehensive field mapping and validation.

    Extracts all asset, liability, and equity fields from Balance Sheet Excel files
    with automatic classification and balance sheet equation validation.
    """

    def __init__(self, **kwargs):
        super().__init__(StatementType.BALANCE, **kwargs)

    def _create_field_mappings(self) -> FieldMappingDict:
        """Create comprehensive field mappings for Balance Sheet items"""

        # Primary field names (standardized)
        primary_names = {
            # Assets
            "total_assets": "Total Assets",
            "current_assets": "Current Assets",
            "cash_and_equivalents": "Cash and Cash Equivalents",
            "cash": "Cash",
            "short_term_investments": "Short-term Investments",
            "accounts_receivable": "Accounts Receivable",
            "net_receivables": "Net Receivables",
            "inventory": "Inventory",
            "prepaid_expenses": "Prepaid Expenses",
            "other_current_assets": "Other Current Assets",
            "non_current_assets": "Non-current Assets",
            "property_plant_equipment": "Property, Plant & Equipment",
            "ppe_net": "PP&E, Net",
            "goodwill": "Goodwill",
            "intangible_assets": "Intangible Assets",
            "long_term_investments": "Long-term Investments",
            "other_assets": "Other Assets",

            # Liabilities
            "total_liabilities": "Total Liabilities",
            "current_liabilities": "Current Liabilities",
            "accounts_payable": "Accounts Payable",
            "short_term_debt": "Short-term Debt",
            "accrued_liabilities": "Accrued Liabilities",
            "deferred_revenue": "Deferred Revenue",
            "other_current_liabilities": "Other Current Liabilities",
            "non_current_liabilities": "Non-current Liabilities",
            "long_term_debt": "Long-term Debt",
            "deferred_tax_liabilities": "Deferred Tax Liabilities",
            "other_liabilities": "Other Liabilities",

            # Equity
            "total_equity": "Total Shareholders' Equity",
            "shareholders_equity": "Shareholders' Equity",
            "common_stock": "Common Stock",
            "preferred_stock": "Preferred Stock",
            "additional_paid_in_capital": "Additional Paid-in Capital",
            "retained_earnings": "Retained Earnings",
            "accumulated_other_comprehensive_income": "Accumulated Other Comprehensive Income",
            "treasury_stock": "Treasury Stock",

            # Working Capital Components
            "working_capital": "Working Capital",
            "net_working_capital": "Net Working Capital"
        }

        # Excel-specific variations
        excel_variations = {
            "total_assets": [
                "Total Assets", "Total Asset", "Assets, Total", "Sum of Assets",
                "Total Company Assets", "Consolidated Total Assets"
            ],
            "current_assets": [
                "Current Assets", "Total Current Assets", "Current Asset Total",
                "Assets, Current", "Short-term Assets", "Liquid Assets"
            ],
            "cash_and_equivalents": [
                "Cash and Cash Equivalents", "Cash & Cash Equivalents", "Cash and Equivalents",
                "Cash & Equivalents", "Cash, Cash Equivalents", "Total Cash"
            ],
            "cash": [
                "Cash", "Cash on Hand", "Cash Balance", "Available Cash",
                "Unrestricted Cash", "Cash (Unrestricted)"
            ],
            "short_term_investments": [
                "Short-term Investments", "Short Term Investments", "Marketable Securities",
                "Short-term Securities", "Temporary Investments", "Liquid Investments"
            ],
            "accounts_receivable": [
                "Accounts Receivable", "Trade Receivables", "Net Receivables",
                "Receivables, Net", "A/R", "Trade Debtors", "Customer Receivables"
            ],
            "inventory": [
                "Inventory", "Inventories", "Stock", "Finished Goods",
                "Raw Materials", "Work in Process", "Merchandise Inventory"
            ],
            "prepaid_expenses": [
                "Prepaid Expenses", "Prepaid Assets", "Prepaids", "Prepaid Costs",
                "Prepaid and Other Assets", "Deferred Charges"
            ],
            "property_plant_equipment": [
                "Property, Plant & Equipment", "Property, Plant and Equipment", "PP&E",
                "Fixed Assets", "Plant and Equipment", "Property and Equipment",
                "Plant, Property & Equipment", "Tangible Assets"
            ],
            "goodwill": [
                "Goodwill", "Goodwill and Intangible Assets", "Purchased Goodwill"
            ],
            "intangible_assets": [
                "Intangible Assets", "Intangibles", "Other Intangible Assets",
                "Patents and Trademarks", "Software", "Intellectual Property"
            ],
            "long_term_investments": [
                "Long-term Investments", "Long Term Investments", "Investment Securities",
                "Non-current Investments", "Strategic Investments"
            ],
            "other_assets": [
                "Other Assets", "Miscellaneous Assets", "Other Non-current Assets",
                "Other Long-term Assets", "Deferred Tax Assets"
            ],

            # Liabilities
            "total_liabilities": [
                "Total Liabilities", "Total Liability", "Liabilities, Total",
                "Sum of Liabilities", "Total Company Liabilities"
            ],
            "current_liabilities": [
                "Current Liabilities", "Total Current Liabilities", "Current Liability Total",
                "Liabilities, Current", "Short-term Liabilities"
            ],
            "accounts_payable": [
                "Accounts Payable", "Trade Payables", "A/P", "Trade Creditors",
                "Vendor Payables", "Supplier Payables"
            ],
            "short_term_debt": [
                "Short-term Debt", "Short Term Debt", "Current Portion of Long-term Debt",
                "Notes Payable", "Short-term Borrowings", "Current Debt"
            ],
            "accrued_liabilities": [
                "Accrued Liabilities", "Accrued Expenses", "Accruals",
                "Accrued Compensation", "Other Accrued Liabilities"
            ],
            "deferred_revenue": [
                "Deferred Revenue", "Unearned Revenue", "Advance Payments",
                "Customer Deposits", "Contract Liabilities"
            ],
            "long_term_debt": [
                "Long-term Debt", "Long Term Debt", "Non-current Debt",
                "Senior Notes", "Term Loans", "Bonds Payable"
            ],
            "deferred_tax_liabilities": [
                "Deferred Tax Liabilities", "Deferred Income Tax", "Tax Liabilities"
            ],

            # Equity
            "total_equity": [
                "Total Shareholders' Equity", "Total Stockholders' Equity", "Total Equity",
                "Shareholders' Equity", "Stockholders' Equity", "Total Shareholder Equity"
            ],
            "shareholders_equity": [
                "Shareholders' Equity", "Stockholders' Equity", "Shareholder Equity",
                "Stockholder Equity", "Owners' Equity", "Net Worth"
            ],
            "common_stock": [
                "Common Stock", "Common Shares", "Ordinary Shares", "Common Equity",
                "Common Stock Outstanding", "Voting Shares"
            ],
            "preferred_stock": [
                "Preferred Stock", "Preferred Shares", "Preference Shares"
            ],
            "additional_paid_in_capital": [
                "Additional Paid-in Capital", "Paid-in Capital", "Share Premium",
                "Capital in Excess of Par", "Additional Capital", "Contributed Capital"
            ],
            "retained_earnings": [
                "Retained Earnings", "Accumulated Earnings", "Undistributed Earnings",
                "Earnings Retained", "Accumulated Profits"
            ],
            "accumulated_other_comprehensive_income": [
                "Accumulated Other Comprehensive Income", "AOCI", "Other Comprehensive Income",
                "Comprehensive Income", "Currency Translation"
            ],
            "treasury_stock": [
                "Treasury Stock", "Treasury Shares", "Own Shares", "Repurchased Stock"
            ]
        }

        # Common abbreviations
        abbreviations = {
            "total_assets": ["TA"],
            "current_assets": ["CA"],
            "cash_and_equivalents": ["CCE", "Cash"],
            "accounts_receivable": ["AR", "A/R"],
            "property_plant_equipment": ["PPE", "PP&E"],
            "total_liabilities": ["TL"],
            "current_liabilities": ["CL"],
            "accounts_payable": ["AP", "A/P"],
            "short_term_debt": ["STD"],
            "long_term_debt": ["LTD"],
            "total_equity": ["TE", "SE"],
            "shareholders_equity": ["SE", "SHE"],
            "retained_earnings": ["RE"],
            "additional_paid_in_capital": ["APIC"]
        }

        # International variations
        international_names = {
            "total_assets": ["Total Assets", "Sum of Assets"],
            "current_assets": ["Current Assets", "Circulating Assets"],
            "property_plant_equipment": ["Fixed Assets", "Tangible Assets"],
            "accounts_receivable": ["Trade Debtors", "Debtors"],
            "inventory": ["Stock", "Stocks"],
            "accounts_payable": ["Trade Creditors", "Creditors"],
            "shareholders_equity": ["Share Capital and Reserves", "Net Assets"]
        }

        return FieldMappingDict(
            primary_names=primary_names,
            excel_variations=excel_variations,
            abbreviations=abbreviations,
            international_names=international_names
        )

    def _get_required_fields(self) -> List[str]:
        """Get list of required fields for Balance Sheet"""
        return [
            "total_assets",           # Must balance with liabilities + equity
            "current_assets",         # Key liquidity metric
            "cash_and_equivalents",   # Most liquid asset
            "total_liabilities",      # Total obligations
            "current_liabilities",    # Short-term obligations
            "shareholders_equity"     # Ownership value
        ]

    def _validate_field_value(self, field_name: str, value: Any, period: str) -> Tuple[bool, str]:
        """Validate Balance Sheet field values with business logic"""
        try:
            if value is None:
                return False, "Value cannot be None"

            # Convert to float for validation
            numeric_value = float(value)

            # General range validation (in millions)
            if not (-500_000 <= numeric_value <= 2_000_000):
                return False, f"Value {numeric_value} is outside reasonable range"

            # Asset validations (should be non-negative)
            asset_fields = [
                "total_assets", "current_assets", "cash_and_equivalents", "cash",
                "short_term_investments", "accounts_receivable", "inventory",
                "prepaid_expenses", "other_current_assets", "non_current_assets",
                "property_plant_equipment", "ppe_net", "goodwill", "intangible_assets",
                "long_term_investments", "other_assets"
            ]

            if field_name in asset_fields:
                if numeric_value < 0:
                    return False, f"{field_name} cannot be negative"

            # Liability validations (should be non-negative)
            liability_fields = [
                "total_liabilities", "current_liabilities", "accounts_payable",
                "short_term_debt", "accrued_liabilities", "deferred_revenue",
                "other_current_liabilities", "non_current_liabilities",
                "long_term_debt", "deferred_tax_liabilities", "other_liabilities"
            ]

            if field_name in liability_fields:
                if numeric_value < 0:
                    return False, f"{field_name} cannot be negative"

            # Equity validations (can be negative in distressed companies)
            equity_fields = [
                "total_equity", "shareholders_equity", "retained_earnings"
            ]

            if field_name in equity_fields:
                # Allow negative equity but warn if extremely negative
                if numeric_value < -100_000:  # Very negative equity
                    return False, f"{field_name} is extremely negative: {numeric_value}M"

            # Stock-related fields (should be non-negative)
            stock_fields = ["common_stock", "preferred_stock", "additional_paid_in_capital"]

            if field_name in stock_fields:
                if numeric_value < 0:
                    return False, f"{field_name} cannot be negative"

            # Treasury stock (should be non-positive as it reduces equity)
            if field_name == "treasury_stock":
                if numeric_value > 0:
                    return False, "Treasury stock should be negative or zero (reduces equity)"

            # Working capital can be negative
            if field_name in ["working_capital", "net_working_capital"]:
                pass  # No specific validation

            return True, ""

        except (ValueError, TypeError) as e:
            return False, f"Value validation failed: {str(e)}"

    def validate_balance_sheet_equation(self, extracted_fields: Dict[str, Dict[str, Any]]) -> List[str]:
        """
        Validate the fundamental balance sheet equation: Assets = Liabilities + Equity

        Returns:
            List of validation warnings/errors
        """
        warnings = []

        # Check each period for balance sheet equation
        all_periods = set()
        for field_data in extracted_fields.values():
            all_periods.update(field_data.keys())

        for period in all_periods:
            period_data = {}
            for field_name, field_values in extracted_fields.items():
                if period in field_values:
                    period_data[field_name] = field_values[period]

            # Validate: Total Assets = Total Liabilities + Total Equity
            if all(key in period_data for key in ["total_assets", "total_liabilities", "shareholders_equity"]):
                assets = period_data["total_assets"]
                liabilities = period_data["total_liabilities"]
                equity = period_data["shareholders_equity"]

                calculated_total = liabilities + equity
                difference = abs(assets - calculated_total)
                tolerance = max(abs(assets) * 0.01, 1.0)  # 1% tolerance or 1M minimum

                if difference > tolerance:
                    warnings.append(
                        f"[{period}] Balance Sheet equation imbalance: "
                        f"Assets({assets:.1f}M) ≠ Liabilities({liabilities:.1f}M) + Equity({equity:.1f}M) "
                        f"Difference: {difference:.1f}M"
                    )

            # Validate: Current Assets <= Total Assets
            if "current_assets" in period_data and "total_assets" in period_data:
                if period_data["current_assets"] > period_data["total_assets"]:
                    warnings.append(
                        f"[{period}] Current Assets ({period_data['current_assets']:.1f}M) "
                        f"exceeds Total Assets ({period_data['total_assets']:.1f}M)"
                    )

            # Validate: Current Liabilities <= Total Liabilities
            if "current_liabilities" in period_data and "total_liabilities" in period_data:
                if period_data["current_liabilities"] > period_data["total_liabilities"]:
                    warnings.append(
                        f"[{period}] Current Liabilities ({period_data['current_liabilities']:.1f}M) "
                        f"exceeds Total Liabilities ({period_data['total_liabilities']:.1f}M)"
                    )

            # Validate: Cash <= Current Assets
            if "cash_and_equivalents" in period_data and "current_assets" in period_data:
                if period_data["cash_and_equivalents"] > period_data["current_assets"]:
                    warnings.append(
                        f"[{period}] Cash & Equivalents ({period_data['cash_and_equivalents']:.1f}M) "
                        f"exceeds Current Assets ({period_data['current_assets']:.1f}M)"
                    )

            # Validate: Working Capital = Current Assets - Current Liabilities
            if all(key in period_data for key in ["current_assets", "current_liabilities"]):
                calculated_wc = period_data["current_assets"] - period_data["current_liabilities"]

                if "working_capital" in period_data:
                    actual_wc = period_data["working_capital"]
                    if abs(calculated_wc - actual_wc) > max(abs(calculated_wc) * 0.05, 1.0):
                        warnings.append(
                            f"[{period}] Working Capital mismatch: "
                            f"Expected {calculated_wc:.1f}M (CA-CL), got {actual_wc:.1f}M"
                        )

            # Validate reasonable debt levels
            if all(key in period_data for key in ["total_liabilities", "total_assets"]):
                debt_ratio = period_data["total_liabilities"] / period_data["total_assets"]
                if debt_ratio > 0.95:  # Very high leverage
                    warnings.append(
                        f"[{period}] Very high debt ratio: {debt_ratio:.1%} "
                        f"(Liabilities/Assets)"
                    )

        return warnings


class CashFlowStatementExtractor(BaseFieldExtractor):
    """
    Cash Flow Statement field extractor with comprehensive activity classification.

    Extracts operating, investing, and financing cash flow components with
    automatic method detection (direct vs indirect) and free cash flow validation.
    """

    def __init__(self, **kwargs):
        super().__init__(StatementType.CASH_FLOW, **kwargs)

    def _create_field_mappings(self) -> FieldMappingDict:
        """Create comprehensive field mappings for Cash Flow Statement items"""

        # Primary field names (standardized)
        primary_names = {
            # Operating Cash Flow (Indirect Method)
            "operating_cash_flow": "Net Cash from Operating Activities",
            "net_income": "Net Income",
            "depreciation_amortization": "Depreciation and Amortization",
            "stock_based_compensation": "Stock-based Compensation",
            "deferred_taxes": "Deferred Taxes",
            "changes_in_working_capital": "Changes in Working Capital",
            "accounts_receivable_change": "Change in Accounts Receivable",
            "inventory_change": "Change in Inventory",
            "accounts_payable_change": "Change in Accounts Payable",
            "accrued_liabilities_change": "Change in Accrued Liabilities",
            "other_operating_activities": "Other Operating Activities",

            # Investing Cash Flow
            "investing_cash_flow": "Net Cash from Investing Activities",
            "capital_expenditures": "Capital Expenditures",
            "capex": "Capital Expenditures",
            "acquisitions": "Acquisitions",
            "divestitures": "Divestitures",
            "purchases_of_investments": "Purchases of Investments",
            "sales_of_investments": "Sales of Investments",
            "purchases_of_securities": "Purchase of Securities",
            "sales_of_securities": "Sale of Securities",
            "other_investing_activities": "Other Investing Activities",

            # Financing Cash Flow
            "financing_cash_flow": "Net Cash from Financing Activities",
            "dividends_paid": "Dividends Paid",
            "share_repurchases": "Share Repurchases",
            "treasury_stock_purchases": "Treasury Stock Purchases",
            "debt_issuance": "Debt Issuance",
            "debt_repayment": "Debt Repayment",
            "net_borrowing": "Net Borrowing",
            "proceeds_from_debt": "Proceeds from Debt",
            "repayment_of_debt": "Repayment of Debt",
            "proceeds_from_stock_issuance": "Proceeds from Stock Issuance",
            "other_financing_activities": "Other Financing Activities",

            # Summary Items
            "net_change_in_cash": "Net Change in Cash",
            "cash_beginning_of_period": "Cash at Beginning of Period",
            "cash_end_of_period": "Cash at End of Period",
            "free_cash_flow": "Free Cash Flow",
            "supplemental_cash_payments": "Supplemental Cash Flow Information"
        }

        # Excel-specific variations
        excel_variations = {
            "operating_cash_flow": [
                "Net Cash Provided by Operating Activities", "Operating Cash Flow",
                "Cash Flow from Operating Activities", "Net Cash from Operations",
                "Cash Generated from Operations", "Operating Activities"
            ],
            "net_income": [
                "Net Income", "Net Earnings", "Profit", "Net Income (Loss)",
                "Consolidated Net Income", "Net Income Available to Common Shareholders"
            ],
            "depreciation_amortization": [
                "Depreciation and Amortization", "D&A", "Depreciation & Amortization",
                "Depreciation", "Amortization", "Non-cash Charges"
            ],
            "stock_based_compensation": [
                "Stock-based Compensation", "Share-based Compensation", "Employee Stock Options",
                "Stock Option Expense", "Equity Compensation", "Share-based Payment"
            ],
            "deferred_taxes": [
                "Deferred Taxes", "Deferred Tax", "Deferred Income Tax",
                "Change in Deferred Tax", "Deferred Tax Benefit"
            ],
            "changes_in_working_capital": [
                "Changes in Working Capital", "Working Capital Changes", "Change in Working Capital",
                "Net Change in Working Capital", "Working Capital", "Changes in Assets and Liabilities"
            ],
            "accounts_receivable_change": [
                "Change in Accounts Receivable", "Accounts Receivable", "A/R Change",
                "Increase in Receivables", "Decrease in Receivables"
            ],
            "inventory_change": [
                "Change in Inventory", "Inventory", "Increase in Inventory",
                "Decrease in Inventory", "Change in Inventories"
            ],
            "accounts_payable_change": [
                "Change in Accounts Payable", "Accounts Payable", "A/P Change",
                "Increase in Payables", "Decrease in Payables"
            ],

            # Investing Activities
            "investing_cash_flow": [
                "Net Cash Used in Investing Activities", "Investing Cash Flow",
                "Cash Flow from Investing Activities", "Net Cash from Investing",
                "Cash Used for Investing", "Investing Activities"
            ],
            "capital_expenditures": [
                "Capital Expenditures", "CapEx", "Capital Spending", "PP&E Purchases",
                "Property, Plant & Equipment", "Purchase of PP&E", "Investment in PP&E"
            ],
            "acquisitions": [
                "Acquisitions", "Business Acquisitions", "Purchase of Businesses",
                "Acquisition of Companies", "Mergers and Acquisitions"
            ],
            "divestitures": [
                "Divestitures", "Sale of Businesses", "Disposal of Businesses",
                "Sale of Subsidiaries", "Asset Sales"
            ],
            "purchases_of_investments": [
                "Purchase of Investments", "Investment Purchases", "Purchase of Securities",
                "Acquisition of Investments", "Investment in Securities"
            ],
            "sales_of_investments": [
                "Sale of Investments", "Investment Sales", "Sale of Securities",
                "Disposal of Investments", "Proceeds from Investment Sales"
            ],

            # Financing Activities
            "financing_cash_flow": [
                "Net Cash Used in Financing Activities", "Financing Cash Flow",
                "Cash Flow from Financing Activities", "Net Cash from Financing",
                "Cash Used for Financing", "Financing Activities"
            ],
            "dividends_paid": [
                "Dividends Paid", "Cash Dividends", "Dividend Payments",
                "Payments of Dividends", "Common Dividends Paid"
            ],
            "share_repurchases": [
                "Share Repurchases", "Stock Repurchases", "Repurchase of Common Stock",
                "Treasury Stock Purchases", "Stock Buybacks", "Share Buybacks"
            ],
            "debt_issuance": [
                "Proceeds from Debt", "Debt Issuance", "Borrowings", "New Debt",
                "Issuance of Debt", "Long-term Debt Proceeds"
            ],
            "debt_repayment": [
                "Repayment of Debt", "Debt Payments", "Principal Payments",
                "Debt Reduction", "Long-term Debt Repayment"
            ],
            "proceeds_from_stock_issuance": [
                "Proceeds from Stock Issuance", "Common Stock Issuance",
                "Equity Issuance", "Sale of Common Stock", "Stock Sales"
            ],

            # Summary
            "net_change_in_cash": [
                "Net Change in Cash", "Net Increase in Cash", "Net Decrease in Cash",
                "Change in Cash and Equivalents", "Net Change in Cash and Cash Equivalents"
            ],
            "cash_beginning_of_period": [
                "Cash at Beginning of Period", "Beginning Cash", "Cash, Beginning",
                "Cash and Equivalents, Beginning", "Opening Cash Balance"
            ],
            "cash_end_of_period": [
                "Cash at End of Period", "Ending Cash", "Cash, Ending",
                "Cash and Equivalents, Ending", "Closing Cash Balance"
            ],
            "free_cash_flow": [
                "Free Cash Flow", "FCF", "Unlevered Free Cash Flow",
                "Operating Cash Flow less CapEx", "Net Free Cash Flow"
            ]
        }

        # Common abbreviations
        abbreviations = {
            "operating_cash_flow": ["OCF", "CFO"],
            "investing_cash_flow": ["ICF", "CFI"],
            "financing_cash_flow": ["FCF", "CFF"],
            "capital_expenditures": ["CapEx", "PPE"],
            "depreciation_amortization": ["D&A", "DA"],
            "free_cash_flow": ["FCF"],
            "net_change_in_cash": ["NCCash"],
            "accounts_receivable_change": ["AR"],
            "accounts_payable_change": ["AP"],
            "stock_based_compensation": ["SBC"]
        }

        # International variations
        international_names = {
            "operating_cash_flow": ["Cash Generated from Operations", "Operating Cash Generation"],
            "investing_cash_flow": ["Cash Used in Investment", "Investment Cash Flow"],
            "financing_cash_flow": ["Cash from Financing", "Financing Cash Generation"],
            "capital_expenditures": ["Capital Investment", "Fixed Asset Investment"],
            "dividends_paid": ["Dividend Distribution", "Shareholder Distributions"]
        }

        return FieldMappingDict(
            primary_names=primary_names,
            excel_variations=excel_variations,
            abbreviations=abbreviations,
            international_names=international_names
        )

    def _get_required_fields(self) -> List[str]:
        """Get list of required fields for Cash Flow Statement"""
        return [
            "operating_cash_flow",      # Core operating performance
            "investing_cash_flow",      # Investment activities
            "financing_cash_flow",      # Financing activities
            "capital_expenditures",     # Key investing activity
            "net_change_in_cash"        # Balance verification
        ]

    def _validate_field_value(self, field_name: str, value: Any, period: str) -> Tuple[bool, str]:
        """Validate Cash Flow Statement field values with business logic"""
        try:
            if value is None:
                return False, "Value cannot be None"

            # Convert to float for validation
            numeric_value = float(value)

            # General range validation (in millions)
            if not (-500_000 <= numeric_value <= 500_000):
                return False, f"Value {numeric_value} is outside reasonable range"

            # Operating cash flow (can be negative but unusual for healthy companies)
            if field_name == "operating_cash_flow":
                if numeric_value < -50_000:
                    return False, f"Operating cash flow {numeric_value}M is extremely negative"

            # Net income validation (already handled in Income Statement)
            if field_name == "net_income":
                pass  # Basic range validation is sufficient

            # Non-cash charges (should be positive additions to cash flow)
            non_cash_fields = ["depreciation_amortization", "stock_based_compensation"]
            if field_name in non_cash_fields:
                if numeric_value < 0:
                    return False, f"{field_name} should be positive (non-cash charge)"

            # Working capital changes (can be positive or negative)
            working_capital_fields = [
                "changes_in_working_capital", "accounts_receivable_change",
                "inventory_change", "accounts_payable_change", "accrued_liabilities_change"
            ]
            if field_name in working_capital_fields:
                pass  # No specific validation - can be positive or negative

            # Capital expenditures (should be negative as cash outflow)
            if field_name in ["capital_expenditures", "capex"]:
                if numeric_value > 0:
                    return False, "Capital expenditures should be negative (cash outflow)"
                if numeric_value < -100_000:
                    return False, f"CapEx {numeric_value}M seems unrealistically high"

            # Acquisitions and divestitures
            if field_name == "acquisitions":
                if numeric_value > 0:
                    return False, "Acquisitions should be negative (cash outflow)"
            elif field_name == "divestitures":
                if numeric_value < 0:
                    return False, "Divestitures should be positive (cash inflow)"

            # Investment activities (purchases negative, sales positive)
            purchase_fields = ["purchases_of_investments", "purchases_of_securities"]
            sale_fields = ["sales_of_investments", "sales_of_securities"]

            if field_name in purchase_fields:
                if numeric_value > 0:
                    return False, f"{field_name} should be negative (cash outflow)"
            elif field_name in sale_fields:
                if numeric_value < 0:
                    return False, f"{field_name} should be positive (cash inflow)"

            # Financing activities
            if field_name == "dividends_paid":
                if numeric_value > 0:
                    return False, "Dividends paid should be negative (cash outflow)"

            if field_name in ["share_repurchases", "treasury_stock_purchases"]:
                if numeric_value > 0:
                    return False, f"{field_name} should be negative (cash outflow)"

            if field_name == "debt_issuance":
                if numeric_value < 0:
                    return False, "Debt issuance should be positive (cash inflow)"

            if field_name == "debt_repayment":
                if numeric_value > 0:
                    return False, "Debt repayment should be negative (cash outflow)"

            if field_name == "proceeds_from_stock_issuance":
                if numeric_value < 0:
                    return False, "Stock issuance proceeds should be positive (cash inflow)"

            # Cash balances (should be positive)
            cash_balance_fields = ["cash_beginning_of_period", "cash_end_of_period"]
            if field_name in cash_balance_fields:
                if numeric_value < 0:
                    return False, f"{field_name} cannot be negative"

            # Net change in cash (can be positive or negative)
            if field_name == "net_change_in_cash":
                pass  # No specific validation

            # Free cash flow (can be negative)
            if field_name == "free_cash_flow":
                pass  # No specific validation

            return True, ""

        except (ValueError, TypeError) as e:
            return False, f"Value validation failed: {str(e)}"

    def validate_cash_flow_relationships(self, extracted_fields: Dict[str, Dict[str, Any]]) -> List[str]:
        """
        Validate logical relationships between Cash Flow Statement components.

        Returns:
            List of validation warnings/errors
        """
        warnings = []

        # Check each period for relationship validation
        all_periods = set()
        for field_data in extracted_fields.values():
            all_periods.update(field_data.keys())

        for period in all_periods:
            period_data = {}
            for field_name, field_values in extracted_fields.items():
                if period in field_values:
                    period_data[field_name] = field_values[period]

            # Validate: Net Change = Operating + Investing + Financing
            cash_flow_components = ["operating_cash_flow", "investing_cash_flow", "financing_cash_flow"]
            if all(comp in period_data for comp in cash_flow_components) and "net_change_in_cash" in period_data:
                calculated_change = (
                    period_data["operating_cash_flow"] +
                    period_data["investing_cash_flow"] +
                    period_data["financing_cash_flow"]
                )
                actual_change = period_data["net_change_in_cash"]

                if abs(calculated_change - actual_change) > max(abs(calculated_change) * 0.05, 1.0):
                    warnings.append(
                        f"[{period}] Cash flow components don't sum to net change: "
                        f"OCF({period_data['operating_cash_flow']:.1f}) + "
                        f"ICF({period_data['investing_cash_flow']:.1f}) + "
                        f"FCF({period_data['financing_cash_flow']:.1f}) = "
                        f"{calculated_change:.1f}M ≠ {actual_change:.1f}M"
                    )

            # Validate: Cash balance equation
            if all(key in period_data for key in ["cash_beginning_of_period", "net_change_in_cash", "cash_end_of_period"]):
                calculated_ending = period_data["cash_beginning_of_period"] + period_data["net_change_in_cash"]
                actual_ending = period_data["cash_end_of_period"]

                if abs(calculated_ending - actual_ending) > max(abs(calculated_ending) * 0.01, 1.0):
                    warnings.append(
                        f"[{period}] Cash balance equation imbalance: "
                        f"Beginning({period_data['cash_beginning_of_period']:.1f}) + "
                        f"Change({period_data['net_change_in_cash']:.1f}) = "
                        f"{calculated_ending:.1f}M ≠ Ending({actual_ending:.1f}M)"
                    )

            # Validate: Free Cash Flow = Operating Cash Flow - CapEx
            if all(key in period_data for key in ["operating_cash_flow", "capital_expenditures"]):
                calculated_fcf = period_data["operating_cash_flow"] + period_data["capital_expenditures"]  # CapEx is negative

                if "free_cash_flow" in period_data:
                    actual_fcf = period_data["free_cash_flow"]
                    if abs(calculated_fcf - actual_fcf) > max(abs(calculated_fcf) * 0.05, 1.0):
                        warnings.append(
                            f"[{period}] Free Cash Flow mismatch: "
                            f"Expected {calculated_fcf:.1f}M (OCF + CapEx), got {actual_fcf:.1f}M"
                        )

            # Validate: Operating cash flow vs Net income relationship
            if "operating_cash_flow" in period_data and "net_income" in period_data:
                ocf = period_data["operating_cash_flow"]
                ni = period_data["net_income"]

                # OCF should generally be higher than Net Income due to non-cash charges
                if ocf < ni * 0.5 and ni > 0:  # OCF significantly below net income
                    warnings.append(
                        f"[{period}] Operating Cash Flow ({ocf:.1f}M) is significantly below "
                        f"Net Income ({ni:.1f}M) - possible working capital issues"
                    )

            # Validate reasonable cash flow patterns
            if "operating_cash_flow" in period_data:
                ocf = period_data["operating_cash_flow"]
                if ocf < 0:
                    warnings.append(
                        f"[{period}] Negative Operating Cash Flow ({ocf:.1f}M) indicates "
                        f"potential operational difficulties"
                    )

            # Validate CapEx reasonableness
            if "capital_expenditures" in period_data and "operating_cash_flow" in period_data:
                capex = abs(period_data["capital_expenditures"])  # Make positive for comparison
                ocf = period_data["operating_cash_flow"]

                if ocf > 0 and capex > ocf * 2:  # CapEx more than 2x OCF
                    warnings.append(
                        f"[{period}] Very high CapEx ({capex:.1f}M) relative to "
                        f"Operating Cash Flow ({ocf:.1f}M) - heavy investment period"
                    )

        return warnings

    def detect_cash_flow_method(self, extracted_fields: Dict[str, Dict[str, Any]]) -> str:
        """
        Detect whether the cash flow statement uses direct or indirect method.

        Returns:
            "indirect", "direct", or "unknown"
        """
        # Check for indicators of indirect method (most common)
        indirect_indicators = [
            "net_income", "depreciation_amortization", "changes_in_working_capital",
            "accounts_receivable_change", "inventory_change", "accounts_payable_change"
        ]

        # Check for indicators of direct method (less common)
        direct_indicators = [
            "cash_received_from_customers", "cash_paid_to_suppliers",
            "cash_paid_for_operating_expenses", "interest_paid", "taxes_paid"
        ]

        indirect_count = sum(1 for indicator in indirect_indicators if indicator in extracted_fields)
        direct_count = sum(1 for indicator in direct_indicators if indicator in extracted_fields)

        if indirect_count >= 3:
            return "indirect"
        elif direct_count >= 2:
            return "direct"
        else:
            return "unknown"


class FinancialStatementFieldExtractor:
    """
    Factory class that coordinates all field extractors and integrates with existing data processing architecture.

    Provides comprehensive field extraction, data quality assessment, missing field reporting,
    and integration with the FinancialCalculator and FinancialVariableRegistry.
    """

    def __init__(self, validate_data: bool = True, min_quality_threshold: float = 0.6):
        """
        Initialize the factory with all extractors and integration settings.

        Args:
            validate_data: Whether to perform data validation across extractors
            min_quality_threshold: Minimum quality score for successful extraction
        """
        self.validate_data = validate_data
        self.min_quality_threshold = min_quality_threshold

        # Initialize all extractors
        self.extractors = {
            StatementType.INCOME: IncomeStatementExtractor(
                validate_data=validate_data,
                min_quality_threshold=min_quality_threshold
            ),
            StatementType.BALANCE: BalanceSheetExtractor(
                validate_data=validate_data,
                min_quality_threshold=min_quality_threshold
            ),
            StatementType.CASH_FLOW: CashFlowStatementExtractor(
                validate_data=validate_data,
                min_quality_threshold=min_quality_threshold
            )
        }

        # Track extraction sessions for comprehensive reporting
        self.extraction_sessions = []

        logger.info("Initialized FinancialStatementFieldExtractor with comprehensive validation")

    def extract_from_file(self,
                         file_path: str,
                         statement_type: StatementType,
                         company_symbol: str,
                         period_type: str = "FY") -> FieldExtractionResult:
        """Extract fields from a specific statement file"""
        if statement_type not in self.extractors:
            raise ValueError(f"No extractor available for statement type: {statement_type}")

        extractor = self.extractors[statement_type]
        result = extractor.extract_fields(file_path, company_symbol, period_type)

        # Track this extraction session
        self.extraction_sessions.append({
            'timestamp': datetime.now(),
            'company_symbol': company_symbol,
            'statement_type': statement_type,
            'result': result
        })

        return result

    def extract_from_company_folder(self,
                                   company_folder: str,
                                   company_symbol: str,
                                   period_type: str = "FY") -> Dict[StatementType, FieldExtractionResult]:
        """
        Extract fields from all financial statements in a company folder.

        Args:
            company_folder: Path to company data folder (e.g., data/companies/MSFT/FY/)
            company_symbol: Company ticker symbol
            period_type: Period type ("FY" or "LTM")

        Returns:
            Dictionary mapping statement types to extraction results
        """
        import os
        from pathlib import Path

        results = {}
        folder_path = Path(company_folder)

        if not folder_path.exists():
            raise FileNotFoundError(f"Company folder not found: {company_folder}")

        # Define expected file patterns
        file_patterns = {
            StatementType.INCOME: ["*Income*Statement*.xlsx", "*Income*.xlsx"],
            StatementType.BALANCE: ["*Balance*Sheet*.xlsx", "*Balance*.xlsx"],
            StatementType.CASH_FLOW: ["*Cash*Flow*.xlsx", "*Cash*.xlsx"]
        }

        for statement_type, patterns in file_patterns.items():
            file_found = False

            for pattern in patterns:
                matching_files = list(folder_path.glob(pattern))
                if matching_files:
                    file_path = str(matching_files[0])  # Use first matching file
                    try:
                        result = self.extract_from_file(file_path, statement_type, company_symbol, period_type)
                        results[statement_type] = result
                        file_found = True
                        logger.info(f"Successfully extracted {statement_type.value} from {file_path}")
                        break
                    except Exception as e:
                        logger.warning(f"Failed to extract {statement_type.value} from {file_path}: {e}")

            if not file_found:
                logger.warning(f"No {statement_type.value} file found in {company_folder}")

        return results

    def batch_extract(self,
                     companies_data: Dict[str, str],
                     period_type: str = "FY") -> Dict[str, Dict[StatementType, FieldExtractionResult]]:
        """
        Perform batch extraction for multiple companies.

        Args:
            companies_data: Dictionary mapping company symbols to their data folders
            period_type: Period type ("FY" or "LTM")

        Returns:
            Nested dictionary: {company_symbol: {statement_type: result}}
        """
        batch_results = {}

        for company_symbol, company_folder in companies_data.items():
            try:
                company_results = self.extract_from_company_folder(
                    company_folder, company_symbol, period_type
                )
                batch_results[company_symbol] = company_results
                logger.info(f"Completed extraction for {company_symbol}: {len(company_results)} statements")
            except Exception as e:
                logger.error(f"Batch extraction failed for {company_symbol}: {e}")
                batch_results[company_symbol] = {}

        return batch_results

    def validate_cross_statement_consistency(self,
                                           results: Dict[StatementType, FieldExtractionResult]) -> List[str]:
        """
        Validate consistency across financial statements for the same company and period.

        Args:
            results: Extraction results for all statement types

        Returns:
            List of validation warnings/errors
        """
        warnings = []

        if len(results) < 2:
            return warnings  # Need at least 2 statements for cross-validation

        # Get common periods across all statements
        all_periods = set()
        for result in results.values():
            all_periods.update(result.available_periods)

        for period in all_periods:
            period_data = {}

            # Collect all field data for this period
            for statement_type, result in results.items():
                for field_name, field_periods in result.extracted_fields.items():
                    if period in field_periods:
                        period_data[f"{statement_type.value}_{field_name}"] = field_periods[period]

            # Cross-statement validations

            # 1. Cash consistency: Balance Sheet cash should match Cash Flow ending cash
            bs_cash = period_data.get(f"{StatementType.BALANCE.value}_cash_and_equivalents")
            cf_ending_cash = period_data.get(f"{StatementType.CASH_FLOW.value}_cash_end_of_period")

            if bs_cash is not None and cf_ending_cash is not None:
                if abs(bs_cash - cf_ending_cash) > max(abs(bs_cash) * 0.05, 1.0):
                    warnings.append(
                        f"[{period}] Cash mismatch: Balance Sheet ({bs_cash:.1f}M) ≠ "
                        f"Cash Flow ending balance ({cf_ending_cash:.1f}M)"
                    )

            # 2. Net Income consistency: Income Statement vs Cash Flow
            is_net_income = period_data.get(f"{StatementType.INCOME.value}_net_income")
            cf_net_income = period_data.get(f"{StatementType.CASH_FLOW.value}_net_income")

            if is_net_income is not None and cf_net_income is not None:
                if abs(is_net_income - cf_net_income) > max(abs(is_net_income) * 0.01, 0.1):
                    warnings.append(
                        f"[{period}] Net Income mismatch: Income Statement ({is_net_income:.1f}M) ≠ "
                        f"Cash Flow Statement ({cf_net_income:.1f}M)"
                    )

            # 3. Free Cash Flow calculation validation
            ocf = period_data.get(f"{StatementType.CASH_FLOW.value}_operating_cash_flow")
            capex = period_data.get(f"{StatementType.CASH_FLOW.value}_capital_expenditures")

            if ocf is not None and capex is not None:
                calculated_fcf = ocf + capex  # CapEx is negative
                cf_fcf = period_data.get(f"{StatementType.CASH_FLOW.value}_free_cash_flow")

                if cf_fcf is not None:
                    if abs(calculated_fcf - cf_fcf) > max(abs(calculated_fcf) * 0.05, 1.0):
                        warnings.append(
                            f"[{period}] Free Cash Flow inconsistency: "
                            f"OCF + CapEx = {calculated_fcf:.1f}M ≠ Reported FCF ({cf_fcf:.1f}M)"
                        )

        return warnings

    def generate_data_quality_report(self,
                                   results: Dict[StatementType, FieldExtractionResult]) -> Dict[str, Any]:
        """
        Generate comprehensive data quality assessment across all statements.

        Returns:
            Detailed data quality report with scores, missing fields, and recommendations
        """
        report = {
            'overall_quality_score': 0.0,
            'statement_scores': {},
            'missing_fields_summary': {},
            'validation_warnings': [],
            'completeness_analysis': {},
            'recommendations': [],
            'extraction_metadata': {}
        }

        if not results:
            report['recommendations'].append("No extraction results available for analysis")
            return report

        # Calculate statement-specific quality scores
        total_score = 0.0
        for statement_type, result in results.items():
            statement_name = statement_type.value
            report['statement_scores'][statement_name] = result.data_quality_score
            report['missing_fields_summary'][statement_name] = result.missing_fields
            total_score += result.data_quality_score

            # Extraction metadata
            report['extraction_metadata'][statement_name] = {
                'fields_extracted': len(result.extracted_fields),
                'periods_available': len(result.available_periods),
                'processing_time': result.processing_time_seconds,
                'success_rate': result.success_rate
            }

        # Overall quality score (average of statement scores)
        report['overall_quality_score'] = total_score / len(results)

        # Cross-statement validation warnings
        report['validation_warnings'] = self.validate_cross_statement_consistency(results)

        # Completeness analysis
        all_required_fields = set()
        for extractor in self.extractors.values():
            all_required_fields.update(extractor._get_required_fields())

        missing_critical_fields = set()
        for statement_type, result in results.items():
            extractor = self.extractors[statement_type]
            required_fields = set(extractor._get_required_fields())
            missing_fields = set(result.missing_fields)
            missing_critical_fields.update(required_fields & missing_fields)

        report['completeness_analysis'] = {
            'total_required_fields': len(all_required_fields),
            'missing_critical_fields': list(missing_critical_fields),
            'critical_field_completion_rate': 1.0 - (len(missing_critical_fields) / len(all_required_fields))
        }

        # Generate recommendations
        if report['overall_quality_score'] < self.min_quality_threshold:
            report['recommendations'].append(
                f"Overall quality score ({report['overall_quality_score']:.1%}) below threshold "
                f"({self.min_quality_threshold:.1%}). Review data sources and field mappings."
            )

        if len(report['validation_warnings']) > 0:
            report['recommendations'].append(
                f"Found {len(report['validation_warnings'])} cross-statement inconsistencies. "
                "Review data integrity and calculation methods."
            )

        if missing_critical_fields:
            report['recommendations'].append(
                f"Missing {len(missing_critical_fields)} critical fields: {', '.join(missing_critical_fields)}. "
                "Consider alternative data sources or manual data entry."
            )

        # Performance recommendations
        avg_processing_time = sum(
            meta['processing_time'] for meta in report['extraction_metadata'].values()
        ) / len(report['extraction_metadata'])

        if avg_processing_time > 5.0:  # More than 5 seconds average
            report['recommendations'].append(
                f"High processing time ({avg_processing_time:.1f}s average). "
                "Consider optimizing field mapping or file structure."
            )

        return report

    def integrate_with_financial_calculator(self,
                                          results: Dict[StatementType, FieldExtractionResult],
                                          company_symbol: str) -> Dict[str, Any]:
        """
        Convert extraction results to format compatible with FinancialCalculator.

        Args:
            results: Extraction results from all statement types
            company_symbol: Company ticker symbol

        Returns:
            Dictionary formatted for FinancialCalculator input
        """
        calculator_data = {
            'company_symbol': company_symbol,
            'financial_data': {},
            'metadata': {
                'data_source': 'FieldExtractor',
                'extraction_timestamp': datetime.now().isoformat(),
                'quality_scores': {}
            }
        }

        # Process each statement's data
        for statement_type, result in results.items():
            statement_name = statement_type.value
            calculator_data['metadata']['quality_scores'][statement_name] = result.data_quality_score

            # Convert extracted fields to calculator format
            for field_name, period_data in result.extracted_fields.items():
                # Use most recent period as primary value
                if result.available_periods:
                    latest_period = result.available_periods[0]  # Assuming sorted with latest first
                    if latest_period in period_data:
                        calculator_data['financial_data'][field_name] = period_data[latest_period]

                # Store historical data for trend analysis
                calculator_data['financial_data'][f"{field_name}_history"] = period_data

        return calculator_data

    def get_all_statistics(self) -> Dict[str, Any]:
        """Get comprehensive statistics from all extractors and processing sessions"""
        stats = {
            'extractor_statistics': {},
            'session_summary': {},
            'performance_metrics': {}
        }

        # Individual extractor statistics
        for statement_type, extractor in self.extractors.items():
            stats['extractor_statistics'][statement_type.value] = extractor.get_statistics()

        # Session summary
        if self.extraction_sessions:
            total_sessions = len(self.extraction_sessions)
            successful_sessions = sum(
                1 for session in self.extraction_sessions
                if session['result'].data_quality_score >= self.min_quality_threshold
            )

            stats['session_summary'] = {
                'total_extractions': total_sessions,
                'successful_extractions': successful_sessions,
                'success_rate': successful_sessions / total_sessions if total_sessions > 0 else 0.0,
                'unique_companies': len(set(s['company_symbol'] for s in self.extraction_sessions)),
                'average_quality_score': sum(s['result'].data_quality_score for s in self.extraction_sessions) / total_sessions if total_sessions > 0 else 0.0
            }

        # Performance metrics
        if self.extraction_sessions:
            processing_times = [s['result'].processing_time_seconds for s in self.extraction_sessions]
            stats['performance_metrics'] = {
                'average_processing_time': sum(processing_times) / len(processing_times),
                'total_processing_time': sum(processing_times),
                'fastest_extraction': min(processing_times),
                'slowest_extraction': max(processing_times)
            }

        return stats

    def export_extraction_results(self,
                                 results: Dict[StatementType, FieldExtractionResult],
                                 output_path: str,
                                 format: str = "json") -> str:
        """
        Export extraction results to various formats for external analysis.

        Args:
            results: Extraction results to export
            output_path: Base path for output files
            format: Export format ("json", "excel", "csv")

        Returns:
            Path to exported file(s)
        """
        import json
        from pathlib import Path

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if format.lower() == "json":
            # Convert results to JSON-serializable format
            export_data = {}
            for statement_type, result in results.items():
                export_data[statement_type.value] = {
                    'extracted_fields': result.extracted_fields,
                    'missing_fields': result.missing_fields,
                    'quality_score': result.data_quality_score,
                    'available_periods': result.available_periods,
                    'processing_metadata': {
                        'extraction_timestamp': result.extraction_timestamp.isoformat(),
                        'processing_time_seconds': result.processing_time_seconds,
                        'success_rate': result.success_rate
                    }
                }

            json_path = output_path.with_suffix('.json')
            with open(json_path, 'w') as f:
                json.dump(export_data, f, indent=2)

            return str(json_path)

        # Additional formats (Excel, CSV) can be implemented as needed
        else:
            raise ValueError(f"Export format '{format}' not yet implemented")

    def clear_session_history(self):
        """Clear extraction session history to free memory"""
        self.extraction_sessions.clear()
        logger.info("Cleared extraction session history")