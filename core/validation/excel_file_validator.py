"""
Excel File Format Validator
============================

Validates Excel financial statement files to ensure they follow the expected structure
and contain required financial data for analysis.

This validator provides standalone validation before data extraction, with detailed
reporting and actionable recommendations for file format improvements.

Key Features:
-------------
- Validates file and directory structure
- Checks for required sheets and column headers
- Validates data ranges and completeness
- Verifies data types and reasonable value ranges
- Generates detailed validation reports with recommendations
- Supports batch processing of company directories
- Integrates with existing ValidationOrchestrator

Usage Example:
--------------
>>> from excel_file_validator import ExcelFileValidator
>>>
>>> # Validate a single file
>>> validator = ExcelFileValidator()
>>> result = validator.validate_file("data/companies/NVDA/FY/Income Statement.xlsx")
>>> print(f"Valid: {result.is_valid}, Errors: {len(result.errors)}")
>>>
>>> # Validate entire company directory
>>> batch_result = validator.validate_company_directory("data/companies/NVDA")
>>> print(f"Files validated: {batch_result.files_validated}")
"""

import logging
import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)


class ValidationSeverity(Enum):
    """Severity levels for validation issues"""
    ERROR = "error"           # Must be fixed for data extraction to work
    WARNING = "warning"       # May impact data quality but not critical
    INFO = "info"            # Informational recommendations


class FileType(Enum):
    """Types of financial statement files"""
    INCOME_STATEMENT = "Income Statement"
    BALANCE_SHEET = "Balance Sheet"
    CASH_FLOW = "Cash Flow Statement"


@dataclass
class ValidationIssue:
    """Represents a single validation issue"""
    severity: ValidationSeverity
    message: str
    file_path: Optional[str] = None
    sheet_name: Optional[str] = None
    location: Optional[str] = None  # Cell reference or row/column
    recommendation: Optional[str] = None

    def __str__(self) -> str:
        parts = [f"[{self.severity.value.upper()}]", self.message]
        if self.location:
            parts.append(f"at {self.location}")
        if self.recommendation:
            parts.append(f"→ {self.recommendation}")
        return " ".join(parts)


@dataclass
class FileValidationResult:
    """Result of validating a single Excel file"""
    file_path: str
    file_type: Optional[FileType]
    is_valid: bool

    # Validation details
    has_valid_structure: bool = False
    has_required_sheets: bool = False
    has_fy_columns: bool = False
    fy_column_count: int = 0
    data_completeness_score: float = 0.0

    # Issues found
    errors: List[ValidationIssue] = field(default_factory=list)
    warnings: List[ValidationIssue] = field(default_factory=list)
    info: List[ValidationIssue] = field(default_factory=list)

    # Metadata
    validation_timestamp: datetime = field(default_factory=datetime.now)
    sheets_found: List[str] = field(default_factory=list)
    periods_found: List[str] = field(default_factory=list)

    def add_issue(self, severity: ValidationSeverity, message: str, **kwargs):
        """Add a validation issue"""
        issue = ValidationIssue(severity=severity, message=message, file_path=self.file_path, **kwargs)
        if severity == ValidationSeverity.ERROR:
            self.errors.append(issue)
        elif severity == ValidationSeverity.WARNING:
            self.warnings.append(issue)
        else:
            self.info.append(issue)

    def get_summary(self) -> Dict[str, Any]:
        """Get summary of validation results"""
        return {
            "is_valid": self.is_valid,
            "file_type": self.file_type.value if self.file_type else "Unknown",
            "error_count": len(self.errors),
            "warning_count": len(self.warnings),
            "info_count": len(self.info),
            "fy_columns": self.fy_column_count,
            "data_completeness": f"{self.data_completeness_score:.1%}",
            "periods": self.periods_found
        }


@dataclass
class BatchValidationResult:
    """Result of batch validation of multiple files"""
    directory_path: str
    files_validated: int = 0
    files_passed: int = 0
    files_failed: int = 0

    # Individual file results
    file_results: List[FileValidationResult] = field(default_factory=list)

    # Overall statistics
    total_errors: int = 0
    total_warnings: int = 0
    missing_files: List[str] = field(default_factory=list)

    # Metadata
    validation_timestamp: datetime = field(default_factory=datetime.now)

    @property
    def is_valid(self) -> bool:
        """Check if all files passed validation"""
        return self.files_failed == 0 and len(self.missing_files) == 0

    def get_summary(self) -> Dict[str, Any]:
        """Get summary of batch validation"""
        return {
            "directory": self.directory_path,
            "is_valid": self.is_valid,
            "files_validated": self.files_validated,
            "files_passed": self.files_passed,
            "files_failed": self.files_failed,
            "total_errors": self.total_errors,
            "total_warnings": self.total_warnings,
            "missing_files": self.missing_files,
            "success_rate": f"{self.files_passed / max(1, self.files_validated) * 100:.1f}%"
        }


class ExcelFileValidator:
    """
    Validates Excel financial statement files for structure and data quality.

    This validator ensures Excel files follow expected patterns for successful
    data extraction and analysis.
    """

    # Expected file name patterns for different statement types
    FILE_PATTERNS = {
        FileType.INCOME_STATEMENT: [
            "Income Statement.xlsx",
            "income_statement.xlsx",
            "income.xlsx"
        ],
        FileType.BALANCE_SHEET: [
            "Balance Sheet.xlsx",
            "balance_sheet.xlsx",
            "balance.xlsx"
        ],
        FileType.CASH_FLOW: [
            "Cash Flow Statement.xlsx",
            "cash_flow_statement.xlsx",
            "cashflow.xlsx",
            "cash_flow.xlsx"
        ]
    }

    # Minimum requirements
    MIN_FY_COLUMNS = 3          # At least 3 years of historical data
    MIN_DATA_ROWS = 10          # Minimum rows of financial data
    MIN_COMPLETENESS = 0.50     # 50% of cells should have data

    def __init__(self, strict_mode: bool = False):
        """
        Initialize the Excel file validator

        Args:
            strict_mode: If True, apply stricter validation rules
        """
        self.strict_mode = strict_mode
        self.logger = logging.getLogger(__name__)

        # Adjust requirements based on mode
        if strict_mode:
            self.MIN_FY_COLUMNS = 5
            self.MIN_DATA_ROWS = 15
            self.MIN_COMPLETENESS = 0.70

    def validate_file(
        self,
        file_path: str,
        expected_type: Optional[FileType] = None
    ) -> FileValidationResult:
        """
        Validate a single Excel file

        Args:
            file_path: Path to Excel file
            expected_type: Expected file type (auto-detected if None)

        Returns:
            FileValidationResult with validation details
        """
        self.logger.info(f"Validating Excel file: {file_path}")

        result = FileValidationResult(
            file_path=file_path,
            file_type=expected_type,
            is_valid=True
        )

        try:
            # Step 1: Check file exists and is accessible
            if not self._validate_file_exists(file_path, result):
                result.is_valid = False
                return result

            # Step 2: Detect or verify file type
            detected_type = self._detect_file_type(file_path)
            if expected_type and detected_type != expected_type:
                result.add_issue(
                    ValidationSeverity.WARNING,
                    f"File type mismatch: expected {expected_type.value}, detected {detected_type.value if detected_type else 'unknown'}",
                    recommendation="Rename file to match expected pattern"
                )
            result.file_type = detected_type or expected_type

            # Step 3: Load and validate Excel workbook
            wb = None
            try:
                wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            except InvalidFileException as e:
                result.add_issue(
                    ValidationSeverity.ERROR,
                    f"Invalid Excel file format: {str(e)}",
                    recommendation="Ensure file is a valid .xlsx Excel file"
                )
                result.is_valid = False
                return result
            except Exception as e:
                result.add_issue(
                    ValidationSeverity.ERROR,
                    f"Failed to open Excel file: {str(e)}",
                    recommendation="Check file permissions and integrity"
                )
                result.is_valid = False
                return result

            try:
                # Step 4: Validate sheet structure
                self._validate_sheet_structure(wb, result)

                # Step 5: Validate data structure
                sheet = wb.active
                data = self._load_sheet_data(sheet)

                # Step 6: Find and validate header row
                header_row_idx = self._find_header_row(data, result)
                if header_row_idx is None:
                    result.add_issue(
                        ValidationSeverity.ERROR,
                        "Could not find header row with period columns",
                        recommendation="Ensure file has a row with period headers (FY, FY-1, FY-2 or LTM, LTM.FQ-1, etc. or Q, Q-1, Q-2)"
                    )
                    result.is_valid = False
                    result.has_fy_columns = False
                    return result

                # Step 7: Validate FY columns
                self._validate_fy_columns(data, header_row_idx, result)

                # Step 8: Validate data rows and completeness
                self._validate_data_rows(data, header_row_idx, result)

                # Step 9: Validate data types and ranges
                self._validate_data_quality(data, header_row_idx, result)

                # Final validation status
                result.is_valid = len(result.errors) == 0
                result.has_valid_structure = result.has_fy_columns and len(result.errors) == 0

            finally:
                # Always close the workbook to release file handles
                if wb is not None:
                    wb.close()

        except Exception as e:
            self.logger.error(f"Unexpected error validating {file_path}: {str(e)}")
            result.add_issue(
                ValidationSeverity.ERROR,
                f"Validation failed with unexpected error: {str(e)}",
                recommendation="Contact support with this error message"
            )
            result.is_valid = False

        self.logger.info(
            f"Validation complete: {file_path} - "
            f"{'PASSED' if result.is_valid else 'FAILED'} "
            f"(Errors: {len(result.errors)}, Warnings: {len(result.warnings)})"
        )

        return result

    def validate_company_directory(
        self,
        company_dir: str,
        validate_fy: bool = True,
        validate_ltm: bool = True
    ) -> BatchValidationResult:
        """
        Validate all Excel files in a company directory

        Args:
            company_dir: Path to company data directory
            validate_fy: Whether to validate FY folder
            validate_ltm: Whether to validate LTM folder

        Returns:
            BatchValidationResult with all validation results
        """
        self.logger.info(f"Validating company directory: {company_dir}")

        batch_result = BatchValidationResult(directory_path=company_dir)

        # Check if directory exists
        if not os.path.exists(company_dir):
            batch_result.missing_files.append(f"Directory not found: {company_dir}")
            return batch_result

        # Validate FY folder
        if validate_fy:
            fy_path = os.path.join(company_dir, "FY")
            if os.path.exists(fy_path):
                self._validate_folder(fy_path, "FY", batch_result)
            else:
                batch_result.missing_files.append("FY folder not found")

        # Validate LTM folder
        if validate_ltm:
            ltm_path = os.path.join(company_dir, "LTM")
            if os.path.exists(ltm_path):
                self._validate_folder(ltm_path, "LTM", batch_result)
            else:
                batch_result.missing_files.append("LTM folder not found")

        # Calculate statistics
        batch_result.total_errors = sum(len(r.errors) for r in batch_result.file_results)
        batch_result.total_warnings = sum(len(r.warnings) for r in batch_result.file_results)
        batch_result.files_passed = sum(1 for r in batch_result.file_results if r.is_valid)
        batch_result.files_failed = batch_result.files_validated - batch_result.files_passed

        self.logger.info(
            f"Batch validation complete: {company_dir} - "
            f"Validated: {batch_result.files_validated}, "
            f"Passed: {batch_result.files_passed}, "
            f"Failed: {batch_result.files_failed}"
        )

        return batch_result

    def generate_validation_report(
        self,
        result: FileValidationResult,
        output_path: Optional[str] = None
    ) -> str:
        """
        Generate a detailed validation report

        Args:
            result: Validation result to report on
            output_path: Optional path to save report (prints if None)

        Returns:
            Report as formatted string
        """
        lines = []
        lines.append("=" * 80)
        lines.append("EXCEL FILE VALIDATION REPORT")
        lines.append("=" * 80)
        lines.append(f"File: {result.file_path}")
        lines.append(f"Type: {result.file_type.value if result.file_type else 'Unknown'}")
        lines.append(f"Status: {'✓ PASSED' if result.is_valid else '✗ FAILED'}")
        lines.append(f"Timestamp: {result.validation_timestamp.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")

        # Summary
        lines.append("SUMMARY")
        lines.append("-" * 80)
        summary = result.get_summary()
        for key, value in summary.items():
            if key != "is_valid":
                lines.append(f"  {key.replace('_', ' ').title()}: {value}")
        lines.append("")

        # Errors
        if result.errors:
            lines.append(f"ERRORS ({len(result.errors)})")
            lines.append("-" * 80)
            for issue in result.errors:
                lines.append(f"  • {issue}")
            lines.append("")

        # Warnings
        if result.warnings:
            lines.append(f"WARNINGS ({len(result.warnings)})")
            lines.append("-" * 80)
            for issue in result.warnings:
                lines.append(f"  • {issue}")
            lines.append("")

        # Info
        if result.info:
            lines.append(f"RECOMMENDATIONS ({len(result.info)})")
            lines.append("-" * 80)
            for issue in result.info:
                lines.append(f"  • {issue}")
            lines.append("")

        lines.append("=" * 80)

        report = "\n".join(lines)

        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(report)
            self.logger.info(f"Validation report saved to: {output_path}")

        return report

    # Private helper methods

    def _validate_file_exists(self, file_path: str, result: FileValidationResult) -> bool:
        """Validate that file exists and is accessible"""
        if not os.path.exists(file_path):
            result.add_issue(
                ValidationSeverity.ERROR,
                f"File not found: {file_path}",
                recommendation="Check file path and ensure file exists"
            )
            return False

        if not os.path.isfile(file_path):
            result.add_issue(
                ValidationSeverity.ERROR,
                f"Path is not a file: {file_path}",
                recommendation="Ensure path points to a file, not a directory"
            )
            return False

        if not file_path.endswith(('.xlsx', '.xls')):
            result.add_issue(
                ValidationSeverity.WARNING,
                f"File does not have Excel extension: {file_path}",
                recommendation="Use .xlsx extension for Excel files"
            )

        return True

    def _detect_file_type(self, file_path: str) -> Optional[FileType]:
        """Detect file type based on file name"""
        file_name = os.path.basename(file_path)

        for file_type, patterns in self.FILE_PATTERNS.items():
            for pattern in patterns:
                if pattern.lower() in file_name.lower():
                    return file_type

        return None

    def _validate_sheet_structure(self, wb, result: FileValidationResult):
        """Validate workbook has expected sheets"""
        result.sheets_found = wb.sheetnames

        if len(wb.sheetnames) == 0:
            result.add_issue(
                ValidationSeverity.ERROR,
                "Workbook has no sheets",
                recommendation="Ensure Excel file contains at least one sheet"
            )
            result.has_required_sheets = False
            return

        # In most cases, we use the active sheet
        result.has_required_sheets = True

        if len(wb.sheetnames) > 1:
            result.add_issue(
                ValidationSeverity.INFO,
                f"Workbook has {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}",
                recommendation="Data will be read from the active sheet"
            )

    def _load_sheet_data(self, sheet) -> List[Tuple]:
        """Load sheet data into list of tuples"""
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data

    def _find_header_row(self, data: List[Tuple], result: FileValidationResult) -> Optional[int]:
        """Find the row containing FY or LTM/Q column headers"""
        for i, row in enumerate(data):
            if row and any(
                str(cell).strip() in ['FY', 'LTM', 'Q'] or
                (str(cell).strip().startswith('FY-') and str(cell).strip()[3:].isdigit()) or
                str(cell).strip().startswith('LTM.FQ-') or
                (str(cell).strip().startswith('Q-') and str(cell).strip()[2:].isdigit())
                for cell in row if cell
            ):
                return i
        return None

    def _validate_fy_columns(
        self,
        data: List[Tuple],
        header_row_idx: int,
        result: FileValidationResult
    ):
        """Validate FY or LTM column structure"""
        headers = data[header_row_idx]

        # Find all FY/LTM columns
        fy_columns = {}
        for col_idx, header in enumerate(headers):
            if header:
                header_str = str(header).strip()
                # Handle FY format
                if header_str == 'FY':
                    fy_columns[col_idx] = 'FY'
                    result.periods_found.append('FY')
                elif header_str.startswith('FY-'):
                    try:
                        years_back = int(header_str[3:])
                        fy_columns[col_idx] = header_str
                        result.periods_found.append(header_str)
                    except ValueError:
                        result.add_issue(
                            ValidationSeverity.WARNING,
                            f"Invalid FY column format: {header_str}",
                            location=f"Column {col_idx}",
                            recommendation="Use format: FY, FY-1, FY-2, etc."
                        )
                # Handle LTM format
                elif header_str == 'LTM':
                    fy_columns[col_idx] = 'LTM'
                    result.periods_found.append('LTM')
                elif header_str.startswith('LTM.FQ-'):
                    fy_columns[col_idx] = header_str
                    result.periods_found.append(header_str)
                # Handle Q format (quarterly)
                elif header_str == 'Q':
                    fy_columns[col_idx] = 'Q'
                    result.periods_found.append('Q')
                elif header_str.startswith('Q-'):
                    try:
                        quarters_back = int(header_str[2:])
                        fy_columns[col_idx] = header_str
                        result.periods_found.append(header_str)
                    except ValueError:
                        pass

        result.fy_column_count = len(fy_columns)
        result.has_fy_columns = result.fy_column_count > 0

        # Check minimum FY columns requirement
        if result.fy_column_count < self.MIN_FY_COLUMNS:
            severity = ValidationSeverity.ERROR if self.strict_mode else ValidationSeverity.WARNING
            result.add_issue(
                severity,
                f"Insufficient period columns: found {result.fy_column_count}, expected at least {self.MIN_FY_COLUMNS}",
                recommendation=f"Add more historical periods to reach {self.MIN_FY_COLUMNS} columns"
            )

        # Sort periods (handle FY, LTM, and Q formats)
        def period_sort_key(p):
            if p in ['FY', 'LTM', 'Q']:
                return 0
            elif p.startswith('FY-'):
                return int(p.split('-')[1])
            elif p.startswith('LTM.FQ-'):
                return int(p.split('-')[1])
            elif p.startswith('Q-'):
                return int(p.split('-')[1])
            return 999

        sorted_periods = sorted(result.periods_found, key=period_sort_key)
        result.periods_found = sorted_periods

        # Check for sequential periods (only for FY format)
        if len(sorted_periods) > 1 and sorted_periods[0].startswith('FY'):
            expected_sequence = ['FY'] + [f'FY-{i}' for i in range(1, len(sorted_periods))]
            if sorted_periods != expected_sequence:
                result.add_issue(
                    ValidationSeverity.WARNING,
                    f"FY columns not in sequential order: {', '.join(sorted_periods)}",
                    recommendation=f"Expected sequence: {', '.join(expected_sequence)}"
                )

    def _validate_data_rows(
        self,
        data: List[Tuple],
        header_row_idx: int,
        result: FileValidationResult
    ):
        """Validate data rows below header"""
        data_rows = data[header_row_idx + 1:]

        if len(data_rows) < self.MIN_DATA_ROWS:
            severity = ValidationSeverity.ERROR if self.strict_mode else ValidationSeverity.WARNING
            result.add_issue(
                severity,
                f"Insufficient data rows: found {len(data_rows)}, expected at least {self.MIN_DATA_ROWS}",
                recommendation=f"Ensure file has at least {self.MIN_DATA_ROWS} rows of financial data"
            )

    def _validate_data_quality(
        self,
        data: List[Tuple],
        header_row_idx: int,
        result: FileValidationResult
    ):
        """Validate data quality and completeness"""
        headers = data[header_row_idx]
        data_rows = data[header_row_idx + 1:]

        # Find period column indices (FY/LTM/Q)
        fy_col_indices = [
            i for i, h in enumerate(headers)
            if h and (str(h).strip() in ['FY', 'LTM', 'Q'] or
                     str(h).strip().startswith('FY-') or
                     str(h).strip().startswith('LTM.FQ-') or
                     str(h).strip().startswith('Q-'))
        ]

        if not fy_col_indices:
            return

        # Calculate data completeness
        total_cells = 0
        filled_cells = 0
        numeric_cells = 0

        for row in data_rows[:50]:  # Check first 50 rows
            for col_idx in fy_col_indices:
                if col_idx < len(row):
                    total_cells += 1
                    cell_value = row[col_idx]

                    if cell_value is not None and str(cell_value).strip():
                        filled_cells += 1

                        # Check if numeric
                        try:
                            # Handle string numbers with commas, parentheses, etc.
                            if isinstance(cell_value, str):
                                cleaned = re.sub(r'[,\(\)\$]', '', cell_value.strip())
                                if cleaned and cleaned != '-':
                                    float(cleaned)
                                    numeric_cells += 1
                            elif isinstance(cell_value, (int, float)):
                                numeric_cells += 1
                        except (ValueError, TypeError):
                            pass

        # Calculate completeness score
        result.data_completeness_score = filled_cells / max(1, total_cells)

        # Validate completeness meets minimum
        if result.data_completeness_score < self.MIN_COMPLETENESS:
            severity = ValidationSeverity.ERROR if self.strict_mode else ValidationSeverity.WARNING
            result.add_issue(
                severity,
                f"Low data completeness: {result.data_completeness_score:.1%} (minimum: {self.MIN_COMPLETENESS:.1%})",
                recommendation="Fill in missing values or mark clearly as N/A"
            )

        # Check numeric data percentage
        numeric_percentage = numeric_cells / max(1, filled_cells)
        if numeric_percentage < 0.8:  # At least 80% should be numeric
            result.add_issue(
                ValidationSeverity.WARNING,
                f"Low numeric data percentage: {numeric_percentage:.1%}",
                recommendation="Ensure financial data cells contain numbers, not text"
            )

    def _validate_folder(
        self,
        folder_path: str,
        folder_type: str,
        batch_result: BatchValidationResult
    ):
        """Validate all Excel files in a folder"""
        expected_files = {
            FileType.INCOME_STATEMENT,
            FileType.BALANCE_SHEET,
            FileType.CASH_FLOW
        }

        found_files = set()

        # Scan folder for Excel files
        for file_name in os.listdir(folder_path):
            if not file_name.endswith(('.xlsx', '.xls')):
                continue

            file_path = os.path.join(folder_path, file_name)
            file_type = self._detect_file_type(file_path)

            if file_type:
                found_files.add(file_type)

            # Validate the file
            result = self.validate_file(file_path, expected_type=file_type)
            batch_result.file_results.append(result)
            batch_result.files_validated += 1

        # Check for missing required files
        missing_types = expected_files - found_files
        for missing_type in missing_types:
            batch_result.missing_files.append(
                f"{folder_type}/{missing_type.value}"
            )


# Convenience functions

def validate_excel_file(file_path: str, strict: bool = False) -> FileValidationResult:
    """
    Quick validation of a single Excel file

    Args:
        file_path: Path to Excel file
        strict: Use strict validation rules

    Returns:
        FileValidationResult
    """
    validator = ExcelFileValidator(strict_mode=strict)
    return validator.validate_file(file_path)


def validate_company(company_dir: str, strict: bool = False) -> BatchValidationResult:
    """
    Quick validation of a company directory

    Args:
        company_dir: Path to company data directory
        strict: Use strict validation rules

    Returns:
        BatchValidationResult
    """
    validator = ExcelFileValidator(strict_mode=strict)
    return validator.validate_company_directory(company_dir)


def print_validation_report(result: FileValidationResult):
    """Print validation report to console"""
    validator = ExcelFileValidator()
    report = validator.generate_validation_report(result)
    print(report)


if __name__ == "__main__":
    # Example usage
    import sys

    if len(sys.argv) < 2:
        print("Usage: python excel_file_validator.py <file_or_directory_path>")
        sys.exit(1)

    path = sys.argv[1]

    if os.path.isfile(path):
        # Validate single file
        result = validate_excel_file(path)
        print_validation_report(result)
        sys.exit(0 if result.is_valid else 1)

    elif os.path.isdir(path):
        # Validate directory
        batch_result = validate_company(path)
        print("\n" + "=" * 80)
        print("BATCH VALIDATION SUMMARY")
        print("=" * 80)
        for key, value in batch_result.get_summary().items():
            print(f"{key.replace('_', ' ').title()}: {value}")

        if batch_result.file_results:
            print("\nIndividual File Results:")
            print("-" * 80)
            for file_result in batch_result.file_results:
                status = "✓" if file_result.is_valid else "✗"
                print(f"{status} {os.path.basename(file_result.file_path)} - "
                      f"Errors: {len(file_result.errors)}, Warnings: {len(file_result.warnings)}")

        sys.exit(0 if batch_result.is_valid else 1)

    else:
        print(f"Error: Path not found: {path}")
        sys.exit(1)
