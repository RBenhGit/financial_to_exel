"""
Diagnostic script to help debug data loading issues

This script will examine your financial statement files and show you:
1. What files are found in your company folder
2. What metrics are available in each file
3. Sample data from each file

Usage: python diagnose_data.py <path_to_company_folder>
"""

import sys
import os
import pandas as pd
from openpyxl import load_workbook

def examine_folder_structure(company_folder):
    """Examine the folder structure"""
    print(f"🔍 Examining folder: {company_folder}")
    print("=" * 60)
    
    if not os.path.exists(company_folder):
        print("❌ Folder does not exist!")
        return False
    
    print(f"📁 Contents of {company_folder}:")
    for item in os.listdir(company_folder):
        item_path = os.path.join(company_folder, item)
        if os.path.isdir(item_path):
            print(f"  📂 {item}/")
            for subitem in os.listdir(item_path):
                print(f"    📄 {subitem}")
        else:
            print(f"  📄 {item}")
    
    return True

def examine_excel_file(file_path):
    """Examine an Excel file and show its contents"""
    print(f"\n📊 Examining: {os.path.basename(file_path)}")
    print("-" * 40)
    
    try:
        # Load with openpyxl (same method as the app)
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        
        # Convert to list of lists, then to DataFrame
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        if len(data) > 1:
            df = pd.DataFrame(data[1:], columns=data[0])
        else:
            df = pd.DataFrame(data)
        
        print(f"✅ File loaded successfully")
        print(f"📏 Shape: {df.shape} (rows x columns)")
        
        # Show the first column (metric names)
        print(f"\n📋 Available metrics (first 15):")
        metric_count = 0
        for idx, row in df.iterrows():
            if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip():
                print(f"  {metric_count + 1:2d}. {row.iloc[0]}")
                metric_count += 1
                if metric_count >= 15:
                    break
        
        if metric_count == 0:
            print("  ⚠️ No metrics found in first column")
        
        # Check for specific metrics we need
        print(f"\n🎯 Looking for required metrics:")
        required_metrics = [
            "EBIT", "Income Tax", "EBT", "Net Income",
            "Depreciation & Amortization", "Total Current Assets", 
            "Total Current Liabilities", "Capital Expenditure",
            "Cash from Operations", "Cash from Financing"
        ]
        
        found_metrics = []
        for metric in required_metrics:
            found = False
            for idx, row in df.iterrows():
                if pd.notna(row.iloc[0]) and metric.lower() in str(row.iloc[0]).lower():
                    found_metrics.append(f"✅ {metric} -> '{row.iloc[0]}'")
                    found = True
                    break
            if not found:
                found_metrics.append(f"❌ {metric} -> NOT FOUND")
        
        for result in found_metrics:
            print(f"  {result}")
            
    except Exception as e:
        print(f"❌ Error loading file: {e}")

def main():
    if len(sys.argv) != 2:
        print("Usage: python diagnose_data.py <path_to_company_folder>")
        print("\nExample:")
        print("  python diagnose_data.py C:/data/AAPL")
        sys.exit(1)
    
    company_folder = sys.argv[1]
    
    print("🔬 Financial Data Diagnostic Tool")
    print("=" * 60)
    
    if not examine_folder_structure(company_folder):
        sys.exit(1)
    
    # Look for financial statement files
    for subfolder in ['FY', 'LTM']:
        subfolder_path = os.path.join(company_folder, subfolder)
        if os.path.exists(subfolder_path):
            print(f"\n📂 Examining {subfolder} folder:")
            print("=" * 40)
            
            for file_name in os.listdir(subfolder_path):
                if file_name.endswith(('.xlsx', '.xls')):
                    file_path = os.path.join(subfolder_path, file_name)
                    examine_excel_file(file_path)
        else:
            print(f"\n❌ {subfolder} folder not found")
    
    print("\n" + "=" * 60)
    print("🎯 Diagnostic complete!")
    print("\nIf you see '❌ NOT FOUND' for required metrics, that's likely why")
    print("the FCF calculations are returning empty results.")
    print("\nCheck that your Excel files have the expected metric names,")
    print("or the metric names may need to be adjusted in the code.")

if __name__ == "__main__":
    main()