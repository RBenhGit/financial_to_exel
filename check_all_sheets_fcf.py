import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

def check_all_sheets_for_fcf():
    """
    Check all sheets for FCFF and FCFE calculations and N/A values
    """
    file_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/V/FCF_Analysis_Visa_Inc_Class_A.xlsx"
    
    try:
        wb = load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            if hasattr(wb[sheet_name], 'iter_rows'):  # Skip chart sheets
                print(f"\n=== Analyzing Sheet: {sheet_name} ===")
                
                # Read with pandas for easier analysis
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    
                    # Look for any N/A values
                    na_positions = []
                    for row_idx, row in df.iterrows():
                        for col_idx, cell in enumerate(row):
                            if isinstance(cell, str) and cell.upper() in ['N/A', '#N/A', 'NA']:
                                na_positions.append((row_idx+1, col_idx+1, cell))
                    
                    if na_positions:
                        print(f"Found N/A values in {sheet_name}:")
                        for row, col, value in na_positions:
                            print(f"  Row {row}, Col {col}: {value}")
                            
                            # Show context around N/A value
                            if row-1 < len(df):
                                context_row = df.iloc[row-1]
                                print(f"    Context: {context_row.tolist()}")
                    
                    # Look for FCF-related content
                    fcf_content = []
                    for row_idx, row in df.iterrows():
                        for col_idx, cell in enumerate(row):
                            if isinstance(cell, str) and any(keyword in cell.upper() for keyword in ['FCFF', 'FCFE', 'FREE CASH FLOW']):
                                fcf_content.append((row_idx+1, col_idx+1, cell))
                    
                    if fcf_content:
                        print(f"FCF-related content in {sheet_name}:")
                        for row, col, content in fcf_content:
                            print(f"  Row {row}, Col {col}: {content}")
                            
                            # Show the data around this cell
                            if row-1 < len(df):
                                data_row = df.iloc[row-1]
                                relevant_data = [val for val in data_row.tolist()[:15] if pd.notna(val)]
                                print(f"    Row data: {relevant_data}")
                
                except Exception as e:
                    print(f"Could not read {sheet_name} with pandas: {e}")
        
        # Now specifically look for the issue the user mentioned
        print(f"\n=== Specific Search for Year 9 FCFF/FCFE Issues ===")
        
        # Load without data_only to see formulas
        wb_formulas = load_workbook(file_path, data_only=False)
        
        for sheet_name in wb_formulas.sheetnames:
            if hasattr(wb_formulas[sheet_name], 'iter_rows'):
                ws = wb_formulas[sheet_name]
                
                # Look for cells that contain N/A
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.upper() in ['N/A', '#N/A', '#REF!', '#DIV/0!']:
                            print(f"Found error/N/A in {sheet_name}!{cell.coordinate}: {cell.value}")
                            
                            # Check if this is in a row that might be related to FCF
                            row_num = cell.row
                            col_a_value = ws.cell(row=row_num, column=1).value
                            if col_a_value:
                                print(f"  Row label: {col_a_value}")
                            
                            # Show the formula if it exists
                            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                                print(f"  Formula: {cell.value}")
        
        wb.close()
        wb_formulas.close()
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_all_sheets_for_fcf()