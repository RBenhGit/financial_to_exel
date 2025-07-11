"""
Improved Financial Calculations Module with robust Excel parsing

This version can handle different Excel file formats from Investing.com exports.
"""

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import logging
from datetime import datetime
from scipy import stats
import yfinance as yf

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ImprovedFinancialCalculator:
    """
    Enhanced financial calculator with robust Excel parsing
    """
    
    def __init__(self, company_folder):
        self.company_folder = company_folder
        self.company_name = os.path.basename(company_folder) if company_folder else "Unknown"
        self.financial_data = {}
        self.fcf_results = {}
        self.ticker_symbol = None
        self.current_stock_price = None
        self.market_cap = None
        
    def load_financial_statements(self):
        """
        Load financial statements with improved Excel parsing
        """
        try:
            # Define file paths
            fy_folder = os.path.join(self.company_folder, 'FY')
            ltm_folder = os.path.join(self.company_folder, 'LTM')
            
            # Load FY statements
            for file_name in os.listdir(fy_folder):
                if file_name.endswith(('.xlsx', '.xls')):
                    file_path = os.path.join(fy_folder, file_name)
                    if 'Income Statement' in file_name or 'income' in file_name.lower():
                        self.financial_data['income_fy'] = self._load_excel_data_robust(file_path)
                        logger.info(f"Loaded income statement: {self.financial_data['income_fy'].shape}")
                    elif 'Balance Sheet' in file_name or 'balance' in file_name.lower():
                        self.financial_data['balance_fy'] = self._load_excel_data_robust(file_path)
                        logger.info(f"Loaded balance sheet: {self.financial_data['balance_fy'].shape}")
                    elif 'Cash Flow' in file_name or 'cash' in file_name.lower():
                        self.financial_data['cashflow_fy'] = self._load_excel_data_robust(file_path)
                        logger.info(f"Loaded cash flow: {self.financial_data['cashflow_fy'].shape}")
            
            # Load LTM statements
            for file_name in os.listdir(ltm_folder):
                if file_name.endswith(('.xlsx', '.xls')):
                    file_path = os.path.join(ltm_folder, file_name)
                    if 'Income Statement' in file_name or 'income' in file_name.lower():
                        self.financial_data['income_ltm'] = self._load_excel_data_robust(file_path)
                    elif 'Balance Sheet' in file_name or 'balance' in file_name.lower():
                        self.financial_data['balance_ltm'] = self._load_excel_data_robust(file_path)
                    elif 'Cash Flow' in file_name or 'cash' in file_name.lower():
                        self.financial_data['cashflow_ltm'] = self._load_excel_data_robust(file_path)
                    
            logger.info("Financial statements loaded successfully")
            
        except Exception as e:
            logger.error(f"Error loading financial statements: {e}")
            raise
    
    def _load_excel_data_robust(self, file_path):
        """
        Robust Excel data loading specifically for Investing.com format
        """
        logger.info(f"Loading: {os.path.basename(file_path)}")
        
        try:
            # Load with openpyxl to get raw data
            wb = load_workbook(filename=file_path)
            sheet = wb.active
            
            # Find the header row (contains 'FY-' pattern)
            header_row_idx = None
            data_start_row = None
            
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                # Look for row with FY-9, FY-8, etc.
                if any(cell and 'FY-' in str(cell) for cell in row):
                    header_row_idx = row_idx
                    data_start_row = row_idx + 1
                    logger.info(f"Found header at row {row_idx}")
                    break
            
            if header_row_idx is None:
                logger.warning("Could not find FY header row, using row 7 as default")
                header_row_idx = 7
                data_start_row = 8
            
            # Extract data starting from the identified row
            data = []
            headers = None
            
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx == header_row_idx:
                    # This is the header row
                    headers = [str(cell) if cell is not None else f'Col_{i}' for i, cell in enumerate(row)]
                elif row_idx > header_row_idx:
                    # This is data
                    data.append(row)
            
            if headers and data:
                df = pd.DataFrame(data, columns=headers)
                logger.info(f"Loaded data with shape: {df.shape}")
                
                # The metric names are typically in column 2 (index 2) for Investing.com files
                if len(df.columns) > 2:
                    # Move the metric names column to be the first column
                    metric_col = df.columns[2]  # Column 2 has the metric names
                    cols = [metric_col] + [col for col in df.columns if col != metric_col]
                    df = df[cols]
                    logger.info(f"Moved metric column '{metric_col}' to first position")
                
                return self._clean_dataframe(df)
            else:
                logger.warning(f"No valid data structure found in {file_path}")
                return pd.DataFrame()
                
        except Exception as e:
            logger.error(f"Error loading Excel file {file_path}: {e}")
            return pd.DataFrame()
    
    def _load_with_openpyxl(self, file_path, skip_rows=0):
        """Load Excel with openpyxl and convert to DataFrame"""
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        
        # Convert to list of lists
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx < skip_rows:
                continue
            data.append(row)
        
        if len(data) > 1:
            # Use first row as headers
            df = pd.DataFrame(data[1:], columns=data[0])
        else:
            df = pd.DataFrame(data)
            
        return df
    
    def _score_dataframe(self, df):
        """
        Score a DataFrame based on how likely it contains financial data
        """
        if df.empty:
            return 0
        
        score = 0
        
        # Check for financial keywords in the first column
        if len(df.columns) > 0:
            first_col = df.iloc[:, 0].astype(str).str.lower()
            
            financial_keywords = [
                'revenue', 'income', 'ebit', 'ebitda', 'net income',
                'total assets', 'current assets', 'liabilities',
                'cash flow', 'depreciation', 'capex', 'capital expenditure',
                'operating', 'financing', 'investing'
            ]
            
            for keyword in financial_keywords:
                if first_col.str.contains(keyword, na=False).any():
                    score += 10
        
        # Check for numeric data in other columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        score += len(numeric_cols) * 2
        
        # Prefer DataFrames with more rows (more historical data)
        score += min(len(df), 50)  # Cap at 50 to avoid huge sheets
        
        # Penalize DataFrames that are too wide (likely formatting issues)
        if len(df.columns) > 20:
            score -= 10
        
        return score
    
    def _clean_dataframe(self, df):
        """Clean and standardize the DataFrame"""
        # Remove completely empty rows and columns
        df = df.dropna(how='all').dropna(axis=1, how='all')
        
        # Clean the first column (metric names)
        if len(df.columns) > 0:
            df.iloc[:, 0] = df.iloc[:, 0].astype(str).str.strip()
            
            # Remove rows where first column is empty, 'nan', or just whitespace
            mask = (df.iloc[:, 0] != '') & (df.iloc[:, 0] != 'nan') & (df.iloc[:, 0].notna())
            df = df[mask]
        
        return df
    
    def _extract_metric_values_robust(self, df, metric_names, reverse=False):
        """
        Enhanced metric extraction that tries multiple name variations
        """
        if df.empty:
            logger.warning(f"DataFrame is empty, cannot extract metrics")
            return []
        
        # If metric_names is a string, convert to list
        if isinstance(metric_names, str):
            metric_names = [metric_names]
        
        # Try each metric name variation
        for metric_name in metric_names:
            try:
                # Find row containing the metric (case-insensitive)
                metric_row = None
                first_col = df.iloc[:, 0].astype(str).str.lower()
                
                for idx, cell_value in enumerate(first_col):
                    if pd.notna(cell_value) and metric_name.lower() in cell_value:
                        metric_row = df.iloc[idx]
                        logger.debug(f"Found '{metric_name}' in row {idx}: '{df.iloc[idx, 0]}'")
                        break
                
                if metric_row is not None:
                    # Extract numeric values (skip the first column which is the metric name)
                    values = []
                    for val in metric_row.iloc[1:]:
                        if pd.notna(val) and val != '':
                            try:
                                # Handle different numeric formats
                                if isinstance(val, str):
                                    # Remove commas, parentheses, and convert to float
                                    clean_val = val.replace(',', '').replace('(', '-').replace(')', '').replace('$', '')
                                    # Remove any other currency symbols or text
                                    import re
                                    clean_val = re.sub(r'[^\d.-]', '', clean_val)
                                    if clean_val and clean_val != '-':
                                        values.append(float(clean_val))
                                else:
                                    values.append(float(val))
                            except (ValueError, TypeError):
                                continue  # Skip invalid values
                    
                    if values:
                        if reverse:
                            values = values[::-1]
                        logger.info(f"Extracted {len(values)} values for '{metric_name}': {values[-3:] if len(values) >= 3 else values}")
                        return values
                        
            except Exception as e:
                logger.debug(f"Error extracting metric '{metric_name}': {e}")
                continue
        
        logger.warning(f"Could not find any of these metrics: {metric_names}")
        return []
    
    def calculate_fcf_to_firm(self):
        """Calculate FCFF with enhanced metric extraction"""
        try:
            income_data = self.financial_data.get('income_fy', pd.DataFrame())
            balance_data = self.financial_data.get('balance_fy', pd.DataFrame())
            cashflow_data = self.financial_data.get('cashflow_fy', pd.DataFrame())
            
            if income_data.empty or balance_data.empty or cashflow_data.empty:
                logger.warning("Missing financial data for FCFF calculation")
                return []
            
            # Try multiple variations of metric names
            ebit_values = self._extract_metric_values_robust(
                income_data, ['ebit', 'operating income', 'earnings before interest and tax'], reverse=True
            )
            
            tax_values = self._extract_metric_values_robust(
                income_data, ['income tax', 'tax expense', 'provision for income tax'], reverse=True
            )
            
            ebt_values = self._extract_metric_values_robust(
                income_data, ['ebt', 'earnings before tax', 'pretax income', 'income before tax'], reverse=True
            )
            
            da_values = self._extract_metric_values_robust(
                cashflow_data, ['depreciation', 'amortization', 'depreciation and amortization'], reverse=True
            )
            
            current_assets_values = self._extract_metric_values_robust(
                balance_data, ['current assets', 'total current assets'], reverse=True
            )
            
            current_liabilities_values = self._extract_metric_values_robust(
                balance_data, ['current liabilities', 'total current liabilities'], reverse=True
            )
            
            capex_values = self._extract_metric_values_robust(
                cashflow_data, ['capex', 'capital expenditure', 'capital expenditures', 'pp&e'], reverse=True
            )
            
            logger.info(f"Extracted data lengths - EBIT: {len(ebit_values)}, Tax: {len(tax_values)}, EBT: {len(ebt_values)}")
            logger.info(f"DA: {len(da_values)}, Current Assets: {len(current_assets_values)}, Current Liabilities: {len(current_liabilities_values)}, CapEx: {len(capex_values)}")
            
            # Calculate tax rates
            tax_rates = []
            for i in range(len(ebt_values)):
                if i < len(tax_values) and ebt_values[i] != 0:
                    tax_rate = abs(tax_values[i]) / abs(ebt_values[i])
                    tax_rates.append(min(tax_rate, 0.35))  # Cap at 35%
                else:
                    tax_rates.append(0.25)  # Default tax rate
            
            # Calculate working capital changes
            working_capital_changes = []
            min_length = min(len(current_assets_values), len(current_liabilities_values))
            for i in range(1, min_length):
                wc_change = (current_assets_values[i] - current_liabilities_values[i]) - \
                           (current_assets_values[i-1] - current_liabilities_values[i-1])
                working_capital_changes.append(wc_change)
            
            # Calculate FCFF
            fcff_values = []
            max_years = min(len(ebit_values), len(tax_rates), len(da_values), len(capex_values), len(working_capital_changes) + 1)
            
            for i in range(max_years - 1):
                try:
                    after_tax_ebit = ebit_values[i+1] * (1 - tax_rates[i+1])
                    da_adj = da_values[i+1] if i+1 < len(da_values) else 0
                    wc_change = working_capital_changes[i] if i < len(working_capital_changes) else 0
                    capex_adj = abs(capex_values[i+1]) if i+1 < len(capex_values) else 0
                    
                    fcff = after_tax_ebit + da_adj - wc_change - capex_adj
                    fcff_values.append(fcff)
                    logger.debug(f"Year {i+1}: FCFF = {fcff/1e6:.1f}M")
                except (IndexError, TypeError) as e:
                    logger.debug(f"Skipping year {i+1} due to missing data: {e}")
                    break
            
            self.fcf_results['FCFF'] = fcff_values
            logger.info(f"FCFF calculated: {len(fcff_values)} years")
            return fcff_values
            
        except Exception as e:
            logger.error(f"Error calculating FCFF: {e}")
            return []
    
    def calculate_all_fcf_types(self):
        """Calculate all FCF types with enhanced error handling"""
        try:
            self.load_financial_statements()
            
            # For now, focus on getting one FCF calculation working
            self.calculate_fcf_to_firm()
            
            # Add simplified LFCF calculation
            self._calculate_simple_lfcf()
            
            logger.info("FCF calculations completed")
            return self.fcf_results
            
        except Exception as e:
            logger.error(f"Error in FCF calculations: {e}")
            return {}
    
    def _calculate_simple_lfcf(self):
        """Simplified LFCF calculation"""
        try:
            cashflow_data = self.financial_data.get('cashflow_fy', pd.DataFrame())
            
            if cashflow_data.empty:
                return []
            
            operating_cash_values = self._extract_metric_values_robust(
                cashflow_data, ['operating cash flow', 'cash from operations', 'operating activities'], reverse=True
            )
            
            capex_values = self._extract_metric_values_robust(
                cashflow_data, ['capex', 'capital expenditure', 'capital expenditures'], reverse=True
            )
            
            lfcf_values = []
            min_length = min(len(operating_cash_values), len(capex_values))
            
            for i in range(min_length):
                lfcf = operating_cash_values[i] - abs(capex_values[i])
                lfcf_values.append(lfcf)
            
            self.fcf_results['LFCF'] = lfcf_values
            logger.info(f"LFCF calculated: {len(lfcf_values)} years")
            return lfcf_values
            
        except Exception as e:
            logger.error(f"Error calculating LFCF: {e}")
            return []