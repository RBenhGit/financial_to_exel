#!/usr/bin/env python3
"""
Detailed FCF Comparison between Excel and App calculations
"""

import pandas as pd
import openpyxl
import os
from financial_calculations import FinancialCalculator

def extract_excel_fcf_data():
    """Extract detailed FCF data from Excel file"""
    
    excel_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/MSFT/FCF_Analysis_Microsoft_Corporation.xlsx"
    
    excel_data = {}
    
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Check FCF DATA sheet specifically
        if 'FCF DATA' in workbook.sheetnames:
            sheet = workbook['FCF DATA']
            print("=== FCF DATA Sheet Analysis ===")
            
            # Scan the sheet for structured data
            for row in range(1, 50):
                for col in range(1, 20):
                    cell = sheet.cell(row, col)
                    if cell.value and isinstance(cell.value, str):
                        if any(keyword in cell.value.lower() for keyword in ['fcf', 'free cash flow']):
                            print(f"Row {row}, Col {col}: {cell.value}")
                            
                            # Get the data row
                            row_data = []
                            for c in range(1, 20):
                                val = sheet.cell(row, c).value
                                row_data.append(val)
                            print(f"  Data: {row_data}")
        
        # Check Data Entry sheet for FCF calculations
        if 'Data Entry' in workbook.sheetnames:
            sheet = workbook['Data Entry']
            print("\n=== Data Entry Sheet Analysis ===")
            
            # Look for year headers
            years = []
            for col in range(1, 20):
                cell = sheet.cell(1, col)  # First row
                if cell.value and isinstance(cell.value, (int, float)) and 2010 <= cell.value <= 2030:
                    years.append((col, cell.value))
            
            print(f"Found years in columns: {years}")
            
            # Look for FCF calculation rows
            fcf_rows = {}
            for row in range(1, 100):
                cell = sheet.cell(row, 1)  # First column
                if cell.value and isinstance(cell.value, str):
                    cell_lower = cell.value.lower()
                    if any(keyword in cell_lower for keyword in 
                          ['free cash flow', 'fcf', 'operating cash flow', 'capital expenditure', 'capex']):
                        print(f"\nRow {row}: {cell.value}")
                        
                        # Extract values for each year
                        values = []
                        for year_col, year in years:
                            val = sheet.cell(row, year_col).value
                            values.append(val)
                        
                        fcf_rows[cell.value] = {
                            'years': [y[1] for y in years],
                            'values': values
                        }
                        print(f"  Values: {values}")
            
            excel_data['fcf_rows'] = fcf_rows
        
        # Check DCF sheet for projected FCF
        if 'DCF' in workbook.sheetnames:
            sheet = workbook['DCF']
            print("\n=== DCF Sheet Analysis ===")
            
            # Look for FCF row (we found B13 earlier)
            fcf_row = sheet[13]  # Row 13
            fcf_values = []
            for cell in fcf_row:
                if cell.value and isinstance(cell.value, (int, float)):
                    fcf_values.append(cell.value)
            
            print(f"DCF FCF values: {fcf_values}")
            excel_data['dcf_fcf'] = fcf_values
        
        return excel_data
        
    except Exception as e:
        print(f"Error extracting Excel data: {e}")
        return {}

def get_app_fcf_data():
    """Get FCF data from the app"""
    
    msft_folder = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/MSFT"
    
    try:
        calc = FinancialCalculator(msft_folder)
        fcf_results = calc.calculate_all_fcf_types()
        
        print("=== App FCF Results ===")
        for fcf_type, values in fcf_results.items():
            print(f"{fcf_type}: {values}")
        
        return fcf_results
        
    except Exception as e:
        print(f"Error getting app FCF data: {e}")
        return {}

def create_comparison_table():
    """Create a detailed comparison table"""
    
    print("\n" + "="*100)
    print("DETAILED FCF COMPARISON: EXCEL vs APP")
    print("="*100)
    
    excel_data = extract_excel_fcf_data()
    app_data = get_app_fcf_data()
    
    # Create comparison table
    print("\n📊 COMPARISON TABLE")
    print("-" * 120)
    print(f"{'Calculation Method':<25} {'Excel Values':<40} {'App Values':<40} {'Difference':<15}")
    print("-" * 120)
    
    # Compare FCFF
    if 'FCFF' in app_data:
        app_fcff = app_data['FCFF']
        excel_fcff = excel_data.get('dcf_fcf', [])
        if excel_fcff:
            print(f"{'FCFF (Free Cash Flow)':<25} {str(excel_fcff[:5]):<40} {str(app_fcff[:5]):<40} {'See Details':<15}")
    
    # Compare FCFE
    if 'FCFE' in app_data:
        app_fcfe = app_data['FCFE']
        print(f"{'FCFE (To Equity)':<25} {'Not Found':<40} {str(app_fcfe[:5]):<40} {'N/A':<15}")
    
    # Compare LFCF
    if 'LFCF' in app_data:
        app_lfcf = app_data['LFCF']
        print(f"{'LFCF (Levered FCF)':<25} {'Not Found':<40} {str(app_lfcf[:5]):<40} {'N/A':<15}")
    
    print("-" * 120)
    
    # Detailed analysis
    print("\n🔍 DETAILED ANALYSIS")
    print("-" * 80)
    
    print("\n1. CALCULATION METHODOLOGIES:")
    print("   Excel (from DCF sheet):")
    print("   - Appears to show projected FCF values for DCF model")
    print("   - Values: [76301.5, 79735.1, 83323.1, 87072.7, 90991.0, ...]")
    print("   - These appear to be FUTURE projections, not historical calculations")
    
    print("\n   App Calculations:")
    print("   - FCFF: EBIT(1-Tax) + D&A - ΔWorking Capital - CapEx")
    print("   - FCFE: Net Income + D&A - ΔWorking Capital - CapEx + Net Borrowing")
    print("   - LFCF: Operating Cash Flow - CapEx")
    print("   - These are HISTORICAL calculations from financial statements")
    
    print("\n2. KEY DIFFERENCES:")
    print("   • TIME ORIENTATION:")
    print("     - Excel: Future projections for DCF valuation")
    print("     - App: Historical analysis from actual financial data")
    
    print("\n   • DATA SOURCE:")
    print("     - Excel: Manually entered or calculated projections")
    print("     - App: Automated extraction from Investing.com financial statements")
    
    print("\n   • CALCULATION SCOPE:")
    print("     - Excel: Single FCF projection (likely FCFF)")
    print("     - App: Three different FCF methodologies")
    
    print("\n3. COMMONALITIES:")
    print("   • Both aim to calculate Free Cash Flow")
    print("   • Both consider core components: Operations, CapEx, Working Capital")
    print("   • Both can be used for valuation purposes")
    
    # Show actual values comparison
    if 'FCFF' in app_data and excel_data.get('dcf_fcf'):
        print("\n📈 VALUE COMPARISON (First 5 periods):")
        excel_values = excel_data['dcf_fcf'][:5]
        app_values = app_data['FCFF'][:5]
        
        print(f"{'Period':<10} {'Excel FCF':<15} {'App FCFF':<15} {'Difference':<15} {'% Diff':<10}")
        print("-" * 70)
        
        for i in range(min(len(excel_values), len(app_values))):
            diff = excel_values[i] - app_values[i]
            pct_diff = (diff / app_values[i] * 100) if app_values[i] != 0 else 0
            print(f"{i+1:<10} {excel_values[i]:<15.1f} {app_values[i]:<15.1f} {diff:<15.1f} {pct_diff:<10.1f}%")

if __name__ == "__main__":
    create_comparison_table()