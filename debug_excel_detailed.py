#!/usr/bin/env python3
"""
Detailed Excel File Structure Analysis
"""

import sys
from pathlib import Path

# Add the project root to Python path
project_root = Path(__file__).parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from openpyxl import load_workbook

def debug_excel_detailed():
    """Detailed analysis of Excel file structure"""
    print("Detailed Excel File Analysis")
    print("=" * 40)
    
    test_file = "data/companies/GOOG/FY/Alphabet Inc Class C - Income Statement.xlsx"
    
    wb = load_workbook(filename=test_file)
    sheet = wb.active
    
    print(f"Worksheet name: {sheet.title}")
    print(f"Max row: {sheet.max_row}")
    print(f"Max column: {sheet.max_column}")
    
    # Convert to list of lists
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    print(f"Total rows: {len(data)}")
    
    # Find header row
    header_row_idx = None
    for i, row in enumerate(data):
        if row and any('FY-' in str(cell) or 'FY' == str(cell) for cell in row if cell):
            header_row_idx = i
            break
    
    if header_row_idx is not None:
        print(f"Header row index: {header_row_idx}")
        headers = data[header_row_idx]
        print(f"Headers: {headers}")
        
        # Print several rows around and after the header
        print("\nRows around header:")
        start_row = max(0, header_row_idx - 2)
        end_row = min(len(data), header_row_idx + 20)
        
        for i in range(start_row, end_row):
            row = data[i]
            marker = " << HEADER" if i == header_row_idx else ""
            print(f"Row {i}: {row}{marker}")
            
            # Show non-empty values in first few columns
            if i > header_row_idx:  # Data rows
                non_empty = []
                for j, cell in enumerate(row[:5]):  # First 5 columns
                    if cell is not None and str(cell).strip():
                        non_empty.append(f"Col{j}: '{cell}'")
                if non_empty:
                    print(f"    Non-empty: {', '.join(non_empty)}")
    
    print("\n" + "=" * 40)
    print("Alternative approach - scanning all rows for variable-like names:")
    
    # Look for rows that might contain financial variable names
    potential_vars = []
    for i, row in enumerate(data):
        if i <= header_row_idx:  # Skip header and above
            continue
            
        # Check all columns for potential variable names
        for j, cell in enumerate(row):
            if cell is not None:
                cell_str = str(cell).strip()
                # Look for terms that suggest financial variables
                if any(term in cell_str.lower() for term in ['revenue', 'income', 'profit', 'expense', 'cost', 'sales']):
                    potential_vars.append((i, j, cell_str))
    
    print(f"Found {len(potential_vars)} potential variable names:")
    for row_idx, col_idx, name in potential_vars[:10]:  # First 10
        print(f"  Row {row_idx}, Col {col_idx}: '{name}'")

if __name__ == "__main__":
    debug_excel_detailed()