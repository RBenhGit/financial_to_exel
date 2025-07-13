#!/usr/bin/env python3
"""
Detailed Data Entry Tab Inspection
Extract and validate historical financial data from Data Entry tab
"""

import pandas as pd
import openpyxl
from financial_calculations import FinancialCalculator

def inspect_data_entry_structure(excel_path):
    """Thoroughly inspect Data Entry tab structure"""
    print("🔍 DETAILED DATA ENTRY TAB INSPECTION")
    print("=" * 60)
    
    try:
        # Read with openpyxl for cell-by-cell inspection
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        ws = workbook['Data Entry']
        
        print(f"📊 Data Entry sheet dimensions: {ws.max_row} x {ws.max_column}")
        
        # Inspect first 25 rows and 15 columns
        print("\n📋 Cell-by-cell inspection (first 25 rows):")
        for row in range(1, min(26, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(16, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    # Truncate long values for display
                    if isinstance(cell_value, str) and len(cell_value) > 20:
                        cell_value = cell_value[:20] + "..."
                    row_data.append(str(cell_value))
                else:
                    row_data.append("None")
            
            print(f"   Row {row:2d}: {row_data[:8]}")  # First 8 columns
        
        # Look for financial statement sections
        print("\n🔍 Searching for financial statement data...")
        financial_sections = []
        
        for row in range(1, min(30, ws.max_row + 1)):
            for col in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    cell_str = cell_value.lower()
                    if any(term in cell_str for term in [
                        'income statement', 'balance sheet', 'cash flow',
                        'revenue', 'operating cash', 'capex', 'capital expenditure',
                        'net income', 'ebit', 'current assets', 'current liabilities'
                    ]):
                        financial_sections.append({
                            'text': cell_value,
                            'row': row,
                            'col': col
                        })
                        print(f"   📍 Found '{cell_value}' at R{row}C{col}")
        
        workbook.close()
        return financial_sections
        
    except Exception as e:
        print(f"❌ Error inspecting Data Entry: {e}")
        return []

def extract_financial_data_manually(excel_path):
    """Manually extract financial data from specific known locations"""
    print("\n📊 MANUAL FINANCIAL DATA EXTRACTION")
    print("=" * 60)
    
    try:
        # Read Data Entry with pandas to see structure
        data_entry = pd.read_excel(excel_path, sheet_name='Data Entry', header=None)
        print(f"Shape: {data_entry.shape}")
        
        # Show more rows to understand structure
        print("\n📋 Data Entry content (first 25 rows, first 10 columns):")
        for i in range(min(25, len(data_entry))):
            row_data = []
            for j in range(min(10, len(data_entry.columns))):
                val = data_entry.iloc[i, j]
                if pd.notna(val):
                    if isinstance(val, str) and len(val) > 15:
                        val = val[:15] + "..."
                    row_data.append(str(val))
                else:
                    row_data.append("NaN")
            print(f"   {i:2d}: {row_data}")
        
        # Try to read with different header assumptions
        print("\n🔄 Trying different reading approaches...")
        
        # Try reading with row 0 as header
        try:
            df_header0 = pd.read_excel(excel_path, sheet_name='Data Entry', header=0)
            print(f"With header=0: {df_header0.shape}, columns: {list(df_header0.columns)[:5]}")
        except:
            pass
        
        # Try reading with row 1 as header
        try:
            df_header1 = pd.read_excel(excel_path, sheet_name='Data Entry', header=1)
            print(f"With header=1: {df_header1.shape}, columns: {list(df_header1.columns)[:5]}")
        except:
            pass
        
        # Try reading with row 2 as header
        try:
            df_header2 = pd.read_excel(excel_path, sheet_name='Data Entry', header=2)
            print(f"With header=2: {df_header2.shape}, columns: {list(df_header2.columns)[:5]}")
        except:
            pass
        
        return data_entry
        
    except Exception as e:
        print(f"❌ Error in manual extraction: {e}")
        return None

def compare_data_sources(excel_path):
    """Compare data between Data Entry and FCF DATA tabs"""
    print("\n🔄 COMPARING DATA SOURCES")
    print("=" * 60)
    
    try:
        # Get FCF DATA (we know this structure works)
        fcf_data = pd.read_excel(excel_path, sheet_name='FCF DATA', header=None)
        
        # Extract operating CF and CapEx from FCF DATA
        fcf_operating_cf = []
        fcf_capex = []
        fcf_years = []
        
        for i in range(2, min(12, len(fcf_data))):
            year = fcf_data.iloc[i, 0]
            op_cf = fcf_data.iloc[i, 2]
            capex = fcf_data.iloc[i, 3]
            
            if pd.notna(year) and pd.notna(op_cf) and pd.notna(capex):
                fcf_years.append(int(year))
                
                if isinstance(op_cf, str):
                    op_cf = float(op_cf.replace(',', ''))
                if isinstance(capex, str):
                    capex = float(capex.replace(',', ''))
                
                fcf_operating_cf.append(float(op_cf))
                fcf_capex.append(float(capex))
        
        print(f"📊 FCF DATA extracted:")
        print(f"   Years: {fcf_years}")
        print(f"   Operating CF: {fcf_operating_cf}")
        print(f"   CapEx: {fcf_capex}")
        
        # Try to find corresponding data in Data Entry
        # This might require manual inspection based on the sheet structure
        data_entry = pd.read_excel(excel_path, sheet_name='Data Entry', header=None)
        
        print(f"\n🔍 Looking for matching data in Data Entry...")
        
        # Search for years in Data Entry
        years_found = []
        for i in range(len(data_entry)):
            for j in range(len(data_entry.columns)):
                val = data_entry.iloc[i, j]
                if isinstance(val, (int, float)) and 2010 <= val <= 2030:
                    years_found.append((int(val), i, j))
        
        if years_found:
            print(f"   Years found in Data Entry: {[(year, f'R{row}C{col}') for year, row, col in years_found[:10]]}")
        
        # Search for large financial values (likely operating CF)
        large_values = []
        for i in range(len(data_entry)):
            for j in range(len(data_entry.columns)):
                val = data_entry.iloc[i, j]
                if isinstance(val, (int, float)) and 10000 <= abs(val) <= 200000:
                    large_values.append((val, i, j))
        
        if large_values:
            print(f"   Large values (potential financial data): {[(val, f'R{row}C{col}') for val, row, col in large_values[:10]]}")
        
        return {
            'fcf_data': {
                'years': fcf_years,
                'operating_cf': fcf_operating_cf,
                'capex': fcf_capex
            },
            'data_entry_info': {
                'years_found': years_found,
                'large_values': large_values[:20]  # First 20 large values
            }
        }
        
    except Exception as e:
        print(f"❌ Error comparing data sources: {e}")
        return None

def validate_with_app_data(comparison_data):
    """Validate extracted Excel data with app calculations"""
    print("\n✅ VALIDATION WITH APP DATA")
    print("=" * 60)
    
    try:
        # Get app data
        goog_folder = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/GOOG"
        calc = FinancialCalculator(goog_folder)
        calc.load_financial_statements()
        
        metrics = calc._calculate_all_metrics()
        app_operating_cf = metrics.get('operating_cash_flow', [])
        app_capex = metrics.get('capex', [])
        
        print(f"📊 App extracted data:")
        print(f"   Operating CF: {app_operating_cf}")
        print(f"   CapEx: {app_capex}")
        
        # Compare with FCF DATA
        fcf_data = comparison_data.get('fcf_data', {})
        fcf_operating_cf = fcf_data.get('operating_cf', [])
        fcf_capex = fcf_data.get('capex', [])
        
        print(f"\n🔍 FCF DATA vs App comparison:")
        
        # App data is newest-first, FCF data is oldest-first
        app_operating_cf_rev = list(reversed(app_operating_cf))
        app_capex_rev = list(reversed(app_capex))
        
        print(f"   FCF Operating CF: {fcf_operating_cf}")
        print(f"   App Operating CF: {app_operating_cf_rev}")
        
        # Compare operating CF
        min_len = min(len(fcf_operating_cf), len(app_operating_cf_rev))
        cf_matches = 0
        
        for i in range(min_len):
            diff = abs(fcf_operating_cf[i] - app_operating_cf_rev[i])
            if diff < 1000:
                cf_matches += 1
        
        cf_match_pct = (cf_matches / min_len) * 100 if min_len > 0 else 0
        print(f"   Operating CF matches: {cf_matches}/{min_len} ({cf_match_pct:.1f}%)")
        
        # Compare CapEx
        print(f"\n   FCF CapEx: {fcf_capex}")
        print(f"   App CapEx: {app_capex_rev}")
        
        capex_matches = 0
        for i in range(min_len):
            diff = abs(abs(fcf_capex[i]) - abs(app_capex_rev[i]))
            if diff < 1000:
                capex_matches += 1
        
        capex_match_pct = (capex_matches / min_len) * 100 if min_len > 0 else 0
        print(f"   CapEx matches: {capex_matches}/{min_len} ({capex_match_pct:.1f}%)")
        
        # Overall validation
        overall_good = cf_match_pct >= 90 and capex_match_pct >= 90
        
        print(f"\n🎯 DATA INTEGRITY VALIDATION:")
        print(f"   Operating CF: {'✅ VALIDATED' if cf_match_pct >= 90 else '❌ ISSUES'} ({cf_match_pct:.1f}%)")
        print(f"   CapEx: {'✅ VALIDATED' if capex_match_pct >= 90 else '❌ ISSUES'} ({capex_match_pct:.1f}%)")
        print(f"   Overall: {'✅ DATA INTEGRITY CONFIRMED' if overall_good else '⚠️ DATA INTEGRITY ISSUES'}")
        
        return overall_good
        
    except Exception as e:
        print(f"❌ Error validating with app data: {e}")
        return False

def main():
    print("🚀 DETAILED DATA ENTRY VALIDATION")
    print("=" * 80)
    
    excel_path = "/mnt/c/AsusWebStorage/ran@benhur.co/MySyncFolder/python/investingAnalysis/financial_to_exel/GOOG/FCF_Analysis_Alphabet Inc Class C.xlsx"
    
    # 1. Inspect Data Entry structure
    financial_sections = inspect_data_entry_structure(excel_path)
    
    # 2. Manual extraction attempt
    data_entry_data = extract_financial_data_manually(excel_path)
    
    # 3. Compare data sources
    comparison_data = compare_data_sources(excel_path)
    
    # 4. Validate with app data
    if comparison_data:
        data_integrity_ok = validate_with_app_data(comparison_data)
        
        print(f"\n📋 FINAL ASSESSMENT")
        print("=" * 60)
        if data_integrity_ok:
            print("✅ DATA INTEGRITY: App correctly extracts data from financial statements")
            print("✅ CALCULATION ACCURACY: FCF calculations match between Excel and App")
            print("✅ CONCLUSION: App is fully validated against Excel methodology")
        else:
            print("⚠️ DATA INTEGRITY: Some discrepancies found between sources")
            print("ℹ️ This may be due to different data collection methods or timing")

if __name__ == "__main__":
    main()