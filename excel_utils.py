"""
Excel Utilities Module

This module provides dynamic Excel data extraction utilities to replace
hardcoded cell references and make the application more robust across
different Excel file formats.
"""

import os
import re
import logging
from typing import Dict, List, Tuple, Optional, Any, Union
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from dataclasses import dataclass
from config import get_excel_config, get_financial_metrics_config

logger = logging.getLogger(__name__)

@dataclass
class CellLocation:
    """Represents a cell location in an Excel worksheet"""
    row: int
    column: int
    value: Any = None

@dataclass
class ExcelMetric:
    """Represents a financial metric found in Excel"""
    name: str
    location: CellLocation
    data_columns: List[int]
    values: List[Any]

class ExcelDataExtractor:
    """
    Dynamic Excel data extractor that can handle various Excel formats
    without relying on hardcoded cell references
    """
    
    def __init__(self, workbook_path: str):
        """
        Initialize the Excel data extractor
        
        Args:
            workbook_path (str): Path to the Excel workbook
        """
        self.workbook_path = workbook_path
        self.workbook = None
        self.worksheet = None
        self.config = get_excel_config()
        self.metrics_config = get_financial_metrics_config()
        
    def __enter__(self):
        """Context manager entry"""
        self.load_workbook()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit"""
        if self.workbook:
            self.workbook.close()
    
    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.workbook = load_workbook(self.workbook_path, read_only=True)
            self.worksheet = self.workbook.active
            logger.info(f"Loaded workbook: {os.path.basename(self.workbook_path)}")
        except Exception as e:
            logger.error(f"Failed to load workbook {self.workbook_path}: {e}")
            raise
    
    def find_company_name(self) -> Optional[str]:
        """
        Dynamically find company name in the Excel sheet
        
        Returns:
            Optional[str]: Company name if found, None otherwise
        """
        if not self.worksheet:
            return None
        
        # Try default positions first
        for row, col in self.config.default_company_name_positions:
            try:
                value = self.worksheet.cell(row, col).value
                if value and isinstance(value, str) and len(value.strip()) > 0:
                    # Basic validation - company names typically don't contain certain patterns
                    if not self._is_likely_company_name(value):
                        continue
                    logger.info(f"Found company name at ({row}, {col}): {value}")
                    return value.strip()
            except Exception as e:
                logger.debug(f"Error checking cell ({row}, {col}): {e}")
                continue
        
        # If not found in default positions, search more broadly
        return self._search_company_name_broadly()
    
    def _is_likely_company_name(self, value: str) -> bool:
        """
        Check if a string is likely to be a company name
        
        Args:
            value (str): String to check
            
        Returns:
            bool: True if likely a company name
        """
        value = value.strip()
        
        # Basic filters
        if len(value) < 2 or len(value) > 100:
            return False
        
        # Skip common non-company patterns
        skip_patterns = [
            r'^period end date',
            r'^fiscal year',
            r'^quarter',
            r'^annual',
            r'^monthly',
            r'^\d{4}',  # Years
            r'^q[1-4]',  # Quarters
            r'^jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec',  # Months
            r'^total|sum|average|net|gross',  # Financial terms
            r'^usd|eur|gbp',  # Currencies
        ]
        
        for pattern in skip_patterns:
            if re.match(pattern, value.lower()):
                return False
        
        # Look for company indicators
        company_indicators = [
            r'\b(inc|corp|corporation|company|ltd|llc|plc)\b',
            r'\b(technologies|tech|systems|group|holdings)\b',
            r'\b(international|global|worldwide)\b'
        ]
        
        for pattern in company_indicators:
            if re.search(pattern, value.lower()):
                return True
        
        # If it contains multiple words and capital letters, likely a company name
        if len(value.split()) > 1 and any(c.isupper() for c in value):
            return True
        
        return False
    
    def _search_company_name_broadly(self) -> Optional[str]:
        """
        Search for company name more broadly in the worksheet
        
        Returns:
            Optional[str]: Company name if found
        """
        if not self.worksheet:
            return None
        
        # Search in first 10 rows and first 10 columns
        for row in range(1, 11):
            for col in range(1, 11):
                try:
                    value = self.worksheet.cell(row, col).value
                    if value and isinstance(value, str):
                        if self._is_likely_company_name(value):
                            logger.info(f"Found company name at ({row}, {col}): {value}")
                            return value.strip()
                except Exception as e:
                    logger.debug(f"Error checking cell ({row}, {col}): {e}")
                    continue
        
        logger.warning("Could not find company name in worksheet")
        return None
    
    def find_period_end_dates(self) -> List[str]:
        """
        Find period end dates in the worksheet
        
        Returns:
            List[str]: List of period end dates
        """
        if not self.worksheet:
            return []
        
        # First try the default location
        try:
            period_row = self.config.default_period_end_row
            period_col = self.config.default_period_end_column
            
            if self.worksheet.cell(period_row, period_col).value == "Period End Date":
                dates = []
                for col in range(self.config.data_start_column, self.config.max_scan_columns):
                    cell_value = self.worksheet.cell(period_row, col).value
                    if cell_value is not None:
                        dates.append(str(cell_value))
                    else:
                        break
                
                if dates:
                    logger.info(f"Found {len(dates)} period end dates at default location")
                    return dates
        except Exception as e:
            logger.debug(f"Error finding period dates at default location: {e}")
        
        # Search for "Period End Date" more broadly
        return self._search_period_end_dates_broadly()
    
    def _search_period_end_dates_broadly(self) -> List[str]:
        """
        Search for period end dates more broadly in the worksheet
        
        Returns:
            List[str]: List of period end dates
        """
        if not self.worksheet:
            return []
        
        # Search for "Period End Date" in first 20 rows
        for row in range(1, 21):
            for col in range(1, 10):
                try:
                    value = self.worksheet.cell(row, col).value
                    if value and isinstance(value, str) and "Period End Date" in value:
                        # Found the row, now extract dates
                        dates = []
                        for data_col in range(col + 1, self.config.max_scan_columns):
                            cell_value = self.worksheet.cell(row, data_col).value
                            if cell_value is not None:
                                dates.append(str(cell_value))
                            else:
                                break
                        
                        if dates:
                            logger.info(f"Found {len(dates)} period end dates at row {row}")
                            return dates
                except Exception as e:
                    logger.debug(f"Error checking cell ({row}, {col}): {e}")
                    continue
        
        logger.warning("Could not find period end dates in worksheet")
        return []
    
    def find_financial_metric(self, metric_name: str) -> Optional[ExcelMetric]:
        """
        Find a financial metric in the worksheet
        
        Args:
            metric_name (str): Name of the financial metric to find
            
        Returns:
            Optional[ExcelMetric]: ExcelMetric object if found, None otherwise
        """
        if not self.worksheet:
            return None
        
        # Search for the metric in the worksheet
        for row in range(1, self.config.max_scan_rows + 1):
            for col in range(1, 10):  # Search in first 10 columns for labels
                try:
                    value = self.worksheet.cell(row, col).value
                    if value and isinstance(value, str) and metric_name in value:
                        # Found the metric, now extract data
                        data_columns = []
                        values = []
                        
                        # Extract historical data (9 years for FY)
                        for j in range(9):
                            data_col = self.config.data_start_column + j
                            data_columns.append(data_col)
                            cell_value = self.worksheet.cell(row, data_col).value
                            values.append(cell_value)
                        
                        location = CellLocation(row, col, value)
                        metric = ExcelMetric(metric_name, location, data_columns, values)
                        
                        logger.info(f"Found metric '{metric_name}' at row {row}")
                        return metric
                        
                except Exception as e:
                    logger.debug(f"Error checking cell ({row}, {col}): {e}")
                    continue
        
        logger.warning(f"Could not find metric '{metric_name}' in worksheet")
        return None
    
    def extract_all_financial_metrics(self, statement_type: str) -> Dict[str, ExcelMetric]:
        """
        Extract all financial metrics for a given statement type
        
        Args:
            statement_type (str): Type of statement ('income', 'balance', 'cashflow')
            
        Returns:
            Dict[str, ExcelMetric]: Dictionary of metric names to ExcelMetric objects
        """
        if not self.worksheet:
            return {}
        
        # Get the appropriate metrics configuration
        if statement_type == 'income':
            metrics_dict = self.metrics_config.income_metrics
        elif statement_type == 'balance':
            metrics_dict = self.metrics_config.balance_metrics
        elif statement_type == 'cashflow':
            metrics_dict = self.metrics_config.cashflow_metrics
        else:
            logger.error(f"Unknown statement type: {statement_type}")
            return {}
        
        extracted_metrics = {}
        
        for metric_name in metrics_dict.keys():
            metric = self.find_financial_metric(metric_name)
            if metric:
                extracted_metrics[metric_name] = metric
        
        logger.info(f"Extracted {len(extracted_metrics)} metrics for {statement_type} statement")
        return extracted_metrics
    
    def extract_ltm_data(self, metric_name: str) -> Optional[Any]:
        """
        Extract LTM (Latest Twelve Months) data for a specific metric
        
        Args:
            metric_name (str): Name of the financial metric
            
        Returns:
            Optional[Any]: LTM value if found, None otherwise
        """
        if not self.worksheet:
            return None
        
        metric = self.find_financial_metric(metric_name)
        if not metric:
            return None
        
        # Extract LTM data from the configured column
        try:
            ltm_value = self.worksheet.cell(metric.location.row, self.config.ltm_column).value
            logger.debug(f"LTM value for '{metric_name}': {ltm_value}")
            return ltm_value
        except Exception as e:
            logger.error(f"Error extracting LTM data for '{metric_name}': {e}")
            return None

def get_company_name_from_excel(file_path: str) -> Optional[str]:
    """
    Utility function to extract company name from Excel file
    
    Args:
        file_path (str): Path to Excel file
        
    Returns:
        Optional[str]: Company name if found
    """
    try:
        with ExcelDataExtractor(file_path) as extractor:
            return extractor.find_company_name()
    except Exception as e:
        logger.error(f"Error extracting company name from {file_path}: {e}")
        return None

def get_period_dates_from_excel(file_path: str) -> List[str]:
    """
    Utility function to extract period dates from Excel file
    
    Args:
        file_path (str): Path to Excel file
        
    Returns:
        List[str]: List of period dates
    """
    try:
        with ExcelDataExtractor(file_path) as extractor:
            return extractor.find_period_end_dates()
    except Exception as e:
        logger.error(f"Error extracting period dates from {file_path}: {e}")
        return []

def extract_financial_data_from_excel(file_path: str, statement_type: str) -> Dict[str, Any]:
    """
    Utility function to extract financial data from Excel file
    
    Args:
        file_path (str): Path to Excel file
        statement_type (str): Type of statement ('income', 'balance', 'cashflow')
        
    Returns:
        Dict[str, Any]: Dictionary of extracted financial data
    """
    try:
        with ExcelDataExtractor(file_path) as extractor:
            metrics = extractor.extract_all_financial_metrics(statement_type)
            
            # Convert to simple dictionary format
            result = {}
            for metric_name, metric in metrics.items():
                result[metric_name] = {
                    'values': metric.values,
                    'location': (metric.location.row, metric.location.column),
                    'data_columns': metric.data_columns
                }
            
            return result
    except Exception as e:
        logger.error(f"Error extracting financial data from {file_path}: {e}")
        return {}

if __name__ == "__main__":
    # Test the Excel utilities
    import sys
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        print(f"Testing Excel utilities with file: {file_path}")
        
        company_name = get_company_name_from_excel(file_path)
        print(f"Company name: {company_name}")
        
        period_dates = get_period_dates_from_excel(file_path)
        print(f"Period dates: {period_dates}")
        
        financial_data = extract_financial_data_from_excel(file_path, 'income')
        print(f"Financial data keys: {list(financial_data.keys())}")
    else:
        print("Usage: python excel_utils.py <path_to_excel_file>")